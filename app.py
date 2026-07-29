# Version 3.4b
from flask import Flask, request, jsonify, render_template_string, send_file, session
import anthropic
import base64
import json
import gspread
from google.oauth2.service_account import Credentials
import os
import time
import uuid
import traceback
import io
import urllib.request
import urllib.error
from datetime import datetime
from zoneinfo import ZoneInfo

def ahora_gt():
    """Devuelve la fecha/hora actual en la zona horaria de Guatemala (America/Guatemala),
    para que las fechas mostradas y guardadas coincidan con la hora real del país,
    sin importar en qué zona horaria corra el servidor."""
    return datetime.now(ZoneInfo("America/Guatemala"))


def _dias_desde_fecha_ddmmaaaa(fecha_str):
    """Calcula cuántos días completos han pasado desde una fecha guardada
    en formato DD/MM/AAAA. Devuelve None si la fecha está vacía o no se
    puede interpretar (por ejemplo, si nunca se revisó)."""
    fecha_str = (fecha_str or "").strip()
    if not fecha_str:
        return None
    try:
        dia, mes, anio = fecha_str.split("/")
        fecha = datetime(int(anio), int(mes), int(dia), tzinfo=ZoneInfo("America/Guatemala"))
        return (ahora_gt() - fecha).days
    except Exception:
        return None

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Image as RLImage

app = Flask(__name__)

# Clave para firmar las cookies de sesión. Se puede fijar con la variable de
# entorno FLASK_SECRET_KEY para que las sesiones sobrevivan a un reinicio del
# servidor; si no se define, se genera una nueva al arrancar (los usuarios
# solo tendrían que volver a iniciar sesión después de cada despliegue).
app.secret_key = os.environ.get("FLASK_SECRET_KEY") or os.urandom(32)

from datetime import timedelta
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=10)
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get("FLASK_INSECURE_COOKIES") != "1"
app.config['SESSION_COOKIE_HTTPONLY'] = True

SHEET_ID = os.environ.get("SHEET_ID", "1cUp1M899jUbHegB1j1S5hkM5BbRl2OSXPLFRQ2PI65c")

# PIN de administrador (Mantenimiento). Vive solo en el servidor y nunca se
# envía al navegador; se puede cambiar sin tocar el código definiendo la
# variable de entorno ADMIN_PIN en Render.
ADMIN_PIN = os.environ.get("ADMIN_PIN", "237546")
# PIN del usuario de apoyo: acceso muy limitado, solo a Registros
# Individuales (buscar y editar teléfono/empadronamiento). Se puede
# cambiar definiendo la variable de entorno APOYO_PIN en Render.
APOYO_PIN = os.environ.get("APOYO_PIN", "445566")
# PIN de "Afiliación al Partido": acceso muy limitado, solo para
# apoyo1/apoyo2, solo para registrar afiliados nuevos. Por defecto son las
# teclas del teclado alfanumérico del teléfono que corresponden a la
# palabra "TOKIO" (T=8, O=6, K=5, I=4, O=6). Se puede cambiar definiendo
# la variable de entorno AFILIACION_PIN en Render.
AFILIACION_PIN = os.environ.get("AFILIACION_PIN", "86546")
HOJA = "DPI"
HOJA_PADRON = "PADRON"

def _cargar_credenciales_google():
    """
    Carga las credenciales de la cuenta de servicio de Google desde una
    variable de entorno, para no dejarlas escritas en el código fuente.

    Soporta dos formas de configurarla:
    1) GOOGLE_CREDENTIALS_JSON = el contenido completo del JSON de la cuenta
       de servicio (como texto).
    2) GOOGLE_CREDENTIALS_FILE = ruta a un archivo .json con esas credenciales
       (útil en entornos donde es más fácil montar un archivo que definir
       una variable de entorno larga).
    """
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if creds_json:
        try:
            return json.loads(creds_json)
        except json.JSONDecodeError as e:
            raise RuntimeError(
                f"GOOGLE_CREDENTIALS_JSON no contiene un JSON valido: {e}"
            )

    creds_file = os.environ.get("GOOGLE_CREDENTIALS_FILE")
    if creds_file:
        with open(creds_file, "r", encoding="utf-8") as f:
            return json.load(f)

    raise RuntimeError(
        "No se encontraron credenciales de Google. Define la variable de "
        "entorno GOOGLE_CREDENTIALS_JSON (contenido del JSON de la cuenta "
        "de servicio) o GOOGLE_CREDENTIALS_FILE (ruta a ese archivo)."
    )


CREDENCIALES_GOOGLE = _cargar_credenciales_google()

# Caché temporal en memoria para servir PDFs recién generados vía una URL
# real del servidor (en vez de un blob: del navegador). Esto es necesario
# porque algunos navegadores de Android descargan un blob: directamente en
# lugar de mostrarlo, pero sí respetan "Content-Disposition: inline" cuando
# el PDF viene de una URL real. Se limpia solo (entradas de más de 10 min).
_pdf_temp_cache = {}


def _guardar_pdf_temporal(pdf_bytes, filename):
    token = uuid.uuid4().hex
    ahora = time.time()
    _pdf_temp_cache[token] = (pdf_bytes, filename, ahora)
    viejos = [k for k, v in _pdf_temp_cache.items() if ahora - v[2] > 600]
    for k in viejos:
        del _pdf_temp_cache[k]
    return token


@app.route("/pdf_temporal/<token>")
def pdf_temporal(token):
    from flask import make_response
    entrada = _pdf_temp_cache.get(token)
    if not entrada:
        return "El PDF ya no está disponible (expiró o ya se abrió). Genéralo de nuevo.", 404
    pdf_bytes, filename, _ts = entrada
    resp = make_response(pdf_bytes)
    resp.headers['Content-Type'] = 'application/pdf'
    resp.headers['Content-Disposition'] = f'inline; filename="{filename}"'
    return resp


# ==== SESIÓN REAL DEL SERVIDOR ====
# A diferencia de la variable "sesionJefe" del navegador (que solo sirve para
# la interfaz), esto es lo que realmente protege los datos: una cookie de
# sesión firmada por el servidor, que ninguna persona puede fabricar sin
# conocer la clave secreta del servidor (app.secret_key), y que se valida en
# cada petición a las rutas sensibles, sin importar qué le diga el navegador.
from functools import wraps


def _iniciar_sesion(nombre, es_admin=False, es_apoyo=False, es_afiliacion=False):
    session['jefe_nombre'] = nombre
    session['es_admin'] = bool(es_admin)
    session['es_apoyo'] = bool(es_apoyo)
    session['es_afiliacion'] = bool(es_afiliacion)
    session.permanent = True


def _cerrar_sesion():
    session.pop('jefe_nombre', None)
    session.pop('es_admin', None)
    session.pop('es_apoyo', None)
    session.pop('es_afiliacion', None)


def _dpi_formato_invalido(cui):
    """True si el CUI no tiene el formato correcto de un DPI guatemalteco:
    debe tener exactamente 13 dígitos, y el código de departamento (los 2
    dígitos justo antes del código de municipio, al final del número) debe
    estar entre 01 y 22, ya que Guatemala solo tiene 22 departamentos — un
    código de 23 en adelante es señal de que el número se leyó o se
    capturó mal. Se usa tanto para resaltar en los reportes impresos como
    para bloquear el registro de una persona nueva con un DPI mal leído."""
    digitos = ''.join(c for c in str(cui or '') if c.isdigit())
    if len(digitos) != 13:
        return True
    codigo_departamento = digitos[9:11]
    try:
        return not (1 <= int(codigo_departamento) <= 22)
    except ValueError:
        return True


def _sesion_actual():
    """Devuelve (nombre, es_admin) de la sesión actual, o (None, False) si
    no hay ninguna sesión iniciada en el servidor."""
    return session.get('jefe_nombre'), bool(session.get('es_admin', False))


def _es_apoyo_sesion():
    """True si la sesión actual es la del usuario de Apoyo (acceso muy
    limitado: solo puede buscar y editar teléfono/empadronamiento desde
    Registros Individuales, nada más)."""
    return bool(session.get('es_apoyo', False))


def _es_afiliacion_sesion():
    """True si la sesión actual es la de apoyo1/apoyo2 en 'Afiliación al
    Partido' — acceso aparte y muy limitado: solo puede registrar
    afiliados nuevos, nada más de la aplicación."""
    return bool(session.get('es_afiliacion', False))


def _puede_ver_todo():
    """True si la sesión actual puede ver los registros de TODOS los jefes
    (no solo los propios): administradores y el usuario de Apoyo. Apoyo
    necesita ver a cualquier persona para poder ayudarla, aunque no tenga
    ningún otro privilegio de administrador."""
    _, es_admin = _sesion_actual()
    return es_admin or _es_apoyo_sesion()


# ==== LÍMITE DE INTENTOS DE PIN ====
# Se controla del lado del servidor (en la sesión, que ya persiste vía
# cookie), no solo en el navegador, para que no se pueda saltar recargando
# la página. Después de MAX_INTENTOS_PIN fallidos, se bloquea por un rato.
MAX_INTENTOS_PIN = 4
BLOQUEO_PIN_SEGUNDOS = 15 * 60  # 15 minutos


def _verificar_bloqueo_pin(clave):
    """Devuelve un mensaje de error si esta clave de intentos está
    bloqueada por demasiados fallos, o None si se puede intentar."""
    estado = session.get(clave, {})
    bloqueado_hasta = estado.get("bloqueado_hasta", 0)
    ahora = time.time()
    if bloqueado_hasta > ahora:
        minutos = max(1, int((bloqueado_hasta - ahora) / 60) + 1)
        return f"Demasiados intentos fallidos. Intente de nuevo en {minutos} minuto(s)."
    return None


def _registrar_intento_pin(clave, exitoso):
    """Registra un intento (éxito o fallo) de PIN. Si fue exitoso, limpia
    el contador. Si falló y ya llegó al máximo, activa el bloqueo temporal."""
    if exitoso:
        session.pop(clave, None)
        return
    estado = session.get(clave, {"fallos": 0, "bloqueado_hasta": 0})
    estado["fallos"] = estado.get("fallos", 0) + 1
    if estado["fallos"] >= MAX_INTENTOS_PIN:
        estado["bloqueado_hasta"] = time.time() + BLOQUEO_PIN_SEGUNDOS
        estado["fallos"] = 0
    session[clave] = estado
    session.permanent = True


def requiere_sesion(f):
    """Exige una sesión de jefe de sector (o administrador) válida, verificada
    por el servidor. Si no la hay, rechaza la petición sin importar lo que
    diga el navegador."""
    @wraps(f)
    def envoltura(*args, **kwargs):
        nombre, _ = _sesion_actual()
        if not nombre:
            return jsonify({"ok": False, "error": "Sesión no iniciada. Ingrese su PIN nuevamente.", "sesion_requerida": True}), 401
        return f(*args, **kwargs)
    return envoltura


def requiere_admin(f):
    """Exige específicamente una sesión de administrador (mismo PIN de
    Mantenimiento), verificada por el servidor."""
    @wraps(f)
    def envoltura(*args, **kwargs):
        nombre, es_admin = _sesion_actual()
        if not nombre or not es_admin:
            return jsonify({"ok": False, "error": "Se requiere sesión de administrador.", "sesion_requerida": True}), 401
        return f(*args, **kwargs)
    return envoltura


def requiere_afiliacion(f):
    """Exige específicamente una sesión de apoyo1/apoyo2 en 'Afiliación al
    Partido', verificada por el servidor — no sirve ninguna otra sesión
    (ni de jefe, ni de admin, ni de apoyo normal) para estas rutas."""
    @wraps(f)
    def envoltura(*args, **kwargs):
        if not _es_afiliacion_sesion():
            return jsonify({"ok": False, "error": "Se requiere sesión de Afiliación al Partido.", "sesion_requerida": True}), 401
        return f(*args, **kwargs)
    return envoltura


@app.route("/logout", methods=["POST"])
def logout():
    _cerrar_sesion()
    return jsonify({"ok": True})


@app.route("/sesion_actual")
def sesion_actual_ruta():
    nombre, es_admin = _sesion_actual()
    return jsonify({"ok": True, "nombre": nombre, "es_admin": es_admin, "es_apoyo": _es_apoyo_sesion(), "es_afiliacion": _es_afiliacion_sesion()})


@app.route("/login_admin", methods=["POST"])
def login_admin():
    """Valida el PIN de administrador (teclado de Mantenimiento) en el
    servidor, para que el PIN nunca viaje al navegador ni aparezca en el
    código fuente visible de la página. Limita a MAX_INTENTOS_PIN intentos
    fallidos antes de bloquear temporalmente."""
    try:
        bloqueo = _verificar_bloqueo_pin('intentos_admin')
        if bloqueo:
            return jsonify({"ok": False, "error": bloqueo, "bloqueado": True})
        pin = (request.json or {}).get("pin", "").strip()
        sh = get_sheet()
        if pin == ADMIN_PIN:
            _iniciar_sesion("ADMINISTRADOR", es_admin=True)
            _registrar_intento_pin('intentos_admin', exitoso=True)
            _registrar_auditoria(sh, "ADMINISTRADOR", "Inicio de sesión (Mantenimiento)", "Acceso correcto")
            return jsonify({"ok": True})
        _registrar_intento_pin('intentos_admin', exitoso=False)
        _registrar_auditoria(sh, "ADMINISTRADOR", "Intento de acceso fallido (Mantenimiento)", "PIN incorrecto")
        return jsonify({"ok": False, "error": "Clave incorrecta"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/login_apoyo", methods=["POST"])
def login_apoyo():
    """Valida el PIN del usuario de Apoyo — un acceso muy limitado, pensado
    para personas que solo necesitan buscar a alguien y corregir su
    teléfono o su empadronamiento, sin ver nada más de la aplicación."""
    try:
        bloqueo = _verificar_bloqueo_pin('intentos_apoyo')
        if bloqueo:
            return jsonify({"ok": False, "error": bloqueo, "bloqueado": True})
        pin = (request.json or {}).get("pin", "").strip()
        sh = get_sheet()
        if pin == APOYO_PIN:
            _iniciar_sesion("APOYO", es_admin=False, es_apoyo=True)
            _registrar_intento_pin('intentos_apoyo', exitoso=True)
            _registrar_auditoria(sh, "APOYO", "Inicio de sesión (Apoyo)", "Acceso correcto")
            return jsonify({"ok": True})
        _registrar_intento_pin('intentos_apoyo', exitoso=False)
        _registrar_auditoria(sh, "APOYO", "Intento de acceso fallido (Apoyo)", "PIN incorrecto")
        return jsonify({"ok": False, "error": "Clave incorrecta"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/login_afiliacion", methods=["POST"])
def login_afiliacion():
    """Valida el acceso a 'Afiliación al Partido' — solo para apoyo1 y
    apoyo2, con un único código compartido entre ambos. Acceso aparte y
    muy limitado: únicamente pueden registrar afiliados nuevos, nada más
    de la aplicación."""
    try:
        nombre = (request.json or {}).get("nombre", "").strip().upper()
        if nombre not in ("APOYO1", "APOYO2"):
            return jsonify({"ok": False, "error": "Usuario no válido"})
        bloqueo = _verificar_bloqueo_pin('intentos_afiliacion')
        if bloqueo:
            return jsonify({"ok": False, "error": bloqueo, "bloqueado": True})
        pin = (request.json or {}).get("pin", "").strip()
        sh = get_sheet()
        if pin == AFILIACION_PIN:
            _iniciar_sesion(nombre, es_admin=False, es_apoyo=False, es_afiliacion=True)
            _registrar_intento_pin('intentos_afiliacion', exitoso=True)
            _registrar_auditoria(sh, nombre, "Inicio de sesión (Afiliación al Partido)", "Acceso correcto")
            return jsonify({"ok": True, "nombre": nombre})
        _registrar_intento_pin('intentos_afiliacion', exitoso=False)
        _registrar_auditoria(sh, nombre, "Intento de acceso fallido (Afiliación al Partido)", "Código incorrecto")
        return jsonify({"ok": False, "error": "Código incorrecto"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def formatear_telefono_gt(telefono):
    """
    Antepone el prefijo internacional de Guatemala (+502) a un numero de
    telefono, si no lo tiene ya. Se usa en todos los puntos donde se guarda
    un telefono en las hojas de Google Sheets.
    """
    t = (telefono or "").strip()
    if not t:
        return ""
    if t.startswith("+502"):
        return t
    # Quitar un '502' inicial sin '+' para no duplicarlo (ej. "502 1234-5678")
    t_sin_espacios = t.replace(" ", "")
    if t_sin_espacios.startswith("502") and len(t_sin_espacios) > 8:
        return "+" + t_sin_espacios
    return "+502" + t


def _formatear_cui_espacios(cui):
    """Muestra el CUI/DPI con el formato de grupos que trae el documento
    físico (4-5-4, ej. '3104 13451 0701'), para que sea más fácil cotejarlo
    a mano contra el DPI físico al imprimir un reporte. Si el CUI no mide
    los 13 dígitos esperados (dato incompleto o con letras), se devuelve
    tal cual venía, sin forzar el formato."""
    c = (cui or "").replace(" ", "").replace("-", "").strip()
    if len(c) == 13 and c.isdigit():
        return f"{c[0:4]} {c[4:9]} {c[9:13]}"
    return cui or "-"


PROMPT = """Eres un experto en lectura de DPI guatemaltecos, y también en detectar
intentos de fraude con documentos falsos o simulados.

PASO 1 - VERIFICAR AUTENTICIDAD:
Evalúa si las dos imágenes (frente y reverso) muestran REALMENTE una tarjeta
física de DPI guatemalteco oficial, con sus elementos característicos:
fotografía de la persona impresa en la tarjeta, diseño oficial del
RENAP/CABAL con holograma y elementos de seguridad, texto impreso (no
escrito a mano), y el formato estándar de tarjeta plástica.

Marca "es_dpi_autentico": false si detectas cualquiera de estas señales de
que NO es un DPI físico real:
- Es una nota escrita a mano (en papel, cuaderno, pizarra, etc.) simulando
  contener datos de un DPI, aunque el texto diga "DPI", "CUI", "Nombre", etc.
- Es una captura de pantalla, un documento de texto, o una imagen generada
  digitalmente en vez de una fotografía de la tarjeta física.
- Es un papel o cartulina sin el diseño oficial, sin holograma, sin foto de
  una persona real impresa en la tarjeta.
- El fondo, iluminación o ángulo sugieren que se fotografió un papel plano
  con texto en vez de una tarjeta plástica con relieve y foto.
- Cualquier otra señal de que se intenta simular un DPI en vez de fotografiar
  uno real.

Si "es_dpi_autentico" es false, explica brevemente el motivo en
"motivo_rechazo".

PASO 2 - EXTRAER DATOS (siempre, sin importar el resultado del paso 1):
Analiza las dos imagenes y extrae los datos que puedas leer, aunque hayas
marcado "es_dpi_autentico" como false — la persona que usa el sistema debe
poder ver qué datos se leyeron para decidir si continúa o no.

REGLA ESTRICTA PARA CAMPOS NO LEGIBLES: si algún dato está tapado, borroso,
cortado fuera de la imagen, o por cualquier motivo no puedes leerlo con
certeza total, devuelve ese campo como cadena vacía "". NUNCA completes,
adivines, infieras ni calcules un dato que no puedas leer directamente en la
imagen — esto aplica especialmente a "fecha_nacimiento", pero también a
todos los demás campos. Es preferible dejar un campo vacío a inventarlo,
aunque parezca "lógico" o puedas deducirlo de otros datos.

Devuelve UNICAMENTE un JSON valido con estas claves exactas:
{
  "es_dpi_autentico": true,
  "motivo_rechazo": "",
  "cui": "",
  "numero_serie": "",
  "primer_nombre": "",
  "segundo_nombre": "",
  "primer_apellido": "",
  "segundo_apellido": "",
  "sexo": "",
  "estado_civil": "",
  "fecha_nacimiento": "",
  "municipio_nacimiento": "",
  "departamento_nacimiento": "",
  "municipio_vecindad": "",
  "departamento_vecindad": "",
  "fecha_expedicion": "",
  "fecha_vencimiento": ""
}
Fechas en formato DD/MM/YYYY. Sexo: M o F. Sin texto extra fuera del JSON."""

HTML = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Cache-Control" content="no-cache, no-store, must-revalidate">
<meta http-equiv="Pragma" content="no-cache">
<meta http-equiv="Expires" content="0">
<title>Lector DPI v12.7</title>
<style>
* { box-sizing:border-box; margin:0; padding:0; }
body { background:#1a5276; font-family:Arial,sans-serif; min-height:100vh; display:flex; flex-direction:column; align-items:center; padding:20px; }
h1 { color:white; font-size:2em; margin:20px 0 5px; text-align:center; }
.sub { color:#aed6f1; margin-bottom:25px; text-align:center; font-size:1.4em; font-weight:bold; }
.card { background:#1f618d; border-radius:20px; padding:25px; width:100%; max-width:500px; }
.btn { width:100%; padding:22px; margin:8px 0; border:none; border-radius:14px; font-size:1.2em; font-weight:bold; cursor:pointer; color:white; }
.overlay { display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.85); z-index:100; overflow-y:auto; padding:15px; }
.box { background:#1f618d; border-radius:20px; padding:25px; max-width:500px; margin:20px auto; }
.box h2 { color:white; text-align:center; margin-bottom:20px; }
.mbtn { width:100%; padding:18px; margin:8px 0; border:none; border-radius:12px; font-size:1.1em; font-weight:bold; cursor:pointer; color:white; }
input,select { width:100%; padding:13px; border-radius:10px; border:none; font-size:1em; margin-bottom:10px; }
.msg { color:#aed6f1; text-align:center; margin-top:10px; }
.teclado { display:grid; grid-template-columns:repeat(3,1fr); gap:10px; margin:15px 0; }
.tecla { padding:18px; font-size:1.3em; font-weight:bold; background:#2980b9; color:white; border:none; border-radius:12px; cursor:pointer; display:flex; flex-direction:column; align-items:center; justify-content:center; line-height:1.1; }
.tecla-letras { font-size:0.32em; font-weight:normal; letter-spacing:1.5px; margin-top:2px; color:#d6eaf8; }
.tabla-wrap { overflow-x:auto; margin-top:10px; max-height:250px; overflow-y:auto; }
table { width:100%; border-collapse:collapse; font-size:0.78em; }
th { background:#154360; color:#aed6f1; padding:8px; text-align:left; }
td { color:white; padding:7px 8px; border-bottom:1px solid #2980b9; }
.jgrid { display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:15px; }
.jbtn { padding:12px 6px; border:none; border-radius:10px; font-size:0.82em; font-weight:bold; cursor:pointer; color:white; background:#566573; }
.jbtn.activo { background:#27ae60; }
.jbtn.tiene { background:#1a5276; }
</style>
</head>
<body>
<!-- Input de archivo COMPARTIDO para las 3 pantallas que piden foto de DPI
     (Presidenta, Coordinadora, Integrante). Antes cada pantalla tenía sus
     propios 2 inputs (Frente/Reverso); en algunos navegadores móviles, el
     SEGUNDO input de archivo de una pantalla dejaba de abrir aunque el
     primero funcionara bien. Usar un solo input compartido para las 6
     situaciones evita ese problema, ya que siempre es el mismo elemento el
     que se abre, solo cambiando a dónde se guarda el resultado. -->
<input type="file" id="fotoInputCompartido" accept="image/*" style="display:none">
<h1>📋 LECTOR DE DPI</h1>
<p class="sub">TOTONICAPÁN</p>
<div class="card">
  <button class="btn" style="background:#7d3c98;font-size:0.85em" id="b6" onclick="abrirLoginAfiliacion()">🎫 Afiliación al Partido</button>
  <button class="btn" style="background:#117864;font-size:0.85em" id="b5" onclick="abrirApoyo()">🤝 Apoyo</button>
  <button class="btn" style="background:#6c3483" id="b2" onclick="show('oConsultas')">📊 Consultas</button>
  <button class="btn" style="background:#1a5276" id="b1" onclick="abrirGrupos()">👥 Grupos</button>
  <button class="btn" style="background:#922b21" id="b3" onclick="show('oMant');clv='';updC();document.getElementById('mantLogin').style.display='block';document.getElementById('mantPanel').style.display='none'">🔧 Mantenimiento</button>
  <button class="btn" style="background:#16a085" id="b4" onclick="abrirReportes()">📈 Reportes</button>
  <button class="btn" style="background:#7b241c;display:none" id="bLogout" onclick="cerrarSesionCompleta()">🚪 Cerrar sesión (<span id="bLogoutNombre"></span>)</button>
</div>

<div class="overlay" id="oLoginJefe">
<div class="box">
  <h2>🔐 Ingreso Jefe de Sector</h2>
  <p style="color:#aed6f1;margin-bottom:6px">Su nombre no aparece en la lista o su PIN no funciona: contacte al administrador.</p>
  <p style="color:#aed6f1;margin-bottom:6px;margin-top:10px">Jefe de Sector:</p>
  <select id="loginJefeSel"><option value="">-- Seleccione --</option></select>
  <p style="color:#aed6f1;margin-bottom:6px;margin-top:10px">PIN:</p>
  <input type="password" id="loginJefePin" inputmode="numeric" placeholder="PIN">
  <button class="mbtn" style="background:#27ae60" onclick="hacerLoginJefe()">Ingresar</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oLoginJefe')">Cancelar</button>
  <p id="msgLoginJefe" class="msg"></p>
</div>
</div>

<div class="overlay" id="oLoginApoyo">
<div class="box">
  <h2>🤝 Ingreso de Apoyo</h2>
  <p style="color:#aed6f1;margin-bottom:10px">Acceso limitado: solo búsqueda y edición de teléfono/empadronamiento.</p>
  <p style="color:#aed6f1;margin-bottom:6px">PIN:</p>
  <input type="password" id="loginApoyoPin" inputmode="numeric" placeholder="PIN">
  <button class="mbtn" style="background:#27ae60" onclick="hacerLoginApoyo()">Ingresar</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oLoginApoyo')">Cancelar</button>
  <p id="msgLoginApoyo" class="msg"></p>
</div>
</div>

<div class="overlay" id="oLoginAfiliacion">
<div class="box">
  <h2>🎫 Afiliación al Partido</h2>
  <p style="color:#aed6f1;margin-bottom:10px">Acceso limitado: solo para registrar afiliados nuevos.</p>
  <p style="color:#aed6f1;margin-bottom:6px">Usuario:</p>
  <select id="loginAfiliacionUsuario">
    <option value="APOYO1">Apoyo 1</option>
    <option value="APOYO2">Apoyo 2</option>
  </select>
  <p style="color:#aed6f1;margin-bottom:6px;margin-top:10px">Código:</p>
  <input type="password" id="loginAfiliacionPin" inputmode="numeric" placeholder="Código">
  <button class="mbtn" style="background:#27ae60" onclick="hacerLoginAfiliacion()">Ingresar</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oLoginAfiliacion')">Cancelar</button>
  <p id="msgLoginAfiliacion" class="msg"></p>
</div>
</div>

<div class="overlay" id="oAfiliacionMenu">
<div class="box">
  <h2>🎫 Afiliación al Partido</h2>
  <p style="color:#aed6f1;font-size:0.85em;margin-bottom:12px">¿Qué quieres hacer?</p>
  <button class="mbtn" style="background:#7d3c98" onclick="irAAfiliacionRegistro()">🎫 Afiliación (registrar nuevo)</button>
  <button class="mbtn" style="background:#1a3a6b" onclick="irAImpresionCarnets()">🖨️ Impresión (carnés ya guardados)</button>
  <button class="mbtn" style="background:#7b241c" onclick="cerrarSesionCompleta()">🚪 Cerrar sesión</button>
</div>
</div>

<div class="overlay" id="oAfiliacionRegistro">
<div class="box">
  <h2>🎫 Registro de Afiliación al Partido</h2>
  <p style="color:#aed6f1;font-size:0.8em;margin-bottom:10px">Toma la foto del frente y el reverso del DPI de la persona que se afilia. Se guarda por separado de los registros normales, y se revisa automáticamente si ya está empadronada.</p>
  <input type="file" id="afF" accept="image/*" style="display:none">
  <input type="file" id="afR" accept="image/*" style="display:none">
  <div id="pasoAfil">
    <label style="display:flex;align-items:center;gap:8px;color:#aed6f1;font-size:0.85em;margin:6px 2px 10px 2px">
      <input type="checkbox" id="afCamaraDirecta" checked style="width:18px;height:18px;margin:0">
      Ir directo a la cámara (desmarca para elegir un archivo/galería)
    </label>
    <button class="mbtn" style="background:#7d3c98" id="btnAfF" onclick="abrirSelectorFotoAfiliacion('afiliacionF')">📷 Foto Frente del DPI</button>
    <button class="mbtn" style="background:#7d3c98" id="btnAfR" onclick="abrirSelectorFotoAfiliacion('afiliacionR')">📷 Foto Reverso del DPI</button>
    <input type="text" id="afTel" inputmode="numeric" placeholder="Teléfono (opcional)">
    <input type="text" id="afDir" placeholder="Dirección / Comunidad">
    <button class="mbtn" style="background:#566573" id="btnAfProc" disabled onclick="procAfiliado()">Analizar y guardar afiliado</button>
  </div>
  <p id="msgAfil" class="msg"></p>
  <button class="mbtn" style="background:#1a3a6b;display:none" id="btnImprimirCarnet" onclick="abrirFirmaCarnet()">🖨️ Imprimir carné de este afiliado</button>
  <button class="mbtn" style="background:#566573" onclick="hide('oAfiliacionRegistro');show('oAfiliacionMenu')">⬅ Menú</button>
  <button class="mbtn" style="background:#7b241c" onclick="cerrarSesionCompleta()">🚪 Cerrar sesión</button>
</div>
</div>

<div class="overlay" id="oImpresionCarnets">
<div class="box">
  <h2>🖨️ Impresión de carnés</h2>
  <p style="color:#aed6f1;font-size:0.8em;margin-bottom:10px">Elige la fecha en que se registraron los afiliados que quieres imprimir. Se genera una lista tamaño carta (10 afiliados por hoja), con espacio para que cada quien firme o ponga su huella.</p>
  <div style="display:flex;gap:8px;align-items:center;margin-bottom:8px">
    <input type="date" id="fechaImpresionCarnets" style="flex:1;padding:8px;border-radius:8px;border:none">
    <button type="button" style="background:#2980b9;color:white;border:none;border-radius:8px;padding:9px 14px;font-size:0.85em;font-weight:bold;cursor:pointer" onclick="document.getElementById('fechaImpresionCarnets').value=new Date().toLocaleDateString('en-CA')">Hoy</button>
  </div>
  <button class="mbtn" style="background:#8e44ad" onclick="buscarAfiliadosParaImprimir()">🔍 Buscar afiliados de esa fecha</button>
  <div id="resImpresionCarnets" style="margin-top:8px"></div>
  <p id="msgImpresionCarnets" class="msg"></p>
  <button class="mbtn" style="background:#566573" onclick="hide('oImpresionCarnets');show('oAfiliacionMenu')">⬅ Menú</button>
  <button class="mbtn" style="background:#7b241c" onclick="cerrarSesionCompleta()">🚪 Cerrar sesión</button>
</div>
</div>

<div class="overlay" id="oFirmaCarnet">
<div class="box">
  <h2>✍️ Firma del afiliado</h2>
  <div style="display:flex;gap:8px;margin-bottom:10px">
    <button class="mbtn" style="background:#1a3a6b;margin:0;flex:1" id="btnModoFirmaDibujar" onclick="cambiarModoFirma('dibujar')">✍️ Firmar en pantalla</button>
    <button class="mbtn" style="background:#566573;margin:0;flex:1" id="btnModoFirmaRecorte" onclick="cambiarModoFirma('recorte')">🪪 Recortar de la foto del DPI</button>
  </div>

  <div id="divFirmaDibujar">
    <p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">Pide a la persona que firme (o presione su huella) con el dedo dentro del recuadro. Se imprimirá en el carné, en el espacio de "Firma del afiliado".</p>
    <div style="background:white;border-radius:8px;padding:4px;touch-action:none">
      <canvas id="canvasFirma" width="600" height="420" style="width:100%;height:auto;display:block;border:2px dashed #7f8c8d;border-radius:6px;touch-action:none"></canvas>
    </div>
    <div style="display:flex;gap:8px;margin-top:8px">
      <button class="mbtn" style="background:#566573;margin:0;flex:1" onclick="limpiarFirma()">🧹 Limpiar</button>
      <button class="mbtn" style="background:#7f8c8d;margin:0;flex:1" onclick="generarCarnetConFirma(true)">Omitir firma</button>
    </div>
  </div>

  <div id="divFirmaRecorte" style="display:none">
    <p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">Mueve y ajusta el recuadro verde hasta que quede justo sobre la firma o la huella que aparece impresa en la foto del FRENTE del DPI (arrastra el centro para moverlo, y la esquina para cambiar su tamaño). Si no se ve bien, mejor usa "Firmar en pantalla".</p>
    <div id="contRecorteDPI" style="background:#0d2137;border-radius:8px;padding:4px;touch-action:none;position:relative">
      <canvas id="canvasRecorteDPI" style="width:100%;height:auto;display:block;border:2px dashed #7f8c8d;border-radius:6px;touch-action:none"></canvas>
    </div>
    <button class="mbtn" style="background:#2980b9;margin-top:8px" onclick="confirmarRecorteDPI()">✅ Confirmar este recuadro</button>
    <div id="previewRecorteDPI" style="display:none;margin-top:8px;text-align:center">
      <p style="color:#aed6f1;font-size:0.78em;margin-bottom:4px">Así se vería:</p>
      <img id="imgPreviewRecorte" style="max-width:60%;background:white;border-radius:6px;border:2px solid #27ae60">
    </div>
  </div>

  <button class="mbtn" style="background:#27ae60" onclick="generarCarnetConFirma(false)">✅ Guardar y generar carné</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oFirmaCarnet')">Cancelar</button>
  <p id="msgFirma" class="msg"></p>
</div>
</div>

<div class="overlay" id="oGrupos">
<div class="box">
  <h2>👥 Grupos</h2>
  <p id="sesionJefeInfo" style="color:#2ecc71;text-align:center;font-size:0.85em;margin-bottom:10px"></p>
  <button class="mbtn" style="background:#27ae60" id="bNuevo" onclick="hide('oGrupos');gN='';gJ='';gC=null;gPs=[];document.getElementById('ngN').value='';show('oNuevoG1')">➕ Nuevo Grupo</button>
  <button class="mbtn" style="background:#117a65" id="bContinuar" onclick="hide('oGrupos');document.getElementById('bgN').value='';document.getElementById('bgD').value='';document.getElementById('resBG').style.display='none';msg('msgBG','','');show('oBuscarG')">🔍 Continuar Grupo</button>
  <button class="mbtn" style="background:#566573" onclick="cerrarSesionJefe()">🔓 Cerrar sesión</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oGrupos')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oConsultas">
<div class="box">
  <h2>📊 Consultas</h2>
  <button class="mbtn" style="background:#2980b9" id="bRegInd" onclick="abrirRegistrosInd()">👤 Registros Individuales</button>
  <button class="mbtn" style="background:#8e44ad" id="bRegGrp" onclick="hide('oConsultas');cargarGrupos()">👥 Registros de Grupos</button>
  <button class="mbtn" style="background:#117a65" id="bRegPres" onclick="abrirPresidentasConsulta()">👩 Presidentas de Comité</button>
  <button class="mbtn" style="background:#1a5276" onclick="abrirConsultaPadron()">🗳️ Consultar Padrón</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oConsultas')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oConsultaPadron">
<div class="box">
  <h2>🗳️ Consultar Padrón</h2>
  <input type="text" id="padronQ" placeholder="Nombre o número de DPI">
  <button class="mbtn" style="background:#1a5276" onclick="buscarPadron()">Buscar</button>
  <div id="resPadron" class="msg" style="max-height:50vh;overflow-y:auto;margin-top:10px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsultaPadron');show('oConsultas')">← Regresar</button>
</div>
</div>

<!-- OVERLAY CONSULTA PRESIDENTAS -->
<div class="overlay" id="oConsPresidentas">
<div class="box">
  <h2>👩 Presidentas de Comité</h2>
  <div style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:12px" id="botonesJefesPres"></div>
  <div id="contPresidentas" class="msg" style="max-height:45vh;overflow-y:auto">Seleccione un Jefe de Sector</div>
  <div id="botonesFijosPres" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsPresidentas'); show(_origenPresCoord==='reportes'?'oReportes':'oConsultas')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oReportes">
<div class="box">
  <h2>📈 Reportes</h2>
  <p id="sesionJefeInfoRep" style="color:#2ecc71;text-align:center;font-size:0.85em;margin-bottom:10px"></p>
  <button class="mbtn" style="background:#117a65" onclick="abrirPresidentasReporte()">👩 Presidentas de Comité</button>
  <button class="mbtn" style="background:#2980b9" onclick="abrirCoordinadorasReporte()">👥 Coordinadoras de Grupo</button>
  <button class="mbtn" style="background:#8e44ad" onclick="abrirIntegrantesReporte()">🧑‍🤝‍🧑 Integrantes</button>
  <button class="mbtn" style="background:#d68910" onclick="abrirPresCoordReporte()">👑 Presidentas + Coordinadoras</button>
  <button class="mbtn" style="background:#117864" onclick="abrirTodoCombinadoReporte()">🌟 Presidentas + Coordinadoras + Integrantes</button>
  <button class="mbtn" style="background:#a04000" onclick="abrirNoEmpadronadosReporte()">😕 No Empadronado y/o No Teléfono</button>
  <button class="mbtn" style="background:#b7950b" onclick="cargarConsObservaciones()">📝 Personas con Observaciones</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oReportes')">✔ Terminar</button>
</div>
</div>

<div class="overlay" id="oConsObservaciones">
<div class="box">
  <h2>📝 Personas con Observaciones</h2>
  <div id="contObservaciones" class="msg" style="max-height:45vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosObservaciones" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsObservaciones');show('oReportes')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oConsTodoCombinado">
<div class="box">
  <h2>🌟 Presidentas + Coordinadoras + Integrantes</h2>
  <div id="contTodoCombinado" class="msg" style="max-height:45vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosTodoCombinado" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsTodoCombinado');show('oReportes')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oSeleccionNoEmpTel">
<div class="box">
  <h2>😕 No Empadronado y/o No Teléfono</h2>
  <p style="color:#aed6f1;font-size:0.85em;margin-bottom:12px">Selecciona qué quieres ver (puedes marcar una o las dos):</p>
  <label style="display:flex;align-items:center;gap:10px;background:#154360;border-radius:8px;padding:12px;margin-bottom:10px;cursor:pointer">
    <input type="checkbox" id="chkNoEmp" checked style="width:22px;height:22px">
    <span style="color:white;font-size:0.95em">Sin empadronar</span>
  </label>
  <label style="display:flex;align-items:center;gap:10px;background:#154360;border-radius:8px;padding:12px;margin-bottom:14px;cursor:pointer">
    <input type="checkbox" id="chkSinTel" style="width:22px;height:22px">
    <span style="color:white;font-size:0.95em">Sin número de teléfono</span>
  </label>
  <div style="background:#154360;border-radius:8px;padding:12px;margin-bottom:14px">
    <p style="color:#aed6f1;font-size:0.85em;margin:0 0 8px 0">📅 Ver solo los agregados en una fecha (opcional):</p>
    <div style="display:flex;gap:8px;align-items:center">
      <input type="date" id="filtroFechaNoEmp" style="flex:1;padding:8px;border-radius:8px;border:none">
      <button type="button" style="background:#2980b9;color:white;border:none;border-radius:8px;padding:9px 14px;font-size:0.85em;font-weight:bold;cursor:pointer" onclick="document.getElementById('filtroFechaNoEmp').value=new Date().toLocaleDateString('en-CA')">Hoy</button>
      <button type="button" style="background:#566573;color:white;border:none;border-radius:8px;padding:9px 12px;font-size:0.85em;font-weight:bold;cursor:pointer" onclick="document.getElementById('filtroFechaNoEmp').value=''">✖</button>
    </div>
    <p style="color:#7fb3d5;font-size:0.7em;margin:6px 0 0 0">Deja vacío para ver todos, sin importar cuándo se agregaron.</p>
  </div>
  <button class="mbtn" style="background:#a04000" onclick="confirmarSeleccionNoEmpTel()">Ver listado</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oSeleccionNoEmpTel');show('oReportes')">← Regresar</button>
  <p id="msgSeleccionNoEmpTel" class="msg"></p>
</div>
</div>

<div class="overlay" id="oConsNoEmpadronados">
<div class="box">
  <h2 id="tituloConsNoEmpadronados">😕 No Empadronado y/o No Teléfono</h2>
  <div id="contNoEmpadronados" class="msg" style="max-height:45vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosNoEmpadronados" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsNoEmpadronados');show('oReportes')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oConsPresCoord">
<div class="box">
  <h2>👑 Presidentas + Coordinadoras</h2>
  <div id="contPresCoord" class="msg" style="max-height:45vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosPresCoord" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsPresCoord');show('oReportes')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oConsIntegrantes">
<div class="box">
  <h2>🧑‍🤝‍🧑 Integrantes</h2>
  <div id="contIntegrantes" class="msg" style="max-height:45vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosIntegrantes" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsIntegrantes');show('oReportes')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oConsCoordinadoras">
<div class="box">
  <h2>👥 Coordinadoras de Grupo</h2>
  <div id="contCoordinadoras" class="msg" style="max-height:45vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosCoordinadoras" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oConsCoordinadoras');show('oReportes')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oMant">
<div class="box">
  <h2>🔧 Mantenimiento</h2>
  <div id="mantLogin">
    <p style="color:#aed6f1;text-align:center;margin-bottom:10px">Ingrese su clave</p>
    <div id="dispClave" style="text-align:center;font-size:2em;letter-spacing:12px;color:white;margin:15px 0;">_ _ _ _ _ _</div>
    <div class="teclado">
      <button class="tecla" id="tk1">1</button><button class="tecla" id="tk2">2<span class="tecla-letras">ABC</span></button><button class="tecla" id="tk3">3<span class="tecla-letras">DEF</span></button>
      <button class="tecla" id="tk4">4<span class="tecla-letras">GHI</span></button><button class="tecla" id="tk5">5<span class="tecla-letras">JKL</span></button><button class="tecla" id="tk6">6<span class="tecla-letras">MNO</span></button>
      <button class="tecla" id="tk7">7<span class="tecla-letras">PQRS</span></button><button class="tecla" id="tk8">8<span class="tecla-letras">TUV</span></button><button class="tecla" id="tk9">9<span class="tecla-letras">WXYZ</span></button>
      <button class="tecla" style="background:#922b21" id="tkDel">⌫</button>
      <button class="tecla" id="tk0">0</button>
      <button class="tecla" style="background:#27ae60" id="tkOk">✓</button>
    </div>
    <p id="errClave" style="color:#e74c3c;text-align:center;display:none">Clave incorrecta</p>
  </div>
  <div id="mantPanel" style="display:none">
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:6px;margin-bottom:14px">
      <button id="tInd" style="background:#2980b9;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('ind')">👤<br>Registros</button>
      <button id="tGrp" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('grp')">👥<br>Grupos</button>
      <button id="tPres" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('pres')">👩<br>Presidentas</button>
      <button id="tJef" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('jef')">🧑‍💼<br>Jefes</button>
      <button id="tAud" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('aud')">📋<br>Auditoría</button>
      <button id="tBor" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('bor')">🗑️<br>Borrado</button>
      <button id="tPad" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('pad')">🔄<br>Padrón</button>
      <button id="tPrueba" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('prueba')">🧪<br>Pruebas</button>
      <button id="tAfil" style="background:#566573;color:white;border:none;border-radius:10px;padding:10px 4px;font-size:0.72em;font-weight:bold;cursor:pointer" onclick="tabM('afil')">🎫<br>Afiliados</button>
    </div>

    <div id="tabInd" style="display:block">
      <input type="text" id="mDPI" inputmode="numeric" placeholder="🔍 Buscar por DPI">
      <button class="mbtn" style="background:#2980b9" onclick="buscarReg()">Buscar registro</button>
      <div id="resReg" style="display:none">
        <div id="datReg" style="background:#154360;border-radius:8px;padding:10px;margin:8px 0;color:white;font-size:0.85em"></div>
        <div id="editRegForm" style="display:none;background:#0d2137;border-radius:8px;padding:10px;margin-bottom:8px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:0 0 4px 2px">CUI (DPI)</p>
          <input type="text" id="editRegCUI" inputmode="numeric" placeholder="Número de CUI/DPI (13 dígitos)" style="margin-bottom:6px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:8px 0 4px 2px">Nombres y apellidos</p>
          <input type="text" id="editRegPNom" placeholder="Primer nombre" style="margin-bottom:6px">
          <input type="text" id="editRegSNom" placeholder="Segundo nombre" style="margin-bottom:6px">
          <input type="text" id="editRegPApe" placeholder="Primer apellido" style="margin-bottom:6px">
          <input type="text" id="editRegSApe" placeholder="Segundo apellido" style="margin-bottom:6px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:8px 0 4px 2px">Datos del DPI</p>
          <input type="text" id="editRegSerie" placeholder="Número de serie" style="margin-bottom:6px">
          <select id="editRegSexo" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:6px">
            <option value="">Sexo: --</option>
            <option value="M">Sexo: Masculino</option>
            <option value="F">Sexo: Femenino</option>
          </select>
          <input type="text" id="editRegEC" placeholder="Estado civil" style="margin-bottom:6px">
          <input type="text" id="editRegFNac" placeholder="Fecha de nacimiento (DD/MM/AAAA)" style="margin-bottom:6px">
          <input type="text" id="editRegMunNac" placeholder="Municipio de nacimiento" style="margin-bottom:6px">
          <input type="text" id="editRegDepNac" placeholder="Departamento de nacimiento" style="margin-bottom:6px">
          <input type="text" id="editRegMunVec" placeholder="Municipio de vecindad" style="margin-bottom:6px">
          <input type="text" id="editRegDepVec" placeholder="Departamento de vecindad" style="margin-bottom:6px">
          <input type="text" id="editRegFExp" placeholder="Fecha de expedición (DD/MM/AAAA)" style="margin-bottom:6px">
          <input type="text" id="editRegFVenc" placeholder="Fecha de vencimiento (DD/MM/AAAA)" style="margin-bottom:6px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:8px 0 4px 2px">Datos de contacto y empadronamiento</p>
          <input type="text" id="editRegDir" placeholder="Dirección / Comunidad" style="margin-bottom:6px">
          <input type="text" id="editRegTel" inputmode="numeric" placeholder="Teléfono" style="margin-bottom:6px">
          <select id="editRegEmp" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:6px">
            <option value="SI">Empadronado: SÍ</option>
            <option value="NO">Empadronado: NO</option>
          </select>
          <input type="text" id="editRegNoEmp" placeholder="No. de Empadronamiento" style="margin-bottom:6px">
          <div style="display:flex;gap:8px">
            <button class="mbtn" style="background:#27ae60;margin:0;flex:1" onclick="guardarEditRegistro()">✅ Guardar</button>
            <button class="mbtn" style="background:#566573;margin:0;flex:1" onclick="cancelarEditRegistro()">✖ Cancelar</button>
          </div>
        </div>
        <button class="mbtn" style="background:#2980b9" id="btnEditReg" onclick="abrirEditRegistro()">✏️ Editar registro</button>
        <button class="mbtn" style="background:#922b21" onclick="elimReg()">🗑️ Eliminar registro</button>
      </div>
    </div>

    <div id="tabPrueba" style="display:none">
      <div style="background:#1c2833;border:2px solid #8e44ad;border-radius:10px;padding:12px">
        <h3 style="color:#bb8fce;margin:0 0 8px 0;font-size:0.95em">🧪 Agregar registro de prueba</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Solo para hacer pruebas. Se guarda ÚNICAMENTE en la hoja DPI (no crea grupo ni presidenta), marcado como Jefe="PRUEBA" para poder identificarlo y borrarlo fácil después.</p>
        <input type="file" id="tpF" accept="image/*" style="display:none">
        <input type="file" id="tpR" accept="image/*" style="display:none">
        <div id="pasoTP">
          <button class="mbtn" style="background:#8e44ad" id="btnTpF" onclick="abrirSelectorFoto('pruebaF')">📷 Foto/Archivo Frente</button>
          <button class="mbtn" style="background:#8e44ad" id="btnTpR" onclick="abrirSelectorFoto('pruebaR')">📷 Foto/Archivo Reverso</button>
          <input type="text" id="tpTel" inputmode="numeric" placeholder="Teléfono (opcional)">
          <input type="text" id="tpDir" placeholder="Dirección / Comunidad">
          <button class="mbtn" style="background:#566573" id="btnTpProc" disabled onclick="procRegistroPrueba()">Analizar y guardar como prueba</button>
        </div>
        <p id="msgTP" class="msg"></p>
        <button class="mbtn" style="background:#6c3483;margin-top:8px" onclick="verRegistrosPrueba()">👁️ Ver registros de prueba ya ingresados</button>
        <div id="listaRegPrueba" style="margin-top:8px;max-height:40vh;overflow-y:auto"></div>
      </div>
    </div>

    <div id="tabAfil" style="display:none">
      <input type="text" id="mDPIAfil" inputmode="numeric" placeholder="🔍 Buscar afiliado por DPI">
      <button class="mbtn" style="background:#7d3c98" onclick="buscarAfiliadoMant()">Buscar afiliado</button>
      <div id="resAfil" style="display:none">
        <div id="datAfil" style="background:#154360;border-radius:8px;padding:10px;margin:8px 0;color:white;font-size:0.85em"></div>
        <div id="editAfilForm" style="display:none;background:#0d2137;border-radius:8px;padding:10px;margin-bottom:8px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:0 0 4px 2px">CUI (DPI)</p>
          <input type="text" id="editAfilCUI" inputmode="numeric" placeholder="Número de CUI/DPI (13 dígitos)" style="margin-bottom:6px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:8px 0 4px 2px">Nombres y apellidos</p>
          <input type="text" id="editAfilPNom" placeholder="Primer nombre" style="margin-bottom:6px">
          <input type="text" id="editAfilSNom" placeholder="Segundo nombre" style="margin-bottom:6px">
          <input type="text" id="editAfilPApe" placeholder="Primer apellido" style="margin-bottom:6px">
          <input type="text" id="editAfilSApe" placeholder="Segundo apellido" style="margin-bottom:6px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:8px 0 4px 2px">Datos del DPI</p>
          <select id="editAfilSexo" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:6px">
            <option value="">Sexo: --</option>
            <option value="M">Sexo: Masculino</option>
            <option value="F">Sexo: Femenino</option>
          </select>
          <input type="text" id="editAfilEC" placeholder="Estado civil" style="margin-bottom:6px">
          <input type="text" id="editAfilFNac" placeholder="Fecha de nacimiento (DD/MM/AAAA)" style="margin-bottom:6px">
          <input type="text" id="editAfilMunNac" placeholder="Municipio de nacimiento" style="margin-bottom:6px">
          <input type="text" id="editAfilDepNac" placeholder="Departamento de nacimiento" style="margin-bottom:6px">
          <input type="text" id="editAfilMunVec" placeholder="Municipio de vecindad" style="margin-bottom:6px">
          <input type="text" id="editAfilDepVec" placeholder="Departamento de vecindad" style="margin-bottom:6px">
          <input type="text" id="editAfilFExp" placeholder="Fecha de expedición (DD/MM/AAAA)" style="margin-bottom:6px">
          <input type="text" id="editAfilFVenc" placeholder="Fecha de vencimiento (DD/MM/AAAA)" style="margin-bottom:6px">
          <p style="color:#7fb3d5;font-size:0.75em;margin:8px 0 4px 2px">Datos de contacto y empadronamiento</p>
          <input type="text" id="editAfilDir" placeholder="Dirección / Comunidad" style="margin-bottom:6px">
          <input type="text" id="editAfilTel" inputmode="numeric" placeholder="Teléfono" style="margin-bottom:6px">
          <select id="editAfilEmp" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:6px">
            <option value="SI">Empadronado: SÍ</option>
            <option value="NO">Empadronado: NO</option>
          </select>
          <input type="text" id="editAfilNoEmp" placeholder="No. de Empadronamiento" style="margin-bottom:6px">
          <div style="display:flex;gap:8px">
            <button class="mbtn" style="background:#27ae60;margin:0;flex:1" onclick="guardarEditAfiliado()">✅ Guardar</button>
            <button class="mbtn" style="background:#566573;margin:0;flex:1" onclick="cancelarEditAfiliado()">✖ Cancelar</button>
          </div>
        </div>
        <button class="mbtn" style="background:#2980b9" id="btnEditAfil" onclick="abrirEditAfiliado()">✏️ Modificar afiliado</button>
        <button class="mbtn" style="background:#922b21" onclick="elimAfiliado()">🗑️ Eliminar afiliado</button>
      </div>
    </div>

    <div id="tabGrp" style="display:none">
      <input type="text" id="mGN" placeholder="🔍 Nombre del grupo">
      <input type="text" id="mGD" inputmode="numeric" placeholder="O DPI coordinadora">
      <button class="mbtn" style="background:#2980b9" onclick="buscarGrpM()">Buscar grupo</button>
      <div id="resGrp" style="display:none">
        <div id="datGrp" style="background:#154360;border-radius:8px;padding:10px;margin:8px 0;color:white;font-size:0.85em"></div>
        <div id="editGrpForm" style="display:none;background:#0d2137;border-radius:8px;padding:10px;margin-bottom:8px">
          <input type="text" id="editGrpNombre" placeholder="Nuevo nombre del grupo" style="margin-bottom:6px">
          <div style="display:flex;gap:8px">
            <button class="mbtn" style="background:#27ae60;margin:0;flex:1" onclick="guardarEditGrupo()">✅ Guardar</button>
            <button class="mbtn" style="background:#566573;margin:0;flex:1" onclick="cancelarEditGrupo()">✖ Cancelar</button>
          </div>
        </div>
        <button class="mbtn" style="background:#2980b9" onclick="abrirEditGrupo()">✏️ Editar nombre</button>
        <button class="mbtn" style="background:#922b21" onclick="elimGrp()">🗑️ Eliminar grupo</button>
      </div>
    </div>

    <div id="tabPres" style="display:none">
      <input type="text" id="mPresN" placeholder="🔍 Buscar por nombre">
      <input type="text" id="mPresD" inputmode="numeric" placeholder="O buscar por DPI">
      <button class="mbtn" style="background:#2980b9" onclick="buscarPresM()">Buscar presidenta</button>
      <div id="resPres" style="display:none">
        <div id="datPres" style="background:#154360;border-radius:8px;padding:10px;margin:8px 0;color:white;font-size:0.85em"></div>
        <button class="mbtn" style="background:#922b21" onclick="elimPres()">🗑️ Eliminar presidenta</button>
      </div>
    </div>

    <div id="tabJef" style="display:none">
      <div style="background:#1c2833;border:2px solid #1a5276;border-radius:10px;padding:12px;margin-bottom:14px">
        <h3 style="color:#5dade2;margin:0 0 8px 0;font-size:0.95em">📧 Reporte diario automático — Correos generales</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Estos correos (hasta 3) reciben TODOS los días a las 8am el listado COMPLETO de personas sin teléfono, de TODOS los jefes de sector juntos.</p>
        <input type="email" id="cfgEmailGral1" placeholder="Correo 1">
        <input type="email" id="cfgEmailGral2" placeholder="Correo 2 (opcional)">
        <input type="email" id="cfgEmailGral3" placeholder="Correo 3 (opcional)">
        <button class="mbtn" style="background:#27ae60" onclick="guardarCorreosGenerales()">💾 Guardar correos generales</button>
      </div>

      <input type="text" id="mJefeNuevo" placeholder="Nombre del nuevo jefe de sector">
      <input type="text" id="mJefePinNuevo" inputmode="numeric" placeholder="PIN (mínimo 4 dígitos)">
      <input type="text" id="mJefeTelNuevo" inputmode="numeric" placeholder="Teléfono (opcional)">
      <input type="text" id="mJefeComunidadNuevo" placeholder="Comunidad a la que pertenece (opcional)">
      <p style="color:#7fb3d5;font-size:0.72em;margin:2px 0 4px 2px">Correos de este jefe (recibe SOLO su propia gente, hasta 3):</p>
      <input type="email" id="mJefeEmail1Nuevo" placeholder="Correo 1 (opcional)">
      <input type="email" id="mJefeEmail2Nuevo" placeholder="Correo 2 (opcional)">
      <input type="email" id="mJefeEmail3Nuevo" placeholder="Correo 3 (opcional)">
      <button class="mbtn" style="background:#27ae60" onclick="agregarJefeMant()">➕ Agregar jefe de sector</button>
      <div id="listaJefesMant" style="margin-top:10px;max-height:45vh;overflow-y:auto">Cargando...</div>
    </div>

    <div id="tabAud" style="display:none">
      <p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">Últimas 300 acciones registradas (más reciente primero)</p>
      <button class="mbtn" style="background:#1a5276;margin-bottom:10px" onclick="abrirEmpadronamientosActualizados()">📄 PDF: DPI con Empadronamiento Actualizado</button>
      <div id="listaAuditoria" style="max-height:55vh;overflow-y:auto">Cargando...</div>
    </div>

    <div id="tabBor" style="display:none">
      <div style="background:#1c2833;border:2px solid #1a5276;border-radius:10px;padding:12px;margin-bottom:14px">
        <h3 style="color:#5dade2;margin:0 0 8px 0;font-size:0.95em">🧹 Quitar espacios y guiones de los DPI</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Revisa TODAS las hojas (DPI, Grupos, Presidentas, Padrón) y corrige los CUI que quedaron guardados con espacios o guiones (ej. "3208 87731 0801" → "3208877310801"). No borra ni cambia ningún otro dato.</p>
        <button class="mbtn" style="background:#1a5276" onclick="limpiarEspaciosCuis()">🧹 Corregir DPI con espacios ahora</button>
      </div>

      <div style="background:#1c2833;border:2px solid #e67e22;border-radius:10px;padding:12px;margin-bottom:14px">
        <h3 style="color:#f5b041;margin:0 0 8px 0;font-size:0.95em">🗑️ Borrar todo lo de UN jefe de sector</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Borra sus Grupos, Presidentas y registros de DPI. El jefe sigue pudiendo iniciar sesión (queda como si empezara de cero). No afecta a los demás jefes.</p>
        <select id="borJefeSel" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:8px"><option value="">-- Seleccione un jefe --</option></select>
        <button class="mbtn" style="background:#e67e22" onclick="borrarDatosJefe()">🗑️ Borrar todo lo de este jefe</button>
      </div>

      <div style="background:#1c2833;border:2px solid #922b21;border-radius:10px;padding:12px">
        <h3 style="color:#e74c3c;margin:0 0 8px 0;font-size:0.95em">⚠️ Reinicio general (TODOS los jefes)</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Borra TODOS los Grupos, Presidentas y registros de DPI, de todos los jefes de sector. No se puede deshacer desde la app. No afecta la lista de Jefes de Sector ni la Auditoría.</p>
        <p style="color:#f5b041;font-size:0.75em;margin-bottom:6px">Para confirmar, escribe exactamente: <b>REINICIAR TODO</b></p>
        <input type="text" id="reinicioConfirmTexto" placeholder="Escribe aquí para confirmar" style="margin-bottom:8px">
        <button class="mbtn" style="background:#922b21" onclick="reinicioGeneralMant()">⚠️ Reiniciar TODO desde cero</button>
      </div>
    </div>

    <div id="tabPad" style="display:none">
      <div style="background:#1c2833;border:2px solid #1a5276;border-radius:10px;padding:12px;margin-bottom:14px">
        <h3 style="color:#5dade2;margin:0 0 8px 0;font-size:0.95em">🔄 Revisar DPI contra Padrón</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Vuelve a comparar a TODAS las personas marcadas como "No empadronado" contra la hoja PADRON. Si alguna ya aparece ahí (por ejemplo, porque el Padrón se actualizó después de haberla registrado), la marca como empadronada automáticamente — en DPI y también en su propia fila de Grupos/Presidentas si la tiene.</p>
        <p style="color:#f5b041;font-size:0.75em;margin-bottom:8px">Solo sube a alguien de NO a SI. Nunca baja a nadie de SI a NO, por seguridad.</p>
        <button class="mbtn" style="background:#1a5276" onclick="revisarPadronMant()">🔄 Revisar ahora</button>
      </div>
      <div id="resPadronMant" style="max-height:38vh;overflow-y:auto"></div>

      <div style="background:#1c2833;border:2px solid #8e44ad;border-radius:10px;padding:12px;margin-top:16px">
        <h3 style="color:#bb8fce;margin:0 0 8px 0;font-size:0.95em">🔧 Corregir Dirección/Comunidad (masivo)</h3>
        <p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">Para cuando varios registros quedaron con una dirección incompleta o mal escrita (ej. "Chocanac" en vez de "Chocanac Aldea Vasquez"). Busca coincidencias EXACTAS (no cambia direcciones parecidas) en las hojas DPI, Grupos y Presidentas.</p>
        <input type="text" id="corrDirActual" placeholder="Texto actual (ej. Chocanac)">
        <input type="text" id="corrDirNuevo" placeholder="Texto correcto (ej. Chocanac Aldea Vasquez)">
        <button class="mbtn" style="background:#8e44ad" onclick="verVistaPreviaCorreccionDir()">🔍 Ver cuántos coinciden (vista previa)</button>
        <div id="resCorrDir" style="margin-top:8px"></div>
      </div>
    </div>

    <p id="msgM" class="msg"></p>
  </div>
  <button class="mbtn" style="background:#c0392b;margin-top:15px" onclick="hide('oMant');clv=''">Cerrar</button>
</div>
</div>

<div class="overlay" id="oEmpActualizados">
<div class="box">
  <h2>📄 Empadronamientos Actualizados</h2>
  <p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">Cambios hechos desde "Editar Empadronamiento" en Registros Individuales</p>
  <div id="listaEmpActualizados" style="max-height:50vh;overflow-y:auto">Cargando...</div>
  <div id="botonesFijosEmpActualizados" style="margin-top:8px"></div>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oEmpActualizados');show('oMant')">← Regresar</button>
</div>
</div>

<div class="overlay" id="oRegistros">
<div class="box">
  <h2>👤 Registros Individuales</h2>
  <div id="contReg" class="msg" style="max-height:38vh;overflow-y:auto">Cargando...</div>
  <button class="mbtn" style="background:#2980b9;margin-top:10px" onclick="hide('oRegistros');show('oEditTel');document.getElementById('busEditTel').value='';document.getElementById('resEditTel').style.display='none';msg('msgEditTel','','')">✏️ Editar teléfono</button>
  <button class="mbtn" style="background:#1a5276" onclick="hide('oRegistros');show('oEditEmp');document.getElementById('busEditEmp').value='';document.getElementById('resEditEmp').style.display='none';document.getElementById('contPendientesEmp').style.display='none';msg('msgEditEmp','','')">🗳️ Editar Empadronamiento</button>
  <button class="mbtn" style="background:#c0392b" onclick="cerrarRegistrosInd()">Cerrar</button>
</div>
</div>

<div class="overlay" id="oVerGrupos">
<div class="box">
  <h2>👥 Registros de Grupos</h2>
  <div id="jGrid" class="jgrid"></div>
  <div id="contGrupos" class="msg">Seleccione un Jefe de Sector</div>
  <button class="mbtn" style="background:#c0392b;margin-top:15px" onclick="hide('oVerGrupos')">Cerrar</button>
</div>
</div>

<div class="overlay" id="oNuevoG1">
<div class="box">
  <h2>➕ Nuevo Grupo</h2>
  <p style="color:#aed6f1;margin-bottom:6px">Nombre del grupo:</p>
  <input type="text" id="ngN" placeholder="Ej: Grupo Sector 1">
  <p style="color:#aed6f1;margin-bottom:6px">Jefe de Sector:</p>
  <p id="ngJDisplay" style="background:#154360;border-radius:8px;padding:10px;color:white;font-weight:bold;margin:0 0 10px 0">-</p>
  <input type="hidden" id="ngJ" value="">
  <p style="color:#aed6f1;margin-bottom:4px;margin-top:8px">Presidenta de Comité:</p>
  <select id="ngP" style="margin-bottom:4px" onchange="selPresidenta(this)">
    <option value="">-- Seleccione Jefe primero --</option>
  </select>
  <button class="mbtn" style="background:#8e44ad;padding:8px" onclick="hide('oNuevoG1');show('oPresidenta')">➕ Registrar nueva Presidenta</button>
  <button class="mbtn" style="background:#27ae60" onclick="irCoord()">Siguiente →</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oNuevoG1')">Cancelar</button>
</div>
</div>

<!-- OVERLAY REGISTRAR PRESIDENTA -->
<div class="overlay" id="oPresidenta">
<div class="box">
  <h2>👩 Registrar Presidenta</h2>
  <div id="infoPresidenta" style="background:#154360;border-radius:10px;padding:10px;margin-bottom:15px;color:#aed6f1;font-size:0.9em;text-align:center">Escanee DPI de la Presidenta</div>
  <input type="file" id="prF" accept="image/*" style="display:none">
  <input type="file" id="prR" accept="image/*" style="display:none">
  <button class="mbtn" style="background:#2980b9" id="bFpr" onclick="abrirSelectorFoto('presF')">📷 Foto Frente</button>
  <button class="mbtn" style="background:#2980b9" id="bRpr" onclick="abrirSelectorFoto('presR')">📷 Foto Reverso</button>
  <input type="text" id="prTel" inputmode="numeric" placeholder="Teléfono Presidenta">
  <input type="text" id="prDir" placeholder="Dirección / Comunidad">
  <button class="mbtn" style="background:#566573" id="bProcPr" disabled onclick="procPresidenta()">Procesar DPI Presidenta</button>
  <p id="msgPr" class="msg"></p>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oPresidenta');show('oNuevoG1')">← Volver</button>
</div>
</div>

<!-- OVERLAY CONFIRMAR PRESIDENTA -->
<div class="overlay" id="oConfPr">
<div class="box">
  <h2>✅ Presidenta Registrada</h2>
  <div id="datPr" style="background:#154360;border-radius:10px;padding:15px;margin-bottom:15px;color:white;font-size:0.9em"></div>
  <p id="msgConfPr" class="msg"></p>
  <button class="mbtn" style="background:#27ae60" onclick="confirmarPresidenta()">✅ Confirmar y continuar</button>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oConfPr');show('oNuevoG1')">❌ Cancelar</button>
</div>
</div>

<div class="overlay" id="oCoord">
<div class="box">
  <h2>📷 DPI Coordinadora</h2>
  <div id="infoCoord" style="background:#154360;border-radius:10px;padding:10px;margin-bottom:15px;color:#aed6f1;font-size:0.9em;text-align:center"></div>
  <input type="file" id="fcF" accept="image/*" style="display:none">
  <input type="file" id="fcR" accept="image/*" style="display:none">
  <button class="mbtn" style="background:#2980b9" id="bFc" onclick="abrirSelectorFoto('coordF')">📷 Foto Frente</button>
  <button class="mbtn" style="background:#2980b9" id="bRc" onclick="abrirSelectorFoto('coordR')">📷 Foto Reverso</button>
  <input type="text" id="fcTel" inputmode="numeric" placeholder="Teléfono coordinadora">
  <input type="text" id="fcDir" placeholder="Dirección / Comunidad coordinadora">
  <button class="mbtn" style="background:#566573" id="bProcC" disabled onclick="procCoord()">Procesar DPI Coordinadora</button>
  <p id="msgC" class="msg"></p>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oCoord')">Cancelar</button>
</div>
</div>

<div class="overlay" id="oConfC">
<div class="box">
  <h2>✅ Coordinadora Registrada</h2>
  <div id="datC" style="background:#154360;border-radius:10px;padding:15px;margin-bottom:15px;color:white;font-size:0.9em"></div>
  <p id="msgConfC" class="msg"></p>
  <button class="mbtn" style="background:#27ae60" onclick="continuarPersonas()">✅ Continuar agregando</button>
  <button class="mbtn" style="background:#8e44ad" onclick="pdfSoloCoord()">📄 PDF solo coordinadora</button>
  <button class="mbtn" style="background:#e67e22" onclick="hide('oConfC')">🚪 Salir</button>
</div>
</div>

<div class="overlay" id="oPersonas">
<div class="box">
  <h2>👥 Agregar Personas</h2>
  <div style="background:#154360;border-radius:10px;padding:10px;margin-bottom:15px;text-align:center">
    <p id="infoGP" style="color:#aed6f1;font-size:0.85em"></p>
    <p id="cntG" style="color:#2ecc71;font-size:1.6em;font-weight:bold">1/10</p>
  </div>
  <input type="file" id="fpF" accept="image/*" style="display:none">
  <input type="file" id="fpR" accept="image/*" style="display:none">
  <div id="pasoF">
    <button class="mbtn" style="background:#2980b9" id="bFp" onclick="abrirSelectorFoto('persF')">📷 Foto Frente</button>
    <button class="mbtn" style="background:#2980b9" id="bRp" onclick="abrirSelectorFoto('persR')">📷 Foto Reverso</button>
    <input type="text" id="fpTel" inputmode="numeric" placeholder="Teléfono (opcional)">
    <input type="text" id="fpDir" placeholder="Dirección / Comunidad">
    <button class="mbtn" style="background:#566573" id="bProcP" disabled onclick="procPersona()">Agregar al grupo</button>
  </div>
  <p id="msgP" class="msg"></p>
  <div id="listaG" style="max-height:180px;overflow-y:auto;margin-top:10px"></div>
  <button class="mbtn" style="background:#8e44ad;margin-top:10px;display:none" id="bPDF" onclick="genPDF()">📄 Generar PDF</button>
  <button class="mbtn" style="background:#c0392b;margin-top:10px" onclick="hide('oPersonas')">Cerrar</button>
</div>
</div>

<div class="overlay" id="oBuscarG">
<div class="box">
  <h2>🔍 Buscar Grupo</h2>
  <input type="text" id="bgN" placeholder="Nombre del grupo">
  <input type="text" id="bgD" inputmode="numeric" placeholder="O DPI coordinadora">
  <button class="mbtn" style="background:#2980b9" onclick="buscarGE()">🔍 Buscar</button>
  <div id="resBG" style="display:none">
    <div id="datBG" style="background:#154360;border-radius:8px;padding:10px;margin:8px 0;color:white;font-size:0.9em"></div>
    <button class="mbtn" style="background:#27ae60" id="bContG" onclick="contGrupoE()">➕ Continuar</button>
    <button class="mbtn" style="background:#8e44ad" onclick="pdfGrupoE()">📄 PDF</button>
  </div>
  <p id="msgBG" class="msg"></p>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oBuscarG')">Cerrar</button>
</div>
</div>

<div class="overlay" id="oPrevia">
<div class="box">
  <h2>📋 Vista Previa</h2>
  <div id="camposP"></div>
  <div id="msgDup" style="display:none;background:#e74c3c;color:white;padding:10px;border-radius:8px;text-align:center;margin-top:10px">⚠️ DPI ya registrado</div>
  <div style="display:flex;gap:10px;margin-top:15px">
    <button class="mbtn" style="flex:1;background:#27ae60" onclick="confirmarDPI()">✅ Guardar</button>
    <button class="mbtn" style="flex:1;background:#c0392b" onclick="hide('oPrevia')">❌ Cancelar</button>
  </div>
</div>
</div>

<!-- OVERLAY VISTA PREVIA PDF -->
<div class="overlay" id="oPDFViewer" style="z-index:9999">
<div class="box" style="padding:10px">
  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px">
    <h2 style="font-size:1em;margin:0">📄 Vista Previa</h2>
    <div style="display:flex;gap:8px">
      <button id="btnDescPDF" style="background:#27ae60;color:white;border:none;border-radius:8px;padding:8px 12px;font-weight:bold;cursor:pointer">⬇️ Descargar</button>
      <button onclick="hide('oPDFViewer');document.getElementById('pdfFrame').style.display='block'" style="background:#c0392b;color:white;border:none;border-radius:8px;padding:8px 12px;font-weight:bold;cursor:pointer">✖ Cerrar</button>
    </div>
  </div>
  <div id="pdfLoadMsg" style="text-align:center;padding:20px;color:#aed6f1;font-size:0.9em">
    <p>⏳ Generando ficha...</p>
  </div>
  <div id="pdfReadyMsg" style="display:none;text-align:center;padding:20px">
    <p style="color:#2ecc71;font-size:0.95em;margin-bottom:15px">✅ Ficha generada correctamente</p>
    <button id="pdfAbrirBtn" style="width:100%;padding:15px;background:#2980b9;color:white;border:none;border-radius:10px;font-size:1em;font-weight:bold;cursor:pointer;margin-bottom:10px">
      📄 Ver PDF
    </button>
    <p style="color:#aed6f1;font-size:0.8em">O usa el botón Descargar para guardar</p>
  </div>
  <iframe id="pdfFrame" style="width:100%;height:65vh;border:none;border-radius:8px;background:white;display:none"></iframe>
</div>
</div>

<!-- OVERLAY EDITAR TELÉFONO -->
<div class="overlay" id="oEditTel">
<div class="box">
  <h2>📞 Editar Teléfono</h2>
  <p style="color:#aed6f1;font-size:0.85em;margin-bottom:10px;text-align:center">Busca la persona por nombre o CUI</p>
  <input type="text" id="busEditTel" placeholder="🔍 Nombre o CUI..." style="margin-bottom:6px">
  <button class="mbtn" style="background:#2980b9" onclick="buscarParaEditarTel()">Buscar</button>
  <div id="resEditTel" style="display:none;margin-top:10px">
    <div id="datEditTel" style="background:#154360;border-radius:8px;padding:10px;margin-bottom:10px;color:white;font-size:0.85em"></div>
    <p style="color:#aed6f1;font-size:0.82em;margin-bottom:6px">Nuevo teléfono:</p>
    <input type="tel" id="nuevoTel" placeholder="+502..." inputmode="numeric">
    <button class="mbtn" style="background:#27ae60" onclick="guardarNuevoTel()">💾 Guardar teléfono</button>
  </div>
  <p id="msgEditTel" class="msg"></p>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oEditTel');show('oRegistros')">Cerrar</button>
</div>
</div>

<div class="overlay" id="oEditEmp">
<div class="box">
  <h2>🗳️ Editar Empadronamiento</h2>
  <p style="color:#aed6f1;font-size:0.85em;margin-bottom:10px;text-align:center">Busca la persona por su número de CUI</p>
  <input type="text" id="busEditEmp" inputmode="numeric" placeholder="🔍 Número de CUI/DPI..." style="margin-bottom:6px">
  <button class="mbtn" style="background:#1a5276" onclick="buscarParaEditarEmp()">Buscar</button>
  <div id="resEditEmp" style="display:none;margin-top:10px">
    <div id="datEditEmp" style="background:#154360;border-radius:8px;padding:10px;margin-bottom:10px;color:white;font-size:0.85em"></div>
    <select id="nuevoEmp" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:6px">
      <option value="SI">Empadronado: SÍ</option>
      <option value="NO">Empadronado: NO</option>
    </select>
    <input type="text" id="nuevoNumEmp" placeholder="No. de Empadronamiento" style="margin-bottom:6px">
    <p style="color:#7fb3d5;font-size:0.75em;margin:4px 0 4px 2px">Observaciones (ej. no vota en este municipio, etc.)</p>
    <textarea id="nuevaObsEmp" placeholder="Observaciones..." rows="2" style="width:100%;padding:8px;border-radius:8px;border:none;margin-bottom:6px;font-family:inherit;resize:vertical"></textarea>
    <button class="mbtn" style="background:#27ae60" onclick="guardarNuevoEmp()">💾 Guardar empadronamiento</button>
    <button class="mbtn" style="background:#16a085" onclick="marcarComoRevisado()">✅ Ya lo revisé hoy (Padrón/TSE)</button>
    <p id="msgRevisado" style="color:#7fb3d5;font-size:0.72em;margin:4px 0 0 2px"></p>
  </div>
  <button class="mbtn" style="background:#5dade2;margin-top:10px" onclick="verPendientesEditarEmp()">📋 Ver los que faltan por empadronar</button>
  <div id="contPendientesEmp" style="display:none;margin-top:8px"></div>
  <p id="msgEditEmp" class="msg"></p>
  <button class="mbtn" style="background:#c0392b" onclick="hide('oEditEmp');show('oRegistros')">Cerrar</button>
</div>
</div>

<!-- OVERLAY ADVERTENCIA DPI NO AUTENTICO -->
<div class="overlay" id="oAdvertenciaDPI">
<div class="box">
  <h2 style="color:#f39c12">⚠️ Advertencia</h2>
  <p style="color:white;font-weight:bold;text-align:center;margin-bottom:10px">SE DETECTÓ QUE NO ES UN DPI Y PUEDE SER QUE LA INFORMACIÓN NO SEA REAL</p>
  <p id="motivoAdvertenciaDPI" style="color:#f5b7b1;text-align:center;font-size:0.85em;margin-bottom:15px"></p>
  <button class="mbtn" style="background:#e67e22" onclick="_continuarPeseAdvertencia()">Continuar de todos modos</button>
  <button class="mbtn" style="background:#c0392b" onclick="_salirPorAdvertencia()">❌ Salir / Tomar otra foto</button>
</div>
</div>

<!-- OVERLAY CONSULTA TSE EMPADRONAMIENTO -->
<div class="overlay" id="oTSE">
<div class="box">
  <h2>🗳️ No Empadronado</h2>
  <div style="background:#154360;border-radius:10px;padding:12px;margin-bottom:12px">
    <p id="tseNombre" style="color:#2ecc71;font-weight:bold;font-size:0.95em;margin-bottom:4px"></p>
    <p id="tseCUI" style="color:#aed6f1;font-size:0.85em"></p>
  </div>
  <p style="color:#f0b27a;font-size:0.85em;margin-bottom:10px;text-align:center">Esta persona no está en el padrón local.<br>Puedes verificar en el TSE:</p>
  <button class="mbtn" style="background:#1a5276" onclick="abrirTSE()">🌐 Abrir consulta TSE</button>
  <div style="margin-top:10px">
    <p style="color:#aed6f1;font-size:0.82em;margin-bottom:6px">Después de consultar, ingresa el No. de empadronamiento:</p>
    <input type="text" id="tseNumEmp" inputmode="numeric" placeholder="No. de empadronamiento (si tiene)">
    <button class="mbtn" style="background:#27ae60" onclick="confirmarTSE('SI')">✅ Está empadronado</button>
    <button class="mbtn" style="background:#922b21" onclick="confirmarTSE('NO')">❌ No está empadronado</button>
  </div>
  <p id="msgTSE" class="msg"></p>
</div>
</div>

<div class="overlay" id="oExito">
<div class="box" style="text-align:center">
  <h2 style="color:#2ecc71">✅ ¡Guardado!</h2>
  <p id="nomG" style="color:white;margin:15px 0;font-size:1.1em"></p>
  <button class="mbtn" style="background:#27ae60" onclick="hide('oExito');resetDPI()">➕ Agregar otro</button>
  <button class="mbtn" style="background:#2980b9" onclick="hide('oExito')">🏠 Inicio</button>
</div>
</div>

<script>

var gN='',gJ='',gC=null,gPs=[],gCF=null,gCR=null,gPF=null,gPR=null;
var dpiD=null,gE=null,mFila=null,mGN='',mGJefe='',clv='',gPres=null,gPresF=null,gPresR=null;
var gTPF=null,gTPR=null;
var gAfF=null,gAfR=null;
var _ultimoAfiliadoGuardado = null;
var _ultimaFotoFrenteAfiliado = null;
var tG={},rJ={};

function show(id){ var e=document.getElementById(id); if(e) e.style.display='block'; }
function hide(id){ var e=document.getElementById(id); if(e) e.style.display='none'; }
function msg(id,t,c){ var e=document.getElementById(id); if(e){ e.textContent=t; e.style.color=c||'#aed6f1'; } }
// Convierte una foto a base64, pero antes la REDIMENSIONA y COMPRIME con un
// <canvas>. Las fotos de cámara sin tocar pueden pesar varios megabytes
// cada una; con varias personas subiendo fotos al mismo tiempo, eso agota
// rápido la memoria del servidor (plan gratis, con RAM limitada). Reducir
// el tamaño aquí no afecta la lectura del DPI (sigue siendo perfectamente
// legible), pero baja el peso enviado en un 80-95%.
function b64(f){
  return new Promise(function(resolve, reject){
    var MAX_DIM = 1600; // suficiente para leer bien el DPI, mucho más liviano que el original
    var reader = new FileReader();
    reader.onload = function(e){
      var img = new Image();
      img.onload = function(){
        var w = img.width, h = img.height;
        if(w > h && w > MAX_DIM){ h = Math.round(h * MAX_DIM / w); w = MAX_DIM; }
        else if(h >= w && h > MAX_DIM){ w = Math.round(w * MAX_DIM / h); h = MAX_DIM; }
        var canvas = document.createElement('canvas');
        canvas.width = w; canvas.height = h;
        var ctx = canvas.getContext('2d');
        ctx.drawImage(img, 0, 0, w, h);
        var dataUrl;
        try{ dataUrl = canvas.toDataURL('image/jpeg', 0.75); }
        catch(err){ dataUrl = e.target.result; } // si algo falla, usar la original sin comprimir
        resolve(dataUrl.split(',')[1]);
      };
      img.onerror = function(){
        // Si no se pudo cargar como imagen (raro), usar el archivo original sin comprimir
        resolve(e.target.result.split(',')[1]);
      };
      img.src = e.target.result;
    };
    reader.onerror = function(){ reject(new Error('No se pudo leer el archivo')); };
    reader.readAsDataURL(f);
  });
}

function updC(){
  var d='';
  for(var i=0;i<6;i++) d+=(i<clv.length?'●':'_')+(i<5?' ':'');
  document.getElementById('dispClave').textContent=d;
}

async function buscarPresM(){
  var nombre=document.getElementById('mPresN').value.trim();
  var dpi=document.getElementById('mPresD').value.trim();
  if(!nombre && !dpi){ alert('Ingrese nombre o DPI'); return; }
  try{
    var url='/buscar_presidenta?nombre='+encodeURIComponent(nombre)+'&dpi='+encodeURIComponent(dpi);
    var r=await fetch(url);
    var d=await r.json();
    if(d.ok && d.presidenta){
      var p=d.presidenta;
      document.getElementById('datPres').innerHTML='<b>Nombre:</b> '+p.nombre+'<br><b>DPI:</b> '+(p.cui||'-')+'<br><b>Jefe de Sector:</b> '+(p.jefe||'-')+'<br><b>Tel:</b> '+(p.telefono||'-');
      document.getElementById('datPres').setAttribute('data-fila', d.numero_fila||'');
      document.getElementById('datPres').setAttribute('data-nombre', p.nombre||'');
      document.getElementById('resPres').style.display='block';
      msg('msgM','','');
    } else {
      document.getElementById('resPres').style.display='none';
      msg('msgM','❌ Presidenta no encontrada','#e74c3c');
    }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function elimPres(){
  var el=document.getElementById('datPres');
  var fila=el.getAttribute('data-fila');
  var nombre=el.getAttribute('data-nombre');
  if(!confirm('¿Eliminar a '+nombre+' de la lista de Presidentas?')) return;
  try{
    var r=await fetch('/eliminar_presidenta',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({numero_fila:fila, nombre:nombre})});
    var d=await r.json();
    if(d.ok){ msg('msgM','✅ Presidenta eliminada','#27ae60'); document.getElementById('resPres').style.display='none'; document.getElementById('mPresN').value=''; document.getElementById('mPresD').value=''; }
    else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

function tabM(t){
  document.getElementById('tabInd').style.display=t==='ind'?'block':'none';
  document.getElementById('tabGrp').style.display=t==='grp'?'block':'none';
  document.getElementById('tabPres').style.display=t==='pres'?'block':'none';
  document.getElementById('tabJef').style.display=t==='jef'?'block':'none';
  document.getElementById('tabAud').style.display=t==='aud'?'block':'none';
  document.getElementById('tabBor').style.display=t==='bor'?'block':'none';
  document.getElementById('tabPad').style.display=t==='pad'?'block':'none';
  document.getElementById('tabPrueba').style.display=t==='prueba'?'block':'none';
  document.getElementById('tabAfil').style.display=t==='afil'?'block':'none';
  document.getElementById('tInd').style.background=t==='ind'?'#2980b9':'#566573';
  document.getElementById('tGrp').style.background=t==='grp'?'#2980b9':'#566573';
  document.getElementById('tPres').style.background=t==='pres'?'#2980b9':'#566573';
  document.getElementById('tJef').style.background=t==='jef'?'#2980b9':'#566573';
  document.getElementById('tAud').style.background=t==='aud'?'#2980b9':'#566573';
  document.getElementById('tBor').style.background=t==='bor'?'#2980b9':'#566573';
  document.getElementById('tPad').style.background=t==='pad'?'#2980b9':'#566573';
  document.getElementById('tPrueba').style.background=t==='prueba'?'#2980b9':'#566573';
  document.getElementById('tAfil').style.background=t==='afil'?'#2980b9':'#566573';
  if(t==='jef'){ cargarJefesMant(); cargarCorreosGenerales(); }
  if(t==='aud') cargarAuditoria();
  if(t==='bor') cargarJefesSelectBorrado();
  if(t==='prueba') verRegistrosPrueba();
}

async function cargarJefesSelectBorrado(){
  var jefes = await obtenerJefes(true);
  var sel = document.getElementById('borJefeSel');
  sel.innerHTML = '<option value="">-- Seleccione un jefe --</option>' +
    jefes.map(function(j){ return '<option>'+j+'</option>'; }).join('');
}

async function revisarPadronMant(){
  var cont = document.getElementById('resPadronMant');
  cont.innerHTML = '<p class="msg">Revisando, puede tardar un momento si hay muchos registros...</p>';
  try{
    var r = await fetch('/revisar_padron');
    var d = await r.json();
    if(d.ok){
      var h = '<p style="color:#2ecc71;font-weight:bold;font-size:0.9em;margin-bottom:8px">✅ '+d.total_actualizados+' persona(s) pasaron de NO a SI empadronado</p>';
      if(d.grupos_actualizados) h += '<p style="color:#aed6f1;font-size:0.75em;margin-bottom:4px">También se actualizaron '+d.grupos_actualizados+' fila(s) en Grupos</p>';
      if(d.presidentas_actualizadas) h += '<p style="color:#aed6f1;font-size:0.75em;margin-bottom:8px">También se actualizaron '+d.presidentas_actualizadas+' fila(s) en Presidentas</p>';
      if(d.nombres && d.nombres.length){
        h += '<div style="background:#154360;border-radius:8px;padding:8px">';
        d.nombres.forEach(function(n){ h += '<p style="color:white;font-size:0.8em;margin:2px 0">• '+n+'</p>'; });
        h += '</div>';
      }
      cont.innerHTML = h;
    } else { cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+(d.error||'Error')+'</p>'; }
  }catch(e){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+e.message+'</p>'; }
}

async function verVistaPreviaCorreccionDir(){
  var actual = document.getElementById('corrDirActual').value.trim();
  var nuevo = document.getElementById('corrDirNuevo').value.trim();
  var cont = document.getElementById('resCorrDir');
  if(!actual || !nuevo){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">Escribe el texto actual y el texto correcto</p>'; return; }
  cont.innerHTML = '<p class="msg">Buscando coincidencias...</p>';
  try{
    var r = await fetch('/vista_previa_correccion_direccion?texto_actual='+encodeURIComponent(actual));
    var d = await r.json();
    if(!d.ok){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+(d.error||'Error')+'</p>'; return; }
    if(d.total === 0){
      cont.innerHTML = '<p class="msg">No se encontró ningún registro con la dirección exacta "'+actual+'"</p>';
      return;
    }
    _corrDirActualVigente = actual;
    var h = '<p style="color:#f7dc6f;font-weight:bold;font-size:0.85em;margin-bottom:8px">Se encontraron '+d.total+' coincidencia(s) de "'+actual+'":</p>';
    h += '<div style="background:#154360;border-radius:8px;padding:8px;max-height:30vh;overflow-y:auto;margin-bottom:8px">';
    d.coincidencias.forEach(function(c){ h += '<p style="color:white;font-size:0.78em;margin:2px 0">• ['+c.hoja+'] '+(c.nombre||'-')+'</p>'; });
    if(d.total > d.coincidencias.length) h += '<p style="color:#aed6f1;font-size:0.75em;margin-top:4px">... y '+(d.total - d.coincidencias.length)+' más</p>';
    h += '</div>';
    h += '<button class="mbtn" style="background:#27ae60" id="btnAplicarCorrDir">✅ Cambiar los '+d.total+' registro(s)</button>';
    cont.innerHTML = h;
    document.getElementById('btnAplicarCorrDir').onclick = aplicarCorreccionDir;
  }catch(e){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+e.message+'</p>'; }
}

var _corrDirActualVigente = '';

async function aplicarCorreccionDir(){
  var actual = _corrDirActualVigente;
  var nuevo = document.getElementById('corrDirNuevo').value.trim();
  if(!confirm('¿Confirmas cambiar TODOS los registros con dirección exacta "'+actual+'" a "'+nuevo+'"? Esta acción no se puede deshacer desde la app.')) return;
  var cont = document.getElementById('resCorrDir');
  cont.innerHTML = '<p class="msg">Aplicando el cambio...</p>';
  try{
    var r = await fetch('/aplicar_correccion_direccion',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({texto_actual:actual, texto_nuevo:nuevo})});
    var d = await r.json();
    if(d.ok){
      var partes = [];
      for(var hoja in d.resumen){ partes.push(hoja+': '+d.resumen[hoja]); }
      cont.innerHTML = '<p style="color:#2ecc71;font-weight:bold;font-size:0.85em">✅ Corregido — '+partes.join(' | ')+'</p>';
      document.getElementById('corrDirActual').value=''; document.getElementById('corrDirNuevo').value='';
    } else { cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+(d.error||'Error')+'</p>'; }
  }catch(e){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+e.message+'</p>'; }
}

async function limpiarEspaciosCuis(){
  if(!confirm('¿Revisar todas las hojas (DPI, Grupos, Presidentas, Padrón) y corregir los CUI que tengan espacios o guiones? No se borra ni cambia ningún otro dato.')) return;
  msg('msgM','Revisando y corrigiendo, puede tardar un momento...','#aed6f1');
  try{
    var r = await fetch('/limpiar_cuis');
    var d = await r.json();
    if(d.ok){
      var partes = [];
      for(var hoja in d.resultados){
        if(hoja.endsWith('_error')) continue;
        partes.push(hoja+': '+d.resultados[hoja]);
      }
      var texto = '✅ Listo — '+partes.join(' | ');
      var huboErrores = Object.keys(d.resultados).some(function(k){ return k.endsWith('_error'); });
      if(huboErrores) texto += ' ⚠️ Algunas hojas tuvieron error: '+JSON.stringify(d.resultados);
      msg('msgM', texto, huboErrores?'#f0b27a':'#2ecc71');
    } else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function borrarDatosJefe(){
  var jefe = document.getElementById('borJefeSel').value;
  if(!jefe){ msg('msgM','Seleccione un jefe de sector','#e74c3c'); return; }
  if(!confirm('¿Seguro que quieres borrar TODOS los Grupos, Presidentas y registros de DPI de '+jefe+'?\\n\\nEsto no se puede deshacer desde la app. El jefe seguirá pudiendo iniciar sesión, solo quedará sin datos.')) return;
  msg('msgM','Borrando...','#aed6f1');
  try{
    var r = await fetch('/eliminar_datos_jefe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_jefe:jefe})});
    var d = await r.json();
    if(d.ok){
      var texto='✅ Borrado: '+d.resumen.GRUPOS+' filas de Grupos, '+d.resumen.PRESIDENTAS+' de Presidentas, '+d.resumen.DPI+' de DPI';
      if(d.errores) texto += ' ⚠️ Hubo errores: '+JSON.stringify(d.errores);
      msg('msgM', texto, d.errores?'#f0b27a':'#2ecc71');
    } else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function reinicioGeneralMant(){
  var texto = document.getElementById('reinicioConfirmTexto').value.trim();
  if(texto !== 'REINICIAR TODO'){ msg('msgM','Escribe exactamente "REINICIAR TODO" para confirmar','#e74c3c'); return; }
  if(!confirm('⚠️ ÚLTIMA CONFIRMACIÓN\\n\\nEsto borrará TODOS los Grupos, Presidentas y registros de DPI de TODOS los jefes de sector, sin excepción.\\n\\nNo se puede deshacer desde la app.\\n\\n¿Continuar?')) return;
  msg('msgM','Reiniciando todo...','#aed6f1');
  try{
    var r = await fetch('/reinicio_general',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({confirmacion:texto})});
    var d = await r.json();
    if(d.ok){
      var texto2='✅ Reinicio: '+d.resumen.GRUPOS+' filas de Grupos, '+d.resumen.PRESIDENTAS+' de Presidentas, '+d.resumen.DPI+' de DPI';
      if(d.errores) texto2 += ' ⚠️ Hubo errores: '+JSON.stringify(d.errores);
      msg('msgM', texto2, d.errores?'#f0b27a':'#2ecc71');
      document.getElementById('reinicioConfirmTexto').value='';
    } else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function cargarAuditoria(){
  document.getElementById('listaAuditoria').innerHTML='Cargando...';
  try{
    var r = await fetch('/ver_auditoria');
    var d = await r.json();
    if(!d.ok || !d.registros.length){ document.getElementById('listaAuditoria').innerHTML='<p class="msg">Sin registros todavía</p>'; return; }
    var h='';
    d.registros.forEach(function(reg){
      var fecha=reg[0]||'', jefe=reg[1]||'', accion=reg[2]||'', detalle=reg[3]||'';
      var esFallo = accion.toLowerCase().indexOf('fallido') >= 0;
      h+='<div style="background:#154360;border-radius:8px;padding:8px 10px;margin-bottom:6px">';
      h+='<p style="color:'+(esFallo?'#e74c3c':'#2ecc71')+';font-size:0.78em;font-weight:bold;margin:0">'+accion+'</p>';
      h+='<p style="color:white;font-size:0.8em;margin:2px 0">'+jefe+'</p>';
      if(detalle) h+='<p style="color:#aed6f1;font-size:0.75em;margin:0">'+detalle+'</p>';
      h+='<p style="color:#7f9bb5;font-size:0.7em;margin:2px 0 0 0">'+fecha+'</p>';
      h+='</div>';
    });
    document.getElementById('listaAuditoria').innerHTML=h;
  }catch(e){ document.getElementById('listaAuditoria').innerHTML='<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>'; }
}

async function cargarCorreosGenerales(){
  try{
    var r = await fetch('/config_correos_admin');
    var d = await r.json();
    if(d.ok && d.emails){
      document.getElementById('cfgEmailGral1').value = d.emails[0] || '';
      document.getElementById('cfgEmailGral2').value = d.emails[1] || '';
      document.getElementById('cfgEmailGral3').value = d.emails[2] || '';
    }
  }catch(e){}
}

async function guardarCorreosGenerales(){
  var emails = [
    document.getElementById('cfgEmailGral1').value.trim(),
    document.getElementById('cfgEmailGral2').value.trim(),
    document.getElementById('cfgEmailGral3').value.trim()
  ];
  try{
    var r = await fetch('/config_correos_admin',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({emails:emails})});
    var d = await r.json();
    if(d.ok){ msg('msgM','✅ Correos generales guardados','#2ecc71'); }
    else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function cargarJefesMant(){
  document.getElementById('listaJefesMant').innerHTML='Cargando...';
  var r = await fetch('/listar_jefes_detalle');
  var d = await r.json();
  var jefes = (d.ok && d.jefes) ? d.jefes : [];
  if(!jefes.length){ document.getElementById('listaJefesMant').innerHTML='<p class="msg">No hay jefes de sector registrados</p>'; return; }
  var h='';
  jefes.forEach(function(j){
    var jEnc = encodeURIComponent(j.nombre);
    var emailsTxt = (j.emails||[]).filter(function(e){return e;}).join(', ') || 'sin correo (no recibe el reporte diario)';
    h+='<div class="jefe-row" data-jefe-original="'+jEnc+'" style="background:#154360;border-radius:8px;padding:10px;margin-bottom:6px;display:flex;align-items:center;justify-content:space-between;gap:8px">';
    h+='<div style="flex:1;min-width:0"><p class="jefe-nombre" style="color:white;font-size:0.85em;margin:0">'+j.nombre+'</p>';
    h+='<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">'+(j.telefono||'sin teléfono')+'</p>';
    h+='<p style="color:#82e0aa;font-size:0.72em;margin:2px 0 0 0">📍 '+(j.comunidad||'sin comunidad asignada')+'</p>';
    h+='<p style="color:#f7dc6f;font-size:0.68em;margin:2px 0 0 0">📧 '+emailsTxt+'</p></div>';
    h+='<button data-jefe="'+jEnc+'" class="btnEditJefe" style="background:#2980b9;color:white;border:none;border-radius:8px;padding:6px 10px;font-size:0.78em;font-weight:bold;cursor:pointer">✏️</button>';
    h+='<button data-jefe="'+jEnc+'" class="btnDelJefe" style="background:#922b21;color:white;border:none;border-radius:8px;padding:6px 10px;font-size:0.78em;font-weight:bold;cursor:pointer">🗑️</button>';
    h+='</div>';
  });
  document.getElementById('listaJefesMant').innerHTML=h;
  document.getElementById('listaJefesMant').querySelectorAll('.btnDelJefe').forEach(function(btn){
    btn.onclick=async function(){
      var nombre=decodeURIComponent(this.getAttribute('data-jefe'));
      if(!confirm('¿Eliminar a '+nombre+' como jefe de sector?\\n\\nEsto NO borra los grupos ya registrados con ese jefe, solo evita que se pueda seleccionar en grupos nuevos.')) return;
      try{
        var r=await fetch('/eliminar_jefe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre:nombre})});
        var d=await r.json();
        if(d.ok){ msg('msgM','✅ Eliminado','#2ecc71'); cargarJefesMant(); }
        else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
      }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
    };
  });
  document.getElementById('listaJefesMant').querySelectorAll('.btnEditJefe').forEach(function(btn){
    btn.onclick=async function(){
      var nombreActual = decodeURIComponent(this.getAttribute('data-jefe'));
      var row = this.closest('.jefe-row');
      row.innerHTML = '<p style="color:#aed6f1;font-size:0.8em;margin:0">Cargando datos actuales...</p>';
      var pinActual = '', telActual = '', emailsActuales = ['','',''], comunidadActual = '';
      try{
        var rp = await fetch('/ver_pin_jefe?nombre='+encodeURIComponent(nombreActual));
        var dp = await rp.json();
        if(dp.ok){ pinActual = dp.pin || ''; telActual = dp.telefono || ''; emailsActuales = dp.emails || ['','','']; comunidadActual = dp.comunidad || ''; }
      }catch(e){}
      row.innerHTML =
        '<div style="flex:1;display:flex;flex-direction:column;gap:6px">' +
        '<input type="text" class="inputEditJefe" value="'+nombreActual.replace(/"/g,'&quot;')+'" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '<input type="text" class="inputEditJefePin" inputmode="numeric" value="'+pinActual.replace(/"/g,'&quot;')+'" placeholder="PIN" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '<input type="text" class="inputEditJefeTel" inputmode="numeric" value="'+telActual.replace(/"/g,'&quot;')+'" placeholder="Teléfono" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '<input type="text" class="inputEditJefeComunidad" value="'+comunidadActual.replace(/"/g,'&quot;')+'" placeholder="Comunidad a la que pertenece" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '<input type="email" class="inputEditJefeEmail1" value="'+(emailsActuales[0]||'').replace(/"/g,'&quot;')+'" placeholder="Correo 1" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '<input type="email" class="inputEditJefeEmail2" value="'+(emailsActuales[1]||'').replace(/"/g,'&quot;')+'" placeholder="Correo 2" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '<input type="email" class="inputEditJefeEmail3" value="'+(emailsActuales[2]||'').replace(/"/g,'&quot;')+'" placeholder="Correo 3" style="padding:6px 8px;border-radius:6px;border:none;font-size:0.85em">' +
        '</div>' +
        '<button class="btnGuardarJefe" style="background:#27ae60;color:white;border:none;border-radius:8px;padding:6px 10px;font-size:0.78em;font-weight:bold;cursor:pointer">✓</button>' +
        '<button class="btnCancelarJefe" style="background:#566573;color:white;border:none;border-radius:8px;padding:6px 10px;font-size:0.78em;font-weight:bold;cursor:pointer">✖</button>';
      var input = row.querySelector('.inputEditJefe');
      var inputPin = row.querySelector('.inputEditJefePin');
      var inputTel = row.querySelector('.inputEditJefeTel');
      var inputComunidad = row.querySelector('.inputEditJefeComunidad');
      var inputE1 = row.querySelector('.inputEditJefeEmail1');
      var inputE2 = row.querySelector('.inputEditJefeEmail2');
      var inputE3 = row.querySelector('.inputEditJefeEmail3');
      input.focus();
      row.querySelector('.btnCancelarJefe').onclick=function(){ cargarJefesMant(); };
      row.querySelector('.btnGuardarJefe').onclick=async function(){
        var nombreNuevo = input.value.trim();
        var pinNuevo = inputPin.value.trim();
        var telNuevo = inputTel.value.trim();
        var comunidadNueva = inputComunidad.value.trim();
        var emailsNuevos = [inputE1.value.trim(), inputE2.value.trim(), inputE3.value.trim()];
        if(!nombreNuevo){ msg('msgM','Escriba un nombre','#e74c3c'); return; }
        try{
          var r=await fetch('/editar_jefe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_actual:nombreActual, nombre_nuevo:nombreNuevo, pin_nuevo:pinNuevo, telefono_nuevo:telNuevo, emails_nuevos:emailsNuevos, comunidad_nueva:comunidadNueva})});
          var d=await r.json();
          if(d.ok){ msg('msgM','✅ Actualizado','#2ecc71'); cargarJefesMant(); }
          else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
        }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
      };
    };
  });
}

async function agregarJefeMant(){
  var nombre = document.getElementById('mJefeNuevo').value.trim();
  var pin = document.getElementById('mJefePinNuevo').value.trim();
  var telefono = document.getElementById('mJefeTelNuevo').value.trim();
  var comunidad = document.getElementById('mJefeComunidadNuevo').value.trim();
  var emails = [
    document.getElementById('mJefeEmail1Nuevo').value.trim(),
    document.getElementById('mJefeEmail2Nuevo').value.trim(),
    document.getElementById('mJefeEmail3Nuevo').value.trim()
  ];
  if(!nombre){ msg('msgM','Escriba un nombre','#e74c3c'); return; }
  if(!pin){ msg('msgM','Escriba un PIN (mínimo 4 dígitos)','#e74c3c'); return; }
  try{
    var r=await fetch('/agregar_jefe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre:nombre, pin:pin, telefono:telefono, emails:emails, comunidad:comunidad})});
    var d=await r.json();
    if(d.ok){
      document.getElementById('mJefeNuevo').value='';
      document.getElementById('mJefePinNuevo').value='';
      document.getElementById('mJefeTelNuevo').value='';
      document.getElementById('mJefeComunidadNuevo').value='';
      document.getElementById('mJefeEmail1Nuevo').value='';
      document.getElementById('mJefeEmail2Nuevo').value='';
      document.getElementById('mJefeEmail3Nuevo').value='';
      msg('msgM','✅ Agregado','#2ecc71'); cargarJefesMant();
    }
    else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

// REGISTROS
async function cargarRegistros(){
  msg('contReg','Cargando...','#aed6f1');
  show('oRegistros');
  try{
    var r=await fetch('/registros');
    var d=await r.json();
    if(d.ok&&d.filas.length>0){
      var esAdmin = sesionJefe && sesionJefe.esAdmin;
      var etiqueta = esAdmin ? 'todos los jefes' : (sesionJefe?sesionJefe.nombre:'');
      var h='<input type="text" placeholder="Buscar nombre o CUI..." oninput="filtrarRegs(this.value)" style="width:100%;padding:8px;border-radius:8px;border:none;background:#0d2137;color:white;font-size:0.9em;margin-bottom:8px">';
      h+='<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+d.filas.length+' registro(s) para '+etiqueta+'</p>';
      h+='<div id="listaRegs" style="max-height:55vh;overflow-y:auto">';
      d.filas.forEach(function(f){
        var cui=f[0]||'';
        var nombre=(f[2]||'')+' '+(f[3]||'')+' '+(f[4]||'')+' '+(f[5]||'');
        h+='<div class="reg-item" data-search="'+cui.toLowerCase()+' '+nombre.toLowerCase()+'" style="background:#154360;border-radius:8px;padding:10px;margin-bottom:6px;display:flex;align-items:center;justify-content:space-between;gap:8px">';
        h+='<div style="flex:1;min-width:0"><p style="color:white;font-size:0.82em;font-weight:bold;margin:0">'+nombre.trim()+'</p>';
        h+='<p style="color:#aed6f1;font-size:0.75em;margin:2px 0">CUI: '+cui+' '+(f[15]==='SI'?'SI':'NO')+'</p></div>';
        h+='<button class="btnPDF" data-cui="'+cui+'" style="padding:7px 10px;background:#8e44ad;color:white;border:none;border-radius:8px;font-size:0.78em;font-weight:bold;cursor:pointer">PDF</button>';
        h+='</div>';
      });
      h+='</div>';
      document.getElementById('contReg').innerHTML=h;
    } else { msg('contReg','No hay registros','#aed6f1'); }
  }catch(e){ msg('contReg','Error: '+e.message,'#e74c3c'); }
}

function filtrarRegs(q){
  var ql=q.toLowerCase();
  document.querySelectorAll('.reg-item').forEach(function(el){
    el.style.display=el.getAttribute('data-search').includes(ql)?'flex':'none';
  });
}

// Muestra un PDF ya generado (como blob URL) dentro del visor. En Android,
// mostrarlo dentro de un iframe suele fallar silenciosamente (el navegador
// solo muestra un ícono genérico sin renderizar el contenido), así que ahí
// se abre directamente en una pestaña nueva en vez de intentar el iframe.
function mostrarPDFViewer(blobUrl, nombreArchivo, urlDirecta){
  document.getElementById('btnDescPDF').onclick=function(){
    var a=document.createElement('a'); a.href=blobUrl; a.download=nombreArchivo||'documento.pdf';
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
  };
  var esAndroid = /Android/i.test(navigator.userAgent);
  if(esAndroid){
    // En Android, un blob: abierto con window.open suele descargarse directo
    // en vez de mostrarse. Usamos una URL real del servidor (con
    // Content-Disposition: inline) que el navegador sí respeta para mostrar
    // la vista previa. Si no hay URL directa disponible, usamos el blob como
    // respaldo (mejor eso que nada).
    document.getElementById('pdfLoadMsg').style.display='none';
    document.getElementById('pdfFrame').style.display='none';
    document.getElementById('pdfReadyMsg').style.display='block';
    var destino = urlDirecta || blobUrl;
    document.getElementById('pdfAbrirBtn').onclick=function(){ window.open(destino,'_blank'); };
    window.open(destino,'_blank');
    return;
  }
  var frame=document.getElementById('pdfFrame');
  var loaded=false;
  frame.onload=function(){ if(!loaded){ loaded=true; frame.style.display='block'; document.getElementById('pdfLoadMsg').style.display='none'; } };
  frame.src=blobUrl;
  setTimeout(function(){
    if(!loaded){ document.getElementById('pdfLoadMsg').style.display='none'; document.getElementById('pdfReadyMsg').style.display='block'; document.getElementById('pdfAbrirBtn').onclick=function(){ window.open(blobUrl,'_blank'); }; }
  },2500);
}

function verFichaReg(cui){
  if(!cui) return;
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  fetch('/ficha_pdf?cui='+encodeURIComponent(cui)).then(function(r){ return r.json(); }).then(function(d){
    if(d.ok&&d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, d.filename||'Ficha.pdf', '/ficha_pdf_directo?cui='+encodeURIComponent(cui));
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'No se pudo generar el PDF (registro no encontrado en la hoja DPI)');
    }
  }).catch(function(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  });
}

// Delegación de clic para todos los botones .btnPDF (registros individuales y grupos)
document.addEventListener('click', function(e){
  var btn = e.target.closest('.btnPDF');
  if(btn){
    var cui = btn.getAttribute('data-cui');
    var nombre = btn.getAttribute('data-nombre');
    if(cui && typeof verFichaReg === 'function') verFichaReg(cui);
    else if(nombre) impGrupo(decodeURIComponent(nombre));
    return;
  }
  var pend = e.target.closest('.pend-emp-item');
  if(pend){
    var cuiP = pend.getAttribute('data-cui');
    if(cuiP) usarPendienteEmp(cuiP);
  }
});

// GRUPOS CONSULTA

// Lista de jefes de sector, ahora administrable desde Mantenimiento en vez de
// estar fija en el código. Se cachea en memoria durante la sesión y se
// refresca cada vez que se agrega o elimina un jefe desde Mantenimiento.
var _jefesCache = null;
async function obtenerJefes(forzarRecarga){
  if(_jefesCache && !forzarRecarga) return _jefesCache;
  try{
    var r = await fetch('/listar_jefes');
    var d = await r.json();
    _jefesCache = (d.ok && d.jefes) ? d.jefes : [];
  }catch(e){ _jefesCache = []; }
  return _jefesCache;
}

async function cargarJefesSelect(){
  var jefes = await obtenerJefes();
  var sel = document.getElementById('loginJefeSel');
  var actual = sel.value;
  sel.innerHTML = '<option value="">-- Seleccione --</option>' +
    jefes.map(function(j){ return '<option>'+j+'</option>'; }).join('') +
    '<option value="ADMINISTRADOR">— ADMINISTRADOR (ver todo) —</option>';
  if(actual) sel.value = actual;
}

// ==== SESIÓN DE JEFE DE SECTOR ====
var sesionJefe = null;       // {nombre: 'X'} una vez autenticado
var _accionPendienteLogin = null;

function actualizarBotonLogout(){
  var btn = document.getElementById('bLogout');
  if(sesionJefe){
    btn.style.display='block';
    document.getElementById('bLogoutNombre').textContent = sesionJefe.esApoyo ? 'APOYO' : (sesionJefe.esAdmin ? 'ADMINISTRADOR' : sesionJefe.nombre);
  } else {
    btn.style.display='none';
  }
}

function _activarModoApoyo(){
  // El usuario de Apoyo no debe ver el menú normal — solo puede buscar y
  // editar teléfono/empadronamiento desde Registros Individuales. Se
  // ocultan los demás botones principales y se entra directo ahí.
  ['b1','b2','b3','b4','b5','b6'].forEach(function(id){
    var el = document.getElementById(id);
    if(el) el.style.display='none';
  });
  cargarRegistros();
}

function _restaurarModoNormal(){
  ['b1','b2','b3','b4','b5','b6'].forEach(function(id){
    var el = document.getElementById(id);
    if(el) el.style.display='block';
  });
  hide('oAfiliacionMenu');
  hide('oAfiliacionRegistro');
  hide('oImpresionCarnets');
  hide('oFirmaCarnet');
}

function abrirLoginAfiliacion(){
  document.getElementById('loginAfiliacionPin').value='';
  msg('msgLoginAfiliacion','','');
  show('oLoginAfiliacion');
}

async function hacerLoginAfiliacion(){
  var usuario = document.getElementById('loginAfiliacionUsuario').value;
  var pin = document.getElementById('loginAfiliacionPin').value.trim();
  if(!pin){ msg('msgLoginAfiliacion','Ingrese el código','#e74c3c'); return; }
  msg('msgLoginAfiliacion','Verificando...','#aed6f1');
  try{
    var r = await fetch('/login_afiliacion',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre:usuario, pin:pin})});
    var d = await r.json();
    if(d.ok){
      sesionJefe = {nombre:d.nombre, esAdmin:false, esApoyo:false, esAfiliacion:true};
      actualizarBotonLogout();
      hide('oLoginAfiliacion');
      if(_ultimoDatosAnalizadosAfiliado){
        // Se estaba recuperando de una sesión vencida a medio guardar: se
        // reintenta guardar de una vez, sin pedir que se vuelvan a tomar
        // las fotos ni a escribir los datos.
        show('oAfiliacionRegistro');
        await _intentarGuardarAfiliado(_ultimoDatosAnalizadosAfiliado);
      } else {
        _activarModoAfiliacion();
      }
    } else {
      msg('msgLoginAfiliacion','❌ '+(d.error||'No se pudo ingresar'),'#e74c3c');
    }
  }catch(e){ msg('msgLoginAfiliacion','❌ '+e.message,'#e74c3c'); }
}

function _activarModoAfiliacion(){
  // apoyo1/apoyo2 tampoco deben ver el menú normal — solo pueden entrar a
  // Afiliación o a Impresión. Se ocultan los demás botones y se muestra
  // el menú con esas 2 opciones.
  ['b1','b2','b3','b4','b5','b6'].forEach(function(id){
    var el = document.getElementById(id);
    if(el) el.style.display='none';
  });
  show('oAfiliacionMenu');
}

function irAAfiliacionRegistro(){
  hide('oAfiliacionMenu');
  gAfF=null; gAfR=null;
  document.getElementById('btnAfF').textContent='📷 Foto Frente del DPI'; document.getElementById('btnAfF').style.background='#7d3c98';
  document.getElementById('btnAfR').textContent='📷 Foto Reverso del DPI'; document.getElementById('btnAfR').style.background='#7d3c98';
  document.getElementById('btnAfProc').disabled=true; document.getElementById('btnAfProc').style.background='#566573';
  document.getElementById('afTel').value=''; document.getElementById('afDir').value='';
  document.getElementById('btnImprimirCarnet').style.display='none';
  _ultimoAfiliadoGuardado = null;
  msg('msgAfil','','');
  show('oAfiliacionRegistro');
}

function irAImpresionCarnets(){
  hide('oAfiliacionMenu');
  document.getElementById('fechaImpresionCarnets').value = new Date().toLocaleDateString('en-CA');
  document.getElementById('resImpresionCarnets').innerHTML = '';
  msg('msgImpresionCarnets','','');
  show('oImpresionCarnets');
}

var _fechaImpresionCarnetsActual = '';

async function buscarAfiliadosParaImprimir(){
  var fechaISO = document.getElementById('fechaImpresionCarnets').value;
  if(!fechaISO){ msg('msgImpresionCarnets','Elige una fecha','#e74c3c'); return; }
  var partes = fechaISO.split('-');
  var fechaFiltro = partes[2]+'/'+partes[1]+'/'+partes[0];
  _fechaImpresionCarnetsActual = fechaFiltro;
  var cont = document.getElementById('resImpresionCarnets');
  cont.innerHTML = '<p class="msg">Buscando...</p>';
  msg('msgImpresionCarnets','','');
  try{
    var r = await fetch('/listar_afiliados_fecha?fecha='+encodeURIComponent(fechaFiltro));
    var d = await r.json();
    if(!d.ok){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+(d.error||'Error')+'</p>'; return; }
    if(d.afiliados.length === 0){
      cont.innerHTML = '<p class="msg">No hay afiliados guardados el '+fechaFiltro+'</p>';
      return;
    }
    var h = '<div style="background:#154360;border-radius:8px;padding:8px;max-height:30vh;overflow-y:auto;margin-bottom:8px">';
    d.afiliados.forEach(function(a){
      h += '<p style="color:white;font-size:0.78em;margin:2px 0">• '+a.primer_nombre+' '+a.primer_apellido+' — DPI: '+a.cui+'</p>';
    });
    h += '</div>';
    h += '<button class="mbtn" style="background:#27ae60" onclick="generarListaAfiliados()">🖨️ Generar lista con estos '+d.afiliados.length+' afiliado(s)</button>';
    cont.innerHTML = h;
  }catch(e){ cont.innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+e.message+'</p>'; }
}

async function generarListaAfiliados(){
  msg('msgImpresionCarnets','Generando la lista...','#aed6f1');
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando la lista...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_lista_afiliados',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({fecha:_fechaImpresionCarnetsActual})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Lista_Afiliados.pdf', d.token?('/pdf_temporal/'+d.token):null);
      msg('msgImpresionCarnets','','');
    } else {
      document.getElementById('pdfLoadMsg').style.display='none';
      msg('msgImpresionCarnets','❌ '+(d.error||'Error al generar la lista'),'#e74c3c');
      hide('oPDFViewer');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='none';
    msg('msgImpresionCarnets','❌ '+e.message,'#e74c3c');
    hide('oPDFViewer');
  }
}

async function procAfiliado(){
  var btn = document.getElementById('btnAfProc');
  btn.disabled=true; btn.textContent='Analizando...'; msg('msgAfil','Analizando el DPI...','#aed6f1');
  try{
    var r = await fetch('/analizar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({frente:gAfF,reverso:gAfR,frente_type:'image/jpeg',reverso_type:'image/jpeg'})});
    var d = await r.json();
    if(!d.ok){
      msg('msgAfil','❌ '+(d.error||'Error al analizar'),'#e74c3c');
      btn.disabled=false; btn.textContent='Analizar y guardar afiliado'; return;
    }
    var datos = d.datos;
    if(!datos.fecha_nacimiento || !datos.fecha_nacimiento.trim()){
      msg('msgAfil','❌ Este DPI no tiene fecha de nacimiento legible. Vuelve a tomar la foto e inténtalo de nuevo.','#e74c3c');
      btn.disabled=false; btn.textContent='Analizar y guardar afiliado'; return;
    }
    var direccion = document.getElementById('afDir').value.trim();
    if(!direccion){
      msg('msgAfil','❌ La dirección / comunidad es obligatoria. Escríbela antes de guardar.','#e74c3c');
      btn.disabled=false; btn.textContent='Analizar y guardar afiliado'; return;
    }
    datos.telefono = document.getElementById('afTel').value.trim();
    datos.direccion = direccion;
    datos.dpi_no_verificado = !!d.no_autentico;
    _ultimaFotoFrenteAfiliado = gAfF; // se guarda desde ya, por si hay que reintentar tras volver a iniciar sesión
    _ultimoDatosAnalizadosAfiliado = datos;
    await _intentarGuardarAfiliado(datos);
  }catch(e){
    msg('msgAfil','❌ '+e.message,'#e74c3c');
    document.getElementById('btnAfProc').disabled=false; document.getElementById('btnAfProc').textContent='Analizar y guardar afiliado';
  }
}

var _ultimoDatosAnalizadosAfiliado = null;

async function _intentarGuardarAfiliado(datos){
  var btn = document.getElementById('btnAfProc');
  msg('msgAfil','Guardando...','#aed6f1');
  try{
    var r2 = await fetch('/guardar_afiliado',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({datos:datos})});
    var d2 = await r2.json();
    if(d2.ok){
      msg('msgAfil','✅ Afiliado guardado: '+(d2.nombre||''),'#2ecc71');
      avisarVozFaltantes(datos.telefono, d2.empadronado);
      _ultimoAfiliadoGuardado = {
        primer_nombre: datos.primer_nombre||'', segundo_nombre: datos.segundo_nombre||'',
        primer_apellido: datos.primer_apellido||'', segundo_apellido: datos.segundo_apellido||'',
        cui: datos.cui||'',
        departamento: datos.departamento_vecindad||datos.departamento_nacimiento||'',
        municipio: datos.municipio_vecindad||datos.municipio_nacimiento||'',
        fecha: new Date().toLocaleDateString('es-GT')
      };
      document.getElementById('btnImprimirCarnet').style.display='block';
      _ultimoDatosAnalizadosAfiliado = null;
      gAfF=null; gAfR=null;
      document.getElementById('btnAfF').textContent='📷 Foto Frente del DPI'; document.getElementById('btnAfF').style.background='#7d3c98';
      document.getElementById('btnAfR').textContent='📷 Foto Reverso del DPI'; document.getElementById('btnAfR').style.background='#7d3c98';
      document.getElementById('afTel').value=''; document.getElementById('afDir').value='';
    } else if(d2.sesion_requerida){
      // La sesión se venció justo antes de guardar (por ejemplo, si Render
      // "se durmió" mientras se tomaban las fotos). NO se pierde el
      // trabajo: los datos ya analizados quedan guardados en
      // _ultimoDatosAnalizadosAfiliado, listos para reintentar apenas se
      // vuelva a iniciar sesión.
      msg('msgAfil','⚠️ Se cerró la sesión (posiblemente por inactividad). Vuelve a iniciar sesión — tus datos ya analizados NO se pierden.','#f39c12');
      document.getElementById('loginAfiliacionPin').value='';
      show('oLoginAfiliacion');
    } else {
      msg('msgAfil','❌ '+(d2.error||'Error al guardar'),'#e74c3c');
    }
  }catch(e){
    msg('msgAfil','❌ '+e.message,'#e74c3c');
  }
  btn.disabled=true; btn.textContent='Analizar y guardar afiliado'; btn.style.background='#566573';
}

async function imprimirCarnetAfiliado(firmaB64){
  if(!_ultimoAfiliadoGuardado){ msg('msgAfil','Primero guarda un afiliado','#e74c3c'); return; }
  var btnCarnet = document.getElementById('btnImprimirCarnet');
  btnCarnet.disabled=true; btnCarnet.textContent='Generando carné...';
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando carné...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_carnet_afiliado',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({datos:_ultimoAfiliadoGuardado, firma_b64:firmaB64||null})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Carnet_Afiliacion.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='none';
      msg('msgAfil','❌ '+(d.error||'Error al generar el carné'),'#e74c3c');
      hide('oPDFViewer');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='none';
    msg('msgAfil','❌ '+e.message,'#e74c3c');
    hide('oPDFViewer');
  }
  btnCarnet.disabled=false; btnCarnet.textContent='🖨️ Imprimir carné de afiliación';
}

// ---- Lienzo de firma (dibujar con el dedo o el mouse) ----
var _firmaDibujando = false;
var _firmaTieneTrazo = false;
var _firmaCtx = null;

function _firmaPosDesdeEvento(canvas, e){
  var rect = canvas.getBoundingClientRect();
  return {
    x: (e.clientX - rect.left) * (canvas.width / rect.width),
    y: (e.clientY - rect.top) * (canvas.height / rect.height)
  };
}

function _rellenarFirmaBlanco(ctx, canvas){
  ctx.save();
  ctx.fillStyle = '#ffffff';
  ctx.fillRect(0, 0, canvas.width, canvas.height);
  ctx.restore();
}

function _inicializarLienzoFirma(){
  var canvas = document.getElementById('canvasFirma');
  var ctx = canvas.getContext('2d');
  _firmaCtx = ctx;
  _rellenarFirmaBlanco(ctx, canvas);
  ctx.lineWidth = 4;
  ctx.lineCap = 'round';
  ctx.lineJoin = 'round';
  ctx.strokeStyle = '#1a3a6b';

  // Se usan "Pointer Events" (en vez de mouse/touch por separado): es una
  // sola API que ya cubre dedo, mouse y lápiz óptico de forma unificada, y
  // evita el problema de que el trazo del dedo a veces no se registrara.
  function iniciar(e){
    _firmaDibujando = true; _firmaTieneTrazo = true;
    var p = _firmaPosDesdeEvento(canvas, e);
    // Se dibuja un punto de una vez al presionar (no solo al arrastrar):
    // así, si alguien solo PRESIONA el dedo sin moverlo (como se haría
    // con una huella), igual queda una marca visible en vez de no dejar
    // nada.
    ctx.beginPath();
    ctx.arc(p.x, p.y, ctx.lineWidth/2, 0, Math.PI*2);
    ctx.fillStyle = ctx.strokeStyle;
    ctx.fill();
    ctx.beginPath(); ctx.moveTo(p.x, p.y);
    try{ canvas.setPointerCapture(e.pointerId); }catch(err){}
    e.preventDefault();
  }
  function dibujar(e){
    if(!_firmaDibujando) return;
    var p = _firmaPosDesdeEvento(canvas, e);
    ctx.lineTo(p.x, p.y); ctx.stroke();
    e.preventDefault();
  }
  function parar(e){ _firmaDibujando = false; }

  canvas.style.touchAction = 'none';
  canvas.addEventListener('pointerdown', iniciar);
  canvas.addEventListener('pointermove', dibujar);
  canvas.addEventListener('pointerup', parar);
  canvas.addEventListener('pointercancel', parar);
  canvas.addEventListener('pointerleave', parar);
  canvas._firmaListo = true;
}

function limpiarFirma(){
  var canvas = document.getElementById('canvasFirma');
  if(_firmaCtx) _rellenarFirmaBlanco(_firmaCtx, canvas);
  _firmaTieneTrazo = false;
}

var _modoFirmaActual = 'dibujar';
var _recorteFirmaDPIBase64 = null;
var _recorteDPICtx = null;
var _recorteDPIImagenBase = null;

function cambiarModoFirma(modo){
  if(modo === 'recorte' && !_ultimaFotoFrenteAfiliado){
    msg('msgFirma','❌ No hay foto del reverso del DPI guardada para este afiliado (fue registrado antes de este cambio). Usa "Firmar en pantalla".','#e74c3c');
    return;
  }
  _modoFirmaActual = modo;
  document.getElementById('btnModoFirmaDibujar').style.background = modo==='dibujar' ? '#1a3a6b' : '#566573';
  document.getElementById('btnModoFirmaRecorte').style.background = modo==='recorte' ? '#1a3a6b' : '#566573';
  document.getElementById('divFirmaDibujar').style.display = modo==='dibujar' ? 'block' : 'none';
  document.getElementById('divFirmaRecorte').style.display = modo==='recorte' ? 'block' : 'none';
  msg('msgFirma','','');
  if(modo === 'recorte' && !_recorteDPICtx){
    _inicializarRecorteDPI();
  }
}

var _recorteRect = null; // {x,y,w,h} en píxeles del canvas
var _recorteModoActual = null; // 'mover' | 'redimensionar' | null
var _recorteArrastreOffset = null;
var _recorteCanvasRef = null;
var _recorteImgRef = null;
var _recorteHandleTam = 22;

function _recorteDibujar(){
  var canvas = _recorteCanvasRef, ctx = _recorteDPICtx, img = _recorteImgRef, r = _recorteRect;
  ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
  ctx.strokeStyle = '#27ae60'; ctx.lineWidth = 3; ctx.setLineDash([]);
  ctx.fillStyle = 'rgba(39,174,96,0.18)';
  ctx.fillRect(r.x, r.y, r.w, r.h);
  ctx.strokeRect(r.x, r.y, r.w, r.h);
  // Manija para redimensionar, en la esquina inferior derecha
  ctx.fillStyle = '#27ae60';
  ctx.fillRect(r.x + r.w - _recorteHandleTam/2, r.y + r.h - _recorteHandleTam/2, _recorteHandleTam, _recorteHandleTam);
}

function _inicializarRecorteDPI(){
  var canvas = document.getElementById('canvasRecorteDPI');
  var img = new Image();
  img.onload = function(){
    var MAX_W = 700;
    var escala = img.width > MAX_W ? MAX_W / img.width : 1;
    canvas.width = Math.round(img.width * escala);
    canvas.height = Math.round(img.height * escala);
    var ctx = canvas.getContext('2d');
    _recorteDPICtx = ctx;
    _recorteDPIImagenBase = img;
    _recorteCanvasRef = canvas;
    _recorteImgRef = img;
    // Recuadro inicial: centrado, de un tamaño razonable para una firma/huella
    var w0 = canvas.width * 0.42, h0 = canvas.height * 0.16;
    _recorteRect = { x: (canvas.width - w0)/2, y: canvas.height * 0.62, w: w0, h: h0 };
    _recorteDibujar();

    function pos(e){
      var rect = canvas.getBoundingClientRect();
      return {
        x: (e.clientX - rect.left) * (canvas.width / rect.width),
        y: (e.clientY - rect.top) * (canvas.height / rect.height)
      };
    }
    function dentroDeManija(p){
      var r = _recorteRect;
      return Math.abs(p.x - (r.x+r.w)) < _recorteHandleTam && Math.abs(p.y - (r.y+r.h)) < _recorteHandleTam;
    }
    function dentroDelRecuadro(p){
      var r = _recorteRect;
      return p.x >= r.x && p.x <= r.x+r.w && p.y >= r.y && p.y <= r.y+r.h;
    }
    function iniciar(e){
      var p = pos(e);
      if(dentroDeManija(p)){ _recorteModoActual = 'redimensionar'; }
      else if(dentroDelRecuadro(p)){ _recorteModoActual = 'mover'; _recorteArrastreOffset = {x: p.x - _recorteRect.x, y: p.y - _recorteRect.y}; }
      else { _recorteModoActual = null; return; }
      try{ canvas.setPointerCapture(e.pointerId); }catch(err){}
      e.preventDefault();
    }
    function arrastrar(e){
      if(!_recorteModoActual) return;
      var p = pos(e);
      var r = _recorteRect;
      if(_recorteModoActual === 'mover'){
        r.x = Math.max(0, Math.min(canvas.width - r.w, p.x - _recorteArrastreOffset.x));
        r.y = Math.max(0, Math.min(canvas.height - r.h, p.y - _recorteArrastreOffset.y));
      } else if(_recorteModoActual === 'redimensionar'){
        r.w = Math.max(30, Math.min(canvas.width - r.x, p.x - r.x));
        r.h = Math.max(20, Math.min(canvas.height - r.y, p.y - r.y));
      }
      _recorteDibujar();
      e.preventDefault();
    }
    function soltar(){ _recorteModoActual = null; }

    canvas.style.touchAction = 'none';
    canvas.addEventListener('pointerdown', iniciar);
    canvas.addEventListener('pointermove', arrastrar);
    canvas.addEventListener('pointerup', soltar);
    canvas.addEventListener('pointercancel', soltar);
  };
  img.src = 'data:image/jpeg;base64,' + _ultimaFotoFrenteAfiliado;
}

function confirmarRecorteDPI(){
  if(!_recorteRect || !_recorteCanvasRef) return;
  var r = _recorteRect;
  var temp = document.createElement('canvas');
  temp.width = r.w; temp.height = r.h;
  var tctx = temp.getContext('2d');
  // Redibujar la imagen base sin el recuadro verde encima, para recortar limpio
  tctx.drawImage(_recorteImgRef, r.x * (_recorteImgRef.width/_recorteCanvasRef.width), r.y * (_recorteImgRef.height/_recorteCanvasRef.height),
    r.w * (_recorteImgRef.width/_recorteCanvasRef.width), r.h * (_recorteImgRef.height/_recorteCanvasRef.height), 0, 0, r.w, r.h);
  _recorteFirmaDPIBase64 = temp.toDataURL('image/png');
  document.getElementById('imgPreviewRecorte').src = _recorteFirmaDPIBase64;
  document.getElementById('previewRecorteDPI').style.display='block';
  msg('msgFirma','✅ Recorte listo — revisa la vista previa abajo','#2ecc71');
}

function abrirFirmaCarnet(){
  var canvas = document.getElementById('canvasFirma');
  if(!canvas._firmaListo) _inicializarLienzoFirma();
  limpiarFirma();
  _recorteFirmaDPIBase64 = null;
  _recorteDPICtx = null;
  _recorteDPIImagenBase = null;
  _recorteRect = null;
  document.getElementById('previewRecorteDPI').style.display='none';
  msg('msgFirma','','');
  cambiarModoFirma('dibujar');
  show('oFirmaCarnet');
}

async function generarCarnetConFirma(omitir){
  var firmaB64 = null;
  if(!omitir){
    if(_modoFirmaActual === 'recorte' && _recorteFirmaDPIBase64){
      firmaB64 = _recorteFirmaDPIBase64;
    } else if(_modoFirmaActual === 'dibujar' && _firmaTieneTrazo){
      var canvas = document.getElementById('canvasFirma');
      firmaB64 = canvas.toDataURL('image/png');
    }
  }
  hide('oFirmaCarnet');
  await imprimirCarnetAfiliado(firmaB64);
}

function cerrarRegistrosInd(){
  // El usuario de Apoyo no tiene ninguna otra pantalla a la que volver
  // (el botón "Cerrar sesión" del menú de fondo queda tapado mientras hay
  // una ventana superpuesta abierta), así que para Apoyo este botón cierra
  // la sesión por completo. Para jefes/administrador normales, solo
  // cierra esta ventana como antes.
  if(sesionJefe && sesionJefe.esApoyo){ cerrarSesionCompleta(); return; }
  hide('oRegistros');
}

function abrirApoyo(){
  document.getElementById('loginApoyoPin').value='';
  msg('msgLoginApoyo','','');
  show('oLoginApoyo');
}

async function hacerLoginApoyo(){
  var pin = document.getElementById('loginApoyoPin').value.trim();
  if(!pin){ msg('msgLoginApoyo','Ingrese el PIN','#e74c3c'); return; }
  msg('msgLoginApoyo','Verificando...','#aed6f1');
  try{
    var r = await fetch('/login_apoyo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pin:pin})});
    var d = await r.json();
    if(d.ok){
      sesionJefe = {nombre:'APOYO', esAdmin:false, esApoyo:true};
      actualizarBotonLogout();
      hide('oLoginApoyo');
      _activarModoApoyo();
    } else {
      msg('msgLoginApoyo','❌ '+(d.error||'No se pudo ingresar'),'#e74c3c');
    }
  }catch(e){ msg('msgLoginApoyo','❌ '+e.message,'#e74c3c'); }
}

async function cerrarSesionCompleta(){
  try{ await fetch('/logout',{method:'POST'}); }catch(e){}
  sesionJefe = null;
  actualizarBotonLogout();
  _restaurarModoNormal();
  // Devolver a un estado limpio: cerrar cualquier panel de Grupos/Reportes/Mantenimiento abierto
  ['oGrupos','oVerGrupos','oNuevoG1','oBuscarG','oReportes','oConsPresidentas','oConsCoordinadoras','oMant','oRegistros','oConsultas'].forEach(function(id){
    var el=document.getElementById(id);
    if(el) el.style.display='none';
  });
  document.getElementById('mantLogin').style.display='block';
  document.getElementById('mantPanel').style.display='none';
  clv=''; updC();
  alert('Sesión cerrada.');
}

// Al cargar la página, revisar si ya existe una sesión activa en el
// servidor (por ejemplo, si la persona recargó la página sin cerrar sesión).
(async function _restaurarSesion(){
  try{
    var r = await fetch('/sesion_actual');
    var d = await r.json();
    if(d.ok && d.nombre){
      sesionJefe = {nombre:d.nombre, esAdmin:!!d.es_admin, esApoyo:!!d.es_apoyo};
      actualizarBotonLogout();
      if(sesionJefe.esApoyo){ _activarModoApoyo(); }
    }
  }catch(e){}
})();

function abrirGrupos(){
  if(sesionJefe){ mostrarInfoSesion(); show('oGrupos'); return; }
  _accionPendienteLogin = function(){ mostrarInfoSesion(); show('oGrupos'); };
  document.getElementById('loginJefeSel').value='';
  document.getElementById('loginJefePin').value='';
  msg('msgLoginJefe','','');
  cargarJefesSelect();
  show('oLoginJefe');
}

function abrirRegistrosInd(){
  if(sesionJefe){ hide('oConsultas'); cargarRegistros(); return; }
  hide('oConsultas');
  _accionPendienteLogin = function(){ cargarRegistros(); };
  document.getElementById('loginJefeSel').value='';
  document.getElementById('loginJefePin').value='';
  msg('msgLoginJefe','','');
  cargarJefesSelect();
  show('oLoginJefe');
}

function abrirRegistrosGrupos(){
  if(sesionJefe){ hide('oConsultas'); cargarGrupos(); return; }
  hide('oConsultas');
  _accionPendienteLogin = function(){ cargarGrupos(); };
  document.getElementById('loginJefeSel').value='';
  document.getElementById('loginJefePin').value='';
  msg('msgLoginJefe','','');
  cargarJefesSelect();
  show('oLoginJefe');
}

function abrirPresidentasConsulta(){
  _origenPresCoord = 'consultas';
  if(sesionJefe){ hide('oConsultas'); cargarConsPresidentas(); return; }
  hide('oConsultas');
  _accionPendienteLogin = function(){ cargarConsPresidentas(); };
  document.getElementById('loginJefeSel').value='';
  document.getElementById('loginJefePin').value='';
  msg('msgLoginJefe','','');
  cargarJefesSelect();
  show('oLoginJefe');
}

function abrirConsultaPadron(){
  if(sesionJefe){ hide('oConsultas'); document.getElementById('padronQ').value=''; document.getElementById('resPadron').innerHTML=''; show('oConsultaPadron'); return; }
  hide('oConsultas');
  _accionPendienteLogin = function(){ document.getElementById('padronQ').value=''; document.getElementById('resPadron').innerHTML=''; show('oConsultaPadron'); };
  document.getElementById('loginJefeSel').value='';
  document.getElementById('loginJefePin').value='';
  msg('msgLoginJefe','','');
  cargarJefesSelect();
  show('oLoginJefe');
}

async function buscarPadron(){
  var q = document.getElementById('padronQ').value.trim();
  if(q.length < 3){ document.getElementById('resPadron').innerHTML='<p class="msg" style="color:#e74c3c">Escriba al menos 3 caracteres</p>'; return; }
  document.getElementById('resPadron').innerHTML='<p class="msg">Buscando...</p>';
  try{
    var r = await fetch('/consultar_padron?q='+encodeURIComponent(q));
    var d = await r.json();
    if(!d.ok){ document.getElementById('resPadron').innerHTML='<p class="msg" style="color:#e74c3c">❌ '+(d.error||'Error')+'</p>'; return; }
    if(!d.encontrado){
      document.getElementById('resPadron').innerHTML='<p class="msg" style="color:#e74c3c">❌ No se encontró en el padrón — probablemente NO está empadronado</p>';
      return;
    }
    var h = '<p style="color:#2ecc71;font-weight:bold;margin-bottom:8px">✅ SÍ aparece en el padrón ('+d.resultados.length+' resultado(s))</p>';
    d.resultados.forEach(function(reg){
      h += '<div style="background:#154360;border-radius:8px;padding:10px;margin-bottom:8px">';
      for(var campo in reg){
        if(!reg[campo]) continue;
        h += '<p style="color:white;font-size:0.85em;margin:2px 0"><b>'+campo+':</b> '+reg[campo]+'</p>';
      }
      h += '</div>';
    });
    document.getElementById('resPadron').innerHTML = h;
  }catch(e){
    document.getElementById('resPadron').innerHTML='<p class="msg" style="color:#e74c3c">❌ '+e.message+'</p>';
  }
}

// ==== REPORTES ====
var _origenPresCoord = 'consultas'; // 'consultas' o 'reportes', para saber a dónde regresar

function abrirReportes(){
  if(sesionJefe){ mostrarInfoSesionReportes(); show('oReportes'); return; }
  _accionPendienteLogin = function(){ mostrarInfoSesionReportes(); show('oReportes'); };
  document.getElementById('loginJefeSel').value='';
  document.getElementById('loginJefePin').value='';
  msg('msgLoginJefe','','');
  cargarJefesSelect();
  show('oLoginJefe');
}

function mostrarInfoSesionReportes(){
  document.getElementById('sesionJefeInfoRep').textContent = sesionJefe ? (sesionJefe.esAdmin ? '👑 Sesión: ADMINISTRADOR (todos los jefes)' : 'Sesión: '+sesionJefe.nombre) : '';
}

function abrirPresidentasReporte(){
  _origenPresCoord = 'reportes';
  hide('oReportes');
  cargarConsPresidentas();
}

function abrirCoordinadorasReporte(){
  hide('oReportes');
  cargarConsCoordinadoras();
}

function abrirIntegrantesReporte(){
  hide('oReportes');
  cargarConsIntegrantes();
}

var _integrantesActuales = [];

function abrirPresCoordReporte(){
  hide('oReportes');
  cargarConsPresCoord();
}

var _presCoordActuales = [];

function abrirTodoCombinadoReporte(){
  hide('oReportes');
  cargarConsTodoCombinado();
}

var _todoCombinadoActuales = [];

function abrirNoEmpadronadosReporte(){
  hide('oReportes');
  document.getElementById('chkNoEmp').checked = true;
  document.getElementById('chkSinTel').checked = false;
  document.getElementById('filtroFechaNoEmp').value = '';
  msg('msgSeleccionNoEmpTel','','');
  show('oSeleccionNoEmpTel');
}

function confirmarSeleccionNoEmpTel(){
  var incluirNoEmp = document.getElementById('chkNoEmp').checked;
  var incluirSinTel = document.getElementById('chkSinTel').checked;
  if(!incluirNoEmp && !incluirSinTel){
    msg('msgSeleccionNoEmpTel','Selecciona al menos una opción','#e74c3c');
    return;
  }
  // El input type=date entrega AAAA-MM-DD; la hoja guarda la fecha de
  // registro como DD/MM/AAAA, así que se convierte antes de mandarla.
  var fechaISO = document.getElementById('filtroFechaNoEmp').value;
  var fechaFiltro = '';
  if(fechaISO){
    var partes = fechaISO.split('-');
    fechaFiltro = partes[2]+'/'+partes[1]+'/'+partes[0];
  }
  hide('oSeleccionNoEmpTel');
  cargarConsNoEmpadronados(incluirNoEmp, incluirSinTel, fechaFiltro);
}

var _noEmpadronadosActuales = [];
var _noEmpadronadosTitulo = 'NO EMPADRONADOS';

var _coordinadorasActuales = [];

// Abre WhatsApp con el número de la persona y el mensaje que se haya
// escrito en la caja de texto correspondiente (se lee en el momento del
// clic, así que si cambias el mensaje entre una persona y otra, cada
// quien recibe el texto que estaba escrito justo cuando le tocaste enviar).
// Recibe el propio botón (this) y lee el teléfono y el id de la caja de
// mensaje desde sus atributos data-, para no tener que anidar comillas.
function _abrirWhatsAppBtn(btn){
  var tel = (btn.dataset.tel||'').replace(/[^0-9]/g,'');
  if(!tel){ alert('Esta persona no tiene teléfono registrado'); return; }
  var campo = document.getElementById(btn.dataset.msgid);
  var mensaje = campo ? campo.value.trim() : '';
  var url = 'https://wa.me/'+tel+(mensaje ? ('?text='+encodeURIComponent(mensaje)) : '');
  window.open(url, '_blank');
}

async function cargarConsCoordinadoras(){
  show('oConsCoordinadoras');
  document.getElementById('contCoordinadoras').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosCoordinadoras').innerHTML='';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var jefeParam = esAdmin ? '' : ('?jefe='+encodeURIComponent(sesionJefe.nombre));
    var r = await fetch('/ver_coordinadoras'+jefeParam);
    var d = await r.json();
    _coordinadorasActuales = (d.ok && d.coordinadoras) ? d.coordinadoras : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_coordinadorasActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_coordinadorasActuales.length+' coordinadora(s) para '+etiqueta+'</p>';
      h += '<div style="background:#0d2137;border-radius:8px;padding:8px;margin-bottom:10px">';
      h += '<p style="color:#7fb3d5;font-size:0.75em;margin:0 0 4px 2px">Mensaje o enlace a enviar por WhatsApp:</p>';
      h += '<input type="text" id="msgWaCoordinadoras" placeholder="Ej. enlace de TikTok..." style="margin:0">';
      h += '</div>';
      _coordinadorasActuales.forEach(function(c){
        h += '<div style="background:#154360;border-radius:8px;padding:10px;margin-bottom:6px">';
        h += '<p style="color:white;font-weight:bold;margin:0">'+(c.nombre||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.78em;margin:2px 0 0 0">Grupo: '+(c.grupo||'-')+' &nbsp;|&nbsp; Presidenta: '+(c.presidenta||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.75em;margin:2px 0 0 0">Jefe: '+(c.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.8em;margin:4px 0 0 0">DPI: '+(c.cui||'-')+' &nbsp; Tel: '+(c.telefono||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.8em;margin:2px 0 0 0">Dirección: '+(c.direccion||'-')+'</p>';
        if(c.telefono) h += '<button class="mbtn" style="background:#25D366;margin:8px 0 0 0;padding:8px;font-size:0.85em" data-tel="'+c.telefono+'" data-msgid="msgWaCoordinadoras" onclick="_abrirWhatsAppBtn(this)">📱 Enviar WhatsApp</button>';
        h += '</div>';
      });
      document.getElementById('contCoordinadoras').innerHTML = h;
      document.getElementById('botonesFijosCoordinadoras').innerHTML =
        '<div style="display:flex;gap:8px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfCoordinadorasSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelCoordinadorasSesion()">📊 Excel</button>'+
        '</div>';
    } else {
      document.getElementById('contCoordinadoras').innerHTML = '<p class="msg">No hay coordinadoras registradas para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contCoordinadoras').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarExcelCoordinadorasSesion(){
  try{
    var r = await fetch('/excel_coordinadoras',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), coordinadoras: _coordinadorasActuales})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Coordinadoras_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function generarPdfCoordinadorasSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_coordinadoras',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), coordinadoras: _coordinadorasActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Coordinadoras_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

var _presidentasActuales = [];

async function cargarConsIntegrantes(){
  show('oConsIntegrantes');
  document.getElementById('contIntegrantes').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosIntegrantes').innerHTML='';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var jefeParam = esAdmin ? '' : ('?jefe='+encodeURIComponent(sesionJefe.nombre));
    var r = await fetch('/ver_integrantes'+jefeParam);
    var d = await r.json();
    _integrantesActuales = (d.ok && d.integrantes) ? d.integrantes : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_integrantesActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_integrantesActuales.length+' integrante(s) para '+etiqueta+' — ordenados por Presidenta, Coordinadora y Nombre</p>';
      _integrantesActuales.forEach(function(p){
        h += '<div style="background:#154360;border-radius:7px;padding:8px;margin-bottom:5px">';
        h += '<p style="color:white;font-weight:bold;margin:0;font-size:0.85em">'+(p.nombre||'-')+(p.es_coordinadora?' <span style="color:#f5b041">(Coord)</span>':'')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Coordinadora: '+(p.coordinadora||'-')+' &nbsp;|&nbsp; Presidenta: '+(p.presidenta||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.7em;margin:2px 0 0 0">Jefe: '+(p.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">DPI: '+(p.cui||'-')+' &nbsp; Tel: '+(p.telefono||'-')+'</p>';
        h += '</div>';
      });
      document.getElementById('contIntegrantes').innerHTML = h;
      document.getElementById('botonesFijosIntegrantes').innerHTML =
        '<div style="display:flex;gap:8px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfIntegrantesSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelIntegrantesSesion()">📊 Excel</button>'+
        '</div>';
    } else {
      document.getElementById('contIntegrantes').innerHTML = '<p class="msg">No hay integrantes registrados para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contIntegrantes').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarExcelIntegrantesSesion(){
  try{
    var r = await fetch('/excel_integrantes',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), integrantes: _integrantesActuales})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Integrantes_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function generarPdfIntegrantesSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_integrantes',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), integrantes: _integrantesActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Integrantes_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

async function cargarConsPresCoord(){
  show('oConsPresCoord');
  document.getElementById('contPresCoord').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosPresCoord').innerHTML='';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var jefeParam = esAdmin ? '' : ('?jefe='+encodeURIComponent(sesionJefe.nombre));
    var r = await fetch('/ver_presidentas_coordinadoras'+jefeParam);
    var d = await r.json();
    _presCoordActuales = (d.ok && d.registros) ? d.registros : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_presCoordActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_presCoordActuales.length+' registro(s) para '+etiqueta+' — ordenados por Comunidad y Presidenta</p>';
      h += '<div style="background:#0d2137;border-radius:8px;padding:8px;margin-bottom:10px">';
      h += '<p style="color:#7fb3d5;font-size:0.75em;margin:0 0 4px 2px">Mensaje o enlace a enviar por WhatsApp:</p>';
      h += '<input type="text" id="msgWaPresCoord" placeholder="Ej. enlace de TikTok..." style="margin:0">';
      h += '</div>';
      _presCoordActuales.forEach(function(p){
        var colorTipo = p.tipo==='Presidenta' ? '#2ecc71' : '#5dade2';
        h += '<div style="background:#154360;border-radius:7px;padding:8px;margin-bottom:5px">';
        h += '<p style="color:white;font-weight:bold;margin:0;font-size:0.85em">'+(p.nombre||'-')+' <span style="color:'+colorTipo+'">('+p.tipo+')</span></p>';
        if(p.grupo) h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Grupo: '+p.grupo+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Presidenta: '+(p.presidenta||'-')+' &nbsp;|&nbsp; Comunidad: '+(p.direccion||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.7em;margin:2px 0 0 0">Jefe: '+(p.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">DPI: '+(p.cui||'-')+' &nbsp; Tel: '+(p.telefono||'-')+'</p>';
        if(p.telefono) h += '<button class="mbtn" style="background:#25D366;margin:8px 0 0 0;padding:8px;font-size:0.85em" data-tel="'+p.telefono+'" data-msgid="msgWaPresCoord" onclick="_abrirWhatsAppBtn(this)">📱 Enviar WhatsApp</button>';
        h += '</div>';
      });
      document.getElementById('contPresCoord').innerHTML = h;
      document.getElementById('botonesFijosPresCoord').innerHTML =
        '<div style="display:flex;gap:8px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfPresCoordSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelPresCoordSesion()">📊 Excel</button>'+
        '</div>';
    } else {
      document.getElementById('contPresCoord').innerHTML = '<p class="msg">No hay presidentas ni coordinadoras registradas para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contPresCoord').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarExcelPresCoordSesion(){
  try{
    var r = await fetch('/excel_presidentas_coordinadoras',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _presCoordActuales})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Presidentas_Coordinadoras_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function generarPdfPresCoordSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_presidentas_coordinadoras',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _presCoordActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Presidentas_Coordinadoras_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

async function cargarConsTodoCombinado(){
  show('oConsTodoCombinado');
  document.getElementById('contTodoCombinado').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosTodoCombinado').innerHTML='';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var jefeParam = esAdmin ? '' : ('?jefe='+encodeURIComponent(sesionJefe.nombre));
    var r = await fetch('/ver_todo_combinado'+jefeParam);
    var d = await r.json();
    _todoCombinadoActuales = (d.ok && d.registros) ? d.registros : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_todoCombinadoActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_todoCombinadoActuales.length+' registro(s) para '+etiqueta+' — ordenados por Presidenta, Coordinadora, Comunidad y Nombre</p>';
      var coloresTipo = {'Presidenta':'#2ecc71','Coordinadora':'#f5b041','Integrante':'#5dade2'};
      _todoCombinadoActuales.forEach(function(p){
        var colorTipo = coloresTipo[p.tipo] || '#5dade2';
        h += '<div style="background:#154360;border-radius:7px;padding:8px;margin-bottom:5px">';
        h += '<p style="color:white;font-weight:bold;margin:0;font-size:0.85em">'+(p.nombre||'-')+' <span style="color:'+colorTipo+'">('+p.tipo+')</span></p>';
        if(p.grupo) h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Grupo: '+p.grupo+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Presidenta: '+(p.presidenta||'-')+' &nbsp;|&nbsp; Comunidad: '+(p.direccion||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.7em;margin:2px 0 0 0">Jefe: '+(p.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">DPI: '+(p.cui||'-')+' &nbsp; Tel: '+(p.telefono||'-')+'</p>';
        h += '</div>';
      });
      document.getElementById('contTodoCombinado').innerHTML = h;
      document.getElementById('botonesFijosTodoCombinado').innerHTML =
        '<div style="display:flex;gap:8px;margin-bottom:6px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfTodoCombinadoSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelTodoCombinadoSesion()">📊 Excel</button>'+
        '</div>'+
        '<button class="mbtn" style="background:#117864;margin:0" onclick="generarOrganigramaSesion()">🗂️ Organigrama de Estructura</button>';
    } else {
      document.getElementById('contTodoCombinado').innerHTML = '<p class="msg">No hay registros para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contTodoCombinado').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarExcelTodoCombinadoSesion(){
  try{
    var r = await fetch('/excel_todo_combinado',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _todoCombinadoActuales})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Presidentas_Coordinadoras_Integrantes_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function generarPdfTodoCombinadoSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_todo_combinado',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _todoCombinadoActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Presidentas_Coordinadoras_Integrantes_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

async function generarOrganigramaSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando organigrama...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_organigrama',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _todoCombinadoActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Organigrama_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar el organigrama');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

var _observacionesActuales = [];

async function cargarConsObservaciones(){
  show('oConsObservaciones');
  document.getElementById('contObservaciones').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosObservaciones').innerHTML='';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var params = [];
    if(!esAdmin) params.push('jefe='+encodeURIComponent(sesionJefe.nombre));
    var r = await fetch('/ver_con_observaciones?'+params.join('&'));
    var d = await r.json();
    _observacionesActuales = (d.ok && d.registros) ? d.registros : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_observacionesActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_observacionesActuales.length+' persona(s) con observaciones para '+etiqueta+'</p>';
      var coloresTipo = {'Presidenta':'#2ecc71','Coordinadora':'#f5b041','Integrante':'#5dade2'};
      _observacionesActuales.forEach(function(p){
        var colorTipo = coloresTipo[p.tipo] || '#5dade2';
        h += '<div style="background:#154360;border-radius:7px;padding:8px;margin-bottom:5px">';
        h += '<p style="color:white;font-weight:bold;margin:0;font-size:0.85em">'+(p.nombre||'-')+' <span style="color:'+colorTipo+'">('+p.tipo+')</span></p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Presidenta: '+(p.presidenta||'-')+' &nbsp;|&nbsp; Comunidad: '+(p.direccion||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.7em;margin:2px 0 0 0">Jefe: '+(p.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">DPI: '+(p.cui||'-')+' &nbsp; Emp: '+(p.empadronado||'-')+'</p>';
        h += '<p style="color:#f7dc6f;font-size:0.78em;margin:3px 0 0 0">📝 '+p.observaciones+'</p>';
        h += '</div>';
      });
      document.getElementById('contObservaciones').innerHTML = h;
      document.getElementById('botonesFijosObservaciones').innerHTML =
        '<div style="display:flex;gap:8px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfObservacionesSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelObservacionesSesion()">📊 Excel</button>'+
        '</div>';
    } else {
      document.getElementById('contObservaciones').innerHTML = '<p class="msg">Nadie tiene observaciones registradas todavía para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contObservaciones').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarPdfObservacionesSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_todo_combinado',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _observacionesActuales, titulo:'PERSONAS CON OBSERVACIONES', modo:'observaciones'})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Reporte_Observaciones_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

async function generarExcelObservacionesSesion(){
  try{
    var r = await fetch('/excel_todo_combinado',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _observacionesActuales, titulo:'PERSONAS CON OBSERVACIONES', modo:'observaciones'})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Reporte_Observaciones_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function cargarConsNoEmpadronados(incluirNoEmp, incluirSinTel, fechaFiltro){
  if(incluirNoEmp === undefined) incluirNoEmp = true;
  if(incluirSinTel === undefined) incluirSinTel = false;
  if(fechaFiltro === undefined) fechaFiltro = '';
  show('oConsNoEmpadronados');
  document.getElementById('contNoEmpadronados').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosNoEmpadronados').innerHTML='';

  var etiquetasFiltro = [];
  if(incluirNoEmp) etiquetasFiltro.push('Sin empadronar');
  if(incluirSinTel) etiquetasFiltro.push('Sin teléfono');
  var textoFiltro = etiquetasFiltro.join(' y/o ');
  if(fechaFiltro) textoFiltro += ' — agregados el ' + fechaFiltro;
  _noEmpadronadosTitulo = textoFiltro.toUpperCase();
  document.getElementById('tituloConsNoEmpadronados').textContent = '😕 ' + textoFiltro;

  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var params = [];
    if(!esAdmin) params.push('jefe='+encodeURIComponent(sesionJefe.nombre));
    params.push('no_emp='+(incluirNoEmp?'SI':''));
    params.push('sin_tel='+(incluirSinTel?'SI':''));
    if(fechaFiltro) params.push('fecha='+encodeURIComponent(fechaFiltro));
    var r = await fetch('/ver_no_empadronados?'+params.join('&'));
    var d = await r.json();
    _noEmpadronadosActuales = (d.ok && d.registros) ? d.registros : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_noEmpadronadosActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_noEmpadronadosActuales.length+' persona(s) para '+etiqueta+' ('+textoFiltro+')</p>';
      var coloresTipo = {'Presidenta':'#2ecc71','Coordinadora':'#f5b041','Integrante':'#5dade2'};
      _noEmpadronadosActuales.forEach(function(p){
        var colorTipo = coloresTipo[p.tipo] || '#5dade2';
        var faltaEmp = incluirNoEmp && (p.empadronado||'').toUpperCase() !== 'SI';
        var faltaTel = incluirSinTel && !((p.telefono||'').trim());
        h += '<div style="background:#154360;border-radius:7px;padding:8px;margin-bottom:5px">';
        h += '<p style="color:white;font-weight:bold;margin:0;font-size:0.85em">'+(p.nombre||'-')+' <span style="color:'+colorTipo+'">('+p.tipo+')</span></p>';
        if(p.grupo) h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Grupo: '+p.grupo+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Presidenta: '+(p.presidenta||'-')+' &nbsp;|&nbsp; Comunidad: '+(p.direccion||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.7em;margin:2px 0 0 0">Jefe: '+(p.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">DPI: '+(p.cui||'-')+' &nbsp; F.Nac: '+(p.fecha_nacimiento||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">Emp: <span style="color:'+(faltaEmp?'#f5b7b1':'#aed6f1')+';font-weight:'+(faltaEmp?'bold':'normal')+'">'+(p.empadronado||'-')+'</span> &nbsp; Tel: <span style="color:'+(faltaTel?'#f5b7b1':'#aed6f1')+';font-weight:'+(faltaTel?'bold':'normal')+'">'+(p.telefono||'FALTA')+'</span></p>';
        if(p.observaciones) h += '<p style="color:#f7dc6f;font-size:0.72em;margin:2px 0 0 0">📝 '+p.observaciones+'</p>';
        h += '</div>';
      });
      document.getElementById('contNoEmpadronados').innerHTML = h;
      document.getElementById('botonesFijosNoEmpadronados').innerHTML =
        '<div style="display:flex;gap:8px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfNoEmpadronadosSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelNoEmpadronadosSesion()">📊 Excel</button>'+
        '</div>';
    } else {
      document.getElementById('contNoEmpadronados').innerHTML = '<p class="msg">🎉 No hay pendientes ('+textoFiltro+') para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contNoEmpadronados').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarExcelNoEmpadronadosSesion(){
  try{
    var r = await fetch('/excel_todo_combinado',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _noEmpadronadosActuales, titulo:_noEmpadronadosTitulo})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Reporte_Pendientes_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function generarPdfNoEmpadronadosSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_todo_combinado',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), registros: _noEmpadronadosActuales, titulo:_noEmpadronadosTitulo})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Reporte_Pendientes_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

var _empActualizadosActuales = [];

function abrirEmpadronamientosActualizados(){
  hide('oMant');
  show('oEmpActualizados');
  cargarEmpadronamientosActualizados();
}

async function cargarEmpadronamientosActualizados(){
  document.getElementById('listaEmpActualizados').innerHTML = '<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosEmpActualizados').innerHTML = '';
  try{
    var r = await fetch('/ver_empadronamientos_actualizados');
    var d = await r.json();
    if(!d.ok){ document.getElementById('listaEmpActualizados').innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+(d.error||'Error')+'</p>'; return; }
    _empActualizadosActuales = d.registros || [];
    if(_empActualizadosActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_empActualizadosActuales.length+' cambio(s) registrado(s)</p>';
      _empActualizadosActuales.forEach(function(reg){
        h += '<div style="background:#154360;border-radius:7px;padding:8px;margin-bottom:5px">';
        h += '<p style="color:white;font-weight:bold;margin:0;font-size:0.85em">CUI: '+(reg.cui||'-')+' — <span style="color:'+(reg.empadronado==='SI'?'#2ecc71':'#e74c3c')+'">'+(reg.empadronado||'-')+'</span></p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">No. Empadronamiento: '+(reg.num_empadronamiento||'-')+(reg.agregado_padron?' &nbsp;|&nbsp; ➕ Agregado a Padrón':'')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.72em;margin:2px 0 0 0">'+(reg.fecha||'-')+' — '+(reg.jefe||'-')+'</p>';
        h += '</div>';
      });
      document.getElementById('listaEmpActualizados').innerHTML = h;
      document.getElementById('botonesFijosEmpActualizados').innerHTML =
        '<button class="mbtn" style="background:#8e44ad;margin:0" onclick="generarPdfEmpActualizados()">📄 Generar PDF</button>';
    } else {
      document.getElementById('listaEmpActualizados').innerHTML = '<p class="msg">Todavía no se ha editado ningún empadronamiento desde Registros Individuales</p>';
    }
  }catch(e){
    document.getElementById('listaEmpActualizados').innerHTML = '<p class="msg" style="color:#e74c3c">❌ '+e.message+'</p>';
  }
}

async function generarPdfEmpActualizados(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_empadronamientos_actualizados',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({registros:_empActualizadosActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Empadronamientos_Actualizados.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

async function cargarConsPresidentas(){
  show('oConsPresidentas');
  document.getElementById('botonesJefesPres').innerHTML='';
  document.getElementById('contPresidentas').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('botonesFijosPres').innerHTML='';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var jefeParam = esAdmin ? '' : ('?jefe='+encodeURIComponent(sesionJefe.nombre));
    var r = await fetch('/ver_presidentas'+jefeParam);
    var d = await r.json();
    _presidentasActuales = (d.ok && d.presidentas) ? d.presidentas : [];
    var etiqueta = esAdmin ? 'todos los jefes' : sesionJefe.nombre;
    if(_presidentasActuales.length){
      var h = '<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+_presidentasActuales.length+' presidenta(s) para '+etiqueta+'</p>';
      h += '<div style="background:#0d2137;border-radius:8px;padding:8px;margin-bottom:10px">';
      h += '<p style="color:#7fb3d5;font-size:0.75em;margin:0 0 4px 2px">Mensaje o enlace a enviar por WhatsApp:</p>';
      h += '<input type="text" id="msgWaPresidentas" placeholder="Ej. enlace de TikTok..." style="margin:0">';
      h += '</div>';
      _presidentasActuales.forEach(function(p){
        h += '<div style="background:#154360;border-radius:8px;padding:10px;margin-bottom:6px">';
        h += '<p style="color:white;font-weight:bold;margin:0">'+(p.nombre||'-')+'</p>';
        if(esAdmin) h += '<p style="color:#f5b041;font-size:0.75em;margin:2px 0 0 0">Jefe: '+(p.jefe||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.8em;margin:4px 0 0 0">DPI: '+(p.cui||'-')+' &nbsp; Tel: '+(p.telefono||'-')+'</p>';
        h += '<p style="color:#aed6f1;font-size:0.8em;margin:2px 0 0 0">Dirección: '+(p.direccion||'-')+'</p>';
        if(p.telefono) h += '<button class="mbtn" style="background:#25D366;margin:8px 0 0 0;padding:8px;font-size:0.85em" data-tel="'+p.telefono+'" data-msgid="msgWaPresidentas" onclick="_abrirWhatsAppBtn(this)">📱 Enviar WhatsApp</button>';
        h += '</div>';
      });
      document.getElementById('contPresidentas').innerHTML = h;
      document.getElementById('botonesFijosPres').innerHTML =
        '<div style="display:flex;gap:8px">'+
        '<button class="mbtn" style="background:#8e44ad;margin:0;flex:1" onclick="generarPdfPresidentasSesion()">📄 PDF</button>'+
        '<button class="mbtn" style="background:#1e8449;margin:0;flex:1" onclick="generarExcelPresidentasSesion()">📊 Excel</button>'+
        '</div>';
    } else {
      document.getElementById('contPresidentas').innerHTML = '<p class="msg">No hay presidentas registradas para '+etiqueta+'</p>';
    }
  }catch(e){
    document.getElementById('contPresidentas').innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function generarExcelPresidentasSesion(){
  try{
    var r = await fetch('/excel_presidentas',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), presidentas: _presidentasActuales})});
    var d = await r.json();
    if(d.ok && d.excel_b64){
      var bytes=atob(d.excel_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'});
      var blobUrl=URL.createObjectURL(blob);
      var a=document.createElement('a'); a.href=blobUrl; a.download='Presidentas_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.xlsx';
      document.body.appendChild(a); a.click(); document.body.removeChild(a);
    } else {
      alert('❌ '+(d.error||'Error al generar Excel'));
    }
  }catch(e){ alert('❌ Error de conexión: '+e.message); }
}

async function generarPdfPresidentasSesion(){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  try{
    var r = await fetch('/pdf_presidentas',{method:'POST',headers:{'Content-Type':'application/json'},
      body:JSON.stringify({jefe_sector: (sesionJefe.esAdmin?'Todos los Jefes de Sector':sesionJefe.nombre), presidentas: _presidentasActuales})});
    var d = await r.json();
    if(d.ok && d.pdf_b64){
      var bytes=atob(d.pdf_b64);
      var arr=new Uint8Array(bytes.length);
      for(var i=0;i<bytes.length;i++) arr[i]=bytes.charCodeAt(i);
      var blob=new Blob([arr],{type:'application/pdf'});
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, 'Presidentas_'+(sesionJefe.esAdmin?'TODOS':sesionJefe.nombre.replace(/ /g,'_'))+'.pdf', d.token?('/pdf_temporal/'+d.token):null);
    } else {
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(d.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

function mostrarInfoSesion(){
  document.getElementById('sesionJefeInfo').textContent = sesionJefe ? (sesionJefe.esAdmin ? '👑 Sesión: ADMINISTRADOR (todos los jefes)' : 'Sesión: '+sesionJefe.nombre) : '';
}

async function hacerLoginJefe(){
  var nombre = document.getElementById('loginJefeSel').value;
  var pin = document.getElementById('loginJefePin').value.trim();
  if(!nombre){ msg('msgLoginJefe','Seleccione su nombre','#e74c3c'); return; }
  if(!pin){ msg('msgLoginJefe','Ingrese su PIN','#e74c3c'); return; }
  msg('msgLoginJefe','Verificando...','#aed6f1');
  try{
    var r = await fetch('/login_jefe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre:nombre, pin:pin})});
    var d = await r.json();
    if(d.ok){
      sesionJefe = {nombre: d.nombre, esAdmin: !!d.es_admin};
      actualizarBotonLogout();
      hide('oLoginJefe');
      var accion = _accionPendienteLogin; _accionPendienteLogin = null;
      if(accion) accion();
    } else {
      msg('msgLoginJefe','❌ '+(d.error||'No se pudo ingresar'),'#e74c3c');
    }
  }catch(e){ msg('msgLoginJefe','❌ '+e.message,'#e74c3c'); }
}

function cerrarSesionJefe(){
  sesionJefe = null;
  hide('oGrupos');
}

async function cargarGrupos(){
  tG={}; rJ={};
  document.getElementById('contGrupos').innerHTML='<p class="msg">Cargando...</p>';
  document.getElementById('jGrid').innerHTML='';
  show('oVerGrupos');
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var jefeParam = (sesionJefe && !esAdmin) ? ('?jefe='+encodeURIComponent(sesionJefe.nombre)) : '';
    var r=await fetch('/ver_grupos'+jefeParam);
    var d=await r.json();
    if(d.ok){ rJ=d.resumen_jefes||{}; d.grupos.forEach(function(g){ if(!tG[g.nombre]) tG[g.nombre]={nombre:g.nombre,coord:g.coordinador,jefe:g.jefe_sector||'',ps:[]}; tG[g.nombre].ps.push(g); }); }
    if(sesionJefe && !esAdmin){
      // Un jefe con sesión iniciada solo ve sus propios grupos: se salta
      // la selección de jefe y se va directo a su lista.
      filtJ(sesionJefe.nombre);
      return;
    }
    var JEFES = await obtenerJefes();
    var h='';
    JEFES.forEach(function(j){
      var c=rJ[j]||0;
      h+='<button class="jbtn '+(c>0?'tiene':'')+'" data-jefe="'+j+'">'+j+'<br><b>('+c+')</b></button>';
    });
    document.getElementById('jGrid').innerHTML=h;
    document.getElementById('jGrid').querySelectorAll('.jbtn').forEach(function(btn){
      btn.onclick=function(){ filtJ(this.getAttribute('data-jefe')); };
    });
    document.getElementById('contGrupos').innerHTML='<p class="msg">Seleccione un Jefe de Sector</p>';
  }catch(e){ document.getElementById('contGrupos').innerHTML='<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>'; }
}

function filtJ(j){
  document.getElementById('jGrid').querySelectorAll('.jbtn').forEach(function(b){ b.className='jbtn '+(rJ[b.getAttribute('data-jefe')]>0?'tiene':''); });
  var el=document.querySelector('[data-jefe="'+j+'"]');
  if(el) el.className='jbtn activo';
  var gs=Object.values(tG).filter(function(g){ return g.jefe===j; });
  if(!gs.length){ document.getElementById('contGrupos').innerHTML='<p class="msg">Sin grupos para '+j+'</p>'; return; }
  // Buscador + lista compacta scrolleable
  var h='<div style="margin-bottom:8px"><input id="busGrp" type="text" placeholder="🔍 Buscar grupo..." style="width:100%;padding:8px;border-radius:8px;border:none;background:#0d2137;color:white;font-size:0.9em" oninput="filtrarGrps()"><p style="color:#aed6f1;font-size:0.82em;margin-top:4px">'+gs.length+' grupo(s)</p></div>';
  h+='<div id="listaGrps" style="max-height:55vh;overflow-y:auto">';
  gs.forEach(function(g,idx){
    var ne=encodeURIComponent(g.nombre);
    h+='<div class="gitem" data-nombre-lower="'+g.nombre.toLowerCase()+'" style="background:#154360;border-radius:8px;padding:10px;margin-bottom:8px;display:flex;align-items:center;justify-content:space-between;gap:8px">';
    h+='<div style="flex:1;min-width:0">';
    h+='<p style="color:#2ecc71;font-weight:bold;font-size:0.85em;margin:0">👥 '+g.nombre+'</p>';
    h+='<p style="color:#aed6f1;font-size:0.75em;margin:2px 0">Coord: '+g.coord+'</p>';
    h+='<p style="color:#f0b27a;font-size:0.75em;margin:0">'+g.ps.length+'/10 personas</p>';
    h+='</div>';
    h+='<button data-nombre="'+ne+'" style="flex-shrink:0;padding:8px 12px;background:#8e44ad;color:white;border:none;border-radius:8px;font-weight:bold;font-size:0.8em;cursor:pointer" class="btnPDF">📄 PDF</button>';
    h+='</div>';
  });
  h+='</div>';
  document.getElementById('contGrupos').innerHTML=h;
  document.getElementById('contGrupos').querySelectorAll('.btnPDF').forEach(function(btn){
    btn.onclick=function(){ impGrupo(decodeURIComponent(this.getAttribute('data-nombre'))); };
  });
}

function filtrarGrps(){
  var q=document.getElementById('busGrp').value.toLowerCase();
  document.querySelectorAll('.gitem').forEach(function(el){
    el.style.display=el.getAttribute('data-nombre-lower').includes(q)?'flex':'none';
  });
}

async function impGrupo(nombre){
  try{
    var r=await fetch('/buscar_grupo?nombre='+encodeURIComponent(nombre));
    var d=await r.json();
    if(!d.ok){ alert('Error'); return; }
    await genPDFData(d.grupo.nombre,d.grupo.coordinador,d.grupo.dpi_coord,d.grupo.jefe_sector||'',d.personas,d.tiene_dir,d.grupo.presidenta||'');
  }catch(e){ alert('Error: '+e.message); }
}

async function genPDFData(nombre,coord,dpiC,jefe,filas,tieneDir,pres){
  document.getElementById('pdfLoadMsg').style.display='block';
  document.getElementById('pdfLoadMsg').textContent='Generando PDF...';
  document.getElementById('pdfReadyMsg').style.display='none';
  document.getElementById('pdfFrame').style.display='none';
  show('oPDFViewer');
  var tieneD=tieneDir!==false;
  var ps=filas.map(function(p){ return {primer_nombre:(p[5]||'').split(' ')[0]||'',segundo_nombre:(p[5]||'').split(' ')[1]||'',primer_apellido:(p[5]||'').split(' ')[2]||'',segundo_apellido:(p[5]||'').split(' ')[3]||'',cui:p[6]||'',direccion:tieneD?(p[7]||''):'',telefono:tieneD?(p[8]||''):(p[7]||''),empadronado:tieneD?(p[9]||'NO'):(p[8]||'NO')}; });
  try{
    var r=await fetch('/generar_pdf_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_grupo:nombre,coordinadora_nombre:coord,coordinadora_dpi:dpiC,jefe_sector:jefe,presidenta:pres||'',personas:ps})});
    var ct=r.headers.get('Content-Type')||'';
    if(r.ok && ct.indexOf('pdf')>=0){
      var token=r.headers.get('X-PDF-Token');
      var blob=await r.blob();
      var blobUrl=URL.createObjectURL(blob);
      mostrarPDFViewer(blobUrl, nombre.replace(/ /g,'_')+'.pdf', token?('/pdf_temporal/'+token):null);
    } else {
      var err=await r.json().catch(function(){return {error:'Error desconocido'};});
      document.getElementById('pdfLoadMsg').style.display='block';
      document.getElementById('pdfLoadMsg').textContent='❌ '+(err.error||'Error al generar PDF');
    }
  }catch(e){
    document.getElementById('pdfLoadMsg').style.display='block';
    document.getElementById('pdfLoadMsg').textContent='❌ Error de conexión: '+e.message;
  }
}

// PRESIDENTAS
async function cargarPresidentas(){
  var j=document.getElementById('ngJ').value;
  var sel=document.getElementById('ngP');
  if(!j){ sel.innerHTML='<option value="">-- Seleccione Jefe primero --</option>'; return; }
  sel.innerHTML='<option value="">Cargando...</option>';
  try{
    var r=await fetch('/ver_presidentas?jefe='+encodeURIComponent(j));
    var d=await r.json();
    if(d.ok && d.presidentas.length>0){
      sel.innerHTML='<option value="">-- Seleccione --</option>';
      d.presidentas.forEach(function(p){ sel.innerHTML+='<option value="'+p.nombre+'">'+p.nombre+'</option>'; });
    } else {
      sel.innerHTML='<option value="">Sin presidentas registradas</option>';
    }
  }catch(e){ sel.innerHTML='<option value="">Error cargando</option>'; }
}

// Abre el selector nativo de foto (Cámara / Galería / Archivos) usando el
// input compartido. "destino" indica a cuál de los 6 casos corresponde
// (presF/presR/coordF/coordR/persF/persR), para que el manejador de cambio
// del input sepa dónde guardar la foto una vez elegida.
//
// NOTA: se intentó dos veces agregar una verificación previa del permiso de
// cámara para avisar si estaba bloqueada, y en ambos intentos terminó
// rompiendo el botón en teléfonos reales (aunque las pruebas automatizadas
// no detectaban ningún problema). Por eso se decidió NO volver a intentarlo:
// es preferible la función básica funcionando siempre a un aviso extra que
// termina rompiendo lo importante. El propio selector nativo del teléfono
// ya gestiona el permiso de cámara por su cuenta al abrirse.
var _fotoDestinoActual = null;
function abrirSelectorFoto(destino){
  _fotoDestinoActual = destino;
  var input = document.getElementById('fotoInputCompartido');
  var esAdmin = sesionJefe && sesionJefe.esAdmin;
  if(esAdmin){
    // Administrador: conserva el menú completo (Galería / Cámara / Archivo),
    // útil para pruebas y para elegir fotos ya existentes.
    input.removeAttribute('capture');
  } else {
    // Jefes de sector normales: va directo a la cámara trasera, sin mostrar
    // el menú de opciones, para agilizar el escaneo de DPI en campo.
    input.setAttribute('capture', 'environment');
  }
  input.value = ''; // limpiar antes de abrir, para poder re-elegir la misma foto si hace falta
  input.click();
}

function abrirSelectorFotoAfiliacion(destino){
  // En 'Afiliación al Partido' la persona puede elegir con un check si
  // quiere ir directo a la cámara (por defecto, para agilizar en campo) o
  // desmarcarlo para poder elegir un archivo/foto ya existente.
  _fotoDestinoActual = destino;
  var input = document.getElementById('fotoInputCompartido');
  var camaraDirecta = document.getElementById('afCamaraDirecta').checked;
  if(camaraDirecta){
    input.setAttribute('capture', 'environment');
  } else {
    input.removeAttribute('capture');
  }
  input.value = '';
  input.click();
}

document.getElementById('fotoInputCompartido').onchange=async function(e){
  if(!e.target.files[0]) return;
  var datos = await b64(e.target.files[0]);
  if(_fotoDestinoActual==='presF'){ gPresF=datos; document.getElementById('bFpr').textContent='✅ Frente OK'; document.getElementById('bFpr').style.background='#27ae60'; chkPr(); }
  else if(_fotoDestinoActual==='presR'){ gPresR=datos; document.getElementById('bRpr').textContent='✅ Reverso OK'; document.getElementById('bRpr').style.background='#27ae60'; chkPr(); }
  else if(_fotoDestinoActual==='coordF'){ gCF=datos; document.getElementById('bFc').textContent='✅ Frente OK'; document.getElementById('bFc').style.background='#27ae60'; chkC(); }
  else if(_fotoDestinoActual==='coordR'){ gCR=datos; document.getElementById('bRc').textContent='✅ Reverso OK'; document.getElementById('bRc').style.background='#27ae60'; chkC(); }
  else if(_fotoDestinoActual==='persF'){ gPF=datos; document.getElementById('bFp').textContent='✅ Frente OK'; document.getElementById('bFp').style.background='#27ae60'; chkP(); }
  else if(_fotoDestinoActual==='persR'){ gPR=datos; document.getElementById('bRp').textContent='✅ Reverso OK'; document.getElementById('bRp').style.background='#27ae60'; chkP(); }
  else if(_fotoDestinoActual==='pruebaF'){ gTPF=datos; document.getElementById('btnTpF').textContent='✅ Frente OK'; document.getElementById('btnTpF').style.background='#27ae60'; chkTP(); }
  else if(_fotoDestinoActual==='pruebaR'){ gTPR=datos; document.getElementById('btnTpR').textContent='✅ Reverso OK'; document.getElementById('btnTpR').style.background='#27ae60'; chkTP(); }
  else if(_fotoDestinoActual==='afiliacionF'){ gAfF=datos; document.getElementById('btnAfF').textContent='✅ Frente OK'; document.getElementById('btnAfF').style.background='#27ae60'; chkAfil(); }
  else if(_fotoDestinoActual==='afiliacionR'){ gAfR=datos; document.getElementById('btnAfR').textContent='✅ Reverso OK'; document.getElementById('btnAfR').style.background='#27ae60'; chkAfil(); }
};
function chkPr(){ if(gPresF&&gPresR){ document.getElementById('bProcPr').disabled=false; document.getElementById('bProcPr').style.background='#154360'; } }
function chkTP(){ if(gTPF&&gTPR){ document.getElementById('btnTpProc').disabled=false; document.getElementById('btnTpProc').style.background='#8e44ad'; } }
function chkAfil(){ if(gAfF&&gAfR){ document.getElementById('btnAfProc').disabled=false; document.getElementById('btnAfProc').style.background='#7d3c98'; } }

async function verRegistrosPrueba(){
  var cont = document.getElementById('listaRegPrueba');
  cont.innerHTML = '<p class="msg">Cargando...</p>';
  try{
    var r = await fetch('/listar_registros_prueba');
    var d = await r.json();
    var regs = (d.ok && d.registros) ? d.registros : [];
    if(!regs.length){ cont.innerHTML = '<p class="msg">No hay registros de prueba guardados</p>'; return; }
    var h = '<p style="color:#aed6f1;font-size:0.78em;margin-bottom:6px">'+regs.length+' registro(s) de prueba — toca el DPI para copiarlo</p>';
    regs.forEach(function(reg){
      h += '<div style="background:#2c1a3d;border-radius:8px;padding:9px;margin-bottom:6px;display:flex;align-items:center;justify-content:space-between;gap:8px">';
      h += '<div style="flex:1;min-width:0">';
      h += '<p style="color:white;font-size:0.82em;margin:0">'+reg.nombre+'</p>';
      h += '<p class="dpi-prueba-copy" data-cui="'+reg.cui+'" style="color:#d2b4de;font-size:0.78em;margin:2px 0 0 0;cursor:pointer;text-decoration:underline">DPI: '+reg.cui+' 📋</p>';
      if(reg.fecha_creacion) h += '<p style="color:#aed6f1;font-size:0.68em;margin:2px 0 0 0">Agregado: '+reg.fecha_creacion+'</p>';
      h += '</div>';
      h += '<button class="btn-del-prueba" data-fila="'+reg.fila+'" data-cui="'+reg.cui+'" style="background:#922b21;color:white;border:none;border-radius:8px;padding:8px 12px;font-size:0.78em;font-weight:bold;cursor:pointer">🗑️</button>';
      h += '</div>';
    });
    cont.innerHTML = h;
    cont.querySelectorAll('.dpi-prueba-copy').forEach(function(el){
      el.onclick = function(){
        var cui = this.getAttribute('data-cui');
        if(navigator.clipboard){ navigator.clipboard.writeText(cui); }
        var original = this.textContent;
        this.textContent = '✅ DPI copiado: '+cui;
        var self = this;
        setTimeout(function(){ self.textContent = original; }, 1500);
      };
    });
    cont.querySelectorAll('.btn-del-prueba').forEach(function(btn){
      btn.onclick = async function(){
        var fila = this.getAttribute('data-fila');
        var cui = this.getAttribute('data-cui');
        if(!confirm('¿Eliminar este registro de prueba (DPI: '+cui+')?')) return;
        try{
          var r = await fetch('/eliminar_dpi',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({numero_fila:fila, cui:cui})});
          var d = await r.json();
          if(d.ok){ verRegistrosPrueba(); }
          else { alert('❌ '+(d.error||'Error al eliminar')); }
        }catch(e){ alert('❌ '+e.message); }
      };
    });
  }catch(e){
    cont.innerHTML = '<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>';
  }
}

async function procRegistroPrueba(){
  var btn = document.getElementById('btnTpProc');
  btn.disabled=true; btn.textContent='Analizando...'; msg('msgTP','Analizando el DPI...','#aed6f1');
  try{
    var r = await fetch('/analizar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({frente:gTPF,reverso:gTPR,frente_type:'image/jpeg',reverso_type:'image/jpeg'})});
    var d = await r.json();
    if(!d.ok){
      msg('msgTP','❌ '+(d.error||'Error al analizar'),'#e74c3c');
      btn.disabled=false; btn.textContent='Analizar y guardar como prueba'; return;
    }
    var datos = d.datos;
    if(!datos.fecha_nacimiento || !datos.fecha_nacimiento.trim()){
      msg('msgTP','❌ Este DPI no tiene fecha de nacimiento legible. Vuelve a tomar la foto e inténtalo de nuevo.','#e74c3c');
      btn.disabled=false; btn.textContent='Analizar y guardar como prueba'; return;
    }
    datos.telefono = document.getElementById('tpTel').value.trim();
    datos.direccion = document.getElementById('tpDir').value.trim();
    if(!datos.direccion){
      msg('msgTP','❌ La dirección / comunidad es obligatoria. Escríbela antes de guardar.','#e74c3c');
      btn.disabled=false; btn.textContent='Analizar y guardar como prueba'; return;
    }
    datos.dpi_no_verificado = !!d.no_autentico;
    msg('msgTP','Guardando...','#aed6f1');
    var r2 = await fetch('/guardar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({datos:datos})});
    var d2 = await r2.json();
    if(d2.ok){
      msg('msgTP','✅ Guardado como prueba: '+(d2.nombre||'')+' — buscalo con DPI: '+(datos.cui||''),'#2ecc71');
      avisarVozFaltantes(datos.telefono, d2.empadronado);
      gTPF=null; gTPR=null;
      document.getElementById('btnTpF').textContent='📷 Foto/Archivo Frente'; document.getElementById('btnTpF').style.background='#8e44ad';
      document.getElementById('btnTpR').textContent='📷 Foto/Archivo Reverso'; document.getElementById('btnTpR').style.background='#8e44ad';
      document.getElementById('tpTel').value=''; document.getElementById('tpDir').value='';
      verRegistrosPrueba();
    } else {
      msg('msgTP','❌ '+(d2.error||'Error al guardar'),'#e74c3c');
    }
    btn.disabled=true; btn.textContent='Analizar y guardar como prueba'; btn.style.background='#566573';
  }catch(e){
    msg('msgTP','❌ '+e.message,'#e74c3c');
    btn.disabled=false; btn.textContent='Analizar y guardar como prueba';
  }
}

// ==== VERIFICACIÓN MANUAL EN EL TSE ====
// Cuando la persona no aparece en el padrón local (hoja PADRON), se pausa el
// registro y se pide verificar manualmente en el sitio del TSE (requiere
// CAPTCHA, por lo que no se puede automatizar). Una vez que la persona
// revisa ahí y confirma el resultado (con o sin número de empadronamiento),
// el flujo de registro continúa normalmente con ese dato.
var _accionPendienteTSE = null;

function _verificarEmpadronamiento(d, nombreCompleto, callback){
  if(d.empadronado){
    callback('SI', d.num_empadronamiento||'');
    return;
  }
  document.getElementById('tseNombre').textContent = nombreCompleto;
  document.getElementById('tseCUI').textContent = 'DPI: '+(d.datos.cui||'-');
  document.getElementById('tseNumEmp').value = '';
  msg('msgTSE','','');
  _accionPendienteTSE = callback;
  show('oTSE');
}

function abrirTSE(){
  window.open('https://consultaempadronamiento.tse.org.gt/', '_blank');
}

function confirmarTSE(estado){
  var numEmp = document.getElementById('tseNumEmp').value.trim();
  if(estado==='SI' && !numEmp){
    msg('msgTSE','Si está empadronada, ingresa el número que te mostró el TSE (o toca "No está empadronado" si no aparece)','#f0b27a');
    return;
  }
  hide('oTSE');
  var callback = _accionPendienteTSE;
  _accionPendienteTSE = null;
  if(callback) callback(estado, estado==='SI'?numEmp:'');
}

// ==== ADVERTENCIA DE DPI POSIBLEMENTE NO AUTÉNTICO ====
// Cuando la IA detecta que la foto no parece ser un DPI físico real (por
// ejemplo, una nota escrita a mano), no se bloquea de una vez — se muestra
// esta advertencia y la persona decide si continúa o prefiere tomar otra
// foto. Si decide continuar, el registro queda marcado como no verificado
// para poder resaltarlo después en los reportes impresos.
var _accionPendienteAdvertencia = null;
var _accionCancelarAdvertencia = null;

function _manejarPosibleNoAutentico(d, callbackContinuar, callbackCancelar){
  if(d.no_autentico){
    document.getElementById('motivoAdvertenciaDPI').textContent = d.motivo_rechazo || '';
    _accionPendienteAdvertencia = callbackContinuar;
    _accionCancelarAdvertencia = callbackCancelar;
    show('oAdvertenciaDPI');
  } else {
    callbackContinuar();
  }
}

function _continuarPeseAdvertencia(){
  hide('oAdvertenciaDPI');
  if(_accionPendienteAdvertencia) _accionPendienteAdvertencia();
}

function _salirPorAdvertencia(){
  hide('oAdvertenciaDPI');
  if(_accionCancelarAdvertencia) _accionCancelarAdvertencia();
}

async function procPresidenta(){
  var btn=document.getElementById('bProcPr');
  btn.disabled=true; btn.textContent='Analizando...'; msg('msgPr','Analizando...','#aed6f1');
  try{
    var r=await fetch('/analizar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({frente:gPresF,reverso:gPresR,frente_type:'image/jpeg',reverso_type:'image/jpeg'})});
    var d=await r.json();
    if(d.ok){
      _manejarPosibleNoAutentico(d, function(){
        var n=(d.datos.primer_nombre||'')+' '+(d.datos.segundo_nombre||'')+' '+(d.datos.primer_apellido||'')+' '+(d.datos.segundo_apellido||'');
        _verificarEmpadronamiento(d, n.trim(), function(estadoEmp, numEmp){
          gPres=d.datos;
          gPres.telefono=document.getElementById('prTel').value.trim();
          gPres.direccion=document.getElementById('prDir').value.trim();
          gPres.empadronado=estadoEmp;
          gPres.num_empadronamiento=numEmp;
          gPres.jefe=document.getElementById('ngJ').value;
          gPres.nombre=n.trim();
          gPres.dpi_no_verificado=!!d.no_autentico;
          document.getElementById('datPr').innerHTML='<b>Nombre:</b> '+n.trim()+'<br><b>DPI:</b> '+(gPres.cui||'-')+'<br><b>Tel:</b> '+(gPres.telefono||'-')+'<br><b>Emp:</b> '+(gPres.empadronado||'-');
          hide('oPresidenta'); show('oConfPr');
          btn.disabled=false; btn.textContent='Procesar DPI Presidenta';
        });
      }, function(){
        btn.disabled=false; btn.textContent='Procesar DPI Presidenta'; msg('msgPr','','');
      });
    } else { msg('msgPr','❌ '+(d.error||'Error'),'#e74c3c'); btn.disabled=false; btn.textContent='Procesar DPI Presidenta'; }
  }catch(e){ msg('msgPr','❌ '+e.message,'#e74c3c'); btn.disabled=false; btn.textContent='Procesar DPI Presidenta'; }
}

// Aviso por voz automático cuando se guarda a alguien sin teléfono y/o sin
// empadronar — usa la síntesis de voz que ya trae el navegador (sin costo,
// sin necesitar internet extra ni ningún servicio externo). Así quien está
// registrando lo escucha de inmediato, sin tener que estar leyendo la
// pantalla mientras sigue tomando fotos.
function avisarVozFaltantes(telefono, empadronado){
  var faltaTel = !telefono || !telefono.trim();
  var faltaEmp = (empadronado||'NO').trim().toUpperCase() !== 'SI';
  if(!faltaTel && !faltaEmp) return;
  if(!('speechSynthesis' in window)) return;
  var texto;
  if(faltaTel && faltaEmp) texto = 'Atención. Esta persona no tiene teléfono y no está empadronada. Sin eso la estrategia no funciona.';
  else if(faltaTel) texto = 'Atención. Esta persona no tiene número de teléfono. Sin eso la estrategia no funciona.';
  else texto = 'Atención. Esta persona no está empadronada. Sin eso la estrategia no funciona.';
  try{
    window.speechSynthesis.cancel();
    var u = new SpeechSynthesisUtterance(texto);
    u.lang = 'es-GT';
    u.rate = 1;
    window.speechSynthesis.speak(u);
  }catch(e){}
}

async function confirmarPresidenta(){
  if(!gPres.fecha_nacimiento || !gPres.fecha_nacimiento.trim()){
    msg('msgConfPr','❌ Este DPI no tiene fecha de nacimiento legible. No se puede guardar el registro sin ese dato — vuelve a tomar la foto (de preferencia más nítida) e inténtalo de nuevo.','#e74c3c');
    return;
  }
  if(!gPres.direccion || !gPres.direccion.trim()){
    msg('msgConfPr','❌ La dirección / comunidad es obligatoria. Regresa y escríbela antes de continuar.','#e74c3c');
    return;
  }
  try{
    var r = await fetch('/registrar_presidenta',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({jefe_sector:gPres.jefe,presidenta:gPres})});
    var d = await r.json();
    if(!d.ok){
      msg('msgConfPr', d.error || 'No se pudo registrar', '#f0b27a');
      return;
    }
    hide('oConfPr');
    show('oNuevoG1');
    avisarVozFaltantes(gPres.telefono, gPres.empadronado);
    await cargarPresidentas();
    // Seleccionar la presidenta recién registrada
    var sel=document.getElementById('ngP');
    for(var i=0;i<sel.options.length;i++){ if(sel.options[i].value===gPres.nombre){ sel.selectedIndex=i; break; } }
    msg('msgPr','','');
    msg('msgConfPr','','');
  }catch(e){ msg('msgConfPr','Error: '+e.message,'#e74c3c'); }
}

// NUEVO GRUPO
function irCoord(){
  gN=document.getElementById('ngN').value.trim();
  gJ=document.getElementById('ngJ').value;
  var selPres=document.getElementById('ngP').value;
  if(!gN){ alert('Ingrese nombre del grupo'); return; }
  if(!gJ){ alert('Seleccione Jefe de Sector'); return; }
  if(!gPres && !selPres){ alert('Seleccione o registre una Presidenta de Comité'); return; }
  if(!gPres && selPres){ gPres={nombre:selPres,cui:'',telefono:'',direccion:''}; }
  hide('oNuevoG1');
  gCF=null; gCR=null;
  document.getElementById('bFc').textContent='📷 Foto Frente'; document.getElementById('bFc').style.background='#2980b9';
  document.getElementById('bRc').textContent='📷 Foto Reverso'; document.getElementById('bRc').style.background='#2980b9';
  document.getElementById('bProcC').disabled=true; document.getElementById('bProcC').style.background='#566573';
  document.getElementById('fcTel').value=''; msg('msgC','','');
  document.getElementById('infoCoord').textContent='Grupo: '+gN+' | Jefe: '+gJ+' | Pres: '+(gPres?gPres.nombre:'');
  show('oCoord');
}

function chkC(){ if(gCF&&gCR){ document.getElementById('bProcC').disabled=false; document.getElementById('bProcC').style.background='#154360'; } }

async function procCoord(){
  var btn=document.getElementById('bProcC');
  btn.disabled=true; btn.textContent='Analizando...'; msg('msgC','Analizando...','#aed6f1');
  try{
    var r=await fetch('/analizar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({frente:gCF,reverso:gCR,frente_type:'image/jpeg',reverso_type:'image/jpeg'})});
    var d=await r.json();
    if(d.ok){
      _manejarPosibleNoAutentico(d, async function(){
        var vr=await fetch('/validar_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_grupo:gN,dpi_coord:d.datos.cui||''})});
        var vd=await vr.json();
        if(!vd.ok){ msg('msgC','❌ '+vd.error,'#e74c3c'); btn.disabled=false; btn.textContent='Procesar DPI Coordinadora'; return; }
        var n=(d.datos.primer_nombre||'')+' '+(d.datos.segundo_nombre||'')+' '+(d.datos.primer_apellido||'')+' '+(d.datos.segundo_apellido||'');
        _verificarEmpadronamiento(d, n.trim(), function(estadoEmp, numEmp){
          gC=d.datos; gC.telefono=document.getElementById('fcTel').value.trim(); gC.direccion=document.getElementById('fcDir').value.trim(); gC.empadronado=estadoEmp; gC.num_empadronamiento=numEmp; gC.es_coordinador='SI'; gC.dpi_no_verificado=!!d.no_autentico;
          document.getElementById('datC').innerHTML='<b>Nombre:</b> '+n.trim()+'<br><b>DPI:</b> '+(gC.cui||'-')+'<br><b>Tel:</b> '+(gC.telefono||'-');
          hide('oCoord'); show('oConfC');
          btn.disabled=false; btn.textContent='Procesar DPI Coordinadora';
        });
      }, function(){
        btn.disabled=false; btn.textContent='Procesar DPI Coordinadora'; msg('msgC','','');
      });
    } else { msg('msgC','❌ '+(d.error||'Error'),'#e74c3c'); btn.disabled=false; btn.textContent='Procesar DPI Coordinadora'; }
  }catch(e){ msg('msgC','❌ '+e.message,'#e74c3c'); btn.disabled=false; btn.textContent='Procesar DPI Coordinadora'; }
}

function _mostrarDuplicadoCoord(rd, cn){
  var elMsg=document.getElementById('msgConfC');
  var cuiDup = rd.cui || (gC?gC.cui:'') || '';
  var txt = 'PERSONA DUPLICADA<br><br>Nombre: '+(rd.nombre||cn)+'<br>CUI: '+cuiDup+'<br><br>'+(rd.error||'Esta persona ya está registrada.')+
    (cuiDup?'<br><br><button class="btnPDF" data-cui="'+cuiDup+'" style="padding:8px 14px;background:#8e44ad;color:white;border:none;border-radius:8px;font-weight:bold;cursor:pointer">PDF registro individual</button>':'');
  elMsg.style.color='#f0b27a';
  elMsg.innerHTML=txt;
}

async function continuarPersonas(){
  var cn=(gC.primer_nombre||'')+' '+(gC.segundo_nombre||'')+' '+(gC.primer_apellido||'')+' '+(gC.segundo_apellido||'');
  if(!gC.fecha_nacimiento || !gC.fecha_nacimiento.trim()){
    msg('msgConfC','❌ Este DPI no tiene fecha de nacimiento legible. No se puede guardar el registro sin ese dato — vuelve a tomar la foto (de preferencia más nítida) e inténtalo de nuevo.','#e74c3c');
    return;
  }
  if(!gC.direccion || !gC.direccion.trim()){
    msg('msgConfC','❌ La dirección / comunidad es obligatoria. Regresa y escríbela antes de continuar.','#e74c3c');
    return;
  }
  var resp = await fetch('/guardar_persona_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_grupo:gN,coord_nombre:cn.trim(),coord_dpi:gC.cui||'',jefe_sector:gJ,presidenta:gPres?gPres.nombre:'',presidenta_dpi:gPres?gPres.cui||'':'',jefe_autenticado:sesionJefe?sesionJefe.nombre:'',persona:gC,numero:1})});
  var rd = await resp.json();
  if(!rd.ok){
    _mostrarDuplicadoCoord(rd, cn.trim());
    return;
  }
  hide('oConfC');
  msg('msgConfC','','');
  avisarVozFaltantes(gC.telefono, gC.empadronado);
  gPs=[]; gPF=null; gPR=null;
  resetFP();
  document.getElementById('infoGP').textContent='Grupo: '+gN+' | Pres: '+(gPres?gPres.nombre:'')+' | Jefe: '+gJ;
  document.getElementById('cntG').textContent='1/10';
  document.getElementById('listaG').innerHTML='';
  document.getElementById('bPDF').style.display='none';
  document.getElementById('pasoF').style.display='block';
  msg('msgP','','');
  actLista();
  show('oPersonas');
}

async function pdfSoloCoord(){
  var cn=(gC.primer_nombre||'')+' '+(gC.segundo_nombre||'')+' '+(gC.primer_apellido||'')+' '+(gC.segundo_apellido||'');
  if(!gC.fecha_nacimiento || !gC.fecha_nacimiento.trim()){
    msg('msgConfC','❌ Este DPI no tiene fecha de nacimiento legible. No se puede guardar el registro sin ese dato — vuelve a tomar la foto (de preferencia más nítida) e inténtalo de nuevo.','#e74c3c');
    return;
  }
  if(!gC.direccion || !gC.direccion.trim()){
    msg('msgConfC','❌ La dirección / comunidad es obligatoria. Regresa y escríbela antes de continuar.','#e74c3c');
    return;
  }
  var resp = await fetch('/guardar_persona_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_grupo:gN,coord_nombre:cn.trim(),coord_dpi:gC.cui||'',jefe_sector:gJ,presidenta:gPres?gPres.nombre:'',presidenta_dpi:gPres?gPres.cui||'':'',jefe_autenticado:sesionJefe?sesionJefe.nombre:'',persona:gC,numero:1})});
  var rd = await resp.json();
  if(!rd.ok){
    _mostrarDuplicadoCoord(rd, cn.trim());
    return;
  }
  hide('oConfC');
  msg('msgConfC','','');
  avisarVozFaltantes(gC.telefono, gC.empadronado);
  await genPDFData(gN,cn.trim(),gC.cui||'',gJ,[{5:cn.trim(),6:gC.cui||'',7:gC.direccion||'',8:gC.telefono||'',9:gC.empadronado||'NO'}],true,gPres?gPres.nombre:'');
}

function chkP(){ if(gPF&&gPR){ document.getElementById('bProcP').disabled=false; document.getElementById('bProcP').style.background='#154360'; } }

async function procPersona(){
  if(gPs.length>=9){ msg('msgP','Grupo completo 10/10','#e74c3c'); return; }
  var btn=document.getElementById('bProcP');
  btn.disabled=true; btn.textContent='Analizando...'; msg('msgP','Analizando...','#aed6f1');
  try{
    var r=await fetch('/analizar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({frente:gPF,reverso:gPR,frente_type:'image/jpeg',reverso_type:'image/jpeg'})});
    var d=await r.json();
    if(d.ok){
      _manejarPosibleNoAutentico(d, function(){
      var nombreParaTSE=(d.datos.primer_nombre||'')+' '+(d.datos.segundo_nombre||'')+' '+(d.datos.primer_apellido||'')+' '+(d.datos.segundo_apellido||'');
      _verificarEmpadronamiento(d, nombreParaTSE.trim(), async function(estadoEmp, numEmp){
      var p=d.datos; p.telefono=document.getElementById('fpTel').value.trim(); p.direccion=document.getElementById('fpDir').value.trim(); p.empadronado=estadoEmp; p.num_empadronamiento=numEmp; p.es_coordinador='NO'; p.dpi_no_verificado=!!d.no_autentico;
      var cn=(gC.primer_nombre||'')+' '+(gC.primer_apellido||'');
      if(!p.fecha_nacimiento || !p.fecha_nacimiento.trim()){
        msg('msgP','❌ Este DPI no tiene fecha de nacimiento legible. No se puede guardar el registro sin ese dato — vuelve a tomar la foto (de preferencia más nítida) e inténtalo de nuevo.','#e74c3c');
        btn.disabled=false; btn.textContent='Agregar al grupo'; return;
      }
      if(!p.direccion){
        msg('msgP','❌ La dirección / comunidad es obligatoria. Escríbela antes de agregar a esta persona.','#e74c3c');
        btn.disabled=false; btn.textContent='Agregar al grupo'; return;
      }
      var resp=await fetch('/guardar_persona_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_grupo:gN,coord_nombre:cn.trim(),coord_dpi:gC.cui||'',jefe_sector:gJ,presidenta:gPres?gPres.nombre:'',presidenta_dpi:gPres?gPres.cui||'':'',jefe_autenticado:sesionJefe?sesionJefe.nombre:'',persona:p,numero:gPs.length+2})});
      var rd=await resp.json();
      if(!rd.ok){
        if(rd.duplicado){
          var elMsg=document.getElementById('msgP');
          if(elMsg){
            var cuiDup=rd.cui||p.cui||'';
            var lineaGrupo = rd.mismo_grupo===false ? ('<br><br><b>⚠️ Esta persona ya está registrada en OTRO grupo:</b> '+(rd.grupo_duplicado||'-')) : '<br><br>Esta persona ya esta registrada en este grupo.';
            var txt='PERSONA DUPLICADA<br><br>Grupo: '+gN+'<br>Jefe de Sector: '+gJ+'<br>Presidenta de Comite: '+(gPres?gPres.nombre:'-')+'<br>Coordinadora: '+cn.trim()+'<br><br>Nombre: '+(rd.nombre||'')+'<br>CUI: '+cuiDup+'<br>Dirección: '+(p.direccion||'-')+lineaGrupo+(cuiDup?'<br><br><button class="btnPDF" data-cui="'+cuiDup+'" style="padding:8px 14px;background:#8e44ad;color:white;border:none;border-radius:8px;font-weight:bold;cursor:pointer">PDF registro individual</button>':'');
            elMsg.style.color='#f0b27a';
            elMsg.innerHTML=txt;
          }
        } else { msg('msgP',rd.error||'Error','#e74c3c'); }
        btn.disabled=false; btn.textContent='Agregar al grupo'; return;
      }
      gPs.push(p);
      resetFP(); actLista();
      var nt=gPs.length+1;
      document.getElementById('cntG').textContent=nt+'/10';
      msg('msgP','✅ Persona '+nt+' guardada','#2ecc71');
      avisarVozFaltantes(p.telefono, p.empadronado);
      if(nt>=10){ document.getElementById('bPDF').style.display='block'; document.getElementById('pasoF').style.display='none'; }
      btn.disabled=false; btn.textContent='Agregar al grupo';
      });
      }, function(){
        btn.disabled=false; btn.textContent='Agregar al grupo'; msg('msgP','','');
      });
    } else { msg('msgP','❌ '+(d.error||'Error'),'#e74c3c'); btn.disabled=false; btn.textContent='Agregar al grupo'; }
  }catch(e){ msg('msgP','❌ '+e.message,'#e74c3c'); btn.disabled=false; btn.textContent='Agregar al grupo'; }
}

function resetFP(){
  gPF=null; gPR=null;
  document.getElementById('fotoInputCompartido').value='';
  document.getElementById('bFp').textContent='📷 Foto Frente'; document.getElementById('bFp').style.background='#2980b9';
  document.getElementById('bRp').textContent='📷 Foto Reverso'; document.getElementById('bRp').style.background='#2980b9';
  document.getElementById('fpTel').value='';
  document.getElementById('fpDir').value='';
  document.getElementById('bProcP').disabled=true; document.getElementById('bProcP').style.background='#566573'; document.getElementById('bProcP').textContent='Agregar al grupo';
}

function actLista(){
  var h='';
  if(gC){ var nc=(gC.primer_nombre||'')+' '+(gC.primer_apellido||''); h+='<div style="background:#1a5276;border:1px solid #2ecc71;border-radius:7px;padding:7px;margin-bottom:5px;color:white;font-size:0.82em">👑 1. '+nc.trim()+' (Coord)</div>'; }
  gPs.forEach(function(p,i){ var np=(p.primer_nombre||'')+' '+(p.primer_apellido||''); h+='<div style="background:#154360;border-radius:7px;padding:7px;margin-bottom:5px;color:white;font-size:0.82em">'+(i+2)+'. '+np.trim()+'</div>'; });
  document.getElementById('listaG').innerHTML=h;
}

async function genPDF(){
  var cn=(gC?(gC.primer_nombre||'')+' '+(gC.segundo_nombre||'')+' '+(gC.primer_apellido||'')+' '+(gC.segundo_apellido||''):'').trim();
  var todas=gC?[gC].concat(gPs):gPs;
  await genPDFData(gN,cn,gC?gC.cui||'':'',gJ,todas.map(function(p){ return {5:(p.primer_nombre||'')+' '+(p.segundo_nombre||'')+' '+(p.primer_apellido||'')+' '+(p.segundo_apellido||''),6:p.cui||'',7:p.direccion||'',8:p.telefono||'',9:p.empadronado||'NO'}; }),true,gPres?gPres.nombre:'');
}

// BUSCAR GRUPO EXISTENTE
async function buscarGE(){
  var n=document.getElementById('bgN').value.trim();
  var d=document.getElementById('bgD').value.trim();
  if(!n&&!d){ msg('msgBG','Ingrese nombre o DPI','#e74c3c'); return; }
  msg('msgBG','Buscando...','#aed6f1');
  document.getElementById('resBG').style.display='none';
  try{
    var p=n?'nombre='+encodeURIComponent(n):'';
    if(d) p+=(p?'&':'')+'dpi='+d;
    if(sesionJefe && !sesionJefe.esAdmin) p+=(p?'&':'')+'jefe_filtro='+encodeURIComponent(sesionJefe.nombre);
    var r=await fetch('/buscar_grupo?'+p);
    var rd=await r.json();
    if(rd.ok&&rd.grupo){
      gE=rd;
      document.getElementById('datBG').innerHTML='<b>Grupo:</b> '+rd.grupo.nombre+'<br><b>Coord:</b> '+rd.grupo.coordinador+'<br><b>Jefe:</b> '+(rd.grupo.jefe_sector||'-')+'<br><b>Personas:</b> '+rd.personas.length+'/10';
      document.getElementById('bContG').style.display=rd.personas.length>=10?'none':'block';
      document.getElementById('resBG').style.display='block';
      msg('msgBG','','');
    } else { msg('msgBG','❌ No encontrado','#e74c3c'); }
  }catch(e){ msg('msgBG','❌ '+e.message,'#e74c3c'); }
}

function contGrupoE(){
  if(!gE) return;
  gN=gE.grupo.nombre; gJ=gE.grupo.jefe_sector||'';
  gC={primer_nombre:gE.grupo.coordinador,primer_apellido:'',segundo_nombre:'',segundo_apellido:'',cui:gE.grupo.dpi_coord,telefono:'',empadronado:'NO',es_coordinador:'SI'};
  gPres=gE.grupo.presidenta?{nombre:gE.grupo.presidenta,cui:gE.grupo.dpi_presidenta||''}:null;
  // Reconstruir gPs con los integrantes YA existentes en el grupo (todas las
  // filas menos la de la coordinadora, que es la No.=1), para que el
  // contador y la lista reflejen la realidad, no solo lo que se agregue en
  // esta sesión. Antes se reiniciaba vacío y el contador se desincronizaba
  // del total real, permitiendo intentar agregar más allá del límite.
  gPs = gE.personas.filter(function(p){ return p[4] !== '1'; }).map(function(p){
    return {primer_nombre:p[5]||'', segundo_nombre:'', primer_apellido:'', segundo_apellido:'', cui:p[6]||'', direccion:p[7]||'', telefono:p[8]||'', empadronado:p[9]||'NO'};
  });
  hide('oBuscarG'); resetFP();
  document.getElementById('infoGP').textContent='Grupo: '+gN+' | Pres: '+(gPres?gPres.nombre:'-')+' | Jefe: '+gJ;
  document.getElementById('cntG').textContent=(gPs.length+1)+'/10';
  document.getElementById('bPDF').style.display='none';
  document.getElementById('pasoF').style.display=(gPs.length+1)>=10?'none':'block';
  msg('msgP','','');
  actLista(); show('oPersonas');
}

async function pdfGrupoE(){
  if(!gE) return;
  await genPDFData(gE.grupo.nombre,gE.grupo.coordinador,gE.grupo.dpi_coord,gE.grupo.jefe_sector||'',gE.personas,gE.tiene_dir,gE.grupo.presidenta||'');
}

// MANTENIMIENTO
var mCUI = null;
var mFilaCompleta = null;
var _editTelCui = null;

async function buscarParaEditarTel(){
  var q=document.getElementById('busEditTel').value.trim();
  if(!q){ msg('msgEditTel','Ingrese un nombre o CUI','#e74c3c'); return; }
  msg('msgEditTel','Buscando...','#aed6f1');
  document.getElementById('resEditTel').style.display='none';
  try{
    var r=await fetch('/buscar_dpi?q='+encodeURIComponent(q));
    var d=await r.json();
    if(d.ok&&d.fila){
      _editTelCui=(d.fila[0]||'').trim();
      var nombreCompleto=[d.fila[2],d.fila[3],d.fila[4],d.fila[5]].filter(Boolean).join(' ');
      document.getElementById('datEditTel').innerHTML='<b>Nombre:</b> '+(nombreCompleto||'-')+'<br><b>CUI:</b> '+_editTelCui+'<br><b>Teléfono actual:</b> '+(d.fila[18]||'-');
      document.getElementById('nuevoTel').value='';
      document.getElementById('resEditTel').style.display='block';
      msg('msgEditTel','','');
    } else { msg('msgEditTel','❌ No encontrado','#e74c3c'); }
  }catch(e){ msg('msgEditTel','❌ '+e.message,'#e74c3c'); }
}

async function guardarNuevoTel(){
  var tel=document.getElementById('nuevoTel').value.trim();
  if(!tel){ msg('msgEditTel','Ingrese el nuevo teléfono','#e74c3c'); return; }
  if(!_editTelCui){ msg('msgEditTel','Vuelva a buscar','#e74c3c'); return; }
  msg('msgEditTel','Guardando...','#aed6f1');
  try{
    var r=await fetch('/actualizar_telefono',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cui:_editTelCui, telefono:tel})});
    var d=await r.json();
    if(d.ok){
      msg('msgEditTel','✅ Teléfono actualizado','#2ecc71');
      document.getElementById('resEditTel').style.display='none';
      document.getElementById('busEditTel').value='';
      _editTelCui=null;
    } else { msg('msgEditTel','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgEditTel','❌ '+e.message,'#e74c3c'); }
}

var _editEmpCui = null;

async function buscarParaEditarEmp(){
  var q=document.getElementById('busEditEmp').value.trim();
  if(!q){ msg('msgEditEmp','Ingrese el número de CUI','#e74c3c'); return; }
  msg('msgEditEmp','Buscando...','#aed6f1');
  document.getElementById('resEditEmp').style.display='none';
  try{
    var r=await fetch('/buscar_dpi?cui='+encodeURIComponent(q));
    var d=await r.json();
    if(d.ok&&d.fila){
      _editEmpCui=(d.fila[0]||'').trim();
      var nombreCompleto=[d.fila[2],d.fila[3],d.fila[4],d.fila[5]].filter(Boolean).join(' ');
      var empActual=(d.fila[15]||'NO').toUpperCase()==='SI'?'SI':'NO';
      document.getElementById('datEditEmp').innerHTML='<b>Nombre:</b> '+(nombreCompleto||'-')+'<br><b>CUI:</b> '+_editEmpCui+'<br><b>Empadronado actual:</b> '+empActual+'<br><b>No. Empadronamiento actual:</b> '+(d.fila[16]||'-')+'<br><b>Observaciones actuales:</b> '+(d.fila[24]||'-')+'<br><b>Última revisión Padrón/TSE:</b> '+(d.fila[25]||'Nunca revisado');
      document.getElementById('nuevoEmp').value=empActual;
      document.getElementById('nuevoNumEmp').value=d.fila[16]||'';
      document.getElementById('nuevaObsEmp').value=d.fila[24]||'';
      document.getElementById('resEditEmp').style.display='block';
      msg('msgEditEmp','','');
      msg('msgRevisado','','');
    } else { msg('msgEditEmp','❌ No encontrado','#e74c3c'); }
  }catch(e){ msg('msgEditEmp','❌ '+e.message,'#e74c3c'); }
}

async function guardarNuevoEmp(){
  if(!_editEmpCui){ msg('msgEditEmp','Vuelva a buscar','#e74c3c'); return; }
  var emp=document.getElementById('nuevoEmp').value;
  var numEmp=document.getElementById('nuevoNumEmp').value.trim();
  var obs=document.getElementById('nuevaObsEmp').value.trim();
  msg('msgEditEmp','Guardando...','#aed6f1');
  try{
    var r=await fetch('/actualizar_empadronamiento_jefe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cui:_editEmpCui, empadronado:emp, num_empadronamiento:numEmp, observaciones:obs})});
    var d=await r.json();
    if(d.ok){
      var texto='✅ Empadronamiento actualizado';
      if(d.agregado_a_padron) texto += ' — también se agregó al Padrón';
      msg('msgEditEmp',texto,'#2ecc71');
      document.getElementById('resEditEmp').style.display='none';
      document.getElementById('busEditEmp').value='';
      _editEmpCui=null;
    } else { msg('msgEditEmp','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgEditEmp','❌ '+e.message,'#e74c3c'); }
}

async function marcarComoRevisado(){
  if(!_editEmpCui){ msg('msgEditEmp','Vuelva a buscar','#e74c3c'); return; }
  msg('msgRevisado','Guardando...','#aed6f1');
  try{
    var r=await fetch('/marcar_revisado',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({cui:_editEmpCui})});
    var d=await r.json();
    if(d.ok){ msg('msgRevisado','✅ Marcado como revisado hoy — no volverá a aparecer en pendientes por 7 días','#2ecc71'); }
    else { msg('msgRevisado','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgRevisado','❌ '+e.message,'#e74c3c'); }
}

async function verPendientesEditarEmp(){
  var cont=document.getElementById('contPendientesEmp');
  cont.style.display='block';
  cont.innerHTML='<p class="msg">Cargando...</p>';
  try{
    var esAdmin = sesionJefe && sesionJefe.esAdmin;
    var params=['no_emp=SI'];
    if(!esAdmin) params.push('jefe='+encodeURIComponent(sesionJefe.nombre));
    var r=await fetch('/ver_no_empadronados?'+params.join('&'));
    var d=await r.json();
    var registros=(d.ok&&d.registros)?d.registros:[];
    var etiqueta = esAdmin ? 'todos los jefes' : (sesionJefe?sesionJefe.nombre:'');
    if(!registros.length){ cont.innerHTML='<p class="msg">🎉 No hay pendientes para '+etiqueta+'</p>'; return; }
    var h='<input type="text" placeholder="Buscar nombre o CUI..." oninput="filtrarPendientesEmp(this.value)" style="width:100%;padding:8px;border-radius:8px;border:none;background:#0d2137;color:white;font-size:0.9em;margin-bottom:8px">';
    h+='<p style="color:#aed6f1;font-size:0.8em;margin-bottom:8px">'+registros.length+' registro(s) para '+etiqueta+'</p>';
    h+='<div id="listaPendEmp" style="max-height:38vh;overflow-y:auto">';
    registros.forEach(function(p){
      var cui=p.cui||'';
      var nombre=p.nombre||'';
      h+='<div class="pend-emp-item" data-search="'+cui.toLowerCase()+' '+nombre.toLowerCase()+'" data-cui="'+cui+'" style="background:#154360;border-radius:8px;padding:10px;margin-bottom:6px;cursor:pointer">';
      h+='<p style="color:white;font-size:0.82em;font-weight:bold;margin:0">'+(nombre||'-')+'</p>';
      h+='<p style="color:#aed6f1;font-size:0.75em;margin:2px 0 0 0">DPI: '+(cui||'-')+' &nbsp; F.Nac: '+(p.fecha_nacimiento||'-')+'</p>';
      h+='</div>';
    });
    h+='</div>';
    cont.innerHTML=h;
  }catch(e){ cont.innerHTML='<p class="msg" style="color:#e74c3c">Error: '+e.message+'</p>'; }
}

function filtrarPendientesEmp(q){
  var ql=q.toLowerCase();
  document.querySelectorAll('.pend-emp-item').forEach(function(el){
    el.style.display=el.getAttribute('data-search').includes(ql)?'block':'none';
  });
}

function usarPendienteEmp(cui){
  document.getElementById('busEditEmp').value=cui;
  document.getElementById('contPendientesEmp').style.display='none';
  buscarParaEditarEmp();
}

async function buscarReg(){
  var dpi=document.getElementById('mDPI').value.trim();
  if(!dpi) return;
  msg('msgM','Buscando...','#aed6f1');
  document.getElementById('resReg').style.display='none';
  document.getElementById('editRegForm').style.display='none';
  try{
    var r=await fetch('/buscar_dpi?cui='+dpi);
    var d=await r.json();
    if(d.ok&&d.fila){
      mFila=d.numero_fila;
      mCUI=(d.fila[0]||'').trim();
      mFilaCompleta=d.fila;
      var f=d.fila;
      document.getElementById('datReg').innerHTML=
        '<b>CUI:</b> '+(f[0]||'-')+'<br>'+
        '<b>Nombre:</b> '+(f[2]||'')+' '+(f[3]||'')+' '+(f[4]||'')+' '+(f[5]||'')+'<br>'+
        '<b>Serie:</b> '+(f[1]||'-')+' &nbsp; <b>Sexo:</b> '+(f[6]||'-')+' &nbsp; <b>E. Civil:</b> '+(f[7]||'-')+'<br>'+
        '<b>F.Nac:</b> '+(f[8]||'-')+' &nbsp; <b>Nacimiento:</b> '+(f[9]||'-')+', '+(f[10]||'-')+'<br>'+
        '<b>Vecindad:</b> '+(f[11]||'-')+', '+(f[12]||'-')+'<br>'+
        '<b>F.Exp:</b> '+(f[13]||'-')+' &nbsp; <b>F.Venc:</b> '+(f[14]||'-')+'<br>'+
        '<b>Dirección:</b> '+(f[17]||'-')+'<br>'+
        '<b>Teléfono:</b> '+(f[18]||'-')+'<br>'+
        '<b>Empadronado:</b> '+(f[15]||'-')+'<br>'+
        '<b>No. Empadronamiento:</b> '+(f[16]||'-');
      document.getElementById('resReg').style.display='block'; msg('msgM','','');
    } else { msg('msgM','❌ No encontrado','#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

function abrirEditRegistro(){
  if(!mFilaCompleta) return;
  var f = mFilaCompleta;
  document.getElementById('editRegCUI').value = f[0]||'';
  document.getElementById('editRegPNom').value = f[2]||'';
  document.getElementById('editRegSNom').value = f[3]||'';
  document.getElementById('editRegPApe').value = f[4]||'';
  document.getElementById('editRegSApe').value = f[5]||'';
  document.getElementById('editRegSerie').value = f[1]||'';
  document.getElementById('editRegSexo').value = (f[6]||'').toUpperCase()==='F' ? 'F' : ((f[6]||'').toUpperCase()==='M' ? 'M' : '');
  document.getElementById('editRegEC').value = f[7]||'';
  document.getElementById('editRegFNac').value = f[8]||'';
  document.getElementById('editRegMunNac').value = f[9]||'';
  document.getElementById('editRegDepNac').value = f[10]||'';
  document.getElementById('editRegMunVec').value = f[11]||'';
  document.getElementById('editRegDepVec').value = f[12]||'';
  document.getElementById('editRegFExp').value = f[13]||'';
  document.getElementById('editRegFVenc').value = f[14]||'';
  document.getElementById('editRegDir').value = f[17]||'';
  document.getElementById('editRegTel').value = f[18]||'';
  document.getElementById('editRegEmp').value = (f[15]||'NO').toUpperCase()==='SI' ? 'SI' : 'NO';
  document.getElementById('editRegNoEmp').value = f[16]||'';
  document.getElementById('editRegForm').style.display='block';
}

function cancelarEditRegistro(){
  document.getElementById('editRegForm').style.display='none';
}

async function guardarEditRegistro(){
  if(!mCUI) return;
  var cuiNuevo = document.getElementById('editRegCUI').value.trim();
  if(!document.getElementById('editRegDir').value.trim()){
    msg('msgM','❌ La dirección / comunidad es obligatoria. No se puede guardar vacía.','#e74c3c');
    return;
  }
  var payload = {
    cui: mCUI,
    cui_nuevo: cuiNuevo,
    primer_nombre: document.getElementById('editRegPNom').value.trim(),
    segundo_nombre: document.getElementById('editRegSNom').value.trim(),
    primer_apellido: document.getElementById('editRegPApe').value.trim(),
    segundo_apellido: document.getElementById('editRegSApe').value.trim(),
    numero_serie: document.getElementById('editRegSerie').value.trim(),
    sexo: document.getElementById('editRegSexo').value,
    estado_civil: document.getElementById('editRegEC').value.trim(),
    fecha_nacimiento: document.getElementById('editRegFNac').value.trim(),
    municipio_nacimiento: document.getElementById('editRegMunNac').value.trim(),
    departamento_nacimiento: document.getElementById('editRegDepNac').value.trim(),
    municipio_vecindad: document.getElementById('editRegMunVec').value.trim(),
    departamento_vecindad: document.getElementById('editRegDepVec').value.trim(),
    fecha_expedicion: document.getElementById('editRegFExp').value.trim(),
    fecha_vencimiento: document.getElementById('editRegFVenc').value.trim(),
    direccion: document.getElementById('editRegDir').value.trim(),
    telefono: document.getElementById('editRegTel').value.trim(),
    empadronado: document.getElementById('editRegEmp').value,
    num_empadronamiento: document.getElementById('editRegNoEmp').value.trim()
  };
  msg('msgM','Guardando...','#aed6f1');
  try{
    var r = await fetch('/actualizar_registro_completo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    var d = await r.json();
    if(d.ok){
      var textoOk = '✅ Registro actualizado';
      if(d.confirmado_por_padron) textoOk += ' — 🗳️ ¡Se confirmó como EMPADRONADO en el padrón!';
      msg('msgM',textoOk,'#2ecc71');
      document.getElementById('editRegForm').style.display='none';
      mCUI = cuiNuevo || mCUI; // si cambio el CUI, buscar con el nuevo de aqui en adelante
      document.getElementById('mDPI').value = mCUI;
      await buscarReg(); // recargar para mostrar los datos ya actualizados
    } else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function elimReg(){
  if(!mFila||!confirm('¿Eliminar? Esto también lo quitará de Grupos y Presidentas si aparece ahí.')) return;
  try{
    var r=await fetch('/eliminar_dpi',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({numero_fila:mFila, cui:mCUI})});
    var d=await r.json();
    if(d.ok){ msg('msgM','✅ Eliminado','#2ecc71'); document.getElementById('resReg').style.display='none'; document.getElementById('mDPI').value=''; mFila=null; mCUI=null; }
    else { msg('msgM','❌ '+d.error,'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

var mFilaAfil=null, mCUIAfil=null, mFilaCompletaAfil=null;

async function buscarAfiliadoMant(){
  var dpi=document.getElementById('mDPIAfil').value.trim();
  if(!dpi) return;
  msg('msgM','Buscando...','#aed6f1');
  document.getElementById('resAfil').style.display='none';
  document.getElementById('editAfilForm').style.display='none';
  try{
    var r=await fetch('/buscar_afiliado?cui='+dpi);
    var d=await r.json();
    if(d.ok&&d.fila){
      mFilaAfil=d.numero_fila;
      mCUIAfil=(d.fila[0]||'').trim();
      mFilaCompletaAfil=d.fila;
      var f=d.fila;
      document.getElementById('datAfil').innerHTML=
        '<b>CUI:</b> '+(f[0]||'-')+'<br>'+
        '<b>Nombre:</b> '+(f[1]||'')+' '+(f[2]||'')+' '+(f[3]||'')+' '+(f[4]||'')+'<br>'+
        '<b>Sexo:</b> '+(f[5]||'-')+' &nbsp; <b>E. Civil:</b> '+(f[6]||'-')+'<br>'+
        '<b>F.Nac:</b> '+(f[7]||'-')+' &nbsp; <b>Nacimiento:</b> '+(f[8]||'-')+', '+(f[9]||'-')+'<br>'+
        '<b>Vecindad:</b> '+(f[10]||'-')+', '+(f[11]||'-')+'<br>'+
        '<b>F.Exp:</b> '+(f[12]||'-')+' &nbsp; <b>F.Venc:</b> '+(f[13]||'-')+'<br>'+
        '<b>Dirección:</b> '+(f[14]||'-')+'<br>'+
        '<b>Teléfono:</b> '+(f[15]||'-')+'<br>'+
        '<b>Empadronado:</b> '+(f[16]||'-')+' &nbsp; <b>No. Empadronamiento:</b> '+(f[17]||'-')+'<br>'+
        '<b>Registrado por:</b> '+(f[18]||'-')+' &nbsp; <b>Fecha:</b> '+(f[19]||'-')+'<br>'+
        '<b>Firma/huella guardada:</b> '+((f[20]||'').trim() ? 'Sí ✅' : 'No');
      document.getElementById('resAfil').style.display='block'; msg('msgM','','');
    } else { msg('msgM','❌ '+(d.error||'No encontrado'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

function abrirEditAfiliado(){
  if(!mFilaCompletaAfil) return;
  var f = mFilaCompletaAfil;
  document.getElementById('editAfilCUI').value = f[0]||'';
  document.getElementById('editAfilPNom').value = f[1]||'';
  document.getElementById('editAfilSNom').value = f[2]||'';
  document.getElementById('editAfilPApe').value = f[3]||'';
  document.getElementById('editAfilSApe').value = f[4]||'';
  document.getElementById('editAfilSexo').value = (f[5]||'').toUpperCase()==='F' ? 'F' : ((f[5]||'').toUpperCase()==='M' ? 'M' : '');
  document.getElementById('editAfilEC').value = f[6]||'';
  document.getElementById('editAfilFNac').value = f[7]||'';
  document.getElementById('editAfilMunNac').value = f[8]||'';
  document.getElementById('editAfilDepNac').value = f[9]||'';
  document.getElementById('editAfilMunVec').value = f[10]||'';
  document.getElementById('editAfilDepVec').value = f[11]||'';
  document.getElementById('editAfilFExp').value = f[12]||'';
  document.getElementById('editAfilFVenc').value = f[13]||'';
  document.getElementById('editAfilDir').value = f[14]||'';
  document.getElementById('editAfilTel').value = f[15]||'';
  document.getElementById('editAfilEmp').value = (f[16]||'NO').toUpperCase()==='SI' ? 'SI' : 'NO';
  document.getElementById('editAfilNoEmp').value = f[17]||'';
  document.getElementById('editAfilForm').style.display='block';
}

function cancelarEditAfiliado(){
  document.getElementById('editAfilForm').style.display='none';
}

async function guardarEditAfiliado(){
  if(!mFilaAfil) return;
  var cuiNuevo = document.getElementById('editAfilCUI').value.trim();
  if(!document.getElementById('editAfilDir').value.trim()){
    msg('msgM','❌ La dirección / comunidad es obligatoria. No se puede guardar vacía.','#e74c3c');
    return;
  }
  var payload = {
    numero_fila: mFilaAfil,
    datos: {
      cui: cuiNuevo,
      primer_nombre: document.getElementById('editAfilPNom').value.trim(),
      segundo_nombre: document.getElementById('editAfilSNom').value.trim(),
      primer_apellido: document.getElementById('editAfilPApe').value.trim(),
      segundo_apellido: document.getElementById('editAfilSApe').value.trim(),
      sexo: document.getElementById('editAfilSexo').value,
      estado_civil: document.getElementById('editAfilEC').value.trim(),
      fecha_nacimiento: document.getElementById('editAfilFNac').value.trim(),
      municipio_nacimiento: document.getElementById('editAfilMunNac').value.trim(),
      departamento_nacimiento: document.getElementById('editAfilDepNac').value.trim(),
      municipio_vecindad: document.getElementById('editAfilMunVec').value.trim(),
      departamento_vecindad: document.getElementById('editAfilDepVec').value.trim(),
      fecha_expedicion: document.getElementById('editAfilFExp').value.trim(),
      fecha_vencimiento: document.getElementById('editAfilFVenc').value.trim(),
      direccion: document.getElementById('editAfilDir').value.trim(),
      telefono: document.getElementById('editAfilTel').value.trim(),
      empadronado: document.getElementById('editAfilEmp').value,
      num_empadronamiento: document.getElementById('editAfilNoEmp').value.trim()
    }
  };
  msg('msgM','Guardando...','#aed6f1');
  try{
    var r = await fetch('/actualizar_afiliado',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)});
    var d = await r.json();
    if(d.ok){
      msg('msgM','✅ Afiliado actualizado','#2ecc71');
      document.getElementById('editAfilForm').style.display='none';
      mCUIAfil = cuiNuevo || mCUIAfil;
      document.getElementById('mDPIAfil').value = mCUIAfil;
      await buscarAfiliadoMant();
    } else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function elimAfiliado(){
  if(!mFilaAfil||!confirm('¿Eliminar a este afiliado? Esta acción no se puede deshacer.')) return;
  try{
    var r=await fetch('/eliminar_afiliado',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({numero_fila:mFilaAfil, cui:mCUIAfil})});
    var d=await r.json();
    if(d.ok){ msg('msgM','✅ Afiliado eliminado','#2ecc71'); document.getElementById('resAfil').style.display='none'; document.getElementById('mDPIAfil').value=''; mFilaAfil=null; mCUIAfil=null; }
    else { msg('msgM','❌ '+d.error,'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function buscarGrpM(){
  var n=document.getElementById('mGN').value.trim();
  var d=document.getElementById('mGD').value.trim();
  if(!n&&!d){ msg('msgM','Ingrese nombre o DPI','#e74c3c'); return; }
  msg('msgM','Buscando...','#aed6f1');
  document.getElementById('resGrp').style.display='none';
  document.getElementById('editGrpForm').style.display='none';
  try{
    var p=n?'nombre='+encodeURIComponent(n):'';
    if(d) p+=(p?'&':'')+'dpi='+d;
    var r=await fetch('/buscar_grupo?'+p);
    var rd=await r.json();
    if(rd.ok&&rd.grupo){
      mGN=rd.grupo.nombre;
      mGJefe=rd.grupo.jefe_sector||'';
      document.getElementById('datGrp').innerHTML='<b>Grupo:</b> '+rd.grupo.nombre+'<br><b>Coord:</b> '+rd.grupo.coordinador+'<br><b>Jefe:</b> '+(rd.grupo.jefe_sector||'-')+'<br><b>Personas:</b> '+rd.personas.length;
      document.getElementById('resGrp').style.display='block'; msg('msgM','','');
    } else { msg('msgM','❌ No encontrado','#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

function abrirEditGrupo(){
  document.getElementById('editGrpNombre').value = mGN||'';
  document.getElementById('editGrpForm').style.display='block';
}

function cancelarEditGrupo(){
  document.getElementById('editGrpForm').style.display='none';
}

async function guardarEditGrupo(){
  var nuevoNombre = document.getElementById('editGrpNombre').value.trim();
  if(!nuevoNombre){ msg('msgM','Escriba el nuevo nombre','#e74c3c'); return; }
  msg('msgM','Guardando...','#aed6f1');
  try{
    var r = await fetch('/editar_nombre_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_actual:mGN, nombre_nuevo:nuevoNombre, jefe_sector:mGJefe||''})});
    var d = await r.json();
    if(d.ok){
      msg('msgM','✅ Nombre actualizado ('+d.filas+' fila(s))','#2ecc71');
      document.getElementById('editGrpForm').style.display='none';
      document.getElementById('mGN').value = nuevoNombre;
      await buscarGrpM();
    } else { msg('msgM','❌ '+(d.error||'Error'),'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}

async function elimGrp(){
  if(!mGN||!confirm('¿Eliminar grupo "'+mGN+'"?')) return;
  try{
    var r=await fetch('/eliminar_grupo',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({nombre_grupo:mGN})});
    var d=await r.json();
    if(d.ok){ msg('msgM','✅ Eliminado ('+d.eliminadas+')','#2ecc71'); document.getElementById('resGrp').style.display='none'; document.getElementById('mGN').value=''; document.getElementById('mGD').value=''; mGN=''; }
    else { msg('msgM','❌ '+d.error,'#e74c3c'); }
  }catch(e){ msg('msgM','❌ '+e.message,'#e74c3c'); }
}


document.addEventListener("DOMContentLoaded", function(){
// MAIN BUTTONS
document.getElementById('b1').onclick=function(){ abrirGrupos(); };
document.getElementById('b2').onclick=function(){ show('oConsultas'); };
document.getElementById('b3').onclick=function(){ show('oMant'); clv=''; updC(); document.getElementById('mantLogin').style.display='block'; document.getElementById('mantPanel').style.display='none'; };
document.getElementById('bNuevo').onclick=function(){
  if(sesionJefe && sesionJefe.esAdmin){
    alert('Para crear un grupo, inicia sesión como el jefe de sector correspondiente (no como Administrador), para que el grupo quede correctamente asignado.');
    return;
  }
  hide('oGrupos'); gN=''; gC=null; gPs=[];
  document.getElementById('ngN').value='';
  gJ = sesionJefe ? sesionJefe.nombre : '';
  document.getElementById('ngJ').value = gJ;
  document.getElementById('ngJDisplay').textContent = gJ || '-';
  cargarPresidentas();
  show('oNuevoG1');
};
document.getElementById('bContinuar').onclick=function(){ hide('oGrupos'); document.getElementById('bgN').value=''; document.getElementById('bgD').value=''; document.getElementById('resBG').style.display='none'; msg('msgBG','',''); show('oBuscarG'); };
document.getElementById('bRegInd').onclick=function(){ abrirRegistrosInd(); };
document.getElementById('bRegGrp').onclick=function(){ abrirRegistrosGrupos(); };

// TECLADO
['1','2','3','4','5','6','7','8','9','0'].forEach(function(n){
  var el=document.getElementById('tk'+n);
  if(el) el.onclick=function(){ if(clv.length<6){ clv+=n; updC(); } };
});
document.getElementById('tkDel').onclick=function(){ clv=clv.slice(0,-1); updC(); };
document.getElementById('tkOk').onclick=async function(){
  document.getElementById('errClave').style.display='none';
  try{
    var r = await fetch('/login_admin',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({pin:clv})});
    var d = await r.json();
    if(d.ok){
      sesionJefe = {nombre:'ADMINISTRADOR', esAdmin:true};
      actualizarBotonLogout();
      document.getElementById('mantLogin').style.display='none';
      document.getElementById('mantPanel').style.display='block';
      document.getElementById('mDPI').value='';
      document.getElementById('mGN').value='';
      document.getElementById('mGD').value='';
      document.getElementById('resReg').style.display='none';
      document.getElementById('resGrp').style.display='none';
      msg('msgM','','');
      tabM('ind');
    } else {
      document.getElementById('errClave').textContent = d.error || 'Clave incorrecta';
      document.getElementById('errClave').style.display='block';
      clv=''; updC();
    }
  }catch(e){
    document.getElementById('errClave').style.display='block';
    document.getElementById('errClave').textContent='Error de conexión';
    clv=''; updC();
  }
};

});
</script>
</body>
</html>"""




# Conexión a Google Sheets reutilizada entre peticiones. Antes, get_sheet()
# volvía a autenticarse con Google Y volvía a abrir la hoja de cálculo en
# CADA llamada (prácticamente cada clic en la app hace varias llamadas a
# get_sheet()). Eso es lento, cuenta contra los límites de la API de Google,
# y bajo uso concurrente de varias personas puede saturar el único proceso
# que corre la app en el plan gratis de Render. Ahora se autentica y se abre
# la hoja una sola vez, y las siguientes llamadas reutilizan esa conexión.
_spreadsheet_cache = None

# Límite de cuántos análisis de DPI (fotos con Claude Vision) pueden correr
# al mismo tiempo. Cada análisis usa varios MB de memoria (las imágenes,
# aunque ya comprimidas del lado del navegador, siguen pesando algo); sin
# este límite, muchas personas subiendo fotos al mismo momento podrían
# agotar la memoria del servidor (plan gratis, con RAM limitada). Si se
# llega al límite, las peticiones de más esperan un poco antes de seguir,
# en vez de arriesgar que el servidor se quede sin memoria para todos.
import threading as _threading_mod
_LIMITE_ANALISIS_SIMULTANEOS = 2
_semaforo_analisis = _threading_mod.Semaphore(_LIMITE_ANALISIS_SIMULTANEOS)

# Candado que obliga a que el registro de presidentas/coordinadoras/
# integrantes ocurra de uno en uno. Sin esto, si dos peticiones llegan casi
# al mismo tiempo (doble toque accidental, o varias personas trabajando a
# la vez), ambas pueden "revisar si ya existe" ANTES de que cualquiera
# termine de guardar, y las dos terminan guardando el mismo duplicado,
# aunque la revisión de duplicados en sí esté bien escrita. Con el candado,
# la segunda petición espera a que la primera termine por completo (revisar
# Y guardar) antes de poder empezar su propia revisión.
_candado_registro_personas = _threading_mod.Lock()


def _conectar_sheets():
    global _spreadsheet_cache
    scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(CREDENCIALES_GOOGLE, scopes=scopes)
    gc = gspread.authorize(creds)
    try:
        _spreadsheet_cache = gc.open_by_key(SHEET_ID)
    except Exception:
        _spreadsheet_cache = gc.open("DPI_PILAS2026")
    return _spreadsheet_cache


def get_sheet():
    global _spreadsheet_cache
    if _spreadsheet_cache is not None:
        return _spreadsheet_cache
    return _conectar_sheets()


@app.route("/")
def index():
    from flask import Response
    resp = Response(HTML, mimetype='text/html')
    resp.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    resp.headers['Pragma'] = 'no-cache'
    resp.headers['Expires'] = '0'
    return resp
def version():
    return "VERSION 3.0 - OK"

@app.route("/analizar", methods=["POST"])
def analizar():
    # Si ya hay demasiados análisis corriendo al mismo tiempo, esperar un
    # poco (hasta 25s) a que se libere un espacio, en vez de sumar uno más y
    # arriesgar que el servidor se quede sin memoria para todos.
    adquirido = _semaforo_analisis.acquire(timeout=25)
    if not adquirido:
        return jsonify({"ok": False, "error": "El servidor está procesando varias fotos a la vez. Espera unos segundos e inténtalo de nuevo."})
    try:
        data = request.json
        frente_type = data.get("frente_type", "image/jpeg")
        reverso_type = data.get("reverso_type", "image/jpeg")
        if frente_type not in ["image/jpeg","image/png","image/gif","image/webp"]:
            frente_type = "image/jpeg"
        if reverso_type not in ["image/jpeg","image/png","image/gif","image/webp"]:
            reverso_type = "image/jpeg"

        api_key = os.environ.get("ANTHROPIC_API_KEY")
        if not api_key:
            return jsonify({"ok": False, "error": "API key no encontrada en el servidor"})

        client = anthropic.Anthropic(api_key=api_key)
        respuesta = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=1000,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": frente_type, "data": data["frente"]}},
                    {"type": "image", "source": {"type": "base64", "media_type": reverso_type, "data": data["reverso"]}},
                    {"type": "text", "text": PROMPT}
                ]
            }]
        )
        texto = respuesta.content[0].text.strip().replace("```json","").replace("```","").strip()
        datos = json.loads(texto)

        # Si la IA detectó que esto NO es un DPI físico real (por ejemplo,
        # notas escritas a mano simulando contener datos de un DPI), NO se
        # bloquea de una vez — se deja que la persona decida si continúa o
        # no (el frontend muestra una advertencia con esa opción), pero
        # queda registrado en auditoría para poder darle seguimiento a
        # quién insiste en intentarlo.
        no_autentico = datos.get("es_dpi_autentico") is False
        motivo_rechazo = datos.get("motivo_rechazo","La imagen no parece ser una fotografía de un DPI físico real.") if no_autentico else ""
        if no_autentico:
            try:
                _nombre_sesion_rechazo, _es_admin_rechazo = _sesion_actual()
                sh_rechazo = get_sheet()
                _registrar_auditoria(sh_rechazo, _nombre_sesion_rechazo or "desconocido", "⚠️ Posible DPI no auténtico detectado", motivo_rechazo)
            except Exception:
                pass

        cui = datos.get("cui","").replace(" ","").replace("-","").strip()

        # CUI especial reservado para pruebas internas (simula una persona
        # no empadronada). Se excluye de la validación de formato para no
        # romper las pruebas, aunque su código de departamento no sea real.
        CUI_PRUEBA_NO_EMP = "9999999999901"
        # Lote de 30 DPI de práctica (CUI 9999000000001 al 9999000000030),
        # usados por jefes de sector para practicar el flujo de la app
        # antes de usar DPI reales. También llevan un código de
        # departamento ficticio ("00"), así que se excluyen igual.
        CUIS_PRUEBA_PRACTICA = {f"99990000000{str(n).zfill(2)}" for n in range(1, 31)}
        cui_es_de_prueba = cui == CUI_PRUEBA_NO_EMP or cui in CUIS_PRUEBA_PRACTICA

        # Si el CUI leído no tiene un formato válido (13 dígitos, con un
        # código de departamento entre 01 y 22), se bloquea el registro
        # por completo — no se guarda nada y se le pide repetir la foto. A
        # diferencia de la advertencia de autenticidad, aquí NO hay opción
        # de "continuar de todos modos": un código de departamento que no
        # existe es un error real, no un caso dudoso que la persona pueda
        # decidir pasar por alto.
        if not cui_es_de_prueba and _dpi_formato_invalido(cui):
            try:
                _nombre_sesion_fmt, _es_admin_fmt = _sesion_actual()
                sh_fmt = get_sheet()
                _registrar_auditoria(sh_fmt, _nombre_sesion_fmt or "desconocido", "⚠️ DPI con formato inválido detectado", f"CUI leído: {cui}")
            except Exception:
                pass
            return jsonify({"ok": False, "cui_invalido": True,
                "error": f"🚫 El número de DPI leído ({cui or 'vacío'}) no tiene un formato válido de CUI guatemalteco. Es probable que la foto no se haya leído bien. Por favor tome la foto de nuevo, con buena luz y asegurándose de que se vea completo el número."})

        # Verificar duplicado y padron
        duplicado = False
        empadronado = False
        num_empadronamiento = ""
        try:
            sh = get_sheet()
            ws_dpi = sh.worksheet(HOJA)
            cuis_existentes = ws_dpi.col_values(1)
            duplicado = cui in cuis_existentes
            if cui == CUI_PRUEBA_NO_EMP:
                # Simular datos de una persona no empadronada
                datos = {
                    "cui": "9999999999901",
                    "numero_serie": "TEST001",
                    "primer_nombre": "PRUEBA",
                    "segundo_nombre": "NO",
                    "primer_apellido": "EMPADRONADO",
                    "segundo_apellido": "TEST",
                    "sexo": "M",
                    "estado_civil": "Soltero",
                    "fecha_nacimiento": "01/01/1990",
                    "municipio_nacimiento": "TOTONICAPAN",
                    "departamento_nacimiento": "TOTONICAPAN",
                    "municipio_vecindad": "TOTONICAPAN",
                    "departamento_vecindad": "TOTONICAPAN",
                    "fecha_expedicion": "01/01/2020",
                    "fecha_vencimiento": "01/01/2030",
                }
                return jsonify({"ok": True, "datos": datos, "empadronado": False,
                                "num_empadronamiento": "", "duplicado": False,
                                "es_prueba": True, "no_autentico": False, "motivo_rechazo": ""})
            try:
                ws_padron = sh.worksheet(HOJA_PADRON)
                dpis_padron = ws_padron.col_values(1)
                cui_limpio = cui.replace(" ","").replace("-","").replace(",","").strip()
                empadronado_encontrado = False
                for celda in dpis_padron:
                    celda_limpia = celda.replace(" ","").replace("-","").strip()
                    if celda_limpia.startswith(cui_limpio) or cui_limpio in celda_limpia.split(",")[0]:
                        empadronado = True
                        num_empadronamiento = cui_limpio
                        empadronado_encontrado = True
                        break
            except:
                pass
        except Exception as sheets_ex:
            print(f"Sheets error (non-fatal): {sheets_ex}")

        return jsonify({"ok": True, "datos": datos, "empadronado": empadronado,
                        "num_empadronamiento": num_empadronamiento, "duplicado": duplicado,
                        "no_autentico": no_autentico, "motivo_rechazo": motivo_rechazo})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "detalle": traceback.format_exc()[-500:]})
    finally:
        _semaforo_analisis.release()


@app.route("/guardar", methods=["POST"])
@requiere_admin
def guardar():
    """Guarda un registro SOLO en la hoja DPI (no en GRUPOS ni en
    PRESIDENTAS) — pensado para pruebas del administrador, no para el flujo
    normal de registro por grupos. Se marca con Jefe='PRUEBA' y
    Rol='PRUEBA' para poder identificarlo y borrarlo fácilmente después
    sin que se mezcle con datos reales de ningún jefe de sector."""
    try:
        datos = request.json["datos"]
        cui = datos.get("cui","").replace(" ","").replace("-","").strip()
        if not datos.get('fecha_nacimiento','').strip():
            return jsonify({"ok": False, "error": "Falta la fecha de nacimiento. No se puede guardar el registro sin ese dato."})
        if not datos.get('direccion','').strip():
            return jsonify({"ok": False, "error": "Falta la dirección / comunidad. No se puede guardar el registro sin ese dato."})
        sh = get_sheet()
        ws_dpi = sh.worksheet(HOJA)
        cuis_existentes = ws_dpi.col_values(1)
        if cui in cuis_existentes:
            return jsonify({"ok": False, "error": "DPI duplicado"})
        empadronado = "NO"
        num_empadronamiento = ""
        try:
            ws_padron = sh.worksheet(HOJA_PADRON)
            dpis_padron = ws_padron.col_values(1)
            cui_limpio = cui.replace(" ","").replace("-","").replace(",","").strip()
            for celda in dpis_padron:
                celda_limpia = celda.replace(" ","").replace("-","").strip()
                if celda_limpia.startswith(cui_limpio) or cui_limpio in celda_limpia.split(",")[0]:
                    empadronado = "SI"
                    num_empadronamiento = cui_limpio
                    break
        except:
            pass
        # 21 columnas exactas segun encabezado de hoja DPI
        cui_guardado = datos.get("cui","").replace(" ","").replace("-","").strip()
        ws_dpi.append_row([
            cui_guardado,                               # A - CUI
            datos.get("numero_serie",""),               # B - Número de serie
            datos.get("primer_nombre",""),              # C - Primer nombre
            datos.get("segundo_nombre",""),             # D - Segundo nombre
            datos.get("primer_apellido",""),            # E - Primer apellido
            datos.get("segundo_apellido",""),           # F - Segundo apellido
            datos.get("sexo",""),                       # G - Sexo
            datos.get("estado_civil",""),               # H - Estado civil
            datos.get("fecha_nacimiento",""),           # I - Fecha de nacimiento
            datos.get("municipio_nacimiento",""),       # J - Municipio nacimiento
            datos.get("departamento_nacimiento",""),    # K - Departamento nacimiento
            datos.get("municipio_vecindad",""),         # L - Municipio vecindad
            datos.get("departamento_vecindad",""),      # M - Departamento vecindad
            datos.get("fecha_expedicion",""),           # N - Fecha expedición
            datos.get("fecha_vencimiento",""),          # O - Fecha vencimiento
            empadronado,                                # P - Empadronado
            num_empadronamiento,                        # Q - No. Empadronamiento
            datos.get("direccion",""),                  # R - Direccion
            formatear_telefono_gt(datos.get("telefono","").strip()),  # S - Telefono
            "PRUEBA",                                    # T - Jefe de Sector (marcado para identificarlo)
            "",                                         # U - Presidenta
            "PRUEBA",                                    # V - Rol (marcado para identificarlo)
            "NO" if datos.get("dpi_no_verificado") else "SI",  # W - DPI Verificado (proviene de una foto real)
            ahora_gt().strftime('%d/%m/%Y'),            # X - Fecha en que se agregó el registro
        ])
        nombre = f"{datos.get('primer_nombre','')} {datos.get('primer_apellido','')}"
        return jsonify({"ok": True, "nombre": nombre, "empadronado": empadronado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/guardar_afiliado", methods=["POST"])
@requiere_afiliacion
def guardar_afiliado():
    """Guarda un afiliado nuevo en la hoja AFILIADOS (totalmente separada
    de la hoja DPI y de los registros normales de Presidentas/
    Coordinadoras/Integrantes). Solo lo puede usar quien tenga sesión de
    'Afiliación al Partido' (apoyo1/apoyo2). Verifica el empadronamiento
    contra el Padrón, igual que el resto de la aplicación."""
    try:
        datos = request.json["datos"]
        cui = datos.get("cui","").replace(" ","").replace("-","").strip()
        if not cui:
            return jsonify({"ok": False, "error": "Falta el DPI"})
        if not datos.get('fecha_nacimiento','').strip():
            return jsonify({"ok": False, "error": "Falta la fecha de nacimiento. No se puede guardar el registro sin ese dato."})
        if not datos.get('direccion','').strip():
            return jsonify({"ok": False, "error": "Falta la dirección / comunidad. No se puede guardar el registro sin ese dato."})
        sh = get_sheet()
        ws_afiliados = _get_ws_afiliados(sh)
        cuis_existentes = [c.replace(" ","").replace("-","").strip() for c in ws_afiliados.col_values(1)]
        if cui in cuis_existentes:
            return jsonify({"ok": False, "error": "Este DPI ya está registrado como afiliado"})

        empadronado = "NO"
        num_empadronamiento = ""
        try:
            ws_padron = sh.worksheet(HOJA_PADRON)
            dpis_padron = ws_padron.col_values(1)
            for celda in dpis_padron:
                celda_limpia = celda.replace(" ","").replace("-","").strip()
                if celda_limpia.startswith(cui) or cui in celda_limpia.split(",")[0]:
                    empadronado = "SI"
                    num_empadronamiento = cui
                    break
        except Exception:
            pass

        nombre_sesion, _ = _sesion_actual()
        ws_afiliados.append_row([
            cui,                                          # CUI
            datos.get("primer_nombre",""),
            datos.get("segundo_nombre",""),
            datos.get("primer_apellido",""),
            datos.get("segundo_apellido",""),
            datos.get("sexo",""),
            datos.get("estado_civil",""),
            datos.get("fecha_nacimiento",""),
            datos.get("municipio_nacimiento",""),
            datos.get("departamento_nacimiento",""),
            datos.get("municipio_vecindad",""),
            datos.get("departamento_vecindad",""),
            datos.get("fecha_expedicion",""),
            datos.get("fecha_vencimiento",""),
            datos.get("direccion",""),
            formatear_telefono_gt(datos.get("telefono","").strip()),
            empadronado,
            num_empadronamiento,
            nombre_sesion or "",
            ahora_gt().strftime('%d/%m/%Y'),
            "",                                           # FirmaB64 (vacía; se llena al firmar el carné)
        ])
        _registrar_auditoria(sh, nombre_sesion or "AFILIACION", "Registró afiliado nuevo", f"CUI: {cui}")
        nombre_completo = f"{datos.get('primer_nombre','')} {datos.get('primer_apellido','')}"
        return jsonify({"ok": True, "nombre": nombre_completo, "empadronado": empadronado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


_CARNET_FRENTE_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAJRA28DASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD7LooooAKKRmVFLMwUDqSelY114t8LWshjuvEmjwuOqvexqfyJppN7AbVFYA8a+DyBjxRoxz/0+x/404eMvCZ6eJdIP/b5H/jVezl2FdG7RWH/AMJj4T6/8JLpGP8Ar8j/AMaP+Ex8J/8AQyaR/wCBkf8AjR7OXYOZG5RWEfGPhMdfEukf+Bkf+NH/AAmPhMD/AJGXSMf9fif40ezn2Fzx7m7RWD/wmfhL/oZtI/8AAyP/ABo/4TLwnjP/AAkmk/8AgZH/AI0eyn2F7WHc3qODWGPGHhQ9PEek/wDgZH/jS/8ACXeFu3iLSv8AwLT/ABo9nPsL2sO5t0Vhnxf4WAyfEWlD/t7T/GgeL/C3/Qw6V/4Fp/jR7OfYPaw7m5RWGPF3hY5/4qHSv/AtP8aX/hLvC/8A0MGl/wDgWn+NHsp9g9rDubdHesT/AIS7wv8A9DDpX/gWn+NB8XeFx18QaWP+3tP8aPZT7B7WHc26Kw/+Ev8AC3/Qw6V/4Fp/jR/wl/hb/oYdK/8AAtP8aPZT7B7WHc3KKxP+Et8L/wDQw6V/4Fp/jR/wlvhj/oYNK/8AAtP8aPZz7B7WHc26KxP+Et8Mf9DBpf8A4Fp/jQPFvhg8DxBpZ/7e0/xo9nPsHtYdzborE/4S3wx/0MGl8f8AT2n+NH/CW+GP+hg0v/wKT/Gj2c+we2h3NuisUeLPDJ/5j+mf+BSf40v/AAlXhr/oPaZ/4FJ/jR7OfYPbU+5s0VjDxV4aP/Me0z/wKT/GlPinw4P+Y7pv/gUn+NHs59he3p/zI2KKyB4m8PHprmmn/t6T/Gj/AISXw/8A9BrTv/AlP8aPZz7B7en/ADI16KyP+En8O99c03/wKT/Gmt4q8MqcHX9LB97tP8aPZy7DVWD2Zs0VjDxV4ZP/ADMGlf8AgXH/AI0HxX4YH/Mw6V/4Fp/jS5JdiuePc2aKxT4t8Lj/AJmLSv8AwLj/AMaQ+LvCw6+ItJH/AG9p/jT9nLsHPHubdFYY8YeFTz/wkek/+Bkf+NKfF3hYHH/CRaV/4Fp/jR7OfYOePc26Kwz4v8LY/wCRi0r/AMC0/wAaUeLfC5HHiLSj/wBvaf40ezn2Fzx7m3RWJ/wl3hf/AKGLSv8AwLT/ABoPi/wt/wBDFpX/AIFp/jR7OfYPaR7m3RWH/wAJf4Vz/wAjHpP/AIFp/jSf8Jh4V7+I9J/8C0/xo9nPsHPHubtFYX/CY+FP+hk0kf8Ab2n+NB8ZeFB18SaT/wCBaf40ezn2D2ke5u0Vgnxn4SAJPiXSMDr/AKYn+NNPjXwj/wBDLpP/AIFp/jR7KfYPaR7nQUVzx8b+Dx18T6QP+3xP8aD448HAc+J9HH/b2n+NHsp9mHtI9zoaK5w+OvBo6+KNH/8AAxP8aP8AhO/Bv/Q0aR/4Fp/jT9lU/lYvaw7nRmiuc/4TvwZ/0NGj/wDgYn+NH/Cd+Df+ho0f/wADE/xo9jU/lYe2h3Ojormx478GnkeJ9IP/AG9p/jR/wnngz/oZ9I/8C0/xp+wqfysXtqf8yOkormv+E+8F/wDQz6Tj/r6T/GmH4g+CR18UaT/4FL/jR7Cp/K/uF7el/MjqKK5f/hYPgoAZ8UaTz0/0pf8AGj/hYPgn/oaNK/8AAlf8aPYVf5X9wfWKX8yOooNcsfiH4IGCfFOkgf8AXytJ/wALF8D/APQ06T/4ErR7Cr/K/uH7en/MjqqK5U/EXwPgn/hKdJ45P+krQfiJ4HHXxTpI/wC3laPYVf5X9wvb0/5kdVRXKH4j+BR/zNWk/wDgStN/4WT4Exn/AISrSff/AEhaPYVP5X9we3p/zI62iuSPxJ8CY/5GrSv/AAIFJ/wszwF38V6V/wCBAo9hV/lf3D9vT/mR11FcifiZ4C7+K9K/7/ik/wCFm+Af+hr0v/v+KPYVf5X9we3p/wAyOvorkP8AhZvgLGf+Eq0zH/XYUjfE7wCOvivTB/22o9hU/lf3D9rDudhRXHN8UPACjnxXpn/f2k/4Wl8Pv+hr03/v5R7Cp/K/uD2sO52VFcb/AMLS+H4/5mrTf+/n/wBak/4Wp8Pf+hr03/v5/wDWo9hV/lf3B7WHc7OiuKPxW+Hg6+LNN/77P+FN/wCFs/Dof8zZp3/fR/wo9hU/lf3D549zt6K4j/hbPw6/6GzT/wDvo/4Uf8LZ+HX/AENmnf8AfR/wo9hU/lY+ePc7eiuHb4t/Dhevi3Tx+Lf4VGfjD8NR18W2P/j3+FHsKn8r+4OZdzvKK4L/AIXF8NP+hssv++X/APiaT/hcnwz/AOhssv8Avl//AIml7Gp/Kx8yO+orgR8Zfhmc48WWfH+y/wD8TTf+FzfDP/obLT/viT/4mj2NT+VhzI9Aorz4/Gj4Zjr4rtf+/cn/AMTQfjT8Mh18VW3/AH6k/wDiafsan8rC6PQaK88Pxq+GQ/5mq3/78yf/ABNB+NfwyH/M0Q/9+Jf/AImj2FT+VhzJHodFeeH41fDMf8zNF6/6iX/4mmn42/DMDP8Awk0XT/nhL/8AE0fV6v8AKxc0e56LRXnLfG74ZrnPiSPpn/j3l/8Aiab/AMLw+Gf/AEMaf+A8v/xNP6vV/lYcyPSKOlebH44/DMA/8VEOPS2l/wDiaa/x0+GajJ8Q5+ltL/8AE0fV6v8AKx8yZ6XRXljftAfCwZz4ifI6/wChy/8AxNNP7QnwqHXxBIO//HnL/wDE1DpTXQdj1WivKH/aG+FSjJ1+X1/49Jf/AImom/aL+FS/8xy4PH/PnJ/hR7OXYdmeuUV5H/w0X8K+P+J1cDP/AE5yf4V1Phj4peAfEcyQaX4msXnc4WGV/Kcn0AbGfwpOEluhWOzooyMUVIBRRRQAV478d/jnpHw83aRpdumr+IWXP2cNiK3z0MpHOf8AZHP0rovj349X4f8AgC51OAodSuG+z2Ktz+8YH5iO4UAn8AO9fAd9d3F7eTXl3O8887mSWR2yzsTkkk9TXfgsH7d3exE58p0njr4g+O/HFyZtf8Q3P2djxZwMY4F9gi8H6nJ965aG0gjIbbub1bmpgwI444oGT619PQwdKmtEcsqkmTRyshyB0p32mQn0NVy3enA4PFdcacexzzbRMtzIad9ocjrxUIZR/wDqpcY5z1rVU49jkqVGkWDdSDg9uBxSG4fHBPTBqHBJAzyacORhhjFaKEOx59SpPuO8xlkJ59aeJmOOcVBnJx3pA2M5+lWox7HBNTb3LIlbduzz2p4uJCAA1VQ2OD+dKc+vFNRh2M/f7lo3MnY89KTznwTnqagLZpFJ6HpT9nHsJc/csGeQ5yTSi5c55qsTkemKYTg++aXs49h3n3LhvJM4BwOlI1w55LHn3qqzHPahSdxFHJHsF59yx9okxt3E0v2pweTntVYkbc55zwKASQeMc0ezh2C8+5ZFy/Tdx70fan6biearde1BAJ+lHJDsUufuWTcOcfMR+NKLycZKyOPXmqpOB+hxQHyenNJwh2NIxm+peXULgceYWB9TThqUw6sfT61QBwvXNLkHrS5Y9hujJ9TTGpvnBJFTJqnP3yKxR1/GglckFR9RTtDsZSwsns2dEt8zf8tc9+tSLeMf4j6VzO454457U9JJVPDMB9aLwXQzeBqvZs6ZLqTP36ebx9pG72rmheTKOJAfrV7Rv7R1XVLbTLCD7Rc3MgjiRRyzH+nvSc6KV2jJ5fiejNA3cqN8rnNZN1cztO3zGvUNb+Eev6Np02o6lqmj29rCu6SR5mAH/jvXPA9a8/8AsSt8xpUnQrfBZmb9thH+8ujOFzck5JPTFDXFwSDuPY1qfYY16ilNtFjGK3+rU+xX9pz7mW91cEHJOKYbmfYFDGtj7LDnhBQLeEE/u160vq0Owv7Tn3Mdbi4AwCcYNO864P8AG3Stjyoh/Av5UBFwcAAfSn9Wh2IeZVH1MhprjJyW+bmlWa8JODJ36VqjBpTgdDS+rw7B/aFTuZRa+bH3ztxilP2/nl+Tng1p5GaQkY9Kf1aHYn+0KvcztuoNgliMe9OEd+cZlPTH3qv5xijv+tL6tT7B/aFbuUvs96XDNdHIx3pXtrl1Ae5JxV0N6Um7JNV9Wp9iXjq/crfZZcbTdSH8aaLIh9xmcn61b/GkLYo+r0+xP12u/tES2y9PMk6f3qQW6KxYM557tUm4YppYHvR7Cn2D6zWf2mNNvCc/Ln8aURxZIMa0bhTS+BweafsafYPb1X9pjvlP8Az9KUEZ3AAGomfgetIWJJo9lBdA9rU/mJzIB0FMMzg7lJBHpUbducU0sPUelHs4dhqpPuTm5cnJYk9aR7h3HzMSO1Vs5PJ96BtJJLUuSHYrml3J2uJMj5uBx1prXDsSWPJquzZ70zJNHJHsUnLuWGum59+Kb9occA9utVy1RlsHvUcsexpFyfUs/aJtv3jx701rlx/ETn3rX8D+FtZ8Zas2m6LFG80cRldpH2oi5A5PuTW345+F2veDtG/tTW77Soo2YJHGszF5W/uqNvJ7/SsJV6MZ8jtc6Y0azjzK9jjftD9mP50fanPOT6VTzxSFs5zW3LHsQua+5bNzJk/OeeetMNy/94+lV92QCPyzRySegFLlj2NFGTJzO/HzH86GupSfvn86r4z3pB1PNFo9jWFKfcti7lBxvPT14pRdSn+M4HvVVcFc0NwAM1No9jthBlg3cgxh2PbrQt1LjBc/nVbIyOOlNzzxU8kex0Riy1JdSkAB2P41E1zOQMyN+dRZB46U1iOBQ4x7G8YMsefKert+dJ9plyQHbB681Bk4yeKbuGT69qjlj2OiFNk/2iTOQ549+tNNxJ/z0b35qufTPFIASam0TZQZMZpSMF29smkaZxyGNRM3SmnPHFS4xLSsS+fIBw55pDPKf4zUZ6HPekA4GKxcEzSxMs0hPDH0600zSZ6n86ibA9aeDkcCslTuxMf5shx8xpRLLz8x9OtIIz34qTGPwrdU0ZSrKOwqlscsTx609ZGUcMefemKRsyaa5zj1rRQiuhyzqSkT+c4y29s/WozNJn7xpqnqODmjjPShxj2HG4rTOT940nmuDncRn3phAz9aacHj3qbI2SZMJpBkBiPxpDM/9456daiGDTgBu+7SaTNUrEN3CJQW3EP6isW8juon5ZmX1Fb554pjqpTBAPtXNWwkJq6No1GjnXuJuhY1Xllk4O4mtjUdPD5eIhWHbtWJPFLC2JQQTXlVKLg9TZTuP+0SAZ3Ent7VLb388T71YhvrVN8KvqaZnIH86zcEx3Z9CfAz9oTxD4SvbfS9euJdW0MkKyStulgHqjH0/unj6V9v6Fq2n65o9rq2l3KXNndRiSKVDkMD/X1FflCjlGypxX1Z+w/8RZotZm8C6jcFra7QzWIY/cmUZZR/vKCfqvvXn4ihZcyLTufYFFA5FFcIz48/bb1+S68d6doIc+RYWXmlc/8ALSQnP/jqrXz4MHJr1L9rOdpvjnrQ5Plx26DPYeSp/rXlS52jOK+nwMVGkjnqPUtKSeowaf0PFQxNgjnHaplHUZr1qTujFoMkdR7Uq568UYORk0u7AI7HrW8TKa0HgHqp5pxPIG7mmKQDweKGPO6tUziqRHjDP1pSM9PWm9cEfjUmSe44p3OaVK41uvHWm/xdcetOcbhjAq3omkXut6va6VpsJmurlxHEg9T3PoB1NTOooq7I+r3djP3bgF4qwltdBd/2efZ2Yxtj+VfXHwz+EPh/wpZQzXdtDqGqlcy3Mqbgp9EB6D9a9ESytVUKIkCjtgV41TOoQlaKudccscldnwEEIbGDx1BobBJGfpX2j8Qvhp4c8YWEkdxaR2l+AfJvYUAdD23Y+8vsfwxXx/4o0W+8O+ILvRdTiEd3aSGNwOh7hge4IwR7Gu3CZjDEaLcwq4B0zNGewzTW65OKeO/Ue+a9U/Zn8NR694/+1zwpLa6bCZnVhkF2+VQQfqT+FdWIxCpQcmY08LzysjyXjGcihTxnPFffx0LSCMHS7Ej/AK91/wAK+ZP2qfDsOk+NLHUrO3jgt9RtcMEUKvmRnB4HHKlfyrzMNm0a1Tksds8ucI3PHhgE7gcU5VBHNN79K779n+3jufi1osM0aSoWlJVwCP8AVN2r0Ktf2cHLsc8MMpOxwpUKByP8KaTwCGA/Gvv3+wtIYfNplmc/9MF/wqNvD2hHg6VYH/t3X/CvG/t2P8p2rLGfAhxuwDQRzn+Vfc+v+AvCeuWrwahoNlIGGA6xBHX6MvIr5P8Ajf4Dl+H3iSKCOV59NvVaSzkf7wwRuRvUjI57giuvD5tTq6dRPAtM4sjHJ4GMUzcuPvLxXa/s+NHefFzRLa5iSSJzKCjKCD+6c8g19jjQdIVPl02z/wC/C/4VOIzZUpWsaRwFz4DaRQM5phmGAcZr0n9o7wQvhPxmL2wh2aVqm6WEAfLFKPvxj25BHscdqX9mK0guvipbx3EaSp9jnOx1yDwOxqvr/NDnRqsHBHmTPLnGCM8imtubqxr7+1zRdJbQb8f2fac2so4iUfwH2r4Jtot6hjjGKzw2MeIbsOdNU1qhkERYjrjpXqfwG1zwf4W1a41zxDeMt6i+TaRLCz7AR8z5A644/E17x+z/AKRp9x8ItAmmsreRmhfcXjBJ/ePXenRNFXOdLsuf+mC/4Vw4rMVrTsOlh22pHx78b/iHP488RQafpckseg2T7ogylTcS95GHoM4UH3PesW1digB+lfa7aDobZYaVYc/9O6/4VyHxn0nTLP4W6/dwafaxSRWu9WSJQQQw6HFaZbmUKFoJbnl5xlc8Sue+x8wHGMcc0nlgqMdaajb1B6ZGacDg8GvsU7q58FJOLsIYwFyDznpUbhx1HtU2R16YoJzTEmyoxxwajZu+auMFOcqK2Ph3bwz/ABB0GCSNXSS+QMrDII9xWVer7ODl2OjDx9rNQ7nNq2OcikJOev619rJ4d0QoM6TYn/t3T/Cj/hGtAfro+nn/ALdk/wAK+efEUU/hPp48N1HrzHxTvJ9MYppPPXIr668V/DDwlrtqySaVDaTEfLPaqI3U/hwfxBr5h8f+Fr/wj4jl0i9O8Ab4ZgMCWM9GH6gj1Fejgs1p4rRaM83G5VUwmr2MMMKAxPRh+dXvCQWTxbo0bgFW1CBWBGQQZFr7UGhaPs/5BVjz/wBO6f4VGYZosJJK17mmX5U8Ym07WPhtpBkfMOPelL99w/Ovtmbw9oZ66TYZ/wCvdP8ACkXw7oRGf7JsD/27p/hXnriKP8p6D4bn/MfFBPHWgnIHP0r1/wDaotLTTdY8PRWdrBbrJbzswjjC5IZOuK8Z3EGvcweK+sU1M8PF4R4ao4MmJxTcknGaYXG3GaYXGa6bnMoEhY54OKbuwpB7Go99IWz1oci1ElLfLUZbpz+tMDDJ+lew/ss+H4dU8Talq91Ak0FlAIow6gqZH69e4VT+dcmKxKoU3NnVhcK69RQR5Bn1/Coi3HXv1r7ybRtJK86ZZf8Afhf8K+U/2jvD6aF8SpZreJYrbUYEuUVQAoYfK4A+qg/8Crz8HnEcTU5LWPUxWTzw9Pnvc86+hprNjknFM3/Ngda9J/Zstbe++KMMN1BFPELKZtsihhkbecHvzXfiK/sYOfY8/D4d1aih3POB9R19aa3IHI/OvvFtB0YLxpdl/wB+F/wpo0LRsc6XY/8Afhf8K8T+34/ynuLIZfzHwe3TJP0pjNg8fSvufxJo2kjw5qZXTbMMLOYgiFQR8h9q+Fl5jU/7Pau3B5isSm0rWOetljoNJs9w/Z78beBfBPhq/utb1B01W7nJeNbd2IjQfIoIGOcsevevMvih471P4g+LH1S7DQWcWUsbTPEUfr/vHqT+Havrr4VaLo7/AA38OSPptlI7aZbsWaBSSTGOc4rxT9sOztLPVvDRtbWG33x3Aby0C5wUx0+pryMPXjPGNvc9mrQccMoo8L+X3wKTOD0rrPg/FHN8UvDUUiK6NfDIIyD8rdq+0ho+mCL/AI8LXpn/AFK/4V6mLzNYeSjY4MPlntVzXPgL07U/kD610XxRaNfiT4jSNBGqajMoAGAPm9K5wkHGD+Zrtp1vaRUgeF5HYQ8DkGkBGMHg96XdnjFfWX7Nem6fc/CmwlnsLWVzPPlnhVj/AKwjqRXNjMasPHmsb0MNzux8mgZxQ2MYFfUn7VVlZWvw5t3t7OCFjqUQzHGF42v6V8tScmlhMZ9YhzWNp4fkdmNY8cfhUatjv+tT2NrNf31tYWylp7mVIYx6sxAH6mvu3w74W0nS9DstPjsrYrbwJGGMS5bAxk8dTWWMx6w9jWlh3LY+DWP4UhPA/Kvs743eFrbVfhlrcNrZwrcRW/2mIpGAcxnfxgegI/GvjHaWGB0q8HjViIto1dLl0YjnOOKjYkHkVMq88mvXvgL8JU8YE69ryyLo0UmyKJSVN04689Qg6cdTx2q8RiI0o80iox10PILW2nuTttoJZm7iNCx/SlvLW6tGH2m2ngz0EsZTP519/aPoOkaPZpZ6Zp1rZwIMKkMYUfpT9T0XTNTtXtr+ytrqFxhkljDqfwNeP/bCvsa+ykfnucsQcDA9qUHg56Gvbv2gvhPa+GID4k8OwsmnF9tzb5JEBJ4Zf9knjHY4rxRUBGTXo08SqkeaIkujGYJ6An8KPLbuCD2r7h+C2maZL8KPDUjWFs7vp8RZmiUknHfivMP2xbC0tNM8Ny2ttDCxupkPloFyNgPYe1ctLMFOr7OwppxVz5vMWcbuKcAFHHpUzAkc17p8Bvg7Bq9lD4n8Uwebay/NZ2bfdkX/AJ6SeoPYdxyetdtavCjHmkcic6rsjwu2tbu6X/RbW4uMdfKiZ8fkKLm1ubYqLm3ngPYSxlCfzr7/ALHSNOsrVLe0s7eCJRhUjjCqB7AVDq2g6Tqlo9rqFhbXULjDJLGGB/OvM/tmN9jX6nLufAZ3d6cRgdulevfHn4Wx+E3GuaGrnSZX2yxEljbsenP90+/Q/WvIlHJ5r06OIjVjzRMfZtOzIlUtnHNIwOADxX1v+zBpmm3PwmtJptPtnle4n3u0QJYiQgEkj0wK2fjH8PbDxT4MubSxsoItRtwZ7J0QKfMA+7kdmHH5HtXJLNIqpyNHSqDSufFzDA+tR7gOrDPTGas3kTQlonUo6khgRypHUV9m/CnQtIm+HPh+WTTbJ3fToWZmgUkkoDzxWuIxsaKuaQhfY+KldT0PIFOYHaOea9o/azsbSw8TaIlpbQwB7SUsI4woJDj0rxdwCBzzWtCt7WPMJojY/rRk4wOKCMEGmucHkH8K2bEIx+bGO9V7m3jnBWRQfepmOB05NNDnPT8awmkwuYN/pssLboQXT+VZ5OOoII4rriRjHHNUL7TorjLL8rjuK46lDrE2jPuYjNn2AFdX8JtZl0P4g6FqkDMrW1/C/HcbxkfiMiuYu7WWBsMDj1qbQnKatasDgiZCP++hXn1l7rTNY7n6yr0paZCSYUJ7qKfXiFnwN+1f/wAl113sdsGP+/KV5du6ZGPwr1H9rEY+Ouu4z92D/wBEpXlmdoxjJNfT4R/u4nNPcmVuxqwmccHNVAcHnvU0bru27jmu+nKzMmT9V+tJ2BGTSZAGKcRn0FdkWQ0KpOc0ev1pufQcUoY4xt/GtLmEoiqRux0FSZAOPeoQST0707dnnFJyJ5CXOM7ec17x+yBoEN3rWq+Ip4yxs0W3gJ7M+Sx+uAB+JrwQnj0r6l/Y+Cr4Av5B959QYH8FWvLzOq40nY2oU05Hr/iPUrTQ9CvdXvW2W1pC80rD+6oz+dfJeo/tDePm1tr6yt9Oi0/flLJ4dxKdtz5zn3Fe9/tNXLQfBvVwpP79oYjj0Mi5r42kjBXoMV5WX4SNaLlI7K1RwtY+4fBPjfSfEPgGw8VTXEGnw3EWZlmlCiJwcMpJ9CP5V86/tJa74R8QeK7K/wDDmqW99ci3MN6YMlflPyHd0JwSOPQV5DKjSRiFpJDEpJEe47c9+OlLDDFbrkBVBr0MNgfYVOe5hUqKcbWLLEE4FfVn7Jfh8ad4GuNblUiXVLglcj/lnHlR+u6vk6KdJLiO3i+eWVwiKOpYnAH51+gHgvSI9B8L6Zo8XC2dskX1IHJ/E5NRm2JvT5U9yMNRtK4T61Zr4oHh7ePths/tgXP/ACz37P515n+1PoI1H4aNqiKTNpVwlyCBz5Z+R/0YH8K4jVvGHl/tmwx+YUt0iGkN83DZj3f+jCK+hfEukw674ev9HuAPKvbaSBv+BKRn9a8Oi3RmpHdNKSsfn49zEpIL+/Arvv2bryN/jPoSAH5jNz/2yevMdUtp9P1C5sLtSlxbSvDIp7MpKkfmK7n9mls/GzQDuwQZv/RL17tfFOdNo5oYeMXc+6pG4GBXwv4o+KHxEs/GWrwW/izUYo7bUZ440DKVVVkYAYI6AACvueL5lGRXluofs+/DvUdXu9UurPUXnu53nlAvWVSzsWOAOgya8Ck4pvmOttpaHT/A/wARX/i34X6Rruqoq3s6MsrKuA5Viu4DtnGa8u/bbjhbwdoMpwJU1JlU99pjbP8AIV7voum2Oh6PbaVptultZ2sYjhiToijpXyX+1v440/xH4jsdA0i5jurbSS7XEsbZVp2wCoPfaBjPqxHatsPHmq3jsS1oc1+zwcfGfw+Bx88v/op6+30GU56V8Qfs7lT8Z/DxH/PSUf8AkJ6+4mwEyK0zD4kKmcJ8a/BkXjXwLd6aiAXsX+kWL/3ZlBwPowyp+tfOX7Le+P4v28UqNHItrco6MMFWA5B+hFfWPhvWrHxBpUeoWMm+F2dDnqroxVlPuCCK8utfAw0H9pCLxDZRbdP1axuJGAHCXA27x/wIHd9d1RQrNQcGEo6nqfiL/kXtQBJGbWUcf7hr4HsCPKj7/KK+9/EQxoV8e32aT/0E18DWQxbR/wC6K9LJlrI4MfeyR9q/s6kf8Kc0D/rnJ/6Neue/ae8XeI/CHh3Srvw5eraXFxemKRmiWTKeWxxhge4Fb37OLq3wc0Af7En/AKNeuw8Q6Do+vW6W+saba38Mbb0SeMOFbGMjPfFeXXajiG5bXOylf2SsfHUfxo+KmB/xUcIx/wBOUX/xNVvEPxR+IniDSp9H1TXI5rK6Xy5o1to1LLkHGQMjpX1o3w98D9vCukf+Aq/4V5R+0/4Y8N6B4AtL3R9FsbG5bUo4zJDCEbbtckZH0H5V6WGrYaVRJR1PMxccSoN30PFYSRGoPUCnbicjFQRSeZEjjowz/jTiT2ODX2cNj8/qR953JN2MEUA8UwE+ntSB8A5FaX0MrEu4Z6VvfDYA/Efw4QcZ1CP+dc7k5PFdF8Lyf+Fk+HBkf8hCP+dcWOf7iXodeAj/ALRD1PsMDEQ57V8x+PfjZ478O/EXWdMsY9MuNNsLxoY4pLc7mUAHBYHOeetfUDL+6GOuK+MPijEp+KPinfg/8TFv/QVr47LcNDEVXGR9/mOLlhaKlE+qvhb4utfHPg60163j8l5MxzwE5MUq8Mvv6j2Irhf2p9CjuPB9rraJ+/sLlVZsf8s5DtI/762/rVH9kB2XQfEFsG/dpfRuq+haPB/9BFdx8f4xL8I/EA/uWwcfVXU/0pQTw2NUY9GRUksZg3J9UfKXhSGRPGWhlhlf7Rt+R/11WvuQEBDg18R+DWLeK9EUnrqNv/6MWvt5VGK7c+fvRZ5/DjfLNHy18dfiZ8QNB+Jmo6LoerQWtjDFEY0a3RmBZAScketcfD8YfiqNh/4SGA4/6dY/m/SvrrVPCHhvVL5r2/0TT7q5YANLLbqzEDgZJGaqv4E8IKePDmlDH/Tqn+FcFDEUIxtKNz169PEyleD0PjPxZ4u8UeMNStrvxJex3LWqNHDsiVAoYgnoOeg/Ks4PXov7Smn2Om/E8Wel2kNrAumwsY4Iwi5LPzgd+leaOSPavrsvcHRTgrI+Px8Zus1N6ku8YyKYHAbpmo9w28cUwHLda7GzkUCVmJFIrEc/hUTNk4pNxx0pXKUNCVz+nJr66/Zt0H+x/hlZ3EiFZ9RdruTI5w3Cf+Ohfzr5P8PWU2sa5Y6TED5l3cJAuP8AaYDNfe9haQ6dpsFlAoWKCJY0A7ADAr53Pa9oqmup9HkmH95zZRtdas7rxBe6JFJm6soopZl9BJu2/wDoBrx/9rrR/P8ADGl67GmXsbkwyED/AJZyD/4pV/OsT4b+KPO/as8UIbgG21BXtEXdn5rcKB+iv+de0fFHQx4i8BaxpAXdJPav5Q/6aL8yf+PAV4mGbw9aMme7iYKrSlE+GsnOfwr1P9lwY+LEROcmwn/9lrys7lOGUqe4PUH0r1L9lqQH4swr62FwP/Qa+rzCalhpM+dwOHca6PrK9cpaOy9QpNfFV18ZfiitzKqeIUVRIwA+zpwAfpX20yK6YboeDXAy/Bj4dTytI3huDc7FjiV+p/4FXyGGnSi37RH0tVTt7h8zz/GT4l31hNY3GtQCKaJopCtsgYqRg844OK4xPljxjouMV9AftEfD3wp4V8Cx6loekRWdz9uiiMiuxO0hsjkkdhXz4GO1h7V9NgXSdNypo8qvGo5LnPuf4Of8kw8NH/qF2/8A6LWvE/20VX+0/DD5w2y5H6x17Z8HWJ+GPhkf9Qu2/wDRa14t+2ihF54YfP8ADcjH/fuvDwumMPRqq9Gx5d8G22/Fbw0c4/05R+hr7gbiD8K+H/g0f+Lr+GQCP+P5ev0NfcB/1HP92ts3f7xE4NWifCnxWGPih4nPf+05f51zRZh0OfWun+LRJ+Jvibkf8hKXn8a5UZxzXu4X+FE56sPeZIDuOK+vf2WW3/CWzGc7bqcfT5zXyCmB+NfXf7KmT8Jrc5/5fLj/ANCrzs4/hGuFjaRR/a7UD4aWxz01OH/0F6+UXfIFfVP7X0mPhtAueupw/wDoL18pAZA5+lGUO1EvEK8j0j9nbQzrvxV012QmDTla8kOOMqMJ/wCPMD+FfXXinWrPw34eutYv2K2tqm+Qgc4yB/WvF/2OtD8nQ9X8QSL81zcLbRn/AGIxk/q36Vsftf6kbT4WNp6HD6hdxQjnkqMuf/QR+deZmNT2tflXQ3w8LI9fkSK5tGRwHilQqw7MpGP5V8CeLtKl0LxPqejSAhrK6khGe6g/KfxGD+Nfavwh1c678MPD+pO5eWSyRJT6ug2N+qmvnf8Aau0L+zfiFDq8a4i1S2DMf+msfyt/47srTK6vJUcGKtE8ktIJLu5itoRulldY0HqWIA/U199+D9HttA8Mafo1qu2Gzt1iHHUgck+5OT+NfD3w0RZviJ4dhcfK2pQZGOvzg197rxBx6Vpm9RtxiFKNtTwb9o/4wal4Mv4PDvhqOBtTkiE088y7lgQkhQF7scHr0H1rH+AHxt1rXvEX/CP+LntmknUta3KJ5eWHJVh06cg+1eZ/tCsZfjJr7SZO1oUGfQRJ/jXBGNchlJVh0KnBFVRy6M6SYOuk7H2z8U/FPgqHwvqWm65runxC5t5ITF5oZzlTjCjJz07V8UR4ZAxPamiJRyeT1yTml7ewrswuF9hFps56k1KWh9zfAoj/AIVJ4ZwSf+JfHXlX7aTD7B4aOeftUwxn/YWvUvgTk/CHwwf+ofH/AFryv9s9WNh4aPb7VMP/ABwV5GH/AN7+ZvW1pnifgjRT4j8W6Toa52XdwqSkdRGOXP8A3yDX3dYW0VpYRwQRrHFEgREUYCqBgAV8g/szxLJ8WNP3gEpBMwyOh2Y/rX2Rj9yRW2bVHKaiRhoKMbnzP8d/jfruj+KLjw14QMEJsyFu7ySMSHzMAlEB447k55z6Vr/s8/GDVfFmqyeG/E3lNfiMy29xEmwSgfeUjpkcHjtmvnvxo5n8aa5M4y0mo3BbPP8Ay0at34GSSW3xb8PPGxBe5MZx3DIwxWjwUPY3LjVblY+xvF2h22v+H77SLsZiu4WibjpkcEe4PP4V8IajaTWV/cWU5xNbSvC4/wBpSQf1FfoO/EO4jtXwr8VAIviX4ljQ/KNSmPT1bP8AWoyuo03EmtFJ3Ppv9lQg/CGw5PFzcf8Aoxq9Um5U4715V+yid3wfsTj/AJebj/0Y1eianq9jYajY2N3OI5r92jtgf42VSxH1wCfwrzcQ37Zs3j8J8r/tMeDxoPjD+2rSLbYauWdgBxHOPvD/AIF9767q+ivg3Hj4Y+HSSSf7Pi6/7oqL4r+E4/GPg290khROV8y2c/wyryp/ofYmtH4UW8tr8N9At542imisYkkRhgqwGCD+NaVq3PTSYoRseCftkxhPEvh8+tnN/wChrXgrMB3r6A/bPGPEHh49vscw/wDH1r55dxg5r3MA/wByjnnuSs6kemKiY7ufypNwzxjp1IpmcHiulzJBWxnn86aSCSSKQ4PHOKaTkCobGkOLAjJOfSms7A4pGO0ZpjNnGazbNUh84EqbHAIqla2RTU7doeV81OMe4qxuJB6jHerGkANqVvk4zMn8xXNXV4MuOjP1HtP+PWL/AHB/KpKjtv8Aj3j/ANwfyqSvmnubHwX+1kP+L566Bz8sH/olK8lB6jFep/teX1vB8d9cikfDbLft/wBMUryM6haHgToP0r6DC1YqmlcwnF3uWvl4IOOeQakRsyZHQD0qtHdWzfdmjPHrUkUiHI3Aj613QmjJxaLaEZyM81IDnseKrLIo6+tSiQHkV205pkEq9c+tB5pMgjNNeeBPvyIPYmtHNLcLXH856HinZPXFUZdTt1Pykv8AQVUbVZDwkagA96xliYIXsmzbBypH619Ofsb3sc/g3WNPDL5trf72Hfa6jB/NTXyH9uuHzukIz2HFer/sr+OYPCvxKW01CdYtP1iP7LLI5wEkzmNifrkf8CrysfV9rCyN6cOVn0z+0RZyX/wi1uOFC7wxrcAAZ4Rgx/QGvif+0d4wkQHuTX6JXkEc8LwyorpIpVlYZDA8EH2rxv8A4Zy8ES6696JdSjtnfcbNJQIx7A43Afj+NcODxXsYtGlSKkeffDT4HxeMfhxb67eateaZqF3I725RFePygcLuU4JzgnII4IrzP4t/DLxF8PZLaXVb2yvbS6kKQTQOQSQM4ZG5HH1HvX3nYWFpp2mwWFlCkFvbxiOKNRwqgYAFfHn7XnieDXPG8Gi2UiyQ6NGySMpyPObBYfgAo+ua0hjKlSduglSSRy/7PehHxJ8XdFgkQtDay/a5sDjEfIz9W2196xg7cV8vfsQ+H2aPXfFEinqtlC3Y4+Z//ZK+hPGPirSfB/h6513XJmisbbb5jKpY8kAAAdTk1y4ufNKxUInjut/s+3d18RJfGKeLnS5fUvt4T7H9079wXO78M17xGxWP5uoFeRt+0r8LSMte6kCP+nF+leg+CPFWi+NfDkWveH5nmsZXZFLoUYMpwQQehrKTk0roGrO58e/tU6EdE+L1/cRJtt9ViS+Tjjcflf8A8eUn8az/ANmiQ/8AC6/D4B6tKM4/6ZNXsv7a/h5p/CukeI4o8tY3TW8xA/5ZyjIz9GUf99V43+y6N3xv0Dp9+XOf+uTV1RnemJH3fD/qxmvF/hJ8brnxF8StV8Fa/aWltJFcTxWE0OQJfLdgVYEn5toyCPQ17YVGzAr87PEWq3WgfFjVdXsH8u5stbnnjIPdZmOPx6fjXJTgp3NFsffvjbQrbxJ4futJu5rmGK4QqXt5mjdfcEH9Oh71+fXjbQb/AMK+K9Q0C/GZ7KYpuAwJE6q4HowIP41+gHhLXLbxL4a0/WrJ99ve26TIfQMM4+o6fhXz5+2R4Pw2m+MraPofsV5gdjlo2P47l/Fa6cJU5J8rIkjzv9nb5vjJ4dPT97J/6KevuCdsQ/lXw9+zuCvxf8N4OMzSD6/unr7enJ8k55q8erzRNM+av2b/ABr/AGf8RfEfgq/m2wXmpXNzYbj0kDtvQfUAN9VPrX0ukIbDFQSOlfn9rc91pfj/AFDU9PkaG7tNWmmhcdVZZWIr7o+Gvii08X+D9P1y0K4uIh5iD/lnIOHQ/Q5FY4ig4JSRSldl3xNEDoGoZ/59Zf8A0A1+fqkCFAp/hHA+lfoL4lI/sO/GCc20nH/ADX57wsDbo2D90da9HJn8RyYyN7H2b+za3/FntAAP/LOT/wBGvVD9pzxl4g8H+F9Lu/Dl6LS4nvvKkYxq+U8tjjBB7gVc/ZrGPg7oHOf3ch/8ivXQ/ETwRovjjToLDWxceVbzedGYZNjBtpHXHTBNcFfl+sNy2N4XVPQ+VP8AhdfxUKZHiJB/25Rf/E1heLfH/jXxbaR2HiLV/tlpHKsyxiBEw4BAOVA7E/nX0rF+z74D6b9V/wDAof8AxNU/E3wE8E6d4d1K/t31Lzra0lmj3XAI3KhI7e1elRr4SMlZHnYhYicWuh826a5MRjJ6cirfY+1ZlrJt2uMgEc/jV0vwMfSvracrxR8NXp2mybfgAGjce1Rbu/8AnFODY7ZrQ53EkVjnOOOldB8MmA+JHh09/wC0Yv51zYJx0zXRfDEk/Enw50/5CMXX61y41/uZeh04Jfv4+p9ox4MY+lfF/wAUsr8UfFS4/wCYk5/8dWvs+Efuxn0rz3W/g/4N1jXrzWb62vHubyXzZtt0yqWxjoOnQV8Xl2LjhqrlI+7zLCSxVFRicv8Asm2E9v4Z1jUnUrHeXqrESPvBFwSPbLEfhXU/tC3SQfCLXy7AeZCsQ57s6gfzrstK06y0bTrfTdOt0trW3UJHEg4Uf5714R+1n4pt3ttO8HWkyyXEsy3d4q/8s41+4D6Etz/wGqpyeJxfOu4nT+q4PlfRHkvhRiPFWiY6/wBo2/8A6MWvt+ENjJNfEfg5QfFeh8gf8TK36/8AXRa+4lGF6V25+7Siebw7qps+Yfjd8TfiBoPxK1PRNC1eO1sIUiMamBWYFkBPJHrmuLb4v/FjKkeIo22+ttHz9flr6O8X/CTwt4l8QT63qQvPtU4UOY5tq4UYHGPQVkn4FeCQcBdQ/wDAo/4Vhhq2DjBKa1O7FLGub9nsfNGq65r/AIk1qTWPEd1Hc3rxLDvRAg2LnHA47moXjjkB3KDXrPxy+H3h7wVoVhfaULlZbi8EDeZNvBXYzdPX5RXlGMjIOK+nwFWnUpfu9j4/MY1YVn7TcrSWMTcoSvt1FVnsJQTtw30NafTFKMEkf1rr5TlVeSMOWKVPvIR6momY/wBK38g8H0qvJbQOxJjX+VKUWdFPEJ7o9A/Zf0Iar8TI76WMmLS4Gnzjje3yr+PJP4V9cSqXQjOK8Z/ZU0NLDwddauVIk1G5O0n/AJ5x/KP13V6J8QfG2ieBtHj1TXZJkt5ZhCnlRl2LEE9B7A18TmlR1cQ0uh9xlkFGgn3PLvCHwEn8P+ObXxUPFclxNDdtcvGbUDzNxO4E7uM5Ne4SLmMg+leSRftF/DgqWM2pjHY2TZNemeGdcsPEmgWmuaZI0lneRCSIsuDj0I7HtiuKr7XRzR2w5ejPjL406EPD/wATNasVQrDJObqDjjZJ83H0JI/Cuh/Zb4+Llrz1sbgfotdh+2BoQS40XxGi/eD2Ux/8fT/2euT/AGVV3fFmInHy2E5H/jte663tMFc8+MOWufW02RCxHUCvkq5+P/xFgupoo20kokrKM2ucgEgd/avreYZgbvkV+euoHOoXQ/6byf8AoRrzssw8K8mpnfWqSgtDrvHvxU8Y+NdJj0rW2sBaLMs5W3g2szKCBk5PHJriix2MfY0Fun5UHlWHTivooUY0YNROKUnN3Z9zfBX/AJJZ4aJ/6Bdv/wCgCvG/20ebvwyf9m5HX3jr2f4OAL8K/DGP+gXb/wDoArxb9tHd9r8MMAdpFyPx/d185hn/ALWds1+7PLPgyAfiv4Y/6/1/ka+4zzBx6V8NfBkkfFnwvuP/ADEEHT2NfdKDMYHtWmbv94hYZWR8IfFcY+KHicf9ROX+dcuS1fbes/CLwJq2q3Wp3+jCW7upDLM/nONzHqcA4/KqLfBH4dDn+wQPpcSf/FV0UM1pwgosU6TbPjQ8V9d/soNu+Edvjte3A/8AHq+ePjbo2meHfiPqejaTB5FpAI9kZYtjcik8nnqa+hf2S/8AkkUOTn/Tbj/0KnmVRVKCkgoq0rGR+2Ev/Fu7VvTVIf8A0GSvlYNjn0Ga+rP2wP8AknEPBx/acH/oL186fDTQ/wDhI/Hei6MVJjuLpfNA/wCea/M//joNLL58lBsdVXkfY3wS0I+Hvhnomnsm2Y24nmBHPmSHe35bsfhWT8afhe3xGk0wtrT6elj5hCCHzA7NjnqOgX9a9HhVY41UDAAwBXlmsfH7wBpeqXenXE2otLazPBIY7UspZTg4PcZBrxrznUcomyjpY6n4V+EH8EeEIvD7ai1+sUruspj2YDHO3GT3zXC/tYaB/aPw+j1eNMzaVcLITj/lm/yN+pU/hXR+AvjD4P8AGmv/ANiaPLefbDE0qrNblAyrjOD6811vi/SItd8N6jpE2DHeWzwn23LgH8DzV0pyp1lKQpx90+GPBF8mm+NtD1CRgscF/C7E9hvGTX35C4e3DAg5XtX54X1rLZ3U1pOpSeCRo3B7Mpwf1FfavwP8YW/i/wACWdwJle9tkWC8QdVcDrj0I5FehmcG0pk0ndHzf+0nYzWnxf1OWSMhLuGGaM44YbAp/VTXNfDfwyPFfjPTdEYyLFPITMyfeWNQSxHpwMZ96+vviT8PNB8dWkMWqRyxTwE+TcwECRAeo5BBB9DVX4afC7w74GklurET3N7Kuxrm4YFgv91QAAB0+tVTzKMKHKtzJ0m5Hk/jP9ncQ2z3PhvWzhBu8i+Xt7SKP5rXz5IAMqT0JHFfbfxu8V2vhX4fajeO6/aJozb2qH+KVwQPy5P0FfEUSAJjPQd66cDXqVYNzFOmovQ+5fgKP+LQ+GB/1D468s/bSXbZeGv+vqb/ANAFer/Af/kkvhjv/wAS6OvLP22P+QX4b+X/AJe5ef8AtmK8qg/9qNa3wHlv7PV+th8X9Fd32rP5kB+rIcD8wK+1mb93X546XqNxpGqWWq2ZxPZzpMn1U5x+lfeHgjxHYeKfDVlrOnTLLBcxhuOqN3U+hByK3zODU1MKOsbHxZ8UNLm0f4j6/YzRlSL6SRMjqjtvUj8DW7+z3p01/wDFnRjChZbZnuJSB91QhHP4kD8a+mviN8LfDfjeRLnUopYbyNdq3EDBX2+hyCCPqOKt/DT4ceHvAkEw0qGR7ifHm3M7bpGA6DOAAPYCm8wj7Hl6kwg1LU6y44tzjrivgX4hXiX/AMQvEV5GQUk1KfaR0IDkA/pX2f8AGPxba+DPA1/qs0iC4MZis4ieZJmB2gfTqfYGvg5SeSxJYnJJ6k0ssi9ZBX1aR9jfslf8kesyCebm4PP/AF0Nc/8Ate315pOkeGdXsJDHc2erCWJwejCNv8K6D9kjJ+Dljn/n4n/9GNWB+2jHnwRo5I6amuP+/T1ytc1do1b5Uj1H4Z+I7Xxf4P0/XbYjFxEPMT+444ZfwINdUqgDpgV8qfsj+NBpXiCfwfezYttQJltNx4WYD5l/4EBn6r719VucD2rCvTdOdi99T5f/AG0dv/CR+Hsn/lzm/D51r53kbso9iRX0D+2k/wDxP/D2cj/RZ/8A0Na+eZGI+brk17mDdqSOSWrJT0FNBwxpM+9IG5zgYH6103GL/Dn0pobk0jMSfqaYSAeuTUtlRQjnI69aY2FUetOY/MTTOCfWpuaJCkrkDFXdFcDU7bpgSpn8xVBiOx496saSc6jb/wDXVf51lV1gw6n6mW3/AB7Rf7g/lUlR23/HtF/uD+VSV8y9zY/Ov9tEkftB69/1ytv/AESleKNzXuP7Zyqf2gdcOMnyrbP/AH5SvF9gIxivRo0m4pktlVVJzwfanxRyk8My496m4B6cZqQHjjp0rrjBohofA1wG4uJP++qtR3d0mP3zHHrVYADBFPU8etbxcl1IaRde6uJAN0pP04qMqzAEkk4qNOePxqQenc8inzSe7LSQ1o8YKn8KXy2KhgRnuKfuOefpSg8Y6Y/WqSBjPJk6ggj60phkAzjP49KlQ846VKenNVyJk3Pb/hP+0VrXh6xg0bxdZy6xYQgJDdRsBcRqOgbPDge5B9zXrdh+0n8N5IVeSTVomIzsazJI/I4r41ODyPyqVGCqPlrmlgoSdy1NdT6O+JH7S73VlLY+DNOntTIpU313gOgPdEGefcn8K+ZNTv5ZppJHkaSSRizuxyWJ6kn3rQL7wQV+tNEELZzEn4itaeEUFZClO57t8C/j14O8CfD2x8O32k6q91G8klxLEiFXdmJyPmz0wOfSs79oL43aF498Hpoeh21/EXuklmNxGFG1cnsTk5x+VeMmxtGP+rXnpig6fbcYXH0NZ/2feVx+0SWxkOzPz6V7t+zb8adI+HfhvUND1+2vZoZbn7RbNbqG25UBgckegP4mvJl0y0285/A0jaPanJ3sPxrWeBc1YzVRJn0R8Uvj14B8Z+ANY8Oiz1WOW8tysLvAuFlUhkJ+b+8BXg/we8VWngz4haZ4jv4JZre0ZjIkWNxyhXjP1rDm0n5iI5s/hUZ0ifHEiGoWAlFWG6ibPrn/AIap8DEFRpWuL3yYk/8Aiq+UPFWoQ6t4k1XVLdWWK8vZbiNX+8FdywB98GqjaTchc7kOO2etQW0kMbmOclHHGCKzjhfZGnOmtD6D/Z7+ONh4I8E/8I54isb+5W3nZrOW3QNiNuSpyRjDZx7H2rqviD8dfAfjDwbqegy6dq6Nd27Iha3XCyDlGzu7MAa+ZIJI2UFWU9uKtR9BVwwUebmuS6mmx1fwx8QWvhXx7o/iC/illtrKQtKsQBfBRl4z9a+iG/aX8FMmP7O1kn08lf8A4qvlUDcMdKci4J4rqqYKNZpsxjPlNjxBeQajreo6hCjJHdXcs6Kw5Adywz7813/wG+KqfDx7+w1O3ubrSrr96iwYLRSjgkAkcMOvuBXl56YBpyqCDmuqeFjUp8jJU7O59Map+0j4QvLCa3XSdaR5Y3RcxIRkqRz81fMaErAB3C4pgHzcjmng/pSwuDjh72M69TnPdfhJ8c/DXhDwPpvh3UdN1V7izRw8kMSshy7MMcjsa61v2l/BeAf7M1z/AMB1/wDiq+XtgznGf6UojUnoBWVTKYVJOVyY4txVrH1DD+0t4N3DfpWuAk/88F/+KqPXv2ivBupaFqOnR6drSyXNtLCha3XGWQgZ+b1NfMoCYp6hQ2RxThksE07mNXMdGrE0WRGinqFANXYnLIDjtis8HLAdKmgfgj3r34LlSR8jiY80my6CCBzg0qkZIzUG7BGOfahHOSau5xuJYzg/XpWl4U1eDQ/Fmk6zdRySQWN0k0ixjLEA84z3rKBwByKR9pFZVaftIuL6l0pezmpLofSD/tI+DEVcaZrhz1HkJx/49UH/AA0x4Lw2/StcVgeB9nQ8ev3q+cSiMSCg696b5MWDmMce1eK8hp3vc+ijxBNKziez+Kv2jr2+ga38J6C9u78far8g7fcIvf6n8K8deS9vdQuNU1S6ku765ffNNIcsxpiBU/hA7Din7iwwOK7sJl1PDarc87HZnVxK5Xoi/pN9Hp+tadqEqu0drdxTuq9SEcMQPfivoST9pTwaq4/s3WuD08hf/iq+biQcDHbBqMKvOQKWNy6GKacuhOX5lLBppK9z6OuP2lvB8YGzStcc+ggUY/8AHqZF+0z4O8sGTSNcVj1XyEOP/Hq+cwiN1RRS+XGesY6ccVwf2DT7np/6xT/lPTfjN8V9F+IOj6fpmlWV/by21957GeMAFdjL1BPOWrzpGII5qBUROQoFSbicZr1sHhVhocqPEx+LeLqc7RPu5oJBFMzg5BFND4OOortucNiX0wajlJCtjqBxQCDjmkY8n6VMtUOLs7nsngb4++E/DfhnTtCl0PWh9jt1jeREjIZh94/e7nJ/GuX+NHxR0z4jnS7PS7C9gs7J2mlN0qqXcgBQACegz+dedtFC+SyAn1IpVRE+4oH0ryoZVTjV9qe5LOZuj7OKsNls7d8/Jt+let/CX4yaZ4F8ILoGrWN/cmGeR4XgUEBGOcckdyfzryokdqjdUbqAfrXRisFDER5WcuEzGrh58zdz2H4t/GHwn468CXGi22napHes6S27SRKFR1YdTnuCw/GuB+DXi2w8CeM117Vba4mtxayw7YAC25sY6444Nc4u1RgAUPtK4YA1zwyunCm6d9Gd0s7nKanY+jLn9pXwgYcLoutkkf3E/wDiq+X5n82eSUA/vJGcA+hJNX3iibOY1pjWsZGVJU+1Th8sjh23Bnas7jU0mrFI9c0p5UnHbFWZbRwDsYEVXaORMbkPT61vKnKx2UcXRnsz6I8C/tB+GNA8G6Tol5ourPNYWccDvEEKsUUAkcjjivP/AI8/E7RviHJo/wDZNjfWwsjKZDcqozvC4AwT/dNeYNz2qPaB0rzIZfGnU9od/t1KNjofA2tW3h7xpo+u3kUklvYXKzyJGAXYAHpmvoiH9pvweEBOja1n0Ea/418sZyMY6UHgfX0oxOCjiGmy6dTkR9VH9prwdkY0TW/c+Wn/AMVTJf2m/B+CV0PWuv8AcT/4qvlkkkDOKUAEdK5v7Jp9yvb+R1XxY8T2fi/x5f8AiGwt54Le5WMKkwAcFUCnOCfSu++Cvxo0DwJ4KTQtS0/UZp1uZZi8KKVwxGBya8WyMYIpCq4yQOa6Z4NTpqHQiMrO57R8cvjD4d8eeC4tG0my1CK5+2Rzs06AKFUNnnJ55rjfgp4t0fwT4xOv6xa3NxHHavHCsCgsHbHPJHbcPxriUCgcinKQacMFGFNwQSqXlc+orj9pXwobOX7No+rmfYfLDRqAWxxk54Ge9fMLyPO7yyndJI5dz7k5P86YufpThwOPxow+ChR2CdVyWh0Hw48QL4Q8caX4iaKSWK1c+ckf3mRlKsB2zzXvrftL+EQMHSda/wC/Sf8AxVfMhIK9Rioyisegqq+Ap1ZcxMKzSszo/iLq+l+IPG2p63o9vPBZ3sgmEcqgMHKjfwCerZP40vgPxfrfgnWf7U0SdVZhtmhcZjmX0YfyPUVzoxjgijecVtKhFw5GTFtO6Ppvw/8AtK+H5LVF8Q6Nf2NyOHe3AljJ9RyCPpirGsftJ+DIbZjptpq1/Nj5UEAjGfdmPH5Gvlh8HAxnNAVfTpXn/wBl0+a5v7ZnWfE7x9rXxA1eO71BVtbK3J+y2UbErHnqxP8AExx1/KuW6qQOM00tgCkzXo06MYR5UZSbbufRPw0/aA8NeF/Bek6Df6Jq0ktjbLA8kQQq2O4yRXKftAfFTQ/iJa6XbaPYX1v9jnaWR7lVGcrjAwTXkQKkZJx7UmeetckcvhCpzoqVXmVmiQgEYrq/ht8SPEPgC9b+zSl1YSuGms5SdjH1Uj7p965ItyNp4pr46nFbVqMaisyYNxd0fVegftKeEJ7UtrGmanYTj+CNBMp+hBH6ik8Q/tLeE7eyZtF0rVNQusfIksYhQH3Ykn8ga+Udx6YxSSMygZFee8sp3ub+1Oq+JXjrXPHeuf2lrEoEaZFtaxk+VbqeoUHqT3J5NcqzENwBTQyjoecU0tz1zXXCEYKyMZXbue5/BD43aH4E8Fw+HtV0rUbiSKaSQS2wQqQzZxyRzyar/Hz4u6D8QPDtlpulafqFvLDeLcM1wqgYCsMcE8/NXirbdvUfSmE+prm+qQU+c0crqxbtL650/UbbULKQw3NtIssTj+FlOQa+mrD9pvwuNPthqel6st55S+f5MaFA+OdpLZx+FfLbkEYHFQuFY9BxVV8NCruVGVtD0j4//EbTviFrGl3Wl2d3bQWUDoftAAZyxB4AJ44rzFjnvinMwx0z26VGSARjr604QUFZCsiQvgY6CmlgGDdT/Kkzlfem8lc5q7isKW+bn60zI6nJz0pC68k9v1pobn+VS2aJCk9jTCcNkd+tL8xbJ61G7DoOam4x5J+gqzpB/wCJnAdpP7xf51TILZ54qxphxew84/eL/OoqP3WLqfqnZnNpCf8AYX+VS1DYc2MB/wCmS/yqavmnuan57ftkRhvj9rh/6Y23/olK8c8lK9n/AGx8j4963zx5Nt/6JWvGsEDjkdua9vD/AMNEPcRoI+ARnj1oEEYx/jS7j6mnBh36/StxCrFGGzjjuKcqIc7R0pucqDSAg8HIx0NAnYkGFb5eAalG0AYPP9KgAJ6de1OGcYHHqaq4rliNQ3WggDPTrTI2P4/Sl3/OQcGqTAl5zyKduI680zzAw+Yc9M012OcY4p8wEuSiYHOehpYmJyD1oZJBbpIR8jkgH6VGo9D3/OqUhWJ2IyacGOR06VCAxBqWOJyPvYpqokFh+QBy3WneZgAL8xPFCwr3J+tPRNuAB9KtVUSIBIecY4709Yiep7U4M1PU/Ln36ZrWNW4mgWIKvy0mDv6dKXcPTvSthuhxVc5DQ3Gex/CqWq6fHdIXXCyj7p/oav8ABYHpkU7GcgcYpytJWZKdmcVKtxaz7GyhHUVct9QlXHzEe1beq6el1FuHEij5T/SuaMMsUhjdGDA8jFeZUUqUvI6YNSRuW+py8AhT+FW49SHG5O/asGIPxhGx9KuIHwCUYfhWlPEPuNwRtpfQNwSV+tSpNGx+Vxg1jJDJnOw+vSpY1bdjB4rqjijKVNGwuD0peMCsuMyKdoJHpVkPPtGDu4rojiYsylSLqnGc9acOT/Oq0UkjHDKOKlEnJ+Uit4VoswnSdh/4U9DgY4+tRrKv9KVWBJAPeuiNRHn1aTJR1/rT4m2tUYbHSnnkA9CP1rZSueXWok5JwOMU9SQBx7VFuyB2FODHrx7VZ50qbJUcj8KGb5hiotzZxilVsEk80XIcGSFietKWx0HtUZcd6Td+Y4oTFykrHIpN2e1RbmCgUqswPvVhykjnG2gMe5qIN6A00k9e1Sw5CwTQx6c9RUO4jqaXdkdaEHISFl4wTmnbzkED2qLOOwNJvz2pk8pMX7Z4zSE4Gc1HnA6/SkyT3ouHKTbj6445pAe2c5FRqSe9DcLkHBphyjyeP60KQeCcUxTnjpmm5wSKQcpIT05pA/OB+dRgjvQSD1pNlcpLnBBoZjxwMVGTkcUAnHX8anmQ1BjyTjHvSr93NLFHLMo8pGb/AHRmrVvo2rTMRFp902ef9WQKzlVgt2WqUn0KjPxgjp3pmQTya1o/DWuSSFRpswP+0MD8zVlPBmusfmt40z/ekFYyxdJdTWFCr0RzUiRtklRVWS3Uk7MiuxTwTqrL88lsmOxck/oKlj8DXfJkvIAAccKTWM8bQ7nbShio7XOHa3YKDkHpUbqV+8Mdq7+PwU5yGvV9PuU//hBYsHdfsR7JXNPGYfuelRr4lfFG55z1pccfhXoD+BreQcXre/7unxeAbTHz3sh+gArneOpLqenTqOS1R58n3jxQDnt+NehHwFZhv+PuYj6DpT18DaeG5uZ8H6cUfXqfcvmPPTnBHt1pikjFelHwJpu3/j6n/SmS+BNOC5W6nBA74oWOp9xOXkedjrS5Nd5/wg9mVbbdzAjpwKjHgfDnF7xn+5VLG0n1JdS3Q4YjJBoINdv/AMIQ7E7L2PA45TFNPge+AytxbsO2TWqxdJ/aM/bNdDigD3pwXnmuxk8E6uOIxBJ24cVC3gvxDyRp5cezqf61pGrTl9oX1l9jkW57dKRugroLjw3rkJPmaTdcHnEZOPyqnPpWoRD99YXKcfxREf0rS8e41iPIyzwBSEVbeMpw67T6EYphAPQj609O5qqyZWP0pNuasEDOeKaQCfenZBzkJyCG28CmnJ7GrBxQcAUnC5amVvmI6d8UcsDxzU5xjpzQoB5rKVIfMiu8ZBGBTQjgklT1q2ckemKRnwuCKn2ZXMVRG+QaQW7ljkirW7FMLnIwcDvScA5iB4JMZBHSmNbybflYZq0WOB/SonchsY7VDSBSZVNtIP4hzTDBIPQ8+tWvMB4H50xmx0NQ4ormZWkimz1FMMcp4zirZkH5d8UwuB3qHYabKvlOMkikaNgucZxVgue4qMsc8nOO1SzRNkLqe2fSoipzyO9WTnuaQsQMkA570mhkGGPY1Y00N9uhOOjr296AeoHerGnc3cXIGGHP41jU+Fi6n6k6d/yD7b/rkv8AIVPVfTP+Qba/9cU/kKsV829zY/Pn9soY+Pmtevk23/ola8ZYnFey/tmZHx81odvJtsf9+Vrxkk8HFe1hn7iM5bilmzx07ijOBkdO9CjkjqKQsfSugnUdken0pVYZ6Z7VGx4zQhz+fNAtybkdqeCSQvTH61CXHFIZM/UGi4+UsZOcBuOuDSh+SPeq5f5TSKegzSTHylwHHGR60nmAjB7cE4qHcT2p6kBCGA5qxWJS2cY4FKGY9+nFRLz9R+op6+/NUmDNCLaqDaOvU1KemR9MCoo8KoAI6U/A2g5IJNcbm7kseSdxAOBT1Hoc1GpwPfPpTgc9G9+a0UhDhnoPmA7Gl3ZwBxTSVPtSAYcGrjNisSBuTwKcccMDxTUJOS2KUEjOOncVsqg7C7+BkYpysMDAOcUzhiQO1LHwelaxmZvccOeD19aiuoQ67x94e1S5B6Gl6oD6UTSnGzEm4u5SIOOntTCCBwSR/KrVxH8m8dRVV2yBxjnmvErU5UpWOuMlJDw5Bzz6EUpJJ6HrTSMnGMH69adbKXmKgbj6dc06c2yJIlxkqQRmpFJPUcD2qZLCdceeVtx/00PzD/gI5qxEtnGgLRvM4HO47VP4CvSo0akyHJIqIrlgFJOecCrcFjczruWPavdnYIP1p639xGmyAiBcYxGuOPr1quzNJksxP1NenTw7S1Zg53HC2WPcJbiIc4wp3fypHWzRQYpJpH/iyoUfzqMjjk0xq05eUXLzbl1bm2X/AJdXbjjdIf6CmyXYIGyFUH4mqhwR9KcD7fpT9pJbGbw8Hui4uoOOBbwfXZSG/kIOYof++MVTBOMjmg5I4FS6sw+qUn0Lg1AqBuihbnqQR/Kj+14g2DZR9OodhWbKpI4qu0cmeDXLUxVWIfUaL6Gymp2xOJIpE/3Wzx9DUsd5aOcCUr2y64/lXPlTnLGlLDHpWMcyqxMKmVUZ7I6hIvMiEkUkcgJwMOM/lTZEdOHVlPXniuYeUqoAJ3fyqaPVr6FNqXD7c/dY7h+RreOctfEjiqZIvss3vlJzn3pC3+FY8WsZGJ4RuHUpx+ladjLBeOEimTcTgK52n9eK7qOZ0qm7sefWy2tT6XJsjAAPNKGI6EV3Oi/CbxpqsMc8OmJHDIu5ZJZkAwe/BJrqtO+AGvuoOoaxYW4PURq0h/oK2ljqEd5GEcHWltE8dz2FNyQcA5r6AsfgFpse03muXUpwQwjhVQffnNatt8EfB1uQZTqFyf8AbuMD/wAdArmlm1BbM2jldZ9D5ryc0oUs2ByT2FfVVh8LvBNkAU0OGUjvO7SZ/M4rbsfDuiWAH2LSrG3x0McCg/nisZZ3TWyNY5PUe7Pkqw0DXNQx9i0m+nBHBSBiD+OK3LH4beM7sgjR5IQehmkVcfhnNfU7RLt+UAdgKgeMDIx1rlnnc38KOqGTR+0z57svg7r8mDd3tlbDvtJc/wAq1IvgzGjj7Trkjr6RwAH8yTXtDpx0qEx8kdD6muWWb12zqhlFFHlcXwk0GPBkur+Ujr86gH8hVuP4aeGIwAbWaTHd52yfyr0Fo8Z47+lRSRAisXmNZ9TRZdRXQ4NPh94YiyRpoODnDyuf61ah8K6HbMGg0q0Rh38oHH511xt3YEqjZ6cCkXTLtydsEh/A1i8XVl1NI4Oktkc99hiiGI40Tv8AKoFVp4jnJHANdW2hai2P9HJ+pApV8K30pO8xp9TmodWb3ZqqMFsjjnhBHPHHaqskW3jHX17V6FH4MLL+8uwPXatPHgewBJkuZ2+mBWfNI1VJHmzQ4wA/BPSopIT2/TuK9RPgvSOhMx/4HUyeFdFjGPs27jHzMTU3YezPIjBg8Dt+VN8kjhfrg9q9iHh/R4xhbGH8s0q6LpYHy2NuB/uCgXKeLLEc9D60/aT0X8q9pGn2Q4W1hH/ABTTYWpPFvFnr90U7DSPGmiYY4NNaN+ynHSvZXs7U8fZov++RUb2sA+7BF6YCjigDx4ROSTtbGCckU1kJH6V65JbwdBGmPTbUD2dtz+5j9fuigTieUMjkYI6Gl8obcE89a9Nksrc5Hkx/98iqr2Nrnm3j/wC+RRdhyHnu30HA4pxX5g2CT656V3E+m2nX7PHk+g6VVl0y0Of3QH40C5TmYiwOB/n3rTtdwAPfHWrE1hbggBcY9DQsKJ0JA60JtDsiROoy1X7XDHJG4H1rL3beM9KsQXix/eXI6U+eS6i5Uag0vTbtSLnTrWTPXfEpz+lVbj4deDtQXE+hWwJ7xAxn/wAdIqSDWLZAA5ZfwrYsNXsHIxcxj6nFXHE1F1G6aZyN78DPBt2rC3+32TNjBiuCQPwbNc9qn7N7OpfSvFEqn+Fbm3z+qkfyr23T7mGXGyVHB9GzW1C4CjjtWyxtVdRKmj5L1v4B/EDT9zWM1jqSD/nnMUY/gwA/WuH1nwZ440ZmGo6BqUarkl1hLpj6rkV94gqwqN4IpOqg/WtI4+fUq1j88Wu7hXweCOoIwaDqM4GdqkfSvvXWPBnhjV4mTUtFsLrdwTJApP54zXAeIP2evA+oq72Md1pcrcg28pKj/gLZH5YraOOv1Hfuj5K/tSXZzGufWkGpHun617l4i/Zn1mBS+i6/a3YA/wBXcxmM/mMivNPE3wp8eaCjSXnh65liUZMlqBMoHvtyfzFarFN7MpOLOa/tNcDMWPxpDqMRJ/dnNZ0qSJIUkVkdeGVhgj2IphPcfnVe3mXyxNIahHjlTmm/bIieSQKzCR2/Ok3Zpe2Y+VGm11EejEfh2pomjyfn6+tZofCkY60zk9PWj2rK5UannpuJ3A4GMf1pomXsRz+lZhDZ4PehVPJPSodVgkaJYBB8wJPvSFskc81nHJ704DGPm5780vaFaGjnPT0ppGQKpb2B4Y8e9OVnxnzD9KftBMvbCSc1Z007byI8cMOtZxnY9M1LZTN9qj/3hUVJ3iyVufqlpX/ILtP+uCf+girNVdHOdIsz/wBME/8AQRVqvnXubn59ftnYHx71k+sFt/6JWvFQcda9l/bWLL8fNYAJx9ntv/RS14edx7k17GHf7tEtFsyR4+9k1G0/oKh20dK3uKxKZJCfSkHvnrSc+tP7U7jsOx0FL93kCm4HrSk5FAMVevp9acCeh9aYCSfen8FsccU7iJVfOcgZ705fu1EpIPX/AOtT1bacjOaExMkAIfPY9aeODioiSerU5SQOTyO9VczNOJtqAcU4Zzww60y3UFFOccUshA9c+1cLkrlWJCdwUj+VPGR15PbFRBlJBXgjqKduBx9apSJHgckfw9cU5RTTwQP5UoZlHHftWikSPwN2Wyc0mT0I70BgeN2O/NG4EY56/nVqQx3B/hPWlLKCB+HWmk56cEdRQdu1s/hWikKw9mK8dR2o3njI74470wM3QjP0p4OCFwSzdBjNXGbEx5Gcc81DLaTSPm3jaRjxtUc1dt40QhrgkcfcH3v/AK1SSXTlfLiAhT0U9fqe9dH1T2y94lS5XoRR2EMSobufLY5ii5YfVug/WnCcRIFto1hXGMr94j3ao+AaAMiuilg6dLZBKbYhYvIScknnJpQwxmlxt7gk0D7uDXRe2xIuG3HkfWhFIBHvTgCRVnTtO1DUJBFY2VzdP/dhiLn9KfNbcLIrEA4phQHPHFei6J8H/Heq8/2O1lGf47uQR/p1/Suu0z9nfVGAbVfEVlAO6wRFz+ZxXNUxVKO8ilFvZHhR6jtSdH46Gvoq1+A3hu3Y/btfvJsdl2oD/Orb/CX4e2yjzJLiQju1wf6Vj/aFDuaKjN9D5r5A45ppJHTjFfT8Xw2+HJjOLENgdfOY/wBadD8N/hvLn/iXLjviZgR+tJ5lQD2Ez5dOcEmjy9+M9q+qP+FS/DO6B2208LeqXTD+ZNV7n4D+Cpl/0PV9QtmPQF0cD9KyljsPLRh7Oa6Hy+0BOeMfWoJISCK+idV/Z8lKMdL8R28hHRZ4Sv6gmuO1r4J+OLEM0VjBfxgdbeYE/kcGs5ewqfCxXa3R5DIjAfMM1BIMAe9dJrXh3WdKkKalpV7ZsP8AntCyj8yMViSwnn8/wrlqUGth3KmeeR7fjTuSuBTnTB4GRmkAArkkmijrPB/jvxV4UdTo2sXEMYIJgdt8R+qHivcPBf7Q9jdBbXxTZmylJwbm3BaI/Veo/DNfM2/t0psjdMdazbuFkffWh65p2t2KX2mXsN3buPlkjbI+nsa0EBY5GTXwb4V8T674av1u9G1Ge1cH5lU5R/ZlPBr6V+FXx20jUkjsPFiRaZdnAF2o/cP9e6H68fSlqJwT2PYBEcdDTDayschCePStSzuLe5hjuLeSOWGRQySRkFXB7gjg1bUAjjpS5hOBz/8AZ9y+P3ePqacNGlf7zqv61v8ATnFJxzxSc2VGCMRdCQfelJ+gxTxotsPvBzj1Na5IpjE+1CkxOKRQTSbNf+WKHHrzTzZ2yjAhjH/Aas5/OkbnGTQmyWis0MY4CL+VNICjIFSkUm05qkTYhZVJztqNuvT8qsFckHBphikZvuk/hVXGkRAMV5qNs54q2ttL/dNOWykY8gYoUkWkZxB6imEEe9ag05j1fmnDSwfvOT9KHJDsYmM9/rSbRjOOtbg0qHuzH8akXTLYYymfqannRPJc50rgU1lOCBXUiwtR0iWnCytx0iX8qTqItUzjpAR1/Cq5QsccnFd19lh/55p+Qo+zQr/yzT8qXOHIcCyP6HiopImJBKN+VeifZ4s4Ea/lSiCL+4v5Uc4uRHmcsLkfNGw/Cq7RMCcKeterGCI8FF/Kk+zQk8xR/wDfIp+0BQR5HNG+OaqyxueO9eyG0tz1hj/75FRtYWjHm3i/74FLnB00eKSo278agaPnIHvXtcmkaa+d1lAfqgqu/hzR3+9p8GfYYp86IdJnjJQ8nHaq8qsDxXtMvhXRZBhrFB9MiqcvgfQ5CcQyL/uyGhSQvZs8bm3bdoGM1VlLbcD8xXr9x8OtKcHZcXS9uSD/AErKuvhjncYNTx6b4v8AA07oThI8yjuZ4TlJiuemD0q9B4k1q2A8q+mwOg3ZH611N18MdUUkxXdrJzkZyD/Ks+bwB4giJxbJIP8AZkBqtCeWRHaePNfhUCUwTj/aTn9K3bD4i7iFu7Bgf70bf41y0/hPxBbgmTS7g44yq7v5VRksLu3Yia2ljPOd6EUC1PVdO8a6LdEB5ZLc/wDTReP0rfs9T066H7i8gk9g4rwmMcADj3q3HJJH0Y1N2hpo9yZlZchgQahdYz/DXkVlq+pWvEF5Kq+m7Ird0/xbqCcTiOYe4waqMrCepueJPA3hfxGrDVtEsbtm6yPEN4+jDn9a8p8Xfs1+H7sNLoGo3elyY4jf99Hn8cMPzr1Sx8XWzEC4hkiPqPmFdBZ6jZ3igwTxvnsDz+VbRryQctj418X/AAI8d6DmWztI9ZtwOXtPvj/gB5/LNeY39jeWU5t721mtpl6xyoUYfgea/SM7SMYzmsbxL4U8PeI7Q2+taRZ3qY482MEr9D1H4VqsT3LU2j87WDEZNIV7V9W+PP2c9CvTJceGr2XSpjyIZMyQn2/vD8z9K8L8Z/Crxp4WZ5L3S3ubRck3Nr+8THqccr+IFbxknsXGrFnDfL3FFDD07UDPJqjRMULjkkU0DBp2flHFIA27kUWExMHd9afjAFKCAOMg/SkJosIfzuJNSWQIukI5+YfzpnJ7VPZrm4UkjAIpTXusFufqhonOi2J/6d4//QRVyqPh850HTz/06x/+gir1fPvc3Pz4/bZH/F/dU4/5dbb/ANFCvEMA969x/bZH/F/NT/69Lb/0WK8Qxz6V6+H/AIaJe4hXkUgUYz+BpwBxzQmQOtbgNIGcDNOGR0PWncA4A60m1uhFAXDpwfWnHFAyetKw9OKbATAwOTmjkd6UDnB5zSr1waBjg2Dj8KdkDacZIphz1pcnbTuQPZsgDHFOjJPGKjB9qBwetDZNjZhJMa8ilznJ6c9KZb8xp34qZydikgHgjpzXlTm1IuwwbdxAX8qevT6VEvJwQetPAA6N74q4zJaHkoAcgihWG1SBwfzpOSo5x6UE9K1UybCvwQOxqRdoOM4qIHbnBzn1oVxnBB+taRkSyeM9T3xRgkikOMelXre0ESrPdcIRlUz8zf8A1q6aUJVHZC5iC3tZJo9y4CA/M5PAp6lYeIclscuev4U+SRmAUAKg6IOgqNlyvevXo4VQ1Zm53EJ6knJPWjPFOVDjNAwTtArpvYQqqC3JPtTgPetXwx4d1jxJqK6fothNeTt1EY4UerHoo+te8eBPgJpuneXfeM75byUc/Y7disQ9mbq34YrlrYqnSV5McYOWx4Domi6prd6LXSbC5vpm/ggQtj6+gr1Lwn8AvEV8En1+9g0iFufLX97MPqB8o/M173DeaH4fsRZaLp9tZwr0jgjCj6nHWsy712eYn58D0rx6+cPaCOiGFf2mYGhfCLwBoJWW5gl1Sded13JlT/wEYX88110GoabpkCwaZZQW0Q4CRRhFH4CubmuJZg2Scjr9KYiTEdCcV5tTF1am7N40oxNu78QXOSBIFB6YqhNq00vDStn61XNlNNjqKsW+kSAjgnmuRuTNVZGdfPNKhxnIrHummxtHcYruV0ZmwdpPGKpaj4ckCmZFJ7kelJcw7o42ynuLOUyw9ehBGQRV9NQvnBYGMD02AZrSi0YYxViLSAo479vSm7j0MQXVwwyyAd+KsQ38sRX5mGRg89K2RohYUHQmI7flUXkPQji1OdQCJWGPQ1cTW7wABZR+NQHR5EGV5/pTW0+dOcE1SlJEtI1l1u3mhaG/tUljbhgy7gfwNcpr3gH4deJ95TTY7Sc9ZLM+UR+A+X9K2ZLGBYh55lBx0VRWfA1xp8shtZCiucnIHNbxxNSHUz9nF9DyTxl8C7qzV5dA1aK8jUZENxhJD9COD+OK8s1nw5rGiymPU9OuLZslR5icH6Hofwr6k1Wa6uIMpKyyA5UjufQ1hHVWMb2ep20dxE33lkG5T7YNdMMTGfxClQ00Pl+RWB5H40meOOle8eIfh14a19TPpMg0q7PRVy0J+o6r+H5V5X4r8F6/4clJv7JjbZwtxH80TfQjp9DzTlTvqjnlFxOeXAwoPT+dPWRl5Xj1pFXIP9aDgYx9CaRJ6J8Nfip4l8GXEcdvctdaaD+8spjlCO+3+4fcflX1f8N/iJ4e8b6cJ9LuPLukXM9nKcSxfh3HuP0r4PByAR9Ku6Xql9pd5Fe6ddS2tzEcxyROVYfiKTpqS0LUu5+iQbcoI6U4ZavDvgx8cLDW44NE8UPHY6pwkdwfliuD2/3W/Q9vSvc4cMuQeD0rBprcfKN2Me1IYmJ7VOKXtU3CxXEDdCR+FOEC45qcUYFFxcqIPs6ZzjJ96cIkH8IqXFJjPtRdlcqGCNQOlKExSmlHWgLAFGKABilFKKQxhVfSinA0UAJgUYxS0ECgQ3GaXpx+FAooGJjpQM5p2KQUCEPpS9KMe9FAAPajODjrSClApgJjIz3oxmlxQQKQxoFL34paTpTEIOe1L35FFJSADg96MD0pcUYFADNtIETHQU7FGKYDBGvpUU1nBMCssKOD1DKDVgCl70XYrIwrzwrod0WaTToNx6lV2n9Kx734f6RKS0BmhPba2QPwNdqfejAo5mLkR5hd/Dq4jGbW9R+Ojrj+VZVz4T1q1GTamQDvG2a9jx7Unlqe1VzE+zR4lJZXULYnikjI/vAipYWkQ5RmUjuK9ins4JlKyxK4PqM1kXnhfS5yWWHyye6HH6UKQuQ42w1i/gI/fs467XOc1rweJUIxPER7rUlz4PkXJt7kEAcBhism40LUbZjvt2ZR3XkVSZDizoYNQtrn7kgz6HrUkkMcqkMoIrj1jeNirBlYdsYrRs7+4hOA+5fRua1U7GTic144+C3g3xSZbh7AaffPk/abPEZJ9WX7rfiM+9fP/j74D+MfDZeewg/tuyU8SWq/vFH+0nX8s19eWeqxsAJgUb17VqRvHKgZGDA9xW0cQ1uUrrY/OCaCaCZoZY2jkQ4ZGUgqfcUxlxjmvur4h/DHwn4yjaTUtPWK8/hu7fCSj6nHzfQ5r5q+I3wS8UeF2e6sIzrOnKSfMgX96g/2k/qM/hXZCpGa0LVRXszyk5J4HT0oAwc4zn9KkZSGIPBBwaQBScCrsa3uAHB7VNaACZcZ69u1MOG7YqW1TMq/Wpn8LEtz9SvDn/Ivad/16Rf+gCr9UPDnPh7Tf+vSL/0AVfr517m5+ff7a+D8fNU9fslt/wCixXiAwDk817d+2sT/AML71TPT7LbY/wC/YrxIc9OK9fD/AAIlickcCjHTNKARgAUBT371uK4YHTpS4x2604ig46Hv0pgNAG7nNKu30NLwRjoKb1HuKAHKuTjNAGD7dKUCjrxQgYYx360AD8qUdckcelABxQxXEb72SOOlL8oPGTmg+tSQWtzcELBbyyE/3EJqGHMjSsvmt1IPvzUjEHntmrumeH9duIUSHSL5sDtCQPzrpNN+HHim7xvskt1J+9NIB+gya8ucXzMfNGxx2ATljx1pQQpY5+8OPavU7D4Q3UgBvdUgjHQiNCxz+OK6HTvhB4ej2/a7q8uT3wQoJ/AU1FmTqI8MJDEAHHrnvUsMckjbYkZyegAJNfSel/DjwlZx4XSIpD1zKxY/rXT6Zo+lWKk2em2lv/uRKDWy0M3UPmTR/BPifVlDWWi3bqwOHdfLX8ziup0T4J+KLuT/AE+ezsI8jJLmRvwC8frX0TAEVAcDHpXlvxd+IiWUT6LoNyDdnInnj5EQ7qD/AHvftXTQpyqy5Yolts8w8X+H9D8MXY0+11GTVtQj/wBc5ULDEfTAzlvxwK52V3lfc5JY9c0sm5yXZixJySTTR09q+io0Y0Y2JHj5QcYpOcnjmlGT1x0rZ8H+GNa8V6smm6PaNPJ/G+MRxD1Zuwq51EldgkY0MTyyLFGjSO5wqqMsx9AK9j+GnwRvdQEWq+Lnk0+zPzCzXidx/tH+Ee3X6V6j8OfhvoHgWyS8uhFf6zjLXDrxH7Rg9B79TWnq2uGR2RJML6V4eLzRR92B1U8O5ass2c+h+GNOXTtBsYLKFR0jUAt7k9Sfc1g6prkt2+N5zntVKUPcnIB561Ys9IdyGIOSc5rwKlWdR3Z2KEY7ETC4mIbc2ewq5aWUr4BBLZ/Ouk07RQQMp2rat9GQIABgjocUlBsUppHN2uju5BYZrUtdG+bcRk4x0rprSyVUAZRkdauJAi9hVLQzc2zmE0UiXzFAPYqe9atpYRFP9UFI6gitXaOwFAxVEttlRbKJRwAKcbOM8FRVjIFNMsYON4zRdiOb1bREikM0S/IeoHb/AOtVJLRRjiuwM0LqQWB9qxdRVLaXfjMZ6H0oHzMqW9oFGGwRUzWsZ/gqMXsP97pU0N3CxwWFK6DUEsVfqlNudMj2AhcfMOladtNEy8EVZCrIhoFdmVc6RBJMskke5cYIFYuteHYWQyWwO3riu0UDYO/FNeNWXbik1cpSseU6lpT29q0u04Vea4zVrCa4tvtccbMAPnA7e9e3a/pgnsJooR8zKQBXP2nhtre0VeA2ORipcWnobRmmtTxDzZ7eRtrsCTzzxXQ6Prg8k215HHcW8gw0UoDKw9CDWh428MNZyNeW6FonPzf7J9PpXHFDG2OeP51rCvKOgnG5J4y+Dmn61bPqfg51tLrG57GRv3beyN/CfY8fSvDdZ0u/0i/ksNStJrS6iOHikXaw/wA+tfSPh/XJ7BkWRiU42sDk11+s+H/DPxB0cQ6tbI0qriK4jwJYj7H09jxWscRd2kZSo9UfGrZB3buSOaQEA89+RXbfE74da14JvsXS/aLByRBdoPkceh/ut7flXEKo5OfpXRFmEo2JwxG3b1r6E+Bfxsex8jw94uuC9qCI7e+c5MQ7CQ9x79u9fPCDgH+tPDsOce3StZUlURMaltGfo/BJHLGkkbKyMoKsDkEGnivlH9n74xT6JJB4a8Szl9KJCW9y5ybU9lY90/8AQfp0+qbeZJY1kRgysMgg5BFefKDi7M33Jce9GKBRnNQIXiiijigYH6UD6UUUAAHvSijtRigBKKWigAozRQaAE60UUd6AD8KXpRSUAJjNL2o6UlABxRS0UABopfek5oASjig0ooASjFAowM0CExRz6U7HNIaAExxRjnmlpCKADoKXv0o5zmgUAJR0pTzR9aAG5HpSk8UUfhQA3mijnr2o5oAMDPSmsgJ6U/Box60wKV1p1rcAiSFG+orIvPDMDDNu5jYdjyK6Sjj0o5mLlTOHu9FvoDnZ5qjuh/pVRHnhPBeMj8K9CKg9RVa6sLa5U+bErcdcc1SkS4I5W11Vydky7vcVcHkzqSjA5qe78OxklrdypP8AC3Iqg1ldWjncpGO46VrGdtjGVM84+JvwX8O+LPNvbSMaXqjAnz4F+WRv9teh+vBr5k8deA/Efgy98jWbJhAWxHdRgmGT6N6+xwa+6IJWIxJz70ahp9lqdpJa31tDc28gw8cqBlYe4NddPE20ZMbxPzz5znFS2oHnLk96+hfi38BHEk2r+DACpy0mnM3P/bM/0P4HtXgU1jdWF81tewSW88bYeKRdrKfQiumU1KDaNYyTZ+n/AIZ/5FvTP+vOL/0AVoVneF+fDWln/pzh/wDQBWjXzz3Ok/Pr9tYf8X81Q9M2lt/6KFeJ19W/tRfDe58T/F6+1SLUYrZGt4EKshJBCDmvNrb4LRBs3GtOwHZIgP5mvUoSSgjGVRJnjqtyN3T0prEgnHIr3OH4NaEigzaheyn/AGdo5/KrafCvwtAoDW803u0x59+K29pFdSPao8CcgHKn9aTLEDnJr6Mh8B+F7dRs0m3OB1cbj+taFt4d0qBD5On20Y/2YgP6Ue1iJ1T5st9PvZwDDZXMuR/DGTWjbeFPEU6B4tJudp5yyhf519ENaqpAjRV7egpwgx2XP0rN1+yF7Vnglp4C8RzorG2SEE4xI4BH4Ctmy+F2oSBWudQgjz1CoWxXr4tj2Gc1PDY3DjKROw+lS68uxDqSZ5vYfCjTQwN1qN1J7KAo/Pmty0+GnheELutHmb1klJz+VdzBp16Qc28mQeTgirMemXXP7thz3NQ602F2ctZ+E9AtyRDpNmpB6mIE/rWxa2NvCcRQxoP9lQAK1H06VMAkDnrmozD5TcsMdqjnkyWEMaoflJx3FXowCucD3z3rNa7SLgVWk1d0OEAH+NJsDok5xhwD6VMJOxrkJNSvJCAJdvfjiqlzNcyEB5pG7/e6UuYZ3y3cCZDzouB3alfVNOiALXkQ47Nn+VectuVeSDkflWR4g1pNJst6kPcPxEp/mfYVtSpyqS5UJux0HxP+IaWdk+laO5NzKmHl6GJT6e5rxLJdvm5JqS6klurh55mLO53MT3NOCCKLLffPT2r6jC4ZUYGc6i2Q2YCMBQRnvUAJycU9gS/OSTXd/Cb4d3njPUDPMWt9It2AubgdWP8AcX39+1OtUS1exUIvYq/C74f6t431PZDutdNib/Sbxlyq/wCyvq3t2719PaTa6D4J0KPSNFt1ijUfM3V5W/vOe5qmJtP8O6TFpWjwJbW0C7URR+p9T71z0ly91KWJ3E9c/wBK+Yx2PdR8sdjvo0LayNHWNTuLt+CcE9jVeys5ppNzZOeg9vWrOmae8rZZSSTXX6Ro6hVJTJxXlpOR0tqJkaZoz5wRxmuo07SEjVQExWnZWARBkc1oRoqjgCrSsZObZVtrVY1HFWQijtSs6r1IqtPdqvCcmncixZLKOScVHJOqqSDmqAllkzuXvUkccjDGOKdguTG4bHIxmgOzDJk7dqa1qzgc1aghVUww5o0AziZFdt0xI7Zp8ISR2Gc5q8YYz1WnRwxqeFAoCxSFmA+4ZpLm0WaMoy5U1fkIC1yfw98ZWPjCzvrizwDZ30tq65zkK3yN9GXB/OqUJNXQrpMoavp9zZSHBJjPRqqQu6n7xz71391AlxE0cgyrdq5XUtKe1kyqkxnoazNYy7lWC5kHfFX4NWeP+LIqh5J447fnSBCCdy96B6HQWmtI3DkZrUt7uKb7rCuSihTeGwMmtK3cR4KnGKdiHbodAVDdRmmmNTwQKp29+DgODWhGyuoYdKdybGbfaTaXVu8M0KMrjBBFeU+MvBBsZGmtkZrfOVIGSvsa9r7c1DPDHKpSRQVPUEVJcZWPnGXSJAwJBBHP/wBar+mXF3pkiyRMduc7c8Yr1TxPoGmx20l0JYrVEGXdyFRR7k8CvCfE3xH8N6fdvZ6ekurXAbYv2f8A1bH0DHr+ANXChOeqRftYnrtrd6Z4k0eTT9Vt4biGVdskMoyD/n1r52+M3wnufCkj6vo4kutEY8k8vbE9m9R6N+fv3nhm+1swi61G3is3f5kgjyWjHoxPevQ9C12C7haxv1SaNxtYOAVYHsRThN03ZhKCmro+LXJOMjpxxTozn5SOQa9n+NvwqXRll8Q+GYTLpbktPAvJtSe4/wBj+X0rxPlX5PzA+vWvRpVFutjiqU2ixllfIAHFe+fs7fFxtNlg8KeJLomzdglndSH/AFJPRGP930Pb6dPBkYSx8EcdajUFGOK3r0OeN0RSq62Z+j0Th1BBBqQH2r5//Zk+KH9q2kfhHXbnN9AmLGaQ8zoB9wnuyjp6j6V78hz6V48ouLszq8xe1LyaQcCl9qkQfSgHijPr0paADHvQKO1BpDCjAzQD/KimAf8A6qQ0tH/6qAE5oooBpALyaBz2pM560UwDNFJ+FFACmjtRnpR0oEJRRzRQApoooNAxKM0vWkoELmilNGaAG0UZpRQAlA60UdKYAcUdO1JS96QB1ooxRQAZo9qO2elJmgANFAo60wE96KWigBopaKKQCdqayK3BAP1p+BSfSgCjPp0EhLBNp9RVSSyljOU5FbPaggHtVKTRDgmYJQ/dIrhPiZ8L9A8bQGWeP7LqKjEd5Eo3D0Df3h7H8MV6lLbxyA5GD61Rltnjc4GR2raNVpaGbhZnX6DEbfQ7CBmDGO2jQkd8KBV2oLAYsYB6Rr/KpxXC9zpWx5j4+8M2WqeJLi5naUOyopCkY4UVz48D6SDnNw2PVq9L1iCOTUZSy5JA5/Cqn2WH+50rphUsjGVO7OHPg3Sz2m4x/HS/8IZo+c+XKc/9NDXc/ZYj/ABSfZYQfu/rVqqT7I4dfBeiDrBIfrIakHhHRFyRaZz6ua7X7NF02Uv2eID7gpe1D2Rw6+FdDDECwRvrk1N/wjulr0sIRn/YrsRBEpyEA/Cn7E7oPypOox+yRw82kWiDCWsY9wgqnLpzn5Y4G/4CleihF/ugfhShB2FLnY/Zo8wm0vUGGUtJiPZDVc6Fqz5K2E30K16xtX0o2jtijmF7JHkUvhPXpAdtieR3Yf41UbwF4glwfIjTP96UV7RtFLgUuew/ZI8Ul+F+uSgHz7IEert/hTU+E2tMR5l7ZKPbe39K9tAFLweafOwVKJ5HafCPCgXGqdsERxdfzNaUHwm0REAmur2THXDhc/kK9Jx6UyZ1RC56AUKTY/ZxPLvGWheBvBvh261nUdPWVIE+VZJGYyMeFUAnGScfzr5K8Qarca5rE+oTxxRmVvliiXaka9lUegFel/tIeOh4k8S/2Lp8wfTtOcgsp4lm6MfcL0H4+teY2No88ixqPnf9B619NluFcY8z3OXETjBaC2cMSq08gykf/jzelU538yQt61pa5JEjLZW2DFCMFv7zdzR4X0O+8Q63a6VpsXmXE7YHoo7sfQAV6VWairHNQptvmZs/DDwPfeNdeWzh3Q2UWHu7nGRGnoPVj2H49BX0zO+leGdFh0XR4Vgt7ddqovX3YnuSepqvoWnab4H8Kw6Np+0yAbp5MYaWTHLH+noK527mNxOS3OeT3xXyeY45zlyR2PXoUbe8yO7llvZyTnrWzoulmVlO0+tQ6TprTFfkJ549xXoOh6YqQoCvOOteZGLerOiUlFC6RpgUhiK6K2tljAOKdbQLEgFTMwVSewrTY5277i9PaobicRr1yaqy3DM5AbA9qbtaTOc5osA2VmkIYEihInY8cirUFuDye9WUiVeAOlAiGG345GBVhUCgAcU4YozQMTGKUfSgUUAJ6UucUUUCOb+J2rnQ/AetaomN9vZSNHk4+crhf/HiK+Yf2afFZ0D4gpptxLtstXAt2yeBMOY2/Mlf+BV63+1nq/2L4eR6crfPqN2keP8AYTLn9VUfjXylbSywzpNCxSRGDI4PKkHII+hr3MBh1Og79TmqytI/RFDuUHNNmiSRCrqCDXM/CrxJ/wAJX4F0vWyV82eHE4HRZVO1x+YJ+hFdVXjVIOnJxOmL5lcwb3TSrFowSvpVeKyHQrXRlc8Gk8tc5wPyrO5Rjf2eBjap6YqSPTgOua1sKPSo554oY2kkdURRlmY4AHuapXewivDZLG24dasLhe9ea+NPjb4J8PF4Ir86tdrx5NjhwD7v90fmfpXivjL48+LNZEkGkrDotu2eYzvmx/vngfgB9a7KOArVeljNzSPp3xN4s0Dw3aG61rVLeyQDIEjfO/8Aur1b8BXifjP9o9PLe38J6QzScgXV9wo9xGDk/iR9K+fb6+vNRuGub25nuZ25aSaQuzfieaitoZbiZYYY3kldgqIoyzE9AB3NerSyynTV56mcqr6HQeLfGHinxjcg6zqlzebmAjtkG2MMegWNeM/rXs/wd+D02i2SeItfgDasV3W9qwyLZff/AKafy+tbXwA+Ey6BFF4j8SW4bWHXNvbsMizX1P8A00P/AI7065r24RIFxgVxY3FRS9nTQ6cW3zM8Q8TWPlymREISU8/7J7iucWWS0kLLu25wRXsPjbRsxvNAmUflgB90+teWX1sY5CGXvjkd68d3O+LudF4X14eUbS82yI42kNghgeMHPWvEfjr8NhoVy3iHQYydHnf95Goz9lc9v9w9vTp6V3srSQsCMc9Mdf8A9ddNoGoW+oWE2maiiTW8ybJEfkFTxg1rQq8j12FOCmj5FtmMcuR07jNW5UyA45B6V1Xxc8FyeEtdxbhm024y9s57DupPqK5bS3SQtBKeD9wnsfSvew9RfC9mePiIOHvIk0y+uNN1CC+sZnt7mCRZIpFOCjA5BFfbXwX8fW3jrwrHebkTUrfEV9CONr4+8B/dbqPxHaviS6g2NtA6V1Xwi8Z3Xgjxlb6kjObOT91eRD/lpEev4jqP/r1z47C21R0YespxPu9TkUtUtIvrbUdPt7+0mWa3uI1kikU8MpGQR+FXRXjs6BRS0me1KOlABSUtJn8KBhnigUdaWgBOtFBooAKXijmjFAgo9qQfWg4oGHXtS9KTvS59aBCc0tHFFAxMcUYFLSd6BCYpRRRQAUmMnNOzmkoGIOTml70YpBQIKUAc0maKBhRS0mcCgQZooHNFMYUUUUhABRjFFFAAaTNLRmmMSigc0CkIMUCjtRzQAlFHX2pM0AL9aOKTOetKKADpTSgbrTutKveh7AjoLMYtYv8AcH8qlqO1/wCPaL/cH8qk71iWYOqf8f0n4fyqvVnVB/p0mPb+VVutax2JYooFFHtTEHfNFJ9aXigBKX0FFJQAYooHHU0GgBaO9HTrQKBiYopRR+FMQYoxRQaQCNnFeU/tFeN28KeD3tbObZqWo5hgwfmRf43/AABwPcivULqZYYHkdgqopJJPAA7mvh34v+LZPGPji91AOTaI3kWi54Eang/8COT+Nejl2H9rUu9kTOXLG5y0KNLKXKkjvW40P9laL9pkP7+6GIxnlV9aXwrYrd38MBxsHzSHtgVV8Yait/rMhhwLeL93EAMAAV9dpTgeRrWqeRkOCzDAJJPQdc19OfBbwfF4M8LnVtRjA1i/QM+4fNDH1CfXuff6V5x+zx4Nj1vWZPEGoJusNNYGNSOJJuo+oHX8q9j8TajvdlDZUDGK+ezPF8q5Vuerh6V3cxNev5J7hgDnJ/yaXSLJmkUhSc9ar2Nq08hdgQxNdv4Y00E5K/Wvm0nJ3O5uyNHw9pqrGgx07muus7ZY1Bx2qtYWixKAo6Vos2yM5Nb7aHO3dkLzbDg1UmleQkZOKZOpZ8hz9KkgQbhmgVxscOQGHJq3CrD+Ae9PiiGATVhQAOlK4hoXAwBS5Io69qKAFpOaOtAz9aBhnNLSUZoAPxprHAzTqjmOEP0ppXYj5h/bA1hp/FGkaIrfJaWjXDjP8cjYH6J+teFHsRjPpXcfHXVH1f4p67cMV2w3H2VMHPyxgJ/ME/jXCjHrxX2GDp8lGKOObuz6C/ZB8U+Vf6j4RuZcLP8A6ZaA/wB8ACRR9RtP/ATX0wR8ua/Pvwlrt34a8Safrlif31lMsoH98dGU/UEj8a+r/EXx08D6XpcNxDeyalczRLItrajcy5GcOx+VT7E59q8nMcFKVTmgtzalNJWZ6kzACub8XeNfDvhe1M+tarb2gxlYy2ZG/wB1ByfwFfMnjf46+L9dV7fTGTRLRj0tzumI95D0/wCAgV5dd3M91K01zPJPM5y0kjlmP1JqaGUt6zY5Vl0PoLxr+0Zhmt/CWlFscfar4YH1WMHP5n8K8c8VeOvFfict/bWt3VxExz5AbZEP+ALgVzJ5HBxQOO9etSwdKktEZObY89eAMUi5JIxSqcZxT7C0ub29itbSGS4nmcJHFGCWdj0AHc1u7RIH2sE1zPHb28TyzSsESNBlnY9AB3NfVXwI+ECeGIY/EGvxLJrUi5iiPK2YPYer+p7dB3NL8BfhNF4Vt49d12NJdclXKp95bRT/AAjsX9W/Aep9nGAuOOK8HHY+75IG9OnfVkKKEXHSnggnrVLVJmhXch/CoLWdmAbP614711NzQnhWVdrdK4zxb4UFwr3Fko3jrH0B9x7120ThlBFQXnmEbYx170b6DTa1PA9asJopWjlRkI6jpWGGktLgSx568rntXtfijRI9QiYsNko/iA615frukXFhIVlQbf4eOKjY6YtSQ7WLC08YeD5dKvGCzj5reQ/8s5McH6dj7V8xa1YXelalLaXSGK4gkKMp6qRX0RYSyQyAqxGTmuT+NmgRahZJ4itUzNGBHdKo6r/C/wCHT8vSvRwlTm91nPXpqSueaRyfbbJZuBIOH+tUJIwrnH1+lJo0/kXphkOI5flPsexq/fwEEjGMV7tN+2pWe6PDTdGrbofQf7I3jpri2n8GajckyQ5nsN39zPzxj6E7h7E+lfRq4I4r88PC2uXfhnxFZaxYMBcWkwkQHo2Oqn2IyD9a++PB2vWPiTw7Y61p8m+3u4hInPK+qn3ByD9K+dxNLkmetCXMrmxzSj6UmaPfNcxohetHNJg+tFABzS9KTHNLQAuaM0UUCEo60lAHagBT2zQB7UYPrSZoHYKXNGaKAFo70gooAKB9KOppBn3oAU0H6dKPpSc0wD/9dFKKKQMSloJ7migBKBRSE455P0oCw7t0o570vpSDrQAlAFGKBQAUZpe1J+FAC/hSUtIPrxQAHP4Un4UvWkxQIKDxRRQAEYoFHGM0UAJzRRRzTASijHvS80gCgdKOgoFDBHQ2nNrFn+4P5VLUVp/x6xf7g/lUtYlmFqn/AB/P+H8qrD6Va1T/AI/n+g/lVWtY7EsDRS0maYhfrRxR0opgJRS0lIA4oB9qUYoFAwH6UvOaOnSigQlJmilz7UAIKDRTJX2xk5pgeRftQeK20LwMdNtZSl5qjGAENgrEPvn+S/8AAq+TLRWabcFzj8q7v4+eKj4o+IF28T7rOyJtbbB4IU/M34tn8AK5zw5pj3k8MSZzIw6c8V9VltDkgrnDi6llY3rry9F8ItOPlur0bEOOQveuN0qxuNS1K3sbSMyz3EixxoOSWJwK3/iLerNrP2KHHkWaCFQDxkdT+dd5+zT4cWTUbnxVeJmK0zDa5HHmEfM34A4/GunF11CLbJw1LQ9V03T7Twf4TtNBsyoaGPMzgYMkh5ZvxNc1IGnuNxJOW7Vs+JpJ5JVmbISXJQ+oqvpFqWcHrnp3r4mtVdWbbPahFRjZGvoWnB2X5feu+0a1EceMc1keHrMKMkV1dpGu0dquKsjCcruxPEoSOqM0rmQr/D61dnkCJgEZqiDuc5pIljRGQd3NW7aLd8zYpIgHAUCraKFUD0obEAGBS+lLyaTFCGLmk69aXPrSUAGaOMUd+KKAuA6UdelJuAHJqC8vbe0gee5mjhhQZeSRgqqPUk8CnGLlshNpbk9ZHi3U00fw3qWqPgi0tZJ8HvtUkD8xXBeNPjj4N0PzILK5fWbtR9yzwYwfeQ/L+Wa8G+IPxg8T+LLWfTnMGn6dMCskEAyZFz0ZzyfwwK9HC5fVqSTasjGdaKWh59fTyXN3NdTtmWZ2lc+rMcn9TVXcRkY61I5ycntxTduOvevqYx5VY5VK7G5IoLNjnFKdvApHwSAPzpsoAx9MUAkjPQUjdhRwBUjFYn/Pemkgkkgk045NWdL0+81TUILHTraS6up2CRQxrlnY9gKmTSV2BHZ2895dRWtrBJPcTOEijjXczsegA7mvq/4C/CWLwpbLreuRJNrkyfKOq2ikcqvqx7t+A75sfAr4TWng2zTVtWWK516ZfmYcraqf4EPr6t36Djr64AAABgV4GOx/N7kDanT6sZGoVQAMAU/tSjpSd+K8c3KOqRF0yvas+EeUuK3JFyuDg1mXUDBsrQBNbzMGA7VfHK8jrWUvGCa0bZwyDnmkNIqahblssBXM+IdOivbZ45EwexrtXAIxWJqcIYEdCKN9Bp8rPGdb0gWkhZIg+w5K9iKy4TBdyT2UkREEoKtGeflPUV6B4ltSHZvXg1xV3ZKJ90XyMDnPSppzdOVzpspI+evHGiyeH/EdxYsDtVt0TEfeQ8g1dX/iYaRHeL99Plk9iK9I+M3h+TUNDg1eOE+daja59U/+sf515p4DkH9pS6bM3yXKYUH+9X02EqaqS6nh4+jpddDJvY9p3AA9ulfQP7IfjHy5bzwdeT435u7IN6/8tEH4YbHs1eK65ZSW8rxyJtKkg1F4O1mfw14o07WrUkPZ3CylQfvqPvL+Kkj8ajMsP1KwVZSjY/QdWyAacDWfod9BqWlW1/ayiS3uIllicfxKwyD+RrQr59nfsGc0opBil+lIYlKKTNLQAtA5NJS96ACj0pOtA5oBDutJR2paAE7fpS0CimA2jr1oo696Qhe/PFH4UCloKE+tFAooEFFFFACUd6WjANADaMijilGKAAdKBS8UUAJ9eKMUtHWgBOD1oo7UfjQAuaSlFHB6mgBtBzS0UCE60UUd6BgaTtS0CgQhoJGeaXtQcUDE4Boo70UAFB4oo+tD2EjobP8A49Isf3B/Kpahsv8Aj0iP+wP5VNWJZh6p/wAfz/h/KqtWtV/4/n+g/lVWtY7EsKKBQKoQ6jFH0NFIBOpFFFLz1oDcQCjA9KP60ZoGBFFLiimIaKOtL2o5oAPwrkfi3r7+G/h/rGrQsq3ENuRAScYkYhV+vJBx7V13avnj9sLXQunaT4cjY7p5WupRn+FPlUfiSf8AvmunB0+eqkJ6I+dcvczM7ktI5LMx6knqa9B+HtuLPTtQ1iZgEtoSV3cZOO1cDp4DSgHPpXe+Kpk0v4fW9gpxJdyAkf7I5/wr7FLlhoeTP36tmcFO0l5eHALyyv8AKAOSxNfU3hnSYPCvgey0leJFi3zH1kblv1OK8N+BOiLrPj+1kmXMNgjXTg85K/d/Uj8q908SXKvIUzXgZtXtHlR62FhrczLqZ7+cZGEjG1Rmt/w7ZfvB8hwaxNIiDS4IyDyOa7vw/agAEV8/TVzqm7I3NLtUijWtZcCPIqvZRHaPSpp2KrgCtWcxXkYSP8xPBqSJFZ/wqCEJvIbqemavW6AUtgJo0CqKeaBRipGBoFHWjp0pgApPel/GuA+KXxItfBMW1tJ1C9nZcqUjKwj/AHpCMZ9hk1pTg6j5URUmqauzvSQK5jxZ498M+FwP7Y1SCCQ9Igd8h/4CuT+lfNXi740eLtfLw290ukWrrjyrMkN+Ln5vyxXnVzM0rs7uzuTlmY5JPuT1r28NkzlrNnkV80s7QR7r4x/aHlcNB4W0nb2Fzfc/lGp/mfwrxnxR4x8T+JJT/bes3d2mSRCz7Yh9EHy/pWKTxTHOcYGOOa9ilgaVLZHHLFTqPVjXYsACKYTzzTjwMceopmT+OetdXKkbU6mgHjuKTHUE5FG0YJz701ugNI6YsUYPXFJ3/Cgg0hypBIqWbJiE5oyABzTh05Gat6HpWoa3qsGl6XayXV3cNtjiQZJP9AO56Cs5SSV2UN0nT73VdSg0/T7aW6urhwkUUYyzH/D3r68+B3wptPBOnf2hqAjudduExLKOVgU/8s0P827/AEp/wP8AhbZ+B9NF3fCK51y4X/SJwMiIf884z6ep7n2xXqQwowBxXz2Px7m+SGx0U6fVkSKFAAXgU8AfjQBS47V5BsGOaUYpO9LSAQruGOlQPCMcsTVioLhXIytAzOuFPm8HAFSwyhSFBGcZpr53EHrTcAfN3HemI0oyGUc1HeW6yr0waitZCT1q4vzClsx3OO16xzuAFee67aGOQleTnPvXtF9axSqdyjpXnfiqwXzH8g7ufTiplqa05W0OWt7ePVNNuNOuk3xzIY2FfNmv2Fx4Z8Wy25JWS0n+U4xkA8H8q+m7Ai1vwHGM9K8q/aT0dYtRs9aiHyTr5UhA7jkGvUy6rf3GZYqHMjn/ABtAlwtvqcAPl3UYfJ9cc1wtyu1siu/0B/7X+H0kJJaWyY8Y6LXE30fLKwr6CtH2lI8PDv2VVxPrX9lPXm1X4ZR2E8habSrh7bk8+WfnT8MMR/wGvYRjFfH/AOyR4j/sv4gXGhzSbYdVtyqA9POjyy/mu8flX1/Gdyivk60eWbR7ad1cUD3pe9IKXFYgIf6UUjHAryvx38cfCHhyeWzsmm1q9iJVktSBGrDqDIePyzWtOlOo7RQnJLc9ULDpmjcPWvk/xL+0V4vu5MaPp2m6XEehdTPJ+Zwv/jtcjcfGL4kXEm9vFNzH7QwxIP0Wu2OWVXuZ+1R9vbh604EGviWx+L3xHt3Ei+KLqXnOJo45F/IrX0N+z78RrrxxpN5bawIV1axZTIYl2rLG2drgdjkEH8PWs6+BnSjzMcaqbseqUUg6Uv4VwmgUE8UEUfhQAhPPIpw5qMuAetKpB5xTAeKXikFL3pDEGf1o9qM8UUAJRmj8KKACjijt1pNy9M0BcXijigc0UAFLik+tLQAneiiigAox60UdqADAopNwz1yfalBp2EGB3pBR9aPrSAQnj0rgvG/xb8GeENU/szU7u4mvFAMkVrF5hiB6bjkAHHbOfatf4m+KIPB/gzUNbmKl4Y8QIx/1krcIv59fYGvlb4UeA9X+Jfiae/1CeZNOE5kv70/eldjkon+0c9f4R+Aruw+HjKLnPYzlOzsj7C0DVbDXNHtdW0ybzrO7jEkL7SMqfUHkH2q971U0fT7TS9MttOsIFt7W2iWKKNeiqBgCrlccrX0NENo+tKD70nvSGAowaXpRSEJR7UHBpMYoA6Gz/wCPSH/cH8qmqGy5tIv9wVMKxLMPVf8Aj+f6D+VVfrVrVf8Aj+f6Cq3atI7EsMGijkUCrEJ3o/lQKUUAIMUuaKKQwxRR9KUUxBRRRQAfhR3oNJQBHOxEZI9K+LP2hNZk1f4paoBJvis2W0jAOQoUfN/48Wr7Ov5UhtZJZW2oilmJOMADJr8/dXuW1HV7y/kbL3U8kzHPdmJ/rXrZTTvNyM60uWJZ8OQmbVbaPHWRQeOvNb/xdnU61b2MeAttCBx6nmrXwr04XOtxysoKx84x+tcz4wuPtPirUZCcjz2Az6A4FfS1NI2PKo+9O57J+zlpaWXhfVNekUiS5lFvGSP4EGTj8T+lb1+3n3TcjrkGrPhG3OkfDLRrIgh2t/OcEYwXJb+oqkgy7NkcmvjMxqc9Vn0FCFomvocG6bGOODXomh24WBQPSuM8MWh83cxLMT1zXommQhYQB3rkprQio9bFu3xGvLVDcsXf5fWp3iVUJJpluit83WrZmEMORluoq1GuwcUiYFSZpXuAdaB0ozxmjNIYHrQKUUgFNAAAqveWkN1C8M8SSxuMMjruUj3Bqx17UoIpqTi7olxUtGeV+L/gl4N1wPLb2baTcnpLZHaufdD8v5AV5D4t+BPizSA8ulPDrMA7Rfu5cf7rHB/AmvrHqKaVVuorvoZnWpeZx1cBSqHwBrOlalpN59l1KxuLOdf4Joyh/XrWe5PevvzXNC0fW7RrXVtNtryE/wAMsYbH0z0P0ryLxn+z7oF6Xn8PX0+lynJEUn72L9fmH5mvZoZxTnpPQ86pl9SHw6ny6+QQQM5600n2xXSePvCWo+DNcGj6nLbSTGIShoHLDaSQM5AIPB4rncdx1r14yU1dGMbp2Y3bk0n06dKcPrTWOM8ZzQzrg7hhgcmm4x3zS556GkABJzmoZ0RQjYOM19Sfsk6PoX/CHTa3bW4OrPcPbXUznLKBgqq/3VIIOO5/CvlzOD0r339j3XfI1nWfDskny3ES3cKn+8h2v+jL+VedmKk6LsbwSvqfTSqAOlKKRSCMinCvlDqD9aUmj2o59KQwI5oGKTvSigQmKMZ7UoFJTAqXUBLbl71RlVlBBzzW0RkYNVbmIMp6YpBYoQvtYYOec/StO3YsBWcIhnGMVctSRgZzQBNMDsOBmufvrJbiJg8QU5OMHmuiPIrI1IspkjX+MZU5xg0dBrc898QaXBZTq0cjDdk7Tzg1wXxcs21TwTdIyt5kC+cgxn7v/wBbNenasjTyGC6/eBPmBU8k46E1kXOkx3enSwq/nW8ykKx6rxgqa6MFLkqpmk9Y2Pnj4OSifUL7SpOEu7c8Y/iHNYHiS2+z38sZ4KsR+ta3hPd4f+IyQSHZ5V00DZHbJFTfEm3WHxBcqB8rNuHHrX1tP3oNHz2IXJUTRzHhfWLjw74r03W7bh7K6jm+oB+YfiMj8a/QnTpo7i0iniYNHIodCDnKkZBr86Z15AIHFfc/wM1E6p8KPDd00rSv9iWJmbqShKH/ANBr5vH0+WVz1qMuaJ2/FBpe1JjivONjzD9pPxRceHPhxOljK0V3qMq2cbqcMqkEuR77QR+NfHgyegGK+vP2hfAGueO9P0qHRJrNGs53kkW4kKhgygDBAPPB/OuZ+Enwn8R+FtUjn1fRPCmoqZAWuJppJJo1/wBhduzP4fjXs4SvTo0r9TnnFyZ4b4d+H/jDxJEsukeHr+4hbpMU8uI/8DfAP4V1+l/s9eObkKb2bStOU9Q8zSMPwUY/Wvr1VA+UDAHAFI6r6VEs0qN6IaopHyZ4r+CLeFPC15ruq+LbYJbp8saWZ/eyHhUBL9Sfbjr2rS/ZE0/UZPG2o6lEpFhDZGG4f+EuzKUX6/KT/wDrrI+OHi68+IPxAtvC+gEz2Ftc/Z7VUPFxOTtaT/dHIB9AT3r6a+HPhOw8GeErTRLJQxjXdcTYwZpT95z9e3oABV4ivJUrT3ZEYrm0OhWl71y/ijx/4P8ADM5tta1+ztbgDPkli8gHuqgkfjV3wx4n0TxLYfb9D1KC+t920vG2dp9CDyD7GvK9lO17HRzLY3Riob26t7O0lurqZIIIYy8kjnCooGSSfQCvNfjl8Tl8BaNDFYxxz6xe5FsknKxqOsjDuBwAO5+hr5p1z4j+Ndetbi11TxHeT2tyNssClUjZf7u1QOPauvD4GdVc3QmVVI9Vh+MfjHxb8ULTR/BtraRaS04GJ4NzvACN8rnPyDGcAeoHJr6LgI25NfG/7PGma3qfxMtRpV3c2lpb4uNReJyqvEpyEb13NgY+p7V9ZeI9f0jw5pUmpa1fQ2VrH96SQ4yfQDqT7DJqsZRUZKECYTvqzaz2pTXh5/aN8G/bTELHWTbA4Nx5KAY/vbd27H4Z9q9ls7qO5toriFw8UiB0YdGUjIP5VyVKE6eskaKSZbAJ6UEcVDNe2lrCZrq4igjHV5HCqPqTUizQzQrLDKskbDKupyGHqDWfLLsVdBkd+tKnJryfxF8cPCGl+LYdBiee9zMIbi7gwYYCTjrn5sHrtr1ATALkmrlRnG10JSRj/ErxAfCvgjVNeSETSWsOYoz0Z2IVc+2SM+1eAfALxv43134pLbX2rXWo2VzFLJdxSHMcIC5VlHRPm2jA65rvPHvxt+HtveXnhjUYLrV4GBgvDBErwjsVyWG4j/ZzyODmu/8AAng3w14VtHHh7TVtVucPI7MzyP6AsxJwM9K6FalTaktWRu9Do1+6KXNct8TPHGheA9F+36xcHzJMrb20fMs7Dso/mTwK+a9W/aA8c3urfbLCa00yzRspaiFZAR6OzDJPrjbUUsJOqrobqJH171o+tZfhLUp9Y8L6Vq11bfZZ7yzinkhH8DOoYjn61g/E34g6H4E0xbvVJWkuJci3tIsGWY+wPQDux4HueKxjSlKXKinJJXOxLDpSj3BFfG3jL42+NvEdy0VpeHRbNjhILJsSEf7Un3ifpge1dt+ynrHiTUPF2oQza3Lc6alr5k9vdXDSOXJwroDnGDnJ9x7V1zwMoQ5mzNVU3Y+k+vaq99dQ2ltJcTyLFFEpd3Y4CqBkknsAKs9R7V87/tT+P1WFvBGk3GZHAbU5EP3F6iHPqeC3tgdzWGHourNRLnKyOR+I3xM8UeP/ABcmgeD7q9g06SXyLSG2cxyXR/vuwwQvcDoByea+nPBOmXOi+E9L0m8vXvbm0tUimndiTI4HJyecZ6Z7Yr5r+FV14f8Ahj4WHjXxIvm61qsZGlWCY877P/f5+4HP8R/hAxnJFaPgf48eK9a+Iem6dc2enR6ZfXKweRGhLoG4BDk8kHHbB9BXbiKDkrQWiMYT11PpxRmhgRUccyLD5kjBVVdzFjgD618/fGD49bJZ9E8DSozLlJtUKhlB7iIHg/7549AetcVHDTqysjWU1FHqfxRtfBmo6Klj4zuLGK0MgkjFxdeSdw4ypyD0JHHrWl4Ks9Cs9BtYPDsdoumIn7j7MwaMj1BHU56n1r4ivU1zVra48RX32y9gEqxy307Fgzt0UM33jweB0A7Cvpn9lC3tovhxLLBfm5eW9dpoegt2AA2Ae4w2e+6u3EYd0qW5lCakz2EdKUHmuO+K3jqx8B+GH1W6Tz7h28q1tg2DNIRnGewAGSfT6ivNvgp8U/FXijWdVvPEcmm2+gWNs008yxeWtu2flG7Jzxu656fSuOGGnKPMjVzSdj3l3CjJOKbFIkih0YMD0INfJvxi+MOq+KJptL8PzT6doikruU7JroerEcqp/uj8fQem/skz6pJ8P7pLwytaR3zLZs/TbtUsF9g2fxzV1MJKnT52JVLux7Rn1NGOOaB70Vxmgh57UYoNFAHQ2P8Ax5w/7gqaobH/AI84f9wVNWJZiatxfN9BVUVZ1b/j+b6CqorVbEsU0h6UtJTQhaKX8KBQAYFAoo460AHFFIKKAAmiijNA2JgmilopiOJ+Nmrtovw01u9j+/8AZWiX2MmEB/8AHq+H4gGOBX13+1XcJD8J7mJj81xdQRL7nduP6Ka+RLcfvF569a+gyiNotnLitj1/4T28UGlX+oyLzFCxyfQCvKdPj/tHxDbwDJNzdKmO/wAzAf1r1/w4Vs/hbqtwACfszDp6jArz34NWKXvxK0dHPypOZSOv3VLD9QK9OtNpNnPhYI+ivEbIka28eAkaBQPQAYArGsV8x1Abb+mau+IZN0jL7modGQ+YOOM9+a+IrS5ptnurRHdeHYOEyABj8666BCIxs7VzuiAeWpAxgcjFdRZEFeetaLRHLJ3YyQskRMh4pLdgTlM81Yn2su1hkGkgRFIwMUkxNEyA7Rk80pIAyTQWA5NUWdruby0yIEOGI/iPp9KLDLqMrjcpyOxp9MRQoAAwBTsc0ALzRR1ooAQ59KOO/wCFHU80cUgEpnmx+b5W8bwM7akxgUYGc4GaYhp61V1W7gsbCe8uZFihhjaSR26KoGST9AKue9eKftU+Ll0vwinhy2kxdaq2JADysCkFv++jhfpurowtF1qqijHEVFTg5Hzv8RvED+KfGWpa7IGCXE37hT1WJeEH/fIGffNc22Mg9/apHfAzgHio88bjwPftX3EIqnFJdD51TcncXngYx70hOOc17L8GPgzdeIWg1zxRDLa6Tw8NqcrLdDsW7qn6n2HNegfEz4FaPrQN94baLR74KAYdv+jy49QPuH3HHqO9efUzClGfJc9KjSk43sfLfXrj1prYx0ra8VeGtY8Nam1hrFhLaSjOCw+WQeqt0YfSsYgfSuqM1JXRuhDkjkcV1Pwi1z/hHfiRomqs+yFbkRTn/pnJ8jZ+gbP4VypJ7dM0Jndx+BqKsOeLRa0P0VtzuiXnJHFSCuU+E+uf8JF8P9F1Zn3STWiCU/8ATRfkf/x5TXVivi6sXCbR2J3QZpQeaTrRxnpUDA80oo4pO9Ahe9GaSjPrQAU2RcqadikPFAFGX5SaktW55qO4I3U63OCMD9KBE8uQhOe1Yd5IMnPNbkw/dnHpWDfgjOAM0xrc56/lyksKSiF2bIcjtWTZMILgW8Vy8gky0iNjj/69WteVhlgADkZz0rmkn8m/WQHBz68damErSOhLQ8W+L9uNM+Jc9xFgLIY7kcdz1/UVb+KUAkktL1eVmhBBNS/tDwka7p13182Bkzj+62f/AGarvjiEXHw/0S9yWxEqsfqK+ywcrxR4OOjZnkt9knpnHH1r6t/Y91U3nwzn09pNzWGoSIFJ+6rhXH6lq+U7zaQ2eo6V9A/sS3DE+KLUt8gNtKoz0P7wf4V5WZxR0YSV0fTIooHQUV4h3AQKTb7Uvag0AJnAryT9pbx3/wAIz4UGiafPs1bVkZNythoYOjv7E/dH4ntXqOr39rpmm3Oo3syw21tE0s0jdFVRkn8hXxTq2oal8UviiZUDibU7oQ2yHnyIRwo/4CoJPvmu7BUFOXPLZGNWVtEesfsneBo2S48a38PILW2nBhwO0kg/9BB/3q0f2hPjBc+Hrl/Cnhi4CaiFH2y7GCbcEZCL/tkck9s+vT1xIbLwf4LMVnFts9JsWKL6iNCefc45+tfHnw58N3HxG+Iiw6hM7LcO97qEoPzFc5YA+pJA9s+1dNK1acqk9kZv3VZGP/ZWt6xpGoeJ5lkazhO64vrljiSQkDaCeXckj+uK9M/ZJ8Qy2viy88NC2VodQiNx5o+8jRjofVSCfx+ta37VskGi+EPDvhnTYI7WykndxFGMKFiUBR+b5/Cud/Zp1DQfC0Ou+NPEF7FbRQotnbgnLuT87hF6seEHHv0rqnNVaLsjJJqR6j8ePCfgb7JdeNfFaXs1xDAtvBDHdmNZm52IoHQkk5I7ZPavJP2d/h1ZeN9Vv7rWoZW0mzUIFjkKb5W5A3DnAXk/UVmfHTxlrfirW7WXUIHsLAQCewsWPzJG3SRx/fYDPsuPXn6U+BOhQeGfhfpUbhY5JoBeXLnj55BuOT7DA/CsG5UKG+rNUlKRc0nw94N+FXhjVdRs4HtLNV8+6kklMkj7RhVBP1wB6mvlXxj4g8T/ABV8axxwW81xJK5Sw0+I5WFP5ZxyzH9BXpf7R3xD07xN4bt9J0C5eazXUnjuZQMJKY0VhtPdcvnPtXX/ALKvhKz0vwSfE80StqGqOwWQjJjgVioUemSCT68elTS/c0/az1bCWr5UfOPj7wqfB2rpolzfx3l+lusl6sS/u4XbkRgnliFwScDrX1F8CNcvb74R6dqWtXALQrKhnfC/uo2IUn6KvX2r5p+K0k83xM8Stdr++/tKYHPYBsL+G0Cu38bePNMtfhVaeA/B7y3iRWkcep3yRsscasfmUEgEl3JGemMjnPHXWpurTinuzOL5Wc94+8V618U/HsWn6aJZLJpvJ0y0JwoHeVx6kAsSeg47Vv8Aj/4gyaV4YtPhv4NvJHsrCD7PeX8R+a5YZLqnomd2T3HHQc+ceHtbk0bTb6z0y1k/tbUdtst0py0cB+9HGBzuc4BPoMDrXWeL9Dh+H3hK20qco/ifWIfMvSOfsdtniFf9piMMfRSOnW/Ywi0mhczZg/C7wbqPjnxXb6ZZRsltEyy3txj5YYgef+BHoB6+wNe+ftM+OpvCvhmPRNLmMWpaoGQSKfmhhHDMPQnO0fie1S+AdW8I/Cf4P6dfX1xGb7UoFvXijIM1zI65Cgf3VBAyeBg+teA/GfVfEOveK/7S8RWqWdzLaRyW9sv/ACxgbLIpzznByfc9ulY61qt2tEabRLvwD8By+MvGEMtzCf7H051mvH7SMOViHqSRz/s59RX1r4/8Z6b4I8KXOu6gDII8JDAhAaaQ/dQf1PYAmuY8HXHg3wL8N9Pmi1KytdLFus32lpBmd2GWbjlmJ7DJ7dq+e/jh471Dx7Nb31pZT23hmzuWgtHkGDNNtyzN77cYHYHnk1hUg69XVaIqL5UVZ4fGPxk8eT3UMPnXEmNzEkW9jDn5Vz2A9OrHJrq/CnwG8Qp8QY7TV4kfw7bSrLNebgBcoMHYq53Ak8HPQZ56V7F+zFa6Onwm0yXT1j82cu164+804Yhg30AAHtj1roPiZ8QPC/gSwMmrXYa7ZS0NlDhppfw7D/aOB/KoniZqXs4Iagt2R/FbxxY+BPCraiY0mupD5NlbdA74746KByfy718u+HvDvjH4v+L7m9eZpXLA3d/MD5NuvZFA/RB+OOtSfFjVvGPiJNM8TeJrVbKwvvN/su1XP7qMbeT3y2Qcnk4zgDFeqeGviN4S8A/BjSYdLltr3WZrYyGzjbLeexO95sfdAPryQABWsKbowvFXkyXLmZyXxNsPC/wm0IaD4fUXnim/hIuNRmwZbaFhhio6RluQAOcZJJ4rQ/ZJ8NX6aje+L7gNDZtAbO2B484lgXYf7I2gZ9c+leW+P7LWfOtda8Q3DvqetRtesjjDrGThGPpuwcDsoHrXq+nfGrQdB+H+kaT4f024vNZjtI7ZbQxFY0lAAJJ6tlucLknPatKsJOlZa3JjZO56j8ZvHs3hXRYNN0SJrrxJqpMOn26LvZc8GQr3x2Hc+wNfKvjXT30rVxo0shv9VV92oyhi5kuXOTGD/Ftzgnuxb2r2qBL3wJ4X1H4m+N5VvfGF9GIbOOTGLYsDsiQdBgZLY6AEepPj3w18S6boPjWLxLr+nT6uYmaVVVwD5xOfMORyRkke+D2pYWn7OLcUOpK563bfDXS9A8Far44+JYGq6y9m0hgmb91bHbtjiUDgsDtUdl6AcZPEfsueDp9Z8Yr4lu4z/Z+j8oxHElwR8oH+6CWP/AfWrnxK+JNv8S9ZstFmvP8AhHPDcT+bPJOC7uw7kIDkjoq9MnJPpuN43bULK1+GvwY0+eGMpsl1OVSpjQ/fkHcE9S7YPZRnFJqajaW7JVrj/wBp34iXAuG8D6PcGOJUDanJG3LE8iHPpjBb1yB61L8GfgnY/wBmxeJ/HigRFPOi0+RtiRpjO+Y/Tnb0A656V5j4+8M/8IH8R4LS7aa/tozb3fmTH5rleDITn1ZXH5V3njD4ha18WtZg8DeDraey0y4YG6mm4eRAeWfBO2MemSWOB7Vbi401GHXdhe71OY+MPi1fGvim00bw3bBdEsW+zaZbQR7BK7EAuFHAzwFHp9TX0p8G/BUfgTwNDp88iG9lY3N9Jn5RIQMgH0UADPsT3r5UnubTwf8AE17jSIftEGjalthWZuZfKbaSSO7EE+2favUo/FnjH426h/wjWl2g0Lw8hB1SeNy7Mn9wvgAk9kAGepOBWeJptwil8JVNpM4P4++J38b/ABH+zWNwjabZP9js2ZwI2Yn55STwAW7/AN1Qagu7iTXH0/4Z+BAX01JPMu7o5UXsw5e4kPaJewPYDvisf4jaINB8batoMETEQXZjt4x8xZGwYx7naVr1PV/DP/CrPgZqVywH/CQ6yI7W5mHWESHmJT6Ku7Pq3PYVt7sIRSFu7nBeAPBNp4z+I/8AYekzzvo1u264u24ZokwGcem9vujsGHXBr6wsdR8M6FqmneC7Ga3tro25a2sYhkpGgzk46Z5PPJ5PPNfIHwy8ea54Pi1Cy8PWsEl5qgjhjkMZeRGBIGxejE7uhzzjrX0d8E/hzceH/O8UeJpXu/E2oAmV5H3mBTjK57seMn2wOOvHjrv4noaUkepDpS980g+lLXknQAzSDrQDxnFGOKTBHQWP/HnD/uCpqhsP+PKH/cFTViWYerf8f7fQfyqr361a1X/j+Y+w/lVXitY7EsXtQDRSiqEIM0celFL0pAAFGOaO9L1oGMNApf60mcUxCk0ZpPpSj1NAB9KRjTu1NbpSA8P/AGvrop4G021A/wBfqSk/8Bjc/wBa+XrVQ0oUnGT1r6T/AGxnUeHNDU9TfPj/AL9//Xr5ssmzcIB619RlS/dI4cUex6iqW/wVvmBA3qijHfLCud/Z2tfN8dCfHEFrK/uDjH9a6LxMAPgdK2c5kj/Dmsr9mUKfFGovk7lsWx+LLW2KlaEgwy2PStfO+5OegbtVnw8ii5TtVTW2xdnPIzirmgD94GA6V8ZL4j2L6HouhglRkc+tdJaqAo6VzGgzLJGCOtdJaykkCtuhzPclnGcU6PG0HFPwDSgDHSkwILoMy7V4zxmi1gSCNY0GFA4qfANAoEAwD0oB/nRiloGHNHalpBxQAUgpe9HWmAlBPFGKTtmkIiupVhhZ3YKqjJYngDHXNfDfxd8Tnxd49v8AVVY/ZQ3k2gPQRLwp/E5b8a+lP2k/FX9geA5rKCTbeapm2jAPKpj9435cf8CFfJekaXfa3qkOm6Xay3V1M2I441yT/gPU9BXv5VSUE6kjxMyrOc1TiVQryukUSF3YhVVQSWPYAdzX0d8DvgqLUQeIvF9uHuRh7bT3GVi9GkHQt/s9B354HSfBL4RWnhWGPVdYWK71th977yW3+yme/q35cdfYY0VFCgcUsfmjfuUzTB4Gy5pkUMKoMkc1Deq3lnaMn0q52ppUHrXhKTbuz1uWysjjvEWiaR4i019N1uwiuoT/AAuPmQ+qkcqfcV86fFH4L6roAk1Lw752q6YOWiC5uIR7gffHuOfUd6+sbqySVemD61kTxzQMUYcetehhsZOm9GZyppnwS3BKkYI4Oe1IBjmvrH4kfCjQPGJe+tCumaxjP2hF+SU/9NFHU+45+tfN3jPwhr3hHUPsmtWTw5JEcy/NFLj+43f6dfUV79DFwrLzMJRcT339j3XRN4c1Xw/LJ89lcrcRKT/yzlHP/jyE/wDAq9+HIr40/Zk1tdJ+KVpbyOEh1OF7Rs92+8n/AI8oH419lx8oK8DM6XJWv3OijK8Q68mgUgpa840A0ZpRRQAhzRQKOnWmAYNMmBC1IPamswHUikBSeMl80INpqSaQdqgzls4piRaYkx4rJ1GMYNacbhhiqOo/dPFCH1OK8QqQGI57VxV0rmY7JOc9CecV3uvDMbc9a4S6XN0QvGT69qyWkjqi9Dz/APaChD6bpVxn5kmZfzX/AOtSa0yyfB/THB+6q++Ks/HkqfCdo3cXS4Of9lqoj978GLY4PygHr/tV9hl7vTR4eYLVnlV5jdnI5r279i6Tb4u8QQgnD2Ebe3En/wBlXiV2PnOVr2X9jRwPHutIeraZn8pU/wAa5MzWjKwmx9ZL0pw601R8opwFfPnoIWmt0paRumaAPC/2ufE8lh4Ws/DFtIVk1SQyXGD0gjIOPxbH4Ka84/ZR0y9ufiTJqEMaG2srNxM7rnBfhQp7McH8AfWuz/aT8BeLfEnjSw1LRdNfULU2i22EZR5LBmJLZIwDuzn2Nem/BnwJD4F8KJYMyS387edezL0aQjGB/sqOB+J716ntYU8PZbs50m5HW65p66tod9pkjbVu7aSAn03qVz+tfH/w38R3Xwu8b3x1LTHnmijeyubffsZSGByCQeMr+INfZoz9K5bxj8PfCHi25W71zRop7lQB56O0chHoWUgkfXNc+FxMaacZrRlzhfVHzP4n1TxN8b/GtlpdjYx2ltb7mjQEstshxulkfHJ4Axx2AHevTPh38AtH0C9XU9eul1u5iOYYjDsgQjoSpJ3n68exr2Dwv4X0Dw1YfYdC0u3sYScsI1+Zz6sx5Y/U1qPGMYx0rSpjtOWCsiVS6s+HPjVdS3nxY8RNOCDHdeSoPZUUKP0H611vg3VfiT8TrW28FWmoeTo9tEkV3cpFtCRAYAkYcscDheM457mvdPGvwg8GeLNe/tnU7O4S8YAStbTmMTY4G4DvjjIwfeux8NaFpPh7S4tL0bT4LG0j6RxrjJ9SerE+pya1qY2PIklqCpu55D8VfhTDD8K7XTPC9o8lxo8puVXGZbncMSk+rng/8BAHavIdB+KfjLR/CcXhTS7yOCGNmSJ1hzcJuYkop9ck44zX2htBqhHoeix6kdTj0fT0vmOTcrbIJCfXdjNZU8daPLNXB0tbo8A+GXwGm1i2bXPHs17HJdfPHZrIVm5/jlY5IY/3evr6V2/jz4Y6bD8LdT8N+EtMit5mCTxDOXnlRgw3OeSSAQCTxnsK9bUgDmo5Iwx5rN42o5XK9krHzN+zz8LtSh1t/FHinS5rNrQ7bC1uEw3md5SO2Og9znsK4X9odb+P4t6oLuOUCQRfZcg/PH5agbfXnI475r7RESgdqhlsLaWRJZII3dDlGZQSv0PatY5hJT5mS6KsfOnwD+D91c3Vv4n8ZWbRxR4ay0+cfMxH3XkB6KOyH6n0N79qPwBqmo6jZ+KdFsJ70CH7NexQIXddpJR9o5I5IOOmBX0HHGF6cU8DBz0NZ/Xp+05yvZK1j4u8B/Brxd4q1GH7Xp9zo2lK2ZLq7jKNt7iNG5JPrjHqe1e8/EX4W2Wo/CseFvD9tHbyWG2bTwT96Rc5DN6uCwJPc57V6wRk5NIVBFE8dOUrgqSsfCGmf8LA8M6pLpmnReI9KvJGCyW9ssqFz9F4b6ivY/hD8GtQvtSTxT8QY5ZJdwkisbly8kjdnmJJ/wC+fz9K+ifKXNOWNR04qqmOcl7qswjStucH8W/AEPjbwnJpqOkF5C3nWUpHypIARg4/hI4P4HtXmHwe+BWowawuq+OLe3SC2fdDYpIJPOcdGcjjb/s9++BwfowjIxTl6VlHF1FHlK9nG9z5x/aO+HPi/XvHcOs6BpT6laS2kcJEciKYmUtwQxHBznP1roPgv8HE8K7Ne8QiK51sj91GvzR2gPof4n/2u3Qete2FRXOfE3VjoPw/1zV4mCS21lI0Rz0kI2p/48RWsMZUmlTIdNLU+YPjz4juvHXxMt/DWkP5tpZXAsbZQflkuGYK7/gcLn0UnvXbXf7N2bbdpnipxJjhbm1ypP8AvKeB+Brk/wBlTwwdX8fy61cqXg0iLeGYZ3TyZVfxA3n8q+uFUBcAVviMTKg1CBMKfNufLWk/s4eKJ9SC6vrWl2tmG+aS33yyMPZSFA/E/nX0D4E8E+H/AAZo407Q7MRhsGad/mlnb+87d/p0HYV0gA/GlA4rjq4yrV3NI0oxPPPiz8NtL8e6fDFcyvZX1qSba8jUMyA9VIP3lOOnHI4p3ww+G+j+BdMeCw33F3MQbm8lA8yUjoOPuqOyj9TXf7RnJpdoFT9Zmo8tx+zjc+etY/Z5vNS8ZXuoHxJFBpd3dPcFFgLTrvYsVGTt6kjP6V7b4O8M6P4U0KHRtGtRBbRc56tIx6uzfxMfWtjGadx+VKpialRWbGoJHJ3/AIA8KX/jCLxZdaSkmrRFSsxZsFlGFYrnBYdjjsPSl+Ing7T/ABp4an0PUHmiikZZFlixvjdTkMM8eox6E11X0pODUKtO6dw5EeZfDT4O+GvBV6dSi87UdSGdlzcgfugf7ijhT78n3r0tBhcYpVpcCipVlUfvDjFIQ0AZpee9Jggn0rMYfyo9xRR/CKAN/T/+PKL/AHRU59qg085sov8AdFT1gWYerf8AH830FVat6t/x/N9BVStY7EsU0Dp0o6UuKoQylo+opTQAmccYopaPpSACKMUA80namAAUo9qB0o70gCkalpDQB8/ftlhv+Ef8Pnov2yUfj5Yr5tsiROpx3r6Y/bIV/wDhFNDcD5BqDq31MZx/I180WufNXA6cdK+pynWkjjxWx7D4pj3fBOYc4WWJqyf2aFA8SX53Y3WbDH4itjUyZvgXet1KtHn25FYP7NkmPGM8YP37WT+lPGPSSKwi0PT9YQreNg5yeK1PDkauQGz9aztbLNcHPGDWn4eP3ePSvk5LU9R7Hf6THGiDaAB7Vu2u1hgcVg6WgEPJJ9Mdq37aMCMbTVI5XuWk6DmnAUyJSBjNP79KBi49aXFJzS8+lMLC4FFFA4pAJ1NFGfagGmIOlJR1oz7UAGMmop5AiEmpea53xnFf3emT2WnXItZ5U2LORny88FgO5Azj3xVQV2TK6Wh84fFUan8TPizJpWhIZ4bFfsyuT+7jAP7yRj2G7I99oxmvb/hN8N9J8E6ZiBftF/MB9pu3XDP/ALK/3V9vzzWn8P8AwdpfhjTEtrG325O53fmSVv7znuf0HauvUALgV118W+X2cNjiw+DtJ1J7jY0CDAFLS8UlcJ6AYpB1pe9HPNACEZqKaBJFIYZFTUe9F7BozA1DS2RjJBnBOSAayNV0yw1myk03WLGK7tn4aOVc/j7H3HNdsVBHSqV3YRTDOMN6itqdeUGJxT3PnLxb8EdW0PUovEXgZzdx2sy3Edm7fvo2UhhtY8OMj2P1r6R0q5+2adb3QjeITRrJ5bqVZMjOCD0I6UWkBiQI3OO9WgMCrr4mVZLm6GcIKD0D8KPrSUtc5qA560uKTNKTSAOKaw3IR7Yp1GKAGxLsjVB2GKbMgcDJwakGKRhkYoEVvLjA61BNJEvAPPtVmSNSpXOKpLYA5LOTTGkOSVQ4HrUV8QVPpUdwpimXHQUXjAxBh0pgcj4k2mNgSRXBX5xLlcnnkDtXeeIeUYdK4O/UNPgEjJ9ax+0dMdjivj05/wCERsx63K9/9k1XtiX+CVsemFweP9o1J8emUeGLKPPzG4GP++TS7RD8ErTp8yL/AOhZr67L37iPHzFanlF8MNjI9a9d/Y4P/Fw9YAHXTDgj/rqleQX53SN69q9f/Y2Un4iau/PGln6f61KwzR+6ycErI+tlzgZpwGTTUztHFL0HFfOnoC4pKX68Ud6QDCik5KgmlAHYU7HtRTCwnXtQMehpe9AoAAOKRhTutHHpQA1Vpe/SlFGfWgBKKWkoAXFIB2pc0ZoATFGKXNFACY9c0U7ANJ1pANopetJmgAo496CaBzTAOvWjGKUUn1oGFc98RfDSeL/BuoeHpLlrX7WihZgu7YysGBI7jI6V0PUUCnGTi7olq+hyPwr8C6f4D8NLpNpK1zM7mW6uWXBmkIxnHYAAAD2+tdfQKOKJyc3djSsJ+NJilo5qQAetAFHSjn60wD8KPalOT1pOOlIYfjR1oxxkZowaYWEAGcUYzS85zijmgABIFL34pD0oyQaAEH1pT1pD64oHTrSYkb+nf8eMP+7U9V9P4sogf7tWKxLMTVv+P1v90VUAq1q3/H6f90VV9K1jsS9xaKOaOlUISlpKU0AH1o4oo/DigYmR2FKMZoxj3oAoEFGSBQTRnikAYNIeKWkPFAHhf7Y0DP4B025GdsOpru9Pmjcfzr5etGHmIc96+xf2nLJrz4P6zsGWg8qcf8BkXP6E18badkuC2OPWvocnqacpy4laXPdtFRb34Ia3bsBlIS44z90g1wvwAuUg+I1kjEASq8f1ypru/hXIt74K1XTjh/Mt5Ex9VOK8i+HV5/Z3xB0qV22bLxFb6FsH+ddWMW6Fg5H0X4hhH2x/96rGhrtdcfUc/pU+vpm4dsA1V0uQrMo/DpXyNR2kepHY9C0V+Aua34ZNnTmuW0eZVKI3Dnp711VmV2DOCTVo53uW4zlQSKkOKamMdqdjNCAM5ooz7UvShggpKZNKkSNI7BVAySTwB61514j+NPgTRp3t21U3sqEhls4zKAR/tfd/WtadGdT4UZzqwh8TPSKTNeN2/wC0N4PlmEbWOsopP3zChH1wHzXXeF/if4M8RTJbWGtQrdOcLBcAxSMfQBsZ/DNazwdaCu4mcMTSk7Jnb59aOKYr7sU/Fc1jcQ4qF4UeQMRnHSpmIVck8CvL/EXxz8DaHq93plxNfzz2spilMFsWTcOoByM4PH4VpTpTqP3UJyS3PUFUClHNcn8O/H+g+OrO4utEefFvKI5Y549jqSMg4yeDzz7GutqJwlB2kNNPYTqKbSkgDJrzDxh8bPCPhjxBd6JqEepvdWjhJPJtwy5KhuCWHYiqp0pVHaKBtRPTRRxXjb/tFeBAcCDWW7f8eo/+Kp8X7QngR2wy6sn1tAf5NW31Kt/KQqse57F1pR9K4Hwn8W/A/iO6S0sdYWK6kOEhukMLOfQbuCfYGu9BBFYzpTp/Eik09hRRiigmsxi0CuV+InjzQvAunQXutPPi4kMUUcEe92IGTxkcAd/cVyOgfHjwTq+s2elRHUYJbuVYY3ntwqBmOFyQxxk4Fbxw9SUeZLQTaR6uOlIOtIh3KDjFP/CsRgKXP51zPxB8Z6R4I0VNW1kXJt3mWFRBHvbcQSOMjj5TXLeFPjZ4O8SeILPRNPGprd3jlIvNtdq5wTyc8cA1pGjOUeZLQLnp4qOVxGhcgkDsKcpyoNKQGGCMisgBG3KDjHHemTEgZBxT16U2TGMmhAQBGbnNObhMd6Y77T8p4pd+4e9MBskW6Pms28CxLgdK0jKcFT2rK1JutNB1OT8TPiFs1wU67rteCPn7V2PiiYhWBHHSuRjO67QY/ix071lFXkdMdjzf9oiZR/ZNqDgKryHn6AVra+I7X4O6VDnaZETCk9eM1x/x3ujL4wjtyQfJhRMH1PP9a6r4rg2fhDw/ZEniAE57HaK+wwcbRSPDx07tnkt4TlhgfXFe2/sWRbvFfiKbHC2US9PWQn+leIXbhucYPf3r6S/YtsFXQtf1TgtLdR249gilv/Z64czlob4Ve6fQy4FOpF6YNOrwTsEH1oFL3opDCgUcUUAHaijrk0D0oAOtGaKAMigQe1JS44prnapNMBWYAVzHi3x74T8LDGt63aWsuMiHdulP/AFy36V5B+0F8X9Q07VLjwl4Xm+zzQ4W+vl++rEZ8uP0IBGW654GME189W9re6relIEub67lbJVFaWVz68ZJr1MNl/PHmnojGdW2iPqDWP2jPCNtkafp2q357NsWJT/30c/pWTH+0vpxYh/Cl3j/AK+1z/6DXlmjfBz4gajB5zaMmnw4zuv51iIH+7yw/ECsLS/Butav4judC8PxQ65PbD97NYvmAf8AbRto68Z6Htmun6ph7bke0kfY/wAMfG+neO/Dv9r2EEtsUmMM0EhBaNwAeo6ggg11ea85+Avgm88EeCzZ6myf2hdzm4uERtyxnAVUB74A5PqTXom4Ac141ZRU2o7G8W7ai9qKbuB96Wsigo6mkJHekLAcU7AO4oyMVBDdQTBjDKkgUlWKMDgjqPrQ11AkyQvNGskgJRGYBmx1wO+KfKwuT+1GaQsMZzXmvxZ+Lmg+BB9jZW1DVmUMtnCwBUHozsfuj06k+neqp0pVHaKFKSW56VmlFfP/AMMvjzdeI/GVto2taXYaba3WVinW4Pyvj5VO7g5PHbkivfkORmqrUZUnaQoyUhxIA5rF0jxd4Z1jVrjSdK1yxvb62BM0EMoZlAOCffB4OOlcx8d/Glp4R8FXMYm/4meoRPBZRK2GBIw0nsFBzn1wK+cvgFqkOjfFPRpJW2x3DNaE5/56LtX/AMe21vRwjqU3NkynZ2Ps0GlBpsfKjNP4rjZohP0pOfzp3U0lIAooGKU0wGnHpQKUcdKM+tIYdaSiloEHPSigfWigBB1oxSijjPNDGjd0/myi/wB2rFQaf/x5Rf7tT1iUYesf8frY67RVRc4GcZxziresf8fx/wB0VUHWtY7EscKO9HNGaoQYzRRnGPej8KQBx3pDwaBk0o+lABS9qBSHFAw7UUcGigQlGaBzQaAOe+IWjrr3g7V9IcgC7s5IgSOhKnB/A4r4DtyUO0j5geSDX6MTfdJI6V8CfEvS00D4h63pScR297IIxz9wncv6EV35fV9nUMq8eaJ6N8Dr8JeS2kg+SVSv5ivJddifSPGF3ECVa2u2APQjDcGut+GWpLZaxEXYbWIH0ql8bbL7P43uJ0OUuo1mXIxyRg/qK97Fe8uY8/CT5Z2PpO1ni1nw5Z3qkN59ukmfqKo6bGVnZDgkdD7VznwU1gap8M7dWJMthIbd8cfL1X9Dj8K6aDKXG4Hg+/6V8riYWme5Td0dhocQkMbZHyD9a6m28wY6YFcloDccHArsLEbohzUR2MprUtRTKzFAckdamqvDDtk3j6VZWgkcBUNzKsUZZiAAMkk4GPrUw6cmvMv2jddk0b4aXwgkKTXrLZoQeQHPzf8AjoataFP2tRRM61T2cHI8a+OHxRv/ABLqdxo2j3TwaJAxQ+WcG6I6sx7pnoPxPt5/4W8J+IvFN20Oh6bNdlfvuoAjT6scAVU0DT5tZ8R2OkQNtkvLhIVOOmTjP4dfwr7m8IeH9M8PaFbaXpkCw28KADA5Y92Y9yepNfQ168MDBRitTxaNKWLk3JnyjJ8EPiDDCZvsFm5xny1u13/TnA/WuI1XStR0q/ew1Syns7tMFopV2sPQj19iK+7tRhZoyY/vDpjvWG/h7TNYvbW81fSbe4nspN9s8qAsje3+HSsKWcOz50dMstS1izI+Bdv4pt/A9sviidpJmwbZJeZYocDash7nv6gYBr0HoKSFAq4HFDkAHPavDq1PaTcrHp04ckUjlPit4mTwl4G1LWsqZoYsW6k8PK3CD35OfoDXwq88s0zzTuXkkYs7N1Zick17l+1v4s+2a5Z+FLWTMVkPtN0AesrD5FP0U5/4HXLfA34dJ40fVLu/U/YbeEwRtz/r3HB/4COfxFe9gYKjS55dTlrS5pWRP+zN4o/sD4kRWNxJts9XX7K+TwJOsZ/PK/8AAq+xojuTnrX543cF3pGrS2zFobuznKMQcFJEbHH0Ir7k+FPimPxd4I03WgymaWPZcqP4Zl4cfnz9CK580o3tURpQlb3Tp5h8h+lfDvxzJ/4W14kz/wA/pHX/AGVr7inI8tvpXw38cTn4t+JTnP8App/9BWpyf+Iy6/wh4H+F3i3xlpDatokNk9sJmhPnXHltuUDPGOnIrcf4CfEdFLfYNObHZb9M/riut/Z8+J3hHwh4HfStdvpbe6N7JKFW3dxsYLg5UH0NeiT/AB5+HAhLLq1wxHO1bKXJ+ny4rsrYnERqNRjoYxpxau2fJmoWN7puo3Gn30L291bSGOWNuqMDg/8A66+uf2afFV54k8AmDUrhri80yf7M0rnLPHtDIWPc4JGf9mvmD4l+JbXxT401PXrW3a2hupAUR8btqqFBOOMkDJHvX0p+y54VvvD3gSW+1KJ4bnVZxcLE4wUiC4TI7E8n6EVGYWdFOW46fxWR696UyRwiE+lOH3cGuN+MXikeEfAWpaujAXKp5VqD3mfhfrjlvoprwqcXOSijqeiPmf8Aab8WN4g+Iktjby7rPSFNrHg8GXOZG/PC/wDAK8vhlZCHRirg5BB5B9avabZXet61b2NuGmvL6dYk3clndsZP4nJ/GvQPj78OovBOqabPpqk6dd2yx7vSeNQH/wC+hhvqW9K+pp8lJKmzkbbdz6a+Cvic+Lfh3perzSh7vy/Juz/02T5WJ+uA3/Aq7Xr2r5a/ZE8U/wBn6/eeFLmTEOor9ots9pkHzAf7yf8AoFfUi9K+extH2VVo6acro8S/a+z/AMIDp/Jx/acef+/cleEfA5tvxd8NnPW8x+aNXu/7X7KPh7Z/3jqUeP8AviSvB/gWA3xa8NgH/l7z/wCONXpYP/dmZT+I+34P9Uv0qWorfiJPpUnavDe5uIPrQQDwahUYmL85PFTikBXdFB69aq3kU52eQwGDz9K0HQMORUJiIkyXO30pgRp8iZbkmsjU36mtu52iPtXM624AYA0dBx3OM8RNuYlfXNc/ZktOX25C+361p63MctkZHPNYeu3P9meEtQ1BmAZYWZT3yeB+pq8NHmmkbt2ieF+IZjrPxMOfmE18EC4zxuA/lXWfG2+8zW0tAcLbxhAP/rVz/wAH7VtS+IcFzMoaO2V7h8jOcDg/mRUXxE1Br3xFdzM28GQgfTtX11HSLZ87inzTSOXu3BIAxj6V9jfsraUum/CHT5tp8y/mlunJ75bav6IK+M5pPMcRxjliAPrX6E+BNJj0TwhpOkxjAtLOKI/UKMn8814GYTuz0qEbRNwDtS0KOKBXmHQLRQDRQFw69qOM0fhRQMPwozRRQISlooOKAFqrqEjQ2c0qJ5jIjMF/vEDOKs5pkq7lwaqLs02J7HwJe3EGs+MJr3WLiS2t77UGlu5UXe0SPJliB3IBP5V9cfDKf4a6bpsdj4R1LRVUgbtk6+fKfV93zsfrXN/ED4CaJ4g1KfU9Gv5NGup2LyR+UJIGY9SFyCufY49q871P9nbxpbl2tL7Rr4D7o8x42P4MuP1r2qlalXglzWOdRcXex9L+ItGsdf0K60jUFke0u4zHKI5CjFT2DDkVH4V8NaL4Z0uPTdE06Cytl52Rryx9WPVj7nmvlGSf4ufDTBmGsafZxnq58+0Ptn5kH6Gvefgl8U7fx1byWF7ElprVtHvliQ/JMnTen4kZHbI9a4K1CcI3TujSMk2eluwRcmvn74v/AB5On3k2h+DTDLPExSfUXUOisOqxjoxHdjkegPWut/ab8V3Phr4dyRWMrRXepyizjkU4KKQS5B9doI/4FXh3wO+GVv4otLrxN4gd00Cw3/uYzta5ZF3MM9kA6kcknAxya3wmHhy+0qE1JvZHQfCL4z+IYvF9tbeJtVk1HTb6RYH8xFBhdiArqQBwCRken0r6luHjhR5JHVEQFmZjgKB1JPpXwH4ba3u/F+mmZobK3n1GJnJbZHChkBPPYAfyr6O+OnxK0vUfhNcyeGL4zx6hfnTGmCMowo3SbcjkFcDPcMa2xeETqRUFuKnU01Ob1/45eJdT8e2+n+DbWB9ONysMUckO+S85wWJ/hB5Ix0HJ9Kp/HX4wXV7dT+GvC140VpGSl5ewv80p7pGw6KOhYdeg46+a+GtR+x6W1noME0/iLVXNoJVX5oIWwPLi/wBuQnBbsvA6nD/iJ4ZtvB81poDzpc6xHEJtSeM5SFmA2Qr67V5J7lh2FdlPC0ozSaMpVGzvv2e9etvBug+IvFesXDx6a/l21tbqfmurgZYhB3IBAJ7buelcwPFXi7xx8VbDVbEudVNyosoYj8ltGDnaP9kDJY9+c1zdzbateeEIdXuQYNJs5VsbJDnEkjZd9vqeCzN6lR6Y9D+A/ifwb4I8P6vr+r3Cy63LJ5NtaIpMpjCg4HGFDMeST/CPpRVpQipTSuwjNvQ9++KPi4eDfA97rRRZLhQIraM9Hlbhc+w5J9ga+TPBPhPX/iT4ruZpblwpc3GpajKNwjB5P1Y4OF9uwFej/FCTxZq/wLsNb8TIyTT621ykRXBit5FbywRjgA9M84IzW18FPGHgTwv8G5xqWowRX5mma6tQf387HhQq9TldoB6DnOK5cOlSpcyWrZc7t2PBZLa1ufFIt9HjdLWS+WK0Vm3NtLhVJPqeD9TX3B4s8T6T4R8OT6zrFzstrcbcLy8j9kUd2Pp/SvjLwL4f8U63rMVz4S0qa4ltLlHSUhTHA2cpvZvl4xn8OleyfEz4Z6zZ/BMW730+r6taX7atfNkt5rOCJNoPJCgg++DxziqxcYVJRTYU24on+FejJ8S9fvPib4vaG5t/OaGw05m3xwInTeOmBngdzlj2ryv4X6Kde+MdlaaapNnDqTXO4dEgjfcD+QUfUiuS0jX9a0zTrqw0zVru1tLxcXEUUhVZBjHOPbj6V3PwstfEmvW0/hjwbZ/YftSgarqjMS/l/wBwuPuJ/sL8zdzjpvKi6UW76E812fVfhLxVo/iWXU49HuPtC6ddfZpnA+RmwDlT3HUZ9jXQ9q5f4ceENN8F+HItH04F8HfPMww00hABY+nQADsABXT9q+dqW5tDrQh+lApaSoGL+FBooFAAeaTvS4oGBzQAgoFFA60AHSigdOtJQAtAI9KKAcEcfnQwRu6d/wAeUX+7Viq+nf8AHjF/u1YrEsxNX/4/T9BVQGresf8AH8f90VUrWOxDFo9qKSqAWigfpQOtIAzS4pBRQMPqKKKAaBCDmlxSUUAHGKM0UCgBsg+XFfH37XOhtpvxGi1dQfK1O2VySP8AlpH8jD8th/GvsPtXiv7Wnhl9Y+Hw1a3BabSJfOYDvE+Ff8vlP0BrWjLlkKWx8r6LdvDdRyqQCpBrsfi35ep+GtG1uMhiuYJCPpkf1rzuzk2tjPQ16FooGteCtT0jfmREE8Ixn5l5/lX0Sk6tA8eT9nVTNv8AZi1hF1LUfDdww2XkXnRA/wB9ev6HP4V61ewPbZUjkHA96+WvBmtS+HvFNhqsLYNtMrNj+Jf4h+IzX13qL293YR3kLrJDNGHRh0IIyCK+fxMW1c9ujIb4UvizbJCQQeSTXoWmSBlUqePrXkOn3H2a7IJOCfTvXf8AhrUQ5WFm4PT61zRfQ1qR6naqcjIpQRwOhNV4Z1CIGbqcCrNWYikkA14d+1ezyeD7EKMot+u/2+R8f1r3FyChrzv4raCPEnhy+0kFVllUNAT0EqnK/gcYPsa68A1GsmzmxcHOk0j5d+Fd1DZfEvQLiZlWIXqKzN0G7K/zIr7mtSBbrj0r8+72Gayv5LeaN4LiFyrKeGRweR7EGvpz4TfGnRtR0q307xPfR6fqcahGlm+WKfHRg3RSe4OOelermuHlUSnE8rLqypNxke0MAeopAqjpWJN4u8NRWv2qTX9KWHGd5u48Y/OvMvFvx90HTdVgtNEtm1iESf6TOpKIE7iMkfM35D3rxaeFqz0SPYliKcd2e1oeay/Fmr2ug6BfaxenEFnA00nPUAdB7k8D60zwn4h0vxLosGsaRdLcWsy8EcMh7qw7MO4rxT9rrxcbfT7DwjayfPdkXN3g8iNThFP1YE/8Aq8Nh3OsoNBOolDmR8+eKdTudc16+1m7wbi8maZ8dASeg9h0H0r2z4TfFbwL4L8E2OiumovcjMt06W3DSscnnPOOAPYV5F4N8K654u1GTT9DtVuJ0i81w0ioqrkDJJOOpFdYfgZ8RRz/AGRbEY7XkX+NfSV4UXFU5Ox59OU78yRh/F7VfD/iDxtc634dW4S2vEWWZJo9hE3RiBnocA/Umu//AGTfFb2HiK78L3Eu23v18+3BPAmQfMB/vL/6BXJ3Pwd+IdpbTXE2hDZCjSMVuYmOAMnADZJ68VxXh7U7nR9cs9Wsm23FpMs0Z91OcfQ9PxpTp06tFwi7lxnJSuz9A5TmBvpXw18cc/8AC2vEuCP+P0/j8q19peGNYtPEHhmy1mxbdb3lusqf7ORyv1ByPwr4w+N6MPi34kH/AE+Z/NFry8qi41ZJnZWknFC+Bvhf4t8Y6J/a2iQWL2vmtD++uNjblxnjHvWy/wABfiTxjT9OY/7N8v8AUV6Z+zL4s8M6L8NTZ6tren2VwL+ZjHPcKjbTtwcE5x/hXqL/ABH8CRpubxbooyP+fxP8a2r4yvCo1GJEKcWtWfGXiDw94i8E+IY7bVrV7C/h2zwsGDA88OrDIPI/MV9H/s6/FG98VSTeHdfkWXUoIvNguAoUzxggMGA43DI5HUH2rzL9pbxt4f8AFmt6bFoNwLtLCKRZLlVIVyxU7VJ6gY69OeKz/wBl6Cef4s2ksRwtvazyS/7pXb/NhW1de3w/NNWZMfdlofY7NhSfavlT9rXxadR8TWnhe1lzb6avnXO08GZxwD/upj/vo19J+MdctfDnhbUNbvCPJs4GlIz94gcL9ScD8a+CNUvLvVtVutSvXMlzdytPK3XLMcmuDLKHNJzfQ2qy0sdf8FvEHh/wt4zj13xDBdTpbxN9mEEYfbIeNxBI6KTj3Nel/Fz4q+BfGvgm60iKLUkvAyzWjyWwAWVfU54BBYfjXCWvwW+IctvFMmixFZEVxm7jBAIzggnIPtUr/BL4j5J/sSHH/X5F/wDFV6dRUJzUnLVGS5kjidC1a60PXLLV7B9tzZzJNGfUqc4Pseh9jX3j4V1m01/w/Y6zYtut7yBZo/YEdD7g5B+lfCXi/wAN6x4W1QaZrdp9lufLWULvVgVOQCCCR2P5V9A/sieKvtOj3/hS4lzJZt9qtQT/AMsnPzgfRsH/AIHXLmdJVKfPHoVTdnYv/thAHwBYA9f7Tj/9FyV4T8Dh/wAXY8NEdPtyj/x017v+1/z8PrA9QNTj/wDRcleA/CC8tNP+JegX1/cxWtrBdh5JZWCog2nkk8Clg1/szRUviPuuIYRR7U+uPj+JPgUIoPi3Rc/9fif41MnxD8EPyvizRD/2+x/414zoTvsbcyOoKjrQBWZoWv6NrkckmkanZ36RkK7W8yyBSegODxWmKzacXZgHeo5idpAqXjHNV5ZF5yaLCZSupWCnOa5PxFPiNgDyfeuk1KUICciuF8RXAdmGe9KT0NKcepgXZilZzIXDg/LjpXAfHvUxZeErTTUPzXcu5sHoq/8A1yK72JDLOcj6V4D8aNZ/tTxlNEjhoLUCCPByDjqfzzXbgIe9zBiJWjY2vhBCun6BrniF+CUFtGf1P9K4bWZRNcyMD1Oa7zViNC8BabpIys0iefL9W5/+t+FebX7gknd7179SfJSPDhH2lW/Y6f4OaB/wlPxR0TS5E3QLOJ5/Qxx/OR+OMfjX3vbrtjH518t/sZeHDPq2r+KZUO2GNbK3J5+ZsM5/ABR/wKvqhBxXzeInzSPWgrCClFIPalBFYFi9faiiikMKKKM0AJk0uaPakzQIU0CiigBGIArifH/xN8LeCpo7fV7uR7yRd6Wtum+Xb6nkBR9SM9q7STkYPSvhT4tyanL8R9ebV1eO7N7ICr8EIDhMf7O0Lj2rswmHjWlZszqTcdj7Pt/FmhtpsN3c6nYWhkjDmOW8izHkZwSGIyPYmsfVvit8P9NQm48U6a7f3YJPOb8kBr4bKRkgnZV3T7O5vZhBY2093Ix4S3iMjfkoNejHLaa3Zk6zPfPif8erTUNJutH8KWErLcxtDLeXaAAIwIOyPnJIJ5bp6VzX7Kun3dz8TfttujC2s7GTz3xwN2FVc+pPP/ATVTwZ8E/GniCSKS7tBodk2C0t6MSY/wBmIfNn/e219N/DvwXo3gjQl0rSI2O47555MGSd+m5j/IDgVGJqUaNNwhqwpqUpXZ5p+1vod5qXgO21G1iaUaZdiaZVGSI2UqW/Alc+2a8x8F/FuPwz8Kx4WtNH82+HnL58rjydshJyV6k/NjHA4619c3EEU8TRyxq6OCrKwyGB4II7iuDtPg58O7fWDqSeGrcybt4iaR2hB9oydv4YxXNQxUIw5ZIuVNt6Hz98L/gvrPjOzh1a+uBpOkSMPLLR7pp17lFPAX0Y/gCK9U+M3wxhj+EUGkeFLBydHmF1FCvzSTDBEh9Wcht3vjA7CvbIokjUKqhQBgADgCnlARg1M8wnKal2GqKSPmr9lvwFcDU5/GGsWM8H2bMOnxTRlCXIw8mD6D5QfUn0rmtC+G3i7xf8Tbwa7pV/Y28l7JNf3c8ZVQpYnCMeGJ6DGR36V9chFFOwOtN5jPmcu4lQVrHgH7SXgnVJPC/h3TvCOiyz6dpkkiG2tU3MgKqFbaOT0bJ9Tk9ayPgd8FL9dRi8Q+NLRYkiIe1058Mxbs8uOMDsvr19K+lWVT1xQqqBxioWPmqfIhqir3MfxJ4f07xBod1o+pwedaXUeyRc4PqCD2IOCD6ivCV/Zpf+2Pn8Vf8AEs3ZAW1/0jb6Zztz74/Cvo/ikwKypYqpTTSZbgmYPgzwlovhPRYtJ0W08i3Q7mLHc8jnq7t3Y4raeIOuDUuKOKylUlJ3bKSSPN9a+Cvw91TUnv59C8qaRt0gt53iRj3O1SAPwxXY+HPD+keHtNTTtG0+CxtU5EcS4BPqT1J9zzWsMGl4qpV6klZslQSGKPalo4oz61kUKR7UdKKKAE5oo69qOlAC4zRRSDFAxDS0A0daBCc0YpaSgBQKQ9elL1ooYI3dO/48Yv8AdqxVfTf+PGL6VYrEsxNY/wCP3/gIqkM1d1j/AI/cf7IqnWsdiWLRRR0piCl69qTpS0wAD2oxigUZzSAMetJijOBQKAFHXijFITQKAFBxRmkwTR3oAXFZviLTbbV9FvNMu03W93A8Eo/2WXB/nWlSOAVIpoD85/F2jT+HPFGo6JcnMtlcPEW6bsHg8+owfxrW8E6j9i1OGYH5CcMPY16p+2H4SNprtn4st1byb0C3ueOFlQfKf+BLx/wCvDNNdopQQe/SvcwFS+h5uMpXVy140sBpXiC5twP3RbfFwcFG5B/pXvvwF8RjWfBB0i4lDXOnjCA8kxHO38jkflXkfja3OreHLTVIjvltV8qXgEhOx/A/zrP+E3iaTw14qt7gyDyHO2cdQUPX8RwfwrHFUeWbi+pvgq3NFH0ZKhEhYAb1OGHt610Xh+7ETKSfbn19a53VZFRkuInDRSAMrL0IPOR7Vc0idHVSvTH61401ys9S90emaa4nuYJNxKoCcdua6NCCoNcX4YuMoMkZ6ZrrbaUMuO9VuYMmOCOlYOt2xyXUZB61vjFRzxLIhVhkGqpz5Xclo8K+KHw3sPFTNqFnKlhq4ABlI/dT46eZjkHtuH4g8Y8J1/wZ4r0KVhqGi3fljgTRJ5sTfRlyPz5r7N1HQ/M3GFtuT0PSs0aRfwsTEcf7r4r2qGZckbPU8+vl8ajutGfFotb9mWNLG5Zj2EDE/wAq6/wb8LPG3iW6Ty9Km0+0J+a6vkMaKPZT8zfgPxr6zs7a9GA28e+6t+CILGufvY5NVWzay9xGVPLrP3mct8O/Cem+CPDUelWJLknzLidhhppMYLH04AAHYAV8vfEPRvGnirxzqmtv4Z1hluJysC/ZXO2JflQdP7oH45r7JlhDgg96rWmmW9vLvRTk88mvPoY10pubV2zvlQTjyo83/Z08Dy+GfCsl/qloYdU1F98iSLh4o1yEQjsepI9x6V6qI0x90U7gdBQCK5a1aVWbkzSnTUFZFW5gRlPyg18f/En4T+KrDxlqCaDoF5e6W8vm20kABVVbnZ1/hOR+Ar7IbHeoWWInJVfyrfC4udB6E1KcZbnkn7Mdr4p0fwpfaF4j0i8sY7efzbNpwBuV8l1HPZhn/gVeV/GP4d+NdS+JWuajpvhrULu0uLgSRTRKCrjYvTn1z+VfWMXlDgYH4VKJYUHzsoBrSGNnCq5paslwjy2ufDR+F/xEA/5E/U+f9lf8aRvhh8QxwfB+qY/3F/xr7cmvrGOFp3nhWIHBcuAo/Go9O1XSdRd47G+tLp0GWWGZXKj3weK6v7UqtX5TLkp3tzHxvpHwZ+IWpzqr6H9gRusl3MiAfgCW/SvoP4KfC638B289xNOLzVLpQs0wXCqo52IOuM8knrx06V6FrOr6To1v9o1S/tbGEnAeeVUUn0ye9LoWsaVrVn9q0m/tr233FPMgkDrkdsjvXNXxtWrDayLhCCla+p45+1Q3ibUtI03w5oOj6je287m4vJLaBnX5OEQkDjklv+AivMPgx8MPEN/47099e8P6hZaZbP8AaZmurdkV9nKpz1y2OPTNfXzwRvyVqREUAYGKinjpU6fJFGzgm7siSJQvIpTGuPuipGIXqcVA97ZxuqPdQI7HaqtIASfQD1rj96Wo5SjHRnj37TngW78R+HrTV9GsZLrUtPk2mKJcvLC/UAdypAOPTdXjnwv0bx54U8b6ZrUfhHXPKilCXCizf5om+Vx09Dn6gV9lFVYeuaEhVR0FdtLHShT5GiXBN3PH/wBp3QtY1/wRZQ6Pp1zfzQ6gkjxwIWYJscbsemSPzr5rn8C+MkOG8Ka0vr/ob/4V97YVecV5L8dfiBdeDPskWm20E1zdbjmUnagGOcDrycda6MFipr3Io58TUjRjzyPmNvAXjcj/AJFPWcdR/ojUkfgjxqPveE9aHriyf/Cvpf4CeK/EHjRdUn1lLYQWrpHF5Ue3LEEsOvIxt/OvWltox/CK2rZlOnLlaDDSjWhzo8T/AGUdH1bSdO13+1tKu9PaWeHyxcQtGXwrZwD1xkfnXuYPFMRFXpTxj9K8itV9rNyOtKwySQLxjk1SvW2rkVauV43A4xWRfzFU+as0DMfWrry0YtyK4PXrjzXAVup4re8QXikuucc+tcokzzXf2dCDGDk5X+VTbmlY6IqyKHi7U08PeEr3Ut22bZshyed7cDH06/hXzv4Xs31rxTbwsQU375Wbn5Ryf8+9dp8e/E4vNSTRbZw0Fn/rCOhkxz+Q4/OsXwLGdK8P3etSMVkuwYoR/sjqfxP8q9zCUrWgeXjq6hFsd481P7dqUhJ+RPlUdgB6VxlwS39Oa0dSkaZ2IPPU12H7PnhIeLfiVZW88JlsLLN3dgjKlVPyqfq2Bj0zXRjqqWhzYGDtdn1Z8BvDSeGPhrpNi0Pl3MsQubkEYPmPyc+4GB+FegDHSooECoAOKlGK+dk7u56qEPNJSmikMTHenZo7UAcUAFHHcUAUUAH4UCkzxQuaBBQaUcUmaBjWAPBrK1nw5oes7f7W0iwvtowpuLdJCPoSOK1h0pRxVRk47MTSZzEXgHwZFKJY/Cuiq46MLGP/AArZtdLsrRAltawwIP4Y0Cj8hV7NHWqdab6i5URoiqMBQBTxj0oLAVFLPHHG0skipGoyzscBR6knpUJOTHoiXrxigfSqmn6lp+oKzaff2t2qHDGCZZAv1weKtihxtuC1D60uKBR9aQwoopcUCEo/DFFHWkMB1ooozmmAClpKXvQAnWijIopAFFH0o71QCUUfhRSEHWilHWjrTAbx3BoH0pQfWk6UgFooHSigBO9GKKOooAPSiilABNDBG7pn/HjF9P61Yqtpv/HjF9P61ZrEsxNX/wCP4/7oqpVzV/8Aj8P+6KpnitY7EsKMUtHFMQlHIpeKKADp2oxRmjjHWgYmaOKMCl6DFACfjS9TRgUfhQIKOlFFABSN0zS/lSHpTA5L4q+E4fGXgu/0SUhZJU3W7n+CVeUP0zwfYmvgi5gubDUJrO4jaGeCQxyowwVYHBBHsRX6RuBtr5O/ax8BtpfiRPGFhF/oeosEugOiTgdf+BAZ+qn1rswdXllYyqw5onnHhm7jeKXTro/6PcJsYHt71w2pQPpmrzQE5MbkdMbh2NbdjMVYY4I71P4qtFv7NdQQZmiG2Qeq+v4V7GNp+1pKa3R5OHl7GryvZnrXwS8Sr4g8NPoV5IGvtPTdDnq8J6fip4+mK6sSz2jbV9flxxmvmvwTrt14b8RWmrWh+aFssvZ0PDKfqK+nmlstZ0u31bTpFktrhN6MO3rn3BrwKsbq59BSndWOt8GaklxCpBwy8Mp7Gu9s7j5Ac814Xp13cafeZX93IvcdHFepeGdUS/skkBG/GGHoawWjCcTsoZ9wGanzkZB4rItZcHBrSgcEY702Zokxk0mxc5IBP0pR0pRTENCKOgFO/OikY7RuNAC0E4Ga88174yeANE1a40u/1ordWzmOZEt5H2MOoyBiuwvNZsoNAfWpZttklt9pMhGMR7d2efar9lPS63Mvbw1s9jjPiX8VNK8H6gNNa0nvb3yhKUjYKqA9ASe5x0x6VhfD/wCMzeKvF9noQ0QWoud5MpuNxXapbptHpivAtR1O98beNNQv3by2ufOu2PXyoo0Z8Y9lQCtj9nmbd8XtHyQvyzfLj/pk1e79QpQoNvex4P16tPEJJ+7c+xpW2x7iQABkmvlP4ufFXXNX164sdE1Cex06GQxw/ZpNjzY43FhycnoPSvSP2hfiPdeG7X/hHLC3Tz7+zZnuGYgxKSV+Uepwea+WoLiWC6juI5NjxOHRjg4IIINZZdhUk5yRrmeJlJqnBn0B8GNO+JMfjS3uNf8A7ch0xY5GlF1MzI524UYJPOSD07VtfE3TPiH4q8WtpOmR3FjoSBU84zBI5CQCznB3MM8AY7Vj/DT4jeIp/hx4p1/WrtLhtOCi1kMSr+8ZTx8oAIyV/Osb4UfEnxn4m+JGl6Xeat5loxeSeNIEUMqoxwSFyOQKJRlzuaS0OZcvJGm29Td+OtjZ+Efhfo3hezZmWS5Lyu3WVlUszke7MP0Ham/si2O+58QaoB8oSK3UgepZj/7LXP8A7WGsibxXpulA/LbWfmHn+KRj/RRXqP7L2lf2f8LobxgRJqNxJcHP90HYv6J+taVJKOD13Y6FJyxl1sjO+PngDWvGmqWFxp+pWFvbW0LIyXLsMOTncMA9sD8K6r4K+E18G+DYtNa7hu55JWnmmi+4zHAGM9gAB+FfNv7QGvTah8T9XjSd2ht5FtkTccDaoBwP97Ne0+N/EzfDf4R6Pp1k6JqstnHb24IyUIQGSQj2z+ZFc9SjN0owT3OmnXjGtKbWx6jrXiLRNFQNq2q2diD93z5ghP0B5NZ9l4+8HXkvk23ibSnfsv2lQT+dfKXhnRfEfxCvtR1Jr0kWsJe5vbli3OCQueuTg8dABXM+Frc6v4hsNNUFjdXEcRHszAH+dUstp21lqS80rN3UdD68+KfjPTdB0CaL+0lgv7q2drPYpYtxgMMcYz3r5n8I67Dp3jLTdX1NppLe3uVmlIy78HOee/SvQ/2ltDlsoLDWFu4Es4o47C3tAp3gAMSc9Mcfyrifgr4Jg8d67e299cXFvZ2cIdzBjczE4C5OQBwT07V2YWnRpUG2cWKqVq9dLsfUPgbxlpPizRJtX03z47WCVona4TZggAk9TxgjmuG8XfHnw9pd29pplrPqjRnBlRwkZPsTyfyrlvjfe2XgDwLYeBvDu+BLwvJOWfMjRg85PqzH8lx0rzr4Xp4E+z6ne+M7jcY0AtbcFwXJByRt6noBk4rmoYKlK9R7dDoxGNrxtTi7M9x8CfGvSfE2uQaRPp1xYTXDbInaQOjN2BwARmvN/wBqS+WXxtDaBhiC2XcPQsSf8K4H4Zqbr4h6HBaqxLahER6gBgT+QBrX+O1xJqHxd1aKFS7LJHAi9ckIox+ddMMPTo1rx7HNKtWr0uWfc93/AGZNLWw+G1vclCr308lwSe4zsH6LXq4xisHwRpi6L4X03S1GPstrHEfqFAJ/PNbua+fxUueq2fTYOHs6MUM4psr7EzTs1WuH5welZJGzZXupjt/+vXOaxeeXuLNkVo6rNGkbkvjjrXA6xfMCwEzOOfvdaUjSCuUdeujJNtDEAntzXP8AjDW08L+GJtQJU3cg8u1U/wATnv8AgOa2bNBNGbqXCKMli3AAHUnNeA/F3xSNf151gb/QbXMcHPX+834kfliuvCUbvmZVaooxOWWOfW9bSFnLNO+ZH6+7Ma6LxHcwKYrK2+W3tk2Rj6VW8NwLp2kyak4IubpSsf8Asp/9esi+m3MST717+Hj7ODm92fNV74itboivdyBuF49T0zX1/wDso+DW8O/D8avdxlb7WmFw2RysI4jX8QS3/Aq+cfgr4Ofxz4+s9Mkjb7DAfPvWA48pT90/7xwv4+1feNrDHBAkUSBERQqqBgADoBXiY2pzSsezRhyxHAU7A7mjvQBXCbijFAxRgelApDEFFFApgL0oOaMe9FIBPpQOD1pTRSATGec0ZFLikpgFH6UCjjvQIDVe8uYbWB7i4lSKGNSzyOwVVA6kk8AVFrWpWek6ZcajqFxHbWltGZJZZDhUUdzXx38aPivqXje8ksrRpbPQY2/c22cNPjo8vr7L0Huea68NhZVn5ETnynrHxH/aA0+wkk0/wdBHqVwvytfTZ+zqf9gcF/rwPrXhfiTxp4m8TzFtc1q6ukJyIS22JfpGMKPyrW+FPwt17xywulP2DSFbD3kiZL46iNf4j79B+len618EdMhv4Lb7TZaT4btESS41GWXde3MhyCpZsJGg4xgdT0Pb1IewovlWrMWpM5L9l2HVp/idDcWEcq2Nvbyi/kX7hRlOxWPqWwQPYn1r61XpWH4M8P6L4c0ODTdCtIre0A3DbyZCf42bqxPqa3cgV5OKqqpO6RvTXKhM0Cmb19adkHvXM0yxaSgsKZJKkalnYAAZJPpQk2BJikLAd6rWN/aX9ol1ZXMNzA4yksTh0bnHBHB5FeNfH34tXnhe+Tw74baEajsEl1cOofyA33VVTxuI5JPQEevG1KhKrLlRMpJK53/xO8d6b4G0A6jeq080jeXbWyNhpn64yegA5J7flVL4L+M9W8caBc6vqWjppsQuDHbFJCwmUDkjIHQ8Z6HHtXyT4q8Va94puIbjxBqkl68ClIiyqoQE5OAoA5459hX03+zNZ67bfD2K51q9nliumDafBJ0gtwMLj0Dct9Metd2Iwio0tdzOM+ZnrH40Zrzv4n/Fjw94IJspN2oauVDCzhYDYD0MjHhR+Z9q838L/HzxHqPi3T7G80LTVsby5jg2RM/mqHYKCGJwSM+gzXJHC1JR5ki3NLQ+iqWmqcjnscU6uYoMijg0cUUDDORzRSUvbNAB1opKUUCEoHrSmge9AB0pO1GOKOooAOOtGRRikoAdSA98UUq0MEbmmHNjF9D/ADqzVbTP+PGL6H+dWaxLMXWB/pn/AAEVTq5rJ/0z/gIqnWsdiWHHpRQaBmmIPwo+tAo5pgFGDR2peKQDfwoFH1pc+1AC4oApDS5GKADFFAIwaKYB1NB60UZB7UgEFc78Q/DNl4t8J32h3wAiuI8K4GTE45Vx7g4P6d66LPtSEAjkU02ncLH53eJ9Iv8Awx4kvdD1JdlzaSlHxnDdww9iCCD6GpLG4AUHOVPBHtX0t+1L8N217R/+Er0mAvqenxEXCKOZrcc5Hqycn6Z9BXytZy4IUn9a+gwWIU48rPNxdD7SKmuWhtrjMa/I3KnGOPT616R8BvGy6XeHw7qk23T7t8RO54hlPf2U9/fBrlJYo72zaGRsd1Poa5SUSWd0Qw2spwRXHjMP7KV1szXB1+ZWe6PrbW7BllLqo3jP8XapfC+ryWN1wSFzgjPXFcN8GPHcXiDTYvDesTg6jEu21kfrOgH3D/tD9R9K6u+sbiO7EaKWZui15ko2Z6yalqex6LfR3lqk0TA5H61s27SNzkCvG/BWvS6e3lSMTGGwwPavWNOvklgSRCCGGQaDGcbGzCSygHrUmOaqW0uRzVsc81JA6orkr5ZBqXrXOfEbVxoXgjWdX/itLKWRef4gp2/rinHVgzwG/wDhP4U8U+M725t/ibptxdX15JKbWJEaTJYsUA8zJI5HTtXZftP+JY9B+HkGg2zhZ9UdYAueRCmC3/sq/ia8I/Zwje8+Mujb8sYRNO3fkRPz+ZFSftLeKF174lXMEThrXS1FnGQ2QWBzIf8Avokf8BFepGT5482yOKVBWko9Ta+D2kKfAHj/AMWTjCW+ly2VuxHG5l3Pj8Ng/wCBVkfs5zD/AIXBoIGQGMq/nE9c23gb4hQeH59Sm0DV7fSY4jcSvIdkezGdxUkZ49q1/wBnIk/GPw6M4/fSf+inrpnibqXmc0MGotPse4ftX22lweFbe/ks4G1OWZLeO4I+dYxucgH/AD1Nee/sueGbPX/GV3eX8UN1bWFsGEckYdS7nC5B4yAGNbX7Y+rk6tomjKflSF7h8njLHaP/AEE11f7G2leR4H1TVmUh7y+8tSe6xqP6s1c6ruGHsmV9XU612ix+1fex6b8NYrCFFQXl5GhVAB8q5c8fVRXl/wCyVYG58e6lqW1ilpYlQfRnYAD8lat39s/Vg2raHo46xxyXDc/3iFH/AKC1dD+x5o62/gbU9YeMh7298tSe6RqOn/AmainU5aOvUdTDp1LnjXx71J7/AOLGuvvDJBMtuuOwjQKf1Br6z+Gsa6H8LtHhnOBaaXG8hPH8G5v5mvibxq883i/Wpp1aOWW/nZ0bggmRuDXr/g34m+INS+FHjRtcuopY7GxitbVkiCHdLuj5IHJxit6tNzpxSIg405Ns8y0yWXxT8S7QTZ36lqqlgOuHkBP6E1237U2pyzfFD7HuYR2VnEiA8gbsuePxH5Vj/s7abJefFPS7prdzbWm+4eQodo2qcfN0+8RW9+0j4a1W48fz69ZW/wBvs7qGMfuSGaNlUKQy9R0yDjHNbxjJ1ErHHUlSUXdnVeGTD4c/Zf1C/jYCfULeV2YH+KR/KUZ9QuK81/Zz0w6l8VdKBztti1ycc42qcZ/EiuVTS/EsluLUW95HbsxYRSybYwfXaSBmvYP2bbax8L3+o6xr99awTyRLBBHuDMATlmOPoBRPD1EpNIhYvDJpOSNP9sG8ZX8PWCg7P30x5/3VH9aj/ZW1TQtJ0vWpNS1WysrmaeNVW4nVGZQp5GTzyTTvj3a6V43lsL/SPEFjHdWkbRGGdmVXUkHIYA4Oexrj/BPwue/vo31bxLo9pbo43iGbzJG9hwAM+pP4VEaE/YcskH1zDuteMkWf2qZ5ZviWkbn92mnxeUe2CWP881i+BNM+GV7o4k8U+ItT03UFciSFU+R1z8pUhG7dcnrXvPxS+HWj/ECK1uINSFnqFsnlpcRqJFZOu1lyM89CDxk1wlj+zc7yj7d4szH3W3s8E/izHH5URrRhS5XobqnGpUutTqvhdpvwxs9J1DxR4Tgnu30tXD3NyX3ghNx2hsAZXuB3rhtB8W+BNf8AiJp7r4Gm/tC/1BD9okvWbEjMDvIzg8849q9h8HfDPRvDfg/UvDFrcXs1rqO/7RJLIvmfMgQ7SAAOB6Vn+D/gb4S8O+ILXW7abVLm5tW3wrcTqUDYIBwFGSM8V57xKTbbPQWFi7WR6lb8p1696nFNjjCqB2ApxIHWvPbu7nclZDHbaCcVkandKikscVZ1W7+zqHHK9K43xDqqxoWY8Ht6UXsNRuyDxBqOYyikFj0rkZ4DcXC7Z9+eWCjp7e9TXlxJevtQHaecjvXL/EbxXb+EtI8i3ZW1a5Q+VH/zxX++R/Idz7VrRoyqysjZ2grs5j43+LUtoD4V02XD4BvXVunpH9fX8vWvH7KI3d2qMP3KENIccH2p1z513O0ju0ksrZLMckk9SavFY7K0FvGct1c+p719DRoJJR6I8XE4ht6bjtSufMO0MAiDCjoPwrEu2aTEacljhQvUn0+tT3kuQAp5Ne1/sr/DR9b1hfGWr27HTrF/9DSReJ5x/GPVU/8AQsehrPG4hQVkPC0bansf7OHgI+DfBaTX0O3VdR2z3W4fNGMfLH+APPuTXqwpkShEAxUgxXz0pczuekkN5FLQc0mcUgF5paO1GfakMT8KAKM4paGITmlpMmlGKBoTB7UcilyD0pDk0AL1oxSd+aUc0BcKa5wuTT+lZ3iG+Gm6Le6gVDC2t5Jtp6HapbH6VUI80kiW7I+Z/wBqXx5NqWunwdp8xFhp7BrwqeJp+oU+yA9P72fQVzvwF+Fc3jnUjq2rI8WgW0mGxkG6cfwKf7o7n8Bz05LwppN/45+IFppjSsZ9UumkuZepVSS8j/lu/Svtlz4e8DeEEMkkGmaPp0IUFjgKB29WYn8STXsVqnsIKnDdmMFzO7J5DpHhvQ2kY22nadZQ5JwEjiRR+gr5A+NfxJuvHWstFbGSDQ7Vj9lhbgyn/nq49T2HYe5Na/xX+IGt/FDU30vQ7aeDQrRWnMRO0sqcmaY9FA7DoPcmuQ+E/hRvGHjrT9HYH7KW867YdoV5b8+F+rVphcMqUXUqbk1J8zsj6k/ZytNTsvhNpZ1WaZ5LgvPCkhyYoWP7tR6DaAQP9qsL40/Gq18JXMmg6BHFqGuDiVnJMNqT0DY5Z/8AZHTv6V0fxY8YQ+E/hvq2oaPLAbq0C2cKRkEQTNhVBA6FQd2D6CvkjwHq0Ok+IW1q7sZdY1NMvYwOC4kumPEj92xknA5LYrDD4dVZOpJaBOfKrI6zxH47+LGl65C2p+INQtb5o0uFtFKKgDcqrRKMDP8AdIzzX15De+RoaX+qMloUthNdFjhYiFy+T2A5rw34X/Dq6t9Qm+I/xPuFW8Um7W3uCAISOfMl7Ajsg6YHfAp/7Rfj211X4UadJ4euXkstduWjMpRkLRxZ3rggHlgo+maK1ONWcYxQ4NxV2U7f45eJdc+IVvZeHdOsv7FafbsuF/ePCPvyu+cRgKC3TgDnNcz8c/i9c+Jpp/D/AIcmkttDUlZZwSr3n+Eft1Pf0riPDNxNc6XB4X8O27zaxrcoiu5AMHy93yQKeynG9z9B0U0tnoOjXPxFtfDkmsxJp4uRBdag7iOM7f8AWMhPAXIIUnrwT1xXdDC0qcrtbGbqSZ7f+yINaXwfqr3kkh0k3YXT43HAIB80r/skkfiG965H9pTRPB2iXc97FJez+JdWm+0FWuspCmeXK46HG1Rn17CvT/AvjPS77xSfCHgzTIpPDmjaezSXMWcMwICrH6gnPzH7xyenJ+X/ABhf6hqvja/v/EscsV1JeYuom+9CoIHlgdtq8Y9q5sPByrOWxpJ+6e4/BD4N+HdU8E2HiDxVYy3d5eMZ4YWmZI1hz8gKgjOQN3PYivRvi/4vj8CeBpbyzSIXb4trCLaNocjg4/uqATj2A71ynxN+L+kaHotvovgaWDUtUmjSK2FuN8dupAC9OGbphPz9DxHx90vX7H4b+C/7evJby9heVb2SRtx811DAE98YZc+1QoSq1U6m1w5lFaHCeDdb0Kx1K98U+KYn1zU1fdaWUvzC4uG5MsxPGxeOO5PTivUPgr4C1W+8SSfEnxwPsoEjXcEU42FnPPmuD9xFH3QfQHoBnnfgTrHwz0PTJ9U8TPDHr0E5MbXETyfu8DaYgARu65PX8K0vGXxFuPifqieF9HlbR/DrNm+u5+JJYwecgdB6J3OM8cVviHK7UVZExtuz6bt2R4UkidXR1DKynIYEZBFSZxXD2fjvwpp+nwWdveO8VvCsUYWNidqgAc49BUcnxQ0VQPLtLuTPfaB/WvBcWdHOjus0DB7153J8S42H7rTGXPQtKP8ACq7fEC/eTCWtug9DkmlysfOj0zIoBxXmy+NNTc4CxD6JVuDxHqcoyZQo9lFHKw50d8D6kUuR6iuNi1m+dRun/QCrcN/clfmuDVKNyXUSOnJHqKMj1Fc8t3LgZkY/jTvtgX70u3Pq1P2bF7VG7uHrRlc8msI6hbp9+6hA95BUZ1exGc6hb/8Af0UuRh7VHQ7l/vUbl9a5r+39KJwNStifaQUn/CQaUDj+0Iev96jlB1EdPuX1pQw9RXMnX9Kx/wAf8X50n/CQaWTgX8f60cgvaHoWlHNjGc56/wAzVqs3w1Kk2iW8qOHVgSGHfk1pVzvc3TujmtfuvK1Uqenlrx+dV4rhHHJwazvGGqWC+L/7Ka8hW+Nqsq25bDsmWG4DuMg9KrxykdDXRCF4oxlOzsb2c8g5pRWVBdlOD0FXYbqNzjIBpOLRSmmWPrS0gYGlzUlCiiiimACijHNHApAHFIaXpSHmgBKXNH4UtABxnpRRgUfSgApO+aUdPSkoGMlQOhBFfHH7SHwxbwlrp17R4CNFvpSSiji1lPJT2U8lfxHYZ+y/rWX4l0XT9e0a60rU7ZLi0uozHLG3cex7EHkEdCBWtCq6crkzipKx+e9rchlx0IqLV7VLmLzVH7xenuPSuv8AjB4A1D4f+JGtJN82nT5azusf6xB1Vu28Z5H0PeuRt5RIm32r6GMo4inZnl1KbpS5omLa3Fzp92s9vI8ciMCrKcFT6j0NfR3wu8f23i+wTStVmSLW41+R+n2kDuPRx3HfrXgV/aq58xeW6kVThlns7lLm2kaKVCGRlOCp7EV49ai6Ts9jvoV+ZXPp/VIbiG88yRlJHVgMbvr7+9dV4J8QmP8A0WRzszkEnp/9avMvhr4/tvFVtHouuSRx6uF2xyN8ouBjp/vfzro5bOTT7kqAwHXB7Vxzi0dt1JHuulXiNGp3Ag85rZgmEg+U15P4M1wNAtrLJlk+6SeSP/rV3VlqQVRlhg1mmZyjY6VTxXjX7WutLp3wwfT9yrJqd1HAu4/wqd7f+ggfjXqsWoRtHljivn79qXQPFHjO+0W08PaZJewWglklYSKqq7bQAdxHYGtIaMi5xH7KOn3F9471y9sgizWulSJA7fdWRyFUn24NdD4U/Zx8Rx+JLG/8QappM1nFcrPdRxNI7zAHJHKgckd/Wux/ZT+H+teENJ1e+8QWgtb2/ljVIzIrkRoDySpIGSx4z2r0P4mePbDwFpdvqGo2F7cwzy+UDbqpCNjI3FiMZwfyrplVbdomWiu2X/HHhePxN4O1Lw6109ol9AYjKiglOQc479OleXeAP2f7Dwh4s0/xCPEt3dy2UnmLEbdUVvlIwTkkDmqN5+0tpZAFn4bvHJ/56zqv8ga5PXf2ifE1yXj02w0+xVhlWZTKw/PAP5VUKFZmE8VTRy37Uep/bfi/fxRuHSzghtx3AOwMf1c19FfAAWuhfBnw+Ly4it/Pha5YyMFH7x2YdfYivjfXLq71vWbnVb+Yz3V3MZZnxjcxOTwOg/lW3aS3MtvDFPdyyJGoVFeQkKo6ADtXdRwDqrlbscOLzKNBc0Vc+kfiVB8G9Z1r+1/Et1FqF7HEIglvcyN8qkkDEZx3NZdt8VfC3hrRYtG8I+H51s4s+WjtsUZOSeSzHJJ6145ptn9qmWCBhJK5AVF5JNd3pnw4vp7ITTCVZ3QmOKOIuc5/iI4FenHAYWkkqkrnz9XNsZXb9mrGZ4p8aJr98bu58KeH2uM8zS2YlkI92br+NU18RasbI2kBtbO2bBaG2tY4lJ9cKvWthPh14pE/lnRJwobBfcoB9+vSuguPhdqkLxRwo1xuXJkUqiqfQ5Of0ruhUwdPRNHmVFj6urTONU6k2mpeS38zRNIYlRpT1ABPHp0qpLbrOckurg9+c16Rpnwu8QuFjmuLOCLPIMhYr7gAV09/8JLa6SNodTa1kVAp8uHKuR/EcnOaHmeFp6NmEcox9bVJniENmsLnzC/tlaseXDjjeAPavXJ/g7eSweS3iqRVIwdltz+Zaq8XwLQNl/Fd+w9BEB/7Mah5vhu5ceH8dLVo8sMUKKJNwZG78cU+2tIruVba2DNK5wqjqxr1T/hRyKuxPEs+09Q1uDn/AMeq7oPwgGkX/wBsXWXncAhQYQoGe/XrUvN8NbcqPDuNUtUeT21tqNipkt55YSpx8jkMv5Vv6R8RfE2kN+9uPtkK9UuBk/8AfXWu7T4XXDQyefqIMufkKgkfiDWbd/CfUXDAX1q4zwGVhWU8Zg6ytKxpHLcyw8uaFzrPh58QdJ8UkWwP2XUFGXtpDyR6qf4hXoUWCo4r5l8S+APEfhQReI9PCebZN5peBs7cdcjrjrn2r6A8Ia3BrPhuy1WNgEuIVkwe2RyPwPFfO4/D04Pmpu6PsspxlWovZ11aSNtiFGe1Zmo3ojUgHkVDquqxxKQHFchq+suc+U3Nedex7drljXdUIQ/N07E1xmoXpv3NuC6gnBIFS3c7XkmzJJbvXOeOfF2keC7AmQpc6s65htM/d9Gf0X9T29a0pU3UZtpBXZL4z8Rad4L0n7RNtl1CRSLa3J6n++3oo/XpXzf4h1i81jU5r+9naeedtzMT+g9B7VB4l8Q6jrmqz3t7O1xcTH5mPQegHoB2FV7VUiUPJ80hHQ84r3MPBQXLE87EVnIvW+IYd5z5jDn2FV7ifZz69ajku8jn6Yq/4N8Oan4w8S2uh6RCZrm4b/gKKOrMeyjqTXRWxKpwsjipUHKXMzoPg94CvviF4ujsEDxWEOJLy4xxFHnoD/eboB+PQGvu3Q9LsdH0m10vToFgs7WIRQxr0VR0/wA96534V+B9N8C+F4dIsgJZjh7u524M8uOW9h2A7CuvHHFfPVqrqSuepGKihAKWjr2o71iUHNLSUCgAHJpetA4pDijqMXPHSkH40CgcigAoHFIfxo/lSQCg+1FANGcUxBS/hTSRigMB1oAdWb4lsRqWh32nk4+020kOfTchX+taG8VHNIirl2VR7mqhK0kxNXR8MfD/AMS3vw48fPqU+krd3FqktpNbSSeWyk8HBwcEEeldNe6x47+OPimHTYIQlrC2VgjJ+zWan/lpI3dsdzyegAr3rxz4A+HHiPVv7U1y2iW8OPMlguWiMuOm4KefTPX3qDU/Gfw++GPhwQ6XBFBEM+RaWseHnf8A3j1PqxJx+Vex9YjNpxj7xz2tuzyz40po3w48JW3w68PEPeXqrcaxeEfvJlB+VT6AkE7ewA9STyXwn1jWbS3utE8D2ck/ijWWET3QXiztl67SeAxJyWPCgDqTxzWvahq3jnxlc38qebfahOWCL0Qdh7KqjGfQV7h8MJIfAGjNbWNlay38+Gurx8lpCOij0Udh+JrarNUaXLLVsiOsrkXxR+Hk/hP9n24sVunvr5dRj1DU5+SHY/KxGecDK8nk4JPWvL/g542i8Cavc6k+iwal58QjUswSSIg9VYg4Bzgj6elezeIPGWr6vaT2N00H2adGilj8sYdSMEc15X/wgejpcbxcXXlZyI9449s4zWFDERUHGYTavdGxe6142+NmtLpSvDpHh6KQNcbCfKQdfnY8yP6LwO+B1rqvj34Y0eL4Z6NpfhueKT+wZMJCJA0kkbDDt7tnDH8a56wgt7K1W1tVEMMfOxTgH39z70rsBnGeTxWbxFpJxWiE5XRW+C9jpWgeH9b8RahJKmvS2c9vpkQjOYcoRvz2ZjgD0APrXk1nYXV3MsENvI8nTG3p7n0r15yT0O7Pb0oSzuJeIoWbPOVXOf0rSOOcW5Nasix03w78UaX4J0JNN0nQ90j4a6uZJf3k7gdTgcAdh2/OvFPFNhqt54n1G7ktJ5mu7qSYMqlg25iev416jZaFqczcWco9CwwP1rSg8I3zcyyRJkeuawpYqVOTl3Ld2rHn/gjRptHuYtUdvLvY+YtvPlcdR/te/at3xUl14ksmtdRvbiX5t6OzltjDvg11sXhJVXMl5yByFSpY/DtmhO6SR+e5AqJ4icp899R8vQ8ZXwPcifEt9H5eeqRksfzrr9C0m10y1EFvGQCcsx5Zj6mu6Gl2UTDEAYjjLHNSm1gVdoRQAOMAVVXF1KqtJiULHMwxT4GEPPfFW4oJR/8AXrXkHYDbxVdmIJzgnPBrm3KIYyY+Rjp+VSLdeWSdveo2KqCc4+tVJpF9enOaLXC5fl1d0+5ECPc1HJ4hvUHyJEnfnJrEvtW06zXNzd20B4H72QL1+prndU8eeGbMMz6tBL1+WHMjfoKpU2xas9AHibVmACTxIep2oP61HJ4i1nAxqM3XPynGP0rymf4raBCMQ219cE9hGFH5k1kXfxdlbP2TRQOeGlm/oBVKk7hySZ7PLf6jO/myX1w7EZP701A8k7DLTMx92Jrwq4+KviVxiFLGH3ERb+ZrJufiD4umHOsyRjPSNFXH5Cr9my/Ytn0hHJngZPHpTluFUEvNGo9Sw4FfLc/iTXrkbZ9a1B1weDcMAfyNUPNkc5d3bPPzMTmk6RSos+qJPEGjQ58zVtPQDPW4Uf1qq/j3wpbj97r9ke3Dbj+lfL5IHpTcjpml7MfsT6Wb4teCbf8A5iE8pBx8lux/pUEnxn8HMSFbUM56i36/rXzcxH19aW3AMgzjFDgkUqSP0/8Ag7qltrPw00PVbQv5Fzbl03jDY3HqK64/rXm/7MeD8BvCftZEf+PtXpFebLdm+x8Wftx3VxZfFzSri1mkgmj0qN45I3Ksp8yTkEcg1U+FX7QMsLQ6V443Sx5CpqUafMv/AF1UDn/eHPqD1p37eTD/AIWppvqNJj/9GSV86Ek16mHinTRlOKZ+i+mX9pqVlDe2VzDc20y7opYmDI49QRU8wk2ZjwfavhP4cfEbxP4Gug+lXfmWTNmayny0T+uB/CfcfrX1R8Mvi/4X8aRxWyXA07VWGDY3LgMx/wBhujj9faidJo53Fo75NRuIGwdyn0PStC212FjiYFfcciqZSOZNrrk+9UbnTWXLwtuH90nmueUBxm0dbbXMNwgaN1Ye1T8elefb7q3fB8yNh+FX7PX7uE4kYSr/ALXX86y5TVVUdj26UtYVn4jtJRtmBiPvyK14LmGdQ0UiuPUHNS0aKSZNwKKTI7GgD3pWGAFHNKMUZFFwAdKSl6daBmi4BxmgUlLikMTvQQD1paBzTA5n4h+DtJ8aeHZ9H1aMmN/njkT78Mg6Op9R+oyK+HPiT4P1vwD4jk0rU4js5aC4VSI7hP7y5/UdQfzr9CTgiuV+IvgvRfGvh+bSNZt96N80cqYEkL9mQ9j+h6GurD4iVNmc4KSPgJJllXAqGSEdV711vxQ+H2t/D/WvsV/GZbSQk2t4i4SdR/Jh3U/qOa5WI71HJz9a9bnjXjZnG4Om7oh2PFIJYGZWUg5BwVNev/D/AOJyXMEWkeKJCwA2x3p6qPST1+v515Qy8ZBFV3BVty8c81w1cPKHodFOsfUVlG1nPHeWMqywN8yMjZVh7Gu5stTE1qs8Y+UcOvoe9fKXgTx7qfhmTyG/0nT2P7y3ft6lT2r3nwB4m0fX5D/ZlzvWQfvLdjh429x6e4riqUuqOuM1I9Ci1ISx7UkKkjGDxVvw7FG93ILgFmP3cniuclheDOBuXP3T2q3p2pmCTAYowPRun51hdpjcUem28aLGFQAKB2qh4l0HTPEOkXGl6taR3VpOu10YfkQexB5BHSoPD2sRXsQQsFlXgr/WtsMDzVptaoxlFPRng13+zT4beRmtdb1SFSchXCPj8cCqv/DM+ljGfEd830t0BP619CKRjNOyprpWMqrqcssHTkeJaX+zt4RtowLmfUbpx1JlC598AV02lfB3wNp7bk0VJj6zyvJ+hOP0r0fIxTGdB1IpfXKz6k/UKF9UYOkeFdC0nH9m6RYWrD+KKBQfzxmtXyFAxjpU+4HuKODWUqk5atm8KNOGkUVHtI2OSgqM2qA52D8quuQB9KbvXrkYpKUiuSPYiiiQY+Xn6VYUAdqYHUt8pFSZFJu40khCFz0pQB6VEkpZ/unAqQSLSGGMdKPamu6gE5qjLdjP3uKYy/gD0pkjIB8zDFZsuooq/M46etZV7qqAHDg9zzSvYajc09Xnge1lhfDK6lD9CMVxvh9IPDXh+DRrW6mmig3bXkI3YJJ7fWmalqyyhlWVScdc9K5y8vXmLRxkk5xx3NHPJrl6FRw8ebmtqbF9qu9iDKMnnDGsmNLi/nKoWI3f3cYqvfHTtF07+1NfvYrG2x8vmtyx9FHUn2Ga8X+JPxYuNRjfS/D/AJllpx+84OJpR/tEfdHsPxNbU6DlqzSU4wO7+JnxItPDMUmlaE8NxqQXElxkPHD7L/eb9B7188atqN5ql7JcXc0k8sh3vI7Elz6kmoLqaS5fc7Z/wpYEH3iOnIr0YRS92JwVKrlqyeFFSMM2NxHHfFI8me447Z7UpfI49O9W/CvhrWvFOvQaVolnLd3MzYCr0Qd2Y9FUdya1qVlSVkYxg5vUj0LStT8R63b6TpNnJdXVw2yOKMcn3PoB1JPSvtz4FfC/T/h3oG35LnV7oA3t3jqf+eadwg/U8n2j+CHwp0z4e6Tvcpea1cKPtd5t/wDHI88hP1PU9gPT0AAwK8qrVc2dUY2EUACloGKM1iaBjiikzQWA70AFFRmVB/EKY91EvfNFhNk+aOM1Sa+Ufw1DJqDgcKKOVsXMjTzRmsWTUZe2BVWXULg/xmq5WHOjoywHeo3uoUPzSov1NcrJdTt96Rj7ZqpK5PVvekoE+0Ouk1SzQczKfpzVKXxDZxnhZG/DFcs7Z6H361XlYn86rkRDqs6a58SgcwwD6sc1mXHiW+bIRlX6Liskk7Rj6/SoX+Y8GjlQObLk2s37ghp3/M1lX2oXDD5pWJ+uat/YryU4it5DxxhTTV8PapM5/wBHZRnq/FWrE+8zk/FGsrpekXOozgsIUztDY3HoB+Jr518Q6pfa1qj3t9KZJWOFUdEHZVHpX0t45+HWt634euLG2uLSOdmR0DscEg5IJxxWL8Ofgkuk3X9p+J7iG5ukP7iC3JMcZ/vEkct6dh716OGq06cHJ7mbjJs5T4eeHF0PTRd3cf8AxMbhfnyOYk/ufX1/LtXUEErhQTnoAM16hbeHNJtxkWaMexf5v51fit7eFQIbeNMcfKoFcVas6krs0jFJHkkOg6vdn91YXBU/xFdo/Wr0XgjVZtvmrDAoH8T5OffFepkE+nrTCOeRzWfMxciPOYPAODm5vyQOojTH6mtG28EaRGd0gmmP+3J/hiuvZV5JFRPtHr1oTbFypGLD4e0i2XEdjAMDqVyf1qwII4zhIgoHYAAVYuriOKNnlkWNF5LMcAD3Ncvqnj7wdpzH7Z4k01SM5VJvMbj2XJq1CT6AbpiBbI/Go2jJUgqAfr1rzrVfjh4LtB/opv79v+mNvtH5uRXJap+0BIUI03w6FPZri4yPyUf1qlh5voNI9smGW3MR0wBjpVeZFcck/QCvm3W/jV4zvnP2WW006PoFhgDH83zXHal4x8U6lxe6/qMo/u/aGVT+AwK0WHa3ZSi2fWGo6npthEZLy/tbdQduZZVT+Zrj9Y+Jvg+wYr/bMVy2SCtujSfqBj9a+ZGaSRiWyxPOScn9aaQehwPxpqjFFch7pq3xr0hUI0/S76duxlZY1/qa5jUfjJq8yn7LpVnCfV3Z/wBOK8wbFMY9KrkitkP2aOz1L4l+Lbz7t/HarnIW3hVf1OTWBqGv63frtvNXvpwedrTtj8s4rLBwuQKQ5wM8VfKWoxHsxckyEsT3JyaaSBgcUjHNBwBzQVYCB/Wmtg//AFqHcdBTS2Og5HrQA9io4zTS2e9NPNJjDYNOwx3B6mlBGcE4HrTeT2pQcdvanYB4OT1GB+tNI96QsAvWmMcipcRDyV/SlhI8wDnrUTZ4wadCx80AD2qJrQEfpR+y+c/AXwmf+nRv/RjV6VXmf7Ln/JBPCg9LVh/5EevTK8iW7LPiH9vI/wDF1bAHtpMf/oySvnTjPWvor9vI/wDF1rAY/wCYTHj/AL+SV86ZIPua9TDfw0RLceSNoOefSiORlYMuQVPBBwR701+1HI4rpIsew/DH46+JfDW2z1x5dd00DAWV/wDSI/8AdkP3vo2fqK+lfh/8RvCvjW3B0bUlNyBl7Sb5J0/4D3HuMivgkuNtTWlzPbXCXFvPJDNGdySRsVZT6gjkGspU0yXBM/Rx0hlQhlB+vNZd7oyOC1vJ5Z9DyK+U/h98f/FegeXaa2Br1ihxmVttwo9pP4v+BA/Wvoj4ffFbwf4yVI9O1NIL1utldYjmz7A8N/wEmsJUmjJxaLt5ZXlr1Rivdk5FQRXV1CcxTsp9jg12RYNncKpXenWlzndGFb+8vBrJxJu0Z1r4k1CHG6QSKOzDP61s2/iy3wPtELp7ryKw7zw/Jy1tKG46PwfzrInsb6CRlkhkHXnHGP5VDiaRm0ek2es6ddHEV1Hn0Jwf1q+JEYZBFeNyOy42GnRaxfWbZiuJE542t/So5Uae07nseaTPTmvLrXxzqkPyyLHKAerLz+la1r8QrXA+02sinuVOR+tJwY1UTO86UtctaeONDlx5lw0BPaRD/Stmz1jTrsD7PewS54wsgNLlaNOZMv4zRTPMXpmjcp6mkGg/v1pCBnmmlhSbu1NDaMjxd4a0nxPos+k6xaJc2sw5U8FT2ZT/AAsOxFfG3xg+EeteBr6S6iSS90V2/c3ir9z0WXH3W9+h7eg+5AMjkUy8sra8tJLW6gimhlUpJHIoZXU9QQeCK0p1nAhxufm2WbG1simMoByO/avpH45/Acx+brvgm1Z15abTU5I94vUf7PX09K+bLqKa3maKUMjqcMrDBGOoI9a9WliozVpHNKk09BkgyNp470mm6neaXfxXmnXMtvNEQVkRsEH6ii4YeWAp/LtVQAZOMdfSubFTV7I0pprc+hfA3xyt7tY7DxlbDeePt9uuCP8AfTv9R+VeqW62GqWS3+k3sF7bP0kifcM/0P1r4nA5H4VveHfEetaJOZ9K1Ce1kUZ+R8A49R0P41yRipOzOlVLbn11plzdWF6rIzED1PQV3On6mLmFTJO3XjH9a+XPC3xymcLB4n0xJucG5tQFb8VPB/AivZvAnjDw5rsWzTNVt5JCcCJjskH/AAE80Ok0rofMpHqEV0YpAqMZEPJBbp9KsfbAFJIwcZrlXW9UnE2eDncv8qhju75W2SyA4PXuRWTug5bnWpfRuv3sHuM1DJcLu6iuNGpPFcOjNyDU39qd2bilzC5GdalxtbKtx3FLJd7lIBwema5IaumOH6dRTV1ZWfAbpz6UuYr2Z18d6qRBHO4jvnrTTcq3y9K5Y6mjkDdz9aadU8s5z196pSRHIdUJkRsq3NPN6MY3Vxr64MdG4OM006znGD1HrT5h+zZ18l8q9CKryagOcGuPn1Zx1br09qoXGtHJy3Tr70cw1SOzudUGPvfrxWVeazGvRhnrXKnUpJsgZJxyMdKjWO+mckK209/SkoylsWoJbmhqOtzZ+XkZ6Z/WsW7vb66SRUJjcHAJ5B//AF1JqlzpWhw+frGo21kmOszhSfoOp/AV534l+NmiaafJ8P2LX0ueZ5xtjHvt+836VvDDN7g6kYno1jplwsD3Wo3SQWyjc8srBEQe5OK4Lxx8XtE0APaeFI49RvQMG9mX9yn+6vBc+5wPrXkfjD4ga94mkJ1C+eSLqkIGyJP91Bx+Jya4+Ul87jnPOa3UKdPzZhKtKWxs+L/FeteJr9rzVL2a5cjALngD0A6KPYVhrkAEggHv1pWAIAz+NOQoq4KknPX0p813qYN3JgFI5XIFTGRiqrj5RwAKaiCVgq4DE8ZOM17X8FfgNq3ipo9W8SCfS9Fb5kXG2e4HYoD91f8AaI+g703iORaAqd9zhvhv8P8AX/iBrQsdJgMdtGQbi7kU+VAPc9yeyjk/rX2f8Kvh9ongDQRp+lxmSeXDXV3IP3k7D19FHZRwP1rf8LeHtH8MaLBo+iWMdpZwj5VXqT3ZieWY+prVGBXHOo5as0UUhgAHagsKXvUbA46VBYpc0xnb1pOppGqrESY13bHWoWc9T9KkcVGTz0/GmZtkbNjrULEY61O2TzimGORzgISfpTYWKrnNRMcdKufYrhjyABTl01j96TFCkg5GZbsW9qilDEeordXS4h1LGpobG3TnylJ96OYfIzlWQtwASfanfYbqXlIH9emK61Io15VFH0FPYAD/AAqWw5Tjv7GvSOdqg+pp8Ph85JllJyeABXVMAwyAKdHCq/MQCafMChcwrHw/bAglGfjqxrWttJtIcbYUBH+zVzcq+gxWR4m8VaF4Y05tQ13UoLK3HAMh5c+iqOWPsBQlKeiL5UjZVUUYCijYmcgDn2r558W/tIxiV4PC2hh1B+W4v2I3D1Ea/wBW/CvRvgP411Pxz4UuNT1WC3iuILtrfMCkIwCqwOCTz82OtXUw1SEeaQ1OLdjstRtUKl1UAjnisuVSF5HFdBdD90foc18Z+Mfj/wCOppJraybTdOjyVDw2+9+vYuSP0qsPSlVWhnU0eh9RtIDx1qrcX9pZoXubuGBB/FLIEA/E18XXXxL8bas3lX3inUsEYCxy+UD/AN8YrImaSZi80ryuTuLOxYk/jXoQwF92YOTPsDV/ip4D01nWbxFbTSL1S2DTH81BH61yGrfH7wvBG32DTtUvH7BkWJfzJJ/SvmzLH27U1xjr0rojgaa3HdntGt/tB6xN8uk6FZ2if3rh2mb9No/nXEat8WfH2oMd2vyWw5wttEkQ/MDP61xQDAZIB/Gg5C9Rk1rHDQjshk+p65q+osWv9Tvbsng+dcO+fzNZkj9CB+nSpyq4JYc1GV9Rx9KfKkWkVyxBxjrSHnt3p7jCZyM59KYxOAazkjZJEbZ6Y9qayORwQKswRPNOsUaM8jsFVVGSxPQAdzX1j8AvgPbaPHb+JPGdqk+pHElvYOAyW3ozjoz+3Rfc9OStVUBpHl/wd/Z+13xekOreIXl0bRnwyKV/0m4X1UH7in+8fwB619JaH8Hvh9oliLW18LabKAuGkuoRPI/uWfJ/pWp4v8f+H/C+q6fo08zXOr6jMkVrYW43SuWOAxHRV9zjocZrrJCFjLOwUAZJJ4FefKrKTDlPgD9o3wxpvhL4q6lpmkxLBZPHHcxQqeIt6glR6DOcDsCK84PAr0D4xagviz4g614gikbbcXLCHJyDEvyp/wCOqD+NcFtxkNjg16PI4pXFF3GFmwMZIoLY5xTs89O2KaT8o6UWHYD1znJ600n5sjnPWlZiMD1700kD8qLFC4z1NCHawYAHB6Hmk38dKQPjPHPSmIdI2WJxgn0qPcxOM5oZst9KjJ+bIJqWxEobntQWB/z1pi5ZsZ616n8HPgv4j8f3Md2VbTtDDfvL6Rc78dVjX+I+/Qevasp1VEaR5eT8uMfpTCcHg5ya+9dP+A3w4svDM2kf2Glw80RRr2dt9wGx94N/Cf8AdAFfCuvWDaXrN9pkjZe0uZIGPqUYr/SohW5xdSqWY8U6EnzVyM81EGIFOgI81ST3qpPQaP0o/ZYbd8BPCx/6d3H/AJFevTq8t/ZRYH4BeGMdoZB/5FevUq8mW7LPiD9vP/kqth/2CY8f9/JK+dGJCjIFfRn7egx8VNPOf+YTH/6Mkr5w6tXqYb4ESx4PH3qaCM4OetKwxgCkDYPPPaukkcWPpSOTjpRxnk5NMdgM0BYPm3deMU6OVo2DKSGU8EHBB9aiZ8d6bnPJoHynp/gj42+OvDLRwtqH9r2SHm3v8yceiv8AeH5ke1e6+B/2hfB2srHBrIm0G6bg+f8AvICfaRRx/wACAr48yDQHVe9ZuCZMqSZ+j2i6tp2rWa3mm31veQN0lgkDqfxFaI2NkEA+tfnP4e8Ta34fuRc6Lqt5p8396CUoG+oHB/GvXPBv7R/izTQsOvWtrrcI6uR5E3/fSjafxWsZUH0MnBo+qr7RtNuQxe2RCe6fKawNQ8Hqw3WlyQeyyDOfxFcb4U/aD8E6wiR6jLPolwxwVuk3R/hIuRj64r0vRtd0rV7dbjTdQtb2JujwSrIp/I1i6biQzi9R8M6xBykQnX/pm2f0PNYl7DdW5CzwvD7OpGfzr2JSjdxmny20M6bJI0cHqGGRSFZnhjykjuOaalzMhykjD6817JdeGdFuMmTToMnnKLtP6Vi6j4C0mUE2ss9u3UDO8frz+tLQLNHBWXiXW7IbYb6ZFzgZYn8a2IPiVrdooEoguQO7rz+OKlv/AIfX6KTbXlvNgfdYFCf5iub1DwprtuzBtOmceqYcH8qVkVzSR0w+LOoRnMljZspGQAzA/SrFv8XpePM0aM+uJyP6V5deWlysu2S3ljZeqsCD75pkOVOMZ9KXKhe1lc920z4p6HPGpntbuBuuNoYfnmti28f+GpwCb1oie0kTD+lfP0CccHIxzV+GOMplycdsY4pciLVVnvX/AAlvh2fhNTgB/wBrK/zFeafFP4deBvHqSXcV9Z2GsEfLeW7r857eYuRv+vB965SIBlBVjjp6GrUA2H69jTirbD9ofPHxA8CeI/Bl95Gp24ktj/qruHLQy/Rux9jg1y4RmAOAOlfXThJ7ZredFmhcYaNwCpHoQa4HxR8KNLv2a50V/sE7Z/dEkxH6d1/UVbXMNVVszwlhgDOPTp+tCr+7zuByP0rofEfhDW9BkI1GzkjTkCXGUb6N0rmJrkRN5cnBzjpjinTXK7sc9VoOYJyA3vUkFzJAymJyGHQrwR+NUzOhBKtx6U5GV1UjofetkozdkZ3lE9F8OfGDxx4fVEt9YluoVI/c3f75fpzyPwNek6F+0TBdKItd0BCx+9JZy4P/AHy2f5186k8dunpVC7nMMuVPPes61FwVzaFW59iab8TvAOqAM+qPYSMcbbuErj/gS5H6101lcaDfoHsNZsbgbf8AlldK36ZzXwvFfy/dc596uw6hPGS6uT2zURVOW5p7Ro+4RZyOu5JEdf7y8iqs2nT5+V8emT+lfG1h4r1mxH+iajeW+P8AnlMyfyNbtp8UPGlucx+I9QI5+/Lu/wDQs1ao0n1H7d9j6lksb/JACnHoef1qI2eodMYI4PzV80r8WPGzAf8AFQ3jYwc7l/wq/D8V/HUaf8hlznkFooyfzxW9LBRqfCyJYlR3R9Dvp16SNsgJ9KYum3xJBlG3PUDpXzlcfE7xxKrBvEN4oz/AVX+Qrn9R8aeIL12W71jUJsjBD3DEE/TOKupg40tZMUcVzbI+o71rKyAa/wBZsLZRwfMuFX+ZrJu/HXgDTPmn8QRXLf3beNpMfiBj9a+VLnUJTku2c9apfbld8bic+tc3NRjsi3Vkz6T1L45eHbbcujeH7u7YdHuXES/kNxP5iuF1340eLNQZlt7iHTojwEtU2kf8DOW/WvKPMPBZic8cmlypHXHNNYhL4UZvme7NbU9av9Qkaa6nlmkPV5GLMffJ5rJIYNuYHdmpRnGDgD3pkjZ47ZrKVWU9xKNhfMctlh04x7VMXwuRg+2O9Rfu2O0dccmtrwj4W1nxNqK2Gh6fcXtwxztiXOB6k9APc0kwsZG5t2cEk1t+CPB3iPxjqgsPD+nSXT5/eSYxFEPV2PC/z9K+iPhn+zVBHbx3fju+j3AhhZWcnI9nk/ov5171o+l6D4fsI9P0i0tLC0T7sUKhR9T6n3PNJybHojyf4OfAHRvDFxHq/iNota1VfmRSn+jwH/ZU/ePozfgBXvMaqi4GKyP7UtU6OD9BTk1aM8KpNQ02HMjWOOaQEcc1knUZHHAA/Wj7XKeS1JQYc6NbjrTHKgYyPzrMSZ2JySTTw+etPlFzll5AM4PFNWTOQOKgDZ4py9aaJbuThVYcjrS+WvpTFYdjipVZSOSKLMV0CooOQo/KpRgDimIw9ao6rrui6WhfUdVsrNR18+dU/maFTnLZDdSEVqzS4I4pOledar8afh3p7Mp8QxXLDtaxPL+oGP1rlNU/aN8MRhv7P0nVLpgOC6pGD+ZJH5V0QwNeW0TCeOox3Z7gDimk+pr5k1n9pHVnG3SvD9rbE/x3MzSfooX+dcXrPxq+IGpqUGsizU9rWBUP5nJ/WuqGU1nvoYTzOktj7KkkjXlnCj1JqGzvrO8keG3uoZ3jHzrG4Yrn1x0r5B8B6X43+J2um1k1rUJLZMNd3U0zNHCvpjOCx7L/AEr6w8D+FtI8JaJHpekW/lxrzJI3Mkz92c9z/LoMVzYrDxoe7e7NsNiJV9bWRtQxBRmldgvNK7hRycAV83/Hv4zCQ3HhrwheEJzHd6jE3X1SIj8i/wCA9axw9CVaVkdbaR1Hxl+Ndh4Wkl0fw+IdR1hcrI2cw2p9Gx95v9kdO57V8u+JfEWt+IdUfUdb1Ce9uGP3pDwg9FXoo9hgVmyv75zyf8agkmjjPzHJIyBX0VDCwoxMJTuW4Q8sqqqkscYAHJNfdPwX8LN4S+H2m6TMm26Kme7/AOuz8sPw4X/gNfOX7Kvgh/EfiYeJ9QgzpelyDyQw4muByo9wnDH32+9fYC4VOPxrzMzrp2gh0ItvmZleM79NL8LarqTYxa2U03Jx91CRX5z6whZN5wSD2NfY37Vfi9NG8Df2BA2bzWTsI7pAhBdvxO1fxPpXx5qZP2ZjnAOK1y6k1TcmXUabMiUEHIP5VoaZfBiIJm/edFJ/i9vrVCYcZHPtVScYAYEgj0rr5nFmcopnVHlTTGzgA4qho1/5oEE7fvBwp/vf/XrTkUleODXRGSkrojbQgG4s27tTWU8YHWpMY6GmFcnk9ef/AK1BVhjD1Yr79qibcTycnNWTg9vwxUToM5HBz0FS0UitJz7e/rSRQyTTJDBE8ksjBURRuZ2PQADqasNES4ABOegFfWX7OHwbi8PW0HinxHbh9alXdbQOvFmpHUj/AJ6Edf7vTrmuPEVVTWpomJ+zn8E4vDEcPifxTAsuuMA1vbtytkD39DJ7/wAPQc811/x1+Kdh8P8ARPs9sY7jXbpCLW3JyEHTzH/2R6dzx61tfFvx7pfw88JS6relZbl8x2Vruw08mOn0HUnsPqK+UfhZ4c1n4yfE241XX5pZrJJBPqU/QEfwwp6ZxgDsoJ9K8qzn78jWKPV/2Z/B1/q2qXHxP8UtLc6heFhYvPydp4aX2z91fQZx1Fdn+0l45i8LeCZdKtpgNV1dGghUH5o4zxJJ7YBwPc+xruNf1PR/BnhGfU7wpaadp8AwqADgcKiD1PAAr4e8e+KtR8ZeKrrXtTO2SY4hiBysMQ+6g+nr3JJ71thKDqz5nshVJGDdbApbgBRmuXblyTnJOa3NXuMQiFRy/XntWQy469xXp19XZE01pciOcdO/WmEkDn6dKnYYQZHFMKrsz3NYtF6MibtxTSMEj1qUgj0NMKgDrkmpsOxGQAelNIz+FPI+Y0w9KlokY2ABz1prDOOaeQScY6V2Pwi8Aan8QfF0OkWStHbJiS8uMZEMWeT/ALx6Aev0NYTlYDqv2cvhFL8Qdb/tLVEkj8O2UgEzAkG5fr5Sn0/vHt06nj7o0y0stNsYLGzt4re2gQRxRRqFRFHQADoKyvC2haT4N8K2+mWMcdpYWMOPmOAqgZLMT+JJPua8Y0P4p6v8RPjlYaH4UuZbfwzp5knupVUbrxVGMtkZCFioAGM5yfQcc7z1JufQlxtWIscY6mvzW8dQJca/qmqQuxSe8llwfRnJ4/OvvT42+Jk8LfDbVr3zAtxJCba2GeTLJ8ox9AS3/Aa+Er9Va0lHQBTxmu3BUeaDbBvU5TBJGakt1HmjI708qvQU6FTvGCOtVOGhaP0c/ZNGPgF4aH+xL/6OevVK8p/ZLOfgJ4d56LMP/Iz16tXiz+JlHxJ+3mSPifpvcf2TH/6Mkr5wwMnPBr6R/byAPxQ0zj/mEpz6/vZK+dSn0r18LG9NA0V2bp7UhbGe+asFEHVAexpCkZGNtdSgQVHcHvTC2fzq6bdD/AKRrSMNgryfQ1XsmMpdOMg5ppIIA71f+xRNzyPxpTZRAdT6UexYXKBIYdefpTguVBxV4afGc/vOntThp5/56fpR7Fi50UApPalHAHFaJsCejj06U9NNYgAuOnpWipshyRnhnGOat2V5eWcomtLme3kHR4ZCjD8RUw09s/eFPFieoZfSiVC6Ickdz4Y+L/xC0aJI4vEMt3EvSO9QT/8AjzfN+ten+Fv2kp40RPEXh9ZD/FNYy4/8cf8A+Kr58S1kBwCvpUi28gJzt/OueWDbDmifY/h344eAdXiQy6r/AGXMTgxXyGMj/gQyuPxrutN1bTNUgFzp1/a3sR6PbzLIv5qa/PspKBgj2OKt6fNeWkoms557aRejxSFG/MVhLByWw9D9Ad6t0IpTGpr4o0D4jePdIIFr4lvmQNu2XLCdSfT5wTj8a7rQ/wBoLxVbOF1PSdLv49vJTfC+frkj9KyeEq9ETofR2o6ZY3OTPawSt6ugJrn77wzokrEnTog3qgK/yrhtM/aB8O3MSjUdH1KzkIG7yikqj15yp/Ste3+LvgO9Of7Xa3YnG2e3df1AIqXhay6Ccomjc+F9MAykJXjjDnFZV54eiVgYmK47dRitGDx34TvAPs3iDT2JOMNKFP5HFWRc2d2DJbXEUy/3o5A38ql0qi3QOUTnpdLdR8rBvXtURttoyOp+8v8AhW1cMp4GRjiqMyg56/iankZm2rlAYXhiWA6k9R/jUiTwght4HHrUU6nB6+nB71nXER5I4J5waLCujdklsrqJoJ3t5o2GCjkEN9Qa4bxZ8KfCuv7pbTOnXP8ACYCGjz/uH+hFTXsnUe+M1S80r0JUk546CjUOdI828Q/BnxTpW6XTfL1KAf8APE/N/wB8n+ma851XT9W0q6aG6tp7eUcmORCpx9CK+l0vbuIZS6mHbhz+dQ6m8erRGDVIYr1PSdQxx7HrTt2LVTufOVreMcFgMDrUF+26YMOh7Yr6AHgHwPqNk8VxBPp1wD+7mgYlfxU5zWHc/Aya4kY6R4o0m5XOFWdWib9Nwq588o2KjKFzxPB/rTjM0ajB47ivSNc+DHj2xUvFozXqD+KzlWYH6AHP6Vxmp+EPE2nEi+0LUrYDqZbZ1A/MVhySXQ2uigLmJhg8ZAGMU2SdCNqMPqKrzWtypwYiMdRimeVcKf8AVNz0OKnlaFzGpYKPLjJ75Pr3rTMRbABC9+tZGltIkeJBgg9D6Vee6b6AHpXsYCUYK8jmrJvYsvCCuM9vWsqX5SBnP0q8bpmTgdu/aqLRTFzt+YHnBqcwqKduUuguXcZfN5gK4PTFZ8IPmAYIPQ1sx6Vf3bKtrazTMQMCKMsT+VdBo3wr8eamVeDw3exxt0kuFEKj8XIryIxkzeTRywYZHJGD2pSxJHHHt3r1/wAOfAbXJVL69q2maWBwEVzcSZ+i/L/49XZ6H8GPA+msH1W81LWJV5KZFvEfwXLH/vqtI0ZvoQ6sF1Pna3id5UWJWeR+AoGSfoBXd+FPg54419o5hpTaZauc/aL8+UoH+6fmP4CvpLwxB4b0CIRaDo1hpxUYDxw5lP1c5Y/ia1LjUfMYNLMWJ7k8VsqDRnLER6HmPhv4D+GNNeOfW7+51mdR88S/uYM/h8xH4j6V6todrZ6LYLY6TZ29hbKOI7dAg/HHU+5rOuda0q1Qvc6jawhTzvlAqkfGvhWM5bX7IY4wHJx+QrVUaj2iYSxMFvI62O5lYgPIzZ75zVhZOB9K4H/hZfgqJjnVHc+q27n+lQS/GPwhCB5cWpTn/ZgA/mRTWCrS+yYyxtJfaPSVyeo5qWMkNnOc9q8YvvjjbqCLDw9OxzwZrkKMfQA/zrIufjb4gfcLXSdMgyOC5dyP1AraGV15dDKWaUV1PoyCT5fX6VZRg3QGvku9+K/ji4UoNXEAP/PGBFP54zWDqPinxLqEe261/U5VHO03LAfkDXTHJpvdmEs4gtkfZdxfWltGzz3UMKL95nkCgfnWVeeOfCVkjtdeI9KQRjLD7UpP5AkmvjGaSSYkyyPIfV2J/nTFHPAxW8ckj1ZhLOZdEfWOofGnwFZhTHqcl4T2t7Z2I+pYAfrXK61+0PpqRkaRoN5PJ2Ny6xr+m4188fMeRRtyM46V1wyijHcwnmtaWx65qv7QPi2Y/wDEv07S7JM/xI0rfqQP0rmr34wfEa5YH/hJJYec4ggjQfotcOQGGPakChSTiuiOBox2ic7xlWW8je1Dxz4z1Bj9r8UatKGGCoumVfyXArCnnlmkMkzvK56s7FifxNR8ZwM0hXPAY4reNGEdkQ605bsUsWwSMY44pP4cFc46U7btAyOtBYggoSpWq5SOYgGWJ9PWur+GHgbUvHPiFdNsQYrePD3d0VysCf1Y9h3+gNZfhTw/qXibX7XRNKh8y5uHwCfuxr/E7egA5r7Y+G/g/TPBnhuDR9PQHb808xXDzyd3b+g7DivIzLGqhHljuengMI68rvYseCfCuleE9Dg0jSLfyrePkk8vI3d3Pdj6/h0rbkcJx0qRiFHNfO/7TPxS+yibwV4fuf8ASXG3UrmNuYlP/LFSOjEfePYcdScfNU6c8RUPqIxjTjZGX+0J8X2vpZ/Cnha7xZLmO+vYm/157xoR/B2JHXoOOvz9KxLew9KVzn5Rx9K6j4d+A9d8ban9k0i3BRMefPIcRQg92Pr6AcmvoaVKGHgc06hx8shC5jGT+grsPg18NdW+IPiQxJvg0qBgb28K5CD+4vYuew7dTxXs2s/s6Ww07RrLSb52vDdH+0r2XhREV6onsRgDkktycV7z4P8ADek+F/D1toujWywWluuAOrOx6ux7se5rjxWYRUfc3HThKT1I/CXh3TPDWi22kaRarbWdsu2NBz7kk9yTkk9zVvXdVstH0u51LUblLa0tozLNK54VR3/z1q85CJknpXyX+0z8TF8Saj/wi+h3IfSbOTN1MjcXMw7A90X9Tz2FeXQoyxFTU6XaC0OF+LXjK58b+L7rWpQ0ducRWcJP+qhX7o+pyWPuTXA6pIdixjoSCavyOBHuOAMVkTEyuzEgHr+HpX0vKqcFFHPe7K0uWx7dBVVwSxwc+9W3Xd7U/StLv9Vv0sdMsrm9uZT8kNvGXdvwHNcc3YsziWVgV4x6djW/pd6Z49kh/ejgn1HrXrHgf9mfxlrSR3PiC6tNAt25MbjzrjH+6p2j8Wz7V694f/Zj8BWCI2oXesanKOSWnES/kgz+tc6xkKb3Bwuj5VIAOcUzBL5Jr6l+Jf7PmiS6PLd+DFls9QhXK20kxeKfH8OWyVY9jnHqO9fL8yPC7RPGyOhKsrDBUjgg120a8KyvEizjuRHkYPHvjrTGUHoD6YqTIHce1d98DvAknj3xelrMrLpdpia/kHde0YP95jx9AT2qqk1CN2Ud5+zB8KDqN3D438QW/wDocLbtMt3XiVwf9cR/dB+76nnsK+mtb1PT9B0S61bU7lLaztIzLNI3YD+ZPQDucVas7S3tLGK1tYEhhhQJHGi4CqBgAD2HFfLP7QHjPUfH3i+DwB4TWS8tbe42OsLZF1cD36bE556Zyewrw5Sded3sbRjZHGeKLrxJ8cfimIdOhZY2ylrG/KWVsDy7+56n1JAHavqv4f8AhbQ/h54Sj0u0KQ21shlubmUhTI2MvK56dvwAA6CqfwZ+Hlh8O/CflSPFJqU4Euo3fQFgPugnoijp+JPWvB/2i/i0fFE8nhjw3ORocb4ubhTj7awPQf8ATMH/AL6PPTFXGLrS5Y7DbsZnx9+J7+Oda/s7TJWXw/ZOfI7faZOhlI9OoUdhz1NeT3EohiMmM47Duaf91cms67kLsP7o4H+NexGEaUOWJjfmZVnZpZWkbqefw9KhbkggdanfBGNuKi254x361izVCFQcUwpnjofWtHTNNvdQuY7SxtZ7q4lOEhhjLux9gOa9c8F/s6eNtaVJ9YNroNu3JE58yfH+4vA/Eg1nOcYbsLniDDkBe1MYHuK+yNE/Zi8F2satqd7qupSj72ZhCh/BRn9azvHv7MmgXOkSzeEp7iw1GNSY4p5jLDKR/Cc/MpPqDj2rm+sQbsVc+SHT1IPeovLBPpzWhqFlcWd3LaXUTxTwu0csbDBRgcFSPUGq+wDk1q9h3I7Wzmu7yK1toXmmmcRxRoMs7E4AHuScV+gXwD+HNv8AD3wNb2EiI2qXGJ9QlHO6Uj7oP91R8o/E968M/Y1+HsWrazP441KEPb6fJ5NgrDhp8fNJ/wABBAHu3tX1N4x1yy8LeFNQ13UCBb2UDSFc8uf4VHuTgD6151WV5cqA8A/bE+IxsrBPAWkz7Z7uMS6m6nlIf4Yvq3U/7IHrWv8AseeBZNA8Hz+J9Qh2Xus7WhVhyluudv8A30SW+m2vIPhX4O1P4u/FO51rWw8lkLj7Xqcv8LZOVhB98AY7KPpX0z8bPG1r8O/ATz2hiXUZ1NtpsGON+PvY/uoOfyHem4XagiTwT9rHxl/bPjOPw3Zy7rHR8iXach7lh83/AHyPl+pavCdbn22u0dXIHWtGeWSeV553aWWRi7uxyzMTkkn1Nc9qkvm3WB91OP8AGvY5VRpWRF7sqY3e2KdCQHGc8GkZeMg/WnRAB1781yS2ND9Fv2RTn4B+H/8AtsP/ACK9es15H+yF/wAkF0Iejz/+jXr1yvDn8TLPir9vAf8AFzdLI76Sn/o2SvnMcA96+jP282x8S9KU540hDx/11kr5waQdAD6V7OEa9mhNi7QSck5z1pxBHRtwpm9COQaDIqtxnBrrUkIUjBA9RSqinqTSeYnO5D1pyzDGAhq1NEO5NtwAOo7UFMYxzTDOB1XPtTlnJxiPj3q1JENslXJ6x9KmDARcDBzzxUInb+6B75pVucHDIOtWponlZLGC8m4jAHTmrCAI3GDkVBHdnHKA0NcNxhAeapSRLTJy3Hzrx04pu3POff8ACozM4BJVW7nmm/aCQMDv3qlJEuLLPI4GCD+lKN3Pf3qLznC5ITnpSrcSt2UVakieVkijvjFLzwQKj81ioK4H4UjSShc4B/CrU4lcrJwSoxQNw71AJJt3JGPpThI+eo/EVSnELMn68e1OXdmqwdznn9KmBbH3u1aRkmQ4NkhGcVPDPcQHME0sR9Ucj+VVQSVOG6UBmAGTVWi+hDps3NP8UeI7MbbfV71VB6GUkfka17bx/wCKISCb4Sr12yRqw/lXGqxzwSOaeGYnrUOjSe6JdNs723+JetqQbi1tJgOo2bc/kasy/Eh5ExJpSAkYO2Q4rzvcCOD2qI59TWUsJRf2Rexfc79vG9tI58y0kTPUiTp7cikTxfp2eRMCeecflXnzNUbkkf4Vm8BRfQl0p9GekR+K9ICndNKpzwGjzkfnUyeKNIY4Fzt7fNGfzry88npSgkc46+tZf2dSZSpzXU9Qh8Q6MvP20YPXcrVZh8UaIMM18gOccqR+PSvJicc/nTTnGQc1X9mwXUlqR7OvjKziwYNbMfrtdulaVr8Tvs5KHX1Yf3JQT+fFeE7hjFRS7pJC5csx5JJ5NS8FGOxpGMmtT6Fb4m+ELmLbrmkaVqQbGT9kRyf++hVaXxN8Fbhy0vgax3c8JaBc/wDfJGK8DAYqNpAx2pU8xWJUkfSl9Ug90Nwn0Z7ZLrHwTJZk8CLn0AkwP/IlVbnV/g64A/4QMNtGBtDrj/yJzXkUU065G7OOxq5DeAffTP0ranhKXU4KzxUdtT0i8174YBla1+HVszJ90tlR+ILEH8as23xB8PWqAWngfTYdvQLDFwPT7tebR3MJ6nGTjBFPzGwGGU10xwNB9Dz6mLxUdz0qX4t3yRLFZ6WkKYwV8zAHPbaKz5viZrDsWWxtAefvM5/rXEAAjPQ9DTguK1jgaK6HJPHVnuzqZfiH4kcjZJaxd8LDkH8yaqy+OPE8rc6jt/3YUH9K5/aAcjrQU5zjqa2WEpLoZvF1H9o05/EmvzMGk1i9JAwNspX+WKpTXl7NzNeTy5/vys38zUWw9AOlCgn+VaKjBbIzdab6iN831pGGB15pwHNLgYp8iM3JjFGck5HpTj92l2H1FOA59aaiJyIwNo9aTrUuMgn8qQLjOO9OwcxEyg/WjkDpU2z0+tByeAvtTsHORHcOw9KQLlTnipsfLg0hXnilYOYiKEnBOBikK7cYqbbkHnHOaNvvRYfMRBD3XFIEBHNSnJ4oK8jmiw+Yrlct8y4oKfJmrGARyMGkdSEHelYfMVyuRUltbS3VxFbwRPLNKwSNEGWdicAAeuaeEz2r339mLwCsh/4TTUochSU05GHGejS/hyo/E+lcmMxMcPTcmdOFoyxFRQR33wQ+HFv4K0Lz7tEk1q8UG7kHPljqIlPoO/qfoK9JHAoRQq9KyPGGvWPhrw7e63qUnl2tpEZHx1bsFX1JOAPc18TUqSxFS76n2tCjGhBRRw37QXxIXwP4bFtp0inXL9WW1Gc+SvRpiPboPU+wNfGU80jSPLK7SSSMXd2OWZjyST3Jrc8c+JdQ8XeJrzXtSb99cN8kYPyxRj7qD2A/M5PeqPh3R73XtatdLsYjLPdSrFGvqx/kB1J7DNe/hcOqELvcyrVUjoPhZ4E1Xx54gSwsVMNrHh7u7ZcrCn9WOOB3+gNfavgrwxpPhXQYNG0e1W3t4uSTy8jd3c/xMfX+lZ/ww8Iaf4N8MW+j2SqzD57ifGDPKerH+QHYACutwAM14+Oxjqy5VsaUKX2pbieWvXAprOFGKGk2jk14F+0H8YobCKfwx4TvBJfNmO9vYmyLcdCiEf8ALTsSPu/XpzUKEq0rI3k0ih+0r8WZIDP4N8NXe2Q5TUbuJvu+sKEd/wC8e3T1r5rP8ulSzsXJycnrkmoWyB8uK+nw2GjRjZHNOZDdNv8Ak5AqqEzwB3wK19K0u+1jUoNN0y0lvLy4bZFDEuWY/wCHqegr6q+CHwN07wq0OteJVh1HWxh448boLQ/7IP3n/wBrt29azxWIjSWpFO8noeTfCf8AZ+1vxKYNU8SGbR9KfDLDtxdTr9D/AKtfdufbvX1J4N8DeG/CVgtnoOk29lHj5mRcySH1dz8zH6mulQKo4FUdc1rStDsXv9X1C2sbVPvSzyBF+nPU+w5r5+pXnWlodagluXFRE4p+B2r578bftBxy6imi/D/S5dWvp5BDDcToQjOTgbI+Gb8dor2jwHaa9a+GrVPEuoLfas4MlzIiKqIzc7FAA+VeBnvjPes6lCUFeQKabsjJ+KnjvT/AOhx6nf2lzdedL5MMcIHzPgnBY8KMA8/oa+HNavZdS1e81GZVSW6nedwv3QXYsQPzr7Z+P+l6dqPwn1z+0XWNbeD7RDIf4ZUOUx9T8v8AwKvhDUb0RfulJ8zIxx0r18tUVTciKmpbWGWeaOGCNpJZGCIi8liTgAe5NfdHwS8Dw+BfBFrppRTfzAT38g/imI5GfRRhR9M96+cf2TPCh8R+Njrl9CGtdFVZVyMq87ZEf5AM31Ar7G+7HgdayzCvd8qJpRu7nmP7Q/jW68O+FU0XRTI+ua232a1SHmRVPDso655Cj3b2qH4D/C+38B6OdU1WOJtduo8zyEgi2Tr5Sn/0I9z7CtnQvB2fHmoeO/EbxzX+TDp0ZbMdjbLkAjtvblie24j1rx39oj4yf22k/hPwjcn+zsmO8vYzzc+scffZ6n+LoOOvLTi5Llj8zW/co/tBfGR/EE1x4V8MXJj0ZCUurpGwbwjqoP8AzzH/AI99K868OfDnxrr9kL7S/Dt5NasMpM4EaOP9ncRu/CvafgJ8EIIIbfxJ4ztBNduBJbadKvyQjs0q93/2TwO+T0+iPKjWPaFAAGAAOla/Wo0PdgDjzH5veIbO+0zUZ9Ov7eW0uLdyksMi4ZWHYiszJ4756V6f+0ndx3nxi19o1TbFIkGQO6RqD+OQRXnuj6Xe6vqcGm6ZaTXl3cPtihiXczn6f17V6Cm3FSZmlYpOGZznAHSvZPhB8Adf8VJFqniEy6JpDYZFZP8ASZx6qp+4PdvwB617F8DvgdpvhaODWvEsUGoa4QGSMjdDaH/Zzwz/AO127ep9vVFUV59fGW0gWo3OR8B/D/wt4Ksfs+gaVHbuwAluH+eaX/ec8/gMD2rqkRRxgVU1nVtN0axk1DVL+2srWP7808gRB+J7+1eLeM/2hrH7SNK8C6ZJq97K4jjuJkZYmYnACJw7n/vn8a41CpVdym1E93VM+grjvit41tfAfhaTWruyubwGVYY4ocDLsDjcx4UcHnn6Vo/De18S2/hmB/Fl/wDa9XnYzThVVUg3dIlCgDCjA+ueao/Gez027+F/iGPVCqW62Mkm5v4XUZQj33BcUqUV7RJjb0Pgrxzqlx4g8U6nr1xBHDJqFy9w0cf3U3HOB6/XvWfoOl3ms6zZ6TYx+bdXkywQr6sxwPwq7qAUW5yACema9o/Yv8HrqnirUPFlym+DTE8i2yOPOkBy34J/6HXr4i1JGcXdH058NPDVn4Q8G6d4fsgPLtIgrPjmRzy7n3LEn8a8t/aZuNX8Va1oHwz8PoZLm9c3l2P4Y41OFZ/RQdzH/dHcivcwoRcCsqz0fStL1TUtfdY1vLxVNzdSkZWJB8qZP3UUAnHqSTXjxn71zS2hkeF9G8OfCz4emIzR29lYxGe8u3GGlfHzO3uTgAfQCvjP4s+Ob/x/4vn1i6V4rRMxWNsTnyYQeP8AgR6k+vsBXa/tGfFJ/GmpnRNGmYeHrOTIYZH2yQf8tD/sD+Efj6Y8cchFyTgDmvXwlDlXPPciT6DdQlMMGFHzMMD/ABrBlTPUVfu5/MkLEgDoBVZimMl1/OtasuZglYhKe9OhH7zp39KczxDq4ogkjaYAHnNYS2KR+hn7IBB+A+i4Of3k/wD6NavXa8e/Y8OfgTpPtPcD/wAitXsNeHU+Jmh8Tft68/FDS84H/EoTn/trJXzjgYz1NfR/7eu7/hZul8cf2Sn/AKNkr5xIOPSvWwq/dol7iAYCjPU/lRtwxAPHWntyAOPyoCZ5zXXyiAJkc8H3pQNxxnHPWnZpV3bjx+NaqAm0BTL5yDx0pyJ2H15pyphfve/NPUZXI4xwRV8pn1GEK3Y05Aufen/cGASfTNLEgI5pqIrCxrtTnGSeM0oXLmpWAJzg8D1qPqc85zWiiQ3qBHfOBimFAT94fhUxjwMKM5NKqHn2NOwrjFUgkr3FPTPpjt9acEAx3B/SlIXb1HHpVWFcRjjC+vFK0aj7uTTttPUACqjG4rjMEBQp4PBpwB7inD0oyG6GtVAdxuNx6YNSr93p0pnVmpefStEgTHLkdaG2+maFJC9M0YPUimJsUqP8mnId31pg4T8aUe3HrTsCZJxltwJpnYcU7ByD7U1sZosNyQyQ4GQM4qI5yDipZCTgD8aao45A9Kqxk5jQpIBI4odRj6U8naMGomc5wF/GpdkOLbFBBxnimO+MgdM8U0nIwKVV3E54ArNzuXyoZu/ebsHB4NPXA4OaGVTj5R+FOCds+9SotjuKuelKP9gke3agr6U4H/OK15RXDbnG4UoGDx0xTlbIzjBpNpHTvT5RXED4HJpyyFTlSR9KMc+1IDkkEdKpIiSjLdFmK6mUg7s/WrKX4H316d6oYz0NJtOa2i2kcFbA0Z9DWinik4DAHPerCoM8YIrDAOcdKesssTfu5GH41ftDzKuV2+FmvsoIUds1RS+lQDdhqni1CFmIYbapTTPPqYWpDoTbefwp3NKssbY+YGpMAHpnNWjnd1uMIHbilCDvUoXg85zQEz7YqkZ3Ih0/+tTVA5BqUoOewFAXJwBQHMNwuDkYpAq8k1IBkilIye1OyFcideBggU3BByDU20kAGlCjBpWDmK4UnORSBTuzg1MvzE5/CgLx7UWHzERQ4U5FKQf1x0qTgjrjFBXJFFiucixySB7fjTSD6H0zVhRngAClwOmBn6UNaC5zU8C+HLjxX4pstDtsoJ2/fSAf6uIcu35dPcivtjRNOtdL0u206yhENtbRLFEg/hUDAFeLfsp+HEj03UfEksf7y4l+zQkjoi8tj6sQP+A17yoAFfF5zinUq8i2R9hkmG5KftHuyKU4FfKX7Vnjk6v4gTwlYzE2OmPuuircSXGOh9QgOPqW9K+j/iJ4hh8L+DtV16XB+x2zOik8NJ0RfxYqK+Ar25lu7mW5nkZ5pXMkjseWYnJJ+pNTleHUpc76Hr1p2Q1YzLII1PJr3D9lDSrWbx3eXUgBksrEmEHszsFLflkf8Crx/S7fbD5jD5n6ewrpPB3iLWPCOux6zo8qJOilGR13RyIeqsPTgehBANfRV8NKdBqO7PmqmPj9YSeyPuqNQqCqOtapZ6TYzX2oXUNpaQrukmlcKiD1JNfO11+01fwWfk/8IpB9uC/e+1nys+u3bn8M/jXkPxA+Ivibxvc+Zrd7/o6nMdnCCkMZ9Qvc+5JNfMU8tqOdpn0ka8ZRTieofGv46XGpNLongu4kt7HBWfUFBWSf1EfdF9+p9h18FeQMM7uetMLA45HPanxxDO5sH2r3sNhY01aKOSviowWoqgvyRx2p9tbT3FxHb28TyyyuEjRRksxOAB9TT4wD2r3X9mH4eXt74ht/GWqWYTTLQM1mZRzPN0DqO6ryd3rjHQ1viakcPTbZxUasq8z1D4F/C608D6WLu7RJ9cuUH2qfGfKHXyk/2R3Pc+2K9OkZI1yxAAGSalwFXI6V8zftP/Em6fUpfBOjXDRW8QH9pSxtgyMRkRZ9AMFvUnHY18rGM8XV1PbilTidN8U/j9p2iyTaV4Sjh1K+TKveOc28R77cf6wj2+X3NfN/iPxFrvirVPtWsajdaldO2EDtnGTwqIOB9AKxtrzTJDDGzyOwVFQZLE9AAOpr6n/Z5+Dq+HvJ8UeKLdW1gjda2jciz/2m7GT/ANB+vT0ZRpYSOm5k5OehZ/Z4+D48LIvifxDCra5cJiGFhn7Eh6j/AK6EdT2HHrXtjssaZ9KlQAL0ryf9qDxXceGfhxJHYSGK81OYWkcithkUgmRh77Rj23V5PNLEVbM2UVBHlX7T3xMGt358H6NOH06zkDXkyNkTzDog9VT9W+gr541aESyI64DYOc9xWi7Ko3HpjvUvhjQdZ8Wa2ml6HYzXt1J91Ixwi/3mPRV9zX0KhChS5TBybZ9ifsp+GU0H4R6fcNHtn1RmvZSRzhuE/wDHFB/E16bqd3aafZzXd7cRW1tCpeSWVgqoo6kk9BTPDllHpmiWOnQgLFa28cCgdgqhR/KvEf2z9WntvDWi6RE7LHe3TyzAHAYRqMA+oy+fwFeBGPt61joXuxPPvjp8YrvxRNPoHh2eS30FDtklXKve/XuI/Re/U+lXf2UfAsGv69N4n1OFZbLTHC2yMMh7jGQT7IMH6keleFySDGcEV9cfsd6hZ3Hw0mtIQBcWt/IJx3O4BlP5cfhXo4pKhStEyi+Z6ntgjVF4/Gue+IPia18J+FNQ126wyWkJZYyceY54RR9WIH410p5XNeAftW6R401+DS9L0DRLy/01Xae5a3AYmQcIpXrgAsc4xz7V5FBKU9Toloj5j1OW81zWJbuQNc3t9cGRgoyXkdskAe5PAr7A+A3wotPAujreXsSTa9dIDczYz5IP/LJD6Due59sVwH7P3wa1nTfEFp4m8VwpafZSZLaxYhpDJjh3xkLjqB1zjpivpgALHn2rtxeI2hExhG71GLtTA/KvGvjD8dNI8MefpPh0w6trC/Kzhs29ue+4j77D+6PxI6VzP7UPxTu9Pu38EeH7hoJSgOpXMbYZQwyIlPYkEFj6ED1r5rSN5pFjQFiSAoAySfQDvV4XBJrnmVKfRG3458ZeI/F16LzxBqcl0U/1cQ+WKL12oOB/OvoP9mD4Vy6XBH4x8Q2xXUJk/wBBt5F5t4yP9YwPR2HT0HueKvwF+CBgltvFHjO1/fqQ9np0oyIj2eUf3vRe3fngfR0apGmFFTisQkuSBMI31YmVjQD0r5a/al+JsOr3X/CGaJcLJZ20m7UJkbKyyr0jB7hTyf8Aa+leq/tKeLrnwp8OriTTpTFf38os4JFOGj3Al2HuFBx6EivimU7V3sTgck5q8uwyl+8kOpLoiLUojMkaKcfNnJPFfbP7LXh0aB8H9KLAedqBe+kIHXefl/8AHAtfGvhjRdW8XaxBpWi2c9zPKwVVjXOxScF2P8IHcniv0V0DTodK0Sy022x5NpbxwJ9EUKP5UsxqJuyCmgv547aB5pnWONFLM7nCqoGSSewxXyR8d/jLceK3l8O+HJJIdBVts03Rr0j+Uft36n0r3b9pe+ksPhBrbRMQ0yR2+QccPIqt+mRXxKcA1WX4eMlzsc5W0ByDWBrN2ZXEUL/IOuO5/wAKsaxelma0gBY/xlf5V6P8BPghq3j+4TVtXWfTvDcbczAbZLoj+GLPbsX6Dtk9OnFV1H3UKCvueR7HxkkD6moyEBO6UfhzW18TNNt/Dvj7XdBtdxgsL+WCLccnYrfLk/TFcy0/oK5PaI1dizJKuMID9TUKOxkHzHrULSk9sClhlYSA4GaznVViU9T9Hv2Ms/8ACg9IBIP7+45/7atXsteK/sUv5nwA0lj1+03P/o1q9qry5as0Z8V/t6Ln4laV6/2Sn/o2SvnAjA4FfSX7eIJ+JGk9f+QSuP8Av7JXze4Ir2sIv3aIk9QxgdaACMZ596ET5OBT8E8qM46g9q7kiGxVBLkEc08R89KcqehApw3MoXd/9etEiWAUng8kd/anKuV6j1NOCk+9CoS/I7Z6/pVpEMRUXPB7/lU0a4ByOf1oij6g08KTwTgj+VVYd7IR0QkD26+tIFxjFSKgPXjmnbVz94Y+tUkZt3IwmPvGk256HGO3rUpGRk9c9KbtBOcd6rlEMZCe/WnBTjinkfIM0obHatFAYwqDjPOKOT1FKM4JpOSc49qtIQ449OlJhduQMc0pB6cdKUKKtIBOM9KfnDcChASucUoA70+UlySEGcc8+hpRnp3p+Pamg4PSnYlzAgj0NORRnnPFCgk4xT1OeNuKfKRF3YMADkHrUbDJ5FSSbSckmoJZAo4NDsi7N7CP7cVE0gFNkd27Y+lNC5zWbmNUu49yS2P4cdKTHPfmjHGKVRlRgYIqdzS1hrDn3+lBPQb6djLU7b3xTUQbGqADnFP+9160nG4c9O9KpwTVpEMNnpTsc4IpQD0zTlBzwa1USW7DE644HvThwecHnvTggHJoYEjAFDRm6qB8MzMoVRnoDwKjwc/jT1X5uhANLtUHOKEjJ1UhuTxkHrT+B1/nTVweoqQKCeDVJMylVGsM4HTtQSQ4xz60/bnpTOSeR3pmMqo7k9vwpCMnk9/SnEOcEg8cU5Vz1GCOfrRY55VbkZB7ZqWO4njxhzj65oVQckj86ZtJ9apHNUUXuaEeosT+8iXj04qzHeQP3K/Wsdu2BQM+4quZo5J0Is31aNxlGBp2M1hq7KPlJH41ZhupVGScj0NNTOadDsahTrkYpAv8qrRX5/jWp45436HBrRSTMJQkgC5PWlK5HrUg2n7pB7cUBSOe9MzbsRbe1OKinjr+lOWMHODTDmIWUFyTx7Um0epqRgOBQyHjgUD5iMJ/jSxgAkk8Dmn7Pyq1o0STataQyJuSSeNGHrlgKiq+WDY4e9JI+xvhfo6aL4C0ewVNrJao0gx/Gw3N+rGukyRUdptSBUAAAGAPSpDzX5xXk51G2fpWFgoUopHhf7Y2qPaeALHTEz/p9+ok/wB2NS2P++tv5V8p2UQnuo4uoJ5x6V9MftpKzaH4bOMqLyYH67F/wNfOvh+EebJMe3yj2r6TKIXgrHFmdb2VJs01XbgLwBx06U5l9P5VKevyr7UFa+lsfCubbuY+pWDzEzQ4DYwQT1rMNtcqwRoinbJNdPs4IxkdqqXse5dw5Yc4rkq4eLdz1cNmVSMeQzIrZUXc2C2KRkO7gHJPStfRNI1LW9Rh03SrKa8u5jhIolyT7n0A7k8Cvpz4OfBaw8MGHWfEPlahrQw0aD5obU/7OfvP/tHp29a48Vi6eGj5nZh6VXEyu9jjvgl8D3vUh8QeNLZorbh7fTXBDy+jS/3V/wBnqe+Oh9jv/H3hfSvGum+CFmBv7giIRwqPLtvlyitjhc4wAPbpXG/G34yWvheOfw94bljudbIKzTD5ks/r/ef27d/SvmGx1S7ttZj1gXEj3sdwLgSuxLGQNu3E9Sc15Sw1XGJ1Kui6HqKcMO1Cnv1P0DOTF9OtfHnjD4SfELVPiPq0dvo7ywXV9LNHfySKsBR3LBi2c9D0xnjpX1L8P/FFj4u8L2euWLDZcJiWPPMMo4dD9D+mD3roCinkAZrx6daeGm0ketb2kdDyT4P/AAX0XwU0eqX7jVdbA4uHTEcB/wCmSnof9o8/TpWr8YfibpngDS1ijCXmt3S4srIHkk8B39Ez+fQeoh+NHxX0rwDYtaR7L3XZUzb2YbhM9Hkx91fbqe3rXyx4QvL3xZ8YNFvteumuri91WBp3fvhwQoHQLxgDoBXTClOtepMlJQ0R9z+HPtg0W0XUJvOvBCgnkC4DSYG447c5rzr9oz4c6n4+0DT10a4t473T5nkWO4YqkquACNwBwRtGO3WvS7HHkqasEgr7VwRqOnO8TblvGx8n+FP2ZdXu7tJfFutwWtqDlrbTyZJG9i7AKv4Bq9vitvAnwg8JPMsVro+nJjc2N0tw+OBnlpHPpz+ArR+J3j7w94B0X7bq0+65kB+y2URBmnI9B2Hqx4H14r4t+JXjbW/H+vNqmsyhUQFba0QnyrdPRQepPdup/ID0Kaq4nWWxi1GB98aXdw3llDcwsDHLGsiH1DAEfoa8T/bA8P3+qeGNK1aztpLiPTp5BcCNCxRHA+c45wCoBPvXSfs4eI11/wCGGlFpN9xYp9hnGeQ0eApP1Taa9Q2gjmuTmeHq3NF78T8//CXgLxX4x1WK10XSLl4sgNcyIUhjHqznj8Bk+gr7N+E/gTT/AAB4Wi0qzbzpmbzbu4IwZpSME47AYwB2H412QQKMAYHpXHfFjx/ovgPQTd3ziW9lUi0s0bEkzevso7t/M8VpVxE8TLlSFGCjqzZ17xl4Y0C6gtdb1yw0+adS0SXEwUuoOCee3asTV/in8PdPhaabxVpjgDO2CXzWP0CZNfFfjHxNq3izX7jW9XmElzN/COEjQdEUdgP/AK/esbzGHJPau2nlkbJyZLqu9kfV3hn452fiT4nWXh7TrDyNLuA6Lc3HEskgUlcKOFU4I5yTkdK9zVvMiGPSvzj02+udM1S11Gyk8u5tZVnif0dSCP1Ffe3wy8WWHjHwnZ63YOMSrtmjB5hlH3kP0P5jBrmxuGVKziOErvU+aPit8KvHur/FjW59P0Oa7tb67M0N2XVYdjYxlieMdMdeOlevfBj4LaZ4NEWraw0Wpa2BlZMfurc/9Mwe/wDtHn0xXrxCk5xmvPfjB8UdG8A6c0bPHd6zIn+j2Ktzz0Z8fdT9T29oWIq1UoRG4pO7JfjL8RrDwBoW5Nl1rFypFnaZzk9N7jsg/U8D26nwTBqcXhTTV1i4e41E2qNdSMOTKRluB6EkD2FfGPhbUb3x18XtFuNfuDczX2pw+cT02hgdgHZcDAFfdCEbB79ajE0fYpLqEHdnl3x/+H154+8MwWmn3cVvfWdx58HnZ8uTKlSrEDI4PBweleOeEf2cNevLsf8ACVana2NmD88VkxllcezEBV+vP0r6wYDr2rlviH4z0HwRoj6nrFyEJBEFuhzLO391B/M9B3p0MRVUfZxHKC3ZT0bRvBXwu8LSG1htdH06FQZp35klPYsx+Z29Bz7V2tlcR3NnFPA+6KRA6MO4IyDXwh8S/iHrnjrVzc6hIYbOMn7NZox8uEev+03qx/QV9cfAHXI9e+FOiXAkDy28AtZueQ8fy8/UAH8arE4aUIKUhRkr2LXxd8NSeLvAeqaFEyrNPEDAzdBKrBkz7EjB+tfDuo+HtesNRfS7vRNRivVfZ5Bt2LE+2Bz9Rmv0PfBPPSmNAjHJFThsY6KsOUbnzD8GP2f57uWHWvHMIhtwQ8elqcM//XYjoP8AYH4ntXtXxT8baL8OfCgmMUP2kp5Wn2EeF3kDgYH3UXjJ/AVY+Jvj/Qvh9ohu9QkEl1KD9lsoyPMmb+ijux4HueK+LvHHinVvGPiGfW9an3zScRxr9yFOyKOwH6nmumjSniZ88tiHaOhheKWXxBr19rWpQxPeXszTTso2gsxycCsZ9FsG5ETr9HNbDj6Ux8hfuivVdCHYjmZiXGg2ZAMbyofQEGqj6Bhv3d116bkroXH19abGAZAcVzVcPBrYFJ3Puf8AYyt/s3wH0uAsGK3NxyOn+sNezV5H+yOAvwT04AYxcT/+hmvXK+bqK0mjrR8X/t3kf8LI0leM/wBkL/6Nkr5xKgjrg19B/t6XCJ8U9Gjc4zo6nPb/AF0lfPueAR09u9e3g2vZpES3E2jIA5NSIvPA5PNKg4IPXoDUiLknLA13xRk2IqlsnPTnrTwCwBA+o9aVSCSPbt3p4xgc8VokZtgVDZAXJz2p8asTjoB709QAQcdakK8Dnr+tWoichhHzc89qcAApJPTvRjPeggY9yea0UBbik5wBz70hAApyjjjAowR1OTVqIWDBI5FNwATjin844NNPQZrRRDQFY46UHJ68UoHakA5qkgugBPQigHC+9OAyKUgKQeOaaRLkkNC85xQV54+tKaVeRkGqSMnIVRlAScUoADZAzikJ4oXB6j3qjJu47OKU544HSmn6daM5Az2obsVFNjlAA60juFXJH6UwvTXbkVlKp2N4Uh0sjE8VXkAJGRUgGRkmkwCcEc1F2zVqxGAWyMfjSKoU5x7VKQAcjIo5YnOKaiS5DWyMZFHpSn19aAvBxVJCchSCecUjBsYBp6kc/lSouTgjkVaRm52GEcjj2pyjJwDzTivr9KXG0ZA6+lWkZOoIpI6GnL6YoC4HPen7RlTn9aZi5ibQcY60rBs5yaUKB0PvijcNxBH41VjCVQQk7cADj0pGDAjJVsrnjnH/ANegYOfakI7569vahIwnVFVecNgc9af8uOAaTr/+unIcfwinYwdUAccdx0pi43kknJNKTmQn2pUXI6d6EZyqDmAOAD27U3gdFp5XHQ0bSFBqrGLqXFyTyOaB8ykng5pQBjOMU7bwOaEiHMj2gkk8gU4Lt6AHNSCM+vNKsYNOxDmRgfNmlwxPXnNTCPChvXpS424wRz3osQ5kaj0X86cMjkVJt7UBfbGKdjPmBHkHIY1YS4lGM81Dt7+tOCsATg4p6kNJltbpWOHTHvU8bxtwG/Ws3pilxyCGKn2quZoydNM0tuTkHNLjA+tUI5ZF4BNWFmIA3DrVKSMnTaLW0scKOak0udLbWLOeRgEiuI3c47KwJ/lVdZVbgdaRwj9eQamsuaDSHSvGabPui0kWW2WRGDKwyCO4qcVxHwU8QR6/4B06bzA1xbxi3uBnkOgxz9Rg/jXcdq/OsTB06jiz9LwlRVKUZI8z/aH8IXPi/wCH1xb2EZk1CxkF3bRr1lKghkHuVJx7gV8e6CCiTKwYMJcMCMEe3tX6FOgYc96818c/Bvwt4kv5tUjSXTNRm5lntcBZW/vOh4J9xgnua9HLMwVB2lscuZ4WWIpNR3PlfAU5HcdKEUPkYOa91b9nyTzML4lBjHrZfN/6HW/oXwH8N2rq2p3+oX7Dqu4RIfwUZ/WvflnOHSvc+VWTYmWlj5uhtZ7q4S3t4JJpXOESNSzMfYDk16b4H+BviHWnjuddJ0aybko4zcMPZei/8C/Kvo/w14V8PeHoQmj6Va2nGC6Jlz9WPzH8TU3iDWdL0HTpdR1W9hsrSIfNJK2B9Pcn0HJrysRndSq+Wkj18Lkcaa5qzMTwZ4K8O+CbFoNFsViLj97cOd0suP7zf0GB7V418c/jaV8/w34Muvm5S61KI9OxSI+vq/5etc/8bPjPd+JI5dF8PPLZaQwxJJ92a5HocfdT/Z6nv6V4dIxZ85xinhcG2/aVtztlNNclLRE8kruSxJJJ5JPJPrmiH7+TUUYY4OMD3qwo7Y/oK9yEL7nFUqQpaR3O8+FvxH1j4f6hJJaIt5p1wQbqydsByONynHytjvg56GvQfGv7R99d6abbwvpT6bO64e6uXWRo/wDcUcZ9zn6V4QD8mMduaqTLySp4zXNiMBRnPna1OnC4mVrE2qXt1qN9NfX9xLdXUzl5ZpXLO7HuSetGl382l6pZ6nakC4tJ0ni9NyMGH8qqO5Ay2APSo8Fup4xke1ZShG1kd6bPtbwz8avAGo6NDeXGv2unSsm6W1uSVkibuvT5uehGc1x3xD/aN0+2gksvBdqb24OV+3XMZWFPdUOGc/XA+tfLq4VR8wNG8Y57Vwxy6kpXZt7R2NXxLruqeINSl1PWL6a9vJT80spyfoPQD0HArJbOOoWlLHGTj0FRSYPyZ5rs5YxVkZ2b3O++DnxIv/h/rckyQm70u6wLu2BwWxnDoegYZ/EHHpj6U0349fDm5tVnk1uS2fHMM9rIHX2+UEH8DXxcAWGQpFKqODn1rkq4KNZ3ZUZ8p9SeN/2ktOit5Lbwhp0t1cEYF1eLsiX3Cfeb8dv4186+Jtd1TxFqs2qaxfTXl3KfnkkOePQDoAOwHArJXPofripo4yTyNox3row+ChS2RNStcgllCABSCTUJQs2VDEnnIrR8lOu0Zp2MAYFdXsGzn9uimtpKwHy478mu4+FnjvXvh7qElzprJPaz4+02cpPlyY6HI+6w/vD8c1zCZPt70r+meaJYSM1Zi+sNHtHiz9pHXL6wNp4f0aLSZnXD3Ms3nuvug2gA+5z9K8Qv7m5vryW9vbiW5uJn3ySyuWdye5J5NMaL5iVGfrUsUPQycDHQd6xpYONN2ijR4hNbl3w1qE+k67Y6taAefZTpOmehZCCAa+xvD/xm8A6lpkd3ca9b6bKVBltrrKvGe46Yb6jNfGoCqAFAAA7VXnyQSB3oxWXwrpX3IpYhqR9RfEb9oXQ7C2e18IRtql6wwLmVClvH74OGc+3A96+afE+vax4l1WTVNbvpr26fgvIeFH91R0VfYVnbsjFOCE44wKzoYKFL4UdMq1xFGSAo4r0f4MfEu9+HmpShoWvNJusG5tg2GDDo6dg3bHcfQGvPjheABSHOORiuuph41I8skZqep9i23x3+G89kLmXWpbVyMmCa1k8we3ygg/gTXD+Ov2kLVLVrTwbYySzNx9svI9qJ7qmcsfrgexr5rkVsnuKjHB/+tXnRyylGVzdVG0afiTW9U17U5dT1e+mvbuU/PLK2T9B2AHYDgVmckZzg08IxGT0oIx2rvjTSVkZ3uRAsepJGaVxzwTj3FOTrSO2BWcojsRyqox8v60xFAlHzdT61LKXI4AOPWmgkEZ2g1zziM+7P2SWVvgxYlRgfaZ+P+BV65XkX7I3HwYsh/wBPU/8A6FXrtfK1tKjOuOx8M/8ABQlJI/ihoUhX5X0YBT7iaTP8xXzhbXssJwfnT+6e30r7J/4KF+GZLrw94c8WQRlhZzyWU7DssgDIT7ZRh/wIV8Xng124aT5bomR6L8O/B+teNpCdHgUwoxWaeRsJCQjON/dQQpAY8Z70zw7oF5rviW08P2LRC6uZfKVi2UXAyWJGeAATkZrF8DeN9f8ABsWpLoM8Vu+oxJDNI0YdgituwM8YPQ5B4NegfA611G707xR4ntjaw3UFo1lYvLKsMa3E3BIZsAbUyfxFerCtdanPJO5xWv6Zd6Hr17o9+FW5s5mikx0JHcex6j2qTQ9NvtXu3tdOh86VIXnYbgMIilmOT6AGvTvHGgST+O/AusajFbXP9sm2tdQEciyxSTxsqPkrwQy4/I1seDtZ0y0+JOsaFo3hjSLaysbW/VZJId88rJG24s5/hJGNowAvFaKo7Enltl4ell8H3XiSTULSKOC5S2S2Z/30rEAkhewAOcn3rJfgda9Ig060174bW14lhZWl7qHisW3mQxBREjxDCL3CAnIFL461PwfpWp614RHhC3W1sg9tbX8RP2z7QnHmOxOCpYHK46dK2hVCx5mXBrXj0G6fwgfEwliNot8LIx879+zfn0xiuwudF0uL4k+BLBLCD7NeWenPcxbflmZz85Yd8962Y9Ghu/C7+HYsQQXPjtrVcfwJt28fQGtnWS2BHkXTgVLY29zf31vYWULTXFxIsUSL1ZicACvV4JPD/ijxNq3geDwvpthbxR3CaZdwRkXMckIYgyPn5w205B9a8osLy7029g1Cyne3urdxJFKvVGHQitIT5loB1fiL4e6tpGm3l3Hqmk6lJp2P7QtrOffLa5OMsCBkA8EjOKTQfhvrusadaXUd9pFtcXyGSxsrm7CXFyozyi++DjJGa0dPM3hfwVq+saw7f2x4mtWt7S2f7/kM2ZLhx2BxhfXrUnirzLb4l+D0gJBjsdKEWO3C9PxJrPnkSzz24ikgkeGVWSSNirq3BBHUGtTxPoN54fvLa1vJIpGuLSK7Qxk42SLuAOe471f+Jqxp8RfEYQAINRnwB/vmvTzZaXqnxC0iLVbGO9tv+EMSUxPxkpASCD2PHB7VcqrjZma1PEbZPNmjj3Ku5gu5jgDJ6n2q3r1gdJ1i5043dre+Q23z7aTfE/AOVbuOa7zUH0rxF8NzrX9gaZpt7p+rwWoNlEY1lhkUna4JO4gj7x5rd0nRdBi8deNdP0/SNFutXtrpV0fTtQfbA0eT5gUEgFgMYBPel9Y8gcTxgE+xo3D/AOsK9Xt9J03/AITHxBqmseDV0pdE0n7W+kO5MUk+Qob/AK5knOAccVieIF03xJ8OJPFUOj2Ol6jp+pJaXC2SGOKaORSVOzJwwIxxWixCZnyM5fQtCvNZsNXvreWFI9KthczhycspbGFwOv1xUnhfQrnXm1BbaeKJrCwkvX8zPzImMqMd+a6z4L/2adE8cDV2uUsDpKee1soaQL5oHyg8Z5HWug8AQ/D9LHxTN4fvPEE14vh+63JewxomwqASCpJznH61m8Q02ilTPHN2MYOcikAJb8aaq+9SxqN456VfM5HRGKijQ1zQ7rSrDSb24kiePVLU3MIQnKqGK4bI65HbNZY2+1dz8SVx4V8BnHXRm/8ARz1r/C3wzC/g7UvFhsNHv7yO7WztItWuFito/l3O5DEBm5AAz6mslVSjdlM8uz70nBPUV7PJonhqbxlpGnvZ6JHP4h02e1ubexuFnhs7znypIyCduSF4z3NcNqekwaP8NbJry0QatqmpSsrsvzxQQDYQPQGQt/3zTVVMhkXgPwdJ4rg1S5bWdP0m10yNJZ5rzdsCsxUfdB7j9al8R+D9K0rS5L218deH9VmRlVbW1MnmvkgcZGOOp9q6D4Nppj+DvHkes3FxbWDWNsJpbeMPIg844KqSAecVzHiWz8DwWkcnhvW9ZvbwzKGiu7FYkCdzuDHnpxUe0lztXBpGP4p0W88O69d6LftE1zasEkaMkqSQDwT9azxkqK941Hw3b6j8QvHeuTWFjqFxpzWy2lrfTrFbmSRBl5CSAQAvAJ5zVGTwhouteI/CS3dro9ldXlxNHqllpN2kkTxxrvVgFY7CwUqa1VdLcykeMxfMwUHknFaevaNf6Dq82lapEsV1Dt8xVYNjKhhyPYiuuvvHOjXkFylx4O0a3uLeVJNKe1tgqx7WGUmBP7xSvr3ruvFS6fqPxH8U6vq+lWN1HoGkx3UVv5exbmV0j2+aRywBJ/DAqvrDT1RzS1PBzk9Pwo4AHP616xp0eleK9A0nXrrRNNsry38Q2+n3C2cPlw3MMmCAydMjBHuDUGpXVnc/EW58N6N4A0++g025uVt7aMkGSQHHmTOesYwTt4UcDtVrE9LGDTPLrWF7y/t7SJ1DzypEpY8AsQBn86ta7plxo2s3uk3MkbzWc7QSOhO0spwcZ5xXqWv6MH0DQdd1DSNF07V4vEEVpJ/ZToY3jOGG4RsVDAj64+tcT8VVA+I/iQD/AKCU/wD6GadOtzSMqnuo5peVGKcMY7eleq6PoGjXnivQdefT4F0NtCOpXkAXEZeBCkike8gXP+9XA+E7CPXvGmmabMBDFfXyRuE42qzjIH4ZrWNdNN9jmkmiPw9oN1rY1D7JLCn2Cze8l8wkbkTGQMDrzWSAcA5/Ovb9O8Q6df6n4x0Gy8M6Zp0VppN7FazW0ZWXYnykOf4s4ByehFWdA8O2+i+F/D3k6B4Y1BNQtkudSm1W5RJWDn7ke5htAXuO4rD63y7oHR5tmeFYzgjj+td3/wAK3C2FjPd+MPD9hLe20dzFBcysjhHGRnj8M1j/ABA0vT9I8W6np+mTJNYxTn7O6PvGwgEDcODjOPwrvPGnh3w5qdl4YutY8YW+jSnQbVfIe0eVioB+YFePUY9qurVdk11MacdWmeZeKvD+p+GdXbTNTSNZQgkR423JIh6Mp7g1nxtxnIr1ePVtL8Q+OI7fSvDlx4hh03SVstLjmT5WZP8AltMCQAvJ6+1aGueG4pP+ET1LV9B0rT7+41lbO5gsthgniypBKoSoPUEfnRHEuNk0EqSlseW6Do9/rVxNb6dEJZIYHuJAzBcRoMsefaluNLMOi2mp/a7RxcyPGLdJMzR7ccsvYHtXsPhjWbS0+I2uaHpvhnRbW0s7e+VGEBMr7ASdzE8qxGCvQDgVzmgaLpviHTvCTzWNvbNqet3S3PkLtzEoVhGD1wBkD0zQsU+bVaEOgraM4bRtDu9U03Vb62kiEemQLPMHJ3MpYL8v41lj72Oa9b03xFb6z4X8cWkWg6Zp3kaePIe0i2N5QlUbH/vdjnr1pQ/h7w5r2leD7jwxpt/BPFbrqF3MpNw0kwBJRs/KF3DAHpTjind3REqStueUgA4xwPr1p2w17H4U0TTLK417RtLttEv9etNTaKOHVgCJrZeAI88b89a858b232HxPf250mTSdsmfsbvuMWQDgHuO4PpWtLEKpJownScFc2dO+HhuPD+n6ve+KtF0yPUEaSCO6ZwxCttPbHUfrWB4r8O3/hu/jtr14JkmiEtvcQPvimjPRlNd/q2m6RqPw88GnVPElvo7RWk+xZbZ5fMBlPI29MY7+tQ2L+HdV1GwsYWk1bSvC+k3Ny7yoYxeSbt23b1Cbiox6A1zwrzTbeppKnG1jzCBfMkRARl2ABPua6aTwhcQ65rWkz6lYwyaTC0kjyOVExUD5UB5JOa6MXNv4t8Harf3Wl6daaho88EkM1lbiEPFI+0xsBwcdQa1vEVnbT+IviPLNbxSPDbI8TMgJjbegyp7HB6iqniZXtsRGkmjz218O3jT6Wt7NbafBqaGSC4uJAIwgJG5sZIGQR0rKeMqxXIIU4yDwa9g+2HUb74c2F7Y6fJby2pdl+zLzgyLj/dwAcevNc7ZXNh4W8Hadqv9i2Gpahqs0x3XsfmRwxRtt2hcjknvVU8TJrVEVKKWzOAYAdcCnYG3PHSvRtKl8Ox6PrnjWDQLaQrcw2tnY3PzwwyOuXbHccHAPSr1paeF7678OeJ9QsLLTrW++0QXNvgra/aIx8jEdkJIyPaqeLtuiPYX6nlaqD8wINO2kV6F8RrO9j0m2uL7w7pdo3nFY7/S2U28qY+4QpIz3BPOK4LaCK6KNT2kbmFRODsyPb6UZKncuafgkmjkjpgdK1Iudb8L/Hl/4F1Z7yGJrqymAF3a7seYo/iU9A4ycevQ+31V4N8X6B4t0tdS0LUI7qPA3x5xJCf7rr1U/wCRmvinH4cVnibUNHvRf6Td3NnIhyJbeQo6fiO1eLmWWRr+/Hc+hybMfZv2cmfoGGBxg07jFfFmifHL4h6ZAkH9qw38a9Ptlurt9Nwwx/E10A/aQ8YrGo/svRGYL94xS9fpvr56WWVU9D6tVotH1jkCs/W9X03SLF77U763srZPvSzyBFH4mvkTWPj58RL+Jo4tRstPUjk2tooP5vuIrzjWdb1bWbk3Or6leX8x/juJmkI+melbUsqlf3mRKr2PqHx5+0RoelB7XwvANYucf8fEmUt1Pt0Z/wAMD3r518b+OvEni+/N3reoyT4J8uIfLHH/ALqjgfXr6k1zLsSoP61b0/Trm8IZE2R/326H6eteth8FCm/dWpx168IxvNlcgsQoBJPQd6nhtmDbnHP930rp9O0y2tYs4Dy45c9fw9KpanbhZd4zg9a9aGHtqz5ytmqnLkp7GVt4PGKRyCAoHt0qeXGOmKqSPjOOtVOSiVhqcq0iSVgq4zzVRyWbGKcCWGSaYVZj8oIINcspuR79KhGmgbBHXH4UmQAT61KISfoKlS2UYJPWsvYyZs60IlTrQY5GXp+lXhCq+lPCKYzjgiqWHfUzli4rYpLC+RnpT2gXPOM4q0VIwRzSMpLDjvWyoo5Z419CMQjjAFM8tg3JzzVpSAeRwOOlJIMscLj2qvZoyeKbI9qgBc9adgc00qD1z1oOMAg01Ezde4uQDg5/OlXhgQOvagLkHcce9C43Ae9OwlVHgsowQM+o603qOD39aVhkYzgUwL8uBwRTSFKsg2gng459aeBJ0LUsS5bPpT88E+tDIVUaM5OQTjrzTd2T09qfIMYAIz600KTRYr24gHGRxTzyORjFJuwgUD9KUHPb2oaNoVhMe1DZ29vrS/NnrnmlYYOPU1LRvGpci2ntzSMvAqQoD70gUcnoMcCpaNoVNSLgtnb2pjY3Zx3qQg4xke1MJwT+VQzqjqMOd2cY9qac96ee1NbPY1nJGgrrlewNRohJxjvUp5HWrGkWNzf6nb2VpE0txcSrFFGvJZ2OAPzNc1a0Y3CKuz7i/ZShaL4L6azLtEs0zr7jeR/Q16tWH4A0CPwv4M0nQIiCLK1SJmH8T4+ZvxYk/jW4a+Nqy5ptnWjn/iL4U0/xv4K1TwtqYxb38BjDgZMTjlHHurAH8K/MXx94V1bwZ4qv/Dmt2xgvbKQo/Hyuv8Lqe6sMEH0Nfq1+FeXfH74NaD8VdGXzmFhrtqhFlqCpkjv5cg/iQnt1BOR3zdGryMTR+bZAIHauo1DxJat8NtK8I2VvNH5N5LfX8r4xNMw2oFA7Kvr3NWviV8NvF/w/1M2XiXR5rZNxWK6T57ef3SQcH6HB9QK5AxtxlT+VelGaktGS0ehfDvx9p+g+HrTSdUs7i5Wx1231W1aLHyBeJV57kAY7Zq/4R8daVZ/Ei78RXFvNJZXkt0JYVIEojnDjjPGQG/SvK9rL1BpoJB4yK3hUcSHFM9gu/FmlW3hOXQNDhvohDri6lZSzlSyose0B8fxZAPHFa9/4+8Nyz6l4hsfDt3F4n1O2eGd5J1a0id12ySouN24jPB6E14rZXs0R25yvvWzZ3UMzbQdrHseK7aU6cjNwsepaR488MQ3fh3XNU0HULnW9FgitV8uZRbyJHkLIR97eATgdM4yazLrxqH0iS3sraaG7HiJ9ZhnYghcj5VI9Qce1cbtyBx9cUu30xXSqcRWselS+OvDFvPqfiHRvD17aeJdShkikd51a1t2kGJJIxjdk5OAema4jwnc6bY+JNPvdYsnvbC3mWSa3XGZQOQvPGM4z7VnL0xSpw3StIRilYhnoHjjxB4B8RXGo6oNM8S/2vdKTE813GYY2xhRtAzsHQKO1SaH4y8MgaNquuaFfXevaHAkFo8M6rBOI8+UZQRkFfbrivPOD7elKDjNNUoWtchk+pzzahfXN7cHdPcytLIfVmOT+prtLXx1bQeIbDVTYTsLbw/8A2SyBxln8pk3j2yQcda4R3GOtM354HFaSjCSszLVHR6Zr0Vr4Hv8Aw+1tI8lzqFvdrKGAVRGGBBHXJzWvdeKvC+q+IPEF1r/h64ktNVuUuIZreRRd2jKMYViMEHuPYVw6yAZ6AelRySKcjrzWco01qXBSZ6Vf/EjT5fEUBj0q6m0FNJ/sieC4nzczwEk7i/QODgj6e9YPiPxJop8Lx+FPCthe22mm6+2XM17IrTTyAbVGF4CqK43ce/GTmnoQD+FZJQvobqNjpPCevQaPo/iPT5LaSV9XsVtY2UgCMhw2TnqOO1L4M16Dw/HrYlt5JjqOly2KbGA2M5X5jnqBiudXaCCDTnIOOQKtKP3jZq6vFp0eg6G9tCI76SKZrsiUtvHmERsR/CcAjA7AHvWV0pFCjjigtxVppdST0BvFPgXU/DWgad4h0TXZrrSbQ2wktLqONGBdmzggn+Kmab4o8LQ2Wp+G7rRtSn8MXdyl1bILhRdW0qrtLBsbWyMjBHSuBOB3FSIyjO6slSgO51t9e/214r0pPA/h2Sy/s9UFpDEPMnkZG3mWVgOTnv0AFX/j/q8Wp/Ee6t7WNI7fTkFqqR/dD5Ly4/7aM35VxunXt3p9wLmxu5rWcAgSQyFGAIwRkc1BIPMdmdizE5LE8k+tVyRTTRLZ1HgHxNo+i6frmm67pd3fWWrQRROLWYRuux93Ug98VHrl58PJtNkXQ9B160v8r5UlxfJJGORnKhcnjPfrXLsq4HI64pqqAfvCj2cea9zGU2ej3fjrTr3xZ4gur/SJ7jQtfjiS6tPNCyoYwux1bpuDAn8apyeKtI0fVdDuPB2imyj0mVp/Nu2Dz3TNwwkZR93GQAOmTXFocgZYAUp5HDAVqqdMwlOR2+ueJPCkei6pbeGPD93Z3WrlRdSXcyyLboG3FIcDOCe55wMVd1D4gaNP4pl1ZdGvXttV04WWs2skyjzAFVQ0RA+UjaCM96853Ac5FISCOop+ypW3Oac5djubnxfo1nHo2leHdNvLfSNP1JNSuPtUqtPdSgjrt+UAKMAe9VdI8WWVv4z1/U7yyuZNO1xLmGeKKQLNHHK+7Kt0yOK5AEYwMfWgAZ4xjrWip07WOeU5ncz+K/Dtt4YtdA0TRr2CC11aLUBLcTq7y7Rg7sAAE8AAcADvXOeLtTj1vxPqesRxNCl7dSTrGxBKBmzgmsobV79elC4I4IzThTpw1Rz1J1JbnqCX+oaF8DZdL1Gzktrm+vjDZGZSrtasEkkK5/hLKvPQ7q870u6n07U7bUbU7bi1mWWNj2ZSCP1FOvL29vhF9rvJ7nyUEcZlkLbFHRRnoPaoCpwCSPzp04winruROUm1ZbHqaeOPBcJ1vULPw7fwaprVnNFcMZlaKJ3HOwf3S3Jz6Vnad4q8LXumaPF4s0a+u7vRY/Kt2tpVCXEQOVjlDdh046ivPs57j86VcE8Ec1PsKXcTr1exd8RXqapq93fxWkFmk8hdYIVCxxjsoA9BV/xbrceuR6KIreSE6dpkVk+8g72QnLD2OaxgqngEHmntE6qMoQO1aJU1bXY5nKprodB4C16z0OTU7bU7W4nsNUtfs1wbZwkyDOQVJ4/Ct5vF+gwadounaTpF7DbaVqy3ymaZWeZeN27HAYkdBwBiuBwCoxinxq2cKP1pSp0pS5mxxq1VGyR2Wj+LLGz+IWo+Ip7OeSyvzcLJEjASKk2eh6ZGarp4qi07T9GtdGhnVtI1Ka8t5Lgg70crtVgO+FwfrXLAZH49aURknhabp0b3uT7SrtY7qbxd4Uh0XX7XSfDt7Z3WtQBZJHnDpG28MVUdk6n16VYsPF/he5k03V9d0S9uNc02KNEeGZVguDH/AKtpAeQRgZx1xXn6wyHohPelEMm77pyan2VHv+IOrV7HZab4l0G9t76HxTpE08lzeterd2TrHMrt95CT1X09Kx/HevN4n8RSamLf7PH5aRRRl9zKiDAy3c+9YxidCAw604RMMDFVGFGMuZMzlKrJWsbHiHWY9V0LQNMSB4zpdtJC7MRiQtIWyPSmeDtWOg6uLtrZbq3kieC5t2YgTROMMue3sfUVmlRgYP5U5UJ6ZP4GqTpKPLch+2bvY6vVvEWiW/h+TQ/C+l3dpb3c8c95LdzB5JNhysYwMBQeadfeLo7rUfE92tjIg1yERqpcExYZTk8c/d/WuTQH+4//AHwaQnA5R+vHyGoUKHf8Rt1+x2dh4u0+D/hGLmXTrp7zQwYmKygRzRZY9CMhst9Kq6drekXOhR6J4h0+7uLa2nea0ltZVSSMPyyHcCCpwD6iuYDZXOxxzg/uzT1JG4hH4/2DS5aC6itiH9k7eHxbpd1LqdjqmikaJeiER29rIEktzEMIwYjBOM5z1pt14s0z7fpUEGhRvoWmo6pZXDb2l8z77s2PvnqOOMVxRl24+STBxk7DUUkzHIEUn/fFJxw973C2JenKdd4g13R38PtoPh6wvLe1luRdzvdyh3LBSFVQowAMnnqa5NlIx3qNZmONsbk+y1M32kICbaTBA7VpTqUaasmTOjXm7uIBCccH604jFIpvOQLZ+mc7TURa6LEeW/r901r9Yp9xfU6r6ErKc5GetI6Bv6ioALxjhYZj34Q9KFjvmcgQzE/7pqXiafcqOCrJ6Ign0aCVmdGaInnAGV/KqM2jXKZCSxMPqQa1pLe9KqfJm/74NRm1vWzi3nI/3DWMqlF9T06NTF09DM/si74zJD/30f8ACpY9GbIMtwi+yirf2e9YZFvOTjvGaPsd8D/x7z8/7BqVOguprKvjJCW2n2UZyVEjDu5q9G0agBSABzj2qkbS9GMW82O+ENL9j1A4H2WcDH9w/wCFaxxFGOzRxVcPiKr967NAXEZYqXwPWobp45YiqnJPoKqpZahI+FtZyxOBiM1qWmkXaAF7aYsR/wA8zx+lV9bpP7RzSwM6a5rM5y4tLg5+Tjr161VktZBjKketdo+n3GM/ZZs5x/qmP9KhbSbhmOLS4P8A2xb/AAqJVKUteY76GMr01ZQOTjtxnpUnl4zgCujbRbs/MLK464x5Lf4VFLoGot92wux04EDf4Vn7SkvtI7VjKkt4swsAnv8AhSsNox1rXbw9quSyabekZ/593/woPhzWyT/xK74+n+jPx+lV9ZpL7Q+acvssx1PqD6UYO30rUbw9rxH/ACCNQ/C1k5/SnJ4b1/IxoupHjr9kk/wpfWaT+0JxqdjJO7OCPam4Bzn1rYbwz4hb/mCap1/585P8KVfC3iInjQ9U47/Ypf8A4mj6zS/mRjKlUf2TJBBzx7dKQjB6de/pW0vhXxKRxoOre+LGX/4ml/4RLxQeP+Ee1jP/AF4y/wDxNH1ql/MifZVf5TFwMhcH3poB5/wre/4RHxX/AA+G9Y54/wCPCX/4mnDwZ4tP/Mta0c9/sEv/AMTS+s0f5kP2Vb+UwQnp2GaTac5yK6AeDvFwGB4X1rr/ANA+X/4mnr4K8XngeFtbx/2D5f8A4mj61R/mQ1RrdjnyPQfpQEJU9Bjr710f/CE+Mj/zK2uE9P8AkHy//E08eBfGZH/Ip677406X/wCJo+t0f5g+r1X0ObRcng/jmnhDzh+OtdAvgXxsDkeEdePPT+z5f/iaePAvjjP/ACKGvZ/7B0v/AMTU/W6P8xSw9X+U5ooRnPPNJsZidzY5rqz4B8cNC0n/AAiet5GBt/s+bJ+ny01PAPjpsj/hENePU5/s6X/4mj65R/mG8NV7HK7M4YD2NPROvb+tdMPAXjnH/IneIOev/Eul/wAKevw98d5+Xwbr+D/1D5P8KTxdH+ZFRo1V0OWYD9fSkwc/qDXWf8K78fnH/FF66ef+gfJ/hTj8NviCef8AhDNc9R/oT/4VH1yh/MjphRq9jkJOmSpGKaRkAgda69vhx8QW6eCtf/8AAF/8KP8AhWvxFIx/whOvYx/z5NSeNofzI6qdGfY41wR95fambQQecV2n/CsviMV/5EjXv/ANuaP+FW/EZiP+KI1wf9upFQ8bQ/mR2wpyXQ4kqMZwaaEJH613CfCv4lN08DaxnpzBj+ZrpPDH7P3xL1mVRc6TDpEHQy306rj/AICm5v0rKpjqCV+Y2UGeR7Qc19Tfsp/CC5sZ4vHPie0MMoXOmWkq4Zcj/XMD0OPuj3z6V2Pwp/Z98M+EbmLVNZm/t7VYyGRpY9sELeqx85PuxPsBXswAA4rwsdmPtVyw2NIwSFooorySwooooA8+/aF/5JBr/wD16n+dfm3B90/WiiuqgAk3VvrUDfe/Ciiutkir96p4fvL9R/SiitKYmdJB/qV+gp8P3G+tFFelHYhk0v8Ax4p/10/pSL/qPyoopohjf+Wi/SkTp/wGiimxdRG/1Y+o/nTW6miiqRLGn7n4/wCFOh/134N/KiiuapubUyuv3z9KjT+lFFJDZIeg+tKf8/pRRVohCNUbffNFFDAQ9B+FKv3vwoopMBR94U40UUyWRv1H1pI/vfhRRQZMnH8NK3QUUUwYnYUDpRRSMpDh90U6PqKKKpGTJV/h/wA96mtP9e30NFFDMi9Z9Hq2n3B9f8KKKl7DZZH/AC0qW06/hRRQQa0fUfSpYf8AXR/71FFYMY8f68/j/On23+tb6f1oooYkaMP3atR/dFFFQxmgn3BT4fvt9BRRUCL8f3qcerfWiigC5B9386uQdR+NFFI0LH8A+tSp0P0oopFE8fX/AD6VaX+D6/0ooqWUWh1/Cnp938aKKQzStPufhVu3+/RRSGix/wAsz9Kev3aKKaKZLHV616H6UUVLKiWl+/T+woopMZKKfH1ooqWNEi/dpRRRUDHD734VKO9FFQxS2HDpTl7UUVDM4jhSmiipNEKOgp3c0UUi0LRRRUlIQUlFFMBf4qO1FFBYoo7UUUAL3o7UUUgAUlFFMBaRaKKAHGk7UUUhsKXtRRQAo6U4dKKKktCH7tC0UUFB3pD0/GiimUN709On4UUUMB1FFFSACiiigD//2Q=="


def _generar_pdf_carnet_afiliado(datos, firma_b64=None):
    """Genera el carné de afiliación en tamaño CR-80 (85.6mm x 54mm, el
    tamaño estándar de una tarjeta de crédito/credencial): una página con
    el frente (el diseño oficial que se subió, con el logo de CABAL) y
    otra página con el reverso, con los datos del afiliado llenados sobre
    el mismo formato de 'DATOS DEL AFILIADO' (Fecha, Nombres, Apellidos,
    DPI No., Departamento, Municipio). Si se manda 'firma_b64' (una firma
    dibujada en pantalla, como imagen PNG en base64), se imprime encima de
    la línea de 'FIRMA DEL AFILIADO'; si no, esa línea queda en blanco
    para firmarse a mano. La línea de 'FIRMA Y SELLO CABAL' siempre queda
    en blanco."""
    import base64 as b64mod
    from reportlab.pdfgen import canvas as _pdfcanvas_carnet
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.units import mm

    ANCHO = 85.6 * mm
    ALTO = 54 * mm

    buffer = io.BytesIO()
    c = _pdfcanvas_carnet.Canvas(buffer, pagesize=(ANCHO, ALTO))

    # ---- Página 1: frente (imagen fija, el diseño oficial) ----
    try:
        frente_img = ImageReader(io.BytesIO(b64mod.b64decode(_CARNET_FRENTE_B64)))
        c.drawImage(frente_img, 0, 0, width=ANCHO, height=ALTO, preserveAspectRatio=False)
    except Exception:
        pass
    c.showPage()

    # ---- Página 2: reverso con los datos del afiliado ----
    margen = 2.8*mm
    azul_oscuro = colors.HexColor('#1a3a6b')
    negro = colors.HexColor('#1c2833')

    # Fondo blanco (por si el PDF no lo asume así por defecto en algún lector)
    c.setFillColor(colors.white)
    c.rect(0, 0, ANCHO, ALTO, stroke=0, fill=1)

    # Barra "DATOS DEL AFILIADO"
    alto_barra = 6.5*mm
    c.setFillColor(azul_oscuro)
    c.rect(margen, ALTO - margen - alto_barra, 40*mm, alto_barra, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 7.5)
    c.drawString(margen + 1.5*mm, ALTO - margen - alto_barra + 2.1*mm, "DATOS DEL AFILIADO")

    # Fecha, a la par de la barra
    c.setFillColor(negro)
    c.setFont('Helvetica', 6.5)
    c.drawString(margen + 41.5*mm, ALTO - margen - alto_barra + 2.1*mm, f"FECHA: {datos.get('fecha','')}")

    nombre_completo = f"{datos.get('primer_nombre','')} {datos.get('segundo_nombre','')}".strip()
    apellido_completo = f"{datos.get('primer_apellido','')} {datos.get('segundo_apellido','')}".strip()

    campos = [
        ("NOMBRES:", nombre_completo),
        ("APELLIDOS:", apellido_completo),
        ("DPI NO.", datos.get('cui','')),
        ("DEPARTAMENTO:", datos.get('departamento','')),
        ("MUNICIPIO:", datos.get('municipio','')),
    ]

    y = ALTO - margen - alto_barra - 4.3*mm
    paso = 4.6*mm
    for etiqueta, valor in campos:
        c.setFont('Helvetica-Bold', 6.3)
        c.setFillColor(negro)
        c.drawString(margen, y, etiqueta)
        ancho_etiqueta = c.stringWidth(etiqueta, 'Helvetica-Bold', 6.3)
        x_linea = margen + ancho_etiqueta + 1.2*mm
        c.setFont('Helvetica', 6.3)
        c.drawString(x_linea, y, (valor or '')[:42])
        c.setStrokeColor(colors.HexColor('#7f8c8d'))
        c.setLineWidth(0.3)
        c.line(x_linea, y - 0.6*mm, ANCHO - margen, y - 0.6*mm)
        y -= paso

    # Líneas de firma — si llegó una firma dibujada en pantalla, se
    # imprime justo encima de la línea de "FIRMA DEL AFILIADO", usando
    # todo el espacio disponible arriba de esa línea para que se vea
    # grande y legible; la de "FIRMA Y SELLO CABAL" siempre queda en
    # blanco (esa la pone el comité).
    y_firma = margen + 3.8*mm
    ancho_firma = (ANCHO - 2*margen - 4*mm) / 2

    if firma_b64:
        try:
            firma_datos = firma_b64.split(',')[-1]  # quitar el prefijo "data:image/png;base64," si viene
            firma_img = ImageReader(io.BytesIO(b64mod.b64decode(firma_datos)))
            alto_firma_img = 10.5*mm
            c.drawImage(firma_img, margen + 0.5*mm, y_firma + 0.3*mm, width=ancho_firma - 1*mm,
                        height=alto_firma_img, preserveAspectRatio=False, mask='auto')
        except Exception:
            pass

    c.setStrokeColor(negro)
    c.setLineWidth(0.4)
    c.line(margen, y_firma, margen + ancho_firma, y_firma)
    c.line(ANCHO - margen - ancho_firma, y_firma, ANCHO - margen, y_firma)
    c.setFont('Helvetica', 5.3)
    c.setFillColor(negro)
    c.drawCentredString(margen + ancho_firma/2, y_firma - 2.6*mm, "FIRMA DEL AFILIADO")
    c.drawCentredString(ANCHO - margen - ancho_firma/2, y_firma - 2.6*mm, "FIRMA Y SELLO CABAL")

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer.read()


@app.route("/pdf_carnet_afiliado", methods=["POST"])
@requiere_afiliacion
def pdf_carnet_afiliado():
    """Genera el carné (frente + reverso) del último afiliado guardado,
    en tamaño CR-80, listo para imprimir en una impresora de tarjetas o en
    papel normal y recortar. Si viene una firma dibujada en pantalla, se
    imprime en el carné Y ADEMÁS se guarda en la hoja AFILIADOS (columna
    FirmaB64), para no perderla y poder reutilizarla después (por ejemplo,
    al imprimir la lista de varios afiliados)."""
    try:
        datos = request.json.get("datos", {})
        firma_b64 = request.json.get("firma_b64")
        pdf_bytes = _generar_pdf_carnet_afiliado(datos, firma_b64)
        filename = f"Carnet_{(datos.get('primer_apellido') or 'afiliado').replace(' ','_')}.pdf"
        token = _guardar_pdf_temporal(pdf_bytes, filename)
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')

        if firma_b64:
            try:
                cui = (datos.get('cui') or '').replace(' ', '').replace('-', '').strip()
                if cui:
                    sh = get_sheet()
                    ws = _get_ws_afiliados(sh)
                    cuis_col = ws.col_values(1)
                    for i, valor in enumerate(cuis_col, start=1):
                        if valor.replace(' ', '').strip() == cui:
                            ws.update_cell(i, len(_AFILIADOS_ENCABEZADO), firma_b64)
                            break
            except Exception:
                pass  # si falla el guardado de la firma, no debe impedir que se entregue el PDF ya generado

        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/listar_afiliados_fecha")
@requiere_afiliacion
def listar_afiliados_fecha():
    """Devuelve los afiliados guardados en una fecha concreta (o todos, si
    no se manda fecha), para poder elegir cuáles imprimir en la hoja tamaño
    carta con varios carnés."""
    try:
        fecha_filtro = request.args.get("fecha", "").strip()
        sh = get_sheet()
        ws = _get_ws_afiliados(sh)
        filas = ws.get_all_values()
        resultado = []
        for fila in filas[1:]:
            if not fila or not fila[0].strip():
                continue
            fecha_fila = fila[19].strip() if len(fila) > 19 else ''
            if fecha_filtro and fecha_fila != fecha_filtro:
                continue
            resultado.append({
                "cui": fila[0].strip(),
                "primer_nombre": fila[1].strip() if len(fila) > 1 else '',
                "segundo_nombre": fila[2].strip() if len(fila) > 2 else '',
                "primer_apellido": fila[3].strip() if len(fila) > 3 else '',
                "segundo_apellido": fila[4].strip() if len(fila) > 4 else '',
                "departamento": fila[11].strip() if len(fila) > 11 else '',
                "municipio": fila[10].strip() if len(fila) > 10 else '',
                "direccion": fila[14].strip() if len(fila) > 14 else '',
                "telefono": fila[15].strip() if len(fila) > 15 else '',
                "empadronado": fila[16].strip() if len(fila) > 16 else 'NO',
                "fecha": fecha_fila,
            })
        return jsonify({"ok": True, "afiliados": resultado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _generar_pdf_hoja_carnets(lista_afiliados):
    """Genera una hoja tamaño CARTA con varios carnés a la vez (frente y
    reverso), acomodados en una cuadrícula de 2 columnas x 4 filas (8
    carnés por hoja), para imprimir varios de un solo tiro y recortarlos.
    Primero salen todas las hojas de FRENTES, y luego, en las mismas
    posiciones de cuadrícula, las hojas de REVERSOS — así, al imprimir a
    doble cara (o al recortar de 2 hojas separadas), cada frente coincide
    con su reverso correspondiente."""
    import base64 as b64mod
    from reportlab.pdfgen import canvas as _pdfcanvas_hoja
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.units import mm

    PAGINA = letter  # 8.5in x 11in
    ANCHO_CARNET = 85.6 * mm
    ALTO_CARNET = 54 * mm
    COLS, FILAS = 2, 4
    GAP_X, GAP_Y = 6*mm, 6*mm
    margen_x = (PAGINA[0] - (COLS*ANCHO_CARNET + (COLS-1)*GAP_X)) / 2
    margen_y_top = (PAGINA[1] - (FILAS*ALTO_CARNET + (FILAS-1)*GAP_Y)) / 2

    buffer = io.BytesIO()
    c = _pdfcanvas_hoja.Canvas(buffer, pagesize=PAGINA)

    try:
        frente_img = ImageReader(io.BytesIO(b64mod.b64decode(_CARNET_FRENTE_B64)))
    except Exception:
        frente_img = None

    por_pagina = COLS * FILAS
    bloques = [lista_afiliados[i:i+por_pagina] for i in range(0, len(lista_afiliados), por_pagina)] or [[]]

    def _posicion(idx):
        col = idx % COLS
        fila = idx // COLS
        x = margen_x + col * (ANCHO_CARNET + GAP_X)
        y = PAGINA[1] - margen_y_top - (fila+1) * ALTO_CARNET - fila * GAP_Y
        return x, y

    # ---- Hojas de FRENTES ----
    for bloque in bloques:
        for idx in range(len(bloque)):
            x, y = _posicion(idx)
            if frente_img:
                c.drawImage(frente_img, x, y, width=ANCHO_CARNET, height=ALTO_CARNET, preserveAspectRatio=False)
        # Líneas guía de corte, tenues
        c.setStrokeColor(colors.HexColor('#bdc3c7'))
        c.setDash(2, 2)
        c.setLineWidth(0.4)
        for i in range(len(bloque)):
            x, y = _posicion(i)
            c.rect(x, y, ANCHO_CARNET, ALTO_CARNET, stroke=1, fill=0)
        c.setDash()
        c.showPage()

    # ---- Hojas de REVERSOS (mismas posiciones, para que coincidan al recortar) ----
    for bloque in bloques:
        for idx, afiliado in enumerate(bloque):
            x, y = _posicion(idx)
            _dibujar_reverso_carnet_en(c, x, y, ANCHO_CARNET, ALTO_CARNET, afiliado)
        c.setStrokeColor(colors.HexColor('#bdc3c7'))
        c.setDash(2, 2)
        c.setLineWidth(0.4)
        for i in range(len(bloque)):
            x, y = _posicion(i)
            c.rect(x, y, ANCHO_CARNET, ALTO_CARNET, stroke=1, fill=0)
        c.setDash()
        c.showPage()

    c.save()
    buffer.seek(0)
    return buffer.read()


def _dibujar_reverso_carnet_en(c, x0, y0, ancho, alto, datos):
    """Dibuja el reverso de UN carné (los mismos datos que
    _generar_pdf_carnet_afiliado dibuja en su propia página), pero
    dentro de una posición (x0,y0) de una hoja más grande — se usa para
    la hoja carta con varios carnés a la vez."""
    from reportlab.lib.units import mm
    margen = 2.8*mm
    azul_oscuro = colors.HexColor('#1a3a6b')
    negro = colors.HexColor('#1c2833')

    c.saveState()
    c.translate(x0, y0)
    c.setFillColor(colors.white)
    c.rect(0, 0, ancho, alto, stroke=0, fill=1)

    alto_barra = 6.5*mm
    c.setFillColor(azul_oscuro)
    c.rect(margen, alto - margen - alto_barra, 40*mm, alto_barra, stroke=0, fill=1)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 7.5)
    c.drawString(margen + 1.5*mm, alto - margen - alto_barra + 2.1*mm, "DATOS DEL AFILIADO")

    c.setFillColor(negro)
    c.setFont('Helvetica', 6.5)
    c.drawString(margen + 41.5*mm, alto - margen - alto_barra + 2.1*mm, f"FECHA: {datos.get('fecha','')}")

    nombre_completo = f"{datos.get('primer_nombre','')} {datos.get('segundo_nombre','')}".strip()
    apellido_completo = f"{datos.get('primer_apellido','')} {datos.get('segundo_apellido','')}".strip()
    campos = [
        ("NOMBRES:", nombre_completo),
        ("APELLIDOS:", apellido_completo),
        ("DPI NO.", datos.get('cui','')),
        ("DEPARTAMENTO:", datos.get('departamento','')),
        ("MUNICIPIO:", datos.get('municipio','')),
    ]
    y = alto - margen - alto_barra - 4.3*mm
    paso = 4.6*mm
    for etiqueta, valor in campos:
        c.setFont('Helvetica-Bold', 6.3)
        c.setFillColor(negro)
        c.drawString(margen, y, etiqueta)
        ancho_etiqueta = c.stringWidth(etiqueta, 'Helvetica-Bold', 6.3)
        x_linea = margen + ancho_etiqueta + 1.2*mm
        c.setFont('Helvetica', 6.3)
        c.drawString(x_linea, y, (valor or '')[:42])
        c.setStrokeColor(colors.HexColor('#7f8c8d'))
        c.setLineWidth(0.3)
        c.line(x_linea, y - 0.6*mm, ancho - margen, y - 0.6*mm)
        y -= paso

    y_firma = margen + 3.8*mm
    ancho_firma = (ancho - 2*margen - 4*mm) / 2
    c.setStrokeColor(negro)
    c.setLineWidth(0.4)
    c.line(margen, y_firma, margen + ancho_firma, y_firma)
    c.line(ancho - margen - ancho_firma, y_firma, ancho - margen, y_firma)
    c.setFont('Helvetica', 5.3)
    c.setFillColor(negro)
    c.drawCentredString(margen + ancho_firma/2, y_firma - 2.6*mm, "FIRMA DEL AFILIADO")
    c.drawCentredString(ancho - margen - ancho_firma/2, y_firma - 2.6*mm, "FIRMA Y SELLO CABAL")
    c.restoreState()


@app.route("/pdf_hoja_carnets", methods=["POST"])
@requiere_afiliacion
def pdf_hoja_carnets():
    """Genera la hoja tamaño CARTA con varios carnés (frente y reverso)
    de los afiliados de la fecha elegida, lista para imprimir y recortar."""
    try:
        fecha_filtro = (request.json or {}).get("fecha", "").strip()
        sh = get_sheet()
        ws = _get_ws_afiliados(sh)
        filas = ws.get_all_values()
        lista = []
        for fila in filas[1:]:
            if not fila or not fila[0].strip():
                continue
            fecha_fila = fila[19].strip() if len(fila) > 19 else ''
            if fecha_filtro and fecha_fila != fecha_filtro:
                continue
            lista.append({
                "cui": fila[0].strip(),
                "primer_nombre": fila[1].strip() if len(fila) > 1 else '',
                "segundo_nombre": fila[2].strip() if len(fila) > 2 else '',
                "primer_apellido": fila[3].strip() if len(fila) > 3 else '',
                "segundo_apellido": fila[4].strip() if len(fila) > 4 else '',
                "departamento": fila[11].strip() if len(fila) > 11 else '',
                "municipio": fila[10].strip() if len(fila) > 10 else '',
                "fecha": fecha_fila,
            })
        if not lista:
            return jsonify({"ok": False, "error": "No hay afiliados guardados para esa fecha"})
        pdf_bytes = _generar_pdf_hoja_carnets(lista)
        token = _guardar_pdf_temporal(pdf_bytes, "Hoja_Carnets.pdf")
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token, "total": len(lista)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _generar_pdf_lista_afiliados(lista_afiliados, fecha_filtro):
    """Genera la 'LISTA DE REGISTRO DE AFILIADOS' en tamaño carta
    horizontal, con el mismo formato ya conocido de 'Lista de Registro de
    Grupo' (encabezado con el logo de CABAL, tabla con No./Nombre/DPI/
    Dirección/Teléfono/Emp./Firma-Huella), en bloques de 10 personas por
    hoja para que cada quien firme o ponga su huella en persona."""
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import Image as RLImage
    import base64 as b64mod

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            rightMargin=0.4*inch, leftMargin=0.4*inch,
                            topMargin=0.3*inch, bottomMargin=0.3*inch)

    titulo_style = ParagraphStyle('titulo', fontSize=11, textColor=colors.HexColor('#1a5276'),
        spaceAfter=1, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
    sub_style = ParagraphStyle('sub', fontSize=8, textColor=colors.HexColor('#2c3e50'),
        spaceAfter=1, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')
    normal_style = ParagraphStyle('normal', fontSize=8, textColor=colors.HexColor('#2c3e50'),
        spaceAfter=1, spaceBefore=0, fontName='Helvetica')
    celda_s = ParagraphStyle('celda', fontSize=7, alignment=TA_CENTER, fontName='Helvetica', leading=8.5)

    try:
        logo_img = RLImage(io.BytesIO(b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7qy654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")),
                            width=1.1*inch, height=1.0*inch)
    except Exception:
        logo_img = None

    story = []
    bloques = [lista_afiliados[i:i+10] for i in range(0, len(lista_afiliados), 10)] or [[]]
    for idx_bloque, bloque in enumerate(bloques):
        if idx_bloque > 0:
            story.append(PageBreak())
        header_content = [
            Paragraph("PARTIDO CABAL", titulo_style),
            Paragraph("MUNICIPIO DE TOTONICAPAN", sub_style),
            Paragraph("LISTA DE REGISTRO DE AFILIADOS", titulo_style),
            Paragraph(f"<b>Fecha:</b> {fecha_filtro or 'Todas'}", sub_style),
            Paragraph(f"<b>Total:</b> {len(lista_afiliados)} afiliado(s)", sub_style),
        ]
        if logo_img:
            header_table = Table([[header_content, logo_img]], colWidths=[9.1*inch, 1.1*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(header_table)
        else:
            for item in header_content:
                story.append(item)
        story.append(Spacer(1, 0.06*inch))

        table_data = [["No.", "Nombre Completo", "DPI", "Dirección/Comunidad", "Teléfono", "Emp.", "Firma / Huella"]]
        for i, a in enumerate(bloque, 1):
            nombre = f"{a.get('primer_nombre','')} {a.get('segundo_nombre','')} {a.get('primer_apellido','')} {a.get('segundo_apellido','')}".strip()
            celda_firma = ''
            firma_b64_fila = (a.get('firma_b64') or '').strip()
            if firma_b64_fila:
                try:
                    firma_datos = firma_b64_fila.split(',')[-1]
                    firma_bytes = b64mod.b64decode(firma_datos)
                    celda_firma = RLImage(io.BytesIO(firma_bytes), width=2.7*inch, height=0.34*inch)
                except Exception:
                    celda_firma = ''
            table_data.append([
                str(i), Paragraph(nombre, celda_s), a.get('cui',''),
                Paragraph(a.get('direccion','') or (a.get('municipio','') or ''), celda_s),
                a.get('telefono',''), a.get('empadronado','NO'), celda_firma
            ])
        for i in range(len(bloque)+1, 11):
            table_data.append([str(i), '', '', '', '', '', ''])

        col_widths = [0.35*inch, 2.5*inch, 1.2*inch, 1.7*inch, 0.95*inch, 0.45*inch, 3.0*inch]
        t = Table(table_data, colWidths=col_widths, rowHeights=[0.32*inch] + [0.42*inch]*10)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a5276')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('FONTSIZE', (0,1), (-1,-1), 7),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#2c3e50')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eaf4fb')]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.08*inch))
        story.append(Paragraph(f"Registrado por: {', '.join(sorted(set(a.get('registrado_por','') for a in bloque if a.get('registrado_por'))))}", normal_style))
        story.append(Spacer(1, 0.08*inch))
        story.append(Paragraph(f"Firma del Encargado de Afiliación: {'_'*55}", normal_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


@app.route("/pdf_lista_afiliados", methods=["POST"])
@requiere_afiliacion
def pdf_lista_afiliados():
    """Genera la lista de afiliados (tabla tamaño carta, la misma
    presentación que 'Lista de Registro de Grupo') de la fecha elegida,
    con espacio para firma/huella de cada persona."""
    try:
        fecha_filtro = (request.json or {}).get("fecha", "").strip()
        sh = get_sheet()
        ws = _get_ws_afiliados(sh)
        filas = ws.get_all_values()
        lista = []
        for fila in filas[1:]:
            if not fila or not fila[0].strip():
                continue
            fecha_fila = fila[19].strip() if len(fila) > 19 else ''
            if fecha_filtro and fecha_fila != fecha_filtro:
                continue
            lista.append({
                "cui": fila[0].strip(),
                "primer_nombre": fila[1].strip() if len(fila) > 1 else '',
                "segundo_nombre": fila[2].strip() if len(fila) > 2 else '',
                "primer_apellido": fila[3].strip() if len(fila) > 3 else '',
                "segundo_apellido": fila[4].strip() if len(fila) > 4 else '',
                "direccion": fila[14].strip() if len(fila) > 14 else '',
                "telefono": fila[15].strip() if len(fila) > 15 else '',
                "empadronado": fila[16].strip() if len(fila) > 16 else 'NO',
                "municipio": fila[10].strip() if len(fila) > 10 else '',
                "registrado_por": fila[18].strip() if len(fila) > 18 else '',
                "fecha": fecha_fila,
                "firma_b64": fila[20].strip() if len(fila) > 20 else '',
            })
        if not lista:
            return jsonify({"ok": False, "error": "No hay afiliados guardados para esa fecha"})
        pdf_bytes = _generar_pdf_lista_afiliados(lista, fecha_filtro)
        token = _guardar_pdf_temporal(pdf_bytes, "Lista_Afiliados.pdf")
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token, "total": len(lista)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/listar_registros_prueba")
@requiere_admin
def listar_registros_prueba():
    """Devuelve todos los registros de la hoja DPI marcados con
    Jefe='PRUEBA' (los creados desde 'Agregar registro de prueba' en
    Mantenimiento), junto con su número de fila, para poder listarlos,
    copiar el DPI, o eliminarlos directamente sin tener que recordarlo."""
    try:
        sh = get_sheet()
        ws = sh.worksheet(HOJA)
        filas = ws.get_all_values()
        registros = []
        for i, fila in enumerate(filas[1:], start=2):
            if not fila or not fila[0].strip():
                continue
            jefe_fila = fila[19].strip().upper() if len(fila) > 19 else ''
            if jefe_fila == 'PRUEBA':
                nombre = f"{fila[2] if len(fila)>2 else ''} {fila[4] if len(fila)>4 else ''}".strip()
                registros.append({
                    "fila": i,
                    "cui": fila[0].strip(),
                    "nombre": nombre or '-',
                    "fecha_creacion": fila[23].strip() if len(fila) > 23 else '',
                })
        registros.sort(key=lambda r: r.get("fecha_creacion",""), reverse=True)
        return jsonify({"ok": True, "registros": registros})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/registros")
@requiere_sesion
def registros():
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        sh = get_sheet()
        ws = sh.worksheet(HOJA)
        filas = ws.get_all_values()
        if len(filas) > 1:
            datos = filas[1:]
            if not _puede_ver_todo():
                # Columna T (índice 19) = Jefe de Sector. Solo se muestran los
                # registros del jefe con sesión iniciada; los que no tengan
                # jefe asignado (registros antiguos) no se muestran a nadie
                # que no sea administrador, para no exponerlos por error.
                datos = [f for f in datos if len(f) > 19 and f[19].strip().upper() == _nombre_sesion.upper()]
            return jsonify({"ok": True, "filas": datos})
        return jsonify({"ok": True, "filas": []})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})



@app.route("/generar_pdf_grupo", methods=["POST"])
@requiere_sesion
def generar_pdf_grupo():
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import io
        from datetime import datetime

        data = request.json
        nombre_grupo = data.get("nombre_grupo", "Grupo")
        coord_nombre = data.get("coordinadora_nombre", "")
        coord_dpi = data.get("coordinadora_dpi", "")
        jefe_sector = data.get("jefe_sector", "")
        personas = data.get("personas", [])

        buffer = io.BytesIO()
        from reportlab.lib.pagesizes import letter, landscape
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                rightMargin=0.4*inch, leftMargin=0.4*inch,
                                topMargin=0.3*inch, bottomMargin=0.3*inch)

        styles = getSampleStyleSheet()
        titulo_style = ParagraphStyle('titulo', parent=styles['Title'],
            fontSize=11, textColor=colors.HexColor('#1a5276'),
            spaceAfter=1, spaceBefore=0, alignment=TA_CENTER)
        sub_style = ParagraphStyle('sub', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=1, spaceBefore=0, alignment=TA_CENTER)
        normal_style = ParagraphStyle('normal', parent=styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#2c3e50'), spaceAfter=1, spaceBefore=0)
        from reportlab.platypus import Image as RLImage

        story = []
        import base64 as b64mod

        # Logo CABAL embedded - usando BytesIO en memoria, colocado a la derecha
        try:
            logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7qy654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
            logo_buffer = io.BytesIO(logo_data)
            logo_img = RLImage(logo_buffer, width=1.1*inch, height=1.0*inch)
        except:
            logo_img = None

        presidenta_pdf = data.get('presidenta','')
        header_content = [
            Paragraph("PARTIDO CABAL", titulo_style),
            Paragraph("MUNICIPIO DE TOTONICAPAN", sub_style),
            Paragraph("LISTA DE REGISTRO DE GRUPO", titulo_style),
            Paragraph(f"<b>Jefe de Sector:</b> {jefe_sector}", sub_style),
        ]
        if presidenta_pdf:
            header_content.append(Paragraph(f"<b>Presidenta de Comité:</b> {presidenta_pdf}", sub_style))
        header_content += [
            Paragraph(f"<b>Grupo:</b> {nombre_grupo}", sub_style),
            Paragraph(f"<b>Coordinadora:</b> {coord_nombre} &nbsp;&nbsp; <b>DPI:</b> {coord_dpi}", sub_style),
            Paragraph(f"<b>Fecha:</b> {ahora_gt().strftime('%d/%m/%Y')}", sub_style),
        ]
        if logo_img:
            header_table = Table([[header_content, logo_img]], colWidths=[9.1*inch, 1.1*inch])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ]))
            story.append(header_table)
        else:
            for item in header_content:
                story.append(item)
        story.append(Spacer(1, 0.04*inch))

        # Tabla — ancho util carta horizontal: 11 - 0.8 = 10.2 pulgadas
        celda_s = ParagraphStyle('celda', fontSize=7, alignment=TA_CENTER, fontName='Helvetica', leading=8.5)
        table_data = [["No.", "Nombre Completo", "DPI", "Dirección/Comunidad", "Teléfono", "Emp.", "Firma / Huella"]]

        for i, p in enumerate(personas[:10], 1):
            nombre = f"{p.get('primer_nombre','')} {p.get('segundo_nombre','')} {p.get('primer_apellido','')} {p.get('segundo_apellido','')}".strip()
            table_data.append([
                str(i), Paragraph(nombre, celda_s), p.get('cui',''), Paragraph(p.get('direccion','') or '', celda_s), p.get('telefono',''), p.get('empadronado','NO'), ''
            ])

        for i in range(len(personas)+1, 11):
            table_data.append([str(i), '', '', '', '', '', ''])

        col_widths = [0.35*inch, 2.5*inch, 1.2*inch, 1.7*inch, 0.95*inch, 0.45*inch, 3.0*inch]
        row_height = 0.42*inch
        t = Table(table_data, colWidths=col_widths, rowHeights=[0.32*inch] + [row_height]*10)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a5276')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 8),
            ('FONTSIZE', (0,1), (-1,-1), 7),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#2c3e50')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eaf4fb')]),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.06*inch))
        story.append(Paragraph(f"Jefe de Sector: {jefe_sector}", normal_style))
        story.append(Spacer(1, 0.06*inch))
        story.append(Paragraph(f"Firma del Jefe de Sector: {'_'*55}", normal_style))
        story.append(Spacer(1, 0.08*inch))
        story.append(Paragraph(f"Firma de la Coordinadora: {'_'*55}", normal_style))

        # Save to Google Sheets
        # No reescribir en Sheets desde aquí — los datos ya fueron guardados por guardar_persona_grupo
        pass

        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        filename = f"{nombre_grupo.replace(' ','_')}.pdf"
        token = _guardar_pdf_temporal(pdf_bytes, filename)
        from flask import send_file
        resp = send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                         as_attachment=True,
                         download_name=filename)
        resp.headers['X-PDF-Token'] = token
        resp.headers['Access-Control-Expose-Headers'] = 'X-PDF-Token'
        return resp
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})



@app.route("/validar_grupo", methods=["POST"])
@requiere_sesion
def validar_grupo():
    try:
        data = request.json
        nombre_grupo = data.get("nombre_grupo","").strip()
        dpi_coord = data.get("dpi_coord","").strip()
        sh = get_sheet()
        try:
            ws = sh.worksheet("GRUPOS")
            filas = ws.get_all_values()
            grupos_vistos = set()
            dpis_vistos = set()
            for fila in filas[1:]:
                if not fila or not fila[0]: continue
                ng = fila[0].strip().lower()
                dc = fila[2].strip() if len(fila) > 2 else ''
                if ng not in grupos_vistos:
                    grupos_vistos.add(ng)
                    if ng == nombre_grupo.lower():
                        return jsonify({"ok": False, "error": f"Ya existe un grupo con el nombre '{nombre_grupo}'"})
                if dc and dc not in dpis_vistos:
                    dpis_vistos.add(dc)
                    if dc == dpi_coord:
                        return jsonify({"ok": False, "error": f"Este coordinador ya tiene un grupo asignado"})
        except:
            pass
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/guardar_persona_grupo", methods=["POST"])
@requiere_sesion
def guardar_persona_grupo():
    try:
        data = request.json
        nombre_grupo = data.get("nombre_grupo","")
        coord_nombre = data.get("coord_nombre","")
        coord_dpi = data.get("coord_dpi","")
        _nombre_sesion, _es_admin = _sesion_actual()
        # El jefe de sector del grupo se toma de la sesión real del servidor,
        # no de lo que envíe el navegador, salvo que sea el administrador
        # (que de todos modos no debería crear grupos desde la interfaz).
        jefe_sector = _nombre_sesion if not _es_admin else data.get("jefe_sector","")
        persona = data.get("persona",{})
        numero = data.get("numero", 1)
        from datetime import datetime
        fecha = ahora_gt().strftime('%d/%m/%Y')
        sh = get_sheet()
        try:
            ws_grupos = sh.worksheet("GRUPOS")
        except:
            ws_grupos = sh.add_worksheet(title="GRUPOS", rows=5000, cols=11)
            ws_grupos.append_row(["Grupo","Coordinador","DPI Coordinador","Jefe de Sector","Presidenta","DPI Presidenta","No.","Nombre Completo","DPI","Direccion","Telefono","Empadronado","Fecha"])
        nombre_p = f"{persona.get('primer_nombre','')} {persona.get('segundo_nombre','')} {persona.get('primer_apellido','')} {persona.get('segundo_apellido','')}".strip()
        presidenta = data.get('presidenta','')
        presidenta_dpi = data.get('presidenta_dpi','')
        cui_persona = persona.get('cui','').replace(" ","").replace("-","").strip()

        # La fecha de nacimiento es obligatoria: sin ella no se guarda el
        # registro (ni en GRUPOS ni en DPI), para evitar filas incompletas
        # que después son difíciles de completar a mano.
        if not persona.get('fecha_nacimiento','').strip():
            return jsonify({"ok": False, "error": "Falta la fecha de nacimiento. No se puede guardar el registro sin ese dato."})
        if not persona.get('direccion','').strip():
            return jsonify({"ok": False, "error": "Falta la dirección / comunidad. No se puede guardar el registro sin ese dato."})

        # Todo lo de aquí en adelante (revisar duplicados Y guardar) ocurre
        # dentro de un candado, para que ninguna otra petición pueda
        # "colarse" entre la revisión y el guardado.
        with _candado_registro_personas:
            # Verificar que esta persona no esté YA registrada como Presidenta
            # de Comité. Una misma persona no puede ser presidenta y, a la vez,
            # coordinadora o integrante de un grupo.
            if cui_persona:
                try:
                    ws_pr_check = sh.worksheet("PRESIDENTAS")
                    for fp in ws_pr_check.get_all_values()[1:]:
                        if len(fp) > 2 and fp[2].strip().replace(" ","").replace("-","") == cui_persona:
                            return jsonify({"ok": False, "duplicado": True, "es_presidenta": True,
                                "error": f"El CUI {cui_persona} ya está registrado como Presidenta de Comité ({fp[1]}). Una misma persona no puede ser presidenta y también coordinadora o integrante.",
                                "nombre": nombre_p, "cui": cui_persona})
                except Exception:
                    pass

            # Verificar duplicado y límite en GRUPOS
            filas_g = ws_grupos.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres_g = any('residenta' in str(h) for h in header_g)
            tiene_dir_g  = any('irecc' in str(h) for h in header_g)
            col_dpi_pers = 8 if tiene_pres_g else (6 if tiene_dir_g else 5)
            col_dpi_coord = 2

            # Verificar duplicado en TODA la hoja GRUPOS (cualquier grupo, de
            # cualquier jefe de sector) y límite de personas en el grupo actual.
            # Antes solo se comparaba contra el propio grupo, así que una persona
            # podía quedar registrada dos veces si la segunda vez era en un
            # grupo distinto (o incluso como coordinadora de otro grupo).
            count_grupo = 0
            grupo_duplicado = None
            for fg in filas_g[1:]:
                if not fg:
                    continue
                if fg[0].strip() == nombre_grupo.strip():
                    count_grupo += 1
                dpi_p = fg[col_dpi_pers].replace(" ","").replace("-","") if len(fg) > col_dpi_pers else ''
                dpi_c = fg[col_dpi_coord].replace(" ","").replace("-","") if len(fg) > col_dpi_coord else ''
                if cui_persona and grupo_duplicado is None and (dpi_p == cui_persona or dpi_c == cui_persona):
                    grupo_duplicado = fg[0].strip()

            if grupo_duplicado:
                mismo_grupo = grupo_duplicado == nombre_grupo.strip()
                texto_error = (f"El CUI {cui_persona} ya esta registrado en este grupo" if mismo_grupo
                               else f"El CUI {cui_persona} ya esta registrado en el grupo '{grupo_duplicado}'")
                return jsonify({"ok": False, "duplicado": True,
                    "error": texto_error, "grupo_duplicado": grupo_duplicado, "mismo_grupo": mismo_grupo,
                    "nombre": nombre_p, "cui": cui_persona})

            if count_grupo >= 10:
                return jsonify({"ok": False, "limite": True,
                    "error": "El grupo ya tiene 10 personas (1 coordinadora + 9 integrantes)"})

            ws_grupos.append_row([nombre_grupo, coord_nombre, coord_dpi.replace(" ","").replace("-","").strip(), jefe_sector, presidenta, presidenta_dpi.replace(" ","").replace("-","").strip(), numero,
                                  nombre_p, cui_persona, persona.get('direccion',''),
                                  formatear_telefono_gt(persona.get('telefono','')), persona.get('empadronado','NO'), fecha])

            jefe_autenticado = _nombre_sesion or "desconocido"
            if numero == 1:
                _registrar_auditoria(sh, jefe_autenticado, "Creó grupo", f"Grupo: {nombre_grupo} | Coordinadora: {coord_nombre}")
            else:
                _registrar_auditoria(sh, jefe_autenticado, "Agregó persona a grupo", f"Grupo: {nombre_grupo} | Persona: {nombre_p}")

            # Guardar también en hoja DPI con TODOS los campos del DPI
            try:
                ws_dpi = sh.worksheet(HOJA)
                cuis_existentes = ws_dpi.col_values(1)
                # Determinar rol según posición en el grupo
                rol_persona = 'Coordinadora' if persona.get('es_coordinador') == 'SI' else 'Integrante'
                cuis_clean = [c.replace(" ","").replace("-","").strip() for c in cuis_existentes]
                if cui_persona and cui_persona not in cuis_clean:
                    # Registro nuevo - guardar todos los datos. Se usa
                    # cui_persona (ya sin espacios ni guiones) en vez del
                    # CUI tal como vino del formulario, para que nunca
                    # queden espacios guardados en medio del número.
                    ws_dpi.append_row([
                            cui_persona,                                    # A CUI
                            persona.get('numero_serie',''),                 # B Serie
                            persona.get('primer_nombre',''),                # C
                            persona.get('segundo_nombre',''),               # D
                            persona.get('primer_apellido',''),              # E
                            persona.get('segundo_apellido',''),             # F
                            persona.get('sexo', persona.get('genero','')), # G Sexo
                            persona.get('estado_civil',''),                 # H Estado civil
                            persona.get('fecha_nacimiento',''),             # I
                            persona.get('municipio_nacimiento',''),         # J
                            persona.get('departamento_nacimiento',''),      # K
                            persona.get('municipio_vecindad',''),           # L
                            persona.get('departamento_vecindad',''),        # M
                            persona.get('fecha_expedicion',''),             # N
                            persona.get('fecha_vencimiento',''),            # O
                            persona.get('empadronado','NO'),                # P Empadronado
                            persona.get('num_empadronamiento',''),          # Q No.Emp
                            persona.get('direccion',''),                    # R Direccion
                            formatear_telefono_gt(persona.get('telefono','')),  # S Telefono
                            jefe_sector,                                    # T Jefe de Sector
                            presidenta,                                     # U Presidenta
                            rol_persona,                                    # V Rol
                            "NO" if persona.get('dpi_no_verificado') else "SI",  # W DPI Verificado
                            ahora_gt().strftime('%d/%m/%Y'),                # X Fecha de registro
                        ])
                else:
                    # Registro existente - actualizar campos vacíos
                        filas_dpi = ws_dpi.get_all_values()
                        for idx_f, fila_d in enumerate(filas_dpi[1:], start=2):
                            if fila_d and fila_d[0].replace(" ","").replace("-","").strip() == cui_persona:
                                # Actualizar solo campos que estén vacíos
                                updates = {}
                                campos = {
                                    'B': persona.get('numero_serie',''),
                                    'G': persona.get('sexo', persona.get('genero','')),
                                    'H': persona.get('estado_civil',''),
                                    'J': persona.get('municipio_nacimiento',''),
                                    'K': persona.get('departamento_nacimiento',''),
                                    'L': persona.get('municipio_vecindad',''),
                                    'M': persona.get('departamento_vecindad',''),
                                    'N': persona.get('fecha_expedicion',''),
                                    'O': persona.get('fecha_vencimiento',''),
                                    'T': jefe_sector,
                                    'U': presidenta,
                                    'V': rol_persona,
                                }
                                for col, val in campos.items():
                                    if val:
                                        col_idx = ord(col) - ord('A')
                                        if col_idx >= len(fila_d) or not fila_d[col_idx].strip():
                                            ws_dpi.update(f'{col}{idx_f}', [[val]])
                                break
            except Exception as e_dpi:
                print(f"Error guardando en DPI: {e_dpi}")

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/registrar_presidenta", methods=["POST"])
@requiere_sesion
def registrar_presidenta():
    try:
        data = request.json
        jefe_sector = data.get("jefe_sector","").strip()
        presidenta = data.get("presidenta",{})
        cui_limpio = presidenta.get("cui","").replace(" ","").replace("-","").strip()
        sh = get_sheet()

        # La fecha de nacimiento es obligatoria, igual que para
        # coordinadoras/integrantes.
        if not presidenta.get('fecha_nacimiento','').strip():
            return jsonify({"ok": False, "error": "Falta la fecha de nacimiento. No se puede guardar el registro sin ese dato."})
        if not presidenta.get('direccion','').strip():
            return jsonify({"ok": False, "error": "Falta la dirección / comunidad. No se puede guardar el registro sin ese dato."})

        # Todo lo de aquí en adelante (revisar duplicados Y guardar) ocurre
        # dentro de un candado, para que ninguna otra petición pueda
        # "colarse" entre la revisión y el guardado (ver nota junto a la
        # definición de _candado_registro_personas más arriba).
        with _candado_registro_personas:
            # Verificar que esta persona no esté ya registrada como
            # coordinadora o integrante de algún grupo (una misma persona no
            # puede ser presidenta y, a la vez, coordinadora/integrante).
            if cui_limpio:
                try:
                    ws_g_check = sh.worksheet("GRUPOS")
                    filas_g_check = ws_g_check.get_all_values()
                    header_check = filas_g_check[0] if filas_g_check else []
                    tiene_pres_check = any('residenta' in str(h) for h in header_check)
                    tiene_dir_check = any('irecc' in str(h) for h in header_check)
                    col_dpi_pers_check = 8 if tiene_pres_check else (6 if tiene_dir_check else 5)
                    col_dpi_coord_check = 2
                    for fg in filas_g_check[1:]:
                        if not fg:
                            continue
                        dpi_p_check = fg[col_dpi_pers_check].replace(" ","").replace("-","") if len(fg) > col_dpi_pers_check else ''
                        dpi_c_check = fg[col_dpi_coord_check].replace(" ","").replace("-","") if len(fg) > col_dpi_coord_check else ''
                        if dpi_p_check == cui_limpio or dpi_c_check == cui_limpio:
                            return jsonify({"ok": False, "duplicado": True, "es_grupo": True,
                                "error": f"El CUI {cui_limpio} ya está registrado como coordinadora/integrante en el grupo '{fg[0].strip()}'. Una misma persona no puede ser presidenta y también coordinadora o integrante."})
                except Exception:
                    pass

                # Si ya existe exactamente esta presidenta (mismo CUI), no
                # duplicar la fila — se trata como ya registrada.
                try:
                    ws_pr_check = sh.worksheet("PRESIDENTAS")
                    for fp in ws_pr_check.get_all_values()[1:]:
                        if len(fp) > 2 and fp[2].strip().replace(" ","").replace("-","") == cui_limpio:
                            return jsonify({"ok": True, "ya_existia": True})
                except Exception:
                    pass

            # Guardar en hoja DPI
            try:
                ws_dpi = sh.worksheet(HOJA)
                cuis = ws_dpi.col_values(1)
                cuis_limpios = [c.replace(" ","").replace("-","").strip() for c in cuis]
                fecha = ahora_gt().strftime('%d/%m/%Y')
                empadronado = presidenta.get("empadronado","NO")
                num_emp_presidenta = presidenta.get("num_empadronamiento","")
                if cui_limpio and cui_limpio not in cuis_limpios:
                    ws_dpi.append_row([
                        cui_limpio, presidenta.get("numero_serie",""),
                        presidenta.get("primer_nombre",""), presidenta.get("segundo_nombre",""),
                        presidenta.get("primer_apellido",""), presidenta.get("segundo_apellido",""),
                        presidenta.get("sexo",""), presidenta.get("estado_civil",""),
                        presidenta.get("fecha_nacimiento",""), presidenta.get("municipio_nacimiento",""),
                        presidenta.get("departamento_nacimiento",""), presidenta.get("municipio_vecindad",""),
                        presidenta.get("departamento_vecindad",""), presidenta.get("fecha_expedicion",""),
                        presidenta.get("fecha_vencimiento",""), empadronado, num_emp_presidenta,
                        presidenta.get("direccion",""), presidenta.get("telefono",""),
                        jefe_sector, presidenta.get("nombre",""), "Presidenta de Comité",
                        "NO" if presidenta.get('dpi_no_verificado') else "SI",
                        fecha,
                    ])
            except Exception as e:
                print(f"Error DPI presidenta: {e}")

            # Guardar en hoja PRESIDENTAS
            try:
                try:
                    ws_pr = sh.worksheet("PRESIDENTAS")
                except:
                    ws_pr = sh.add_worksheet(title="PRESIDENTAS", rows=1000, cols=8)
                    ws_pr.append_row(["Jefe de Sector","Nombre","DPI","Telefono","Direccion","Empadronado","Fecha"])
                nombre = presidenta.get("nombre","") or f"{presidenta.get('primer_nombre','')} {presidenta.get('segundo_nombre','')} {presidenta.get('primer_apellido','')} {presidenta.get('segundo_apellido','')}".strip()
                ws_pr.append_row([
                    jefe_sector, nombre, cui_limpio,
                    formatear_telefono_gt(presidenta.get('telefono','')), presidenta.get('direccion',''),
                    empadronado, ahora_gt().strftime('%d/%m/%Y')
                ])
            except Exception as e:
                print(f"Error PRESIDENTAS: {e}")

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/pdf_presidentas", methods=["POST"])
@requiere_sesion
def pdf_presidentas():
    try:
        data = request.json
        jefe = data.get("jefe_sector","")
        presidentas = data.get("presidentas",[])

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.4*inch, bottomMargin=0.4*inch)
        styles = getSampleStyleSheet()
        titulo_s = ParagraphStyle('t', fontSize=13, textColor=colors.HexColor('#1a5276'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
        sub_s = ParagraphStyle('s', fontSize=9, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')
        norm_s = ParagraphStyle('n', fontSize=8.5, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=0, fontName='Helvetica')

        story = []
        import base64 as b64mod
        try:
            logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7qy654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
            logo_buffer = io.BytesIO(logo_data)
            logo_img = RLImage(logo_buffer, width=0.8*inch, height=0.7*inch)
        except Exception:
            logo_img = None

        hdr = [Paragraph("PARTIDO CABAL", titulo_s),
               Paragraph("MUNICIPIO DE TOTONICAPAN", sub_s),
               Paragraph("PRESIDENTAS DE COMITÉ", titulo_s),
               Paragraph(f"<b>Jefe de Sector:</b> {jefe}", sub_s),
               Paragraph(f"<b>Total:</b> {len(presidentas)} presidenta(s)", sub_s),
               Paragraph(f"<b>Fecha:</b> {ahora_gt().strftime('%d/%m/%Y')}", sub_s)]
        ht = None
        if logo_img:
            ht = Table([[hdr, logo_img]], colWidths=[6.7*inch, 0.8*inch])
            ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
            story.append(ht)
        else:
            for item in hdr:
                story.append(item)
        story.append(Spacer(1,0.1*inch))
        story.append(HRFlowable(width="100%",thickness=2,color=colors.HexColor('#1a5276')))
        story.append(Spacer(1,0.1*inch))

        # Tabla de presidentas. La columna "Jefe de Sector" solo se agrega
        # si hay presidentas de más de un jefe en la lista (viendo "todos
        # los jefes"), para no repetir el mismo dato en cada fila cuando ya
        # se está viendo solo un jefe específico.
        celda_s = ParagraphStyle('celda', fontSize=8, alignment=TA_CENTER, fontName='Helvetica', leading=10)
        mostrar_jefe_col = len(set(p.get('jefe','').strip() for p in presidentas)) > 1
        if mostrar_jefe_col:
            table_data = [["No.", "Nombre", "DPI", "Teléfono", "Dirección", "Emp.", "Jefe de Sector"]]
        else:
            table_data = [["No.", "Nombre", "DPI", "Teléfono", "Dirección", "Emp."]]
        for i, p in enumerate(presidentas, 1):
            fila = [
                str(i),
                Paragraph(p.get('nombre','-'), celda_s),
                p.get('cui','-'),
                p.get('telefono','-'),
                Paragraph(p.get('direccion','-'), celda_s),
                p.get('empadronado','-'),
            ]
            if mostrar_jefe_col:
                fila.append(Paragraph(p.get('jefe','-'), celda_s))
            table_data.append(fila)

        if mostrar_jefe_col:
            col_widths = [0.3*inch, 1.7*inch, 1.3*inch, 1.0*inch, 1.6*inch, 0.45*inch, 1.15*inch]
        else:
            col_widths = [0.35*inch, 2.0*inch, 1.5*inch, 1.2*inch, 2.0*inch, 0.55*inch]
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a5276')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#eaf4fb')]),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        story.append(t)
        story.append(Spacer(1,0.15*inch))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#bdc3c7')))
        story.append(Spacer(1,0.05*inch))
        story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO",
            ParagraphStyle('pie',fontSize=7,textColor=colors.HexColor('#7f8c8d'),alignment=TA_CENTER,fontName='Helvetica')))

        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        token = _guardar_pdf_temporal(pdf_bytes, f"Presidentas_{jefe.replace(' ','_')}.pdf")
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/pdf_coordinadoras", methods=["POST"])
@requiere_sesion
def pdf_coordinadoras():
    try:
        data = request.json
        jefe = data.get("jefe_sector","")
        coordinadoras = data.get("coordinadoras",[])

        OFICIO_HORIZONTAL = (13*inch, 8.5*inch)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=OFICIO_HORIZONTAL,
                                rightMargin=0.4*inch, leftMargin=0.4*inch,
                                topMargin=0.35*inch, bottomMargin=0.35*inch)
        titulo_s = ParagraphStyle('t', fontSize=14, textColor=colors.HexColor('#1a5276'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
        sub_s = ParagraphStyle('s', fontSize=9.5, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')

        story = []
        import base64 as b64mod
        try:
            logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7yq654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
            logo_buffer = io.BytesIO(logo_data)
            logo_img = RLImage(logo_buffer, width=0.8*inch, height=0.7*inch)
        except Exception:
            logo_img = None

        hdr = [Paragraph("PARTIDO CABAL", titulo_s),
               Paragraph("MUNICIPIO DE TOTONICAPAN", sub_s),
               Paragraph("COORDINADORAS DE GRUPO", titulo_s),
               Paragraph(f"<b>Jefe de Sector:</b> {jefe}", sub_s),
               Paragraph(f"<b>Total:</b> {len(coordinadoras)} coordinadora(s)", sub_s),
               Paragraph(f"<b>Fecha:</b> {ahora_gt().strftime('%d/%m/%Y')}", sub_s)]
        ht = None
        if logo_img:
            ht = Table([[hdr, logo_img]], colWidths=[11.4*inch, 0.8*inch])
            ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
            story.append(ht)
        else:
            for item in hdr:
                story.append(item)
        story.append(Spacer(1,0.12*inch))
        story.append(HRFlowable(width="100%",thickness=2,color=colors.HexColor('#1a5276')))
        story.append(Spacer(1,0.12*inch))

        celda_s = ParagraphStyle('celda', fontSize=8, alignment=TA_CENTER, fontName='Helvetica', leading=9.5)
        table_data = [["No.", "Grupo", "Coordinadora", "DPI", "Teléfono", "Dirección", "Emp.", "Presidenta"]]
        for i, c in enumerate(coordinadoras, 1):
            table_data.append([
                str(i),
                Paragraph(c.get('grupo','-'), celda_s),
                Paragraph(c.get('nombre','-'), celda_s),
                c.get('cui','-'),
                c.get('telefono','-'),
                Paragraph(c.get('direccion','-'), celda_s),
                c.get('empadronado','-'),
                Paragraph(c.get('presidenta','-'), celda_s),
            ])

        col_widths = [0.35*inch, 1.9*inch, 1.9*inch, 1.3*inch, 1.05*inch, 2.2*inch, 0.5*inch, 3.0*inch]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a5276')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#eaf4fb')]),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        story.append(t)
        story.append(Spacer(1,0.15*inch))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#bdc3c7')))
        story.append(Spacer(1,0.05*inch))
        story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO",
            ParagraphStyle('pie',fontSize=7,textColor=colors.HexColor('#7f8c8d'),alignment=TA_CENTER,fontName='Helvetica')))

        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        token = _guardar_pdf_temporal(pdf_bytes, f"Coordinadoras_{jefe.replace(' ','_')}.pdf")
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _generar_excel_lista(titulo_hoja, jefe, columnas, filas, nombre_col_jefe=None):
    """Genera un archivo Excel (.xlsx) en memoria a partir de una lista de
    diccionarios, con encabezado con estilo. Reutilizado por presidentas y
    coordinadoras para no duplicar el formato. Devuelve los bytes del archivo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = titulo_hoja[:31]  # Excel limita el nombre de la hoja a 31 caracteres

    ws.merge_cells('A1:' + chr(64 + len(columnas)) + '1')
    ws['A1'] = f"{titulo_hoja} — {jefe}"
    ws['A1'].font = Font(bold=True, size=14, color='1A5276')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:' + chr(64 + len(columnas)) + '2')
    ws['A2'] = f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO"
    ws['A2'].font = Font(italic=True, size=9, color='7F8C8D')
    ws['A2'].alignment = Alignment(horizontal='center')

    header_row = 4
    borde = Border(*(Side(style='thin', color='BDC3C7'),) * 4)
    for col_i, nombre_col in enumerate(columnas, 1):
        celda = ws.cell(row=header_row, column=col_i, value=nombre_col)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill('solid', fgColor='1A5276')
        celda.alignment = Alignment(horizontal='center', vertical='center')
        celda.border = borde

    for i, fila in enumerate(filas, header_row + 1):
        for col_i, valor in enumerate(fila, 1):
            celda = ws.cell(row=i, column=col_i, value=valor)
            celda.border = borde
            celda.alignment = Alignment(horizontal='center', vertical='center')
            if i % 2 == 0:
                celda.fill = PatternFill('solid', fgColor='EAF4FB')

    for col_i in range(1, len(columnas) + 1):
        max_len = max([len(str(columnas[col_i-1]))] + [len(str(f[col_i-1])) for f in filas]) if filas else len(str(columnas[col_i-1]))
        ws.column_dimensions[chr(64 + col_i)].width = min(max(max_len + 4, 12), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@app.route("/excel_presidentas", methods=["POST"])
@requiere_sesion
def excel_presidentas():
    try:
        data = request.json
        jefe = data.get("jefe_sector", "")
        presidentas = data.get("presidentas", [])
        columnas = ["No.", "Nombre", "DPI", "Teléfono", "Dirección", "Empadronado"]
        if any('jefe' in p for p in presidentas):
            columnas.append("Jefe de Sector")
        filas = []
        for i, p in enumerate(presidentas, 1):
            fila = [i, p.get('nombre','-'), p.get('cui','-'), p.get('telefono','-'), p.get('direccion','-'), p.get('empadronado','-')]
            if "Jefe de Sector" in columnas:
                fila.append(p.get('jefe','-'))
            filas.append(fila)
        excel_bytes = _generar_excel_lista("PRESIDENTAS DE COMITÉ", jefe, columnas, filas)
        import base64 as b64enc
        return jsonify({"ok": True, "excel_b64": b64enc.b64encode(excel_bytes).decode('utf-8')})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/excel_coordinadoras", methods=["POST"])
@requiere_sesion
def excel_coordinadoras():
    try:
        data = request.json
        jefe = data.get("jefe_sector", "")
        coordinadoras = data.get("coordinadoras", [])
        columnas = ["No.", "Grupo", "Coordinadora", "DPI", "Teléfono", "Dirección", "Empadronado", "Presidenta"]
        if any('jefe' in c for c in coordinadoras):
            columnas.append("Jefe de Sector")
        filas = []
        for i, c in enumerate(coordinadoras, 1):
            fila = [i, c.get('grupo','-'), c.get('nombre','-'), c.get('cui','-'), c.get('telefono','-'), c.get('direccion','-'), c.get('empadronado','-'), c.get('presidenta','-')]
            if "Jefe de Sector" in columnas:
                fila.append(c.get('jefe','-'))
            filas.append(fila)
        excel_bytes = _generar_excel_lista("COORDINADORAS DE GRUPO", jefe, columnas, filas)
        import base64 as b64enc
        return jsonify({"ok": True, "excel_b64": b64enc.b64encode(excel_bytes).decode('utf-8')})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _clase_canvas_numerado():
    """Devuelve una clase de canvas de ReportLab que numera las páginas
    como 'Página X de Y' abajo a la derecha. Se necesita una clase (no una
    simple función) porque el total de páginas solo se sabe al terminar de
    armar todo el documento; esta clase guarda cada página en memoria y
    recién al final, cuando ya sabe cuántas hay en total, dibuja el número
    en cada una."""
    from reportlab.pdfgen import canvas as _pdfcanvas

    class _CanvasNumerado(_pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            _pdfcanvas.Canvas.__init__(self, *args, **kwargs)
            self._paginas_guardadas = []

        def showPage(self):
            self._paginas_guardadas.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            total = len(self._paginas_guardadas)
            for estado in self._paginas_guardadas:
                self.__dict__.update(estado)
                self.setFont('Helvetica', 8)
                self.setFillColor(colors.HexColor('#7f8c8d'))
                ancho_pagina = self._pagesize[0]
                self.drawRightString(ancho_pagina - 0.4*inch, 0.18*inch, f"Página {self._pageNumber} de {total}")
                _pdfcanvas.Canvas.showPage(self)
            _pdfcanvas.Canvas.save(self)

    return _CanvasNumerado


def _generar_pdf_reporte_bytes(jefe, registros, titulo_reporte, modo=""):
    """Construye el PDF con la plantilla OFICIAL del sistema (logo de
    CABAL, encabezado 'PARTIDO CABAL / MUNICIPIO DE TOTONICAPAN', tabla con
    todas las columnas incluyendo Teléfono, agrupado por Presidenta con
    salto de página entre cada una). La usan tanto la ruta /pdf_todo_combinado
    (cuando alguien lo pide desde la app) como el reporte diario automático
    por correo, para que ambos se vean exactamente igual. Devuelve los
    bytes del PDF ya armado."""
    OFICIO_HORIZONTAL = (13*inch, 8.5*inch)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=OFICIO_HORIZONTAL,
                            rightMargin=0.4*inch, leftMargin=0.4*inch,
                            topMargin=0.35*inch, bottomMargin=0.4*inch)
    titulo_s = ParagraphStyle('t', fontSize=14, textColor=colors.HexColor('#1a5276'),
        spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
    sub_s = ParagraphStyle('s', fontSize=9.5, textColor=colors.HexColor('#2c3e50'),
        spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')

    story = []
    import base64 as b64mod
    try:
        logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7yq654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
        logo_buffer_bytes = logo_data
        logo_img_ok = True
    except Exception:
        logo_img_ok = False

    def _agregar_encabezado(total_actual):
        if logo_img_ok:
            logo_img = RLImage(io.BytesIO(logo_buffer_bytes), width=0.8*inch, height=0.7*inch)
        else:
            logo_img = None
        hdr = [Paragraph("PARTIDO CABAL", titulo_s),
               Paragraph("MUNICIPIO DE TOTONICAPAN", sub_s),
               Paragraph(titulo_reporte, titulo_s),
               Paragraph(f"<b>Jefe de Sector:</b> {jefe}", sub_s),
               Paragraph(f"<b>Total:</b> {total_actual} registro(s)", sub_s),
               Paragraph(f"<b>Fecha:</b> {ahora_gt().strftime('%d/%m/%Y')}", sub_s)]
        if logo_img:
            ht = Table([[hdr, logo_img]], colWidths=[11.4*inch, 0.8*inch])
            ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
            story.append(ht)
        else:
            for item in hdr:
                story.append(item)
        story.append(Spacer(1,0.12*inch))
        story.append(HRFlowable(width="100%",thickness=2,color=colors.HexColor('#1a5276')))
        story.append(Spacer(1,0.12*inch))

    celda_s = ParagraphStyle('celda', fontSize=8, alignment=TA_CENTER, fontName='Helvetica', leading=9.5)
    modo = (modo or "").strip().lower()
    # Cuando el reporte junta a TODOS los jefes de sector (en vez de uno
    # solo), se agrega una columna "Jefe" en la tabla — si no, al mezclar a
    # varios jefes en un mismo PDF no hay forma de saber de quién es cada
    # grupo con solo mirar el encabezado (que dice "TODOS").
    incluir_jefe = 'todos' in (jefe or '').strip().lower()
    if modo == "observaciones":
        if incluir_jefe:
            encabezado_tabla = ["No.", "Jefe", "Tipo", "Nombre", "DPI", "Comunidad", "Presidenta", "Observaciones"]
            col_widths = [0.3*inch, 1.3*inch, 0.85*inch, 1.5*inch, 1.3*inch, 1.35*inch, 1.35*inch, 3.15*inch]
        else:
            encabezado_tabla = ["No.", "Tipo", "Nombre", "DPI", "Comunidad", "Presidenta", "Observaciones"]
            col_widths = [0.35*inch, 0.95*inch, 1.7*inch, 1.35*inch, 1.55*inch, 1.55*inch, 3.6*inch]
    else:
        if incluir_jefe:
            encabezado_tabla = ["No.", "Jefe", "Tipo", "Nombre", "DPI", "F. Nacimiento", "Teléfono", "Dirección", "Emp.", "Coordinadora", "Verif."]
            col_widths = [0.3*inch, 1.2*inch, 0.85*inch, 1.6*inch, 1.3*inch, 0.95*inch, 0.9*inch, 1.9*inch, 0.4*inch, 1.45*inch, 0.5*inch]
        else:
            encabezado_tabla = ["No.", "Tipo", "Nombre", "DPI", "F. Nacimiento", "Teléfono", "Dirección", "Emp.", "Coordinadora", "Verif."]
            col_widths = [0.3*inch, 1.0*inch, 1.9*inch, 1.3*inch, 1.0*inch, 0.95*inch, 2.3*inch, 0.4*inch, 1.75*inch, 0.5*inch]

    # La columna DPI está en la posición 3 normalmente, pero se corre a la
    # 4 cuando se agrega la columna "Jefe" al inicio de la tabla.
    col_idx_dpi = 4 if incluir_jefe else 3

    def _estilo_tabla(grupo_filas):
        comandos = [
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a5276')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#eaf4fb')]),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]
        for idx_fila, p in enumerate(grupo_filas, start=1):
            if _dpi_formato_invalido(p.get('cui','')):
                comandos.append(('BACKGROUND',(col_idx_dpi,idx_fila),(col_idx_dpi,idx_fila),colors.HexColor('#f9e79f')))
            if p.get('dpi_no_verificado'):
                comandos.append(('BACKGROUND',(0,idx_fila),(-1,idx_fila),colors.HexColor('#aed6f1')))
        return TableStyle(comandos)

    grupos_por_presidenta = []
    presidenta_actual = None
    for reg in registros:
        if reg.get('presidenta') != presidenta_actual:
            grupos_por_presidenta.append([])
            presidenta_actual = reg.get('presidenta')
        grupos_por_presidenta[-1].append(reg)

    if not grupos_por_presidenta:
        grupos_por_presidenta = [[]]

    for idx_grupo, grupo in enumerate(grupos_por_presidenta):
        if idx_grupo > 0:
            story.append(PageBreak())
        _agregar_encabezado(len(grupo))
        table_data = [encabezado_tabla]
        for i, p in enumerate(grupo, 1):
            if modo == "observaciones":
                fila = [str(i)]
                if incluir_jefe:
                    fila.append(Paragraph(p.get('jefe','-') or '-', celda_s))
                fila += [
                    Paragraph(p.get('tipo','-'), celda_s),
                    Paragraph(p.get('nombre','-'), celda_s),
                    p.get('cui','-'),
                    Paragraph(p.get('direccion','-'), celda_s),
                    Paragraph(p.get('presidenta','-'), celda_s),
                    Paragraph(p.get('observaciones','-') or '-', celda_s),
                ]
                table_data.append(fila)
            else:
                fila = [str(i)]
                if incluir_jefe:
                    fila.append(Paragraph(p.get('jefe','-') or '-', celda_s))
                fila += [
                    Paragraph(p.get('tipo','-'), celda_s),
                    Paragraph(p.get('nombre','-'), celda_s),
                    p.get('cui','-'),
                    p.get('fecha_nacimiento','-') or '-',
                    p.get('telefono','-') or 'FALTA',
                    Paragraph(p.get('direccion','-'), celda_s),
                    p.get('empadronado','-'),
                    Paragraph(p.get('coordinadora','-'), celda_s),
                    'NO' if p.get('dpi_no_verificado') else 'SI',
                ]
                table_data.append(fila)
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(_estilo_tabla(grupo))
        story.append(t)
    story.append(Spacer(1,0.15*inch))
    story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#bdc3c7')))
    story.append(Spacer(1,0.05*inch))
    story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO",
        ParagraphStyle('pie',fontSize=7.5,textColor=colors.HexColor('#7f8c8d'),alignment=TA_CENTER,fontName='Helvetica')))

    doc.build(story, canvasmaker=_clase_canvas_numerado())
    buffer.seek(0)
    return buffer.read()



_ORGANIGRAMA_ESTILO_NIVEL = {
    0: dict(fondo=colors.HexColor('#1a5276'), texto=colors.white,        f1=12, f2=8.5, w=2.6*inch, h=0.55*inch),
    1: dict(fondo=colors.HexColor('#2980b9'), texto=colors.white,        f1=10, f2=7.5, w=2.55*inch, h=1.22*inch),
    2: dict(fondo=colors.HexColor('#5dade2'), texto=colors.white,        f1=9,  f2=7,   w=2.3*inch, h=1.05*inch),
    3: dict(fondo=colors.HexColor('#eaf2f8'), texto=colors.HexColor('#1c2833'), f1=8.5, f2=7, w=2.05*inch, h=0.92*inch),
}
_ORGANIGRAMA_GAP_X = 0.22*inch
_ORGANIGRAMA_GAP_Y_NIVEL = [0.45*inch, 0.42*inch, 0.4*inch]


_ORGANIGRAMA_GAP_VERT_INTEGRANTES = 0.1*inch


def _organigrama_ancho_nodo(nodo):
    est = _ORGANIGRAMA_ESTILO_NIVEL[nodo['nivel']]
    if not nodo['hijos']:
        return est['w']
    if nodo['nivel'] == 2:
        # Las integrantes (hijos de una coordinadora) se dibujan apiladas
        # EN COLUMNA, una debajo de otra, no en fila — así que no
        # 'ensanchan' el árbol sin importar cuántas sean. El ancho de esta
        # rama es simplemente el mayor entre la caja de la coordinadora y
        # la de una integrante.
        return max(est['w'], _ORGANIGRAMA_ESTILO_NIVEL[3]['w'])
    total = sum(_organigrama_ancho_nodo(h) for h in nodo['hijos']) + _ORGANIGRAMA_GAP_X * (len(nodo['hijos']) - 1)
    return max(total, est['w'])


def _organigrama_alto_total(nodo):
    est = _ORGANIGRAMA_ESTILO_NIVEL[nodo['nivel']]
    if not nodo['hijos']:
        return est['h']
    if nodo['nivel'] == 2:
        # Altura de la columna de integrantes apiladas debajo
        est_hijo = _ORGANIGRAMA_ESTILO_NIVEL[3]
        n = len(nodo['hijos'])
        altura_columna = n * est_hijo['h'] + max(0, n - 1) * _ORGANIGRAMA_GAP_VERT_INTEGRANTES
        return est['h'] + _ORGANIGRAMA_GAP_Y_NIVEL[2] + altura_columna
    return est['h'] + _ORGANIGRAMA_GAP_Y_NIVEL[min(nodo['nivel'], len(_ORGANIGRAMA_GAP_Y_NIVEL)-1)] + max(_organigrama_alto_total(h) for h in nodo['hijos'])


def _organigrama_envolver_texto(c, texto, ancho_max, tam_fuente, fuente='Helvetica-Bold'):
    palabras = texto.split()
    lineas, actual = [], ''
    for palabra in palabras:
        prueba = (actual + ' ' + palabra).strip()
        if c.stringWidth(prueba, fuente, tam_fuente) <= ancho_max - 8:
            actual = prueba
        else:
            if actual:
                lineas.append(actual)
            actual = palabra
    if actual:
        lineas.append(actual)
    return lineas or ['']


_ORGANIGRAMA_COLOR_SIN_EMPADRONAR = colors.HexColor('#f4d03f')  # amarillo
_ORGANIGRAMA_COLOR_SIN_TELEFONO = colors.HexColor('#58d68d')     # verde
_ORGANIGRAMA_COLOR_AMBOS_FALTAN = colors.HexColor('#e67e22')     # naranja (le faltan las 2 cosas)
_ORGANIGRAMA_COLOR_TEXTO_ALERTA = colors.HexColor('#1c2833')     # texto oscuro, se lee mejor sobre estos colores claros


def _organigrama_dibujar_caja(c, x_centro, y_top, nodo):
    est = _ORGANIGRAMA_ESTILO_NIVEL[nodo['nivel']]
    w, h = est['w'], est['h']
    x0 = x_centro - w / 2
    y0 = y_top - h

    sin_emp = bool(nodo.get('sin_empadronar'))
    sin_tel = bool(nodo.get('sin_telefono'))
    if sin_emp and sin_tel:
        color_fondo, color_texto = _ORGANIGRAMA_COLOR_AMBOS_FALTAN, _ORGANIGRAMA_COLOR_TEXTO_ALERTA
    elif sin_emp:
        color_fondo, color_texto = _ORGANIGRAMA_COLOR_SIN_EMPADRONAR, _ORGANIGRAMA_COLOR_TEXTO_ALERTA
    elif sin_tel:
        color_fondo, color_texto = _ORGANIGRAMA_COLOR_SIN_TELEFONO, _ORGANIGRAMA_COLOR_TEXTO_ALERTA
    else:
        color_fondo, color_texto = est['fondo'], est['texto']

    c.setFillColor(color_fondo)
    c.roundRect(x0, y0, w, h, 4, stroke=0, fill=1)
    c.setStrokeColor(colors.HexColor('#7f8c8d'))
    c.setLineWidth(0.6)
    c.roundRect(x0, y0, w, h, 4, stroke=1, fill=0)

    # 'detalle' puede ser un solo texto o una lista de líneas (ej. DPI y
    # Teléfono en líneas separadas)
    detalle = nodo.get('detalle')
    lineas_detalle = detalle if isinstance(detalle, list) else ([detalle] if detalle else [])

    c.setFillColor(color_texto)
    lineas_nombre = _organigrama_envolver_texto(c, nodo['nombre'], w, est['f1'])[:2]
    interlineado = est['f1'] + 1.6
    interlineado_detalle = est['f2'] + 1.4
    bloque_nombre_h = len(lineas_nombre) * interlineado
    bloque_detalle_h = len(lineas_detalle) * interlineado_detalle + (3 if lineas_detalle else 0)
    y_inicio = y0 + h/2 + (bloque_nombre_h + bloque_detalle_h)/2 - est['f1']*0.75

    c.setFont('Helvetica-Bold', est['f1'])
    yy = y_inicio
    for linea in lineas_nombre:
        c.drawCentredString(x_centro, yy, linea)
        yy -= interlineado
    if lineas_detalle:
        c.setFont('Helvetica', est['f2'])
        yy -= 1
        for linea_d in lineas_detalle:
            c.drawCentredString(x_centro, yy, linea_d)
            yy -= interlineado_detalle


def _organigrama_dibujar_arbol(c, nodo, x_centro, y_top):
    _organigrama_dibujar_caja(c, x_centro, y_top, nodo)
    est = _ORGANIGRAMA_ESTILO_NIVEL[nodo['nivel']]
    if not nodo['hijos']:
        return

    if nodo['nivel'] == 2:
        # Integrantes en columna: una caja debajo de la otra, conectadas
        # por una sola línea vertical, en vez de en fila horizontal —
        # así no se aprietan aunque una coordinadora tenga muchas.
        est_hijo = _ORGANIGRAMA_ESTILO_NIVEL[3]
        c.setStrokeColor(colors.HexColor('#7f8c8d'))
        c.setLineWidth(0.9)
        y_padre_abajo = y_top - est['h']
        y_cursor_top = y_padre_abajo - _ORGANIGRAMA_GAP_Y_NIVEL[2]
        c.line(x_centro, y_padre_abajo, x_centro, y_cursor_top)
        for idx, hijo in enumerate(nodo['hijos']):
            _organigrama_dibujar_caja(c, x_centro, y_cursor_top, hijo)
            if idx < len(nodo['hijos']) - 1:
                y_linea_desde = y_cursor_top - est_hijo['h']
                y_linea_hasta = y_linea_desde - _ORGANIGRAMA_GAP_VERT_INTEGRANTES
                c.line(x_centro, y_linea_desde, x_centro, y_linea_hasta)
                y_cursor_top = y_linea_hasta
        return

    y_hijos_top = y_top - est['h'] - _ORGANIGRAMA_GAP_Y_NIVEL[min(nodo['nivel'], len(_ORGANIGRAMA_GAP_Y_NIVEL)-1)]

    anchos_hijos = [_organigrama_ancho_nodo(h) for h in nodo['hijos']]
    ancho_total_hijos = sum(anchos_hijos) + _ORGANIGRAMA_GAP_X * (len(nodo['hijos']) - 1)
    x_cursor = x_centro - ancho_total_hijos / 2
    centros_hijos = []
    for hijo, ancho_h in zip(nodo['hijos'], anchos_hijos):
        cx = x_cursor + ancho_h / 2
        centros_hijos.append(cx)
        x_cursor += ancho_h + _ORGANIGRAMA_GAP_X

    c.setStrokeColor(colors.HexColor('#7f8c8d'))
    c.setLineWidth(0.9)
    y_padre_abajo = y_top - est['h']
    y_barra = y_padre_abajo - _ORGANIGRAMA_GAP_Y_NIVEL[min(nodo['nivel'], len(_ORGANIGRAMA_GAP_Y_NIVEL)-1)] / 2
    c.line(x_centro, y_padre_abajo, x_centro, y_barra)
    if len(centros_hijos) > 1:
        c.line(min(centros_hijos), y_barra, max(centros_hijos), y_barra)
    for cx in centros_hijos:
        c.line(cx, y_barra, cx, y_hijos_top)

    for hijo, cx in zip(nodo['hijos'], centros_hijos):
        _organigrama_dibujar_arbol(c, hijo, cx, y_hijos_top)


def _organigrama_dibujar_encabezado(c, logo_img, ancho_pagina, alto_pagina, titulo_reporte, subtitulo):
    margen = 0.4 * inch
    if logo_img:
        c.drawImage(logo_img, ancho_pagina - margen - 0.7*inch, alto_pagina - margen - 0.62*inch,
                    width=0.7*inch, height=0.62*inch, mask='auto')
    c.setFillColor(colors.HexColor('#1a5276'))
    c.setFont('Helvetica-Bold', 15)
    c.drawCentredString(ancho_pagina/2, alto_pagina - margen - 0.24*inch, "PARTIDO CABAL")
    c.setFont('Helvetica', 9.5)
    c.setFillColor(colors.HexColor('#2c3e50'))
    c.drawCentredString(ancho_pagina/2, alto_pagina - margen - 0.4*inch, "MUNICIPIO DE TOTONICAPAN")
    c.setFont('Helvetica-Bold', 13)
    c.setFillColor(colors.HexColor('#1a5276'))
    c.drawCentredString(ancho_pagina/2, alto_pagina - margen - 0.62*inch, titulo_reporte)
    c.setFont('Helvetica', 9)
    c.setFillColor(colors.HexColor('#2c3e50'))
    c.drawCentredString(ancho_pagina/2, alto_pagina - margen - 0.8*inch,
                         f"Fecha: {ahora_gt().strftime('%d/%m/%Y')}" + (f"   |   {subtitulo}" if subtitulo else ""))
    c.setStrokeColor(colors.HexColor('#1a5276'))
    c.setLineWidth(1.4)
    c.line(margen, alto_pagina - margen - 0.92*inch, ancho_pagina - margen, alto_pagina - margen - 0.92*inch)


def _organigrama_dibujar_pie(c, ancho_pagina, pagina_num, total_paginas):
    margen = 0.4*inch
    x = margen
    for color, texto in [
        (_ORGANIGRAMA_COLOR_SIN_EMPADRONAR, "Sin empadronar"),
        (_ORGANIGRAMA_COLOR_SIN_TELEFONO, "Sin teléfono"),
        (_ORGANIGRAMA_COLOR_AMBOS_FALTAN, "Ambos"),
    ]:
        c.setFillColor(color)
        c.setStrokeColor(colors.HexColor('#7f8c8d'))
        c.setLineWidth(0.5)
        c.roundRect(x, margen*0.5 - 5, 12, 10, 2, stroke=1, fill=1)
        c.setFont('Helvetica', 7.5)
        c.setFillColor(colors.HexColor('#7f8c8d'))
        c.drawString(x + 16, margen*0.5, texto)
        x += 16 + c.stringWidth(texto, 'Helvetica', 7.5) + 14
    c.drawCentredString(ancho_pagina/2, margen*0.5,
                         f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO")
    c.drawRightString(ancho_pagina - margen, margen*0.5, f"Página {pagina_num} de {total_paginas}")


def _organigrama_nombre_grupo_base(nombre_grupo):
    """Quita el número final de un nombre de grupo (ej. 'Las Guadalupe 10'
    -> 'Las Guadalupe', 'Nueva Alianza 5' -> 'Nueva Alianza') para mostrar
    en la Presidenta el nombre 'base' del grupo, sin el número que
    distingue a cada Coordinadora en particular — la Presidenta no tiene
    un grupo propio guardado en ningún lado, así que esto es una máscara
    calculada, no un dato guardado."""
    import re
    return re.sub(r'[\s\-]*\d+\s*$', '', nombre_grupo or '').strip()


def _organigrama_numero_grupo(nombre_grupo):
    """Extrae el número final de un nombre de grupo (ej. 'Las Guadalupe
    10' -> 10), para poder ordenar a las Coordinadoras por el número de
    su grupo (1, 2, 3...) en vez de dejarlas en el orden en que aparecen
    en la hoja. Si el nombre no termina en número, devuelve None."""
    import re
    m = re.search(r'(\d+)\s*$', nombre_grupo or '')
    return int(m.group(1)) if m else None


def _organigrama_construir_grupos_de_presidenta(miembros):
    """De la lista de registros de una presidenta, arma su lista de
    'grupos' (cada Coordinadora+sus integrantes cuenta como 1 grupo, y las
    integrantes sin coordinadora asignada cuentan como 1 grupo extra) —
    esto es lo que se usa para no pasar de 3 grupos por hoja."""
    pres_reg = next((m for m in miembros if m.get('tipo') == 'Presidenta'), None)
    dpi_pres = pres_reg.get('cui', '-') if pres_reg else '-'
    tel_pres = (pres_reg.get('telefono') or 'Sin tel.') if pres_reg else '-'
    fnac_pres = (pres_reg.get('fecha_nacimiento') or 'Sin fecha') if pres_reg else '-'
    dir_pres = (pres_reg.get('direccion') or 'Sin comunidad') if pres_reg else '-'

    coords = {}
    orden_coords = []
    integrantes_directos = []
    for m in miembros:
        if m.get('tipo') == 'Presidenta':
            continue
        if m.get('tipo') == 'Coordinadora':
            nombre_c = m.get('nombre', '-')
            if nombre_c not in coords:
                coords[nombre_c] = {'reg': m, 'integrantes': []}
                orden_coords.append(nombre_c)
            else:
                coords[nombre_c]['reg'] = m
        else:
            coord_de = (m.get('coordinadora') or '').strip()
            if coord_de and coord_de != '-':
                if coord_de not in coords:
                    coords[coord_de] = {'reg': None, 'integrantes': []}
                    orden_coords.append(coord_de)
                coords[coord_de]['integrantes'].append(m)
            else:
                integrantes_directos.append(m)

    # Ordenar las coordinadoras por el número de su propio grupo (1, 2,
    # 3...), en vez de dejarlas en el orden en que aparecen en la hoja.
    # Las que no tienen número en su grupo van primero.
    def _clave_orden_coord(nombre_c):
        info_c = coords[nombre_c]
        nombre_grupo = (info_c['reg'].get('grupo') or '').strip() if info_c['reg'] else ''
        numero = _organigrama_numero_grupo(nombre_grupo)
        return (0, 0) if numero is None else (1, numero)
    orden_coords.sort(key=_clave_orden_coord)

    grupos = []
    nombres_grupos_presidenta = []  # nombres de "Grupo" (nombre_grupo) que aparecen bajo esta presidenta
    for nombre_c in orden_coords:
        info = coords[nombre_c]
        dpi_c = info['reg'].get('cui', '-') if info['reg'] else '-'
        tel_c = (info['reg'].get('telefono') or 'Sin tel.') if info['reg'] else '-'
        fnac_c = (info['reg'].get('fecha_nacimiento') or 'Sin fecha') if info['reg'] else '-'
        emp_c = (info['reg'].get('empadronado') or 'NO') if info['reg'] else 'NO'
        nombre_grupo_c = (info['reg'].get('grupo') or '').strip() if info['reg'] else ''
        if nombre_grupo_c:
            nombre_grupo_base = _organigrama_nombre_grupo_base(nombre_grupo_c)
            if nombre_grupo_base and nombre_grupo_base not in nombres_grupos_presidenta:
                nombres_grupos_presidenta.append(nombre_grupo_base)
        hijos_integrantes = []
        for integ in info['integrantes']:
            hijos_integrantes.append({'nivel': 3, 'nombre': integ.get('nombre', '-'),
                                       'detalle': [f"DPI: {integ.get('cui', '-')}", f"Tel: {integ.get('telefono') or 'Sin tel.'}",
                                                   f"F.Nac: {integ.get('fecha_nacimiento') or 'Sin fecha'}"],
                                       'sin_empadronar': (integ.get('empadronado') or 'NO').strip().upper() != 'SI',
                                       'sin_telefono': not (integ.get('telefono') or '').strip(),
                                       'hijos': []})
        detalle_coord = [f'Grupo: {nombre_grupo_c[:26]}{"…" if len(nombre_grupo_c) > 26 else ""}'] if nombre_grupo_c else []
        detalle_coord += [f'DPI: {dpi_c}', f'Tel: {tel_c}', f'F.Nac: {fnac_c}']
        grupos.append({'nivel': 2, 'nombre': nombre_c, 'detalle': detalle_coord,
                        'sin_empadronar': (emp_c or 'NO').strip().upper() != 'SI',
                        'sin_telefono': not (info['reg'].get('telefono') or '').strip() if info['reg'] else True,
                        'hijos': hijos_integrantes})

    if integrantes_directos:
        hijos_directos = []
        for integ in integrantes_directos:
            hijos_directos.append({'nivel': 3, 'nombre': integ.get('nombre', '-'),
                                    'detalle': [f"DPI: {integ.get('cui', '-')}", f"Tel: {integ.get('telefono') or 'Sin tel.'}",
                                                f"F.Nac: {integ.get('fecha_nacimiento') or 'Sin fecha'}"],
                                    'sin_empadronar': (integ.get('empadronado') or 'NO').strip().upper() != 'SI',
                                    'sin_telefono': not (integ.get('telefono') or '').strip(),
                                    'hijos': []})
        grupos.append({'nivel': 2, 'nombre': 'Integrantes directas', 'detalle': 'Sin coordinadora asignada', 'hijos': hijos_directos})

    emp_pres = ((pres_reg.get('empadronado') or 'NO') if pres_reg else 'NO').strip().upper() != 'SI'
    tel_falta_pres = not (pres_reg.get('telefono') or '').strip() if pres_reg else True
    return dpi_pres, tel_pres, fnac_pres, dir_pres, grupos, nombres_grupos_presidenta, emp_pres, tel_falta_pres


def _organigrama_construir_nodo_jefe(nombre_jefe, estructura):
    """Arma el nodo raíz (Jefe de Sector) con sus Presidentas como hijos, a
    partir de una 'estructura' (lista de (nombre_presidenta, dpi, tel,
    fnac, dir, nombres_grupos, sin_emp_pres, tel_falta_pres,
    grupos_de_esa_presidenta)). Se usa tanto para calcular si algo cabe
    bien en una hoja como para dibujarlo de verdad."""
    nodo_jefe = {'nivel': 0, 'nombre': nombre_jefe.upper(), 'detalle': 'JEFE DE SECTOR', 'hijos': []}
    for nombre_presidenta, dpi_pres, tel_pres, fnac_pres, dir_pres, nombres_grupos_pres, sin_emp_pres, tel_falta_pres, grupos_pagina in estructura:
        detalle_pres = []
        if nombres_grupos_pres:
            etiqueta_grupos = 'Grupo: ' if len(nombres_grupos_pres) == 1 else 'Grupos: '
            texto_grupos = ', '.join(nombres_grupos_pres)
            if len(texto_grupos) > 30:
                texto_grupos = texto_grupos[:28] + '…'
            detalle_pres.append(etiqueta_grupos + texto_grupos)
        detalle_pres += [f'Comunidad: {dir_pres}', f'DPI: {dpi_pres}', f'Tel: {tel_pres}', f'F.Nac: {fnac_pres}']
        nodo_pres = {'nivel': 1, 'nombre': nombre_presidenta, 'detalle': detalle_pres,
                     'sin_empadronar': sin_emp_pres, 'sin_telefono': tel_falta_pres, 'hijos': grupos_pagina}
        nodo_jefe['hijos'].append(nodo_pres)
    return nodo_jefe


def _organigrama_agrupar_por_presidenta(unidades):
    """Reagrupa una lista plana de unidades (nombre_presidenta, dpi, tel,
    fnac, dir, nombres_grupos, sin_emp_pres, tel_falta_pres, grupo_o_None)
    en la 'estructura' que espera _organigrama_construir_nodo_jefe,
    conservando el orden de aparición."""
    por_presidenta = {}
    orden = []
    for nombre_presidenta, dpi_pres, tel_pres, fnac_pres, dir_pres, nombres_grupos_pres, sin_emp_pres, tel_falta_pres, grupo in unidades:
        if nombre_presidenta not in por_presidenta:
            por_presidenta[nombre_presidenta] = (dpi_pres, tel_pres, fnac_pres, dir_pres, nombres_grupos_pres, sin_emp_pres, tel_falta_pres, [])
            orden.append(nombre_presidenta)
        if grupo is not None:
            por_presidenta[nombre_presidenta][7].append(grupo)
    return [(np,) + por_presidenta[np] for np in orden]


def _generar_pdf_organigrama_bytes(jefe, registros, titulo_reporte="ORGANIGRAMA DE ESTRUCTURA"):
    """Genera el reporte en forma de organigrama gráfico real (cajas
    conectadas con líneas): Jefe de Sector arriba, cada Presidenta debajo,
    sus Coordinadoras debajo de ella, y las Integrantes de cada
    Coordinadora debajo de esa coordinadora. Usa el mismo logo y
    encabezado oficial que los demás reportes. En vez de un número fijo de
    grupos por hoja, va agregando grupos a cada página mientras el árbol
    siga cabiendo bien (sin encogerse demasiado); en cuanto ya no cabe más,
    pasa a una hoja nueva — así se aprovecha el espacio real de cada hoja.
    Nunca mezcla dos Jefes de Sector distintos en la misma hoja; si un
    mismo jefe necesita más de una hoja, repite su encabezado marcando
    '(continuación)'. La caja de cada Presidenta muestra su propia
    Comunidad (tomada de su dirección registrada)."""
    import base64 as b64mod
    from reportlab.pdfgen import canvas as _pdfcanvas_organigrama
    from reportlab.lib.utils import ImageReader

    buffer = io.BytesIO()
    pagesize = landscape(letter)
    c = _pdfcanvas_organigrama.Canvas(buffer, pagesize=pagesize)

    try:
        logo_img = ImageReader(io.BytesIO(b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7qy654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")))
    except Exception:
        logo_img = None

    incluir_todos = 'todos' in (jefe or '').strip().lower()
    if incluir_todos:
        por_jefe = {}
        orden_jefes = []
        for r in registros:
            j = (r.get('jefe') or 'Sin Jefe Asignado').strip() or 'Sin Jefe Asignado'
            if j not in por_jefe:
                por_jefe[j] = []
                orden_jefes.append(j)
            por_jefe[j].append(r)
    else:
        orden_jefes = [jefe]
        por_jefe = {jefe: registros}

    # El área disponible para dibujar el árbol es la misma en todas las
    # páginas (depende solo del tamaño de hoja y los márgenes), así que se
    # calcula una sola vez y se reutiliza tanto para decidir cuántos
    # grupos caben por hoja como para dibujarlos después.
    margen = 0.4*inch
    area_x0, area_x1 = margen, pagesize[0] - margen
    area_y_top = pagesize[1] - margen - 1.05*inch
    area_y_bottom = margen + 0.3*inch
    area_w = area_x1 - area_x0
    area_h = area_y_top - area_y_bottom
    _ORGANIGRAMA_ESCALA_MINIMA = 0.45  # por debajo de esto, mejor pasar a otra hoja que verse chiquito

    paginas = []
    for nombre_jefe in orden_jefes:
        regs_jefe = por_jefe[nombre_jefe]
        grupos_pres = {}
        orden_presidentas = []
        for r in regs_jefe:
            p = (r.get('presidenta') or 'Sin Presidenta Asignada').strip() or 'Sin Presidenta Asignada'
            if p not in grupos_pres:
                grupos_pres[p] = []
                orden_presidentas.append(p)
            grupos_pres[p].append(r)

        if not orden_presidentas:
            paginas.append((nombre_jefe, [], False))
            continue

        # Se van agregando los GRUPOS de UNA MISMA presidenta a la página
        # actual mientras el árbol siga cabiendo sin encogerse demasiado
        # (así se aprovecha el espacio si a una presidenta le caben varios
        # grupos en una sola hoja). Pero en cuanto se pasa a la SIGUIENTE
        # presidenta, siempre se empieza una hoja nueva — nunca se mezclan
        # dos presidentas distintas en la misma página.
        primera_pagina_de_este_jefe = True
        for nombre_presidenta in orden_presidentas:
            dpi_pres, tel_pres, fnac_pres, dir_pres, grupos, nombres_grupos_pres, sin_emp_pres, tel_falta_pres = _organigrama_construir_grupos_de_presidenta(grupos_pres[nombre_presidenta])
            if not grupos:
                unidades_pres = [(nombre_presidenta, dpi_pres, tel_pres, fnac_pres, dir_pres, nombres_grupos_pres, sin_emp_pres, tel_falta_pres, None)]
            else:
                unidades_pres = [(nombre_presidenta, dpi_pres, tel_pres, fnac_pres, dir_pres, nombres_grupos_pres, sin_emp_pres, tel_falta_pres, g) for g in grupos]

            pagina_actual = []
            for unidad in unidades_pres:
                candidata = pagina_actual + [unidad]
                nodo_prueba = _organigrama_construir_nodo_jefe(nombre_jefe, _organigrama_agrupar_por_presidenta(candidata))
                ancho_prueba = _organigrama_ancho_nodo(nodo_prueba)
                alto_prueba = _organigrama_alto_total(nodo_prueba)
                escala_prueba = min(area_w / ancho_prueba, area_h / alto_prueba, 1.0) if ancho_prueba and alto_prueba else 1.0
                if escala_prueba >= _ORGANIGRAMA_ESCALA_MINIMA or not pagina_actual:
                    # Cabe bien (o es la primera unidad de la hoja — esa
                    # siempre se acepta, para no quedar en un bucle
                    # infinito con un grupo enorme que nunca "cabe bien"
                    # solo).
                    pagina_actual = candidata
                else:
                    paginas.append((nombre_jefe, _organigrama_agrupar_por_presidenta(pagina_actual), not primera_pagina_de_este_jefe))
                    primera_pagina_de_este_jefe = False
                    pagina_actual = [unidad]
            if pagina_actual:
                paginas.append((nombre_jefe, _organigrama_agrupar_por_presidenta(pagina_actual), not primera_pagina_de_este_jefe))
                primera_pagina_de_este_jefe = False

    total_paginas = len(paginas)
    for idx_pag, (nombre_jefe, estructura, es_continuacion) in enumerate(paginas):
        subt = f"Jefe de Sector: {nombre_jefe}" + ("  (continuación)" if es_continuacion else "")
        _organigrama_dibujar_encabezado(c, logo_img, pagesize[0], pagesize[1], titulo_reporte, subt)

        nodo_jefe = _organigrama_construir_nodo_jefe(nombre_jefe, estructura)

        if nodo_jefe['hijos']:
            ancho_arbol = _organigrama_ancho_nodo(nodo_jefe)
            alto_arbol = _organigrama_alto_total(nodo_jefe)
            escala = min(area_w / ancho_arbol, area_h / alto_arbol, 1.0) if ancho_arbol and alto_arbol else 1.0
            c.saveState()
            c.translate(area_x0 + area_w/2, area_y_top)
            c.scale(escala, escala)
            _organigrama_dibujar_arbol(c, nodo_jefe, 0, 0)
            c.restoreState()
        else:
            c.saveState()
            c.translate(area_x0 + area_w/2, area_y_top)
            _organigrama_dibujar_caja(c, 0, 0, nodo_jefe)
            c.setFont('Helvetica-Oblique', 10)
            c.setFillColor(colors.HexColor('#7f8c8d'))
            c.drawCentredString(0, -0.9*inch, "(Sin presidentas ni integrantes registrados para este jefe)")
            c.restoreState()

        _organigrama_dibujar_pie(c, pagesize[0], idx_pag + 1, total_paginas)
        c.showPage()

    c.save()
    buffer.seek(0)
    return buffer.read()



def _obtener_comunidades_jefes(sh):
    """Devuelve un diccionario {NOMBRE_JEFE_EN_MAYUSCULAS: comunidad} para
    poder mostrar, en el organigrama, a qué comunidad pertenece cada jefe
    de sector."""
    resultado = {}
    try:
        ws = _get_ws_jefes(sh)
        for fila in ws.get_all_values()[1:]:
            if fila and fila[0].strip():
                resultado[fila[0].strip().upper()] = fila[6].strip() if len(fila) > 6 else ''
    except Exception:
        pass
    return resultado


@app.route("/pdf_organigrama", methods=["POST"])
@requiere_sesion
def pdf_organigrama():
    try:
        data = request.json
        jefe = data.get("jefe_sector", "")
        registros = data.get("registros", [])
        titulo_reporte = data.get("titulo", "ORGANIGRAMA DE ESTRUCTURA")

        pdf_bytes = _generar_pdf_organigrama_bytes(jefe, registros, titulo_reporte)
        filename = f"Organigrama_{jefe.replace(' ','_')}.pdf"
        token = _guardar_pdf_temporal(pdf_bytes, filename)
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/pdf_todo_combinado", methods=["POST"])
@requiere_sesion
def pdf_todo_combinado():
    try:
        data = request.json
        jefe = data.get("jefe_sector","")
        registros = data.get("registros",[])
        titulo_reporte = data.get("titulo", "PRESIDENTAS + COORDINADORAS + INTEGRANTES")
        modo = data.get("modo", "").strip().lower()

        pdf_bytes = _generar_pdf_reporte_bytes(jefe, registros, titulo_reporte, modo)
        filename = f"Presidentas_Coordinadoras_Integrantes_{jefe.replace(' ','_')}.pdf"
        token = _guardar_pdf_temporal(pdf_bytes, filename)
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/excel_todo_combinado", methods=["POST"])
@requiere_sesion
def excel_todo_combinado():
    try:
        data = request.json
        jefe = data.get("jefe_sector", "")
        registros = data.get("registros", [])
        titulo_reporte = data.get("titulo", "PRESIDENTAS + COORDINADORAS + INTEGRANTES")
        modo = data.get("modo", "").strip().lower()
        if modo == "observaciones":
            columnas = ["No.", "Tipo", "Nombre", "DPI", "Comunidad", "Presidenta", "Observaciones"]
            filas = []
            for i, p in enumerate(registros, 1):
                filas.append([i, p.get('tipo','-'), p.get('nombre','-'), p.get('cui','-'), p.get('direccion','-'),
                              p.get('presidenta','-'), p.get('observaciones','-') or '-'])
        else:
            columnas = ["No.", "Tipo", "Nombre", "DPI", "F. Nacimiento", "Teléfono", "Dirección", "Empadronado", "Coordinadora", "DPI Verificado"]
            filas = []
            for i, p in enumerate(registros, 1):
                filas.append([i, p.get('tipo','-'), p.get('nombre','-'), p.get('cui','-'), p.get('fecha_nacimiento','-') or '-', p.get('telefono','-'), p.get('direccion','-'),
                              p.get('empadronado','-'), p.get('coordinadora','-'), 'NO' if p.get('dpi_no_verificado') else 'SI'])
        excel_bytes = _generar_excel_lista(titulo_reporte, jefe, columnas, filas)
        import base64 as b64enc
        return jsonify({"ok": True, "excel_b64": b64enc.b64encode(excel_bytes).decode('utf-8')})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/pdf_integrantes", methods=["POST"])
@requiere_sesion
def pdf_integrantes():
    try:
        data = request.json
        jefe = data.get("jefe_sector","")
        integrantes = data.get("integrantes",[])

        # Tamaño oficio horizontal (13 x 8.5 pulgadas), ya que esta tabla
        # lleva más columnas que las de presidentas/coordinadoras.
        OFICIO_HORIZONTAL = (13*inch, 8.5*inch)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=OFICIO_HORIZONTAL,
                                rightMargin=0.4*inch, leftMargin=0.4*inch,
                                topMargin=0.35*inch, bottomMargin=0.35*inch)
        titulo_s = ParagraphStyle('t', fontSize=14, textColor=colors.HexColor('#1a5276'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
        sub_s = ParagraphStyle('s', fontSize=9.5, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')

        story = []
        import base64 as b64mod
        try:
            logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7yq654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
            logo_buffer = io.BytesIO(logo_data)
            logo_img = RLImage(logo_buffer, width=0.8*inch, height=0.7*inch)
        except Exception:
            logo_img = None

        hdr = [Paragraph("PARTIDO CABAL", titulo_s),
               Paragraph("MUNICIPIO DE TOTONICAPAN", sub_s),
               Paragraph("REPORTE GENERAL DE INTEGRANTES", titulo_s),
               Paragraph(f"<b>Jefe de Sector:</b> {jefe}", sub_s),
               Paragraph(f"<b>Total:</b> {len(integrantes)} integrante(s)", sub_s),
               Paragraph(f"<b>Fecha:</b> {ahora_gt().strftime('%d/%m/%Y')}", sub_s)]
        ht = None
        if logo_img:
            ht = Table([[hdr, logo_img]], colWidths=[11.4*inch, 0.8*inch])
            ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
            story.append(ht)
        else:
            for item in hdr:
                story.append(item)
        story.append(Spacer(1,0.12*inch))
        story.append(HRFlowable(width="100%",thickness=2,color=colors.HexColor('#1a5276')))
        story.append(Spacer(1,0.12*inch))

        celda_s = ParagraphStyle('celda', fontSize=8, alignment=TA_CENTER, fontName='Helvetica', leading=9.5)
        table_data = [["No.", "Nombre", "DPI", "Teléfono", "Dirección", "Emp.", "Coordinadora", "Presidenta"]]
        for i, p in enumerate(integrantes, 1):
            table_data.append([
                str(i),
                Paragraph(p.get('nombre','-'), celda_s),
                p.get('cui','-'),
                p.get('telefono','-'),
                Paragraph(p.get('direccion','-'), celda_s),
                p.get('empadronado','-'),
                Paragraph(p.get('coordinadora','-'), celda_s),
                Paragraph(p.get('presidenta','-'), celda_s),
            ])

        # Ancho util en oficio horizontal: 13 - 0.8 = 12.2 pulgadas
        col_widths = [0.4*inch, 2.2*inch, 1.5*inch, 1.2*inch, 2.8*inch, 0.6*inch, 1.85*inch, 1.85*inch]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a5276')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#eaf4fb')]),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        story.append(t)
        story.append(Spacer(1,0.15*inch))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#bdc3c7')))
        story.append(Spacer(1,0.05*inch))
        story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO",
            ParagraphStyle('pie',fontSize=7.5,textColor=colors.HexColor('#7f8c8d'),alignment=TA_CENTER,fontName='Helvetica')))

        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        filename = f"Integrantes_{jefe.replace(' ','_')}.pdf"
        token = _guardar_pdf_temporal(pdf_bytes, filename)
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/excel_integrantes", methods=["POST"])
@requiere_sesion
def excel_integrantes():
    try:
        data = request.json
        jefe = data.get("jefe_sector", "")
        integrantes = data.get("integrantes", [])
        columnas = ["No.", "Nombre", "DPI", "Teléfono", "Dirección", "Empadronado", "Coordinadora", "Presidenta"]
        filas = []
        for i, p in enumerate(integrantes, 1):
            filas.append([i, p.get('nombre','-'), p.get('cui','-'), p.get('telefono','-'), p.get('direccion','-'),
                          p.get('empadronado','-'), p.get('coordinadora','-'), p.get('presidenta','-')])
        excel_bytes = _generar_excel_lista("REPORTE GENERAL DE INTEGRANTES", jefe, columnas, filas)
        import base64 as b64enc
        return jsonify({"ok": True, "excel_b64": b64enc.b64encode(excel_bytes).decode('utf-8')})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/pdf_presidentas_coordinadoras", methods=["POST"])
@requiere_sesion
def pdf_presidentas_coordinadoras():
    try:
        data = request.json
        jefe = data.get("jefe_sector","")
        registros = data.get("registros",[])

        OFICIO_HORIZONTAL = (13*inch, 8.5*inch)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=OFICIO_HORIZONTAL,
                                rightMargin=0.4*inch, leftMargin=0.4*inch,
                                topMargin=0.35*inch, bottomMargin=0.35*inch)
        titulo_s = ParagraphStyle('t', fontSize=14, textColor=colors.HexColor('#1a5276'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
        sub_s = ParagraphStyle('s', fontSize=9.5, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=2, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')

        story = []
        import base64 as b64mod
        try:
            logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7yq654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
            logo_buffer = io.BytesIO(logo_data)
            logo_img = RLImage(logo_buffer, width=0.8*inch, height=0.7*inch)
        except Exception:
            logo_img = None

        hdr = [Paragraph("PARTIDO CABAL", titulo_s),
               Paragraph("MUNICIPIO DE TOTONICAPAN", sub_s),
               Paragraph("PRESIDENTAS Y COORDINADORAS", titulo_s),
               Paragraph(f"<b>Jefe de Sector:</b> {jefe}", sub_s),
               Paragraph(f"<b>Total:</b> {len(registros)} registro(s)", sub_s),
               Paragraph(f"<b>Fecha:</b> {ahora_gt().strftime('%d/%m/%Y')}", sub_s)]
        ht = None
        if logo_img:
            ht = Table([[hdr, logo_img]], colWidths=[11.4*inch, 0.8*inch])
            ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
                ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
                ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
            story.append(ht)
        else:
            for item in hdr:
                story.append(item)
        story.append(Spacer(1,0.12*inch))
        story.append(HRFlowable(width="100%",thickness=2,color=colors.HexColor('#1a5276')))
        story.append(Spacer(1,0.12*inch))

        celda_s = ParagraphStyle('celda', fontSize=8, alignment=TA_CENTER, fontName='Helvetica', leading=9.5)
        table_data = [["No.", "Tipo", "Nombre", "DPI", "Teléfono", "Dirección", "Emp.", "Grupo", "Presidenta"]]
        for i, p in enumerate(registros, 1):
            table_data.append([
                str(i),
                p.get('tipo','-'),
                Paragraph(p.get('nombre','-'), celda_s),
                p.get('cui','-'),
                p.get('telefono','-'),
                Paragraph(p.get('direccion','-'), celda_s),
                p.get('empadronado','-'),
                Paragraph(p.get('grupo','-') or '-', celda_s),
                Paragraph(p.get('presidenta','-'), celda_s),
            ])

        # Ancho util en oficio horizontal: 13 - 0.8 = 12.2 pulgadas
        col_widths = [0.3*inch, 0.9*inch, 1.9*inch, 1.2*inch, 1.0*inch, 2.0*inch, 0.5*inch, 1.7*inch, 2.7*inch]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a5276')),
            ('TEXTCOLOR',(0,0),(-1,0),colors.white),
            ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#eaf4fb')]),
            ('TOPPADDING',(0,0),(-1,-1),5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ]))
        story.append(t)
        story.append(Spacer(1,0.15*inch))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#bdc3c7')))
        story.append(Spacer(1,0.05*inch))
        story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO",
            ParagraphStyle('pie',fontSize=7.5,textColor=colors.HexColor('#7f8c8d'),alignment=TA_CENTER,fontName='Helvetica')))

        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        filename = f"Presidentas_Coordinadoras_{jefe.replace(' ','_')}.pdf"
        token = _guardar_pdf_temporal(pdf_bytes, filename)
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/excel_presidentas_coordinadoras", methods=["POST"])
@requiere_sesion
def excel_presidentas_coordinadoras():
    try:
        data = request.json
        jefe = data.get("jefe_sector", "")
        registros = data.get("registros", [])
        columnas = ["No.", "Tipo", "Nombre", "DPI", "Teléfono", "Dirección", "Empadronado", "Grupo", "Presidenta"]
        filas = []
        for i, p in enumerate(registros, 1):
            filas.append([i, p.get('tipo','-'), p.get('nombre','-'), p.get('cui','-'), p.get('telefono','-'),
                          p.get('direccion','-'), p.get('empadronado','-'), p.get('grupo','-') or '-', p.get('presidenta','-')])
        excel_bytes = _generar_excel_lista("PRESIDENTAS Y COORDINADORAS", jefe, columnas, filas)
        import base64 as b64enc
        return jsonify({"ok": True, "excel_b64": b64enc.b64encode(excel_bytes).decode('utf-8')})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_presidentas")
@requiere_sesion
def ver_presidentas():
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe = "" if _es_admin else _nombre_sesion
        sh = get_sheet()
        try:
            ws = sh.worksheet("PRESIDENTAS")
        except:
            return jsonify({"ok": True, "presidentas": []})
        filas = ws.get_all_values()
        presidentas = []
        for fila in filas[1:]:
            if not fila or not fila[0]:
                continue
            if not jefe or fila[0].strip().upper() == jefe.upper():
                presidentas.append({
                    "jefe": fila[0].strip(),
                    "nombre": fila[1].strip(),
                    "cui": fila[2].strip() if len(fila) > 2 else '',
                    "telefono": fila[3].strip() if len(fila) > 3 else '',
                    "direccion": fila[4].strip() if len(fila) > 4 else '',
                    "empadronado": fila[5].strip() if len(fila) > 5 else '',
                })
        presidentas.sort(key=lambda p: (p["direccion"].upper(), p["nombre"].upper()))
        return jsonify({"ok": True, "presidentas": presidentas})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ficha_pdf_directo")
@requiere_sesion
def ficha_pdf_directo():
    """Misma ficha pero devuelve PDF directo para Google Docs Viewer"""
    from flask import make_response
    try:
        cui = request.args.get('cui','').strip()
        # Reutilizar la misma lógica llamando a ficha_pdf internamente
        # pero devolviendo el PDF directamente
        response = ficha_pdf()
        data = response.get_json()
        if data and data.get('ok') and data.get('pdf_b64'):
            import base64 as b64dec
            pdf_bytes = b64dec.b64decode(data['pdf_b64'])
            resp = make_response(pdf_bytes)
            resp.headers['Content-Type'] = 'application/pdf'
            resp.headers['Content-Disposition'] = f'inline; filename=Ficha_{cui}.pdf'
            return resp
        return jsonify({"ok": False, "error": "No se pudo generar"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ficha_pdf")
@requiere_sesion
def ficha_pdf():
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        cui = request.args.get('cui','').strip()
        sh = get_sheet()
        ws_dpi = sh.worksheet(HOJA)
        filas = ws_dpi.get_all_values()
        # Limpiar CUI para buscar sin importar espacios
        cui_c = cui.replace(' ','').replace('-','').strip()
        fila = None
        for f in filas[1:]:
            if f and f[0].replace(' ','').replace('-','').strip() == cui_c:
                fila = f
                break
        if not fila:
            return jsonify({"ok": False, "error": "Registro no encontrado"})
        if not _es_admin:
            jefe_fila = fila[19].strip().upper() if len(fila) > 19 else ''
            if jefe_fila != _nombre_sesion.upper():
                return jsonify({"ok": False, "error": "Registro no encontrado"})

        # Detectar índices correctos usando encabezado de hoja DPI
        header_dpi = filas[0] if filas else []
        def get_col(header, keywords):
            for k in keywords:
                for i, h in enumerate(header):
                    if k.lower() in str(h).lower():
                        return i
            return -1

        idx_serie    = get_col(header_dpi, ['serie']) 
        idx_pnombre  = get_col(header_dpi, ['primer nombre','primer_nombre'])
        idx_snombre  = get_col(header_dpi, ['segundo nombre','segundo_nombre'])
        idx_papell   = get_col(header_dpi, ['primer apellido','primer_apellido'])
        idx_sapell   = get_col(header_dpi, ['segundo apellido','segundo_apellido'])
        idx_sexo     = get_col(header_dpi, ['sexo','genero'])
        idx_ecivil   = get_col(header_dpi, ['civil'])
        idx_fnac     = get_col(header_dpi, ['nacimiento','nac'])
        idx_munnic   = get_col(header_dpi, ['municipio nac','municipio_nac'])
        idx_depnac   = get_col(header_dpi, ['departamento nac','departamento_nac'])
        idx_munvec   = get_col(header_dpi, ['municipio vec','municipio_vec'])
        idx_depvec   = get_col(header_dpi, ['departamento vec','departamento_vec'])
        idx_fexp     = get_col(header_dpi, ['expedicion','expedición'])
        idx_fvenc    = get_col(header_dpi, ['vencimiento'])
        idx_emp      = get_col(header_dpi, ['empadronado'])
        idx_noemp    = get_col(header_dpi, ['no. empadron','num empadron','empadronamiento'])
        idx_dir      = get_col(header_dpi, ['direcc','comunidad'])
        idx_tel      = get_col(header_dpi, ['telefono','teléfono'])
        idx_jefe     = get_col(header_dpi, ['jefe','jefa sector','jefe sector'])
        idx_pres     = get_col(header_dpi, ['presidenta','nombre jefa','nombre_jefa'])
        idx_rol      = get_col(header_dpi, ['rol','tipo','cargo'])

        def gf(idx): return fila[idx].strip() if idx >= 0 and idx < len(fila) else ''
        # Usar cui limpio para el PDF
        cui = cui_c

        # Buscar en GRUPOS para obtener grupo, coordinadora, jefe y presidenta
        coord = ''
        grupo = ''
        presidenta = gf(idx_pres)
        jefe = gf(idx_jefe)
        rol = gf(idx_rol) if idx_rol >= 0 else ''
        try:
            ws_g = sh.worksheet("GRUPOS")
            gs = ws_g.get_all_values()
            header_g = gs[0] if gs else []
            # Detectar formato de GRUPOS
            tiene_pres_g = any('residenta' in str(h) for h in header_g)
            tiene_dir_g  = any('irecc' in str(h) for h in header_g)
            # Detectar índices en GRUPOS
            if tiene_pres_g:
                # Grupo,Coord,DPI_Coord,Jefe,Pres,DPI_Pres,No,Nombre,DPI,Dir,Tel,Emp,Fecha
                gi_grupo=0; gi_coord=1; gi_dpi_coord=2; gi_jefe=3
                gi_pres=4; gi_dpi_pers=8; gi_nombre_pers=7
            elif tiene_dir_g:
                # Grupo,Coord,DPI_Coord,Jefe,No,Nombre,DPI,Dir,Tel,Emp,Fecha
                gi_grupo=0; gi_coord=1; gi_dpi_coord=2; gi_jefe=3
                gi_pres=-1; gi_dpi_pers=6; gi_nombre_pers=5
            else:
                # Grupo,Coord,DPI_Coord,Jefe,No,Nombre,DPI,Tel,Emp,Fecha
                gi_grupo=0; gi_coord=1; gi_dpi_coord=2; gi_jefe=3
                gi_pres=-1; gi_dpi_pers=5; gi_nombre_pers=4

            def gg(fg, idx): return fg[idx].strip() if idx >= 0 and idx < len(fg) else ''

            # Limpiar CUI para comparación
            cui_c = cui.replace(' ','').replace('-','').strip()

            # Primera pasada: buscar si es COORDINADORA (tiene prioridad)
            for fg in gs[1:]:
                if not fg or not fg[0]: continue
                dpi_coord = gg(fg, gi_dpi_coord).replace(' ','').replace('-','')
                if dpi_coord == cui_c:
                    grupo = gg(fg, gi_grupo)
                    coord = gg(fg, gi_coord)
                    if not jefe: jefe = gg(fg, gi_jefe)
                    if gi_pres >= 0 and not presidenta:
                        presidenta = gg(fg, gi_pres)
                    rol = 'Coordinadora'
                    break

            # Segunda pasada: buscar como integrante (siempre, aunque rol ya esté)
            if not grupo:
                for fg in gs[1:]:
                    if not fg or not fg[0]: continue
                    dpi_pers = gg(fg, gi_dpi_pers).replace(' ','').replace('-','')
                    if dpi_pers == cui_c:
                        grupo = gg(fg, gi_grupo)
                        coord = gg(fg, gi_coord)
                        if not jefe: jefe = gg(fg, gi_jefe)
                        if gi_pres >= 0 and not presidenta:
                            presidenta = gg(fg, gi_pres)
                        if not rol: rol = 'Integrante'
                        break

            # Si aún no encontró por DPI exacto, buscar por nombre completo
            if not grupo:
                p1 = gf(idx_pnombre).upper() if idx_pnombre>=0 else ''
                p3 = gf(idx_papell).upper() if idx_papell>=0 else ''
                gi_nombres = [7, 5, 4]
                for fg in gs[1:]:
                    if not fg or not fg[0]: continue
                    encontrado = False
                    for gi_n in gi_nombres:
                        nombre_g = gg(fg, gi_n).upper()
                        if not nombre_g: continue
                        if p1 and p3 and p1 in nombre_g and p3 in nombre_g:
                            encontrado = True
                            break
                    if encontrado:
                        grupo = gg(fg, gi_grupo)
                        coord = gg(fg, gi_coord)
                        if not jefe: jefe = gg(fg, gi_jefe)
                        if gi_pres >= 0 and not presidenta:
                            presidenta = gg(fg, gi_pres)
                        if not rol: rol = 'Integrante'
                        break
        except Exception as eg:
            print(f"Error GRUPOS: {eg}")

        # Determinar rol si no está definido
        if not rol:
            if coord and grupo:
                rol = 'Integrante'
            elif jefe and presidenta and not grupo:
                rol = 'Coordinadora'
            elif jefe and not presidenta:
                rol = 'Presidenta de Comité'
            else:
                rol = 'Integrante'


        # Generar PDF
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        from reportlab.platypus import Image as RLImage
        import base64 as b64mod

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.35*inch, bottomMargin=0.35*inch)
        styles = getSampleStyleSheet()
        titulo_s = ParagraphStyle('t', fontSize=12, textColor=colors.HexColor('#1a5276'),
            spaceAfter=1, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica-Bold')
        sub_s = ParagraphStyle('s', fontSize=8, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=1, spaceBefore=0, alignment=TA_CENTER, fontName='Helvetica')
        sec_s = ParagraphStyle('sec', fontSize=9, textColor=colors.white,
            spaceAfter=0, spaceBefore=0, fontName='Helvetica-Bold')
        norm_s = ParagraphStyle('n', fontSize=8.5, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=0, fontName='Helvetica')
        pie_s = ParagraphStyle('pie', fontSize=7, textColor=colors.HexColor('#7f8c8d'),
            spaceAfter=0, alignment=TA_CENTER, fontName='Helvetica')

        story = []
        logo_data = b64mod.b64decode("/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0aHBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCAEYATsDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqGe7gt5beOWQK9xJ5UQP8TbWbH5Kx/Cmk3sBNRRRSAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAK8f+KniGS38VaXb2kn7zTdtycHpISCAR9FB+jV6+zKiF2IVVGST0Ar5g13U21rXb3UW3f6RKzqG6hP4R+AwPwr3chwyq13OS0ivz0/K5y4qfLFJdT6ZsruG/sbe8t23QzxrLG3qrDI/nU9effCPV1vfDMunM/76xlIC/8ATN8sp/PePwr0GvKxVB4etKk+j/4Y3hLmipBRRRXOWFFFFABRRSMyopZmCqBkknAFAC0VxeufFTwnoe+M6h9uuF48mxHmnPoW+6PxNeaa78ctbvA8WjWMGnRnIEsp86X6gcKPxDV6mFybGYnWMLLu9F/XoQ5xR73PcQ2sDz3E0cMSDLSSMFVR7k1gWHjXSdZ1U6fopl1NkP7+e2XMEA9WkOFJ9AuSfTAJHh/hjwl4k+KGoC+1nUbttKSQ77qZycnusKn5c+4GB79K+hNG0XT9A0yLTtMtkt7aIcKo5J7knuT3JqsbgqGC/dynz1Oy2Xq+r8tPMItyL9FFFeSWFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAcp8R9VbSvBd4Y22zXOLZD/vfe/8AHQ34189nrXpXxg1V5tas9LVh5NvF5rAHq7nv9ABj/eNeamvuMjoeywqk95a/5Hl4mfNUt2Ot+Gms/wBk+NLZHbEF4DbP7E8qf++gB/wI19BV8nBmQh0dkdTlWU4IPYivp7w7qy654esdSXbmeIFwvRXHDD8GBFeZxHhuWca666P5f1+B0YSejiadFYus+LdB8Pg/2nqcEMmMiIHdIfogyf0rzzWfjao3R6HpTOegnvW2j6hF5P4kV4+Gy7E4n+HB277L72dEqsI7s9drmdc+IHhjw+WjvNUie4X/AJd7f97Jn0KrnH44rwPWvGXiPX9y6hq0xhbrBD+6jx6YXr+JNc8FVBhQAPQCvfw3DXXET+S/zf8AkYyxP8qPVtb+ON7Luj0PSkt16Ce8be34IpwPxJ+leb614l1zxCSdW1W5ukJz5RbbGP8AgC4X9Kzz1pm1mKoqszMQqqoyST0AHc19HhsuwuF1pwSffd/ezNzlLdkRwo7AD8hXpXw9+FM+vmLVtdSS30vIeK3Pyvcj1PdUP5ntgc10fw9+EotzFrHiaENOMNBYNgqno0nq3+z0HfJ6ewV4Ob8QWvRwj9Zf5f5/d3N6dPqyO3t4bS3jt7eJIoYlCRxooCqo4AAHQVJRRXxzdzcKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKQkKpYnAAyTVe+1Gy0y3M9/dwW0Q/jmkCj9a848YfEzSrnRrzTNHaaea4jMRuAm2NVPDYJwScZAwO+c11YbB1sRJKnF279EZ1KsYLVnmWu6nJrOuXuoyMSbiYsuey9FH4KAPwrMNPPSmGv0SnBQiox2R497u7Gt0q7Dr+sWmmNptrqd1b2bOXMUL7Mk4z8w5xwOM4/OqTdKjarlCM1aSuWm1sM2gEkDk8k9zTW6U49aa1aFIYaYetPNXtE0HUvEmqpp2mQGSZuWc5CRL/AHnbsP1PQZpynGEXKbskXFNuyKFpZXWo30VlY28lxdTHbHFGMsx/oPUngDk1774A+GNp4YEepal5d1rBGVbGUtsjkJ6n1b8sDrteDfA2m+DrMiAeffyj9/duo3N/sr/dX2/PNdRXxebZ5LE3o0NIfi/+B/T7HdTpcur3CiiivnTYKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiikYkKSoyccDOM0ALUc9xDawtNcTRxRLyzyMFA/E15f4p8eeKdPuPsp0tNJ3fdkf8AfF/91vufhg155qGoXuqTebqF3PdODkGZywH0HQfhXtYXJalZKc5JJ9tf+B+JxVcbGLsld/cexar8UPD9huS0aXUZR2t1+T/vs4H5ZrhdX+KHiDUMpZmHToj/AM8hvk/76YY/ICuNPWmNXvYfKMLR15eZ+ev4bHFPF1Z9begXU895OZ7qeW4mP/LSZy7fmahPenmmHoa9WKSVkY9Rh6Uw089KYatFoa3So2qRulRtWiNBh601qcetdr4H+HV34qZb28MlrpAP38Yef2T2/wBr8s9ssRiKeHpupVdkaQi5OyMfwj4M1LxhfGO2Hk2UbYnvGXKp7AfxN7du/v8AQ3h7w5pvhjTFsdNg2J1kkbl5Wx95j3P6DtgVesLC00yyisrG3jt7aJdqRxjAAqxXwuZZrVxsrbQWy/VnoUqSgvMKKKK8o1CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAIbuztr+2a3u7eOeF/vRyKGB/A15v4h+FStuuNAmCHqbSdiV/4C/UfQ5+or06iunDYythpXpv5dDKpRhUXvI+Zb+xu9Muza39tLbTgZ8uQYJHqOxHuOKrNX0tqmkafrNobXUbWO4hPIDjlT6g9QfcV5d4i+FN3a7rjQpjdRdfs0zASL/ut0b8cH3NfT4PO6NW0avuv8P+B8/vPOq4OcNY6o82NMPQ1NcQTW0zQ3EMkMqHDRyIVZfqDUJ6Gvdi01dHL1GHpTDTz0phq0WhrdKjapG6VG1aI0JbKaC21G2nuYBPbxyq8sJHEig5Kn6jivq23aF7aJrfZ5BQGPZ93bjjHtivks9a+gfhXrf9reDYrdz++09vsxz3QAFD9Np2/8BNfN8SUHKnCsumj+Z2YWWridvRRRXx52hRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFISFBJOAOSaWuY+IGqtpPg+8eJ9k1xi3jPpu6499u7HvWlGk6tSNOO7diZSUYuT6HiGv6rJreu3moyZ/fSkoD/Cg4UfgABWWehp5ph71+j04KEVCOyPCbbd2MPSmGtvw94a1HxPqH2SwjAVeZZ3B8uIe59fQdT9MkaniT4ca54f3Txp/aNkP+W1uh3KP9pOSPqMj6Vm8XQhV9lKaUuxrGnNrmS0OObpUbU/cGXIII9RTGrtQDD1rvvhJrbad4tGnt/qNQQxn2dQWU/wDoQ+rCuBPWn211PY3cF3bP5dxBIskb4zhgcj9RWGLw6xFCVJ9V+PT8TWEuWSZ9bUVW06+i1PTLW/gOYriJZV+hGas1+ZtNOzPVCiiikAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFeR/F7U2k1Gx0xX/dwxmd1B6sxwM/QA/99V650r5s8Q6n/bGv32olsrNKShP9wcL/AOOgV7eRUPaYj2j2ivxf9M48bPlp8vcyzXS+EPBN74ruPMJa301GxJc45Yjqqep9+g9+la/gv4dza4Y9R1ZXg008pFkrJP6H2T36nt617PBbw2tvHBbxJFDGoVERcKoHYCvTzLOFRvSoO8ur7f8ABMMPhXL3p7FbSdIsdE0+Ox0+3WGBOw5JPck9ST6mrtFFfJSk5Nyk7tnpJW0Rx/ib4caJ4iL3Cx/Yb5v+Xi3AG4/7a9G+vB968d8S+Btc8MFpLq38+yHS7twWQD/aHVPx49zX0lQQCCCMg9q9XBZxiMLaN+aPZ/ozGpQhPXZnyMeeR0ppr3vxP8J9H1jfcaWRpd4ef3aZhc/7SdvquPxrxzxD4W1rwxNs1SzZIicJcx/NE/phux9jg19dgs0w+L0i7S7Pf5dzknRlDc9Z+DOsC78NXOlvJuksZtyKf4Y35H1+YP8AnXpVfN/w11lNF8cWbytshus2sh/38bf/AB4LX0hXyeeYb2OLbW0tf8/xOyhLmh6BRRRXjmwUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAc1491T+yvCF4ytia4AtogDyWfg499u4/hXKeCvhuE8vU9fhBfhobJuQvoz+p/wBnt354HpM1lb3NxBPNEJHtyWi3chWIxuA6ZxkZ6jJ9TU9dtPGzo0HSpaN7v9DGVFTnzS6BRRRXEbBRRRQAUUUUAFRzQxXMLwzxJLE42ujqGVh6EHrUlFGwHl3ib4OWV0WuvDk4sLnr9nkJMLH2PVPwyPavRdJku5dJtX1CHyb3ygJ0yCA4GGwR1Gc4PpirlFdVbGVq8Iwqu/Ls3v8AeTGCi7oKKKK5SgooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiuf8Y68dA0F5ov+PqY+VD/ALLEH5vwAJ+uKulTlUmoR3ZM5qEXJ7IreJfHOn+H3NsiG7vh1hRsBP8Aebt9OTXB3PxM8RTPmI2duv8AdSEsfzJ/pXP6dp95reqR2luDLcTuSXck+5Zj+pP9a9X0z4caBZQAXUBvp8fNJMSB+Cg4A/M+9fQSpYLARSqrmk/69Dy41MRiW3B2RxNr8T/EEEg89bO5TurRlD+YP9K77wz4207xIfIANrfAZNvIc7h6qf4h+vtWfrHwy0e8gZtNDWFz/CVYtGfYqT/LFHgjwR/YJe+1ERyaiSVj2ncsS9Mg+p9fTj1zz4meX1aDnTXLLov60sbUY4mFRRk7o7aiisrxJrMegaBdag+CyLiNf7znhR+fX2zXjQhKclGO7O9tJXZwHiz4kanp3iK5sdJ+yG3t8Rs0sZcs4+90I6HjHtW/4A8Y3HiWK7t9REIvYCHXykKq0Z46EnkHr9RXkWlaRfa9qBtLJfMuCjSHe3XAycn1J4+pFW/COttoHia0vC22Bm8q4B4/dt1z9OG/CvrK+WUHh3Tppc8Un5/09TyqeJn7RSk/dZ9D0UUV8ieseR+JPidr2k+I7+wtoNPMFvKUQyROWI9yHFZJ+MHiYH/j30v/AL8v/wDF1heNv+R01j/r5avTvDnw98L3/hrTbu60zzJ57ZJJH8+QbmIyejYr66pDAYXD06lWlfmS29PU86Lq1JtRlsccvxi8SK2WtNLde4Eci/rvNd14P+JNj4nuRYXFubHUWBKRl96S4GTtbA56nBHT1qDW/hV4bl0q4NhA9jcohZJVmdlBAz8wYkEfrXiemz3Ntq1nPZAm6jnRoQOpfcMD3yeMe9KnhcvzCjN0I8so/L06tWL56tKS5ne59V0UUV8kdxyXxD8Wv4T8Pia1MZ1C4kEdusi7hxyzEZHAHH1IrzbRvjFr39t2i6u1j/ZzShZykBUqh4LZ3HpnPTtWb8VPEH9teL5beJs2unj7OnPVwfnP5/L/AMBrltW0O+0aOxe+jCC+tluYgD/A2cA+h7kdsivtcuyrDfVYxrxXPP7/AJeiOOdWXN7uyPrGiuO+GfiE+IfB1uZSftVkfssxP8RUDa3vlSMn1zXY18fXoyo1JU5bp2OtO6ucV8TPFuo+ENCtbzTI7Z5pbkRN9oRmULtJ4AI54Fc/8NPiPrni/wAR3Gn6nDYpDHaNMpt4nVtwdRzljxhjT/jp/wAirp//AF+j/wBAauO+Bn/I8Xv/AGDn/wDRkdfRYfB0JZPOs4rmV9fmjNyfPY+g6KKK+YNQooooAKKKKACiiigAooooAKKKKACiiigAooooAK81+KsjedpkX8IWRvx+WvSq4H4o2Ly6bZXyAlYJCj4HQNjBPtlcfiK78sko4uDf9aHLjU3QlYz/AIVWyNd6ldHG9ESMfRiSf/QRXp1eL+Ateh0TXGS6bZbXSiNnJ4Rs/KT7dR+Ne0Vrm8JxxLlLZ7EYCUXRSXQKKRmCqWYgADJJ7VW0/U7LVbcz2F1HcRK5QvG2QGHUf59Qa8yztc7Lq9i1XlPxZ1kyXVro0ZGyMefKf9o5Cj2wMn/gQr1G6uYrO0mup3CQwoZHY9lAyTXzfquoSanqd1qFwx3zyFzuPQdh9AMAewr2sjw3tK7qvaP5s48dU5Ycq6np3wn0VYbC51qVD5k5MMJP/PMH5iPqwx/wCuK+IOiDRvFdwE/1F1/pEY9NxO4f99A/gRWVb+KtXs4Et7XW7iGFBhI0lACj2qrqGt3uqtH9v1KS7aPIQyuCVz1x+Qr3aGDxEMXLESkrPprt0OKVam6Sglqj3H4e62dZ8KQeY5a5tD9nlJPJwPlPvlSOfXNdVXhvwy1v+y/FK2ksm23v18kgnA8zqh/mv/Aq9yr5jNcN9XxLS2eqPTw1Tnpruj5w8bf8jprH/Xy1XLD4l+JdMsYLG2msvIgQRx+Zb5YKOmTuGap+Nv8AkdNY/wCvlq9e8L+FfD9z4V0qefRNPlmktY3eSS3VmYlQSSSK+jxOIoUcJSdeHMml+RxU4TlUlyux5Jq3xB8TazZyWlzfpHbyDbIlvEI949CeTj2zzWl8MLDQrjxFDcalqCJewvm0s3TAkbHDbjwSDnC9cjNetz+B/C1xGY5NA0/B7pAqH8CMEV8/+KtKt9F8TX+nWsvmQQTYjbOSo4IGfVc4z6ilhMRh8bTnhqCdNtdEv6/rc0nCdNqcnc+n6xvFmtr4d8MX2p5HmRR4hB7yNwo/Mj8M1H4M1WXWvB+mX85LTSRbZGP8TKSpP4lc/jXmfxo13z9Rs9Dhf93br584B/5aMMKD9Fyf+BivncDgnWxioS6PX0W/+R11KlocyOG8JaI/ibxXZWEjMUkk8y4djklF+ZufU4xn1Nev/F3w6upeE11GGPNxpZLgKP8Alk2A4/DCt9FNeFWWtXWj3TT6fqT2dxt2F4nAbacHHP0FW7jxx4guoXgn8TXcsMilHjaZcMp4IPHQ19jisBiauKp16cklDpr8+nVaHLTnFQcWtzqfhF4gfSvF66fJKBaaivlFWOAJByhHv1XHfcPSvoSvjqOYo6TQSFXRgyOjcqwOQQfUGvq7wvrkXiPw1YarGRmeIeYo/hkHDr+DAivF4lwnJUjiIrSWj9V/wPyN6EtOU4P46f8AIrad/wBfo/8AQGrj/gZ/yPF7/wBg5/8A0ZHXYfHT/kVtOH/T5/7I1cf8DP8AkeL3/sHP/wCjI66ML/yIp/P80N/xD6Dooor483CiiigAooooAKKKKACiiigAooooAKKKKACiiigAqG8tIL+zmtLmMPDKpR1PcGpqKabTugavoeHeJ/CN94dmdyjz6eT8lyozgej46H36H9KraZ4w17SYFgs9QYwL92OVRIqj2zyB7ZxXvRAIwRkViXXhDw9euXm0e0Lk5LImwn8VxXuUs3hOHJiYc39dmebPASjLmoyseNav4r1rWIjFf6ixgPWJAI0P1A6/jmus+Guh63DqB1Lc9ppzrh45F5ufTCnoB13fgMgmu/svDOh6c6yWmlWkci9HEQLD8TzWrWeJzSEqTo0KfKn/AFsXSwclNTqSuzgfinrQtNGi0uNwJbtt0ntGvP6tj8jXMfC/RI9R1ybUJ4xJDZp8gYZBkbgexwN35g16veaNpeoTCa9060uZQu0PNCrkD0yR05P51NZ2Fnp0TRWVpBbRs24pDGEBPTOB34H5VlTzCNLBvDwWr3f9eWhrLDuVb2knoh32O1/59of++BTJdOsp4XiktYWR1KsNg5B61ZorzeaXc6bI+ZdX0+40PWrmxlYpNbS4V1OCe6sPTIwR9a+hfDetReINAtNRjIzImJFH8Ljhh+efwxUl5oOkahOZ7zS7O4mIAMksKs2B2yRmrFlp9npsBhsbSC2iLbikMYQE+uB34Fepj8xhi6UE42lHr+Zz0aDpSeujPnrxyAvjbVxn/l4J/QUtp4/8T6fZw2drq3l28KBI0MEbbVHQZK5r3m58N6JeXD3FzpNlLNIcvI8KlmPucc1D/wAIj4c/6Alh/wB+F/wrvjnOGdKNOrS5rJb2fQy+qzUm4ytc8Nn+I/i2WJo21x1BHJSGJT+YXIrH0fRtU8T6h5GmwSXUsj5knYkopPJZ37evqfc19GJ4U8OxsGXQ9OyO5tkP9K1Yoo4IxHFGkaDoqKAB+FDz2jSi1hqKi36L8lqUsNJ/HK5mW0Vl4S8LJG8m2z062y8hHJCjJOPUnJ+pr5pvLi78R+IZZkVpby/uSUj3ZJZ2+VQT2GQB6AV9S3dnbX9q9reW8VxbvjfFKgZWwcjIPuAao2vhjQLK5S5tdF0+CeM5SSO2RWU+xA4riy7M4YTnnKPNOXX+vM2qUnOyWyF0rQNP0rSrWwitYGWCNU3GMZYgck+5OT+NXPsFn/z6Qf8AfsVYorypVJSbbZqkkeJfGzw7FaXFhrltEkUcw+yzhFCjeMsh47kbhn/ZAp3wQ8RLHdXnh+aQYm/0m2BP8QGHH4jafwavY72wtNStHtL61hurd8bopkDq2DkZB46iqNn4W8P6fdx3VnomnW9xHkpLFbIrLkYOCBkcEj8a9hZrCeAeEqxba2f4r/L0M/Z2nzI4P464HhPTyTj/AE4f+gNXiGj+IdT8O3r3mkXxtLh4zEzhFbKkgkYYEdQPyr621HStP1eBYNSsba8hVt4juIg6humcEdeTWZ/whHhT/oWtI/8AAKP/AArpy3OqGGwv1erT5t+1vxCVNt3ufOp+K3jcf8zE/wD4Dw//ABFN/wCFr+OCf+Rhk/8AAeH/AOIr6M/4Qfwp/wBC1pH/AIBR/wCFH/CD+FP+ha0j/wAAo/8ACuj+28t/6Bl90f8AIfI+5twsWgjZjklQSfwp9IAAAAMAdBS18maBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH/9k=")
        logo_buffer = io.BytesIO(logo_data)
        logo_img = RLImage(logo_buffer, width=0.8*inch, height=0.7*inch)

        hdr = [Paragraph("PARTIDO CABAL", titulo_s),
               Paragraph("MUNICIPIO DE TOTONICAPAN", sub_s),
               Paragraph("FICHA DE REGISTRO INDIVIDUAL", titulo_s)]
        ht = Table([[hdr, logo_img]], colWidths=[6.7*inch, 0.8*inch])
        ht.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(1,0),(1,0),'RIGHT'),
            ('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0),
            ('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
        story.append(ht)
        story.append(Spacer(1,0.06*inch))
        story.append(HRFlowable(width="100%",thickness=2,color=colors.HexColor('#1a5276')))
        story.append(Spacer(1,0.06*inch))

        def sh_hdr(txt):
            t=Table([[Paragraph(txt,sec_s)]],colWidths=[7.5*inch])
            t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#1a5276')),
                ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),('LEFTPADDING',(0,0),(-1,-1),8)]))
            return t

        def fila_d(lbl,val):
            return [Paragraph(f"<b>{lbl}</b>",norm_s),Paragraph(str(val or '-'),norm_s)]

        def tabla_d(datos):
            t=Table(datos,colWidths=[2.0*inch,5.5*inch],rowHeights=[0.22*inch]*len(datos))
            t.setStyle(TableStyle([('FONTSIZE',(0,0),(-1,-1),8.5),('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.white,colors.HexColor('#eaf4fb')]),
                ('GRID',(0,0),(-1,-1),0.3,colors.HexColor('#d5d8dc')),
                ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),('LEFTPADDING',(0,0),(-1,-1),6)]))
            return t

        nombre_completo = f"{fila[2]} {fila[3]} {fila[4]} {fila[5]}".strip() if len(fila)>5 else ''

        story.append(sh_hdr("📋  DATOS PERSONALES"))
        story.append(tabla_d([
            fila_d("CUI / DPI:", cui),
            fila_d("Número de Serie:", gf(idx_serie) if idx_serie>=0 else gf(1)),
            fila_d("Primer Nombre:", gf(idx_pnombre) if idx_pnombre>=0 else gf(2)),
            fila_d("Segundo Nombre:", gf(idx_snombre) if idx_snombre>=0 else gf(3)),
            fila_d("Primer Apellido:", gf(idx_papell) if idx_papell>=0 else gf(4)),
            fila_d("Segundo Apellido:", gf(idx_sapell) if idx_sapell>=0 else gf(5)),
            fila_d("Sexo:", gf(idx_sexo) if idx_sexo>=0 else gf(6)),
            fila_d("Estado Civil:", gf(idx_ecivil) if idx_ecivil>=0 else gf(7)),
            fila_d("Fecha de Nacimiento:", gf(idx_fnac) if idx_fnac>=0 else gf(8)),
        ]))
        story.append(Spacer(1,0.06*inch))

        story.append(sh_hdr("📍  LUGAR DE NACIMIENTO Y VECINDAD"))
        story.append(tabla_d([
            fila_d("Municipio Nacimiento:", gf(idx_munnic) if idx_munnic>=0 else gf(9)),
            fila_d("Departamento Nacimiento:", gf(idx_depnac) if idx_depnac>=0 else gf(10)),
            fila_d("Municipio Vecindad:", gf(idx_munvec) if idx_munvec>=0 else gf(11)),
            fila_d("Departamento Vecindad:", gf(idx_depvec) if idx_depvec>=0 else gf(12)),
            fila_d("Fecha Expedición:", gf(idx_fexp) if idx_fexp>=0 else gf(13)),
            fila_d("Fecha Vencimiento:", gf(idx_fvenc) if idx_fvenc>=0 else gf(14)),
        ]))
        story.append(Spacer(1,0.06*inch))

        story.append(sh_hdr("📞  CONTACTO Y UBICACIÓN"))
        story.append(tabla_d([
            fila_d("Dirección / Comunidad:", gf(idx_dir) if idx_dir>=0 else gf(17)),
            fila_d("Teléfono:", gf(idx_tel) if idx_tel>=0 else gf(18)),
        ]))
        story.append(Spacer(1,0.06*inch))

        # Determinar rol basado en la información disponible
        rol = gf(idx_rol) if idx_rol >= 0 else ''
        if not rol:
            if not jefe and not presidenta and not coord and not grupo:
                rol = 'Jefe de Sector'
            elif jefe and not presidenta and not coord:
                rol = 'Presidenta de Comité'
            elif jefe and presidenta and not grupo:
                rol = 'Coordinadora'
            else:
                rol = 'Integrante'

        story.append(sh_hdr("🗳️  DATOS POLÍTICOS / PARTIDO"))
        datos_pol = [
            fila_d("Empadronado:", gf(idx_emp) if idx_emp>=0 else gf(15)),
            fila_d("No. Empadronamiento:", gf(idx_noemp) if idx_noemp>=0 else gf(16)),
            fila_d("Rol:", rol),
        ]
        # Mostrar campos según rol
        if rol in ('Presidenta de Comité', 'Coordinadora', 'Integrante'):
            datos_pol.append(fila_d("Jefe de Sector:", jefe or '-'))
        if rol in ('Coordinadora', 'Integrante'):
            datos_pol.append(fila_d("Presidenta de Comité:", presidenta or '-'))
        if rol == 'Integrante':
            datos_pol.append(fila_d("Coordinadora:", coord or '-'))
            datos_pol.append(fila_d("Grupo:", grupo or '-'))
        story.append(tabla_d(datos_pol))
        story.append(Spacer(1,0.08*inch))
        story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor('#bdc3c7')))
        story.append(Spacer(1,0.04*inch))
        story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO", pie_s))

        doc.build(story)
        buffer.seek(0)
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(buffer.read()).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "filename": f"Ficha_{cui}.pdf"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/verificar_grupos")
@requiere_admin
def verificar_grupos():
    """Verifica y muestra cuántas personas tiene cada grupo, detecta duplicados"""
    try:
        sh = get_sheet()
        ws_g = sh.worksheet("GRUPOS")
        filas_g = ws_g.get_all_values()
        header_g = filas_g[0] if filas_g else []
        tiene_pres = any('residenta' in str(h) for h in header_g)
        tiene_dir  = any('irecc' in str(h) for h in header_g)

        if tiene_pres:
            col_dpi = 8
        elif tiene_dir:
            col_dpi = 6
        else:
            col_dpi = 5

        grupos = {}
        duplicados = {}

        for i, fila in enumerate(filas_g[1:], start=2):
            if not fila or not fila[0]: continue
            nombre = fila[0].strip()
            dpi = fila[col_dpi].strip() if len(fila) > col_dpi else ''
            dpi_clean = dpi.replace(' ','').replace('-','')

            if nombre not in grupos:
                grupos[nombre] = []
                duplicados[nombre] = []

            if dpi_clean and dpi_clean in [d.replace(' ','') for d in grupos[nombre]]:
                duplicados[nombre].append(f"Fila {i}: DPI {dpi} DUPLICADO")
            grupos[nombre].append(dpi)

        resultado = {}
        for nombre, dpis in grupos.items():
            resultado[nombre] = {
                "total": len(dpis),
                "duplicados": duplicados.get(nombre, []),
                "excede": len(dpis) > 10
            }

        return jsonify({"ok": True, "grupos": resultado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/limpiar_duplicados_grupo", methods=["POST"])
@requiere_admin
def limpiar_duplicados_grupo():
    """Elimina filas duplicadas en un grupo dejando solo la primera ocurrencia de cada DPI"""
    try:
        nombre_grupo = request.json.get("nombre_grupo","").strip()
        sh = get_sheet()
        ws_g = sh.worksheet("GRUPOS")
        filas_g = ws_g.get_all_values()
        header_g = filas_g[0] if filas_g else []
        tiene_pres = any('residenta' in str(h) for h in header_g)
        tiene_dir  = any('irecc' in str(h) for h in header_g)
        col_dpi = 8 if tiene_pres else (6 if tiene_dir else 5)

        dpis_vistos = set()
        filas_borrar = []

        for i, fila in enumerate(filas_g[1:], start=2):
            if not fila or fila[0].strip() != nombre_grupo: continue
            dpi = fila[col_dpi].strip().replace(' ','').replace('-','') if len(fila) > col_dpi else ''
            if dpi in dpis_vistos:
                filas_borrar.append(i)
            else:
                dpis_vistos.add(dpi)

        for i in reversed(filas_borrar):
            ws_g.delete_rows(i)

        return jsonify({"ok": True, "eliminadas": len(filas_borrar)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _grupos_columnas_direccion(header):
    """Detecta en qué columna (índice) está 'Dirección' en la hoja GRUPOS,
    según el formato de encabezado que tenga (varía según cuándo se creó
    esa hoja). Devuelve -1 si esa hoja no tiene columna de Dirección
    propia (algunas versiones antiguas no la tienen)."""
    tiene_dir_bg = any('irecc' in str(h) for h in header)
    tiene_pres_bg = any('residenta' in str(h) for h in header)
    if tiene_pres_bg:
        gi_dpi_pers = 8
    elif tiene_dir_bg:
        gi_dpi_pers = 6
    else:
        gi_dpi_pers = 5
    if tiene_dir_bg:
        return gi_dpi_pers + 1
    return -1


@app.route("/vista_previa_correccion_direccion")
@requiere_admin
def vista_previa_correccion_direccion():
    """Busca (sin cambiar nada todavía) cuántos registros tienen
    exactamente el texto viejo en su Dirección/Comunidad, en las 3 hojas
    donde se guarda (DPI, Grupos, Presidentas), para poder confirmar antes
    de aplicar el cambio masivo."""
    try:
        texto_actual = request.args.get("texto_actual", "").strip()
        if not texto_actual:
            return jsonify({"ok": False, "error": "Falta el texto a buscar"})
        sh = get_sheet()
        coincidencias = []

        ws_dpi = sh.worksheet(HOJA)
        for fila in ws_dpi.get_all_values()[1:]:
            if fila and len(fila) > 17 and fila[17].strip().lower() == texto_actual.lower():
                nombre = f"{fila[2] if len(fila)>2 else ''} {fila[4] if len(fila)>4 else ''}".strip()
                coincidencias.append({"hoja": "DPI", "nombre": nombre or fila[0], "cui": fila[0]})

        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            if filas_g:
                col_dir = _grupos_columnas_direccion(filas_g[0])
                if col_dir != -1:
                    for fila in filas_g[1:]:
                        if fila and len(fila) > col_dir and fila[col_dir].strip().lower() == texto_actual.lower():
                            coincidencias.append({"hoja": "GRUPOS", "nombre": fila[col_dir-1] if col_dir-1 < len(fila) else '-', "cui": ''})
        except Exception:
            pass

        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            for fila in ws_p.get_all_values()[1:]:
                if fila and len(fila) > 4 and fila[4].strip().lower() == texto_actual.lower():
                    coincidencias.append({"hoja": "PRESIDENTAS", "nombre": fila[1] if len(fila)>1 else '-', "cui": fila[2] if len(fila)>2 else ''})
        except Exception:
            pass

        return jsonify({"ok": True, "total": len(coincidencias), "coincidencias": coincidencias[:100]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/aplicar_correccion_direccion", methods=["POST"])
@requiere_admin
def aplicar_correccion_direccion():
    """Reemplaza, en las 3 hojas (DPI, Grupos, Presidentas), a quien
    tenga EXACTAMENTE el texto viejo en su Dirección/Comunidad, por el
    texto nuevo. Usa coincidencia exacta (no 'contiene') para no alterar
    por accidente direcciones parecidas pero distintas. Cada hoja se
    actualiza en una sola llamada por lotes, para no toparse con el
    límite de escrituras de Google Sheets."""
    try:
        texto_actual = (request.json.get("texto_actual") or "").strip()
        texto_nuevo = (request.json.get("texto_nuevo") or "").strip()
        if not texto_actual or not texto_nuevo:
            return jsonify({"ok": False, "error": "Faltan el texto actual y/o el texto nuevo"})
        sh = get_sheet()
        resumen = {}

        ws_dpi = sh.worksheet(HOJA)
        celdas_dpi = []
        for i, fila in enumerate(ws_dpi.get_all_values()[1:], start=2):
            if fila and len(fila) > 17 and fila[17].strip().lower() == texto_actual.lower():
                celdas_dpi.append(gspread.Cell(i, 18, texto_nuevo))  # columna R = 18
        if celdas_dpi:
            ws_dpi.update_cells(celdas_dpi)
        resumen["DPI"] = len(celdas_dpi)

        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            celdas_g = []
            if filas_g:
                col_dir = _grupos_columnas_direccion(filas_g[0])
                if col_dir != -1:
                    for i, fila in enumerate(filas_g[1:], start=2):
                        if fila and len(fila) > col_dir and fila[col_dir].strip().lower() == texto_actual.lower():
                            celdas_g.append(gspread.Cell(i, col_dir + 1, texto_nuevo))
            if celdas_g:
                ws_g.update_cells(celdas_g)
            resumen["GRUPOS"] = len(celdas_g)
        except Exception:
            resumen["GRUPOS"] = 0

        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            filas_p = ws_p.get_all_values()
            celdas_p = []
            for i, fila in enumerate(filas_p[1:], start=2):
                if fila and len(fila) > 4 and fila[4].strip().lower() == texto_actual.lower():
                    celdas_p.append(gspread.Cell(i, 5, texto_nuevo))
            if celdas_p:
                ws_p.update_cells(celdas_p)
            resumen["PRESIDENTAS"] = len(celdas_p)
        except Exception:
            resumen["PRESIDENTAS"] = 0

        _nombre_sesion, _es_admin = _sesion_actual()
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Corrigió dirección de forma masiva",
                              f"'{texto_actual}' -> '{texto_nuevo}' | {resumen}")
        return jsonify({"ok": True, "resumen": resumen})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/revisar_padron")
@requiere_admin
def revisar_padron():
    """Vuelve a comparar a TODAS las personas marcadas como 'No empadronado'
    en la hoja DPI contra la hoja PADRON — útil cuando el Padrón se
    actualiza DESPUÉS de haber registrado a alguien, y esa persona ya
    debería contar como empadronada pero el sistema todavía no se había
    dado cuenta. Usa el mismo criterio de comparación que se usa al
    registrar a alguien nuevo. Solo SUBE a alguien de NO a SI — nunca baja
    a nadie de SI a NO, por seguridad (si alguien salió del Padrón después
    de haber sido confirmado, eso hay que revisarlo a mano)."""
    try:
        sh = get_sheet()
        ws_padron = sh.worksheet(HOJA_PADRON)
        dpis_padron_limpios = [
            c.replace(" ","").replace("-","").strip()
            for c in ws_padron.col_values(1) if c and c.strip()
        ]

        def _buscar_en_padron(cui_limpio):
            for celda_limpia in dpis_padron_limpios:
                if celda_limpia.startswith(cui_limpio) or cui_limpio in celda_limpia.split(",")[0]:
                    return celda_limpia
            return None

        # 1. Hoja DPI: revisar a quien esté en NO, y subir a SI a quien
        # ya aparezca en el Padrón.
        ws_dpi = sh.worksheet(HOJA)
        filas = ws_dpi.get_all_values()
        celdas_dpi = []
        cuis_recien_empadronados = set()
        nombres_actualizados = []
        for i, fila in enumerate(filas[1:], start=2):
            if not fila or not fila[0].strip():
                continue
            estado_actual = fila[15].strip().upper() if len(fila) > 15 else "NO"
            if estado_actual == "SI":
                continue
            cui_c = fila[0].replace(" ","").replace("-","").strip()
            if not cui_c:
                continue
            match = _buscar_en_padron(cui_c)
            if match:
                celdas_dpi.append(gspread.Cell(i, 16, "SI"))   # P - Empadronado
                celdas_dpi.append(gspread.Cell(i, 17, match))  # Q - No. Empadronamiento
                cuis_recien_empadronados.add(cui_c)
                nombre = f"{fila[2] if len(fila)>2 else ''} {fila[4] if len(fila)>4 else ''}".strip()
                nombres_actualizados.append(nombre or cui_c)
        if celdas_dpi:
            ws_dpi.update_cells(celdas_dpi)

        # 2. Hoja GRUPOS: subir a SI su propia fila si es una de las
        # personas recién confirmadas.
        cnt_grupos = 0
        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres = any('residenta' in str(h) for h in header_g)
            tiene_dir  = any('irecc' in str(h) for h in header_g)
            if tiene_pres:
                col_dpi_idx, col_emp_idx = 8, 11
            elif tiene_dir:
                col_dpi_idx, col_emp_idx = 6, 9
            else:
                col_dpi_idx, col_emp_idx = 5, 7
            celdas_g = []
            for i, fila in enumerate(filas_g[1:], start=2):
                if not fila:
                    continue
                dpi_persona = fila[col_dpi_idx].replace(" ","").replace("-","") if len(fila) > col_dpi_idx else ''
                if dpi_persona in cuis_recien_empadronados:
                    celdas_g.append(gspread.Cell(i, col_emp_idx + 1, "SI"))
                    cnt_grupos += 1
            if celdas_g:
                ws_g.update_cells(celdas_g)
        except Exception:
            pass

        # 3. Hoja PRESIDENTAS
        cnt_pres = 0
        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            filas_p = ws_p.get_all_values()
            celdas_p = []
            for i, fila in enumerate(filas_p[1:], start=2):
                if fila and len(fila) > 2:
                    dpi_p = fila[2].replace(" ","").replace("-","")
                    if dpi_p in cuis_recien_empadronados:
                        celdas_p.append(gspread.Cell(i, 6, "SI"))
                        cnt_pres += 1
            if celdas_p:
                ws_p.update_cells(celdas_p)
        except Exception:
            pass

        _nombre_sesion, _es_admin = _sesion_actual()
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Revisó DPI contra Padrón",
                              f"{len(cuis_recien_empadronados)} persona(s) pasaron de NO a SI")

        return jsonify({
            "ok": True,
            "total_actualizados": len(cuis_recien_empadronados),
            "nombres": nombres_actualizados[:300],
            "grupos_actualizados": cnt_grupos,
            "presidentas_actualizadas": cnt_pres,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/limpiar_cuis")
@requiere_admin
def limpiar_cuis():
    """Limpia los CUIs con espacios (o guiones) en todas las hojas. Junta
    todos los cambios de cada hoja y los manda en UNA sola llamada por hoja
    (ws.update_cells), en vez de una llamada por cada celda — así es mucho
    más rápido y no se arriesga a toparse con el límite de solicitudes de
    Google Sheets cuando hay cientos de filas por corregir."""
    try:
        sh = get_sheet()
        resultados = {}

        # Limpiar en hoja DPI
        try:
            ws = sh.worksheet(HOJA)
            filas = ws.get_all_values()
            celdas = []
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[0]:
                    cui_original = fila[0]
                    cui_limpio = cui_original.replace(' ','').replace('-','').strip()
                    if cui_limpio != cui_original:
                        celdas.append(gspread.Cell(i, 1, cui_limpio))
            if celdas:
                ws.update_cells(celdas)
            resultados['DPI'] = f'{len(celdas)} CUIs corregidos'
        except Exception as e:
            resultados['DPI_error'] = str(e)

        # Limpiar en hoja GRUPOS (columna DPI coord y DPI persona)
        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres = any('residenta' in str(h) for h in header_g)
            col_dpi_coord = 2
            col_dpi_pers = 8 if tiene_pres else 6
            celdas_g = []
            for i, fila in enumerate(filas_g[1:], start=2):
                if not fila: continue
                for col in [col_dpi_coord, col_dpi_pers]:
                    if col < len(fila) and fila[col]:
                        orig = fila[col]
                        limpio = orig.replace(' ','').replace('-','').strip()
                        if limpio != orig:
                            celdas_g.append(gspread.Cell(i, col + 1, limpio))
            if celdas_g:
                ws_g.update_cells(celdas_g)
            resultados['GRUPOS'] = f'{len(celdas_g)} CUIs corregidos'
        except Exception as e:
            resultados['GRUPOS_error'] = str(e)

        # Limpiar en hoja PRESIDENTAS
        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            filas_p = ws_p.get_all_values()
            celdas_p = []
            for i, fila in enumerate(filas_p[1:], start=2):
                if fila and len(fila) > 2 and fila[2]:
                    orig = fila[2]
                    limpio = orig.replace(' ','').replace('-','').strip()
                    if limpio != orig:
                        celdas_p.append(gspread.Cell(i, 3, limpio))
            if celdas_p:
                ws_p.update_cells(celdas_p)
            resultados['PRESIDENTAS'] = f'{len(celdas_p)} CUIs corregidos'
        except Exception as e:
            resultados['PRESIDENTAS_error'] = str(e)

        # Limpiar en hoja PADRON
        try:
            ws_pad = sh.worksheet(HOJA_PADRON)
            filas_pad = ws_pad.get_all_values()
            celdas_pad = []
            for i, fila in enumerate(filas_pad[1:], start=2):
                if fila and fila[0]:
                    orig = fila[0]
                    limpio = orig.replace(' ','').replace('-','').strip()
                    if limpio != orig:
                        celdas_pad.append(gspread.Cell(i, 1, limpio))
            if celdas_pad:
                ws_pad.update_cells(celdas_pad)
            resultados['PADRON'] = f'{len(celdas_pad)} CUIs corregidos'
        except Exception as e:
            resultados['PADRON_error'] = str(e)

        _nombre_sesion, _es_admin = _sesion_actual()
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Limpió espacios/guiones en DPI", str(resultados))

        return jsonify({"ok": True, "resultados": resultados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/actualizar_telefono", methods=["POST"])
@requiere_sesion
def actualizar_telefono():
    try:
        cui = request.json.get("cui","").strip()
        telefono = formatear_telefono_gt(request.json.get("telefono","").strip())
        if not cui or not telefono:
            return jsonify({"ok": False, "error": "Datos incompletos"})

        cui_c = cui.replace(" ","").replace("-","").strip()
        sh = get_sheet()
        actualizados = {}

        # 1. Actualizar en hoja DPI
        try:
            ws_dpi = sh.worksheet(HOJA)
            filas = ws_dpi.get_all_values()
            header = filas[0] if filas else []
            idx_tel = next((i for i,h in enumerate(header) if 'telefono' in str(h).lower() or 'teléfono' in str(h).lower()), 18)
            col_tel = chr(65 + idx_tel)
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[0].replace(" ","").strip() == cui_c:
                    ws_dpi.update(f'{col_tel}{i}', [[telefono]])
                    actualizados['DPI'] = True
                    break
        except Exception as e:
            actualizados['DPI_error'] = str(e)

        # 2. Actualizar en hoja GRUPOS
        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres = any('residenta' in str(h) for h in header_g)
            tiene_dir  = any('irecc' in str(h) for h in header_g)
            if tiene_pres:
                col_tel_g = 10  # Tel en formato con Presidenta
            elif tiene_dir:
                col_tel_g = 8   # Tel en formato con Dir
            else:
                col_tel_g = 7   # Tel en formato antiguo
            col_letra = chr(65 + col_tel_g)
            cnt = 0
            for i, fila in enumerate(filas_g[1:], start=2):
                if not fila: continue
                dpi_p = fila[col_tel_g-2].replace(" ","") if len(fila) > col_tel_g-2 else ''
                dpi_c = fila[2].replace(" ","") if len(fila) > 2 else ''
                if dpi_p == cui_c or dpi_c == cui_c:
                    ws_g.update(f'{col_letra}{i}', [[telefono]])
                    cnt += 1
            actualizados['GRUPOS'] = cnt
        except Exception as e:
            actualizados['GRUPOS_error'] = str(e)

        # 3. Actualizar en hoja PRESIDENTAS
        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            filas_p = ws_p.get_all_values()
            for i, fila in enumerate(filas_p[1:], start=2):
                if fila and len(fila) > 2 and fila[2].replace(" ","") == cui_c:
                    ws_p.update(f'D{i}', [[telefono]])
                    actualizados['PRESIDENTAS'] = True
                    break
        except Exception as e:
            actualizados['PRESIDENTAS_error'] = str(e)

        return jsonify({"ok": True, "actualizados": actualizados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/marcar_revisado", methods=["POST"])
@requiere_sesion
def marcar_revisado():
    """Guarda la fecha de HOY como la última vez que alguien confirmó a
    mano, en el Padrón y en la página del TSE, que el estado de esta
    persona seguía igual. Así, mientras esa fecha siga 'vigente' (menos de
    7 días), esta persona no vuelve a aparecer en la lista de pendientes
    por revisar — evita repetir cada día el mismo trabajo con la misma
    gente."""
    try:
        cui = request.json.get("cui", "").strip()
        cui_c = cui.replace(" ", "").replace("-", "").strip()
        if not cui_c:
            return jsonify({"ok": False, "error": "Falta el CUI"})
        sh = get_sheet()
        ws_dpi = sh.worksheet(HOJA)
        filas = ws_dpi.get_all_values()
        for i, fila in enumerate(filas[1:], start=2):
            if fila and fila[0].replace(" ", "").strip() == cui_c:
                ws_dpi.update(f'Z{i}', [[ahora_gt().strftime('%d/%m/%Y')]])
                _nombre_sesion, _es_admin = _sesion_actual()
                _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Marcó como revisado (Padrón/TSE)", f"CUI: {cui_c}")
                return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "No se encontró a esa persona en la hoja DPI"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/actualizar_empadronamiento_jefe", methods=["POST"])
@requiere_sesion
def actualizar_empadronamiento_jefe():
    """Edita SOLO el estado de empadronamiento (SI/NO) y el número de
    empadronamiento de una persona (buscada por CUI), desde la pantalla de
    Registros Individuales. Se propaga a DPI, a su propia fila en GRUPOS y
    a su propia fila en PRESIDENTAS. Si se marca como empadronado (SI) y
    esa persona no aparece todavía en la hoja PADRON, se agrega ahí
    también, para que quede como referencia local ya confirmada. Si se
    marca como NO, no se toca PADRON — no se borran registros existentes
    del padrón automáticamente, por seguridad."""
    try:
        data = request.json
        cui = data.get("cui","").strip()
        empadronado = data.get("empadronado","NO").strip().upper()
        if empadronado not in ("SI","NO"):
            empadronado = "NO"
        num_emp = data.get("num_empadronamiento","").strip()
        observaciones = data.get("observaciones","").strip()
        if not cui:
            return jsonify({"ok": False, "error": "Falta el CUI"})
        cui_c = cui.replace(" ","").replace("-","").strip()
        sh = get_sheet()
        actualizados = {}

        # Solo se bloquea si se intenta CAMBIAR el estado de una persona que
        # ya está marcada como empadronada (SI) a otra cosa — así se evita
        # modificar por accidente un empadronamiento ya confirmado. Pero si
        # el estado se deja igual (sigue en SI), sí se permite pasar, para
        # poder anotar una observación sin tocar el empadronamiento.
        try:
            ws_dpi_check = sh.worksheet(HOJA)
            filas_check = ws_dpi_check.get_all_values()
            fila_encontrada = None
            for fila_c in filas_check[1:]:
                if fila_c and fila_c[0].replace(" ","").strip() == cui_c:
                    fila_encontrada = fila_c
                    break
            if fila_encontrada is None:
                return jsonify({"ok": False, "error": "No se encontró a esa persona en la hoja DPI"})
            estado_actual = (fila_encontrada[15].strip().upper() if len(fila_encontrada) > 15 else "NO")
            if estado_actual == "SI" and empadronado != "SI":
                return jsonify({"ok": False, "error": "Esta persona ya está marcada como EMPADRONADA. Para evitar cambios accidentales, solo se puede editar cuando el estado actual es NO."})
        except Exception as e:
            return jsonify({"ok": False, "error": f"No se pudo verificar el estado actual: {e}"})

        # 1. Hoja DPI
        try:
            ws_dpi = sh.worksheet(HOJA)
            filas = ws_dpi.get_all_values()
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[0].replace(" ","").strip() == cui_c:
                    ws_dpi.update(f'P{i}', [[empadronado]])
                    ws_dpi.update(f'Q{i}', [[num_emp]])
                    # Solo se toca la columna de observaciones si de verdad
                    # se mandó algo desde el formulario. Así, si alguien
                    # guarda sin escribir nada en observaciones, no se borra
                    # por accidente lo que ya hubiera en esa columna.
                    if observaciones:
                        ws_dpi.update(f'Y{i}', [[observaciones]])
                    actualizados['DPI'] = True
                    break
        except Exception as e:
            actualizados['DPI_error'] = str(e)

        # 2. Hoja GRUPOS (columna Emp de su propia fila)
        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres = any('residenta' in str(h) for h in header_g)
            tiene_dir  = any('irecc' in str(h) for h in header_g)
            if tiene_pres:
                col_dpi_idx, col_emp_idx = 8, 11
            elif tiene_dir:
                col_dpi_idx, col_emp_idx = 6, 9
            else:
                col_dpi_idx, col_emp_idx = 5, 7
            cnt = 0
            for i, fila in enumerate(filas_g[1:], start=2):
                if not fila:
                    continue
                dpi_persona = fila[col_dpi_idx].replace(" ","") if len(fila) > col_dpi_idx else ''
                if dpi_persona == cui_c:
                    ws_g.update(f'{chr(65+col_emp_idx)}{i}', [[empadronado]])
                    cnt += 1
            actualizados['GRUPOS'] = cnt
        except Exception as e:
            actualizados['GRUPOS_error'] = str(e)

        # 3. Hoja PRESIDENTAS
        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            filas_p = ws_p.get_all_values()
            for i, fila in enumerate(filas_p[1:], start=2):
                if fila and len(fila) > 2 and fila[2].replace(" ","") == cui_c:
                    ws_p.update(f'F{i}', [[empadronado]])
                    actualizados['PRESIDENTAS'] = True
                    break
        except Exception as e:
            actualizados['PRESIDENTAS_error'] = str(e)

        # 4. Hoja PADRON: si se confirma como empadronado, agregarlo si
        # todavía no aparece (no se borra si se marca como NO).
        agregado_a_padron = False
        if empadronado == "SI":
            try:
                ws_padron = sh.worksheet(HOJA_PADRON)
                ya_esta = any(cui_c in (celda or '').replace(" ","").replace("-","") for celda in ws_padron.col_values(1))
                if not ya_esta:
                    ws_padron.append_row([cui_c])
                    agregado_a_padron = True
                actualizados['PADRON'] = True
            except Exception as e:
                actualizados['PADRON_error'] = str(e)

        _nombre_sesion, _es_admin = _sesion_actual()
        detalle_aud = f"CUI: {cui_c} | Emp: {empadronado} | No.Emp: {num_emp}"
        if observaciones:
            detalle_aud += f" | Obs: {observaciones}"
        if agregado_a_padron:
            detalle_aud += " | ➕ Agregado a PADRON"
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Editó empadronamiento (Registros Individuales)", detalle_aud)

        return jsonify({"ok": True, "actualizados": actualizados, "agregado_a_padron": agregado_a_padron})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/consultar_padron")
@requiere_sesion
def consultar_padron():
    """Busca en la hoja PADRON por nombre o por DPI (coincidencia parcial,
    sin distinguir mayúsculas/minúsculas). Como no se sabe de antemano qué
    columnas exactas tiene esa hoja (puede variar), se devuelven TODAS las
    columnas de cada fila que coincida, usando los encabezados de la propia
    hoja como nombres de campo — así se muestra cualquier dato disponible,
    sea cual sea la estructura real."""
    try:
        q = request.args.get("q","").strip()
        if not q or len(q) < 3:
            return jsonify({"ok": False, "error": "Escriba al menos 3 caracteres (nombre o DPI)"})
        sh = get_sheet()
        ws_padron = sh.worksheet(HOJA_PADRON)
        filas = ws_padron.get_all_values()
        if not filas:
            return jsonify({"ok": True, "encontrado": False, "resultados": []})

        encabezados = filas[0]
        q_dpi = q.replace(" ","").replace("-","").strip()
        q_nombre = q.strip().upper()
        resultados = []
        for fila in filas[1:]:
            if not fila or not any(c.strip() for c in fila):
                continue
            coincide_dpi = q_dpi.isdigit() and any(q_dpi in (c or '').replace(" ","").replace("-","") for c in fila)
            coincide_nombre = (not q_dpi.isdigit()) and any(q_nombre in (c or '').upper() for c in fila)
            if coincide_dpi or coincide_nombre:
                registro = {}
                for idx, valor in enumerate(fila):
                    clave = encabezados[idx].strip() if idx < len(encabezados) and encabezados[idx].strip() else f"Columna {idx+1}"
                    registro[clave] = valor
                resultados.append(registro)
                if len(resultados) >= 30:
                    break

        return jsonify({"ok": True, "encontrado": len(resultados) > 0, "resultados": resultados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/actualizar_registro_completo", methods=["POST"])
@requiere_admin
def actualizar_registro_completo():
    """Edita todos los datos de una persona (buscada por su CUI actual):
    el CUI mismo (por si el escaneo lo leyó mal), los campos extraídos del
    DPI (nombres, apellidos, sexo, estado civil, fechas, municipios/
    departamentos, número de serie) más dirección, teléfono, empadronado y
    número de empadronamiento. Como una misma persona puede aparecer en
    varias hojas y columnas (DPI siempre; GRUPOS si es coordinadora o
    integrante — y ahí su CUI se repite en la columna "DPI Coord" de TODAS
    las filas de su grupo si es coordinadora, o en "DPI Presidenta" de
    TODAS las filas de sus grupos si es presidenta; PRESIDENTAS si es
    presidenta), los cambios se propagan automáticamente a todas esas
    columnas, para no dejar CUI's viejos regados por distintas hojas."""
    try:
        data = request.json
        cui_actual = data.get("cui","").strip()
        cui_nuevo_raw = data.get("cui_nuevo","").strip()
        direccion = data.get("direccion","").strip()
        telefono = formatear_telefono_gt(data.get("telefono","").strip())
        empadronado = data.get("empadronado","NO").strip().upper()
        if empadronado not in ("SI","NO"):
            empadronado = "NO"
        num_emp = data.get("num_empadronamiento","").strip()

        numero_serie = data.get("numero_serie","").strip()
        primer_nombre = data.get("primer_nombre","").strip()
        segundo_nombre = data.get("segundo_nombre","").strip()
        primer_apellido = data.get("primer_apellido","").strip()
        segundo_apellido = data.get("segundo_apellido","").strip()
        sexo = data.get("sexo","").strip()
        estado_civil = data.get("estado_civil","").strip()
        fecha_nacimiento = data.get("fecha_nacimiento","").strip()
        municipio_nacimiento = data.get("municipio_nacimiento","").strip()
        departamento_nacimiento = data.get("departamento_nacimiento","").strip()
        municipio_vecindad = data.get("municipio_vecindad","").strip()
        departamento_vecindad = data.get("departamento_vecindad","").strip()
        fecha_expedicion = data.get("fecha_expedicion","").strip()
        fecha_vencimiento = data.get("fecha_vencimiento","").strip()

        nombre_completo = " ".join(x for x in [primer_nombre, segundo_nombre, primer_apellido, segundo_apellido] if x)

        if not cui_actual:
            return jsonify({"ok": False, "error": "Falta el CUI"})
        cui_c = cui_actual.replace(" ","").replace("-","").strip()
        cui_nuevo_c = cui_nuevo_raw.replace(" ","").replace("-","").strip() or cui_c
        cambia_cui = cui_nuevo_c != cui_c

        # Si el CUI va a cambiar, evitar chocar con otra persona que ya
        # tenga ese CUI nuevo registrado.
        sh = get_sheet()
        if cambia_cui:
            try:
                ws_check = sh.worksheet(HOJA)
                for fila_chk in ws_check.get_all_values()[1:]:
                    if fila_chk and fila_chk[0].replace(" ","").strip() == cui_nuevo_c:
                        return jsonify({"ok": False, "error": f"El CUI {cui_nuevo_c} ya está registrado en otra persona"})
            except Exception:
                pass

        # Volver a verificar contra el padrón (hoja PADRON) usando el CUI
        # correcto (el nuevo, si se corrigió). Si el CUI SÍ aparece ahí, se
        # confirma como empadronado automáticamente, sin importar lo que
        # decía el formulario. Si NO aparece, se respeta lo que la persona
        # escribió en el formulario — no se fuerza a "NO", porque el padrón
        # local puede estar incompleto y quizás ya se verificó por otro
        # medio (por ejemplo, consulta directa al TSE).
        empadronado_antes = empadronado
        confirmado_por_padron = False
        try:
            ws_padron = sh.worksheet(HOJA_PADRON)
            dpis_padron = ws_padron.col_values(1)
            for celda in dpis_padron:
                celda_limpia = celda.replace(" ","").replace("-","").strip()
                if celda_limpia.startswith(cui_nuevo_c) or cui_nuevo_c in celda_limpia.split(",")[0]:
                    empadronado = "SI"
                    if not num_emp:
                        num_emp = cui_nuevo_c
                    if empadronado_antes != "SI":
                        confirmado_por_padron = True
                    break
        except Exception:
            pass

        actualizados = {}

        # 1. Hoja DPI (siempre debería estar aquí). Es la fuente de verdad,
        # así que se actualiza la fila completa en UNA sola llamada (en vez
        # de 18 llamadas separadas, una por columna). Esto es más rápido y
        # evita que un límite de solicitudes de Google Sheets tumbe solo
        # algunas de las columnas a medio guardar. Si esta llamada falla,
        # se reporta como error real — antes el error se ignoraba
        # silenciosamente y la app decía "guardado" aunque el CUI nuevo
        # nunca quedara escrito en la hoja.
        cui_para_guardar = cui_nuevo_c if cambia_cui else cui_c
        try:
            ws_dpi = sh.worksheet(HOJA)
            filas = ws_dpi.get_all_values()
            fila_encontrada = None
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[0].replace(" ","").strip() == cui_c:
                    fila_encontrada = i
                    break
            if fila_encontrada is None:
                return jsonify({"ok": False, "error": f"No se encontró el registro con CUI {cui_c} para actualizar"})
            fila_valores = [
                cui_para_guardar, numero_serie, primer_nombre, segundo_nombre,
                primer_apellido, segundo_apellido, sexo, estado_civil,
                fecha_nacimiento, municipio_nacimiento, departamento_nacimiento,
                municipio_vecindad, departamento_vecindad, fecha_expedicion,
                fecha_vencimiento, empadronado, num_emp, direccion, telefono
            ]
            ws_dpi.update(f'A{fila_encontrada}:S{fila_encontrada}', [fila_valores])
            actualizados['DPI'] = True
        except Exception as e:
            return jsonify({"ok": False, "error": f"No se pudo guardar en la hoja principal (DPI): {e}"})

        # 2. Hoja GRUPOS. El CUI de esta persona puede repetirse en 3
        # columnas distintas: su propio "DPI" (una fila), "DPI Coord" (si es
        # coordinadora, repetido en TODAS las filas de su grupo), y "DPI
        # Presidenta" (si es presidenta, repetido en TODAS las filas de
        # todos sus grupos). Se revisan y actualizan las 3 por separado.
        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres = any('residenta' in str(h) for h in header_g)
            tiene_dir  = any('irecc' in str(h) for h in header_g)
            if tiene_pres:
                col_dpi_idx, col_dir_idx, col_tel_idx, col_emp_idx = 8, 9, 10, 11
                col_dpi_pres_idx = 5
            elif tiene_dir:
                col_dpi_idx, col_dir_idx, col_tel_idx, col_emp_idx = 6, 7, 8, 9
                col_dpi_pres_idx = -1
            else:
                col_dpi_idx, col_dir_idx, col_tel_idx, col_emp_idx = 5, -1, 6, 7
                col_dpi_pres_idx = -1
            col_nombre_idx = col_dpi_idx - 1
            col_dpi_coord_idx = 2
            cnt = 0
            for i, fila in enumerate(filas_g[1:], start=2):
                if not fila:
                    continue
                dpi_persona = fila[col_dpi_idx].replace(" ","") if len(fila) > col_dpi_idx else ''
                dpi_coord = fila[col_dpi_coord_idx].replace(" ","") if len(fila) > col_dpi_coord_idx else ''
                dpi_pres = fila[col_dpi_pres_idx].replace(" ","") if col_dpi_pres_idx >= 0 and len(fila) > col_dpi_pres_idx else ''
                es_su_propia_fila = dpi_persona == cui_c
                es_coordinadora_de_este_grupo = dpi_coord == cui_c
                es_presidenta_de_este_grupo = col_dpi_pres_idx >= 0 and dpi_pres == cui_c

                if es_su_propia_fila:
                    if col_dir_idx >= 0:
                        ws_g.update(f'{chr(65+col_dir_idx)}{i}', [[direccion]])
                    ws_g.update(f'{chr(65+col_tel_idx)}{i}', [[telefono]])
                    ws_g.update(f'{chr(65+col_emp_idx)}{i}', [[empadronado]])
                    if cambia_cui:
                        ws_g.update(f'{chr(65+col_dpi_idx)}{i}', [[cui_nuevo_c]])
                    if nombre_completo:
                        ws_g.update(f'{chr(65+col_nombre_idx)}{i}', [[nombre_completo]])
                        if len(fila) > 6 and fila[6].strip() == '1':
                            ws_g.update(f'B{i}', [[nombre_completo]])
                    cnt += 1

                if es_coordinadora_de_este_grupo and cambia_cui:
                    ws_g.update(f'C{i}', [[cui_nuevo_c]])
                    if not es_su_propia_fila:
                        cnt += 1

                if es_presidenta_de_este_grupo and cambia_cui:
                    ws_g.update(f'{chr(65+col_dpi_pres_idx)}{i}', [[cui_nuevo_c]])
                    if not es_su_propia_fila:
                        cnt += 1
            actualizados['GRUPOS'] = cnt
        except Exception as e:
            actualizados['GRUPOS_error'] = str(e)

        # 3. Hoja PRESIDENTAS (si aparece como presidenta). Formato:
        # Jefe de Sector, Nombre, DPI, Telefono, Direccion, Empadronado, Fecha
        try:
            ws_p = sh.worksheet("PRESIDENTAS")
            filas_p = ws_p.get_all_values()
            for i, fila in enumerate(filas_p[1:], start=2):
                if fila and len(fila) > 2 and fila[2].replace(" ","") == cui_c:
                    ws_p.update(f'D{i}', [[telefono]])
                    ws_p.update(f'E{i}', [[direccion]])
                    ws_p.update(f'F{i}', [[empadronado]])
                    if cambia_cui:
                        ws_p.update(f'C{i}', [[cui_nuevo_c]])
                    if nombre_completo:
                        ws_p.update(f'B{i}', [[nombre_completo]])
                    actualizados['PRESIDENTAS'] = True
                    break
        except Exception as e:
            actualizados['PRESIDENTAS_error'] = str(e)

        _nombre_sesion, _es_admin = _sesion_actual()
        detalle_aud = f"CUI: {cui_c}"
        if cambia_cui:
            detalle_aud += f" → {cui_nuevo_c}"
        detalle_aud += f" | Nombre: {nombre_completo} | Dir: {direccion} | Tel: {telefono} | Emp: {empadronado} | No.Emp: {num_emp}"
        if confirmado_por_padron:
            detalle_aud += " | ✅ Confirmado como empadronado por el padrón"
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Editó registro completo", detalle_aud)

        return jsonify({"ok": True, "actualizados": actualizados, "confirmado_por_padron": confirmado_por_padron, "empadronado_final": empadronado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/actualizar_empadronamiento", methods=["POST"])
@requiere_admin
def actualizar_empadronamiento():
    try:
        data = request.json
        cui = data.get("cui","").strip()
        empadronado = data.get("empadronado","NO")
        num_emp = data.get("num_empadronamiento","").strip()
        sh = get_sheet()
        ws = sh.worksheet(HOJA)
        filas = ws.get_all_values()
        header = filas[0] if filas else []

        # Encontrar índices de columnas
        idx_emp = next((i for i,h in enumerate(header) if 'empadronado' in str(h).lower()), 15)
        idx_noemp = next((i for i,h in enumerate(header) if 'empadronamiento' in str(h).lower() and 'no' in str(h).lower()), 16)

        for i, fila in enumerate(filas[1:], start=2):
            if fila and fila[0].strip() == cui:
                # Actualizar columna empadronado
                col_emp = chr(65 + idx_emp)
                col_noemp = chr(65 + idx_noemp)
                ws.update(f'{col_emp}{i}', [[empadronado]])
                if num_emp:
                    ws.update(f'{col_noemp}{i}', [[num_emp]])
                return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "CUI no encontrado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/buscar_presidenta")
@requiere_admin
def buscar_presidenta():
    try:
        nombre = request.args.get('nombre','').strip().lower()
        dpi = request.args.get('dpi','').strip()
        sh = get_sheet()
        try:
            ws = sh.worksheet("PRESIDENTAS")
        except:
            return jsonify({"ok": False, "error": "No existe hoja PRESIDENTAS"})
        filas = ws.get_all_values()
        for i, fila in enumerate(filas[1:], start=2):
            if not fila: continue
            nombre_fila = fila[1].strip().lower() if len(fila) > 1 else ''
            dpi_fila = fila[2].strip() if len(fila) > 2 else ''
            if (nombre and nombre in nombre_fila) or (dpi and dpi == dpi_fila):
                return jsonify({"ok": True, "numero_fila": i, "presidenta": {
                    "nombre": fila[1].strip() if len(fila) > 1 else '',
                    "cui": fila[2].strip() if len(fila) > 2 else '',
                    "telefono": fila[3].strip() if len(fila) > 3 else '',
                    "direccion": fila[4].strip() if len(fila) > 4 else '',
                    "jefe": fila[0].strip() if len(fila) > 0 else '',
                }})
        return jsonify({"ok": False, "error": "Presidenta no encontrada"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/eliminar_presidenta", methods=["POST"])
@requiere_admin
def eliminar_presidenta():
    try:
        numero_fila = request.json.get("numero_fila")
        nombre = request.json.get("nombre","").strip()
        sh = get_sheet()
        ws = sh.worksheet("PRESIDENTAS")
        if numero_fila:
            ws.delete_rows(int(numero_fila))
        else:
            # Buscar por nombre
            filas = ws.get_all_values()
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[1].strip().lower() == nombre.lower():
                    ws.delete_rows(i)
                    break
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# Nombres originales, usados solo para sembrar la hoja JEFES_SECTOR la primera
# vez que se crea (así no se pierden los jefes que ya existían antes de que
# esta lista se moviera de estar fija en el código a la hoja de cálculo).
# El PIN inicial de cada uno es provisional; el administrador debe avisarles
# su PIN real o cambiarlo desde Mantenimiento antes de que empiecen a usarlo.
_JEFES_SECTOR_INICIALES = [
    ("JUAN VASQUEZ", "1001"), ("WALTER XURUX", "1002"), ("JUAN CARLOS RODAS", "1003"),
    ("GREGORIO SAY", "1004"), ("CARLOS TACAM", "1005"), ("ISABEL VELASQUEZ", "1006"),
]


def _get_ws_jefes(sh):
    """Obtiene (o crea) la hoja JEFES_SECTOR, con columnas Nombre, PIN,
    Telefono, Email1, Email2, Email3 (hasta 3 correos por jefe, usados para
    el envío automático del reporte diario), y Comunidad (para mostrar en
    el organigrama a qué comunidad pertenece el jefe)."""
    try:
        ws = sh.worksheet("JEFES_SECTOR")
    except Exception:
        ws = sh.add_worksheet(title="JEFES_SECTOR", rows=200, cols=7)
        ws.append_row(["Nombre", "PIN", "Telefono", "Email1", "Email2", "Email3", "Comunidad"])
        for nombre_j, pin_j in _JEFES_SECTOR_INICIALES:
            ws.append_row([nombre_j, pin_j, "", "", "", "", ""])
        return ws

    # Migración: si la hoja ya existía de una versión anterior (sin PIN, sin
    # Teléfono, sin las columnas de Email, o sin Comunidad), se completa lo
    # que falte. No nos basamos en contar cuántas columnas tiene el
    # encabezado (Google Sheets a veces reporta una columna "vacía" como
    # presente por formato), sino en revisar directamente el contenido de
    # cada fila.
    try:
        if ws.col_count < 7:
            ws.resize(cols=7)
        filas = ws.get_all_values()
        if filas:
            encabezado = filas[0]
            if len(encabezado) < 2 or not encabezado[1].strip():
                ws.update_cell(1, 2, "PIN")
            if len(encabezado) < 3 or not encabezado[2].strip():
                ws.update_cell(1, 3, "Telefono")
            if len(encabezado) < 4 or not encabezado[3].strip():
                ws.update_cell(1, 4, "Email1")
            if len(encabezado) < 5 or not encabezado[4].strip():
                ws.update_cell(1, 5, "Email2")
            if len(encabezado) < 6 or not encabezado[5].strip():
                ws.update_cell(1, 6, "Email3")
            if len(encabezado) < 7 or not encabezado[6].strip():
                ws.update_cell(1, 7, "Comunidad")
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[0].strip():
                    pin_actual = fila[1].strip() if len(fila) > 1 else ''
                    if not pin_actual:
                        ws.update_cell(i, 2, "1000")
    except Exception:
        pass
    return ws


def _get_ws_config_correos(sh):
    """Obtiene (o crea) la hoja CONFIG_CORREOS, con una sola fila de datos
    que guarda hasta 3 correos 'generales' (por ejemplo, los del
    administrador), que reciben el reporte combinado de TODOS los jefes de
    sector, además de los correos específicos de cada jefe."""
    try:
        ws = sh.worksheet("CONFIG_CORREOS")
    except Exception:
        ws = sh.add_worksheet(title="CONFIG_CORREOS", rows=5, cols=3)
        ws.append_row(["Email1", "Email2", "Email3"])
        ws.append_row(["", "", ""])
    return ws


_AFILIADOS_ENCABEZADO = [
    "CUI", "PrimerNombre", "SegundoNombre", "PrimerApellido", "SegundoApellido",
    "Sexo", "EstadoCivil", "FechaNacimiento", "MunicipioNac", "DeptNac",
    "MunicipioVec", "DeptVec", "FechaExpedicion", "FechaVencimiento",
    "Direccion", "Telefono", "Empadronado", "NumEmpadronamiento",
    "RegistradoPor", "FechaRegistro", "FirmaB64",
]


def _get_ws_afiliados(sh):
    """Obtiene (o crea) la hoja AFILIADOS, completamente separada de la
    hoja DPI — aquí se guardan únicamente las personas registradas desde
    'Afiliación al Partido' (por apoyo1/apoyo2), sin mezclarse con los
    registros normales de Presidentas/Coordinadoras/Integrantes."""
    try:
        ws = sh.worksheet("AFILIADOS")
    except Exception:
        ws = sh.add_worksheet(title="AFILIADOS", rows=200, cols=len(_AFILIADOS_ENCABEZADO))
        ws.append_row(_AFILIADOS_ENCABEZADO)
        return ws

    # Migración: si la hoja ya existía de antes de agregar la columna de
    # firma, se completa el encabezado que falte, sin tocar los datos ya
    # guardados.
    try:
        if ws.col_count < len(_AFILIADOS_ENCABEZADO):
            ws.resize(cols=len(_AFILIADOS_ENCABEZADO))
        encabezado = ws.row_values(1)
        if len(encabezado) < len(_AFILIADOS_ENCABEZADO) or not encabezado[-1].strip():
            ws.update_cell(1, len(_AFILIADOS_ENCABEZADO), "FirmaB64")
    except Exception:
        pass
    return ws


@app.route("/buscar_afiliado")
@requiere_admin
def buscar_afiliado():
    """Busca un afiliado por CUI en la hoja AFILIADOS, para poder verlo,
    modificarlo o eliminarlo desde Mantenimiento."""
    try:
        cui = request.args.get('cui', '').strip().replace(' ', '').replace('-', '')
        if not cui:
            return jsonify({"ok": False, "error": "Falta el DPI"})
        sh = get_sheet()
        ws = _get_ws_afiliados(sh)
        filas = ws.get_all_values()
        for i, fila in enumerate(filas[1:], start=2):
            if fila and fila[0].strip().replace(' ', '') == cui:
                return jsonify({"ok": True, "fila": fila, "numero_fila": i})
        return jsonify({"ok": False, "fila": None, "error": "No se encontró ningún afiliado con ese DPI"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/actualizar_afiliado", methods=["POST"])
@requiere_admin
def actualizar_afiliado():
    """Guarda los cambios hechos a un afiliado desde Mantenimiento (una
    sola llamada por lotes, para evitar el límite de escrituras de Google
    Sheets)."""
    try:
        numero_fila = request.json.get("numero_fila")
        datos = request.json.get("datos", {})
        if not numero_fila:
            return jsonify({"ok": False, "error": "Falta el número de fila"})
        cui_nuevo = datos.get('cui', '').replace(' ', '').replace('-', '').strip()
        if not cui_nuevo:
            return jsonify({"ok": False, "error": "El DPI no puede quedar vacío"})
        sh = get_sheet()
        ws = _get_ws_afiliados(sh)
        fila_actual = ws.row_values(int(numero_fila))
        firma_actual = fila_actual[20].strip() if len(fila_actual) > 20 else ''
        registrado_por_actual = fila_actual[18].strip() if len(fila_actual) > 18 else ''
        fecha_registro_actual = fila_actual[19].strip() if len(fila_actual) > 19 else ''
        fila_nueva = [
            cui_nuevo,
            datos.get('primer_nombre', ''), datos.get('segundo_nombre', ''),
            datos.get('primer_apellido', ''), datos.get('segundo_apellido', ''),
            datos.get('sexo', ''), datos.get('estado_civil', ''),
            datos.get('fecha_nacimiento', ''),
            datos.get('municipio_nacimiento', ''), datos.get('departamento_nacimiento', ''),
            datos.get('municipio_vecindad', ''), datos.get('departamento_vecindad', ''),
            datos.get('fecha_expedicion', ''), datos.get('fecha_vencimiento', ''),
            datos.get('direccion', ''), formatear_telefono_gt(datos.get('telefono', '').strip()),
            datos.get('empadronado', 'NO'), datos.get('num_empadronamiento', ''),
            registrado_por_actual, fecha_registro_actual, firma_actual,
        ]
        ws.update(f'A{numero_fila}:U{numero_fila}', [fila_nueva])
        _nombre_sesion, _es_admin = _sesion_actual()
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Modificó un afiliado", f"CUI: {cui_nuevo}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/eliminar_afiliado", methods=["POST"])
@requiere_admin
def eliminar_afiliado():
    """Elimina un afiliado de la hoja AFILIADOS."""
    try:
        numero_fila = request.json.get("numero_fila")
        cui = request.json.get("cui", "").strip()
        if not numero_fila:
            return jsonify({"ok": False, "error": "Falta el número de fila"})
        sh = get_sheet()
        ws = _get_ws_afiliados(sh)
        ws.delete_rows(int(numero_fila))
        _nombre_sesion, _es_admin = _sesion_actual()
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Eliminó un afiliado", f"CUI: {cui}")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})





def _get_ws_auditoria(sh):
    """Obtiene (o crea) la hoja AUDITORIA, donde se registra cada acción relevante."""
    try:
        ws = sh.worksheet("AUDITORIA")
    except Exception:
        ws = sh.add_worksheet(title="AUDITORIA", rows=5000, cols=4)
        ws.append_row(["Fecha/Hora", "Jefe de Sector", "Acción", "Detalle"])
    return ws


def _registrar_auditoria(sh, jefe, accion, detalle=""):
    """Agrega una línea al registro de auditoría. No interrumpe la operación
    principal si falla (por eso va envuelto en try/except)."""
    try:
        ws = _get_ws_auditoria(sh)
        ws.append_row([ahora_gt().strftime('%d/%m/%Y %H:%M:%S'), jefe or '-', accion, detalle])
    except Exception:
        pass


@app.route("/listar_jefes")
def listar_jefes():
    try:
        sh = get_sheet()
        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()
        jefes = [f[0].strip() for f in filas[1:] if f and f[0].strip()]
        return jsonify({"ok": True, "jefes": jefes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/listar_jefes_detalle")
@requiere_sesion
def listar_jefes_detalle():
    """Igual que /listar_jefes, pero además incluye el teléfono y los
    correos de cada jefe de sector (usado en Mantenimiento para poder
    editarlos)."""
    try:
        sh = get_sheet()
        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()
        jefes = []
        for f in filas[1:]:
            if f and f[0].strip():
                jefes.append({
                    "nombre": f[0].strip(),
                    "telefono": f[2].strip() if len(f) > 2 else '',
                    "emails": [
                        f[3].strip() if len(f) > 3 else '',
                        f[4].strip() if len(f) > 4 else '',
                        f[5].strip() if len(f) > 5 else '',
                    ],
                    "comunidad": f[6].strip() if len(f) > 6 else '',
                })
        return jsonify({"ok": True, "jefes": jefes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/debug_migracion_jefes")
@requiere_admin
def debug_migracion_jefes():
    """Ruta temporal de diagnóstico: muestra exactamente qué pasa al intentar
    migrar la hoja JEFES_SECTOR, sin ocultar ningún error. Se puede borrar
    una vez confirmado que la migración funciona correctamente."""
    try:
        sh = get_sheet()
        ws = sh.worksheet("JEFES_SECTOR")
        info = {"col_count_antes": ws.col_count, "filas_antes": ws.get_all_values()}
        try:
            if ws.col_count < 2:
                ws.resize(cols=2)
            info["resize_ok"] = True
        except Exception as e:
            info["resize_error"] = str(e)
        try:
            ws.update_cell(1, 2, "PIN")
            info["update_header_ok"] = True
        except Exception as e:
            info["update_header_error"] = str(e)
        try:
            ws.update_cell(2, 2, "1000")
            info["update_fila_ok"] = True
        except Exception as e:
            info["update_fila_error"] = str(e)
        info["filas_despues"] = ws.get_all_values()
        return jsonify({"ok": True, "info": info})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "tipo": type(e).__name__})


@app.route("/agregar_jefe", methods=["POST"])
@requiere_admin
def agregar_jefe():
    try:
        nombre = request.json.get("nombre", "").strip()
        pin = request.json.get("pin", "").strip()
        telefono = formatear_telefono_gt(request.json.get("telefono", "").strip())
        emails = request.json.get("emails", ["", "", ""])
        emails = [(e or "").strip() for e in (emails + ["", "", ""])[:3]]
        comunidad = (request.json.get("comunidad") or "").strip()
        if not nombre:
            return jsonify({"ok": False, "error": "Nombre vacío"})
        if not pin or not pin.isdigit() or len(pin) < 4:
            return jsonify({"ok": False, "error": "El PIN debe tener al menos 4 dígitos numéricos"})
        sh = get_sheet()
        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()
        existentes = [f[0].strip().upper() for f in filas[1:] if f and f[0].strip()]
        if nombre.upper() in existentes:
            return jsonify({"ok": False, "error": "Ese jefe de sector ya existe"})
        ws.append_row([nombre.upper(), pin, telefono] + emails + [comunidad])
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _fila_esta_vacia(fila):
    """Una fila se considera realmente vacía solo si NINGUNA de sus columnas
    tiene contenido. Antes se revisaba solo la primera columna (Grupo, o
    Jefe de Sector), lo que hacía que se saltaran por error filas con datos
    reales en las demás columnas pero con esa primera columna en blanco
    (por ejemplo, un grupo sin nombre guardado), dejándolas sin borrar."""
    return not fila or not any((c or '').strip() for c in fila)


def _limpiar_grupos_presidentas_dpi(sh, jefe_filtro=None):
    """Borra filas de GRUPOS, PRESIDENTAS y DPI. Si jefe_filtro se indica,
    solo borra las filas de ESE jefe de sector (sin tocar a los demás). Si
    jefe_filtro es None, borra TODO (usado en el reinicio general). Nunca
    toca JEFES_SECTOR ni AUDITORIA. Devuelve cuántas filas se borraron de
    cada hoja, y cualquier error real que haya ocurrido (en vez de
    ocultarlo en silencio, para poder diagnosticar si algo falla)."""
    resumen = {"GRUPOS": 0, "PRESIDENTAS": 0, "DPI": 0}
    errores = {}
    jefe_norm = jefe_filtro.strip().upper() if jefe_filtro else None

    # GRUPOS: columna D (índice 3) = Jefe de Sector.
    # En vez de borrar fila por fila (una llamada a Google Sheets por cada
    # una — con muchas filas eso agota rápido el límite de "escrituras por
    # minuto" de Google y deja el borrado a medias), se arma en memoria la
    # lista de filas que SÍ deben quedar, se limpia la hoja completa y se
    # vuelve a escribir todo de una sola vez. Esto siempre son 2 llamadas
    # a Google Sheets por hoja, sin importar si se borran 5 filas o 500.
    try:
        ws = sh.worksheet("GRUPOS")
        filas = ws.get_all_values()
        encabezado = filas[0] if filas else []
        filas_que_quedan = []
        cantidad_borrada = 0
        for fila in filas[1:]:
            if _fila_esta_vacia(fila):
                continue
            jefe_fila = fila[3].strip().upper() if len(fila) > 3 else ''
            if jefe_norm is None or jefe_fila == jefe_norm:
                cantidad_borrada += 1
            else:
                filas_que_quedan.append(fila)
        if cantidad_borrada:
            ws.clear()
            ws.update('A1', [encabezado] + filas_que_quedan)
        resumen["GRUPOS"] = cantidad_borrada
    except Exception as e:
        errores["GRUPOS"] = str(e)

    # PRESIDENTAS: columna A (índice 0) = Jefe de Sector
    try:
        ws = sh.worksheet("PRESIDENTAS")
        filas = ws.get_all_values()
        encabezado = filas[0] if filas else []
        filas_que_quedan = []
        cantidad_borrada = 0
        for fila in filas[1:]:
            if _fila_esta_vacia(fila):
                continue
            jefe_fila = fila[0].strip().upper() if len(fila) > 0 else ''
            if jefe_norm is None or jefe_fila == jefe_norm:
                cantidad_borrada += 1
            else:
                filas_que_quedan.append(fila)
        if cantidad_borrada:
            ws.clear()
            ws.update('A1', [encabezado] + filas_que_quedan)
        resumen["PRESIDENTAS"] = cantidad_borrada
    except Exception as e:
        errores["PRESIDENTAS"] = str(e)

    # DPI: columna T (índice 19) = Jefe de Sector. En un reinicio general
    # (jefe_norm es None) se borran TODOS los registros, para que de verdad
    # sea "empezar de 0"; al borrar solo un jefe, se dejan intactos los
    # registros sin jefe asignado o de otros jefes.
    try:
        ws = sh.worksheet(HOJA)
        filas = ws.get_all_values()
        encabezado = filas[0] if filas else []
        filas_que_quedan = []
        cantidad_borrada = 0
        for fila in filas[1:]:
            if _fila_esta_vacia(fila):
                continue
            if jefe_norm is None:
                cantidad_borrada += 1
                continue
            jefe_fila = fila[19].strip().upper() if len(fila) > 19 else ''
            if jefe_fila == jefe_norm:
                cantidad_borrada += 1
            else:
                filas_que_quedan.append(fila)
        if cantidad_borrada:
            ws.clear()
            ws.update('A1', [encabezado] + filas_que_quedan)
        resumen["DPI"] = cantidad_borrada
    except Exception as e:
        errores["DPI"] = str(e)

    return resumen, errores


@app.route("/eliminar_datos_jefe", methods=["POST"])
@requiere_admin
def eliminar_datos_jefe():
    """Borra TODOS los grupos, presidentas y registros de DPI de un jefe de
    sector específico, sin afectar a los demás jefes. El jefe sigue
    pudiendo iniciar sesión normalmente (no se toca JEFES_SECTOR), solo
    queda sin datos, como si estuviera empezando de cero."""
    try:
        nombre_jefe = request.json.get("nombre_jefe", "").strip()
        if not nombre_jefe:
            return jsonify({"ok": False, "error": "Falta indicar el jefe de sector"})
        sh = get_sheet()
        resumen, errores = _limpiar_grupos_presidentas_dpi(sh, jefe_filtro=nombre_jefe)
        _nombre_sesion, _es_admin = _sesion_actual()
        detalle_aud = f"Jefe: {nombre_jefe} | Grupos: {resumen['GRUPOS']} | Presidentas: {resumen['PRESIDENTAS']} | DPI: {resumen['DPI']}"
        if errores:
            detalle_aud += f" | ERRORES: {errores}"
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Borrado de datos de un jefe", detalle_aud)
        return jsonify({"ok": True, "resumen": resumen, "errores": errores if errores else None})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/reinicio_general", methods=["POST"])
@requiere_admin
def reinicio_general():
    """Borra TODOS los grupos, presidentas y registros de DPI, de TODOS los
    jefes de sector. No toca JEFES_SECTOR (todos pueden seguir iniciando
    sesión) ni AUDITORIA (se conserva el historial). Exige escribir una
    frase exacta de confirmación, para evitar un borrado accidental."""
    try:
        confirmacion = request.json.get("confirmacion", "").strip()
        if confirmacion != "REINICIAR TODO":
            return jsonify({"ok": False, "error": "Confirmación incorrecta"})
        sh = get_sheet()
        resumen, errores = _limpiar_grupos_presidentas_dpi(sh, jefe_filtro=None)
        _nombre_sesion, _es_admin = _sesion_actual()
        detalle_aud = f"Grupos: {resumen['GRUPOS']} | Presidentas: {resumen['PRESIDENTAS']} | DPI: {resumen['DPI']}"
        if errores:
            detalle_aud += f" | ERRORES: {errores}"
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "⚠️ REINICIO GENERAL", detalle_aud)
        return jsonify({"ok": True, "resumen": resumen, "errores": errores if errores else None})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/eliminar_jefe", methods=["POST"])
@requiere_admin
def eliminar_jefe():
    try:
        nombre = request.json.get("nombre", "").strip().upper()
        if not nombre:
            return jsonify({"ok": False, "error": "Nombre vacío"})
        sh = get_sheet()
        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()
        for i, fila in enumerate(filas[1:], start=2):
            if fila and fila[0].strip().upper() == nombre:
                ws.delete_rows(i)
                return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "No se encontró ese jefe de sector"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/editar_jefe", methods=["POST"])
@requiere_admin
def editar_jefe():
    try:
        nombre_actual = request.json.get("nombre_actual", "").strip().upper()
        nombre_nuevo = request.json.get("nombre_nuevo", "").strip().upper()
        pin_nuevo = request.json.get("pin_nuevo", "").strip()
        telefono_nuevo = request.json.get("telefono_nuevo", "").strip()
        emails_nuevos = request.json.get("emails_nuevos", None)
        comunidad_nueva = request.json.get("comunidad_nueva", None)
        if not nombre_actual or not nombre_nuevo:
            return jsonify({"ok": False, "error": "Nombre vacío"})
        if pin_nuevo and (not pin_nuevo.isdigit() or len(pin_nuevo) < 4):
            return jsonify({"ok": False, "error": "El PIN debe tener al menos 4 dígitos numéricos"})
        sh = get_sheet()
        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()

        # Evitar duplicados (a menos que sea el mismo nombre sin cambios reales)
        if nombre_nuevo != nombre_actual:
            existentes = [f[0].strip().upper() for f in filas[1:] if f and f[0].strip()]
            if nombre_nuevo in existentes:
                return jsonify({"ok": False, "error": "Ya existe un jefe de sector con ese nombre"})

        for i, fila in enumerate(filas[1:], start=2):
            if fila and fila[0].strip().upper() == nombre_actual:
                # Se calculan primero los valores finales de cada campo
                # (usando el actual si no vino uno nuevo), y se manda TODA
                # la fila en UNA sola llamada, en vez de una llamada por
                # cada campo por separado. Esto evita toparse con el límite
                # de "escrituras por minuto" de Google Sheets, sobre todo
                # cuando además hay que renombrar al jefe en varios grupos.
                pin_final = pin_nuevo if pin_nuevo else (fila[1].strip() if len(fila) > 1 else '')
                tel_final = formatear_telefono_gt(telefono_nuevo) if telefono_nuevo else (fila[2].strip() if len(fila) > 2 else '')
                if emails_nuevos is not None:
                    emails_final = [(e or "").strip() for e in (list(emails_nuevos) + ["", "", ""])[:3]]
                else:
                    emails_final = [
                        fila[3].strip() if len(fila) > 3 else '',
                        fila[4].strip() if len(fila) > 4 else '',
                        fila[5].strip() if len(fila) > 5 else '',
                    ]
                comunidad_final = comunidad_nueva.strip() if comunidad_nueva is not None else (fila[6].strip() if len(fila) > 6 else '')
                ws.update(f'A{i}:G{i}', [[nombre_nuevo, pin_final, tel_final] + emails_final + [comunidad_final]])

                # Actualizar también el nombre del jefe en los grupos ya
                # existentes, para que no queden grupos "huérfanos" con el
                # nombre anterior. Se juntan todos los cambios y se mandan
                # en UNA sola llamada (en vez de una por cada grupo), por la
                # misma razón de arriba.
                try:
                    ws_g = sh.worksheet("GRUPOS")
                    filas_g = ws_g.get_all_values()
                    celdas_g = []
                    for j, fg in enumerate(filas_g[1:], start=2):
                        if len(fg) > 3 and fg[3].strip().upper() == nombre_actual:
                            celdas_g.append(gspread.Cell(j, 4, nombre_nuevo))
                    if celdas_g:
                        ws_g.update_cells(celdas_g)
                except Exception:
                    pass
                return jsonify({"ok": True})
        return jsonify({"ok": False, "error": "No se encontró ese jefe de sector"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/login_jefe", methods=["POST"])
def login_jefe():
    """Valida el nombre + PIN de un jefe de sector para iniciar sesión en la
    sección de Grupos. Registra el intento (exitoso o fallido) en AUDITORIA.
    También acepta un inicio de sesión especial como ADMINISTRADOR (mismo PIN
    que Mantenimiento), que ve todo sin filtro por jefe de sector. Limita a
    MAX_INTENTOS_PIN intentos fallidos antes de bloquear temporalmente."""
    try:
        bloqueo = _verificar_bloqueo_pin('intentos_jefe')
        if bloqueo:
            return jsonify({"ok": False, "error": bloqueo, "bloqueado": True})

        nombre = request.json.get("nombre", "").strip().upper()
        pin = request.json.get("pin", "").strip()
        if not nombre or not pin:
            return jsonify({"ok": False, "error": "Complete nombre y PIN"})
        sh = get_sheet()

        if nombre == "ADMINISTRADOR":
            if pin == ADMIN_PIN:
                _iniciar_sesion("ADMINISTRADOR", es_admin=True)
                _registrar_intento_pin('intentos_jefe', exitoso=True)
                _registrar_auditoria(sh, "ADMINISTRADOR", "Inicio de sesión", "Acceso correcto (sin filtro)")
                return jsonify({"ok": True, "nombre": "ADMINISTRADOR", "es_admin": True})
            else:
                _registrar_intento_pin('intentos_jefe', exitoso=False)
                _registrar_auditoria(sh, "ADMINISTRADOR", "Intento de acceso fallido", "PIN incorrecto")
                return jsonify({"ok": False, "error": "PIN incorrecto"})

        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()
        for fila in filas[1:]:
            if fila and fila[0].strip().upper() == nombre:
                pin_real = fila[1].strip() if len(fila) > 1 else ""
                if pin_real and pin == pin_real:
                    _iniciar_sesion(nombre, es_admin=False)
                    _registrar_intento_pin('intentos_jefe', exitoso=True)
                    _registrar_auditoria(sh, nombre, "Inicio de sesión", "Acceso correcto")
                    return jsonify({"ok": True, "nombre": nombre})
                else:
                    _registrar_intento_pin('intentos_jefe', exitoso=False)
                    _registrar_auditoria(sh, nombre, "Intento de acceso fallido", "PIN incorrecto")
                    return jsonify({"ok": False, "error": "PIN incorrecto"})
        _registrar_intento_pin('intentos_jefe', exitoso=False)
        return jsonify({"ok": False, "error": "Jefe de sector no encontrado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_pin_jefe")
@requiere_admin
def ver_pin_jefe():
    """Solo para uso desde Mantenimiento: permite ver el PIN, teléfono y
    correos actuales de un jefe al momento de editarlo."""
    try:
        nombre = request.args.get("nombre", "").strip().upper()
        sh = get_sheet()
        ws = _get_ws_jefes(sh)
        filas = ws.get_all_values()
        for fila in filas[1:]:
            if fila and fila[0].strip().upper() == nombre:
                return jsonify({
                    "ok": True,
                    "pin": fila[1].strip() if len(fila) > 1 else "",
                    "telefono": fila[2].strip() if len(fila) > 2 else "",
                    "emails": [
                        fila[3].strip() if len(fila) > 3 else "",
                        fila[4].strip() if len(fila) > 4 else "",
                        fila[5].strip() if len(fila) > 5 else "",
                    ],
                    "comunidad": fila[6].strip() if len(fila) > 6 else "",
                })
        return jsonify({"ok": False, "error": "No encontrado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/config_correos_admin", methods=["GET", "POST"])
@requiere_admin
def config_correos_admin():
    """GET: devuelve los hasta 3 correos 'generales' configurados (los que
    reciben el reporte combinado de TODOS los jefes). POST: los guarda."""
    try:
        sh = get_sheet()
        ws = _get_ws_config_correos(sh)
        if request.method == "GET":
            filas = ws.get_all_values()
            fila = filas[1] if len(filas) > 1 else []
            emails = [fila[i].strip() if len(fila) > i else "" for i in range(3)]
            return jsonify({"ok": True, "emails": emails})
        else:
            emails = request.json.get("emails", ["", "", ""])
            emails = [(e or "").strip() for e in (list(emails) + ["", "", ""])[:3]]
            ws.update('A2:C2', [emails])
            return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_empadronamientos_actualizados")
@requiere_admin
def ver_empadronamientos_actualizados():
    """Revisa TODA la hoja de Auditoría (no solo las últimas 300 filas) y
    devuelve, ya estructurados, los casos donde se editó el empadronamiento
    de alguien desde 'Editar Empadronamiento' en Registros Individuales —
    con su CUI, el nuevo estado, el número, si se agregó al Padrón, quién
    hizo el cambio y cuándo."""
    try:
        sh = get_sheet()
        ws = _get_ws_auditoria(sh)
        filas = ws.get_all_values()
        registros = []
        for fila in filas[1:]:
            if len(fila) < 4:
                continue
            fecha, jefe, accion, detalle = fila[0], fila[1], fila[2], fila[3]
            if "Editó empadronamiento" not in accion:
                continue
            # El detalle tiene el formato:
            # "CUI: X | Emp: Y | No.Emp: Z" (+ " | ➕ Agregado a PADRON" si aplica)
            cui = ""
            emp = ""
            num_emp = ""
            agregado_padron = "Agregado a PADRON" in detalle
            for parte in detalle.split("|"):
                parte = parte.strip()
                if parte.startswith("CUI:"):
                    cui = parte.replace("CUI:", "").strip()
                elif parte.startswith("Emp:"):
                    emp = parte.replace("Emp:", "").strip()
                elif parte.startswith("No.Emp:"):
                    num_emp = parte.replace("No.Emp:", "").strip()
            registros.append({
                "fecha": fecha, "jefe": jefe, "cui": cui,
                "empadronado": emp, "num_empadronamiento": num_emp,
                "agregado_padron": agregado_padron,
            })
        # Más reciente primero
        registros.reverse()
        return jsonify({"ok": True, "registros": registros})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/pdf_empadronamientos_actualizados", methods=["POST"])
@requiere_admin
def pdf_empadronamientos_actualizados():
    try:
        data = request.json
        registros = data.get("registros", [])

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                rightMargin=0.5*inch, leftMargin=0.5*inch,
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        titulo_s = ParagraphStyle('t', fontSize=15, textColor=colors.HexColor('#1a5276'),
            spaceAfter=4, alignment=TA_CENTER, fontName='Helvetica-Bold')
        sub_s = ParagraphStyle('s', fontSize=10, textColor=colors.HexColor('#2c3e50'),
            spaceAfter=2, alignment=TA_CENTER, fontName='Helvetica')

        story = []
        story.append(Paragraph("DPI con Empadronamiento Actualizado", titulo_s))
        story.append(Paragraph("Cambios realizados desde Editar Empadronamiento — Registros Individuales", sub_s))
        story.append(Paragraph(f"{len(registros)} cambio(s) registrado(s)", sub_s))
        story.append(Spacer(1, 0.15*inch))
        story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1a5276')))
        story.append(Spacer(1, 0.15*inch))

        celda_s = ParagraphStyle('celda', fontSize=9, alignment=TA_CENTER, fontName='Helvetica', leading=11)
        filas = [["No.", "Fecha", "CUI", "Nuevo estado", "No. Empadronamiento", "¿Se agregó a Padrón?", "Realizado por"]]
        for i, r in enumerate(registros, 1):
            filas.append([
                str(i),
                r.get('fecha','-'),
                r.get('cui','-'),
                r.get('empadronado','-'),
                r.get('num_empadronamiento','-') or '-',
                "Sí" if r.get('agregado_padron') else "No",
                Paragraph(r.get('jefe','-'), celda_s),
            ])

        t = Table(filas, colWidths=[0.35*inch, 1.3*inch, 1.4*inch, 0.9*inch, 1.4*inch, 1.3*inch, 1.6*inch], repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a5276')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#bdc3c7')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#eaf4fb')]),
            ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        story.append(t)

        story.append(Spacer(1, 0.2*inch))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#bdc3c7')))
        story.append(Spacer(1, 0.05*inch))
        story.append(Paragraph(f"Generado el {ahora_gt().strftime('%d/%m/%Y %H:%M')} — Sistema DPI-2026 — Partido CABAL-TOTO",
            ParagraphStyle('pie', fontSize=7.5, textColor=colors.HexColor('#7f8c8d'), alignment=TA_CENTER)))

        doc.build(story)
        buffer.seek(0)
        pdf_bytes = buffer.read()
        token = _guardar_pdf_temporal(pdf_bytes, "Empadronamientos_Actualizados.pdf")
        import base64 as b64enc
        pdf_b64 = b64enc.b64encode(pdf_bytes).decode('utf-8')
        return jsonify({"ok": True, "pdf_b64": pdf_b64, "token": token})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_auditoria")
@requiere_admin
def ver_auditoria():
    """Devuelve el registro de auditoría (más recientes primero), solo para
    uso desde Mantenimiento (administrador)."""
    try:
        sh = get_sheet()
        ws = _get_ws_auditoria(sh)
        filas = ws.get_all_values()
        registros = filas[1:]
        registros.reverse()
        return jsonify({"ok": True, "registros": registros[:300]})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})



@app.route("/eliminar_dpi", methods=["POST"])
@requiere_admin
def eliminar_dpi():
    try:
        numero_fila = request.json.get("numero_fila")
        cui = request.json.get("cui","").strip().replace(" ","").replace("-","")
        if not numero_fila:
            return jsonify({"ok": False, "error": "Número de fila no proporcionado"})
        sh = get_sheet()
        eliminados = {}

        # 1. Eliminar de hoja DPI
        try:
            ws_dpi = sh.worksheet(HOJA)
            ws_dpi.delete_rows(int(numero_fila))
            eliminados["DPI"] = True
        except Exception as e:
            eliminados["DPI_error"] = str(e)

        if cui:
            # 2. Eliminar de hoja GRUPOS (todas las filas donde aparezca el CUI)
            try:
                ws_g = sh.worksheet("GRUPOS")
                filas_g = ws_g.get_all_values()
                header_g = filas_g[0] if filas_g else []
                tiene_pres = any('residenta' in str(h) for h in header_g)
                tiene_dir  = any('irecc' in str(h) for h in header_g)
                if tiene_pres:
                    col_dpi_pers = 8; col_dpi_coord = 2
                elif tiene_dir:
                    col_dpi_pers = 6; col_dpi_coord = 2
                else:
                    col_dpi_pers = 5; col_dpi_coord = 2
                filas_borrar = []
                for i, fila in enumerate(filas_g[1:], start=2):
                    if not fila: continue
                    dpi_p = fila[col_dpi_pers].strip().replace(" ","").replace("-","") if len(fila) > col_dpi_pers else ''
                    dpi_c = fila[col_dpi_coord].strip().replace(" ","").replace("-","") if len(fila) > col_dpi_coord else ''
                    if dpi_p == cui or dpi_c == cui:
                        filas_borrar.append(i)
                for i in reversed(filas_borrar):
                    ws_g.delete_rows(i)
                eliminados["GRUPOS"] = len(filas_borrar)
            except Exception as e:
                eliminados["GRUPOS_error"] = str(e)

            # 3. Eliminar de hoja PRESIDENTAS si aplica
            try:
                ws_p = sh.worksheet("PRESIDENTAS")
                filas_p = ws_p.get_all_values()
                filas_borrar_p = []
                for i, fila in enumerate(filas_p[1:], start=2):
                    if fila and len(fila) > 2 and fila[2].strip().replace(" ","").replace("-","") == cui:
                        filas_borrar_p.append(i)
                for i in reversed(filas_borrar_p):
                    ws_p.delete_rows(i)
                eliminados["PRESIDENTAS"] = len(filas_borrar_p)
            except Exception as e:
                eliminados["PRESIDENTAS_error"] = str(e)

        return jsonify({"ok": True, "eliminados": eliminados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/editar_nombre_grupo", methods=["POST"])
@requiere_admin
def editar_nombre_grupo():
    """Renombra un grupo (actualiza la columna Grupo en todas sus filas de
    GRUPOS). Se compara también contra el jefe de sector del grupo
    encontrado, por si existieran dos grupos con el mismo nombre exacto
    bajo jefes distintos, para no renombrar el equivocado."""
    try:
        nombre_actual = request.json.get("nombre_actual","").strip()
        nombre_nuevo = request.json.get("nombre_nuevo","").strip()
        jefe_sector = request.json.get("jefe_sector","").strip()
        if not nombre_actual or not nombre_nuevo:
            return jsonify({"ok": False, "error": "Faltan datos"})
        sh = get_sheet()
        ws_g = sh.worksheet("GRUPOS")
        filas_g = ws_g.get_all_values()
        cnt = 0
        for i, fila in enumerate(filas_g[1:], start=2):
            if not fila:
                continue
            coincide_nombre = fila[0].strip() == nombre_actual
            coincide_jefe = (not jefe_sector) or (len(fila) > 3 and fila[3].strip().upper() == jefe_sector.upper())
            if coincide_nombre and coincide_jefe:
                ws_g.update_cell(i, 1, nombre_nuevo)
                cnt += 1
        if cnt == 0:
            return jsonify({"ok": False, "error": "No se encontró el grupo"})
        _nombre_sesion, _es_admin = _sesion_actual()
        _registrar_auditoria(sh, _nombre_sesion or "ADMINISTRADOR", "Editó nombre de grupo",
            f"'{nombre_actual}' → '{nombre_nuevo}' ({cnt} fila(s))")
        return jsonify({"ok": True, "filas": cnt})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/eliminar_grupo", methods=["POST"])
@requiere_admin
def eliminar_grupo():
    try:
        nombre_grupo = request.json.get("nombre_grupo","").strip()
        sh = get_sheet()
        eliminados = {}

        # 1. Obtener CUIs de todas las personas del grupo antes de borrar
        cuis_grupo = []
        try:
            ws_g = sh.worksheet("GRUPOS")
            filas_g = ws_g.get_all_values()
            header_g = filas_g[0] if filas_g else []
            tiene_pres = any('residenta' in str(h) for h in header_g)
            tiene_dir  = any('irecc' in str(h) for h in header_g)
            col_dpi = 8 if tiene_pres else (6 if tiene_dir else 5)
            col_dpi_coord = 2
            filas_borrar = []
            for i, fila in enumerate(filas_g[1:], start=2):
                if fila and fila[0].strip().lower() == nombre_grupo.lower():
                    filas_borrar.append(i)
                    if len(fila) > col_dpi and fila[col_dpi].strip():
                        cuis_grupo.append(fila[col_dpi].strip())
                    if len(fila) > col_dpi_coord and fila[col_dpi_coord].strip():
                        cuis_grupo.append(fila[col_dpi_coord].strip())
            for i in reversed(filas_borrar):
                ws_g.delete_rows(i)
            eliminados["GRUPOS"] = len(filas_borrar)
        except Exception as e:
            eliminados["GRUPOS_error"] = str(e)

        # 2. Eliminar de DPI todos los CUIs del grupo
        try:
            ws_dpi = sh.worksheet(HOJA)
            filas_dpi = ws_dpi.get_all_values()
            filas_dpi_borrar = []
            for i, fila in enumerate(filas_dpi[1:], start=2):
                if fila and fila[0].strip() in cuis_grupo:
                    filas_dpi_borrar.append(i)
            for i in reversed(filas_dpi_borrar):
                ws_dpi.delete_rows(i)
            eliminados["DPI"] = len(filas_dpi_borrar)
        except Exception as e:
            eliminados["DPI_error"] = str(e)

        return jsonify({"ok": True, "eliminados": eliminados})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_coordinadoras")
@requiere_sesion
def ver_coordinadoras():
    """Devuelve una coordinadora por grupo (la persona No.1 de cada grupo),
    filtrado opcionalmente por jefe de sector. Usa la misma detección de
    formato de columnas que /buscar_grupo, para no depender de una posición
    fija que pueda cambiar si la hoja GRUPOS se reestructura."""
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        sh = get_sheet()
        try:
            ws = sh.worksheet("GRUPOS")
        except Exception:
            return jsonify({"ok": True, "coordinadoras": []})

        filas = ws.get_all_values()
        if len(filas) <= 1:
            return jsonify({"ok": True, "coordinadoras": []})

        header_bg = filas[0]
        tiene_dir_bg = any('irecc' in str(h) for h in header_bg)
        tiene_pres_bg = any('residenta' in str(h) for h in header_bg)

        if tiene_pres_bg:
            gi_jefe=3; gi_pres=4; gi_dpi_pers=8
        elif tiene_dir_bg:
            gi_jefe=3; gi_pres=-1; gi_dpi_pers=6
        else:
            gi_jefe=3; gi_pres=-1; gi_dpi_pers=5

        gi_no = gi_dpi_pers - 2
        gi_nombre = gi_dpi_pers - 1
        if tiene_dir_bg:
            gi_dir = gi_dpi_pers + 1
            gi_tel = gi_dir + 1
        else:
            gi_dir = -1
            gi_tel = gi_dpi_pers + 1
        gi_emp = gi_tel + 1

        def gc(f, i): return f[i].strip() if i >= 0 and i < len(f) else ''

        # Presidenta por jefe de sector, para cuando la hoja GRUPOS no la
        # tiene directamente (formato antiguo sin columna Presidenta).
        presidenta_por_jefe = {}
        if not tiene_pres_bg:
            try:
                ws_pr = sh.worksheet("PRESIDENTAS")
                for fp in ws_pr.get_all_values()[1:]:
                    if fp and len(fp) > 1:
                        presidenta_por_jefe[fp[0].strip().upper()] = fp[1].strip()
            except Exception:
                pass

        vistos = set()
        coordinadoras = []
        for fila in filas[1:]:
            if not fila or not fila[0].strip():
                continue
            nombre_grupo = fila[0].strip()
            jefe_fila = gc(fila, gi_jefe)
            if jefe_filtro and jefe_fila.upper() != jefe_filtro:
                continue
            if gc(fila, gi_no) != '1':
                continue
            if nombre_grupo in vistos:
                continue
            vistos.add(nombre_grupo)
            presidenta = gc(fila, gi_pres) if gi_pres >= 0 else presidenta_por_jefe.get(jefe_fila.upper(), '')
            coordinadoras.append({
                "grupo": nombre_grupo,
                "jefe": jefe_fila,
                "presidenta": presidenta or '(sin presidenta)',
                "nombre": gc(fila, gi_nombre),
                "cui": gc(fila, gi_dpi_pers),
                "telefono": gc(fila, gi_tel),
                "direccion": gc(fila, gi_dir),
                "empadronado": gc(fila, gi_emp),
            })
        coordinadoras.sort(key=lambda c: (c["presidenta"].upper(), c["direccion"].upper(), c["nombre"].upper()))
        return jsonify({"ok": True, "coordinadoras": coordinadoras})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_presidentas_coordinadoras")
@requiere_sesion
def ver_presidentas_coordinadoras():
    """Reporte combinado: todas las presidentas Y todas las coordinadoras
    juntas en una sola lista, ordenadas primero por comunidad (dirección) y
    luego por presidenta."""
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        sh = get_sheet()
        filas_out = []

        # --- Presidentas ---
        try:
            ws_pr = sh.worksheet("PRESIDENTAS")
            for fila in ws_pr.get_all_values()[1:]:
                if not fila or not fila[0]:
                    continue
                jefe_fila = fila[0].strip()
                if jefe_filtro and jefe_fila.upper() != jefe_filtro:
                    continue
                nombre = fila[1].strip() if len(fila) > 1 else ''
                filas_out.append({
                    "tipo": "Presidenta",
                    "nombre": nombre,
                    "cui": fila[2].strip() if len(fila) > 2 else '',
                    "telefono": fila[3].strip() if len(fila) > 3 else '',
                    "direccion": fila[4].strip() if len(fila) > 4 else '',
                    "empadronado": fila[5].strip() if len(fila) > 5 else '',
                    "presidenta": nombre,
                    "grupo": '',
                    "jefe": jefe_fila,
                })
        except Exception:
            pass

        # --- Coordinadoras (misma lógica que /ver_coordinadoras) ---
        try:
            ws = sh.worksheet("GRUPOS")
            filas_g = ws.get_all_values()
            if len(filas_g) > 1:
                header_bg = filas_g[0]
                tiene_dir_bg = any('irecc' in str(h) for h in header_bg)
                tiene_pres_bg = any('residenta' in str(h) for h in header_bg)

                if tiene_pres_bg:
                    gi_jefe=3; gi_pres=4; gi_dpi_pers=8
                elif tiene_dir_bg:
                    gi_jefe=3; gi_pres=-1; gi_dpi_pers=6
                else:
                    gi_jefe=3; gi_pres=-1; gi_dpi_pers=5

                gi_no = gi_dpi_pers - 2
                gi_nombre = gi_dpi_pers - 1
                if tiene_dir_bg:
                    gi_dir = gi_dpi_pers + 1
                    gi_tel = gi_dir + 1
                else:
                    gi_dir = -1
                    gi_tel = gi_dpi_pers + 1
                gi_emp = gi_tel + 1

                def gc(f, i): return f[i].strip() if i >= 0 and i < len(f) else ''

                presidenta_por_jefe = {}
                if not tiene_pres_bg:
                    try:
                        ws_pr2 = sh.worksheet("PRESIDENTAS")
                        for fp in ws_pr2.get_all_values()[1:]:
                            if fp and len(fp) > 1:
                                presidenta_por_jefe[fp[0].strip().upper()] = fp[1].strip()
                    except Exception:
                        pass

                vistos = set()
                for fila in filas_g[1:]:
                    if not fila or not fila[0].strip():
                        continue
                    nombre_grupo = fila[0].strip()
                    jefe_fila = gc(fila, gi_jefe)
                    if jefe_filtro and jefe_fila.upper() != jefe_filtro:
                        continue
                    if gc(fila, gi_no) != '1':
                        continue
                    if nombre_grupo in vistos:
                        continue
                    vistos.add(nombre_grupo)
                    presidenta = gc(fila, gi_pres) if gi_pres >= 0 else presidenta_por_jefe.get(jefe_fila.upper(), '')
                    filas_out.append({
                        "tipo": "Coordinadora",
                        "nombre": gc(fila, gi_nombre),
                        "cui": gc(fila, gi_dpi_pers),
                        "telefono": gc(fila, gi_tel),
                        "direccion": gc(fila, gi_dir),
                        "empadronado": gc(fila, gi_emp),
                        "presidenta": presidenta or '(sin presidenta)',
                        "grupo": nombre_grupo,
                        "jefe": jefe_fila,
                    })
        except Exception:
            pass

        filas_out.sort(key=lambda p: (p["direccion"].upper(), p["presidenta"].upper()))
        return jsonify({"ok": True, "registros": filas_out})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


def _recolectar_registros_combinados(jefe_filtro):
    """Junta TODAS las presidentas, coordinadoras e integrantes en una sola
    lista (sin ordenar todavía), cada una marcada con su tipo. Se usa tanto
    para el reporte combinado completo como para el de 'No Empadronados'."""
    sh = get_sheet()
    registros = []

    # Fecha de nacimiento (columna I) y si el DPI fue verificado como real
    # (columna W) por CUI, tomadas de la hoja DPI, ya que ni GRUPOS ni
    # PRESIDENTAS guardan esos datos directamente.
    fnac_por_cui = {}
    no_verificado_por_cui = {}
    fecha_registro_por_cui = {}
    observaciones_por_cui = {}
    fecha_revision_por_cui = {}
    try:
        ws_dpi_fnac = sh.worksheet(HOJA)
        for fd in ws_dpi_fnac.get_all_values()[1:]:
            if fd and fd[0].strip() and len(fd) > 8:
                cui_limpio = fd[0].strip().replace(" ","").replace("-","")
                fnac_por_cui[cui_limpio] = fd[8].strip()
                # Columna W (índice 22) = "NO" si se guardó pese a la
                # advertencia de que la foto no parecía un DPI físico real.
                no_verificado_por_cui[cui_limpio] = len(fd) > 22 and fd[22].strip().upper() == "NO"
                # Columna X (índice 23) = fecha (DD/MM/AAAA) en que se
                # agregó el registro. Los registros creados antes de que
                # se empezara a guardar esta columna quedan sin fecha.
                fecha_registro_por_cui[cui_limpio] = fd[23].strip() if len(fd) > 23 else ''
                # Columna Y (índice 24) = observaciones libres (ej. "no
                # vota en este municipio, pero se registra igual").
                observaciones_por_cui[cui_limpio] = fd[24].strip() if len(fd) > 24 else ''
                # Columna Z (índice 25) = fecha (DD/MM/AAAA) de la última
                # vez que alguien confirmó a mano en el Padrón y en la
                # página del TSE que el estado de esta persona seguía
                # igual. Sirve para no tener que revisar todos los días a
                # quien ya se revisó recientemente.
                fecha_revision_por_cui[cui_limpio] = fd[25].strip() if len(fd) > 25 else ''
    except Exception:
        pass

    # 1. Coordinadoras e integrantes (de GRUPOS)
    try:
        ws = sh.worksheet("GRUPOS")
        filas = ws.get_all_values()
        if len(filas) > 1:
            header_bg = filas[0]
            tiene_dir_bg = any('irecc' in str(h) for h in header_bg)
            tiene_pres_bg = any('residenta' in str(h) for h in header_bg)
            if tiene_pres_bg:
                gi_jefe=3; gi_pres=4; gi_dpi_pers=8
            elif tiene_dir_bg:
                gi_jefe=3; gi_pres=-1; gi_dpi_pers=6
            else:
                gi_jefe=3; gi_pres=-1; gi_dpi_pers=5
            gi_no = gi_dpi_pers - 2
            gi_nombre = gi_dpi_pers - 1
            if tiene_dir_bg:
                gi_dir = gi_dpi_pers + 1
                gi_tel = gi_dir + 1
            else:
                gi_dir = -1
                gi_tel = gi_dpi_pers + 1
            gi_emp = gi_tel + 1

            def gc(f, i): return f[i].strip() if i >= 0 and i < len(f) else ''

            coord_por_grupo = {}
            for fila in filas[1:]:
                if fila and fila[0].strip() and gc(fila, gi_no) == '1':
                    coord_por_grupo[fila[0].strip()] = gc(fila, gi_nombre)

            presidenta_por_jefe = {}
            if not tiene_pres_bg:
                try:
                    ws_pr0 = sh.worksheet("PRESIDENTAS")
                    for fp0 in ws_pr0.get_all_values()[1:]:
                        if fp0 and len(fp0) > 1:
                            presidenta_por_jefe[fp0[0].strip().upper()] = fp0[1].strip()
                except Exception:
                    pass

            for fila in filas[1:]:
                if not fila or not fila[0].strip():
                    continue
                nombre_grupo = fila[0].strip()
                jefe_fila = gc(fila, gi_jefe)
                if jefe_filtro and jefe_fila.upper() != jefe_filtro:
                    continue
                presidenta = gc(fila, gi_pres) if gi_pres >= 0 else presidenta_por_jefe.get(jefe_fila.upper(), '')
                es_coord = gc(fila, gi_no) == '1'
                cui_persona = gc(fila, gi_dpi_pers)
                registros.append({
                    "tipo": "Coordinadora" if es_coord else "Integrante",
                    "presidenta": presidenta or '(sin presidenta)',
                    "coordinadora": coord_por_grupo.get(nombre_grupo, ''),
                    "grupo": nombre_grupo,
                    "jefe": jefe_fila,
                    "nombre": gc(fila, gi_nombre),
                    "cui": cui_persona,
                    "telefono": gc(fila, gi_tel),
                    "direccion": gc(fila, gi_dir),
                    "empadronado": gc(fila, gi_emp),
                    "fecha_nacimiento": fnac_por_cui.get(cui_persona.replace(" ","").replace("-",""), ''),
                    "dpi_no_verificado": no_verificado_por_cui.get(cui_persona.replace(" ","").replace("-",""), False),
                    "fecha_registro": fecha_registro_por_cui.get(cui_persona.replace(" ","").replace("-",""), ''),
                    "observaciones": observaciones_por_cui.get(cui_persona.replace(" ","").replace("-",""), ''),
                    "fecha_revision": fecha_revision_por_cui.get(cui_persona.replace(" ","").replace("-",""), ''),
                })
    except Exception:
        pass

    # 2. Presidentas (de PRESIDENTAS)
    try:
        ws_pr = sh.worksheet("PRESIDENTAS")
        for fp in ws_pr.get_all_values()[1:]:
            if not fp or not fp[0].strip():
                continue
            jefe_p = fp[0].strip()
            if jefe_filtro and jefe_p.upper() != jefe_filtro:
                continue
            nombre_p = fp[1].strip() if len(fp) > 1 else ''
            cui_p = fp[2].strip() if len(fp) > 2 else ''
            registros.append({
                "tipo": "Presidenta",
                "presidenta": nombre_p,
                "coordinadora": '',
                "grupo": '',
                "jefe": jefe_p,
                "nombre": nombre_p,
                "cui": cui_p,
                "telefono": fp[3].strip() if len(fp) > 3 else '',
                "direccion": fp[4].strip() if len(fp) > 4 else '',
                "empadronado": fp[5].strip() if len(fp) > 5 else '',
                "fecha_nacimiento": fnac_por_cui.get(cui_p.replace(" ","").replace("-",""), ''),
                "dpi_no_verificado": no_verificado_por_cui.get(cui_p.replace(" ","").replace("-",""), False),
                "fecha_registro": fecha_registro_por_cui.get(cui_p.replace(" ","").replace("-",""), ''),
                "observaciones": observaciones_por_cui.get(cui_p.replace(" ","").replace("-",""), ''),
                "fecha_revision": fecha_revision_por_cui.get(cui_p.replace(" ","").replace("-",""), ''),
            })
    except Exception:
        pass

    return registros


def _ordenar_registros_combinados(registros):
    """Orden: primero por la comunidad de cada presidenta, luego su
    nombre -> dentro de eso, sus coordinadoras ordenadas de la misma
    forma (primero por la comunidad de CADA coordinadora, luego su
    nombre), y bajo cada coordinadora, la coordinadora misma antes que
    sus integrantes (también por comunidad y luego nombre entre
    ellos). El mismo patrón "comunidad, luego nombre" se repite en
    cada nivel: Presidenta -> Coordinadora -> Integrantes."""
    presidenta_comunidad = {}
    coordinadora_comunidad = {}
    for reg in registros:
        if reg["tipo"] == "Presidenta":
            presidenta_comunidad[reg["presidenta"].upper()] = reg["direccion"].upper()
        elif reg["tipo"] == "Coordinadora":
            coordinadora_comunidad[(reg["presidenta"].upper(), reg["coordinadora"].upper())] = reg["direccion"].upper()

    orden_tipo = {"Presidenta": 0, "Coordinadora": 0, "Integrante": 1}
    registros.sort(key=lambda p: (
        presidenta_comunidad.get(p["presidenta"].upper(), ''),
        p["presidenta"].upper(),
        coordinadora_comunidad.get((p["presidenta"].upper(), p["coordinadora"].upper()), ''),
        p["coordinadora"].upper(),
        orden_tipo.get(p["tipo"], 1),
        p["direccion"].upper(),
        p["nombre"].upper()
    ))
    return registros


def _ordenar_por_jefe_luego_jerarquia(registros):
    """Igual que _ordenar_registros_combinados (Presidenta -> Coordinadora ->
    Integrantes, por comunidad y nombre), pero agrupando primero por Jefe de
    Sector. Como Python ordena de forma estable, basta con aplicar primero
    el orden jerárquico normal y luego reordenar por jefe: el orden interno
    de cada jefe se conserva."""
    registros = _ordenar_registros_combinados(registros)
    registros.sort(key=lambda r: (r.get("jefe","") or "Sin Jefe Asignado").upper())
    return registros


def _enviar_correo_via_brevo(destinatarios, asunto, cuerpo_texto, adjuntos):
    """Manda UN correo (con uno o varios archivos adjuntos) usando la API
    de Brevo (https://api.brevo.com), que se comunica por HTTPS normal —
    a diferencia de Gmail por SMTP, esto SÍ funciona desde Render, porque
    Render no bloquea el tráfico HTTPS (si lo bloqueara, la app entera
    dejaría de funcionar, ya que también usa HTTPS para hablar con Google
    Sheets y con la IA).

    'adjuntos' es una lista de tuplas (bytes_del_archivo, nombre_archivo) —
    se pueden mandar varios archivos distintos en el mismo correo.

    Necesita las variables de entorno BREVO_API_KEY (la API key generada
    en Brevo) y BREVO_SENDER_EMAIL (el correo verificado en Brevo como
    'remitente'). Devuelve (True, "") si se envió bien, o (False, "mensaje
    de error") si algo falló — nunca lanza la excepción hacia arriba, para
    que un correo fallido no tumbe el envío de los demás."""
    api_key = os.environ.get("BREVO_API_KEY", "").strip()
    remitente = os.environ.get("BREVO_SENDER_EMAIL", "").strip()
    destinatarios = [d.strip() for d in destinatarios if d and d.strip()]
    if not api_key or not remitente:
        return False, "Faltan las variables de entorno BREVO_API_KEY / BREVO_SENDER_EMAIL en Render"
    if not destinatarios:
        return False, "No hay destinatarios configurados"
    try:
        payload = {
            "sender": {"email": remitente},
            "to": [{"email": d} for d in destinatarios],
            "subject": asunto,
            "textContent": cuerpo_texto,
            "attachment": [
                {"content": base64.b64encode(contenido).decode("ascii"), "name": nombre}
                for (contenido, nombre) in adjuntos
            ],
        }
        peticion = urllib.request.Request(
            "https://api.brevo.com/v3/smtp/email",
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "accept": "application/json",
                "api-key": api_key,
                "content-type": "application/json",
            },
            method="POST",
        )
        with urllib.request.urlopen(peticion, timeout=25) as resp:
            resp.read()
        return True, ""
    except urllib.error.HTTPError as e:
        try:
            detalle = e.read().decode("utf-8", errors="replace")
        except Exception:
            detalle = str(e)
        return False, f"Brevo rechazó el envío ({e.code}): {detalle}"
    except Exception as e:
        return False, str(e)


def _filtrar_solo_pendientes_manteniendo_estructura(registros):
    """Para el organigrama del correo diario: arma la estructura de abajo
    hacia arriba y solo deja una Presidenta o Coordinadora si ELLA MISMA
    tiene algo pendiente (falta empadronarse y/o teléfono) O tiene a
    alguien pendiente debajo. Si una Presidenta/Coordinadora no tiene
    absolutamente nada que revisar, se quita por completo — no tiene
    sentido mandarla solo para mostrar una rama vacía."""
    def _esta_pendiente(persona):
        falta_emp = (persona.get('empadronado') or '').strip().upper() != 'SI'
        falta_tel = not (persona.get('telefono') or '').strip()
        return falta_emp or falta_tel

    por_presidenta = {}
    orden_presidentas = []
    for r in registros:
        p = (r.get('presidenta') or '').strip()
        if p not in por_presidenta:
            por_presidenta[p] = []
            orden_presidentas.append(p)
        por_presidenta[p].append(r)

    resultado = []
    for nombre_p in orden_presidentas:
        miembros = por_presidenta[nombre_p]
        pres_reg = next((m for m in miembros if m.get('tipo') == 'Presidenta'), None)

        por_coord = {}
        orden_coords = []
        integrantes_directos = []
        for m in miembros:
            if m.get('tipo') == 'Presidenta':
                continue
            if m.get('tipo') == 'Coordinadora':
                nombre_c = m.get('nombre', '-')
                if nombre_c not in por_coord:
                    por_coord[nombre_c] = {'reg': m, 'integrantes': []}
                    orden_coords.append(nombre_c)
                else:
                    por_coord[nombre_c]['reg'] = m
            else:
                coord_de = (m.get('coordinadora') or '').strip()
                if coord_de and coord_de != '-':
                    if coord_de not in por_coord:
                        por_coord[coord_de] = {'reg': None, 'integrantes': []}
                        orden_coords.append(coord_de)
                    por_coord[coord_de]['integrantes'].append(m)
                else:
                    integrantes_directos.append(m)

        presidenta_tiene_algo_pendiente = pres_reg is not None and _esta_pendiente(pres_reg)
        aportes_de_esta_presidenta = []

        for nombre_c in orden_coords:
            info = por_coord[nombre_c]
            integrantes_pendientes = [i for i in info['integrantes'] if _esta_pendiente(i)]
            coordinadora_pendiente = info['reg'] is not None and _esta_pendiente(info['reg'])
            if integrantes_pendientes or coordinadora_pendiente:
                if info['reg'] is not None:
                    aportes_de_esta_presidenta.append(info['reg'])
                aportes_de_esta_presidenta.extend(integrantes_pendientes)
                presidenta_tiene_algo_pendiente = True

        integrantes_directos_pendientes = [i for i in integrantes_directos if _esta_pendiente(i)]
        if integrantes_directos_pendientes:
            aportes_de_esta_presidenta.extend(integrantes_directos_pendientes)
            presidenta_tiene_algo_pendiente = True

        if presidenta_tiene_algo_pendiente:
            if pres_reg is not None:
                resultado.append(pres_reg)
            resultado.extend(aportes_de_esta_presidenta)

    return resultado


@app.route("/cron/reporte_diario_sin_telefono")
def cron_reporte_diario_sin_telefono():
    """Pensado para ser llamado por un disparador EXTERNO (por ejemplo,
    cron-job.org) todos los días a las 8:00am hora de Guatemala, ya que el
    plan gratuito de Render 'duerme' el servicio y no puede despertarse
    solo. Genera el ORGANIGRAMA gráfico completo de cada jefe (con los
    colores que marcan quién no está empadronado y/o no tiene teléfono) y
    lo manda por correo: a cada jefe le llega SOLO el organigrama de su
    propia gente, y a los correos generales (configurados en Mantenimiento)
    les llega el organigrama completo de todos los jefes.

    Protegida con un token en vez de sesión, porque quien la llama no es
    una persona con sesión iniciada sino el disparador externo. El token se
    configura en Render con la variable de entorno CRON_SECRET_TOKEN, y
    debe coincidir con el que se manda en la URL (?token=...)."""
    token_esperado = os.environ.get("CRON_SECRET_TOKEN", "").strip()
    token_recibido = request.args.get("token", "").strip()
    if not token_esperado or token_recibido != token_esperado:
        return jsonify({"ok": False, "error": "Token inválido o no configurado"}), 403

    try:
        sh = get_sheet()
        todos_registros = _recolectar_registros_combinados("")  # "" = todos los jefes
        todos_registros = _ordenar_por_jefe_luego_jerarquia(todos_registros)

        total_sin_tel = sum(1 for r in todos_registros if not r.get("telefono","").strip())
        total_sin_emp = sum(1 for r in todos_registros if r.get("empadronado","").strip().upper() != "SI")

        fecha_hoy = ahora_gt().strftime('%d/%m/%Y')
        resultado = {"enviados": [], "fallidos": [],
                     "total_sin_telefono": total_sin_tel,
                     "total_sin_empadronar": total_sin_emp}

        # 1. Un correo por cada Jefe de Sector, con el ORGANIGRAMA de SU
        # propia gente (colores: amarillo=sin empadronar, verde=sin
        # teléfono, naranja=ambos).
        ws_jefes = _get_ws_jefes(sh)
        filas_jefes = ws_jefes.get_all_values()
        por_jefe = {}
        for r in todos_registros:
            por_jefe.setdefault(r.get("jefe","") or "Sin Jefe Asignado", []).append(r)

        for fila in filas_jefes[1:]:
            if not fila or not fila[0].strip():
                continue
            nombre_jefe = fila[0].strip()
            emails_jefe = [fila[i].strip() for i in range(3, 6) if len(fila) > i and fila[i].strip()]
            if not emails_jefe:
                continue  # este jefe no tiene correo configurado, se omite

            regs_jefe = por_jefe.get(nombre_jefe, [])
            sin_tel_jefe = sum(1 for r in regs_jefe if not r.get("telefono","").strip())
            sin_emp_jefe = sum(1 for r in regs_jefe if r.get("empadronado","").strip().upper() != "SI")
            regs_jefe_filtrado = _filtrar_solo_pendientes_manteniendo_estructura(regs_jefe)
            if not regs_jefe_filtrado:
                # Nadie pendiente para este jefe hoy — no tiene caso
                # mandarle un correo con un organigrama vacío.
                resultado["enviados"].append({"jefe": nombre_jefe, "emails": [], "error": None, "sin_pendientes": True})
                continue
            pdf_organigrama = _generar_pdf_organigrama_bytes(nombre_jefe, regs_jefe_filtrado, "ORGANIGRAMA — PENDIENTES")
            adjuntos = [(pdf_organigrama, f"Organigrama_{nombre_jefe.replace(' ','_')}.pdf")]
            cuerpo = (f"Buen día,\n\nAdjunto el organigrama de estructura correspondiente a {nombre_jefe}, actualizado al {fecha_hoy}.\n\n"
                      f"Los colores marcan quién necesita atención:\n"
                      f"🟨 Amarillo = sin empadronar ({sin_emp_jefe} persona(s))\n"
                      f"🟩 Verde = sin número de teléfono ({sin_tel_jefe} persona(s))\n"
                      f"🟧 Naranja = le faltan las dos cosas\n\n"
                      f"Este correo se envía automáticamente todos los días.")
            ok, err = _enviar_correo_via_brevo(emails_jefe, f"Organigrama diario — {nombre_jefe} — {fecha_hoy}", cuerpo, adjuntos)
            (resultado["enviados"] if ok else resultado["fallidos"]).append({"jefe": nombre_jefe, "emails": emails_jefe, "error": (err[:200] if err else None) if not ok else None})

        # 2. Un correo con el organigrama COMPLETO (todos los jefes) a los
        # correos generales.
        ws_config = _get_ws_config_correos(sh)
        filas_config = ws_config.get_all_values()
        fila_config = filas_config[1] if len(filas_config) > 1 else []
        emails_generales = [fila_config[i].strip() for i in range(3) if len(fila_config) > i and fila_config[i].strip()]
        if emails_generales:
            todos_registros_filtrado = _filtrar_solo_pendientes_manteniendo_estructura(todos_registros)
            if not todos_registros_filtrado:
                resultado["enviados"].append({"jefe": "TODOS (general)", "emails": [], "error": None, "sin_pendientes": True})
            else:
                pdf_general = _generar_pdf_organigrama_bytes("TODOS LOS JEFES DE SECTOR", todos_registros_filtrado, "ORGANIGRAMA — PENDIENTES")
                adjuntos_gral = [(pdf_general, "Organigrama_TODOS.pdf")]
                cuerpo_gral = (f"Buen día,\n\nAdjunto el organigrama de estructura completo (todos los jefes de sector), actualizado al {fecha_hoy}.\n\n"
                               f"Los colores marcan quién necesita atención:\n"
                               f"🟨 Amarillo = sin empadronar ({total_sin_emp} persona(s))\n"
                               f"🟩 Verde = sin número de teléfono ({total_sin_tel} persona(s))\n"
                               f"🟧 Naranja = le faltan las dos cosas\n\n"
                               f"Este correo se envía automáticamente todos los días.")
                ok, err = _enviar_correo_via_brevo(emails_generales, f"Organigrama diario — TODOS — {fecha_hoy}", cuerpo_gral, adjuntos_gral)
                (resultado["enviados"] if ok else resultado["fallidos"]).append({"jefe": "TODOS (general)", "emails": emails_generales, "error": (err[:200] if err else None) if not ok else None})

        _registrar_auditoria(sh, "SISTEMA (cron)", "Envió organigrama diario",
                              f"{len(resultado['enviados'])} correo(s) enviados, {len(resultado['fallidos'])} fallido(s)")
        return jsonify({"ok": True, "resultado": resultado})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_todo_combinado")
@requiere_sesion
def ver_todo_combinado():
    """Combina TODAS las presidentas, coordinadoras e integrantes en una
    sola lista, cada una marcada con su tipo (Presidenta/Coordinadora/
    Integrante), ordenadas jerárquicamente: primero por presidenta, luego
    por coordinadora, luego por comunidad y nombre. Filtrado opcionalmente
    por jefe de sector (según sesión)."""
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        registros = _recolectar_registros_combinados(jefe_filtro)
        registros = _ordenar_registros_combinados(registros)
        return jsonify({"ok": True, "registros": registros})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_con_observaciones")
@requiere_sesion
def ver_con_observaciones():
    """Devuelve solo a las personas que tienen algo escrito en el campo de
    Observaciones (ej. 'empadronada pero vota en otro municipio'), para dar
    seguimiento o imprimir un listado de por qué quedaron marcadas así."""
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        registros = _recolectar_registros_combinados(jefe_filtro)
        registros = [r for r in registros if r.get("observaciones","").strip()]
        registros = _ordenar_registros_combinados(registros)
        return jsonify({"ok": True, "registros": registros})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_no_empadronados")
@requiere_sesion
def ver_no_empadronados():
    """Igual que /ver_todo_combinado, pero filtra según lo que se pida:
    personas sin empadronar, sin teléfono, o ambas condiciones a la vez
    (con criterio OR: se incluye a quien le falte AL MENOS una de las
    cosas marcadas). Útil para dar seguimiento a quién falta por
    empadronarse y/o a quién falta conseguirle un número de teléfono."""
    try:
        incluir_no_emp = request.args.get("no_emp", "SI").strip().upper() == "SI"
        incluir_sin_tel = request.args.get("sin_tel", "").strip().upper() == "SI"
        # Si viene en SI, se muestran también los que ya se revisaron hace
        # menos de 7 días (por defecto se OCULTAN, para no repetir cada
        # día el mismo trabajo con la misma gente ya confirmada en
        # Padrón/TSE).
        incluir_revisados_recientes = request.args.get("incluir_revisados", "").strip().upper() == "SI"
        # Fecha (DD/MM/AAAA) para ver solo lo agregado ese día. Si viene
        # vacía, no se filtra por fecha (se ven todos, sin importar cuándo
        # se agregaron).
        fecha_filtro = request.args.get("fecha", "").strip()
        # Si por algún motivo no se pidió ninguno de los dos, se mantiene
        # el comportamiento de siempre: solo no empadronados.
        if not incluir_no_emp and not incluir_sin_tel:
            incluir_no_emp = True

        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        registros = _recolectar_registros_combinados(jefe_filtro)

        def _le_falta_algo(r):
            dias_revision = _dias_desde_fecha_ddmmaaaa(r.get("fecha_revision", ""))
            revisado_recientemente = (not incluir_revisados_recientes) and dias_revision is not None and dias_revision < 7
            falta_emp = incluir_no_emp and r.get("empadronado","").strip().upper() != "SI" and not revisado_recientemente
            falta_tel = incluir_sin_tel and not r.get("telefono","").strip()
            return falta_emp or falta_tel

        registros = [r for r in registros if _le_falta_algo(r)]
        if fecha_filtro:
            registros = [r for r in registros if r.get("fecha_registro","").strip() == fecha_filtro]
        registros = _ordenar_registros_combinados(registros)
        return jsonify({"ok": True, "registros": registros, "incluyo_no_emp": incluir_no_emp, "incluyo_sin_tel": incluir_sin_tel})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_integrantes")
@requiere_sesion
def ver_integrantes():
    """Devuelve TODAS las personas de TODOS los grupos (coordinadoras e
    integrantes por igual), con su presidenta y coordinadora, ordenadas
    jerárquicamente: primero por presidenta, luego por coordinadora, luego
    por nombre. Filtrado opcionalmente por jefe de sector (según sesión)."""
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        sh = get_sheet()
        try:
            ws = sh.worksheet("GRUPOS")
        except Exception:
            return jsonify({"ok": True, "integrantes": []})

        filas = ws.get_all_values()
        if len(filas) <= 1:
            return jsonify({"ok": True, "integrantes": []})

        header_bg = filas[0]
        tiene_dir_bg = any('irecc' in str(h) for h in header_bg)
        tiene_pres_bg = any('residenta' in str(h) for h in header_bg)

        if tiene_pres_bg:
            gi_jefe=3; gi_pres=4; gi_dpi_pers=8
        elif tiene_dir_bg:
            gi_jefe=3; gi_pres=-1; gi_dpi_pers=6
        else:
            gi_jefe=3; gi_pres=-1; gi_dpi_pers=5

        gi_no = gi_dpi_pers - 2
        gi_nombre = gi_dpi_pers - 1
        if tiene_dir_bg:
            gi_dir = gi_dpi_pers + 1
            gi_tel = gi_dir + 1
        else:
            gi_dir = -1
            gi_tel = gi_dpi_pers + 1
        gi_emp = gi_tel + 1

        def gc(f, i): return f[i].strip() if i >= 0 and i < len(f) else ''

        # Nombre de coordinadora por grupo (fila No.=1), y presidenta por
        # jefe de sector (si la hoja no la tiene directamente), para poder
        # anotar ambos datos en cada integrante.
        coord_por_grupo = {}
        for fila in filas[1:]:
            if fila and fila[0].strip() and gc(fila, gi_no) == '1':
                coord_por_grupo[fila[0].strip()] = gc(fila, gi_nombre)

        presidenta_por_jefe = {}
        if not tiene_pres_bg:
            try:
                ws_pr = sh.worksheet("PRESIDENTAS")
                for fp in ws_pr.get_all_values()[1:]:
                    if fp and len(fp) > 1:
                        presidenta_por_jefe[fp[0].strip().upper()] = fp[1].strip()
            except Exception:
                pass

        integrantes = []
        for fila in filas[1:]:
            if not fila or not fila[0].strip():
                continue
            nombre_grupo = fila[0].strip()
            jefe_fila = gc(fila, gi_jefe)
            if jefe_filtro and jefe_fila.upper() != jefe_filtro:
                continue
            presidenta = gc(fila, gi_pres) if gi_pres >= 0 else presidenta_por_jefe.get(jefe_fila.upper(), '')
            integrantes.append({
                "presidenta": presidenta or '(sin presidenta)',
                "coordinadora": coord_por_grupo.get(nombre_grupo, ''),
                "grupo": nombre_grupo,
                "jefe": jefe_fila,
                "nombre": gc(fila, gi_nombre),
                "cui": gc(fila, gi_dpi_pers),
                "telefono": gc(fila, gi_tel),
                "direccion": gc(fila, gi_dir),
                "empadronado": gc(fila, gi_emp),
                "es_coordinadora": gc(fila, gi_no) == '1',
            })

        integrantes.sort(key=lambda p: (p["presidenta"].upper(), p["coordinadora"].upper(), p["direccion"].upper(), p["nombre"].upper()))
        return jsonify({"ok": True, "integrantes": integrantes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/ver_grupos")
@requiere_sesion
def ver_grupos():
    try:
        sh = get_sheet()
        try:
            ws = sh.worksheet("GRUPOS")
        except:
            return jsonify({"ok": True, "grupos": [], "resumen_jefes": {}})
        filas = ws.get_all_values()
        if len(filas) <= 1:
            return jsonify({"ok": True, "grupos": [], "resumen_jefes": {}})
        grupos = []
        conteo = {}
        resumen_jefes = {}
        grupos_por_jefe = {}

        # Detect format by checking header row
        header = filas[0] if filas else []
        # New format has Jefe de Sector at col 3
        # Old format: Grupo(0),Coordinador(1),DPI Coord(2),No.(3),Nombre(4),DPI(5),Tel(6),Emp(7),Fecha(8)
        # New format: Grupo(0),Coordinador(1),DPI Coord(2),Jefe(3),No.(4),Nombre(5),DPI(6),Tel(7),Emp(8),Fecha(9)
        tiene_jefe = len(header) >= 4 and 'Jefe' in header[3] if header else False
        tiene_pres = any('residenta' in str(h) for h in header) if header else False
        tiene_dir  = any('irecc'     in str(h) for h in header) if header else False

        for fila in filas[1:]:
            if not fila or not fila[0]:
                continue
            nombre = fila[0].strip()

            if tiene_jefe:
                jefe = fila[3].strip() if len(fila) > 3 else ''
                if tiene_pres:
                    # Formato nuevo con Presidenta: Grupo,Coord,DPI,Jefe,Pres,DPIPres,No,Nombre,DPI,Dir,Tel,Emp,Fecha
                    presidenta = fila[4].strip() if len(fila) > 4 else ''
                    no        = fila[6].strip() if len(fila) > 6 else ''
                    nombre_p  = fila[7].strip() if len(fila) > 7 else ''
                    dpi       = fila[8].strip() if len(fila) > 8 else ''
                    direccion = fila[9].strip() if len(fila) > 9 else ''
                    tel       = fila[10].strip() if len(fila) > 10 else ''
                    emp       = fila[11].strip() if len(fila) > 11 else ''
                elif tiene_dir:
                    # Formato con Dir sin Presidenta: Grupo,Coord,DPI,Jefe,No,Nombre,DPI,Dir,Tel,Emp,Fecha
                    presidenta = ''
                    no        = fila[4].strip() if len(fila) > 4 else ''
                    nombre_p  = fila[5].strip() if len(fila) > 5 else ''
                    dpi       = fila[6].strip() if len(fila) > 6 else ''
                    direccion = fila[7].strip() if len(fila) > 7 else ''
                    tel       = fila[8].strip() if len(fila) > 8 else ''
                    emp       = fila[9].strip() if len(fila) > 9 else ''
                else:
                    # Formato sin Dir ni Presidenta
                    presidenta = ''
                    no        = fila[4].strip() if len(fila) > 4 else ''
                    nombre_p  = fila[5].strip() if len(fila) > 5 else ''
                    dpi       = fila[6].strip() if len(fila) > 6 else ''
                    direccion = ''
                    tel       = fila[7].strip() if len(fila) > 7 else ''
                    emp       = fila[8].strip() if len(fila) > 8 else ''
            else:
                jefe = ''; presidenta = ''
                no        = fila[3].strip() if len(fila) > 3 else ''
                nombre_p  = fila[4].strip() if len(fila) > 4 else ''
                dpi       = fila[5].strip() if len(fila) > 5 else ''
                direccion = ''
                tel       = fila[6].strip() if len(fila) > 6 else ''
                emp       = fila[7].strip() if len(fila) > 7 else ''

            _ = direccion

            conteo[nombre] = conteo.get(nombre, 0) + 1
            if nombre not in grupos_por_jefe:
                grupos_por_jefe[nombre] = jefe

            grupos.append({
                "nombre": nombre,
                "coordinador": fila[1].strip() if len(fila) > 1 else '',
                "dpi_coord": fila[2].strip() if len(fila) > 2 else '',
                "jefe_sector": jefe,
                "presidenta": presidenta,
                "no": no,
                "nombre_persona": nombre_p,
                "dpi": dpi,
                "direccion": direccion,
                "telefono": tel,
                "empadronado": emp,
                "total": conteo[nombre]
            })

        # Build summary by jefe
        for nombre_g, jefe in grupos_por_jefe.items():
            if jefe:
                if jefe not in resumen_jefes:
                    resumen_jefes[jefe] = 0
                resumen_jefes[jefe] += 1

        # Para grupos sin presidenta, buscar en hoja PRESIDENTAS
        try:
            ws_pr = sh.worksheet("PRESIDENTAS")
            filas_pr = ws_pr.get_all_values()
            pres_por_jefe = {}
            for fp in filas_pr[1:]:
                if fp and len(fp) > 1:
                    pres_por_jefe[fp[0].strip().upper()] = fp[1].strip()
            for g in grupos:
                if not g.get("presidenta"):
                    jefe_u = g.get("jefe_sector","").strip().upper()
                    if jefe_u in pres_por_jefe:
                        g["presidenta"] = pres_por_jefe[jefe_u]
        except: pass

        # Filtro por jefe de sector, determinado por la sesión del servidor
        # (no por lo que envíe el navegador): un jefe normal solo ve sus
        # propios grupos; el administrador los ve todos.
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()
        if jefe_filtro:
            grupos = [g for g in grupos if g.get("jefe_sector","").strip().upper() == jefe_filtro]
            resumen_jefes = {k: v for k, v in resumen_jefes.items() if k.strip().upper() == jefe_filtro}

        return jsonify({"ok": True, "grupos": grupos, "resumen_jefes": resumen_jefes})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/buscar_grupo")
@requiere_sesion
def buscar_grupo():
    try:
        nombre = request.args.get('nombre', '').strip().lower()
        dpi = request.args.get('dpi', '').strip()
        sh = get_sheet()
        try:
            ws = sh.worksheet("GRUPOS")
        except:
            return jsonify({"ok": False, "error": "No existe la hoja GRUPOS"})

        filas = ws.get_all_values()
        if len(filas) <= 1:
            return jsonify({"ok": False, "error": "No hay grupos registrados"})

        # Detectar si tiene columna Dirección
        header_bg = filas[0] if filas else []
        tiene_dir_bg = any('irecc' in str(h) for h in header_bg)

        header_pres = any('residenta' in str(h) for h in header_bg)
        tiene_pres_bg = header_pres

        # Detectar índices según formato
        if tiene_pres_bg:
            gi_dpi_coord=2; gi_jefe=3; gi_pres=4; gi_dpi_pres=5; gi_dpi_pers=8
        elif tiene_dir_bg:
            gi_dpi_coord=2; gi_jefe=3; gi_pres=-1; gi_dpi_pres=-1; gi_dpi_pers=6
        else:
            gi_dpi_coord=2; gi_jefe=3; gi_pres=-1; gi_dpi_pres=-1; gi_dpi_pers=5

        # Índices de los datos de cada persona, calculados en relación al de DPI
        # para no depender de una posición fija (evita que un cambio de columnas
        # como agregar Presidenta/DPI Presidenta desalinee los datos).
        gi_nombre_pers = gi_dpi_pers - 1
        gi_no_pers = gi_dpi_pers - 2
        if tiene_dir_bg:
            gi_dir_pers = gi_dpi_pers + 1
            gi_tel_pers = gi_dir_pers + 1
        else:
            gi_dir_pers = -1
            gi_tel_pers = gi_dpi_pers + 1
        gi_emp_pers = gi_tel_pers + 1

        def gc(f,i): return f[i].strip() if i>=0 and i<len(f) else ''

        # Limpiar DPI buscado (quitar espacios)
        dpi_limpio = dpi.replace(' ','').replace('-','').strip() if dpi else ''

        # Filtro por jefe de sector, determinado por la sesión del servidor:
        # un jefe normal solo encuentra sus propios grupos; el administrador
        # (incluyendo la búsqueda desde Mantenimiento) los encuentra todos.
        _nombre_sesion, _es_admin = _sesion_actual()
        jefe_filtro = "" if _es_admin else _nombre_sesion.upper()

        grupo_info = None
        personas = []
        for fila in filas[1:]:
            if len(fila) < 3:
                continue
            nombre_grupo = fila[0].strip()
            dpi_coord = gc(fila, gi_dpi_coord).replace(' ','')
            dpi_pers  = gc(fila, gi_dpi_pers).replace(' ','')
            jefe = gc(fila, gi_jefe)
            presidenta_bg = gc(fila, gi_pres) if gi_pres>=0 else ''
            dpi_presidenta_bg = gc(fila, gi_dpi_pres) if gi_dpi_pres>=0 else ''

            # Buscar por nombre, DPI coordinador, DPI persona o DPI jefe en col T
            col_t = fila[19].strip() if len(fila) > 19 else ''
            if grupo_info is not None:
                # Ya se identificó un grupo (por el primer resultado que
                # coincidió); de aquí en adelante solo se agregan personas
                # de ESE MISMO grupo, aunque otros grupos también coincidan
                # con la búsqueda (por ejemplo, buscar solo "A" coincide con
                # el nombre de muchos grupos distintos — antes esto mezclaba
                # a todas esas personas en un solo resultado).
                match = nombre_grupo == grupo_info["nombre"] and jefe == grupo_info["jefe_sector"]
            else:
                match = (
                    (nombre and nombre in nombre_grupo.lower()) or
                    (dpi_limpio and dpi_limpio == dpi_coord) or
                    (dpi_limpio and dpi_limpio == dpi_pers)
                )
                if jefe_filtro and jefe.strip().upper() != jefe_filtro:
                    match = False
            if match:
                if grupo_info is None:
                    grupo_info = {
                        "nombre": nombre_grupo,
                        "coordinador": gc(fila, 1),
                        "dpi_coord": dpi_coord,
                        "jefe_sector": jefe,
                        "presidenta": presidenta_bg,
                        "dpi_presidenta": dpi_presidenta_bg
                    }
                # Normalizamos cada fila a un formato fijo de 10 columnas
                # (nombre_grupo, coordinador, dpi_coord, jefe, no, nombre, dpi,
                # direccion, telefono, empadronado) para que el generador de
                # PDF siempre lea el dato correcto, sin importar el layout
                # real de la hoja GRUPOS.
                fila_normalizada = [
                    nombre_grupo,
                    gc(fila, 1),
                    dpi_coord,
                    jefe,
                    gc(fila, gi_no_pers),
                    gc(fila, gi_nombre_pers),
                    dpi_pers,
                    gc(fila, gi_dir_pers) if gi_dir_pers >= 0 else '',
                    gc(fila, gi_tel_pers),
                    gc(fila, gi_emp_pers),
                ]
                personas.append(fila_normalizada)

        if grupo_info:
            # Si no tiene presidenta (o le falta el DPI), buscar en hoja PRESIDENTAS por jefe de sector
            if not grupo_info.get("presidenta") or not grupo_info.get("dpi_presidenta"):
                try:
                    ws_pr = sh.worksheet("PRESIDENTAS")
                    filas_pr = ws_pr.get_all_values()
                    jefe_s = grupo_info.get("jefe_sector","").strip().upper()
                    for fp in filas_pr[1:]:
                        if fp and fp[0].strip().upper() == jefe_s:
                            if not grupo_info.get("presidenta"):
                                grupo_info["presidenta"] = fp[1].strip() if len(fp) > 1 else ''
                            if not grupo_info.get("dpi_presidenta"):
                                grupo_info["dpi_presidenta"] = fp[2].strip() if len(fp) > 2 else ''
                            break
                except: pass
            # Las filas ya quedaron normalizadas a 10 columnas con Direccion
            # incluida (vacía si la hoja no la tiene), así que el frontend
            # siempre debe tratarlas como si "tiene_dir" fuera verdadero.
            return jsonify({"ok": True, "grupo": grupo_info, "personas": personas, "total": len(personas), "tiene_dir": True})
        return jsonify({"ok": False, "error": "Grupo no encontrado"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/buscar_dpi")
@requiere_sesion
def buscar_dpi():
    try:
        _nombre_sesion, _es_admin = _sesion_actual()
        def es_propio(fila):
            if _puede_ver_todo():
                return True
            return len(fila) > 19 and fila[19].strip().upper() == _nombre_sesion.upper()

        cui = request.args.get('cui','').strip()
        q = request.args.get('q','').strip()
        sh = get_sheet()
        ws = sh.worksheet(HOJA)
        filas = ws.get_all_values()
        if cui:
            cui_limpio = cui.replace(' ','').replace('-','').strip()
            for i, fila in enumerate(filas[1:], start=2):
                if fila and fila[0].strip().replace(' ','') == cui_limpio and es_propio(fila):
                    return jsonify({"ok": True, "fila": fila, "numero_fila": i})
            return jsonify({"ok": False, "fila": None})
        elif q:
            # Búsqueda combinada por CUI o por nombre (usada por "Editar
            # Teléfono", que permite buscar por cualquiera de los dos).
            q_limpio = q.replace(' ','').replace('-','').strip()
            q_lower = q.strip().lower()
            for i, fila in enumerate(filas[1:], start=2):
                if not fila:
                    continue
                if not es_propio(fila):
                    continue
                cui_fila = fila[0].strip().replace(' ','') if len(fila) > 0 else ''
                # Se filtran los campos vacíos ANTES de unir con espacios,
                # para no dejar espacios dobles cuando falta el segundo
                # nombre o segundo apellido (eso hacía fallar la búsqueda
                # por nombre para personas sin segundo nombre/apellido).
                partes_nombre = [fila[j].strip() for j in (2,3,4,5) if len(fila) > j and fila[j].strip()]
                nombre_fila = ' '.join(partes_nombre).lower()
                nombre_fila_normalizado = ' '.join(nombre_fila.split())
                q_lower_normalizado = ' '.join(q_lower.split())
                if (q_limpio and cui_fila == q_limpio) or (q_lower_normalizado and q_lower_normalizado in nombre_fila_normalizado):
                    return jsonify({"ok": True, "fila": fila, "numero_fila": i})
            return jsonify({"ok": False, "fila": None})
        else:
            return jsonify({"ok": False, "error": "Falta el CUI o nombre a buscar"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    # threaded=True permite atender varias peticiones al mismo tiempo. Sin
    # esto, el servidor de Flask procesa una sola petición a la vez: si
    # varias personas usan la app junto (varios jefes + apoyos), cada quien
    # queda esperando su turno detrás de la persona anterior, lo que se
    # siente como que "la app no deja entrar a los demás".
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
